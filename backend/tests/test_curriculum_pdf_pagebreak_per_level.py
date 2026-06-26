"""Iteration 43 — Verify curriculum PDF:
- Each level gets its own page (PageBreak between levels).
- Page count == number of distinct levels.
- All pages portrait A4.
- Within a level: terms stacked vertically (KeepTogether per term).
- Excel still works; student gets 403; scoped exports still portrait.
"""
import os
import io
import pytest
import requests
from pypdf import PdfReader

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    try:
        with open("/app/frontend/.env") as f:
            for line in f:
                if line.startswith("REACT_APP_BACKEND_URL="):
                    BASE_URL = line.split("=", 1)[1].strip().strip('"').rstrip("/")
                    break
    except Exception:
        pass

DEPT_ID = "698e500997fef774e66e93a8"
EXPORT_URL = f"{BASE_URL}/api/curriculum/department/{DEPT_ID}/export"
CURRICULUM_URL = f"{BASE_URL}/api/curriculum/by-department/{DEPT_ID}"


def _login(username: str, password: str) -> str:
    r = requests.post(f"{BASE_URL}/api/auth/login",
                      json={"username": username, "password": password}, timeout=30)
    assert r.status_code == 200, f"login failed {username}: {r.status_code} {r.text[:200]}"
    return r.json()["access_token"]


@pytest.fixture(scope="module")
def admin_token():
    return _login("admin", "admin123")


@pytest.fixture(scope="module")
def student_token():
    return _login("234", "234")


@pytest.fixture(scope="module")
def distinct_levels(admin_token):
    """Compute distinct level count by querying curriculum-by-department endpoint
    or, failing that, by parsing the xlsx export (level column).
    Returns a sorted list of distinct levels and the total course count."""
    headers = {"Authorization": f"Bearer {admin_token}"}
    levels = set()
    total_courses = 0

    # Try the structured endpoint first
    r = requests.get(CURRICULUM_URL, headers=headers, timeout=30)
    if r.status_code == 200:
        data = r.json()
        # response shape may vary; try a few shapes
        courses = None
        if isinstance(data, dict):
            for key in ("courses", "items", "data", "rows"):
                if key in data and isinstance(data[key], list):
                    courses = data[key]
                    break
            if courses is None:
                # Maybe data is a grid: {levels: [{level:1, terms:{1:[...],2:[...]}}, ...]}
                if "levels" in data and isinstance(data["levels"], list):
                    for lvl_block in data["levels"]:
                        lv = lvl_block.get("level")
                        terms = lvl_block.get("terms") or {}
                        # only count levels that actually have any courses
                        has_any = False
                        if isinstance(terms, dict):
                            for arr in terms.values():
                                if isinstance(arr, list) and len(arr) > 0:
                                    has_any = True
                                    total_courses += len(arr)
                        elif isinstance(terms, list):
                            for t in terms:
                                arr = t.get("courses") or []
                                if arr:
                                    has_any = True
                                    total_courses += len(arr)
                        if has_any and lv is not None:
                            levels.add(int(lv))
                    return sorted(levels), total_courses
        elif isinstance(data, list):
            courses = data

        if courses:
            for c in courses:
                lv = c.get("level")
                if lv is None:
                    continue
                try:
                    levels.add(int(lv))
                    total_courses += 1
                except Exception:
                    pass
            if levels:
                return sorted(levels), total_courses

    # Fallback: parse xlsx
    r = requests.get(f"{EXPORT_URL}?format=xlsx", headers=headers, timeout=60)
    assert r.status_code == 200
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    for sheet in wb.worksheets:
        # find header row containing "المستوى" / "level"
        header_row_idx = None
        level_col_idx = None
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for j, v in enumerate(row):
                if v is None:
                    continue
                s = str(v).strip().lower()
                if "level" in s or "المستوى" in str(v):
                    header_row_idx = i
                    level_col_idx = j
                    break
            if header_row_idx:
                break
        if header_row_idx and level_col_idx is not None:
            for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                v = row[level_col_idx]
                if v is None or str(v).strip() == "":
                    continue
                try:
                    levels.add(int(float(str(v).strip())))
                    total_courses += 1
                except Exception:
                    pass
    return sorted(levels), total_courses


# --- 1. Basic PDF validity ---
def test_pdf_valid(admin_token):
    r = requests.get(f"{EXPORT_URL}?format=pdf",
                     headers={"Authorization": f"Bearer {admin_token}"}, timeout=60)
    assert r.status_code == 200, r.text[:200]
    assert r.headers.get("content-type", "").startswith("application/pdf")
    assert r.content[:4] == b"%PDF"
    assert b"%%EOF" in r.content[-1024:]
    assert len(r.content) > 1000


# --- 2. Every page portrait A4 ---
def test_pdf_every_page_portrait_a4(admin_token):
    r = requests.get(f"{EXPORT_URL}?format=pdf",
                     headers={"Authorization": f"Bearer {admin_token}"}, timeout=60)
    reader = PdfReader(io.BytesIO(r.content))
    assert len(reader.pages) >= 1
    for i, page in enumerate(reader.pages):
        w = float(page.mediabox.width)
        h = float(page.mediabox.height)
        assert abs(w - 595.27) < 1.5, f"page {i+1} width={w} not portrait A4"
        assert abs(h - 841.89) < 1.5, f"page {i+1} height={h} not portrait A4"
        assert h > w, f"page {i+1} landscape (w={w} h={h})"


# --- 3. Page count == number of distinct levels (KEY ASSERTION) ---
def test_pdf_page_count_equals_distinct_levels(admin_token, distinct_levels):
    levels, total = distinct_levels
    assert len(levels) > 0, "Could not determine distinct level count from API/xlsx"
    r = requests.get(f"{EXPORT_URL}?format=pdf",
                     headers={"Authorization": f"Bearer {admin_token}"}, timeout=60)
    reader = PdfReader(io.BytesIO(r.content))
    n_pages = len(reader.pages)
    print(f"\n[INFO] distinct levels={levels} total_courses={total} pdf_pages={n_pages}")
    assert n_pages == len(levels), (
        f"Expected {len(levels)} pages (one per level {levels}), "
        f"got {n_pages}. PageBreak grouping mismatch."
    )


# --- 4. Excel export unchanged ---
def test_excel_export_still_works(admin_token):
    r = requests.get(f"{EXPORT_URL}?format=xlsx",
                     headers={"Authorization": f"Bearer {admin_token}"}, timeout=60)
    assert r.status_code == 200, r.text[:200]
    ctype = r.headers.get("content-type", "")
    assert "spreadsheet" in ctype or "xlsx" in ctype
    assert r.content[:2] == b"PK"


# --- 5. Permission gate: student → 403 ---
def test_student_forbidden(student_token):
    r = requests.get(f"{EXPORT_URL}?format=pdf",
                     headers={"Authorization": f"Bearer {student_token}"}, timeout=30)
    assert r.status_code == 403


# --- 6. Scoped exports remain valid portrait PDFs ---
@pytest.mark.parametrize("qs", [
    "format=pdf&level=1",
    "format=pdf&term=1",
    "format=pdf&level=1&term=2",
])
def test_scoped_exports_portrait(admin_token, qs):
    r = requests.get(f"{EXPORT_URL}?{qs}",
                     headers={"Authorization": f"Bearer {admin_token}"}, timeout=60)
    assert r.status_code == 200, f"{qs} → {r.status_code} {r.text[:200]}"
    assert r.content[:4] == b"%PDF"
    reader = PdfReader(io.BytesIO(r.content))
    assert len(reader.pages) >= 1
    for i, p in enumerate(reader.pages):
        assert float(p.mediabox.height) > float(p.mediabox.width), \
            f"{qs} page {i+1} landscape"


# --- 7. Scoped to single level → exactly 1 page ---
def test_scoped_single_level_one_page(admin_token, distinct_levels):
    levels, _ = distinct_levels
    if not levels:
        pytest.skip("No levels detected")
    lv = levels[0]
    r = requests.get(f"{EXPORT_URL}?format=pdf&level={lv}",
                     headers={"Authorization": f"Bearer {admin_token}"}, timeout=60)
    assert r.status_code == 200
    reader = PdfReader(io.BytesIO(r.content))
    # Single-level scope should fit on 1 page since dataset is small
    assert len(reader.pages) == 1, \
        f"Single-level scope level={lv} produced {len(reader.pages)} pages, expected 1"
