from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Body, Request, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, Response
from starlette.middleware.gzip import GZipMiddleware
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
from enum import Enum
import uuid
from datetime import datetime, timedelta, timezone
from jose import JWTError, jwt
from passlib.context import CryptContext
from bson import ObjectId
from pymongo.errors import DuplicateKeyError
import json
import pandas as pd
from io import BytesIO
from functools import lru_cache
import time as time_module

# توقيت اليمن (عدن) UTC+3
YEMEN_TIMEZONE = timezone(timedelta(hours=3))

def get_yemen_time():
    """الحصول على الوقت الحالي بتوقيت اليمن (UTC+3)"""
    return datetime.now(YEMEN_TIMEZONE)

def get_yemen_date_start():
    """الحصول على بداية اليوم بتوقيت اليمن"""
    now = get_yemen_time()
    return now.replace(hour=0, minute=0, second=0, microsecond=0)


def normalize_semester_date(d):
    """تحويل تاريخ الفصل من D-M-YYYY (أو DD-MM-YYYY) إلى YYYY-MM-DD لمقارنته بـ lectures.date.
    يدعم أيضاً صيغة YYYY-MM-DD إذا كانت محفوظة هكذا.
    """
    if not d or not isinstance(d, str):
        return None
    s = d.strip()
    # YYYY-MM-DD
    if len(s) == 10 and s[4] == '-' and s[7] == '-':
        return s
    # D-M-YYYY أو DD-MM-YYYY
    parts = s.split('-')
    if len(parts) == 3:
        try:
            day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
            return f"{year:04d}-{month:02d}-{day:02d}"
        except Exception:
            return None
    return None


async def get_active_semester_with_dates(db_inst):
    """جلب الفصل الدراسي المفعّل مع تواريخ مُحوَّلة لصيغة YYYY-MM-DD.
    يبحث في كلا الحقلين status='active' و is_active=True ضماناً.
    Returns: dict {id, name, start_date, end_date} أو None
    """
    sem = await db_inst.semesters.find_one({
        "$or": [{"status": "active"}, {"is_active": True}]
    })
    if not sem:
        return None
    return {
        "id": str(sem["_id"]),
        "name": sem.get("name"),
        "start_date": normalize_semester_date(sem.get("start_date")),
        "end_date": normalize_semester_date(sem.get("end_date")),
    }


# PDF imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import arabic_reshaper
from bidi.algorithm import get_display

# استيراد النماذج من الملفات المنفصلة
from models.permissions import (
    UserRole, Permission, DEFAULT_PERMISSIONS, ALL_PERMISSIONS,
    FULL_PERMISSION_MAPPING, ScopeType, user_has_permission
)
from models.users import (
    UserBase, UserCreate, UserLogin, UserResponse, Token,
    UserPermissionScope, UserPermissionsUpdate, UserPermissionResponse
)
from models.roles import RoleCreate, RoleUpdate, RoleResponse
from models.departments import DepartmentBase, DepartmentCreate, DepartmentResponse
from models.university import (
    UniversityBase, UniversityCreate, UniversityUpdate, UniversityResponse,
    FacultyBase, FacultyCreate, FacultyUpdate, FacultyResponse
)
from models.students import StudentBase, StudentCreate, StudentResponse
from models.teachers import TeacherBase, TeacherCreate, TeacherUpdate, TeacherResponse
from models.courses import CourseBase, CourseCreate, CourseResponse, AttendanceStatus
from models.lectures import (
    LectureStatus, ACTIVE_LECTURE_STATUSES,
    LectureCreate, LectureUpdate, LectureResponse, GenerateLecturesRequest
)
from models.attendance import (
    AttendanceRecord, AttendanceSessionCreate, AttendanceResponse,
    AttendanceStats, SingleAttendanceCreate, OfflineSyncData
)
from models.semesters import SemesterStatus, SemesterCreate, SemesterUpdate, SemesterResponse
from models.settings import SemesterDates, AcademicYearConfig, SystemSettings, SettingsUpdate
from models.notifications import NotificationType, NotificationBase, NotificationCreate, NotificationResponse
from models.activity_logs import ActivityLogType, ActivityLog, ActivityLogResponse
from models.enrollments import EnrollmentCreate, EnrollmentResponse
from models.auth import ActivateStudentAccount, ChangePasswordRequest, ForceChangePasswordRequest

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# JWT Configuration
SECRET_KEY = os.environ.get('SECRET_KEY', 'ahgaff-university-secure-key-2026-x9f8k2m5n7p3q1w4')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 8  # 8 hours - يوم عمل واحد

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Security
security = HTTPBearer()

# MongoDB connection - مُحسَّن للإنتاج على Railway
# - maxPoolSize كبير لخدمة نداءات متوازية كثيرة
# - minPoolSize = 5: إبقاء 5 اتصالات جاهزة دائماً (تقضي على latency الأول)
# - compressors = 'zstd,snappy': ضغط البيانات بين السيرفر وقاعدة البيانات
# - retryWrites: إعادة المحاولة تلقائياً عند أخطاء الشبكة العابرة
# - socketTimeoutMS: منع تعليق الاتصالات الميتة
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
try:
    client = AsyncIOMotorClient(
        mongo_url,
        serverSelectionTimeoutMS=5000,
        maxPoolSize=100,
        minPoolSize=5,
        maxIdleTimeMS=60000,
        socketTimeoutMS=30000,
        connectTimeoutMS=5000,
        retryWrites=True,
        retryReads=True,
        compressors='zstd,snappy,zlib',
        appname='ahgaff-attendance',
    )
    db = client[os.environ.get('DB_NAME', 'ahgaff_attendance')]
except Exception as e:
    print(f"Warning: MongoDB connection failed: {e}")
    client = None
    db = None

# تهيئة قاعدة البيانات للـ routes
from routes.deps import set_database
if db is not None:
    set_database(db)

# استيراد الـ routes المنفصلة
from routes.auth import router as auth_router
from routes.users import router as users_router
from routes.roles import router as roles_router
from routes.departments import router as departments_router
from routes.students import router as students_router
from routes.teachers import router as teachers_router
from routes.courses import router as courses_router
from routes.notifications import router as notifications_router
from routes.teaching_load import router as teaching_load_router
from routes.attendance_approval import router as attendance_approval_router, create_change_requests_for_diff
from routes.weekly_schedule import router as weekly_schedule_router
from routes.study_plans import router as study_plans_router
from routes.admin_tools import router as admin_tools_router
from routes.dashboard import router as dashboard_router
from routes.calendar_events import router as calendar_router
from routes.global_search import router as global_search_router
from routes.entity_details import router as entity_details_router
from routes.archives import router as archives_router
from routes.archive_pdf import router as archive_pdf_router
from routes.curriculum import router as curriculum_router
from routes.student_status import router as student_status_router
from routes.alumni import router as alumni_router
from routes.student_transfer import router as student_transfer_router

# Create the main app
app = FastAPI(title="نظام حضور جامعة الأحقاف")

# Gzip compression - ضغط الردود لتسريع النقل
app.add_middleware(GZipMiddleware, minimum_size=500)

# ==================== Security Headers (Fast approach) ====================
from starlette.types import ASGIApp, Receive, Scope, Send

class SecurityHeadersMiddleware:
    def __init__(self, app: ASGIApp):
        self.app = app

    async def __call__(self, scope: Scope, receive: Receive, send: Send):
        if scope["type"] != "http":
            await self.app(scope, receive, send)
            return

        async def send_with_headers(message):
            if message["type"] == "http.response.start":
                headers = dict(message.get("headers", []))
                extra = [
                    (b"x-content-type-options", b"nosniff"),
                    (b"x-frame-options", b"DENY"),
                    (b"x-xss-protection", b"1; mode=block"),
                    (b"referrer-policy", b"strict-origin-when-cross-origin"),
                    (b"permissions-policy", b"camera=(), microphone=(), geolocation=()"),
                ]
                message["headers"] = list(message.get("headers", [])) + extra
            await send(message)

        await self.app(scope, receive, send_with_headers)

app.add_middleware(SecurityHeadersMiddleware)

# ==================== Rate Limiting ====================
_login_attempts = {}  # {ip: [(timestamp, success), ...]}
RATE_LIMIT_WINDOW = 300  # 5 دقائق
RATE_LIMIT_MAX_ATTEMPTS = 10  # أقصى عدد محاولات فاشلة

def check_rate_limit(ip: str) -> bool:
    """التحقق من تجاوز حد المحاولات - True = مسموح"""
    now = time_module.time()
    if ip in _login_attempts:
        _login_attempts[ip] = [a for a in _login_attempts[ip] if now - a[0] < RATE_LIMIT_WINDOW]
        failed = [a for a in _login_attempts[ip] if not a[1]]
        if len(failed) >= RATE_LIMIT_MAX_ATTEMPTS:
            return False
    return True

def record_login_attempt(ip: str, success: bool):
    """تسجيل محاولة دخول"""
    if ip not in _login_attempts:
        _login_attempts[ip] = []
    _login_attempts[ip].append((time_module.time(), success))

# CORS - النطاقات المسموحة
# قائمة Whitelist دائمة (مدمجة في الكود) - النطاقات الأساسية للتطبيق
_DEFAULT_ORIGINS = [
    "https://students.ahgaff.net",
    "https://ahgaffstudent-web-production.up.railway.app",
    "https://app.ahgaff.net",
    "https://api.ahgaff.net",
    "http://localhost:3000",
    "http://localhost:8081",
]
ALLOWED_ORIGINS = os.environ.get("ALLOWED_ORIGINS", "")
if ALLOWED_ORIGINS.strip() == "*":
    origins = ["*"]
else:
    # دمج: النطاقات الافتراضية + ما يأتي من env (إن وُجد)
    extra = [o.strip() for o in ALLOWED_ORIGINS.split(",") if o.strip()]
    origins = list(dict.fromkeys(_DEFAULT_ORIGINS + extra))  # dedupe with order

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
    max_age=86400,  # 🚀 24 ساعة - يقلل CORS preflight requests بشكل كبير
)

# ==================== Simple Cache ====================
_cache = {}
CACHE_TTL = 300  # 5 دقائق

def get_cached(key):
    if key in _cache:
        data, ts = _cache[key]
        if time_module.time() - ts < CACHE_TTL:
            return data
        del _cache[key]
    return None

def set_cached(key, data):
    _cache[key] = (data, time_module.time())

def clear_cache(prefix=None):
    if prefix:
        keys_to_del = [k for k in _cache if k.startswith(prefix)]
        for k in keys_to_del:
            del _cache[k]
    else:
        _cache.clear()

# Health check MUST be first - before any middleware
@app.get("/health")
async def root_health_check():
    """Root health check endpoint for Railway deployment"""
    return {"status": "ok"}

@app.get("/api/health")
async def api_health_check_direct():
    """Direct API health check endpoint"""
    return {"status": "ok", "message": "Server is running"}

@app.get("/")
async def root():
    """Root endpoint"""
    return {"status": "ok", "message": "نظام حضور جامعة الأحقاف"}

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ==================== Deploy Guide Public Endpoint ====================
from fastapi.responses import HTMLResponse, PlainTextResponse
import os as _os_for_guide

@api_router.get("/deploy-guide", response_class=HTMLResponse, include_in_schema=False)
async def get_deploy_guide():
    """عرض دليل النشر على Google Cloud (DEPLOY_GCP.md) كصفحة HTML."""
    # ابحث عن الملف في عدة مسارات محتملة
    _here = _os_for_guide.path.dirname(__file__)
    candidates = [
        _os_for_guide.path.join(_here, "_deploy_guide_content.md"),  # ← مضمَّن مع الكود (يعمل في كل البيئات)
        "/app/DEPLOY_GCP.md",
        _os_for_guide.path.join(_here, "..", "..", "DEPLOY_GCP.md"),
        _os_for_guide.path.join(_here, "..", "DEPLOY_GCP.md"),
        "./DEPLOY_GCP.md",
    ]
    content = None
    for p in candidates:
        try:
            if _os_for_guide.path.exists(p):
                with open(p, "r", encoding="utf-8") as f:
                    content = f.read()
                break
        except Exception:
            continue
    if content is None:
        return HTMLResponse(
            "<h1>الدليل غير متوفر</h1><p>لم يتم العثور على ملف DEPLOY_GCP.md</p>",
            status_code=404,
        )

    # تحويل بسيط من Markdown إلى HTML (بدون مكتبات خارجية)
    import html as _html
    escaped = _html.escape(content)
    page = f"""<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>دليل نشر نظام أحقاف على Google Cloud Run</title>
  <style>
    body {{
      font-family: 'Segoe UI', Tahoma, Arial, sans-serif;
      max-width: 920px;
      margin: 0 auto;
      padding: 32px 20px;
      background: #f7f9fc;
      color: #1a1a1a;
      line-height: 1.75;
    }}
    .toolbar {{
      position: sticky; top: 0; background: #1565c0; color: #fff;
      padding: 12px 16px; border-radius: 8px; margin-bottom: 20px;
      display: flex; gap: 12px; align-items: center; box-shadow: 0 2px 8px rgba(0,0,0,0.15);
    }}
    .toolbar h2 {{ margin: 0; flex: 1; font-size: 18px; }}
    .toolbar a, .toolbar button {{
      background: #fff; color: #1565c0; padding: 8px 14px; border-radius: 6px;
      text-decoration: none; font-weight: 600; border: none; cursor: pointer; font-size: 14px;
    }}
    pre {{
      background: #1e1e2e; color: #cdd6f4; padding: 16px;
      border-radius: 8px; overflow-x: auto; direction: ltr; text-align: left;
      font-family: 'Courier New', monospace; font-size: 13px;
    }}
    .container {{
      background: #fff; padding: 32px; border-radius: 12px;
      box-shadow: 0 4px 16px rgba(0,0,0,0.06);
    }}
  </style>
</head>
<body>
  <div class="toolbar">
    <h2>📘 دليل نشر نظام أحقاف على Google Cloud Run</h2>
    <button onclick="copyContent()">📋 نسخ كل الدليل</button>
    <a href="/api/deploy-guide/raw" download="DEPLOY_GCP.md">⬇️ تحميل .md</a>
  </div>
  <div class="container">
    <pre id="content">{escaped}</pre>
  </div>
  <script>
    function copyContent() {{
      const text = document.getElementById('content').innerText;
      navigator.clipboard.writeText(text).then(() => {{
        alert('✅ تم نسخ الدليل بالكامل إلى الحافظة');
      }}).catch(() => {{
        alert('❌ فشل النسخ. حدد النص يدوياً وانسخه.');
      }});
    }}
  </script>
</body>
</html>"""
    return HTMLResponse(content=page)


@api_router.get("/deploy-guide/raw", response_class=PlainTextResponse, include_in_schema=False)
async def get_deploy_guide_raw():
    """تنزيل الدليل كملف Markdown خام."""
    _here = _os_for_guide.path.dirname(__file__)
    candidates = [
        _os_for_guide.path.join(_here, "_deploy_guide_content.md"),
        "/app/DEPLOY_GCP.md",
        _os_for_guide.path.join(_here, "..", "..", "DEPLOY_GCP.md"),
        _os_for_guide.path.join(_here, "..", "DEPLOY_GCP.md"),
        "./DEPLOY_GCP.md",
    ]
    for p in candidates:
        try:
            if _os_for_guide.path.exists(p):
                with open(p, "r", encoding="utf-8") as f:
                    return PlainTextResponse(
                        f.read(),
                        media_type="text/markdown; charset=utf-8",
                        headers={"Content-Disposition": "attachment; filename=DEPLOY_GCP.md"},
                    )
        except Exception:
            continue
    return PlainTextResponse("الدليل غير متوفر", status_code=404)


# ==================== Trash Helper ====================
async def save_to_trash(item_type: str, item_name: str, backup_data: dict, deleted_by: str):
    """حفظ عنصر محذوف في سلة المحذوفات"""
    trash_item = {
        "item_type": item_type,
        "item_name": item_name,
        "backup_data": backup_data,
        "deleted_by": deleted_by,
        "deleted_at": get_yemen_time(),
        "expires_at": get_yemen_time() + timedelta(days=30),
    }
    await db.trash.insert_one(trash_item)

async def restore_from_trash_helper(backup_data: dict):
    """استعادة عنصر من سلة المحذوفات"""
    backup_type = backup_data.get("backup_type")
    
    if backup_type == "course_backup":
        course_data = backup_data.get("course", {})
        course_data.pop("_id", None)
        course_data.pop("id", None)
        result = await db.courses.insert_one(course_data)
        new_id = str(result.inserted_id)
        
        student_id_map = {}
        for s in backup_data.get("students", []):
            old_id = s.pop("_id", None)
            existing = await db.students.find_one({"student_id": s.get("student_id")})
            if existing:
                student_id_map[old_id] = str(existing["_id"])
            else:
                res = await db.students.insert_one(s)
                student_id_map[old_id] = str(res.inserted_id)
        
        for e in backup_data.get("enrollments", []):
            e.pop("_id", None)
            e["course_id"] = new_id
            old_sid = e.get("student_id")
            if old_sid in student_id_map:
                e["student_id"] = student_id_map[old_sid]
            await db.enrollments.insert_one(e)
        
        for l in backup_data.get("lectures", []):
            l.pop("_id", None)
            l["course_id"] = new_id
            await db.lectures.insert_one(l)
        
        for a in backup_data.get("attendance", []):
            a.pop("_id", None)
            a["course_id"] = new_id
            old_sid = a.get("student_id")
            if old_sid in student_id_map:
                a["student_id"] = student_id_map[old_sid]
            await db.attendance.insert_one(a)
        
        return {"message": "تم استعادة المقرر بنجاح", "new_id": new_id}
    
    elif backup_type == "teacher_backup":
        teacher_data = backup_data.get("teacher", {})
        teacher_data.pop("_id", None)
        teacher_data.pop("user_id", None)
        result = await db.teachers.insert_one(teacher_data)
        new_id = str(result.inserted_id)
        
        for c in backup_data.get("teaching_load", []):
            course_code = c.get("code")
            if course_code:
                await db.courses.update_many({"code": course_code, "teacher_id": None}, {"$set": {"teacher_id": new_id}})
        
        return {"message": "تم استعادة المعلم بنجاح", "new_id": new_id}
    
    elif backup_type == "student_backup":
        student_data = backup_data.get("student", {})
        student_data.pop("_id", None)
        student_data.pop("user_id", None)
        
        existing = await db.students.find_one({
            "student_id": student_data.get("student_id"),
            "department_id": student_data.get("department_id"),
        })
        if existing:
            raise HTTPException(status_code=400, detail=f"الطالب برقم {student_data.get('student_id')} موجود بالفعل في نفس القسم")
        
        result = await db.students.insert_one(student_data)
        new_id = str(result.inserted_id)
        
        for e in backup_data.get("enrollments", []):
            e.pop("_id", None)
            e["student_id"] = new_id
            await db.enrollments.insert_one(e)
        
        for a in backup_data.get("attendance", []):
            a.pop("_id", None)
            a["student_id"] = new_id
            await db.attendance.insert_one(a)
        
        return {"message": "تم استعادة الطالب بنجاح", "new_id": new_id}
    
    raise HTTPException(status_code=400, detail="نوع النسخة الاحتياطية غير معروف")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ==================== Helper Functions ====================

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    # Truncate password to 72 bytes for bcrypt compatibility
    password_bytes = password.encode('utf-8')[:72]
    return pwd_context.hash(password_bytes.decode('utf-8'))

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = get_yemen_time() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt


def parse_fields(fields: Optional[str]) -> Optional[set]:
    """يحلل قيمة ?fields=a,b,c إلى set من الأسماء.
    يرجع None إذا لم تُحدَّد (= نُرجع كل الحقول كما هو حالياً).
    "id" دائماً مضمَّن إن طُلب أي حقل."""
    if not fields:
        return None
    parts = {p.strip() for p in fields.split(",") if p.strip()}
    if not parts:
        return None
    parts.add("id")
    return parts


def apply_fields(items: list, allowed: Optional[set]) -> list:
    """يُرجع نسخة من القائمة بحقول مختارة فقط (تخفيف حجم الاستجابة).
    إذا كان allowed=None يُرجع القائمة كما هي."""
    if not allowed:
        return items
    return [{k: v for k, v in (it or {}).items() if k in allowed} for it in items]


# ==================== Activity Logging Function ====================

# ترجمة أنواع الأنشطة إلى العربية
ACTION_TRANSLATIONS = {
    "login": "تسجيل دخول",
    "logout": "تسجيل خروج",
    "password_change": "تغيير كلمة المرور",
    "view_page": "عرض صفحة",
    "view_report": "عرض تقرير",
    "create_user": "إنشاء مستخدم",
    "create_student": "إنشاء طالب",
    "create_course": "إنشاء مقرر",
    "create_department": "إنشاء قسم",
    "create_faculty": "إنشاء كلية",
    "create_lecture": "إنشاء محاضرة",
    "update_user": "تعديل مستخدم",
    "update_student": "تعديل طالب",
    "update_course": "تعديل مقرر",
    "update_department": "تعديل قسم",
    "update_faculty": "تعديل كلية",
    "update_lecture": "تعديل محاضرة",
    "delete_user": "حذف مستخدم",
    "delete_student": "حذف طالب",
    "delete_course": "حذف مقرر",
    "delete_department": "حذف قسم",
    "delete_faculty": "حذف كلية",
    "delete_lecture": "حذف محاضرة",
    "record_attendance": "تسجيل حضور",
    "update_attendance": "تعديل حضور",
    "export_data": "تصدير بيانات",
    "import_data": "استيراد بيانات",
    "create_role": "إنشاء دور",
    "update_role": "تعديل دور",
    "delete_role": "حذف دور",
    "enroll_students": "تسجيل طلاب في مقرر",
    "unenroll_students": "إلغاء تسجيل طلاب",
    "safe_delete_student": "حذف طالب (آمن)",
    "safe_delete_teacher": "حذف معلم (آمن)",
    "safe_delete_course": "حذف مقرر (آمن)",
    "restore_student": "استعادة طالب",
    "restore_teacher": "استعادة معلم",
    "restore_course": "استعادة مقرر",
    "override_lecture_status": "تغيير حالة محاضرة",
}

async def log_activity(
    user: dict,
    action: str,
    entity_type: str = None,
    entity_id: str = None,
    entity_name: str = None,
    details: dict = None,
    ip_address: str = None,
    user_agent: str = None
):
    """تسجيل نشاط المستخدم في قاعدة البيانات"""
    try:
        action_ar = ACTION_TRANSLATIONS.get(action, action)
        
        log_entry = {
            "user_id": str(user.get("_id", user.get("id", ""))),
            "username": user.get("username", "غير معروف"),
            "user_role": user.get("role", "غير محدد"),
            "action": action,
            "action_ar": action_ar,
            "entity_type": entity_type,
            "entity_id": entity_id,
            "entity_name": entity_name,
            "details": details,
            "ip_address": ip_address,
            "user_agent": user_agent,
            "faculty_id": user.get("faculty_id"),
            "department_id": user.get("department_id"),
            "timestamp": get_yemen_time()
        }
        
        await db.activity_logs.insert_one(log_entry)
        logger.info(f"Activity logged: {action} by {user.get('username')}")
    except Exception as e:
        logger.error(f"Failed to log activity: {e}")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="بيانات الاعتماد غير صالحة",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if user is None:
        raise credentials_exception
    
    # احصل على صلاحيات المستخدم - من الدور المخصص أو من جدول الأدوار أو الافتراضية
    user_permissions = []
    
    if user.get("role_id"):
        try:
            role_doc = await db.roles.find_one({"_id": ObjectId(user["role_id"])})
            if role_doc:
                user_permissions = list(role_doc.get("permissions", []))
        except Exception:
            pass
    
    if not user_permissions:
        user_role = user.get("role", "employee")
        role_doc = await db.roles.find_one({"system_key": user_role})
        if role_doc:
            user_permissions = list(role_doc.get("permissions", []))
        else:
            user_permissions = list(DEFAULT_PERMISSIONS.get(user_role, []))
    
    custom_permissions = user.get("custom_permissions") or []
    # دمج الصلاحيات المخصصة المباشرة أيضاً (من إدارة المستخدمين)
    direct_permissions = user.get("permissions") or []
    all_extra = list(set(custom_permissions) | set(direct_permissions))
    for perm in all_extra:
        if perm not in user_permissions:
            user_permissions.append(perm)
    
    # توسيع الصلاحيات: manage_lectures → reschedule_lecture, generate_lectures, إلخ
    expanded = set(user_permissions)
    for perm in user_permissions:
        if perm in FULL_PERMISSION_MAPPING:
            expanded.update(FULL_PERMISSION_MAPPING[perm])
    user_permissions = list(expanded)
    
    return {
        "id": str(user["_id"]),
        "username": user["username"],
        "full_name": user["full_name"],
        "role": user["role"],
        "email": user.get("email"),
        "phone": user.get("phone"),
        "is_active": user.get("is_active", True),
        "permissions": user_permissions,
        # 🔒 ضرورية لفحص ملكية القسم/الكلية في endpoints الحماية
        "faculty_id": str(user["faculty_id"]) if user.get("faculty_id") else None,
        "department_id": str(user["department_id"]) if user.get("department_id") else None,
    }

def get_user_permissions(user: dict) -> List[str]:
    """الحصول على صلاحيات المستخدم مع توسيع الصلاحيات الرئيسية"""
    base_permissions = user.get("permissions") or DEFAULT_PERMISSIONS.get(user["role"], [])
    # توسيع الصلاحيات: manage_lectures يشمل reschedule_lecture, generate_lectures, إلخ
    expanded = set(base_permissions)
    for perm in base_permissions:
        if perm in FULL_PERMISSION_MAPPING:
            expanded.update(FULL_PERMISSION_MAPPING[perm])
    return list(expanded)

TEACHER_ONLY_PERMISSIONS = ["record_attendance", "take_attendance", "edit_attendance"]

def has_permission(user: dict, permission: str) -> bool:
    """التحقق من أن المستخدم لديه صلاحية معينة"""
    if user["role"] == UserRole.ADMIN:
        # المدير لا يحصل تلقائياً على صلاحيات التحضير
        if permission in TEACHER_ONLY_PERMISSIONS:
            return permission in user.get("permissions", [])
        return True
    permissions = get_user_permissions(user)
    return permission in permissions

def require_permission(permission: str):
    """Dependency للتحقق من الصلاحيات"""
    async def check_permission(current_user: dict = Depends(get_current_user)):
        if not has_permission(current_user, permission):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"ليس لديك صلاحية: {permission}"
            )
        return current_user
    return check_permission

def generate_qr_code(student_id: str) -> str:
    """Generate unique QR code for student"""
    return f"SHARIA-{student_id}-{uuid.uuid4().hex[:8].upper()}"

async def get_active_lecture_ids(course_id: str = None) -> set:
    """
    جلب IDs المحاضرات الفعّالة فقط (غير الملغاة)
    المحاضرات المجدولة والمنعقدة تُحسب، الملغاة لا تُحسب
    """
    query = {"status": {"$in": ACTIVE_LECTURE_STATUSES}}
    if course_id:
        query["course_id"] = course_id
    
    lectures = await db.lectures.find(query, {"_id": 1}).to_list(10000)
    return {str(lec["_id"]) for lec in lectures}

async def filter_attendance_by_active_lectures(records: list, course_id: str = None) -> list:
    """
    فلترة سجلات الحضور لتشمل فقط المحاضرات الفعّالة (غير الملغاة)
    المحاضرات التعويضية تُحسب تلقائياً لأنها ليست ملغاة
    """
    active_lecture_ids = await get_active_lecture_ids(course_id)
    return [r for r in records if r.get("lecture_id") in active_lecture_ids]

# ==================== Backend Scoping Helper ====================

async def get_user_scope_filter(current_user: dict, scope_type: str = "students") -> dict:
    """
    إرجاع فلتر Query بناءً على دور المستخدم ونطاقه
    
    الأدوار ونطاقاتها:
    - admin: لا يوجد فلتر (يرى كل شيء)
    - dean: يرى فقط بيانات كليته (faculty_id)
    - department_head: يرى فقط بيانات قسمه (department_id)
    - teacher: يرى فقط بيانات مقرراته
    - registrar/registration_manager: يرى فقط بيانات كليته
    
    scope_type: "students", "courses", "departments", "teachers", "attendance"
    """
    query = {}
    role = current_user.get("role", "")
    user_id = current_user.get("id")
    
    # Admin يرى كل شيء
    if role == UserRole.ADMIN:
        return query
    
    # جلب بيانات المستخدم الكاملة للحصول على faculty_id و department_id
    user_data = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user_data:
        return query
    
    faculty_id = user_data.get("faculty_id")
    department_id = user_data.get("department_id")
    # دعم الأقسام المتعددة
    department_ids = user_data.get("department_ids", [])
    if not department_ids and department_id:
        department_ids = [department_id]
    teacher_record_id = user_data.get("teacher_record_id")
    
    # Dean (عميد) - يرى بيانات كليته
    if role == "dean" and faculty_id:
        if scope_type == "students":
            # جلب أقسام الكلية أولاً
            dept_ids = await get_faculty_department_ids(faculty_id)
            if dept_ids:
                query["department_id"] = {"$in": dept_ids}
        elif scope_type == "courses":
            dept_ids = await get_faculty_department_ids(faculty_id)
            if dept_ids:
                query["department_id"] = {"$in": dept_ids}
        elif scope_type == "departments":
            query["faculty_id"] = faculty_id
        elif scope_type == "teachers":
            dept_ids = await get_faculty_department_ids(faculty_id)
            if dept_ids:
                query["department_id"] = {"$in": dept_ids}
    
    # Department Head (رئيس قسم) - يرى بيانات أقسامه (قسم واحد أو أقسام متعددة)
    elif role == "department_head":
        if department_ids and len(department_ids) > 0:
            if scope_type in ["students", "courses", "teachers"]:
                if len(department_ids) == 1:
                    query["department_id"] = department_ids[0]
                else:
                    query["department_id"] = {"$in": department_ids}
            elif scope_type == "departments":
                if len(department_ids) == 1:
                    query["_id"] = ObjectId(department_ids[0])
                else:
                    query["_id"] = {"$in": [ObjectId(did) for did in department_ids]}
        elif department_id:
            if scope_type in ["students", "courses", "teachers"]:
                query["department_id"] = department_id
            elif scope_type == "departments":
                query["_id"] = ObjectId(department_id)
    
    # Custom role with faculty (no department) - يرى بيانات الكلية كلها
    elif role == "custom" and faculty_id and not department_id:
        if scope_type == "students":
            dept_ids = await get_faculty_department_ids(faculty_id)
            if dept_ids:
                query["department_id"] = {"$in": dept_ids}
        elif scope_type == "courses":
            dept_ids = await get_faculty_department_ids(faculty_id)
            if dept_ids:
                query["department_id"] = {"$in": dept_ids}
        elif scope_type == "departments":
            query["faculty_id"] = faculty_id
        elif scope_type == "teachers":
            dept_ids = await get_faculty_department_ids(faculty_id)
            if dept_ids:
                query["department_id"] = {"$in": dept_ids}
    
    # Custom role with department(s) - يرى بيانات أقسامه (قسم واحد أو أقسام متعددة)
    elif role == "custom" and (department_ids or department_id):
        # استخدام الأقسام المتعددة إن وُجدت
        user_dept_ids = department_ids if department_ids else [department_id]
        if scope_type in ["students", "courses", "teachers"]:
            if len(user_dept_ids) == 1:
                query["department_id"] = user_dept_ids[0]
            else:
                query["department_id"] = {"$in": user_dept_ids}
        elif scope_type == "departments":
            if len(user_dept_ids) == 1:
                query["_id"] = ObjectId(user_dept_ids[0])
            else:
                query["_id"] = {"$in": [ObjectId(did) for did in user_dept_ids]}
    
    # Registration Manager / Registrar - نطاق مرن:
    # (1) إن كانت له أقسام محددة (department_ids) → يرى تلك الأقسام فقط
    # (2) إن كانت له كلية فقط (faculty_id) → يرى كل أقسام الكلية
    # (3) إن كان له كل من faculty_id و department_ids → أولوية للأقسام (أكثر تحديداً)
    elif role in ["registration_manager", "registrar"]:
        # 🔧 الحالة (1): أقسام محددة → استخدمها مباشرة
        if department_ids and len(department_ids) > 0:
            if scope_type in ["students", "courses", "teachers"]:
                if len(department_ids) == 1:
                    query["department_id"] = department_ids[0]
                else:
                    query["department_id"] = {"$in": department_ids}
            elif scope_type == "departments":
                if len(department_ids) == 1:
                    query["_id"] = ObjectId(department_ids[0])
                else:
                    query["_id"] = {"$in": [ObjectId(did) for did in department_ids]}
        # الحالة (2): كلية كاملة بدون تحديد أقسام
        elif faculty_id:
            if scope_type == "students":
                dept_ids = await get_faculty_department_ids(faculty_id)
                if dept_ids:
                    query["department_id"] = {"$in": dept_ids}
            elif scope_type == "courses":
                dept_ids = await get_faculty_department_ids(faculty_id)
                if dept_ids:
                    query["department_id"] = {"$in": dept_ids}
            elif scope_type == "departments":
                query["faculty_id"] = faculty_id
            elif scope_type == "teachers":
                dept_ids = await get_faculty_department_ids(faculty_id)
                if dept_ids:
                    query["department_id"] = {"$in": dept_ids}
    
    # Teacher - يرى فقط مقرراته
    elif role == UserRole.TEACHER:
        if scope_type == "courses":
            # استخدام teacher_record_id أو user_id
            if teacher_record_id:
                query["teacher_id"] = teacher_record_id
            else:
                query["teacher_id"] = user_id
        elif scope_type == "students":
            # جلب طلاب مقرراته فقط
            course_ids = await get_teacher_course_ids(user_id, teacher_record_id)
            if course_ids:
                # جلب طلاب هذه المقررات
                enrolled_students = await db.enrollments.find(
                    {"course_id": {"$in": course_ids}}
                ).distinct("student_id")
                if enrolled_students:
                    query["_id"] = {"$in": [ObjectId(sid) for sid in enrolled_students]}
                else:
                    query["_id"] = {"$in": []}  # لا يوجد طلاب
    
    # Employee (موظف) — يخضع لنفس قواعد scope حسب ما هو مُسنَد له:
    # إذا كان له قسم → يرى قسمه/أقسامه فقط (مثل department_head)
    # إذا كان له كلية فقط → يرى الكلية كلها (مثل dean بحدود الكلية)
    elif role == "employee":
        user_dept_ids = department_ids if department_ids else ([department_id] if department_id else [])
        if user_dept_ids:
            if scope_type in ["students", "courses", "teachers"]:
                if len(user_dept_ids) == 1:
                    query["department_id"] = user_dept_ids[0]
                else:
                    query["department_id"] = {"$in": user_dept_ids}
            elif scope_type == "departments":
                if len(user_dept_ids) == 1:
                    query["_id"] = ObjectId(user_dept_ids[0])
                else:
                    query["_id"] = {"$in": [ObjectId(did) for did in user_dept_ids]}
        elif faculty_id:
            # موظف الكلية: يرى كل أقسامها
            if scope_type in ["students", "courses", "teachers"]:
                dept_ids = await get_faculty_department_ids(faculty_id)
                if dept_ids:
                    query["department_id"] = {"$in": dept_ids}
            elif scope_type == "departments":
                query["faculty_id"] = faculty_id

    # 🔒 Fail-safe أمني حرج: إذا كان للدور نطاق محدود لكن لم يُطبَّق أي فلتر
    # (مثلاً employee/department_head بلا قسم مسنَد) → نمنع رؤية أي شيء بدلاً من إظهار الكل
    SCOPED_ROLES = {"dean", "department_head", "registrar", "registration_manager",
                    "custom", "employee", UserRole.TEACHER}
    if role in SCOPED_ROLES and not query:
        logging.warning(
            f"RBAC: user '{user_data.get('username')}' (role={role}) has no scope data "
            f"(faculty_id={faculty_id}, department_ids={department_ids}) — blocking all results."
        )
        # فلتر مستحيل التحقق: ObjectId غير موجود إطلاقاً
        query["_id"] = ObjectId("000000000000000000000000")
    
    return query

async def get_faculty_department_ids(faculty_id: str) -> list:
    """جلب معرفات الأقسام التابعة لكلية معينة"""
    departments = await db.departments.find(
        {"faculty_id": faculty_id},
        {"_id": 1}
    ).to_list(100)
    return [str(d["_id"]) for d in departments]

async def get_teacher_course_ids(user_id: str, teacher_record_id: str = None) -> list:
    """جلب معرفات المقررات التي يدرسها المعلم"""
    query = {"is_active": True}
    if teacher_record_id:
        query["teacher_id"] = teacher_record_id
    else:
        query["teacher_id"] = user_id
    
    courses = await db.courses.find(query, {"_id": 1}).to_list(100)
    return [str(c["_id"]) for c in courses]

# تسجيل دالة فلتر النطاق في deps لاستخدامها من ملفات الـ routes الأخرى
try:
    from routes.deps import set_scope_filter as _set_scope_filter
    _set_scope_filter(get_user_scope_filter)
except Exception as _e:
    logging.warning(f"Could not register scope_filter to deps: {_e}")

# ==================== Auth Routes ====================
# تم نقل هذه الـ routes إلى routes/auth.py
# الـ endpoints التالية متاحة عبر auth_router:
# - POST /api/auth/login
# - GET /api/auth/me

# ==================== User Management Routes ====================

@api_router.post("/users", response_model=UserResponse)
async def create_user(user: UserCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_users"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بإنشاء مستخدمين")
    
    existing = await db.users.find_one({"username": user.username})
    if existing:
        raise HTTPException(status_code=400, detail="اسم المستخدم موجود مسبقاً")
    
    user_dict = user.dict()
    user_dict["password"] = get_password_hash(user.password)
    user_dict["created_at"] = get_yemen_time()
    user_dict["is_active"] = True
    
    # التأكد من أن permissions قائمة وليست None
    if user_dict.get("permissions") is None:
        user_dict["permissions"] = []
    
    # التعامل مع الدور الجديد (role_id)
    if user.role_id:
        role = await db.roles.find_one({"_id": ObjectId(user.role_id)})
        if role:
            system_key = role.get("system_key") or ""
            # 🔧 إذا كان الدور المخصّص بلا system_key، نحاول مطابقته باسم الدور
            # مع دور نظام أصلي (مثلاً: "رئيس قسم" → department_head)
            if not system_key:
                ROLE_NAME_MAP = {
                    "رئيس قسم": "department_head",
                    "عميد كلية": "dean",
                    "عميد": "dean",
                    "مسجل": "registrar",
                    "مدير التسجيل": "registration_manager",
                    "موظف": "employee",
                    "مدير النظام": "admin",
                }
                role_name = (role.get("name") or "").strip()
                if role_name in ROLE_NAME_MAP:
                    system_key = ROLE_NAME_MAP[role_name]
                else:
                    system_key = "custom"
            # 🔒 منع قطعي: لا يُسمح بإنشاء حسابات teacher/student من هذه الشاشة
            if system_key in ["teacher", "student"]:
                raise HTTPException(
                    status_code=400,
                    detail=f"لا يمكن إنشاء حساب بدور '{system_key}' من شاشة المستخدمين. استخدم شاشة 'إدارة {'المعلمين' if system_key == 'teacher' else 'الطلاب'}' المخصصة."
                )
            user_dict["role_id"] = user.role_id
            user_dict["role"] = system_key
            user_dict["permissions"] = role.get("permissions", [])

    # 🔒 تحقق إلزامي: الأدوار محدودة النطاق يجب أن يكون لها نطاق مُسنَد (يمنع سيناريو zain)
    _final_role = user_dict.get("role", "")
    _has_dept = bool(user_dict.get("department_id") or user_dict.get("department_ids"))
    _has_faculty = bool(user_dict.get("faculty_id"))
    if _final_role == "department_head" and not _has_dept:
        raise HTTPException(
            status_code=400,
            detail="رئيس القسم يجب أن يكون له قسم مُسنَد. اختر قسماً واحداً على الأقل."
        )
    if _final_role in ["dean", "registrar", "registration_manager"] and not _has_faculty:
        role_ar = {"dean": "العميد", "registrar": "المسجِّل", "registration_manager": "مدير التسجيل"}[_final_role]
        raise HTTPException(
            status_code=400,
            detail=f"{role_ar} يجب أن يكون له كلية مُسنَدة. اختر الكلية أولاً."
        )
    elif user.role in ["teacher", "student"]:
        # نفس الحماية لو مُرّر role مباشرة
        raise HTTPException(
            status_code=400,
            detail=f"لا يمكن إنشاء حساب بدور '{user.role}' من شاشة المستخدمين. استخدم شاشة 'إدارة {'المعلمين' if user.role == 'teacher' else 'الطلاب'}' المخصصة."
        )
    elif not user.role:
        user_dict["role"] = "employee"  # افتراضي
        user_dict["permissions"] = []
    
    # دعم الأقسام المتعددة - ضبط department_id للتوافق
    if user_dict.get("department_ids") and len(user_dict["department_ids"]) > 0:
        user_dict["department_id"] = user_dict["department_ids"][0]
    
    result = await db.users.insert_one(user_dict)
    
    # إرجاع البيانات بالشكل المطلوب
    return {
        "id": str(result.inserted_id),
        "username": user_dict["username"],
        "full_name": user_dict["full_name"],
        "role": user_dict["role"],
        "role_id": user_dict.get("role_id"),
        "email": user_dict.get("email"),
        "phone": user_dict.get("phone"),
        "created_at": user_dict["created_at"],
        "is_active": user_dict.get("is_active", True),
        "permissions": user_dict.get("permissions", []),
        "university_id": user_dict.get("university_id"),
        "faculty_id": user_dict.get("faculty_id"),
        "department_id": user_dict.get("department_id"),
        "department_ids": user_dict.get("department_ids", []),
    }

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(role: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_users"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    query = {}
    if role:
        query["role"] = role
    else:
        # 🔧 شاشة "إدارة المستخدمين" لا تعرض الطلاب/المعلمين (لديها شاشات مستقلة)
        # استبعادهم في DB يمنع امتلاء الـ limit بهم وإخفاء الحسابات الإدارية
        query["role"] = {"$nin": ["student", "teacher"]}
    
    # رفع الحد الأقصى لاستيعاب أكبر قاعدة بيانات (الإداريون قليلون عادة، الـ limit للحماية فقط)
    users = await db.users.find(query).to_list(10000)
    
    # جلب أسماء الكليات والأقسام للعرض
    result = []
    for u in users:
        user_data = {
            "id": str(u["_id"]),
            "username": u["username"],
            "full_name": u["full_name"],
            "role": u["role"],
            "role_id": u.get("role_id"),
            "email": u.get("email"),
            "phone": u.get("phone"),
            "created_at": u["created_at"],
            "is_active": u.get("is_active", True),
            "permissions": u.get("permissions") or DEFAULT_PERMISSIONS.get(u["role"], []),
            "university_id": u.get("university_id"),
            "faculty_id": u.get("faculty_id"),
            "department_id": u.get("department_id"),
            "department_ids": u.get("department_ids", []),
        }
        
        # جلب اسم الكلية إذا موجود
        if u.get("faculty_id"):
            try:
                faculty = await db.faculties.find_one({"_id": ObjectId(u["faculty_id"])})
                user_data["faculty_name"] = faculty["name"] if faculty else None
            except:
                user_data["faculty_name"] = None
        
        # جلب أسماء الأقسام - دعم الأقسام المتعددة
        dept_ids_to_fetch = u.get("department_ids") or []
        if not dept_ids_to_fetch and u.get("department_id"):
            dept_ids_to_fetch = [u.get("department_id")]
        
        if dept_ids_to_fetch and len(dept_ids_to_fetch) > 0:
            try:
                dept_names = []
                for did in dept_ids_to_fetch:
                    if did:  # التأكد من أن المعرف ليس None أو فارغاً
                        dept = await db.departments.find_one({"_id": ObjectId(did)})
                        if dept:
                            dept_names.append(dept["name"])
                user_data["department_name"] = " | ".join(dept_names) if dept_names else None
                user_data["department_names"] = dept_names
            except Exception as e:
                logger.error(f"Error fetching department names: {e}")
                user_data["department_name"] = None
                user_data["department_names"] = []
        else:
            user_data["department_name"] = None
            user_data["department_names"] = []
        
        result.append(user_data)
    
    return result

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_users"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    result = await db.users.delete_one({"_id": ObjectId(user_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    return {"message": "تم حذف المستخدم بنجاح"}


# ═══════════════════════════════════════════════════════════════
# 🛠️ Admin Tools: إصلاح/حذف مستخدم بالـ username (للحالات الطارئة)
# يُستخدم عندما لا يظهر المستخدم في القائمة لخطأ في role
# ═══════════════════════════════════════════════════════════════
@api_router.get("/admin/find-user-by-username/{username}")
async def admin_find_user_by_username(username: str, current_user: dict = Depends(get_current_user)):
    """جلب مستخدم بـ username (للتشخيص). admin فقط."""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح — admin فقط")
    u = await db.users.find_one({"username": {"$regex": f"^{username}$", "$options": "i"}})
    if not u:
        raise HTTPException(status_code=404, detail=f"المستخدم '{username}' غير موجود في قاعدة البيانات")
    return {
        "id": str(u["_id"]),
        "username": u.get("username"),
        "full_name": u.get("full_name"),
        "role": u.get("role"),
        "role_id": u.get("role_id"),
        "faculty_id": u.get("faculty_id"),
        "department_id": u.get("department_id"),
        "is_active": u.get("is_active", True),
        "created_at": u.get("created_at"),
    }


class AdminFixUserRequest(BaseModel):
    username: str
    action: str  # "delete" | "fix-role"
    new_role: Optional[str] = None  # مطلوب لو action=fix-role


@api_router.post("/admin/fix-user")
async def admin_fix_user(req: AdminFixUserRequest, current_user: dict = Depends(get_current_user)):
    """حذف أو إصلاح role لمستخدم لا يظهر في الواجهة. admin فقط."""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح — admin فقط")

    u = await db.users.find_one({"username": {"$regex": f"^{req.username}$", "$options": "i"}})
    if not u:
        raise HTTPException(status_code=404, detail=f"المستخدم '{req.username}' غير موجود")

    if req.action == "delete":
        await db.users.delete_one({"_id": u["_id"]})
        return {"message": f"✅ تم حذف المستخدم '{u['username']}' بنجاح", "deleted_id": str(u["_id"])}

    if req.action == "fix-role":
        if not req.new_role:
            raise HTTPException(status_code=400, detail="new_role مطلوب لتعديل الدور")
        valid_roles = ["admin", "dean", "department_head", "registrar", "registration_manager", "employee"]
        if req.new_role not in valid_roles:
            raise HTTPException(status_code=400, detail=f"دور غير مدعوم. المسموح: {', '.join(valid_roles)}")
        await db.users.update_one({"_id": u["_id"]}, {"$set": {"role": req.new_role}})
        return {"message": f"✅ تم تعديل دور '{u['username']}' إلى '{req.new_role}'", "user_id": str(u["_id"])}

    raise HTTPException(status_code=400, detail="action غير صالح — استخدم 'delete' أو 'fix-role'")

class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    password: Optional[str] = None
    role_id: Optional[str] = None  # معرف الدور الجديد
    # حقول النطاق
    university_id: Optional[str] = None
    faculty_id: Optional[str] = None
    department_id: Optional[str] = None  # للتوافق مع القديم
    department_ids: Optional[List[str]] = None  # قائمة الأقسام المتعددة

@api_router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(user_id: str, data: UserUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_users"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # حماية مدير النظام - لا يمكن تغيير دوره أو صلاحياته
    if user.get("role") == UserRole.ADMIN or user.get("username") == "admin":
        # لا يمكن تغيير الدور أو الصلاحيات للمدير
        if data.role_id:
            raise HTTPException(status_code=403, detail="لا يمكن تغيير دور مدير النظام")
    
    update_data = {}
    if data.full_name:
        update_data["full_name"] = data.full_name
    if data.email:
        update_data["email"] = data.email
    if data.phone:
        update_data["phone"] = data.phone
    if data.password:
        update_data["password"] = get_password_hash(data.password)
    
    # تحديث حقول النطاق
    if data.university_id is not None:
        update_data["university_id"] = data.university_id if data.university_id else None
    if data.faculty_id is not None:
        update_data["faculty_id"] = data.faculty_id if data.faculty_id else None
    if data.department_id is not None:
        update_data["department_id"] = data.department_id if data.department_id else None
    
    # دعم الأقسام المتعددة
    if data.department_ids is not None:
        # تخزين قائمة الأقسام
        update_data["department_ids"] = data.department_ids if data.department_ids else []
        # إذا كانت هناك أقسام متعددة، نضع أول قسم كـ department_id للتوافق
        if data.department_ids and len(data.department_ids) > 0:
            update_data["department_id"] = data.department_ids[0]
        else:
            update_data["department_id"] = None
    
    # تحديث الدور إذا تم تحديده (فقط لغير المدير)
    if data.role_id and user.get("role") != UserRole.ADMIN:
        role = await db.roles.find_one({"_id": ObjectId(data.role_id)})
        if role:
            # 🔧 استنتاج system_key الصحيح من system_key أو من الاسم العربي
            # لمنع تخزين role="custom" الذي يُخفي المستخدم من فلتر القائمة
            sys_key = (role.get("system_key") or "").strip()
            if not sys_key:
                ROLE_NAME_MAP = {
                    "رئيس قسم": "department_head", "عميد كلية": "dean", "عميد": "dean",
                    "مسجل": "registrar", "مدير التسجيل": "registration_manager",
                    "موظف": "employee", "مدير النظام": "admin",
                    "Department Head": "department_head", "Dean": "dean",
                    "Registrar": "registrar", "Employee": "employee",
                }
                sys_key = ROLE_NAME_MAP.get((role.get("name") or "").strip(), "custom")
            update_data["role_id"] = data.role_id
            update_data["role"] = sys_key
            update_data["permissions"] = role.get("permissions", [])
    
    # 🔒 تحقق إلزامي: لا يُسمح بحفظ مستخدم محدود النطاق بدون scope data (يمنع سيناريو zain)
    # نحسب الحالة النهائية بعد التحديث
    _final_role = update_data.get("role", user.get("role", ""))
    _final_dept_ids = update_data.get("department_ids", user.get("department_ids", []))
    _final_dept_id = update_data.get("department_id", user.get("department_id"))
    _final_faculty_id = update_data.get("faculty_id", user.get("faculty_id"))
    _has_dept = bool(_final_dept_ids) or bool(_final_dept_id)
    _has_faculty = bool(_final_faculty_id)
    if _final_role == "department_head" and not _has_dept:
        raise HTTPException(
            status_code=400,
            detail="رئيس القسم يجب أن يكون له قسم مُسنَد. اختر قسماً واحداً على الأقل."
        )
    if _final_role in ["dean", "registrar", "registration_manager"] and not _has_faculty:
        role_ar = {"dean": "العميد", "registrar": "المسجِّل", "registration_manager": "مدير التسجيل"}[_final_role]
        raise HTTPException(
            status_code=400,
            detail=f"{role_ar} يجب أن يكون له كلية مُسنَدة. اختر الكلية أولاً."
        )
    
    if update_data:
        await db.users.update_one(
            {"_id": ObjectId(user_id)},
            {"$set": update_data}
        )
    
    updated = await db.users.find_one({"_id": ObjectId(user_id)})
    
    # إرجاع صلاحيات كاملة للمدير
    user_permissions = updated.get("permissions", DEFAULT_PERMISSIONS.get(updated["role"], []))
    if updated["role"] == UserRole.ADMIN:
        user_permissions = list(DEFAULT_PERMISSIONS.get(UserRole.ADMIN, []))
    
    # توسيع الصلاحيات
    expanded = set(user_permissions)
    for perm in user_permissions:
        if perm in FULL_PERMISSION_MAPPING:
            expanded.update(FULL_PERMISSION_MAPPING[perm])
    user_permissions = list(expanded)
    
    return {
        "id": str(updated["_id"]),
        "username": updated["username"],
        "full_name": updated["full_name"],
        "role": updated["role"],
        "role_id": updated.get("role_id"),
        "email": updated.get("email"),
        "phone": updated.get("phone"),
        "created_at": updated["created_at"],
        "is_active": updated.get("is_active", True),
        "permissions": user_permissions,
        "university_id": updated.get("university_id"),
        "faculty_id": updated.get("faculty_id"),
        "department_id": updated.get("department_id"),
        "department_ids": updated.get("department_ids", [])
    }

# ==================== User Management Actions ====================

@api_router.post("/users/{user_id}/reset-password")
async def reset_user_password(
    user_id: str,
    new_password: str = Body(..., embed=True),
    current_user: dict = Depends(get_current_user)
):
    """إعادة تعيين كلمة مرور المستخدم"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_users"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بإعادة تعيين كلمة المرور")
    
    # التحقق من وجود المستخدم
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # تشفير كلمة المرور الجديدة
    hashed_password = pwd_context.hash(new_password)
    
    # تحديث كلمة المرور
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"password": hashed_password, "updated_at": get_yemen_time()}}
    )
    
    # تسجيل النشاط
    await log_activity(
        user=current_user,
        action="reset_password",
        entity_type="user",
        entity_id=user_id,
        entity_name=user.get('username')
    )
    
    return {"message": "تم إعادة تعيين كلمة المرور بنجاح"}

@api_router.post("/users/{user_id}/toggle-active")
async def toggle_user_active(
    user_id: str,
    current_user: dict = Depends(get_current_user)
):
    """تفعيل/إيقاف المستخدم"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_users"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل حالة المستخدم")
    
    # التحقق من وجود المستخدم
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # لا يمكن إيقاف حساب المدير الرئيسي
    if user.get("role") == "admin" and user.get("username") == "admin":
        raise HTTPException(status_code=400, detail="لا يمكن إيقاف حساب مدير النظام الرئيسي")
    
    # تبديل الحالة
    new_status = not user.get("is_active", True)
    
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"is_active": new_status, "updated_at": get_yemen_time()}}
    )
    
    # تسجيل النشاط
    action = "activate_user" if new_status else "deactivate_user"
    await log_activity(
        user=current_user,
        action=action,
        entity_type="user",
        entity_id=user_id,
        entity_name=user.get('username')
    )
    
    return {"message": f"تم {'تفعيل' if new_status else 'إيقاف'} المستخدم بنجاح", "is_active": new_status}

# ==================== Roles Management Routes (إدارة الأدوار) ====================

@api_router.get("/roles")
async def get_all_roles(current_user: dict = Depends(get_current_user)):
    """الحصول على جميع الأدوار"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    roles = await db.roles.find().to_list(100)
    
    result = []
    for role in roles:
        # حساب عدد المستخدمين لكل دور (بواسطة role_id أو role القديم)
        system_key = role.get("system_key", "")
        users_count = await db.users.count_documents({
            "$or": [
                {"role_id": str(role["_id"])},
                {"role": system_key} if system_key else {"_id": None}
            ]
        })
        result.append({
            "id": str(role["_id"]),
            "name": role["name"],
            "description": role.get("description", ""),
            "permissions": role.get("permissions", []),
            "is_system": role.get("is_system", False),
            "system_key": system_key,
            "users_count": users_count,
            "created_at": role.get("created_at", get_yemen_time())
        })
    
    return result

@api_router.post("/roles")
async def create_role(role_data: RoleCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء دور جديد"""
    if current_user["role"] != "admin" and not has_permission(current_user, "manage_roles"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # التحقق من عدم وجود دور بنفس الاسم
    existing = await db.roles.find_one({"name": role_data.name})
    if existing:
        raise HTTPException(status_code=400, detail="يوجد دور بهذا الاسم")
    
    # التحقق من صحة الصلاحيات
    valid_permissions = [p["key"] for p in ALL_PERMISSIONS]
    for perm in role_data.permissions:
        if perm not in valid_permissions:
            raise HTTPException(status_code=400, detail=f"صلاحية غير صالحة: {perm}")
    
    role_doc = {
        "name": role_data.name,
        "description": role_data.description or "",
        "permissions": role_data.permissions,
        "is_system": False,
        "created_at": get_yemen_time(),
        "created_by": current_user["id"]
    }
    
    result = await db.roles.insert_one(role_doc)
    
    return {
        "id": str(result.inserted_id),
        "message": "تم إنشاء الدور بنجاح"
    }

@api_router.get("/roles/{role_id}")
async def get_role(role_id: str, current_user: dict = Depends(get_current_user)):
    """الحصول على تفاصيل دور معين"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    role = await db.roles.find_one({"_id": ObjectId(role_id)})
    if not role:
        raise HTTPException(status_code=404, detail="الدور غير موجود")
    
    users_count = await db.users.count_documents({"role_id": role_id})
    
    return {
        "id": str(role["_id"]),
        "name": role["name"],
        "description": role.get("description", ""),
        "permissions": role.get("permissions", []),
        "is_system": role.get("is_system", False),
        "users_count": users_count,
        "created_at": role.get("created_at", get_yemen_time())
    }

@api_router.put("/roles/{role_id}")
async def update_role(role_id: str, role_data: RoleUpdate, current_user: dict = Depends(get_current_user)):
    """تحديث دور"""
    if current_user["role"] != "admin" and not has_permission(current_user, "manage_roles"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    role = await db.roles.find_one({"_id": ObjectId(role_id)})
    if not role:
        raise HTTPException(status_code=404, detail="الدور غير موجود")
    
    update_data = {}
    
    # للأدوار النظامية: يمكن تعديل الصلاحيات فقط، لا يمكن تغيير الاسم
    if role.get("is_system"):
        if role_data.name and role_data.name != role.get("name"):
            raise HTTPException(status_code=400, detail="لا يمكن تغيير اسم دور نظامي")
        # السماح بتعديل الصلاحيات
        if role_data.permissions is not None:
            valid_permissions = [p["key"] for p in ALL_PERMISSIONS]
            for perm in role_data.permissions:
                if perm not in valid_permissions:
                    raise HTTPException(status_code=400, detail=f"صلاحية غير صالحة: {perm}")
            update_data["permissions"] = role_data.permissions
        if role_data.description is not None:
            update_data["description"] = role_data.description
    else:
        # للأدوار غير النظامية: يمكن تعديل كل شيء
        if role_data.name:
            # التحقق من عدم وجود دور آخر بنفس الاسم
            existing = await db.roles.find_one({"name": role_data.name, "_id": {"$ne": ObjectId(role_id)}})
            if existing:
                raise HTTPException(status_code=400, detail="يوجد دور آخر بهذا الاسم")
            update_data["name"] = role_data.name
        
        if role_data.description is not None:
            update_data["description"] = role_data.description
        
        if role_data.permissions is not None:
            valid_permissions = [p["key"] for p in ALL_PERMISSIONS]
            for perm in role_data.permissions:
                if perm not in valid_permissions:
                    raise HTTPException(status_code=400, detail=f"صلاحية غير صالحة: {perm}")
            update_data["permissions"] = role_data.permissions
    
    if update_data:
        update_data["updated_at"] = get_yemen_time()
        await db.roles.update_one({"_id": ObjectId(role_id)}, {"$set": update_data})
    
    return {"message": "تم تحديث الدور بنجاح"}

@api_router.delete("/roles/{role_id}")
async def delete_role(role_id: str, current_user: dict = Depends(get_current_user)):
    """حذف دور"""
    if current_user["role"] != "admin" and not has_permission(current_user, "manage_roles"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    role = await db.roles.find_one({"_id": ObjectId(role_id)})
    if not role:
        raise HTTPException(status_code=404, detail="الدور غير موجود")
    
    if role.get("is_system"):
        raise HTTPException(status_code=400, detail="لا يمكن حذف دور نظامي")
    
    # التحقق من عدم وجود مستخدمين بهذا الدور
    users_count = await db.users.count_documents({"role_id": role_id})
    if users_count > 0:
        raise HTTPException(status_code=400, detail=f"لا يمكن حذف الدور، يوجد {users_count} مستخدم مرتبط به")
    
    await db.roles.delete_one({"_id": ObjectId(role_id)})
    
    return {"message": "تم حذف الدور بنجاح"}

@api_router.post("/roles/init")
async def init_default_roles(current_user: dict = Depends(get_current_user)):
    """إنشاء الأدوار الافتراضية"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    default_roles = [
        {
            "name": "مدير النظام",
            "description": "صلاحيات كاملة على النظام",
            "permissions": list(DEFAULT_PERMISSIONS.get(UserRole.ADMIN, [])),
            "is_system": True,
            "system_key": "admin"
        },
        {
            "name": "أستاذ",
            "description": "صلاحيات الأستاذ الافتراضية",
            "permissions": list(DEFAULT_PERMISSIONS.get(UserRole.TEACHER, [])),
            "is_system": True,
            "system_key": "teacher"
        },
        {
            "name": "موظف",
            "description": "صلاحيات الموظف الافتراضية",
            "permissions": list(DEFAULT_PERMISSIONS.get(UserRole.EMPLOYEE, [])),
            "is_system": True,
            "system_key": "employee"
        },
        {
            "name": "طالب",
            "description": "صلاحيات الطالب الافتراضية",
            "permissions": list(DEFAULT_PERMISSIONS.get(UserRole.STUDENT, [])),
            "is_system": True,
            "system_key": "student"
        },
    ]
    
    created = 0
    updated = 0
    for role in default_roles:
        existing = await db.roles.find_one({"system_key": role["system_key"]})
        if not existing:
            role["created_at"] = get_yemen_time()
            await db.roles.insert_one(role)
            created += 1
        else:
            # تحديث الصلاحيات للأدوار الموجودة لتطابق الافتراضية
            if set(existing.get("permissions", [])) != set(role["permissions"]):
                await db.roles.update_one(
                    {"system_key": role["system_key"]},
                    {"$set": {"permissions": role["permissions"]}}
                )
                updated += 1
    
    return {"message": f"تم إنشاء {created} دور وتحديث {updated} دور"}

# ==================== Permissions Management Routes ====================

class PermissionUpdate(BaseModel):
    permissions: List[str]

class UserRoleUpdate(BaseModel):
    role_id: str  # معرف الدور المخصص

@api_router.get("/permissions/all")
async def get_all_permissions(current_user: dict = Depends(get_current_user)):
    """الحصول على قائمة جميع الصلاحيات المتاحة"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # جلب الأدوار المخصصة
    roles = await db.roles.find().to_list(100)
    roles_list = []
    for r in roles:
        roles_list.append({
            "id": str(r["_id"]),
            "name": r["name"],
            "permissions": r.get("permissions", []),
            "is_system": r.get("is_system", False)
        })
    
    return {
        "permissions": ALL_PERMISSIONS,
        "roles": roles_list,
        "default_permissions": DEFAULT_PERMISSIONS
    }

@api_router.get("/users/{user_id}/permissions")
async def get_user_permissions_endpoint(user_id: str, current_user: dict = Depends(get_current_user)):
    """الحصول على صلاحيات مستخدم معين"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # جلب صلاحيات المستخدم من الدور المسند له
    user_permissions = []
    role_name = ""
    role_id = user.get("role_id")
    
    if role_id:
        role = await db.roles.find_one({"_id": ObjectId(role_id)})
        if role:
            user_permissions = role.get("permissions", [])
            role_name = role.get("name", "")
    else:
        # استخدام الصلاحيات الافتراضية للدور القديم
        user_permissions = DEFAULT_PERMISSIONS.get(user.get("role", ""), [])
        role_name = user.get("role", "")
    
    return {
        "user_id": user_id,
        "role_id": role_id,
        "role_name": role_name,
        "permissions": user_permissions
    }

@api_router.put("/users/{user_id}/role")
async def assign_role_to_user(user_id: str, data: UserRoleUpdate, current_user: dict = Depends(get_current_user)):
    """إسناد دور لمستخدم"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # التحقق من وجود الدور
    role = await db.roles.find_one({"_id": ObjectId(data.role_id)})
    if not role:
        raise HTTPException(status_code=404, detail="الدور غير موجود")
    
    # تحديث دور المستخدم — استنتاج system_key الصحيح من الاسم العربي إن لم يكن موجوداً
    sys_key = (role.get("system_key") or "").strip()
    if not sys_key:
        ROLE_NAME_MAP = {
            "رئيس قسم": "department_head", "عميد كلية": "dean", "عميد": "dean",
            "مسجل": "registrar", "مدير التسجيل": "registration_manager",
            "موظف": "employee", "مدير النظام": "admin",
            "Department Head": "department_head", "Dean": "dean",
            "Registrar": "registrar", "Employee": "employee",
        }
        sys_key = ROLE_NAME_MAP.get((role.get("name") or "").strip(), "custom")
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"role_id": data.role_id, "role": sys_key}}
    )
    
    return {
        "message": "تم إسناد الدور بنجاح",
        "role_id": data.role_id,
        "role_name": role["name"],
        "permissions": role.get("permissions", [])
    }

@api_router.delete("/users/{user_id}/permissions/reset")
async def reset_user_permissions(user_id: str, current_user: dict = Depends(get_current_user)):
    """إعادة تعيين صلاحيات المستخدم إلى الافتراضية حسب دوره"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_users"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$unset": {"permissions": 1}}
    )
    
    default_perms = DEFAULT_PERMISSIONS.get(user["role"], [])
    return {
        "message": "تم إعادة تعيين الصلاحيات إلى الافتراضية",
        "permissions": default_perms
    }

# ==================== Department Routes ====================

@api_router.post("/departments", response_model=DepartmentResponse)
async def create_department(dept: DepartmentCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_departments") and not has_permission(current_user, "add_department"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    dept_dict = dept.dict()
    dept_dict["created_at"] = get_yemen_time()
    
    result = await db.departments.insert_one(dept_dict)
    dept_dict["id"] = str(result.inserted_id)
    
    return dept_dict

@api_router.get("/departments", response_model=List[DepartmentResponse])
async def get_departments(current_user: dict = Depends(get_current_user)):
    # تطبيق فلتر النطاق بناءً على دور المستخدم
    scope_filter = await get_user_scope_filter(current_user, "departments")
    
    depts = await db.departments.find(scope_filter).to_list(100)
    result = []
    for d in depts:
        dept_data = {
            "id": str(d["_id"]),
            "name": d["name"],
            "code": d["code"],
            "description": d.get("description"),
            "faculty_id": d.get("faculty_id"),
            "created_at": d["created_at"]
        }
        # جلب اسم الكلية إذا موجود
        if d.get("faculty_id"):
            try:
                faculty = await db.faculties.find_one({"_id": ObjectId(d["faculty_id"])})
                dept_data["faculty_name"] = faculty["name"] if faculty else None
            except:
                dept_data["faculty_name"] = None
        result.append(dept_data)
    return result

@api_router.delete("/departments/{dept_id}")
async def delete_department(dept_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_departments") and not has_permission(current_user, "delete_department"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    dept = await db.departments.find_one({"_id": ObjectId(dept_id)})
    if not dept:
        raise HTTPException(status_code=404, detail="القسم غير موجود")
    
    # التحقق من عدم وجود بيانات مرتبطة (الفعالة وغير الفعالة)
    students_count = await db.students.count_documents({"department_id": dept_id})
    if students_count > 0:
        raise HTTPException(status_code=400, detail=f"لا يمكن حذف القسم - يوجد {students_count} طالب مرتبط به. احذف الطلاب أولاً")
    
    teachers_count = await db.teachers.count_documents({"department_id": dept_id})
    if teachers_count > 0:
        raise HTTPException(status_code=400, detail=f"لا يمكن حذف القسم - يوجد {teachers_count} معلم مرتبط به. احذف المعلمين أولاً")
    
    courses_count = await db.courses.count_documents({"department_id": dept_id})
    if courses_count > 0:
        raise HTTPException(status_code=400, detail=f"لا يمكن حذف القسم - يوجد {courses_count} مقرر مرتبط به. احذف المقررات أولاً")
    
    result = await db.departments.delete_one({"_id": ObjectId(dept_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="القسم غير موجود")
    
    return {"message": "تم حذف القسم بنجاح"}

@api_router.put("/departments/{dept_id}", response_model=DepartmentResponse)
async def update_department(dept_id: str, dept: DepartmentCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_departments") and not has_permission(current_user, "edit_department"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    existing = await db.departments.find_one({"_id": ObjectId(dept_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="القسم غير موجود")
    
    await db.departments.update_one(
        {"_id": ObjectId(dept_id)},
        {"$set": dept.dict()}
    )
    
    return {
        "id": dept_id,
        **dept.dict(),
        "created_at": existing["created_at"]
    }

@api_router.get("/departments/{dept_id}/details")
async def get_department_details(dept_id: str, current_user: dict = Depends(get_current_user)):
    """Get department details with students, courses, and teachers"""
    dept = await db.departments.find_one({"_id": ObjectId(dept_id)})
    if not dept:
        raise HTTPException(status_code=404, detail="القسم غير موجود")
    
    # Get courses in this department
    courses = await db.courses.find({"department_id": dept_id}).to_list(500)
    course_ids = [str(c["_id"]) for c in courses]
    
    # Get teachers directly from teachers collection by department_ids
    dept_teachers = await db.teachers.find({
        "$or": [
            {"department_ids": dept_id},
            {"department_id": dept_id}
        ]
    }).to_list(500)
    
    teachers = []
    seen_teacher_ids = set()
    for t in dept_teachers:
        tid = str(t["_id"])
        if tid not in seen_teacher_ids:
            seen_teacher_ids.add(tid)
            teachers.append({
                "id": tid,
                "full_name": t.get("full_name", ""),
                "username": t.get("teacher_id", t.get("username", "")),
                "academic_title": t.get("academic_title", ""),
                "specialization": t.get("specialization", ""),
            })
    
    # Also add teachers from courses who aren't in the department's teachers list
    for c in courses:
        tid = c.get("teacher_id")
        if tid and tid not in seen_teacher_ids:
            try:
                teacher = await db.teachers.find_one({"_id": ObjectId(tid)})
                if not teacher:
                    teacher = await db.users.find_one({"_id": ObjectId(tid), "role": "teacher"})
                if teacher:
                    seen_teacher_ids.add(tid)
                    teachers.append({
                        "id": tid,
                        "full_name": teacher.get("full_name", ""),
                        "username": teacher.get("teacher_id", teacher.get("username", "")),
                        "academic_title": teacher.get("academic_title", ""),
                        "specialization": teacher.get("specialization", ""),
                    })
            except:
                pass
    
    # Get students enrolled in courses of this department
    student_ids = set()
    for course_id in course_ids:
        enrollments = await db.enrollments.find({"course_id": course_id}).to_list(1000)
        for e in enrollments:
            student_ids.add(e["student_id"])
    
    # Get student details
    students = []
    for sid in student_ids:
        try:
            student = await db.students.find_one({"_id": ObjectId(sid)})
            if student:
                students.append({
                    "id": str(student["_id"]),
                    "student_id": student["student_id"],
                    "full_name": student["full_name"],
                    "level": student.get("level", 1),
                    "section": student.get("section", "")
                })
        except:
            pass
    
    # Format courses
    courses_data = []
    for c in courses:
        teacher_name = ""
        if c.get("teacher_id"):
            teacher = await db.users.find_one({"_id": ObjectId(c["teacher_id"])})
            if teacher:
                teacher_name = teacher["full_name"]
        
        # Count enrolled students
        enrolled_count = await db.enrollments.count_documents({"course_id": str(c["_id"])})
        
        courses_data.append({
            "id": str(c["_id"]),
            "name": c["name"],
            "code": c.get("code", ""),
            "level": c.get("level", 1),
            "section": c.get("section", ""),
            "teacher_name": teacher_name,
            "students_count": enrolled_count
        })
    
    return {
        "id": str(dept["_id"]),
        "name": dept["name"],
        "code": dept.get("code", ""),
        "description": dept.get("description", ""),
        "students_count": len(students),
        "courses_count": len(courses),
        "teachers_count": len(teachers),
        "students": students,
        "courses": courses_data,
        "teachers": teachers
    }

@api_router.get("/departments/stats")
async def get_departments_stats(current_user: dict = Depends(get_current_user)):
    """Get all departments with student counts (مُحسَّن: استعلامات مُجمَّعة بدلاً من N+1)"""
    # تطبيق فلتر النطاق بناءً على دور المستخدم
    scope_filter = await get_user_scope_filter(current_user, "departments")
    
    depts = await db.departments.find(scope_filter).to_list(100)
    if not depts:
        return []
    
    dept_ids = [str(d["_id"]) for d in depts]
    
    # استعلامات مُجمَّعة (3 استعلامات إجمالاً بدلاً من 1 + N + N*M)
    # 1) عدد الطلاب لكل قسم مباشرة من جدول students
    students_pipeline = [
        {"$match": {"department_id": {"$in": dept_ids}}},
        {"$group": {"_id": "$department_id", "count": {"$sum": 1}}}
    ]
    students_counts = {}
    async for row in db.students.aggregate(students_pipeline):
        students_counts[row["_id"]] = row["count"]
    
    # 2) عدد المقررات لكل قسم
    courses_pipeline = [
        {"$match": {"department_id": {"$in": dept_ids}}},
        {"$group": {"_id": "$department_id", "count": {"$sum": 1}}}
    ]
    courses_counts = {}
    async for row in db.courses.aggregate(courses_pipeline):
        courses_counts[row["_id"]] = row["count"]
    
    # 3) أسماء الكليات دفعة واحدة
    fac_ids_set = set()
    for d in depts:
        if d.get("faculty_id"):
            try:
                fac_ids_set.add(ObjectId(d["faculty_id"]))
            except Exception:
                pass
    
    faculty_names = {}
    if fac_ids_set:
        async for f in db.faculties.find(
            {"_id": {"$in": list(fac_ids_set)}},
            {"name": 1}
        ):
            faculty_names[str(f["_id"])] = f.get("name")
    
    result = []
    for dept in depts:
        dept_id = str(dept["_id"])
        result.append({
            "id": dept_id,
            "name": dept["name"],
            "code": dept.get("code", ""),
            "description": dept.get("description", ""),
            "faculty_id": dept.get("faculty_id"),
            "faculty_name": faculty_names.get(dept.get("faculty_id", "")),
            "default_program_code": dept.get("default_program_code", ""),
            "students_count": students_counts.get(dept_id, 0),
            "courses_count": courses_counts.get(dept_id, 0),
        })
    
    return result

@api_router.get("/departments/dashboard")
async def get_department_head_dashboard(current_user: dict = Depends(get_current_user)):
    """لوحة تحكم رئيس القسم - إحصائيات الأقسام + التنبيهات"""
    
    # جلب الأقسام التي يديرها المستخدم
    user_data = await db.users.find_one({"_id": ObjectId(current_user["id"])})
    if not user_data:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # جلب قائمة الأقسام
    department_ids = user_data.get("department_ids") or []
    if not department_ids and user_data.get("department_id"):
        department_ids = [user_data.get("department_id")]
    
    # إذا كان المستخدم admin أو لديه faculty_id، جلب كل الأقسام
    if current_user["role"] == UserRole.ADMIN:
        depts = await db.departments.find({}).to_list(100)
    elif user_data.get("faculty_id") and not department_ids:
        depts = await db.departments.find({"faculty_id": user_data["faculty_id"]}).to_list(100)
    elif department_ids:
        depts = await db.departments.find({"_id": {"$in": [ObjectId(did) for did in department_ids]}}).to_list(100)
    else:
        depts = []
    
    # جلب إعدادات النظام
    settings = await db.settings.find_one({"type": "general"})
    warning_threshold = settings.get("warning_threshold", 15) if settings else 15
    deprivation_threshold = settings.get("deprivation_threshold", 25) if settings else 25
    
    departments_data = []
    total_students = 0
    total_courses = 0
    all_warnings = []
    
    for dept in depts:
        dept_id = str(dept["_id"])
        
        # جلب المقررات في هذا القسم
        courses = await db.courses.find({"department_id": dept_id, "is_active": True}).to_list(100)
        course_ids = [str(c["_id"]) for c in courses]
        
        # جلب الطلاب المسجلين
        student_ids = set()
        for course_id in course_ids:
            enrollments = await db.enrollments.find({"course_id": course_id}).to_list(1000)
            for e in enrollments:
                student_ids.add(e["student_id"])
        
        dept_students_count = len(student_ids)
        dept_courses_count = len(courses)
        total_students += dept_students_count
        total_courses += dept_courses_count
        
        # حساب نسبة الحضور اليومي
        today = get_yemen_time().replace(hour=0, minute=0, second=0, microsecond=0)
        today_attendance = 0
        today_total = 0
        
        for course_id in course_ids:
            # جلب محاضرات اليوم
            today_lectures = await db.lectures.find({
                "course_id": course_id,
                "date": {"$gte": today}
            }).to_list(100)
            
            for lecture in today_lectures:
                lecture_id = str(lecture["_id"])
                records = await db.attendance.find({
                    "lecture_id": lecture_id,
                    "course_id": course_id
                }).to_list(1000)
                
                today_total += len(records)
                today_attendance += sum(1 for r in records if r["status"] == AttendanceStatus.PRESENT)
        
        attendance_rate = round((today_attendance / today_total * 100), 1) if today_total > 0 else 0
        
        # جلب التنبيهات (طلاب قريبون من الحرمان)
        dept_warnings = []
        for course in courses:
            cid = str(course["_id"])
            active_lecture_ids = await get_active_lecture_ids(cid)
            total_lectures = len(active_lecture_ids)
            
            if total_lectures == 0:
                continue
            
            enrollments = await db.enrollments.find({"course_id": cid}).to_list(1000)
            
            for enrollment in enrollments:
                student_id = enrollment["student_id"]
                
                records = await db.attendance.find({
                    "student_id": student_id,
                    "course_id": cid,
                    "lecture_id": {"$in": list(active_lecture_ids)}
                }).to_list(1000)
                
                absent_count = sum(1 for r in records if r["status"] == AttendanceStatus.ABSENT)
                absence_rate = (absent_count / total_lectures) * 100
                
                if absence_rate >= warning_threshold:
                    student = await db.students.find_one({"_id": ObjectId(student_id)})
                    if student:
                        warning_data = {
                            "student_id": student.get("student_id"),
                            "student_name": student.get("full_name"),
                            "course_name": course["name"],
                            "course_code": course.get("code", ""),
                            "department_id": dept_id,
                            "department_name": dept["name"],
                            "total_lectures": total_lectures,
                            "absent_count": absent_count,
                            "absence_rate": round(absence_rate, 2),
                            "remaining_allowed": max(0, int((deprivation_threshold / 100) * total_lectures) - absent_count),
                            "status": "محروم" if absence_rate >= deprivation_threshold else "إنذار"
                        }
                        dept_warnings.append(warning_data)
                        all_warnings.append(warning_data)
        
        # جلب اسم الكلية
        faculty_name = None
        if dept.get("faculty_id"):
            try:
                faculty = await db.faculties.find_one({"_id": ObjectId(dept["faculty_id"])})
                faculty_name = faculty["name"] if faculty else None
            except:
                pass
        
        departments_data.append({
            "id": dept_id,
            "name": dept["name"],
            "code": dept.get("code", ""),
            "faculty_name": faculty_name,
            "students_count": dept_students_count,
            "courses_count": dept_courses_count,
            "today_attendance_rate": attendance_rate,
            "warnings_count": len([w for w in dept_warnings if w["status"] == "إنذار"]),
            "deprivations_count": len([w for w in dept_warnings if w["status"] == "محروم"])
        })
    
    return {
        "departments": departments_data,
        "summary": {
            "total_departments": len(departments_data),
            "total_students": total_students,
            "total_courses": total_courses,
            "total_warnings": len([w for w in all_warnings if w["status"] == "إنذار"]),
            "total_deprivations": len([w for w in all_warnings if w["status"] == "محروم"])
        },
        "warnings": sorted(all_warnings, key=lambda x: x["absence_rate"], reverse=True)[:20],  # أعلى 20 تنبيه
        "thresholds": {
            "warning": warning_threshold,
            "deprivation": deprivation_threshold
        }
    }

# ==================== Notification Routes (مسارات الإشعارات) ====================

# ==================== FCM Token Registration (تسجيل رمز الإشعارات) ====================

@api_router.post("/fcm/register")
async def register_fcm_token(request: Request, current_user: dict = Depends(get_current_user)):
    """تسجيل رمز FCM للإشعارات"""
    body = await request.json()
    token = body.get("token")
    platform = body.get("platform", "android")
    
    if not token:
        raise HTTPException(status_code=400, detail="الرمز مطلوب")
    
    user_id = current_user["id"]
    
    # حذف الرمز القديم إذا كان موجوداً لنفس الجهاز
    await db.fcm_tokens.delete_many({"token": token})
    
    # تسجيل الرمز الجديد
    await db.fcm_tokens.insert_one({
        "user_id": user_id,
        "token": token,
        "platform": platform,
        "created_at": get_yemen_time()
    })
    
    return {"message": "تم تسجيل رمز الإشعارات بنجاح"}

@api_router.delete("/fcm/unregister")
async def unregister_fcm_token(request: Request, current_user: dict = Depends(get_current_user)):
    """إلغاء تسجيل رمز FCM"""
    body = await request.json()
    token = body.get("token")
    
    if token:
        await db.fcm_tokens.delete_many({"token": token})
    
    return {"message": "تم إلغاء تسجيل الرمز"}



async def create_student_notification(
    student_id: str,
    student_db_id: str,
    notification_type: NotificationType,
    course_name: str,
    course_id: str,
    absence_rate: float,
    remaining_allowed: int
):
    """إنشاء إشعار للطالب"""
    
    # جلب بيانات الطالب
    student = await db.students.find_one({"_id": ObjectId(student_db_id)})
    if not student:
        return None
    
    # تحديد العنوان والرسالة حسب النوع
    if notification_type == NotificationType.WARNING:
        title = f"⚠️ إنذار غياب - {course_name}"
        message = f"نسبة غيابك في مقرر {course_name} وصلت إلى {absence_rate:.1f}%. يتبقى لك {remaining_allowed} غياب مسموح قبل الحرمان."
    elif notification_type == NotificationType.DEPRIVATION:
        title = f"🚫 تنبيه حرمان - {course_name}"
        message = f"تجاوزت نسبة الغياب المسموحة في مقرر {course_name}. نسبة غيابك الحالية: {absence_rate:.1f}%."
    else:
        title = f"📢 إشعار - {course_name}"
        message = f"إشعار متعلق بمقرر {course_name}"
    
    # التحقق من عدم وجود إشعار مشابه خلال آخر 24 ساعة
    yesterday = get_yemen_time() - timedelta(days=1)
    existing_notification = await db.notifications.find_one({
        "student_id": student_db_id,
        "course_id": course_id,
        "type": notification_type.value,
        "created_at": {"$gte": yesterday}
    })
    
    if existing_notification:
        # تحديث الإشعار الموجود إذا تغيرت النسبة
        if existing_notification.get("absence_rate") != absence_rate:
            await db.notifications.update_one(
                {"_id": existing_notification["_id"]},
                {"$set": {
                    "message": message,
                    "absence_rate": absence_rate,
                    "remaining_allowed": remaining_allowed,
                    "is_read": False,
                    "updated_at": get_yemen_time()
                }}
            )
        return None
    
    # إنشاء إشعار جديد
    notification = {
        "student_id": student_db_id,
        "user_id": student.get("user_id"),
        "title": title,
        "message": message,
        "type": notification_type.value,
        "course_id": course_id,
        "course_name": course_name,
        "absence_rate": absence_rate,
        "remaining_allowed": remaining_allowed,
        "is_read": False,
        "created_at": get_yemen_time()
    }
    
    result = await db.notifications.insert_one(notification)
    
    # إرسال إشعار Firebase Push للطالب
    try:
        from services.firebase_service import send_notification_to_many
        user_id = student.get("user_id")
        if user_id:
            tokens_docs = await db.fcm_tokens.find({"user_id": user_id}).to_list(100)
            tokens = [doc["token"] for doc in tokens_docs]
            if tokens:
                await send_notification_to_many(tokens, title, message)
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"Firebase push failed for student: {e}")
    
    return str(result.inserted_id)

async def check_and_create_absence_notifications(student_id: str, course_id: str):
    """التحقق من نسبة الغياب وإنشاء إشعارات إذا لزم الأمر"""
    
    # جلب الطالب
    student = await db.students.find_one({"student_id": student_id})
    if not student:
        return
    
    student_db_id = str(student["_id"])
    
    # جلب المقرر
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        return
    
    # جلب الإعدادات
    settings = await db.settings.find_one({"type": "general"})
    warning_threshold = settings.get("warning_threshold", 15) if settings else 15
    deprivation_threshold = settings.get("deprivation_threshold", 25) if settings else 25
    
    # حساب نسبة الغياب
    active_lecture_ids = await get_active_lecture_ids(course_id)
    total_lectures = len(active_lecture_ids)
    
    if total_lectures == 0:
        return
    
    records = await db.attendance.find({
        "student_id": student_id,
        "course_id": course_id,
        "lecture_id": {"$in": list(active_lecture_ids)}
    }).to_list(1000)
    
    absent_count = sum(1 for r in records if r["status"] == AttendanceStatus.ABSENT)
    absence_rate = (absent_count / total_lectures) * 100
    
    # حساب الغيابات المتبقية
    max_allowed = int((deprivation_threshold / 100) * total_lectures)
    remaining_allowed = max(0, max_allowed - absent_count)
    
    # إنشاء إشعار حسب النسبة
    if absence_rate >= deprivation_threshold:
        await create_student_notification(
            student_id=student_id,
            student_db_id=student_db_id,
            notification_type=NotificationType.DEPRIVATION,
            course_name=course["name"],
            course_id=course_id,
            absence_rate=absence_rate,
            remaining_allowed=0
        )
    elif absence_rate >= warning_threshold:
        await create_student_notification(
            student_id=student_id,
            student_db_id=student_db_id,
            notification_type=NotificationType.WARNING,
            course_name=course["name"],
            course_id=course_id,
            absence_rate=absence_rate,
            remaining_allowed=remaining_allowed
        )

@api_router.get("/notifications")
async def get_notifications(
    limit: int = 50,
    unread_only: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """جلب إشعارات المستخدم الحالي"""
    
    # البحث عن الطالب المرتبط بهذا المستخدم
    user = await db.users.find_one({"_id": ObjectId(current_user["id"])})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # البحث عن الطالب
    student = await db.students.find_one({"user_id": str(user["_id"])})
    
    if not student:
        # إذا لم يكن طالباً، ربما يكون admin أو معلم - إرجاع قائمة فارغة
        return {"notifications": [], "unread_count": 0}
    
    student_db_id = str(student["_id"])
    
    # جلب الإشعارات
    query = {"student_id": student_db_id}
    if unread_only:
        query["is_read"] = False
    
    notifications = await db.notifications.find(query).sort("created_at", -1).limit(limit).to_list(limit)
    
    # حساب عدد الإشعارات غير المقروءة
    unread_count = await db.notifications.count_documents({"student_id": student_db_id, "is_read": False})
    
    result = []
    for n in notifications:
        result.append({
            "id": str(n["_id"]),
            "title": n["title"],
            "message": n["message"],
            "type": n["type"],
            "course_id": n.get("course_id"),
            "course_name": n.get("course_name"),
            "absence_rate": n.get("absence_rate"),
            "remaining_allowed": n.get("remaining_allowed"),
            "is_read": n.get("is_read", False),
            "created_at": n["created_at"]
        })
    
    return {"notifications": result, "unread_count": unread_count}

@api_router.get("/notifications/count")
async def get_unread_notifications_count(current_user: dict = Depends(get_current_user)):
    """جلب عدد الإشعارات غير المقروءة لأي مستخدم"""
    user_id = current_user["id"]
    
    count = await db.notifications.count_documents({
        "user_id": user_id,
        "is_read": False
    })
    
    return {"count": count}

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_as_read(
    notification_id: str,
    current_user: dict = Depends(get_current_user)
):
    """تحديد إشعار كمقروء"""
    
    notification = await db.notifications.find_one({"_id": ObjectId(notification_id)})
    if not notification:
        raise HTTPException(status_code=404, detail="الإشعار غير موجود")
    
    await db.notifications.update_one(
        {"_id": ObjectId(notification_id)},
        {"$set": {"is_read": True}}
    )
    
    return {"message": "تم تحديد الإشعار كمقروء"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_as_read(current_user: dict = Depends(get_current_user)):
    """تحديد جميع الإشعارات كمقروءة"""
    
    user = await db.users.find_one({"_id": ObjectId(current_user["id"])})
    if not user:
        return {"message": "لا توجد إشعارات"}
    
    # بناء استعلام البحث بناءً على نوع المستخدم
    query = {"is_read": False}
    
    student = await db.students.find_one({"user_id": str(user["_id"])})
    if student:
        query["student_id"] = str(student["_id"])
    else:
        # للمعلمين والمستخدمين الآخرين - استخدام user_id
        query["user_id"] = str(user["_id"])
    
    result = await db.notifications.update_many(
        query,
        {"$set": {"is_read": True}}
    )
    
    return {"message": f"تم تحديد {result.modified_count} إشعار كمقروء"}

class ManualNotificationCreate(BaseModel):
    student_id: str  # معرف الطالب (من جدول students)
    title: str
    message: str
    type: str = "warning"  # warning, deprivation, info, reminder
    course_id: Optional[str] = None

@api_router.post("/notifications/manual")
async def create_manual_notification(
    data: ManualNotificationCreate,
    current_user: dict = Depends(get_current_user)
):
    """إنشاء إنذار/إشعار يدوي لطالب معين"""
    
    # التحقق من الصلاحيات - يجب أن يكون admin أو لديه صلاحية send_notifications
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_users"):
        user = await db.users.find_one({"_id": ObjectId(current_user["id"])})
        user_permissions = user.get("permissions", []) if user else []
        
        if Permission.SEND_NOTIFICATIONS not in user_permissions:
            raise HTTPException(status_code=403, detail="غير مصرح لك بإرسال إنذارات. تحتاج صلاحية 'إرسال إشعارات وإنذارات للطلاب'")
    
    # جلب بيانات الطالب
    student = await db.students.find_one({"_id": ObjectId(data.student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    # جلب اسم المقرر إذا تم تحديده
    course_name = None
    if data.course_id:
        course = await db.courses.find_one({"_id": ObjectId(data.course_id)})
        if course:
            course_name = course["name"]
    
    # إنشاء الإشعار
    notification = {
        "student_id": data.student_id,
        "user_id": student.get("user_id"),
        "title": data.title,
        "message": data.message,
        "type": data.type,
        "course_id": data.course_id,
        "course_name": course_name,
        "is_read": False,
        "is_manual": True,  # تمييز الإنذار اليدوي
        "sent_by": current_user["id"],
        "sent_by_name": current_user.get("full_name", ""),
        "created_at": get_yemen_time()
    }
    
    result = await db.notifications.insert_one(notification)
    
    # إرسال push notification عبر Firebase
    try:
        from services.firebase_service import send_notification_to_many
        user_id = student.get("user_id")
        if user_id:
            tokens_docs = await db.fcm_tokens.find({"user_id": user_id}).to_list(100)
            tokens = [doc["token"] for doc in tokens_docs if doc.get("token")]
            if tokens:
                await send_notification_to_many(tokens, data.title, data.message)
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"Firebase push failed: {e}")
    
    # تسجيل النشاط
    await log_activity(
        user=current_user,
        action="send_manual_notification",
        entity_type="notification",
        entity_id=str(result.inserted_id),
        entity_name=student.get('full_name', student.get('student_id'))
    )
    
    return {
        "message": f"تم إرسال الإنذار للطالب {student.get('full_name')} بنجاح",
        "notification_id": str(result.inserted_id)
    }


# ==================== النتائج النهائية - Final Results ====================

class FinalResultItem(BaseModel):
    student_number: str  # رقم القيد
    result: str  # "pass" | "fail" | "ناجح" | "راسب"
    grade: Optional[str] = None  # درجة اختيارية
    notes: Optional[str] = None


class FinalResultsPayload(BaseModel):
    results: List[FinalResultItem]


def _normalize_result(value: str) -> Optional[str]:
    """تحويل قيمة النتيجة إلى 'pass' أو 'fail'"""
    if not value:
        return None
    v = str(value).strip().lower()
    pass_keywords = {"pass", "passed", "p", "ناجح", "ناجحة", "نجاح", "✓", "yes", "y", "1", "true"}
    fail_keywords = {"fail", "failed", "f", "راسب", "راسبة", "رسوب", "✗", "no", "n", "0", "false"}
    if v in pass_keywords:
        return "pass"
    if v in fail_keywords:
        return "fail"
    return None


async def _send_final_results_to_students(course_id: str, items: List[dict], current_user: dict) -> dict:
    """إنشاء إشعارات النتيجة النهائية وإرسال Push للطلاب"""
    try:
        course_oid = ObjectId(course_id)
    except Exception:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    course = await db.courses.find_one({"_id": course_oid})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")

    course_name = course.get("name", "")
    course_code = course.get("code", "")

    # خريطة طلاب المقرر برقم القيد
    student_numbers = list({str(it.get("student_number", "")).strip() for it in items if it.get("student_number")})
    students_cursor = db.students.find({"student_id": {"$in": student_numbers}})
    students_by_number = {}
    async for s in students_cursor:
        students_by_number[str(s.get("student_id", "")).strip()] = s

    sent = 0
    failed: List[dict] = []
    push_sent = 0

    from services.firebase_service import send_notification_to_many

    for it in items:
        student_number = str(it.get("student_number", "")).strip()
        result_value = _normalize_result(it.get("result", ""))
        grade = it.get("grade")
        notes = it.get("notes")

        if not student_number:
            failed.append({"student_number": student_number, "reason": "رقم القيد مفقود"})
            continue
        if result_value not in ("pass", "fail"):
            failed.append({"student_number": student_number, "reason": "النتيجة غير صالحة (يجب أن تكون ناجح/راسب)"})
            continue

        student = students_by_number.get(student_number)
        if not student:
            failed.append({"student_number": student_number, "reason": "الطالب غير موجود"})
            continue

        result_ar = "ناجح" if result_value == "pass" else "راسب"
        title = f"النتيجة النهائية - {course_name}"
        message_parts = [f"رقم القيد: {student_number}", f"المقرر: {course_name}"]
        if course_code:
            message_parts.append(f"الرمز: {course_code}")
        message_parts.append(f"النتيجة: {result_ar}")
        if grade:
            message_parts.append(f"الدرجة: {grade}")
        if notes:
            message_parts.append(f"ملاحظات: {notes}")
        message = " | ".join(message_parts)

        student_db_id = str(student["_id"])
        notification = {
            "student_id": student_db_id,
            "user_id": student.get("user_id"),
            "title": title,
            "message": message,
            "type": NotificationType.INFO.value,
            "course_id": course_id,
            "course_name": course_name,
            "is_read": False,
            "is_manual": True,
            "is_final_result": True,
            "final_result": result_value,
            "final_grade": grade,
            "sent_by": current_user["id"],
            "sent_by_name": current_user.get("full_name", ""),
            "created_at": get_yemen_time(),
        }
        await db.notifications.insert_one(notification)
        sent += 1

        # إرسال Push عبر Firebase
        try:
            user_id = student.get("user_id")
            if user_id:
                tokens_docs = await db.fcm_tokens.find({"user_id": user_id}).to_list(100)
                tokens = [doc["token"] for doc in tokens_docs if doc.get("token")]
                if tokens:
                    await send_notification_to_many(tokens, title, message)
                    push_sent += 1
        except Exception as e:
            logging.getLogger(__name__).error(f"Final result push failed for {student_number}: {e}")

    await log_activity(
        user=current_user,
        action="send_final_results",
        entity_type="course",
        entity_id=course_id,
        entity_name=course_name,
        details={"sent": sent, "failed_count": len(failed)},
    )

    return {
        "message": f"تم إرسال {sent} إشعار نتيجة نهائية",
        "sent": sent,
        "failed_count": len(failed),
        "failed": failed,
        "push_sent": push_sent,
    }


def _can_send_final_results(current_user: dict) -> bool:
    if current_user.get("role") == UserRole.ADMIN:
        return True
    return (
        has_permission(current_user, "send_notifications")
        or has_permission(current_user, "manage_courses")
        or has_permission(current_user, "manage_grades")
    )


@api_router.post("/courses/{course_id}/send-final-results")
async def send_final_results_json(
    course_id: str,
    payload: FinalResultsPayload,
    current_user: dict = Depends(get_current_user),
):
    """إرسال النتيجة النهائية (ناجح/راسب) لطلاب المقرر عبر الإشعارات (JSON)"""
    if not _can_send_final_results(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك بإرسال النتائج النهائية")

    items = [item.dict() for item in payload.results]
    if not items:
        raise HTTPException(status_code=400, detail="لا توجد نتائج للإرسال")

    return await _send_final_results_to_students(course_id, items, current_user)


@api_router.post("/courses/{course_id}/send-final-results/upload")
async def send_final_results_upload(
    course_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """إرسال النتيجة النهائية لطلاب المقرر من ملف Excel/CSV"""
    if not _can_send_final_results(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك بإرسال النتائج النهائية")

    contents = await file.read()
    filename = (file.filename or "").lower()
    try:
        if filename.endswith(".csv"):
            df = pd.read_csv(BytesIO(contents))
        else:
            df = pd.read_excel(BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"تعذر قراءة الملف: {e}")

    # قبول أعمدة عربية أو إنجليزية
    column_map_candidates = {
        "student_number": ["student_number", "رقم القيد", "رقم_القيد", "الرقم", "id", "student_id", "رقم الطالب"],
        "result": ["result", "النتيجة", "النتيجه", "status", "الحالة"],
        "grade": ["grade", "الدرجة", "الدرجه", "score", "الدرجة النهائية"],
        "notes": ["notes", "ملاحظات", "ملاحظه", "ملاحظة"],
    }

    cols_lower = {c: str(c).strip() for c in df.columns}
    resolved = {}
    for key, candidates in column_map_candidates.items():
        for c, original in cols_lower.items():
            if original in candidates or original.lower() in [x.lower() for x in candidates]:
                resolved[key] = c
                break

    if "student_number" not in resolved or "result" not in resolved:
        raise HTTPException(
            status_code=400,
            detail="الأعمدة المطلوبة غير موجودة. يجب توفر: 'رقم القيد' و 'النتيجة'",
        )

    items = []
    for _, row in df.iterrows():
        sn = row.get(resolved["student_number"])
        rs = row.get(resolved["result"])
        if pd.isna(sn) or pd.isna(rs):
            continue
        # الأرقام قد تأتي كـ float — تحويلها إلى string نظيفة
        if isinstance(sn, float) and sn.is_integer():
            sn = str(int(sn))
        items.append({
            "student_number": str(sn).strip(),
            "result": str(rs).strip(),
            "grade": (None if "grade" not in resolved or pd.isna(row.get(resolved["grade"])) else str(row.get(resolved["grade"])).strip()),
            "notes": (None if "notes" not in resolved or pd.isna(row.get(resolved["notes"])) else str(row.get(resolved["notes"])).strip()),
        })

    if not items:
        raise HTTPException(status_code=400, detail="الملف لا يحتوي على بيانات صالحة")

    return await _send_final_results_to_students(course_id, items, current_user)


# ==================== النتائج على مستوى القسم - Department-Level Results ====================

DEPT_RESULT_VALUES = {
    "ناجح": "ناجح",
    "ناجحة": "ناجح",
    "نجاح": "ناجح",
    "pass": "ناجح",
    "passed": "ناجح",
    "p": "ناجح",
    "دور ثان": "دور ثان",
    "دور ثاني": "دور ثان",
    "دور 2": "دور ثان",
    "second": "دور ثان",
    "second_round": "دور ثان",
    "second round": "دور ثان",
    "راجع التسجيل": "راجع التسجيل",
    "راجع شؤون الطلاب": "راجع التسجيل",
    "راجع التسجيل والقبول": "راجع التسجيل",
    "review": "راجع التسجيل",
    "review_registration": "راجع التسجيل",
}


def _normalize_dept_result(value: str) -> Optional[str]:
    if not value:
        return None
    v = str(value).strip().lower()
    # Build lowercased lookup
    for k, mapped in DEPT_RESULT_VALUES.items():
        if v == k.lower():
            return mapped
    return None


@api_router.get("/template/department-final-results")
async def get_department_final_results_template(current_user: dict = Depends(get_current_user)):
    """تحميل نموذج Excel لإرسال النتائج على مستوى القسم"""
    if not _can_send_final_results(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    data = {
        "رقم القيد": ["12345", "12346", "12347"],
        "النتيجة": ["ناجح", "دور ثان", "راجع التسجيل"],
    }
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=department_final_results_template.xlsx"},
    )



@api_router.get("/departments/{department_id}/distinct-levels")
async def get_distinct_levels_in_department(
    department_id: str,
    current_user: dict = Depends(get_current_user),
):
    """يُرجع قائمة المستويات الموجودة فعلياً في القسم (مرتبة تصاعدياً)."""
    levels = await db.students.distinct(
        "level",
        {"department_id": department_id, "level": {"$exists": True, "$ne": None}},
    )
    # تنظيف وفرز - تجاهل القيم غير الصحيحة
    cleaned = sorted({int(lv) for lv in levels if isinstance(lv, (int, float)) and lv})
    return {"department_id": department_id, "levels": cleaned}


@api_router.get("/departments/{department_id}/distinct-sections")
async def get_distinct_sections_in_department(
    department_id: str,
    level: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
):
    """يُرجع قائمة الشعب الموجودة فعلياً في القسم (اختيارياً مع فلتر المستوى)."""
    query: dict = {
        "department_id": department_id,
        "section": {"$exists": True, "$ne": None, "$nin": ["", None]},
    }
    if level is not None:
        query["level"] = level
    sections = await db.students.distinct("section", query)
    cleaned = sorted({str(s).strip() for s in sections if s and str(s).strip()})
    return {
        "department_id": department_id,
        "level": level,
        "sections": cleaned,
    }


@api_router.post("/departments/{department_id}/send-final-results/upload")
async def send_department_final_results_upload(
    department_id: str,
    file: UploadFile = File(...),
    level: Optional[int] = None,
    section: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """إرسال النتيجة النهائية لطلاب قسم كامل من ملف Excel/CSV (عمودان: رقم القيد، النتيجة).
    فلاتر اختيارية: level (المستوى)، section (الشعبة).
    """
    if not _can_send_final_results(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك بإرسال النتائج")

    try:
        dept_oid = ObjectId(department_id)
    except Exception:
        raise HTTPException(status_code=404, detail="القسم غير موجود")

    department = await db.departments.find_one({"_id": dept_oid})
    if not department:
        raise HTTPException(status_code=404, detail="القسم غير موجود")
    department_name = department.get("name", "")

    contents = await file.read()
    filename = (file.filename or "").lower()
    try:
        if filename.endswith(".csv"):
            df = pd.read_csv(BytesIO(contents))
        else:
            df = pd.read_excel(BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"تعذر قراءة الملف: {e}")

    # Resolve column names (Arabic or English)
    student_col = None
    result_col = None
    for c in df.columns:
        cl = str(c).strip().lower()
        if cl in ("رقم القيد", "student_number", "student_id", "رقم_القيد", "id"):
            student_col = c
        elif cl in ("النتيجة", "result", "status", "الحالة"):
            result_col = c
    if student_col is None or result_col is None:
        raise HTTPException(
            status_code=400,
            detail="الأعمدة المطلوبة غير موجودة. يجب توفر: 'رقم القيد' و 'النتيجة'",
        )

    # Build rows
    rows: List[dict] = []
    for _, row in df.iterrows():
        sn = row.get(student_col)
        rs = row.get(result_col)
        if pd.isna(sn) or pd.isna(rs):
            continue
        if isinstance(sn, float) and sn.is_integer():
            sn = str(int(sn))
        rows.append({"student_number": str(sn).strip(), "result": str(rs).strip()})

    if not rows:
        raise HTTPException(status_code=400, detail="الملف لا يحتوي على بيانات صالحة")

    # Fetch all students in this department (+ optional level/section filters) keyed by student_id
    student_numbers = list({r["student_number"] for r in rows if r["student_number"]})
    student_filter = {
        "department_id": department_id,
        "student_id": {"$in": student_numbers},
    }
    if level is not None:
        student_filter["level"] = level
    if section:
        student_filter["section"] = section
    students_cursor = db.students.find(student_filter)
    students_by_number: dict = {}
    async for s in students_cursor:
        students_by_number[str(s.get("student_id", "")).strip()] = s

    # بناء عنوان واضح يعكس الفلاتر المختارة
    title_parts = [f"النتيجة النهائية - {department_name}"]
    if level is not None:
        title_parts.append(f"المستوى {level}")
    if section:
        title_parts.append(f"الشعبة {section}")
    base_title = " - ".join(title_parts)

    sent = 0
    push_sent = 0
    failed: List[dict] = []
    from services.firebase_service import send_notification_to_many

    for r in rows:
        sn = r["student_number"]
        result_norm = _normalize_dept_result(r["result"])
        if result_norm is None:
            failed.append({"student_number": sn, "reason": f"نتيجة غير صالحة: {r['result']}"})
            continue
        student = students_by_number.get(sn)
        if not student:
            # رسالة الفشل تذكر سبب الاستبعاد (قسم/مستوى/شعبة)
            reason = "الطالب غير موجود في هذا القسم"
            if level is not None or section:
                reason += " (أو لا يطابق فلتر المستوى/الشعبة)"
            failed.append({"student_number": sn, "reason": reason})
            continue

        message = f"رقم القيد: {sn} | القسم: {department_name}"
        if level is not None:
            message += f" | المستوى: {level}"
        if section:
            message += f" | الشعبة: {section}"
        message += f" | النتيجة: {result_norm}"

        student_db_id = str(student["_id"])
        notification = {
            "student_id": student_db_id,
            "user_id": student.get("user_id"),
            "title": base_title,
            "message": message,
            "type": NotificationType.INFO.value,
            "department_id": department_id,
            "department_name": department_name,
            "level": level,
            "section": section,
            "is_read": False,
            "is_manual": True,
            "is_final_result": True,
            "is_department_result": True,
            "final_result": result_norm,
            "sent_by": current_user["id"],
            "sent_by_name": current_user.get("full_name", ""),
            "created_at": get_yemen_time(),
        }
        await db.notifications.insert_one(notification)
        sent += 1

        try:
            user_id = student.get("user_id")
            if user_id:
                tokens_docs = await db.fcm_tokens.find({"user_id": user_id}).to_list(100)
                tokens = [doc["token"] for doc in tokens_docs if doc.get("token")]
                if tokens:
                    await send_notification_to_many(tokens, base_title, message)
                    push_sent += 1
        except Exception as e:
            logging.getLogger(__name__).error(f"Department final result push failed for {sn}: {e}")

    await log_activity(
        user=current_user,
        action="send_department_final_results",
        entity_type="department",
        entity_id=department_id,
        entity_name=department_name,
        details={
            "sent": sent,
            "failed_count": len(failed),
            "level": level,
            "section": section,
        },
    )

    return {
        "message": f"تم إرسال {sent} إشعار نتيجة نهائية",
        "department_name": department_name,
        "level": level,
        "section": section,
        "sent": sent,
        "failed_count": len(failed),
        "failed": failed,
        "push_sent": push_sent,
    }


@api_router.get("/template/final-results")
async def get_final_results_template(current_user: dict = Depends(get_current_user)):
    """تحميل نموذج Excel لإدخال النتائج النهائية"""
    if not _can_send_final_results(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    data = {
        "رقم القيد": ["12345", "12346", "12347"],
        "النتيجة": ["ناجح", "راسب", "ناجح"],
        "الدرجة": ["85", "45", "90"],
        "ملاحظات": ["", "", ""],
    }
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=final_results_template.xlsx"},
    )

@api_router.get("/students/{student_id}/notifications")
async def get_student_notifications(
    student_id: str,
    current_user: dict = Depends(get_current_user)
):
    """جلب إشعارات طالب معين (للمدير أو من لديه صلاحية)"""
    
    # التحقق من الصلاحيات
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students") and not has_permission(current_user, "view_attendance"):
        user = await db.users.find_one({"_id": ObjectId(current_user["id"])})
        user_permissions = user.get("permissions", []) if user else []
        
        if Permission.SEND_NOTIFICATIONS not in user_permissions:
            raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # جلب الإشعارات
    notifications = await db.notifications.find(
        {"student_id": student_id}
    ).sort("created_at", -1).limit(50).to_list(50)
    
    result = []
    for n in notifications:
        result.append({
            "id": str(n["_id"]),
            "title": n["title"],
            "message": n["message"],
            "type": n["type"],
            "course_id": n.get("course_id"),
            "course_name": n.get("course_name"),
            "is_read": n.get("is_read", False),
            "is_manual": n.get("is_manual", False),
            "sent_by_name": n.get("sent_by_name"),
            "created_at": n["created_at"]
        })
    
    return {"notifications": result, "count": len(result)}


@api_router.delete("/students/{student_id}/notifications/{notification_id}")
async def delete_student_notification(
    student_id: str,
    notification_id: str,
    current_user: dict = Depends(get_current_user)
):
    """حذف إشعار محدد أُرسل لطالب — مخصّص لمن لديه صلاحية إرسال الإشعارات.
    
    لا يلمس endpoint المستلم (المستخدم لحذف إشعاراته الخاصة).
    """
    # نفس فحص الصلاحيات المستخدم في GET أعلاه + صلاحية SEND_NOTIFICATIONS
    if current_user["role"] != UserRole.ADMIN:
        user_doc = await db.users.find_one({"_id": ObjectId(current_user["id"])})
        user_perms = (user_doc or {}).get("permissions", [])
        if Permission.SEND_NOTIFICATIONS not in user_perms:
            raise HTTPException(status_code=403, detail="غير مصرح لك بحذف الإشعارات")
    
    result = await db.notifications.delete_one(
        {"_id": ObjectId(notification_id), "student_id": student_id}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="الإشعار غير موجود")
    return {"message": "تم حذف الإشعار", "id": notification_id}


@api_router.get("/students/notifications/unread-counts")
async def get_students_unread_notifications_counts(
    current_user: dict = Depends(get_current_user)
):
    """جلب عدد الإشعارات غير المقروءة لكل طالب (للمدير ومن لديه صلاحية SEND_NOTIFICATIONS).
    
    يُرجع: {"counts": {student_id: count, ...}, "total": number}
    """
    if current_user["role"] != UserRole.ADMIN:
        user_doc = await db.users.find_one({"_id": ObjectId(current_user["id"])})
        user_perms = (user_doc or {}).get("permissions", [])
        if Permission.SEND_NOTIFICATIONS not in user_perms:
            # نُرجع كائناً فارغاً بدل 403 لتجنّب تعطيل واجهة قائمة الطلاب
            return {"counts": {}, "total": 0}
    
    pipeline = [
        {"$match": {"is_read": {"$ne": True}, "student_id": {"$exists": True, "$ne": None}}},
        {"$group": {"_id": "$student_id", "count": {"$sum": 1}}},
    ]
    cursor = db.notifications.aggregate(pipeline)
    counts: dict = {}
    total = 0
    async for row in cursor:
        sid = row.get("_id")
        c = int(row.get("count", 0))
        if sid:
            counts[str(sid)] = c
            total += c
    return {"counts": counts, "total": total}


# ==================== Student Routes ====================

@api_router.post("/students/bulk-change-level")
async def bulk_change_level(request: Request, current_user: dict = Depends(get_current_user)):
    """تغيير مستوى مجموعة من الطلاب
    section_mode:
        - "keep"  (افتراضي): الإبقاء على شعبة كل طالب كما هي
        - "set"   : تعيين شعبة موحّدة لجميع الطلاب (new_section)
        - "clear" : إزالة الشعبة من جميع الطلاب
    """
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    data = await request.json()
    student_ids = data.get("student_ids", [])
    new_level = data.get("new_level")
    section_mode = (data.get("section_mode") or "keep").lower()
    new_section = (data.get("new_section") or "").strip()
    if not student_ids or new_level is None:
        raise HTTPException(status_code=400, detail="بيانات ناقصة")
    if section_mode not in ("keep", "set", "clear"):
        raise HTTPException(status_code=400, detail="section_mode غير صالح")
    if section_mode == "set" and not new_section:
        raise HTTPException(status_code=400, detail="يجب تحديد شعبة جديدة عند اختيار 'set'")

    update_set = {"level": int(new_level)}
    if section_mode == "set":
        update_set["section"] = new_section
    elif section_mode == "clear":
        update_set["section"] = ""

    updated = 0
    for sid in student_ids:
        result = await db.students.update_one({"_id": ObjectId(sid)}, {"$set": update_set})
        if result.modified_count > 0:
            updated += 1

    section_msg = ""
    if section_mode == "set":
        section_msg = f" - الشعبة: {new_section}"
    elif section_mode == "clear":
        section_msg = " - بلا شعبة"
    return {"message": f"تم تغيير مستوى {updated} طالب إلى المستوى {new_level}{section_msg}", "updated": updated}

@api_router.post("/students", response_model=StudentResponse)
async def create_student(student: StudentCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # ✅ السماح بتكرار رقم القيد فقط لو القسم مختلف
    existing = await db.students.find_one({
        "student_id": student.student_id,
        "department_id": student.department_id,
    })
    if existing:
        raise HTTPException(status_code=400, detail="رقم الطالب موجود مسبقاً في نفس القسم")
    
    student_dict = student.dict()
    student_dict["qr_code"] = generate_qr_code(student.student_id)
    student_dict["created_at"] = get_yemen_time()
    student_dict["is_active"] = True
    
    # تعبئة faculty_id تلقائياً من القسم
    if student_dict.get("department_id"):
        student_dict["faculty_id"] = await _resolve_faculty_id(student_dict["department_id"])

    # تعبئة program_code تلقائياً من القسم إن لم يكن محدداً
    if not student_dict.get("program_code") and student_dict.get("department_id"):
        try:
            dept = await db.departments.find_one({"_id": ObjectId(student_dict["department_id"])})
            if dept and dept.get("default_program_code"):
                student_dict["program_code"] = dept["default_program_code"]
        except Exception:
            pass

    # تعبئة enrollment_year تلقائياً من المستوى إن لم تكن محددة
    if not student_dict.get("enrollment_year") and student_dict.get("level"):
        try:
            from routes.admin_tools import compute_enrollment_year_from_level
            ey = await compute_enrollment_year_from_level(db, int(student_dict["level"]))
            if ey:
                student_dict["enrollment_year"] = ey
        except Exception:
            pass
    
    # توليد الرقم المرجعي تلقائياً (قبل إنشاء الـ user حتى نستخدمه كـ fallback)
    try:
        from routes.admin_tools import generate_reference_for_new_student
        ref = await generate_reference_for_new_student(db, student_dict)
        if ref:
            student_dict["reference_number"] = ref
    except Exception:
        pass

    # Create user account for student if password provided
    user_id = None
    if student.password:
        # ✅ لو رقم القيد مستخدم في users (بسبب طالب آخر بنفس الرقم في قسم آخر)،
        #    استخدم الرقم المرجعي كـ username
        desired_username = student.student_id
        if await db.users.find_one({"username": desired_username}):
            ref_num = student_dict.get("reference_number")
            if ref_num:
                desired_username = ref_num
            else:
                raise HTTPException(
                    status_code=400,
                    detail="رقم القيد مستخدم لطالب آخر، ولا يمكن توليد الرقم المرجعي. أكمل بيانات (الكلية/البرنامج/سنة الالتحاق) ثم أعد المحاولة."
                )
        user_data = {
            "username": desired_username,
            "password": get_password_hash(student.password),
            "full_name": student.full_name,
            "role": UserRole.STUDENT,
            "email": student.email,
            "phone": student.phone,
            "created_at": get_yemen_time(),
            "is_active": True
        }
        user_result = await db.users.insert_one(user_data)
        user_id = str(user_result.inserted_id)
    
    student_dict["user_id"] = user_id
    del student_dict["password"]

    result = await db.students.insert_one(student_dict)
    student_dict["id"] = str(result.inserted_id)

    return student_dict

@api_router.get("/students", response_model=List[StudentResponse])
async def get_students(
    department_id: Optional[str] = None,
    level: Optional[int] = None,
    section: Optional[str] = None,
    status: Optional[str] = None,
    is_active: Optional[bool] = None,
    include_alumni: bool = False,  # 🆕 افتراضياً يُستثنى الخريجون من القائمة
    include_enrollments: bool = False,  # 🚀 أداء: غير مفعّل افتراضياً لأن جلب enrollments مكلف جداً
    current_user: dict = Depends(get_current_user)
):
    # بناء فلتر الاستعلام الأساسي
    # ملاحظة: لا نفلتر على is_active لأن الطلاب المتخرجين/المفصولين/المجمَّدين
    # يجب أن يظهروا في القائمة (الفلترة تكون من واجهة المستخدم حسب الحالة)
    query: dict = {}
    if is_active is not None:
        query["is_active"] = is_active
    if status:
        query["status"] = status
    # 🆕 استثناء الخريجين من قائمة الطلاب الافتراضية
    if not include_alumni:
        query["is_alumni"] = {"$ne": True}
    
    # تطبيق فلتر النطاق بناءً على دور المستخدم
    scope_filter = await get_user_scope_filter(current_user, "students")
    query.update(scope_filter)
    
    # تطبيق الفلاتر الإضافية من المستخدم
    if department_id:
        # التحقق من أن المستخدم له حق الوصول لهذا القسم
        if "$in" in query.get("department_id", {}):
            # إذا كان هناك فلتر نطاق على الأقسام، تأكد من أن القسم المطلوب ضمنه
            if department_id in query["department_id"]["$in"]:
                query["department_id"] = department_id
            # وإلا نترك فلتر النطاق كما هو (لن يُرجع شيء من هذا القسم)
        elif not query.get("department_id"):
            query["department_id"] = department_id
    if level:
        query["level"] = level
    if section:
        query["section"] = section
    
    students = await db.students.find(query).to_list(5000)
    
    # 🚀 جلب enrollments + courses (مكلف) فقط إذا طُلب صراحة
    student_course_ids = {}
    courses_map = {}
    if include_enrollments and students:
        student_ids = [str(s["_id"]) for s in students]
        enrollments = await db.enrollments.find(
            {"student_id": {"$in": student_ids}},
            {"student_id": 1, "course_id": 1}
        ).to_list(50000)
        all_course_ids = set()
        for e in enrollments:
            sid = e["student_id"]
            cid = e["course_id"]
            student_course_ids.setdefault(sid, []).append(cid)
            all_course_ids.add(cid)
        if all_course_ids:
            course_obj_ids = []
            for cid in all_course_ids:
                try:
                    course_obj_ids.append(ObjectId(cid))
                except:
                    pass
            if course_obj_ids:
                course_docs = await db.courses.find(
                    {"_id": {"$in": course_obj_ids}},
                    {"name": 1, "code": 1}
                ).to_list(2000)
                for c in course_docs:
                    courses_map[str(c["_id"])] = {"name": c.get("name", ""), "code": c.get("code", "")}
    
    return [{
        "id": str(s["_id"]),
        "student_id": s["student_id"],
        "full_name": s["full_name"],
        "department_id": s["department_id"],
        "faculty_id": s.get("faculty_id"),
        "level": s["level"],
        "section": s["section"],
        "phone": s.get("phone"),
        "email": s.get("email"),
        "user_id": s.get("user_id"),
        "qr_code": s["qr_code"],
        "created_at": s["created_at"],
        "is_active": s.get("is_active", True),
        "status": s.get("status") or ("active" if s.get("is_active", True) else "inactive"),
        "status_changed_at": s.get("status_changed_at"),
        "status_reason": s.get("status_reason"),
        "graduation_date": s.get("graduation_date"),
        "graduated_from_level": s.get("graduated_from_level"),
        "expulsion_date": s.get("expulsion_date"),
        "frozen_at": s.get("frozen_at"),
        "reference_number": s.get("reference_number"),
        "program_code": s.get("program_code"),
        "enrollment_year": s.get("enrollment_year"),
        "enrolled_courses": [
            {"name": courses_map.get(cid, {}).get("name", ""), "code": courses_map.get(cid, {}).get("code", "")}
            for cid in student_course_ids.get(str(s["_id"]), [])
            if cid in courses_map
        ] if include_enrollments else [],
    } for s in students]

@api_router.get("/students/me", response_model=StudentResponse)
async def get_my_student_record(current_user: dict = Depends(get_current_user)):
    """الحصول على بيانات الطالب الحالي - للطلاب فقط"""
    if current_user["role"] != UserRole.STUDENT:
        raise HTTPException(status_code=403, detail="هذا الـ endpoint للطلاب فقط")
    
    # البحث عن سجل الطالب المرتبط بهذا المستخدم
    student = await db.students.find_one({"user_id": current_user["id"]})
    if not student:
        raise HTTPException(status_code=404, detail="لم يتم العثور على سجل الطالب")
    
    return {
        "id": str(student["_id"]),
        "student_id": student["student_id"],
        "full_name": student["full_name"],
        "department_id": student["department_id"],
        "level": student["level"],
        "section": student["section"],
        "phone": student.get("phone"),
        "email": student.get("email"),
        "user_id": student.get("user_id"),
        "qr_code": student["qr_code"],
        "created_at": student["created_at"],
        "is_active": student.get("is_active", True)
    }

@api_router.get("/students/me/courses")
async def get_my_courses(
    fields: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """جلب مقررات الطالب الحالي من التسجيل"""
    if current_user["role"] != UserRole.STUDENT:
        raise HTTPException(status_code=403, detail="هذا الـ endpoint للطلاب فقط")
    
    student = await db.students.find_one({"user_id": current_user["id"]})
    if not student:
        raise HTTPException(status_code=404, detail="لم يتم العثور على سجل الطالب")
    
    student_id = str(student["_id"])
    
    # الفصل النشط
    active_sem = await db.semesters.find_one({"status": "active"})
    active_sem_id = str(active_sem["_id"]) if active_sem else None
    
    # جلب التسجيلات (مفلترة بالفصل النشط إن وُجد)
    enroll_query: dict = {"student_id": student_id}
    if active_sem_id:
        enroll_query["semester_id"] = active_sem_id
    enrollments = await db.enrollments.find(enroll_query).to_list(100)
    course_ids = [e["course_id"] for e in enrollments]
    
    # إذا لا يوجد تسجيلات، جلب المقررات حسب القسم والمستوى (مفلترة بالفصل النشط)
    if not course_ids:
        query: dict = {
            "department_id": student["department_id"],
            "level": student["level"],
        }
        if active_sem_id:
            query["semester_id"] = active_sem_id
        if student.get("section"):
            query["$or"] = [
                {"section": {"$in": [None, "", student["section"]]}},
                {"section": {"$exists": False}},
            ]
        courses_list = await db.courses.find(query).to_list(100)
    else:
        courses_list = []
        for cid in course_ids:
            try:
                course = await db.courses.find_one({"_id": ObjectId(cid)})
                if course:
                    courses_list.append(course)
            except:
                pass
    
    allowed = parse_fields(fields)
    need_teacher_name = (allowed is None) or ("teacher_name" in allowed)
    need_semester_name = (allowed is None) or ("semester_name" in allowed)

    result = []
    for c in courses_list:
        teacher_name = "غير محدد"
        if need_teacher_name and c.get("teacher_id"):
            try:
                # أولاً: ابحث في collection المعلمين (الجديد)
                teacher = await db.teachers.find_one({"_id": ObjectId(c["teacher_id"])})
                if teacher:
                    teacher_name = teacher.get("full_name", "غير محدد")
                else:
                    # احتياط: collection users (القديم)
                    user_doc = await db.users.find_one({"_id": ObjectId(c["teacher_id"])})
                    if user_doc:
                        teacher_name = user_doc.get("full_name", user_doc.get("username", "غير محدد"))
            except:
                pass

        semester_name = None
        if need_semester_name and c.get("semester_id"):
            try:
                sem = await db.semesters.find_one({"_id": ObjectId(c["semester_id"])})
                if sem:
                    semester_name = sem.get("name")
            except:
                pass

        result.append({
            "id": str(c["_id"]),
            "name": c.get("name", ""),
            "code": c.get("code", ""),
            "department_id": c.get("department_id", ""),
            "teacher_id": c.get("teacher_id"),
            "teacher_name": teacher_name,
            "level": c.get("level"),
            "section": c.get("section", ""),
            "credit_hours": c.get("credit_hours", 3),
            "room": c.get("room", ""),
            "semester_id": c.get("semester_id"),
            "semester_name": semester_name,
            "academic_year": c.get("academic_year"),
        })

    return apply_fields(result, allowed)



@api_router.get("/students/{student_id}/courses")
async def get_student_courses_admin(
    student_id: str,
    current_user: dict = Depends(get_current_user),
):
    """جلب مقررات طالب محدد - فقط التسجيلات الصريحة في الفصل النشط (بدون استنتاج)."""
    # التحقق من الصلاحيات
    if current_user["role"] not in [UserRole.ADMIN] and not has_permission(current_user, "manage_students") and not has_permission(current_user, "manage_courses") and not has_permission(current_user, "manage_teachers"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    try:
        student = await db.students.find_one({"_id": ObjectId(student_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرّف الطالب غير صحيح")
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    # الفصل النشط
    active_sem = await db.semesters.find_one({"status": "active"})
    active_sem_id = str(active_sem["_id"]) if active_sem else None
    active_sem_oid = active_sem["_id"] if active_sem else None
    
    # جلب جميع تسجيلات الطالب الصريحة (بدون فلتر semester_id على enrollment
    # لأن enrollments قديمة كثيرة لها semester_id=None، لكن الـ course الخاص بها
    # موسوم بـ semester_id الصحيح)
    enrollments = await db.enrollments.find({"student_id": student_id}).to_list(500)
    course_ids = [e["course_id"] for e in enrollments if e.get("course_id")]
    
    courses_list: list = []
    if course_ids:
        # جلب المقررات لتسجيلات الطالب، مع فلترة بالفصل النشط على مستوى المقرر
        course_filter: dict = {"_id": {"$in": [ObjectId(cid) for cid in course_ids if cid]}}
        if active_sem_id:
            # يقبل string + ObjectId
            course_filter["semester_id"] = {"$in": [active_sem_id, active_sem_oid]}
        async for c in db.courses.find(course_filter):
            courses_list.append(c)
    
    # ⚠️ لا يوجد fallback / inference - إذا لا توجد تسجيلات → قائمة فارغة
    inferred = False
    
    result = []
    for c in courses_list:
        teacher_name = None
        if c.get("teacher_id"):
            try:
                t = await db.teachers.find_one({"_id": ObjectId(c["teacher_id"])})
                if t:
                    teacher_name = t.get("full_name")
            except Exception:
                pass
        sem_name = None
        if c.get("semester_id"):
            try:
                sem = await db.semesters.find_one({"_id": ObjectId(c["semester_id"])})
                if sem:
                    sem_name = sem.get("name")
            except Exception:
                pass
        result.append({
            "id": str(c["_id"]),
            "name": c.get("name", ""),
            "code": c.get("code", ""),
            "level": c.get("level"),
            "section": c.get("section", ""),
            "credit_hours": c.get("credit_hours", 3),
            "teacher_id": c.get("teacher_id"),
            "teacher_name": teacher_name,
            "semester_id": c.get("semester_id"),
            "semester_name": sem_name,
        })
    
    return {
        "student_id": student_id,
        "total_courses": len(result),
        "is_inferred": inferred,
        "courses": result,
    }


@api_router.get("/students/{student_id}", response_model=StudentResponse)
async def get_student(student_id: str, current_user: dict = Depends(get_current_user)):
    student = await db.students.find_one({"_id": ObjectId(student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    return {
        "id": str(student["_id"]),
        "student_id": student["student_id"],
        "full_name": student["full_name"],
        "department_id": student["department_id"],
        "level": student["level"],
        "section": student["section"],
        "phone": student.get("phone"),
        "email": student.get("email"),
        "user_id": student.get("user_id"),
        "qr_code": student["qr_code"],
        "created_at": student["created_at"],
        "is_active": student.get("is_active", True)
    }

@api_router.get("/students/qr/{qr_code}", response_model=StudentResponse)
async def get_student_by_qr(qr_code: str, current_user: dict = Depends(get_current_user)):
    student = await db.students.find_one({"qr_code": qr_code})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    return {
        "id": str(student["_id"]),
        "student_id": student["student_id"],
        "full_name": student["full_name"],
        "department_id": student["department_id"],
        "level": student["level"],
        "section": student["section"],
        "phone": student.get("phone"),
        "email": student.get("email"),
        "user_id": student.get("user_id"),
        "qr_code": student["qr_code"],
        "created_at": student["created_at"],
        "is_active": student.get("is_active", True)
    }

@api_router.post("/admin/cleanup-orphan-enrollments")
async def cleanup_orphan_enrollments(current_user: dict = Depends(get_current_user)):
    """
    تنظيف التسجيلات (enrollments) التي تشير لطلاب محذوفين مسبقاً (ghost enrollments).
    يُصحّح أعداد الطلاب على بطاقات المقررات.
    """
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    # جلب جميع IDs الطلاب الموجودين فعلياً
    valid_student_ids = set()
    async for s in db.students.find({}, {"_id": 1}):
        valid_student_ids.add(str(s["_id"]))

    # جلب جميع enrollments والتحقق من كل واحدة
    orphan_ids = []
    async for e in db.enrollments.find({}, {"_id": 1, "student_id": 1}):
        if e.get("student_id") not in valid_student_ids:
            orphan_ids.append(e["_id"])

    deleted = 0
    if orphan_ids:
        res = await db.enrollments.delete_many({"_id": {"$in": orphan_ids}})
        deleted = res.deleted_count

    # كذلك نظّف سجلات الحضور المعزولة
    orphan_attendance_ids = []
    async for a in db.attendance.find({}, {"_id": 1, "student_id": 1}):
        if a.get("student_id") not in valid_student_ids:
            orphan_attendance_ids.append(a["_id"])
    deleted_attendance = 0
    if orphan_attendance_ids:
        res = await db.attendance.delete_many({"_id": {"$in": orphan_attendance_ids}})
        deleted_attendance = res.deleted_count

    await log_activity(current_user, "cleanup_orphan_enrollments", "system", "", "", {
        "enrollments_deleted": deleted,
        "attendance_deleted": deleted_attendance,
    })

    return {
        "message": f"تم حذف {deleted} تسجيل و {deleted_attendance} سجل حضور غير صالحين",
        "enrollments_deleted": deleted,
        "attendance_deleted": deleted_attendance,
    }


@api_router.delete("/students/{student_id}")
async def delete_student(student_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    student = await db.students.find_one({"_id": ObjectId(student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    # حذف التسجيلات في المقررات (لمنع ghost enrollments)
    enrollments_deleted = await db.enrollments.delete_many({"student_id": student_id})
    # حذف سجلات الحضور
    attendance_deleted = await db.attendance.delete_many({"student_id": student_id})
    # حذف حساب المستخدم إن وُجد
    if student.get("user_id"):
        await db.users.delete_one({"_id": ObjectId(student["user_id"])})
    # حذف الطالب نهائياً
    await db.students.delete_one({"_id": ObjectId(student_id)})
    
    await log_activity(current_user, "delete_student", "student", student_id, student.get("full_name", ""), {
        "enrollments_deleted": enrollments_deleted.deleted_count,
        "attendance_deleted": attendance_deleted.deleted_count,
    })
    
    return {
        "message": "تم حذف الطالب بنجاح",
        "deleted": {
            "enrollments": enrollments_deleted.deleted_count,
            "attendance": attendance_deleted.deleted_count,
        }
    }

@api_router.get("/students/{student_id}/backup-info")
async def get_student_backup_info(student_id: str, current_user: dict = Depends(get_current_user)):
    """معلومات الطالب قبل الحذف"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    student = await db.students.find_one({"_id": ObjectId(student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    enrollments = await db.enrollments.find({"student_id": student_id}).to_list(100)
    course_ids = [e["course_id"] for e in enrollments]
    courses = []
    for cid in course_ids:
        c = await db.courses.find_one({"_id": ObjectId(cid)})
        if c:
            courses.append(c)
    attendance_count = await db.attendance.count_documents({"student_id": student_id})
    
    return {
        "student_name": student.get("full_name", ""),
        "student_id_num": student.get("student_id", ""),
        "courses_count": len(courses),
        "courses_names": [c.get("name", "") for c in courses],
        "attendance_count": attendance_count,
        "has_user_account": bool(student.get("user_id")),
    }

@api_router.post("/students/{student_id}/safe-delete")
async def safe_delete_student(student_id: str, current_user: dict = Depends(get_current_user)):
    """حذف آمن للطالب مع نسخة احتياطية"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    student = await db.students.find_one({"_id": ObjectId(student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    def clean_doc(doc):
        d = {}
        for k, v in doc.items():
            if k == '_id' or isinstance(v, ObjectId):
                d[k] = str(v)
            elif hasattr(v, 'isoformat'):
                d[k] = v.isoformat()
            else:
                d[k] = v
        return d
    
    enrollments = await db.enrollments.find({"student_id": student_id}).to_list(100)
    attendance = await db.attendance.find({"student_id": student_id}).to_list(50000)
    
    backup = {
        "backup_type": "student_backup",
        "backup_date": get_yemen_time().isoformat(),
        "student": clean_doc(student),
        "enrollments": [clean_doc(e) for e in enrollments],
        "attendance": [clean_doc(a) for a in attendance],
    }
    
    # حذف التسجيلات
    await db.enrollments.delete_many({"student_id": student_id})
    # حذف سجلات الحضور
    await db.attendance.delete_many({"student_id": student_id})
    # حذف حساب المستخدم
    if student.get("user_id"):
        await db.users.delete_one({"_id": ObjectId(student["user_id"])})
    # حذف الطالب
    await db.students.delete_one({"_id": ObjectId(student_id)})
    
    # حفظ في سلة المحذوفات
    await save_to_trash("student", student.get("full_name", ""), backup, current_user.get("username", "admin"))
    
    await log_activity(current_user, "safe_delete_student", "student", student_id, student.get("full_name", ""), {"enrollments": len(enrollments), "attendance": len(attendance)})
    
    return {
        "message": "تم حذف الطالب بنجاح",
        "backup": backup,
        "deleted": {"enrollments": len(enrollments), "attendance": len(attendance)}
    }

@api_router.post("/students/restore")
async def restore_student(request: Request, current_user: dict = Depends(get_current_user)):
    """استعادة طالب من نسخة احتياطية"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    data = await request.json()
    if data.get("backup_type") != "student_backup":
        raise HTTPException(status_code=400, detail="ملف النسخة الاحتياطية غير صالح")
    
    student_data = data.get("student", {})
    old_student_id = student_data.pop("_id", None)
    student_data.pop("user_id", None)
    
    # Check if student with same student_id already exists in same department (allow duplicates across departments)
    existing = await db.students.find_one({
        "student_id": student_data.get("student_id"),
        "department_id": student_data.get("department_id"),
    })
    if existing:
        raise HTTPException(status_code=400, detail=f"الطالب برقم {student_data.get('student_id')} موجود بالفعل في نفس القسم")
    
    result = await db.students.insert_one(student_data)
    new_student_id = str(result.inserted_id)
    
    # Restore enrollments
    restored_enrollments = 0
    for e in data.get("enrollments", []):
        e.pop("_id", None)
        e["student_id"] = new_student_id
        await db.enrollments.insert_one(e)
        restored_enrollments += 1
    
    # Restore attendance
    restored_attendance = 0
    for a in data.get("attendance", []):
        a.pop("_id", None)
        a["student_id"] = new_student_id
        await db.attendance.insert_one(a)
        restored_attendance += 1
    
    await log_activity(current_user, "restore_student", "student", new_student_id, student_data.get("full_name", ""))
    
    return {
        "message": "تم استعادة الطالب بنجاح",
        "new_student_id": new_student_id,
        "restored": {"enrollments": restored_enrollments, "attendance": restored_attendance}
    }

class StudentUpdate(BaseModel):
    student_id: Optional[str] = None  # رقم القيد (قابل للتعديل)
    full_name: Optional[str] = None
    department_id: Optional[str] = None
    level: Optional[int] = None
    section: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None

@api_router.put("/students/{student_id}", response_model=StudentResponse)
async def update_student(student_id: str, data: StudentUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    student = await db.students.find_one({"_id": ObjectId(student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    
    # التحقق من تفرد رقم القيد الجديد إذا تم تغييره
    new_sid = update_data.get("student_id")
    if new_sid and new_sid != student.get("student_id"):
        # ✅ تعديل رقم القيد للمدير العام (admin) فقط
        if current_user["role"] != UserRole.ADMIN:
            raise HTTPException(
                status_code=403,
                detail="تغيير رقم القيد مسموح للمدير العام فقط"
            )
        new_sid = new_sid.strip()
        # ✅ السماح بتكرار رقم القيد في قسم آخر فقط (نفس قاعدة الإنشاء)
        target_dept = update_data.get("department_id") or student.get("department_id")
        exists = await db.students.find_one({
            "student_id": new_sid,
            "department_id": target_dept,
            "_id": {"$ne": ObjectId(student_id)}
        })
        if exists:
            raise HTTPException(
                status_code=400,
                detail=f"رقم القيد '{new_sid}' مستخدم بالفعل في نفس القسم للطالب: {exists.get('full_name', '')}"
            )
        update_data["student_id"] = new_sid
        # مزامنة اسم المستخدم في جدول users (إذا كان الاسم السابق = رقم القيد)
        if student.get("user_id"):
            user = await db.users.find_one({"_id": ObjectId(student["user_id"])})
            if user and user.get("username") == student.get("student_id"):
                # تأكد أن الرقم الجديد غير مستخدم كـ username
                user_exists = await db.users.find_one({
                    "username": new_sid,
                    "_id": {"$ne": ObjectId(student["user_id"])}
                })
                if not user_exists:
                    await db.users.update_one(
                        {"_id": ObjectId(student["user_id"])},
                        {"$set": {"username": new_sid}}
                    )
    elif new_sid and new_sid == student.get("student_id"):
        # نفس القيمة - لا حاجة للتحقق ولا التحديث
        update_data.pop("student_id", None)
    
    if update_data:
        await db.students.update_one(
            {"_id": ObjectId(student_id)},
            {"$set": update_data}
        )
        await log_activity(current_user, "update_student", "student", student_id, student.get("full_name", ""), update_data)
    
    updated = await db.students.find_one({"_id": ObjectId(student_id)})
    return {
        "id": str(updated["_id"]),
        "student_id": updated["student_id"],
        "full_name": updated["full_name"],
        "department_id": updated["department_id"],
        "level": updated["level"],
        "section": updated["section"],
        "phone": updated.get("phone"),
        "email": updated.get("email"),
        "user_id": updated.get("user_id"),
        "qr_code": updated["qr_code"],
        "created_at": updated["created_at"],
        "is_active": updated.get("is_active", True)
    }

# ==================== Student Account Activation (تفعيل حساب الطالب) ====================

@api_router.post("/students/{student_id}/activate")
async def activate_student_account(student_id: str, current_user: dict = Depends(get_current_user)):
    """تفعيل حساب للطالب - الرقم الجامعي يكون اسم المستخدم وكلمة المرور الافتراضية"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بتفعيل حسابات الطلاب")
    
    # البحث عن الطالب
    student = await db.students.find_one({"_id": ObjectId(student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    # التحقق من عدم وجود حساب مسبق
    if student.get("user_id"):
        raise HTTPException(status_code=400, detail="الطالب لديه حساب مفعل مسبقاً")
    
    # تحديد اسم المستخدم: رقم القيد افتراضياً، أو الرقم المرجعي لو رقم القيد مستخدم
    desired_username = student["student_id"]
    existing_user = await db.users.find_one({"username": desired_username})
    if existing_user:
        ref_num = student.get("reference_number")
        if ref_num and not await db.users.find_one({"username": ref_num}):
            desired_username = ref_num
        else:
            raise HTTPException(
                status_code=400,
                detail="رقم القيد مستخدم لطالب آخر، ولا يوجد رقم مرجعي صالح. أكمل بيانات (الكلية/البرنامج/سنة الالتحاق) ثم أعد المحاولة."
            )
    
    # إنشاء حساب للطالب
    user_dict = {
        "username": desired_username,
        "password": get_password_hash(student["student_id"]),  # الرقم الجامعي ككلمة مرور افتراضية
        "full_name": student["full_name"],
        "role": UserRole.STUDENT,
        "email": student.get("email"),
        "phone": student.get("phone"),
        "student_record_id": str(student["_id"]),  # ربط بسجل الطالب
        "must_change_password": True,  # إجبار على تغيير كلمة المرور
        "is_active": True,
        "created_at": get_yemen_time()
    }
    
    result = await db.users.insert_one(user_dict)
    user_id = str(result.inserted_id)
    
    # ربط الحساب بسجل الطالب
    await db.students.update_one(
        {"_id": ObjectId(student_id)},
        {"$set": {"user_id": user_id}}
    )
    
    return {
        "message": "تم تفعيل حساب الطالب بنجاح",
        "username": desired_username,
        "user_id": user_id,
        "must_change_password": True
    }

@api_router.post("/students/{student_id}/deactivate")
async def deactivate_student_account(student_id: str, current_user: dict = Depends(get_current_user)):
    """إلغاء تفعيل حساب الطالب"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    student = await db.students.find_one({"_id": ObjectId(student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    if not student.get("user_id"):
        raise HTTPException(status_code=400, detail="الطالب ليس لديه حساب مفعل")
    
    # حذف حساب المستخدم
    await db.users.delete_one({"_id": ObjectId(student["user_id"])})
    
    # إزالة ربط الحساب من سجل الطالب
    await db.students.update_one(
        {"_id": ObjectId(student_id)},
        {"$unset": {"user_id": ""}}
    )
    
    return {"message": "تم إلغاء تفعيل حساب الطالب"}

@api_router.post("/students/bulk-activate")
async def bulk_activate_students(current_user: dict = Depends(get_current_user)):
    """تفعيل حسابات جميع الطلاب دفعة واحدة"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # جلب الطلاب الذين ليس لديهم حسابات
    students = await db.students.find({"user_id": {"$exists": False}}).to_list(5000)
    students += await db.students.find({"user_id": None}).to_list(5000)
    students += await db.students.find({"user_id": ""}).to_list(5000)
    
    # إزالة التكرار
    seen = set()
    unique_students = []
    for s in students:
        sid = str(s["_id"])
        if sid not in seen:
            seen.add(sid)
            unique_students.append(s)
    
    activated = 0
    failed = 0
    
    for student in unique_students:
        try:
            username = student["student_id"]
            
            # التحقق من عدم وجود حساب مسبق
            existing = await db.users.find_one({"username": username})
            if existing:
                # ربط الحساب الموجود
                await db.students.update_one(
                    {"_id": student["_id"]},
                    {"$set": {"user_id": str(existing["_id"])}}
                )
                activated += 1
                continue
            
            # إنشاء حساب جديد
            user_doc = {
                "username": username,
                "password": get_password_hash(username),
                "full_name": student.get("full_name", ""),
                "role": UserRole.STUDENT,
                "is_active": True,
                "must_change_password": True,
                "created_at": get_yemen_time()
            }
            result = await db.users.insert_one(user_doc)
            
            await db.students.update_one(
                {"_id": student["_id"]},
                {"$set": {"user_id": str(result.inserted_id)}}
            )
            activated += 1
        except Exception as e:
            failed += 1
    
    return {
        "message": f"تم تفعيل {activated} حساب طالب",
        "activated": activated,
        "failed": failed,
        "total": len(unique_students)
    }

@api_router.post("/students/bulk-deactivate")
async def bulk_deactivate_students(current_user: dict = Depends(get_current_user)):
    """إلغاء تفعيل حسابات جميع الطلاب دفعة واحدة"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # جلب الطلاب الذين لديهم حسابات
    students = await db.students.find({"user_id": {"$exists": True, "$ne": None, "$ne": ""}}).to_list(5000)
    
    deactivated = 0
    
    for student in students:
        try:
            if student.get("user_id"):
                await db.users.delete_one({"_id": ObjectId(student["user_id"])})
                await db.students.update_one(
                    {"_id": student["_id"]},
                    {"$unset": {"user_id": ""}}
                )
                deactivated += 1
        except Exception:
            pass
    
    return {
        "message": f"تم إلغاء تفعيل {deactivated} حساب طالب",
        "deactivated": deactivated
    }



@api_router.post("/students/{student_id}/reset-password")
async def reset_student_password(student_id: str, current_user: dict = Depends(get_current_user)):
    """إعادة تعيين كلمة مرور الطالب"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    student = await db.students.find_one({"_id": ObjectId(student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    if not student.get("user_id"):
        raise HTTPException(status_code=400, detail="الطالب ليس لديه حساب مفعل")
    
    # إعادة كلمة المرور للرقم الجامعي
    await db.users.update_one(
        {"_id": ObjectId(student["user_id"])},
        {"$set": {
            "password": get_password_hash(student["student_id"]),
            "must_change_password": True
        }}
    )
    
    return {
        "message": "تم إعادة تعيين كلمة المرور بنجاح",
        "new_password": student["student_id"]
    }

# ==================== Teacher Routes (إدارة المعلمين) ====================

@api_router.get("/teachers")
async def get_teachers(
    department_id: Optional[str] = None,
    cross_university: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على جميع المعلمين.
    
    - `cross_university=true`: يتجاهل فلتر النطاق ويُرجع كل أساتذة الجامعة
      لكن فقط إذا كان المستخدم admin أو لديه صلاحية `cross_university_assignment`.
    """
    # 🌐 وضع "عابر للجامعة" — يتطلب إذناً صريحاً
    role = current_user.get("role", "")
    if cross_university and (
        role == UserRole.ADMIN or has_permission(current_user, "cross_university_assignment")
    ):
        scope_filter: dict = {}  # لا قيود — كل أساتذة الجامعة
        logging.info(f"cross_university teachers fetch by '{current_user.get('username')}' (role={role})")
    else:
        # تطبيق فلتر النطاق بناءً على دور المستخدم
        scope_filter = await get_user_scope_filter(current_user, "teachers")
    query = {**scope_filter}
    
    # تطبيق فلتر القسم الإضافي
    if department_id:
        if "$in" in query.get("department_id", {}):
            if department_id in query["department_id"]["$in"]:
                query["department_id"] = department_id
        elif not query.get("department_id"):
            query["department_id"] = department_id

    # 🔧 استثناء "المعلمون العابرون للأقسام":
    # إذا كان المستخدم محدوداً بقسم/كلية، نضيف للقائمة كل معلم لديه teaching_load
    # في أحد الأقسام المسموح بها (حتى لو قسمه الإداري الأصلي مختلف).
    # هذا يضمن أن المعلم الذي يدرّس في كليتك يظهر لك حتى لو هو إدارياً في كلية أخرى.
    extra_teacher_ids: set = set()
    role = current_user.get("role", "")
    if role != UserRole.ADMIN and not department_id:
        allowed_dept_ids: list = []
        dept_q = scope_filter.get("department_id")
        if isinstance(dept_q, dict) and "$in" in dept_q:
            allowed_dept_ids = list(dept_q["$in"])
        elif isinstance(dept_q, str):
            allowed_dept_ids = [dept_q]
        if allowed_dept_ids:
            # جلب معلمي teaching_loads المرتبطين بأقسام المستخدم
            try:
                # نحتاج جلب course_ids في تلك الأقسام أولاً
                course_ids_in_scope = []
                async for c in db.courses.find(
                    {"department_id": {"$in": allowed_dept_ids}, "is_active": True},
                    {"_id": 1},
                ):
                    course_ids_in_scope.append(str(c["_id"]))
                if course_ids_in_scope:
                    async for tl in db.teaching_loads.find(
                        {"course_id": {"$in": course_ids_in_scope}},
                        {"teacher_id": 1},
                    ):
                        tid = tl.get("teacher_id")
                        if tid:
                            extra_teacher_ids.add(tid)
                    # كذلك جلب معلمي المقررات مباشرة (في حال غاب teaching_load)
                    async for c in db.courses.find(
                        {"department_id": {"$in": allowed_dept_ids}, "is_active": True, "teacher_id": {"$ne": ""}},
                        {"teacher_id": 1},
                    ):
                        tid = c.get("teacher_id")
                        if tid:
                            extra_teacher_ids.add(tid)
            except Exception:
                pass

    # دمج: إما معلمون من نفس القسم/الكلية، أو معلمون عابرون لها
    if extra_teacher_ids:
        try:
            extra_obj_ids = [ObjectId(t) for t in extra_teacher_ids]
        except Exception:
            extra_obj_ids = []
        if extra_obj_ids:
            # إذا كان query فيه شرط department_id، نحوّله إلى $or
            if "department_id" in query:
                dept_clause = {"department_id": query.pop("department_id")}
                query["$or"] = [dept_clause, {"_id": {"$in": extra_obj_ids}}]
            else:
                # لا شرط قسم → نضمّن المعلمين العابرين فقط (مع إبقاء أي فلاتر أخرى)
                query.setdefault("$or", []).append({"_id": {"$in": extra_obj_ids}})

    teachers = await db.teachers.find(query).to_list(1000)
    
    # جلب المقررات المسندة لكل المعلمين دفعة واحدة
    teacher_ids = [str(t["_id"]) for t in teachers]
    all_courses = await db.courses.find({
        "teacher_id": {"$in": teacher_ids},
        "is_active": True,
    }).to_list(5000)
    # تجميع المقررات حسب المعلم
    courses_by_teacher = {}
    for c in all_courses:
        tid = c.get("teacher_id", "")
        if tid not in courses_by_teacher:
            courses_by_teacher[tid] = []
        courses_by_teacher[tid].append({
            "name": c.get("name", ""),
            "code": c.get("code", ""),
            "section": c.get("section", ""),
            "level": c.get("level", 1),
        })
    
    # 🔧 توحيد المنطق مع /teacher-courses: نستخدم UNION (courses.teacher_id ∪ teaching_loads.course_id)
    # ضمن الفصل النشط، مع fallback لـ course.credit_hours إن لم يوجد teaching_load.
    # هذا يضمن تطابقاً 100% مع عرض /teacher-courses في عمودي "النصاب" و "المقررات".
    active_sem = await db.semesters.find_one({"status": "active"})
    active_sem_id = str(active_sem["_id"]) if active_sem else None
    active_sem_name = active_sem.get("name", "") if active_sem else ""
    current_sem_hours_by_teacher: dict = {}
    current_sem_courses_by_teacher: dict = {}
    if active_sem_id and teacher_ids:
        # 1) خريطة (teacher_id, course_id) → weekly_hours من teaching_loads النشطة
        loads_map: dict = {}  # (tid, cid) -> hours
        async for tl in db.teaching_loads.find({
            "teacher_id": {"$in": teacher_ids},
            "semester_id": active_sem_id,
        }):
            key = (tl.get("teacher_id", ""), tl.get("course_id", ""))
            loads_map[key] = tl.get("weekly_hours", 0) or 0
        # 2) خريطة (teacher_id, course_id) → credit_hours من courses النشطة في الفصل
        # هذه تغطي الحالة التي يكون فيها course.teacher_id محدّداً لكن teaching_load غائب.
        active_courses_map: dict = {}  # (tid, cid) -> credit_hours
        async for c in db.courses.find({
            "teacher_id": {"$in": teacher_ids},
            "semester_id": active_sem_id,
            "is_active": True,
        }):
            tid_c = c.get("teacher_id", "")
            cid_c = str(c["_id"])
            active_courses_map[(tid_c, cid_c)] = c.get("credit_hours", 0) or 0
        # 3) UNION: كل المفاتيح (teacher, course) من المصدرين
        union_keys: set = set(loads_map.keys()) | set(active_courses_map.keys())
        for tid_u, cid_u in union_keys:
            # الأولوية لـ teaching_load.weekly_hours، ثم credit_hours احتياطياً
            hours = loads_map.get((tid_u, cid_u))
            if hours is None or hours == 0:
                hours = active_courses_map.get((tid_u, cid_u), 0)
            current_sem_hours_by_teacher[tid_u] = current_sem_hours_by_teacher.get(tid_u, 0) + (hours or 0)
            current_sem_courses_by_teacher[tid_u] = current_sem_courses_by_teacher.get(tid_u, 0) + 1
    
    result = []
    for teacher in teachers:
        tid = str(teacher["_id"])
        teacher_dict = {
            "id": str(teacher["_id"]),
            "teacher_id": teacher.get("teacher_id", ""),
            "full_name": teacher.get("full_name", ""),
            "department_id": teacher.get("department_id"),
            "department_ids": teacher.get("department_ids", [teacher.get("department_id")] if teacher.get("department_id") else []),
            "faculty_id": teacher.get("faculty_id"),
            "email": teacher.get("email"),
            "phone": teacher.get("phone"),
            "specialization": teacher.get("specialization"),
            "academic_title": teacher.get("academic_title"),
            "user_id": teacher.get("user_id"),
            "weekly_hours": teacher.get("weekly_hours", 12),
            "teaching_load": teacher.get("teaching_load"),
            "created_at": teacher.get("created_at", get_yemen_time()),
            "is_active": teacher.get("is_active", True),
            # حقول مُحسّبة من الفصل النشط (لاتساق العرض مع صفحة /teacher-courses)
            "current_semester_id": active_sem_id,
            "current_semester_name": active_sem_name,
            "current_semester_hours": current_sem_hours_by_teacher.get(tid, 0),
            "current_semester_courses_count": current_sem_courses_by_teacher.get(tid, 0),
        }
        teacher_dict["assigned_courses"] = courses_by_teacher.get(tid, [])
        result.append(teacher_dict)
    return result

@api_router.post("/teachers", response_model=TeacherResponse)
async def create_teacher(request: Request, current_user: dict = Depends(get_current_user)):
    """إضافة معلم جديد"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_teachers"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    body = await request.json()
    
    # دعم department_ids من الفرونت اند (مصفوفة) وتحويلها إلى department_id
    if "department_ids" in body and body["department_ids"]:
        if isinstance(body["department_ids"], list) and len(body["department_ids"]) > 0:
            body["department_id"] = body["department_ids"][0]
    
    teacher = TeacherCreate(**{k: v for k, v in body.items() if k in TeacherCreate.__fields__})
    
    # التحقق من عدم تكرار الرقم الوظيفي
    existing = await db.teachers.find_one({"teacher_id": teacher.teacher_id})
    if existing:
        raise HTTPException(status_code=400, detail="الرقم الوظيفي مستخدم مسبقاً")
    
    teacher_dict = teacher.dict()
    teacher_dict["created_at"] = get_yemen_time()
    teacher_dict["is_active"] = True
    
    # حفظ department_ids كمصفوفة
    if "department_ids" in body and isinstance(body["department_ids"], list):
        teacher_dict["department_ids"] = body["department_ids"]
    elif teacher_dict.get("department_id"):
        teacher_dict["department_ids"] = [teacher_dict["department_id"]]
    
    # تعبئة faculty_id تلقائياً من القسم
    if teacher_dict.get("department_id"):
        teacher_dict["faculty_id"] = await _resolve_faculty_id(teacher_dict["department_id"])
    
    result = await db.teachers.insert_one(teacher_dict)
    teacher_dict["id"] = str(result.inserted_id)
    
    return teacher_dict

@api_router.get("/teachers/me")
async def get_my_teacher_profile(current_user: dict = Depends(get_current_user)):
    """جلب البروفايل الكامل للمعلم المسجل دخوله (للاستخدام في تطبيق المعلم).

    يُرجع: البيانات الشخصية + أسماء الأقسام + الفصل الحالي.
    """
    if current_user["role"] != UserRole.TEACHER:
        raise HTTPException(status_code=403, detail="هذا الـ endpoint للمعلمين فقط")

    # ابحث في collection teachers عبر user_id أو teacher_record_id
    teacher = None
    user_record = await db.users.find_one({"_id": ObjectId(current_user["id"])})
    teacher_record_id = (user_record or {}).get("teacher_record_id")
    if teacher_record_id:
        try:
            teacher = await db.teachers.find_one({"_id": ObjectId(teacher_record_id)})
        except Exception:
            pass
    if not teacher:
        teacher = await db.teachers.find_one({"user_id": current_user["id"]})

    if not teacher:
        raise HTTPException(status_code=404, detail="لم يتم العثور على سجل المعلم")

    # جمع معرفات الأقسام (يدعم القديم single + الجديد multiple)
    dept_ids = teacher.get("department_ids") or []
    if not dept_ids and teacher.get("department_id"):
        dept_ids = [teacher["department_id"]]

    # جلب أسماء الأقسام
    department_names: list = []
    if dept_ids:
        try:
            obj_ids = [ObjectId(d) for d in dept_ids if d]
            depts_cursor = db.departments.find({"_id": {"$in": obj_ids}})
            async for d in depts_cursor:
                department_names.append({
                    "id": str(d["_id"]),
                    "name": d.get("name", ""),
                    "faculty_id": d.get("faculty_id"),
                })
        except Exception:
            pass

    # الفصل الدراسي الفعّال
    current_semester = None
    sem = await db.semesters.find_one({"status": "active"})
    if sem:
        current_semester = {
            "id": str(sem["_id"]),
            "name": sem.get("name"),
            "academic_year": sem.get("academic_year"),
        }

    return {
        "id": str(teacher["_id"]),
        "user_id": teacher.get("user_id"),
        "teacher_id": teacher.get("teacher_id", ""),
        "full_name": teacher.get("full_name", ""),
        "username": (user_record or {}).get("username", ""),
        "email": teacher.get("email") or (user_record or {}).get("email"),
        "phone": teacher.get("phone"),
        "specialization": teacher.get("specialization"),
        "academic_title": teacher.get("academic_title"),
        "weekly_hours": teacher.get("weekly_hours", 12),
        "teaching_load": teacher.get("teaching_load"),
        "department_id": teacher.get("department_id"),
        "department_ids": dept_ids,
        "departments": department_names,
        "current_semester": current_semester,
        "is_active": teacher.get("is_active", True),
        "created_at": teacher.get("created_at"),
    }


@api_router.get("/teachers/{teacher_id}")
async def get_teacher(teacher_id: str, current_user: dict = Depends(get_current_user)):
    """الحصول على معلم محدد"""
    teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
    if not teacher:
        raise HTTPException(status_code=404, detail="المعلم غير موجود")
    
    return {
        "id": str(teacher["_id"]),
        "teacher_id": teacher.get("teacher_id", ""),
        "full_name": teacher.get("full_name", ""),
        "department_id": teacher.get("department_id"),
        "department_ids": teacher.get("department_ids", [teacher.get("department_id")] if teacher.get("department_id") else []),
        "email": teacher.get("email"),
        "phone": teacher.get("phone"),
        "specialization": teacher.get("specialization"),
        "academic_title": teacher.get("academic_title"),
        "user_id": teacher.get("user_id"),
        "weekly_hours": teacher.get("weekly_hours", 12),
        "teaching_load": teacher.get("teaching_load"),
        "created_at": teacher.get("created_at", get_yemen_time()),
        "is_active": teacher.get("is_active", True)
    }

@api_router.get("/teachers/{teacher_id}/courses")
async def get_teacher_courses(
    teacher_id: str,
    semester_id: Optional[str] = None,
    include_all: bool = False,
    current_user: dict = Depends(get_current_user),
):
    """جلب مقررات المعلم في الفصل النشط افتراضياً.
    
    Query params:
    - semester_id: لجلب مقررات فصل محدد (نشط أو مؤرشف)
    - include_all=true: لجلب كل المقررات من جميع الفصول (سلوك قديم)

    🔧 يستخدم UNION (courses.teacher_id ∪ teaching_loads.course_id) لضمان عدم
    إخفاء أي مقرر مرتبط بالمعلم سواءً عبر العلاقة المباشرة أو العبء التدريسي.
    """
    # التحقق من وجود المعلم
    teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
    if not teacher:
        raise HTTPException(status_code=404, detail="المعلم غير موجود")
    
    # تحديد الفصل المستهدف: نشط افتراضياً، أو semester_id إن مُرّر
    from routes._active_semester import (
        get_active_semester as _get_act_sem,
        get_teacher_active_course_ids as _get_tcr_ids,
        get_courses_lecture_counts as _get_lec_counts,
    )

    target_sem = None
    semester_label = ""
    if not include_all:
        if semester_id:
            try:
                _s = await db.semesters.find_one({"_id": ObjectId(semester_id)})
                if _s:
                    target_sem = {
                        "id": str(_s["_id"]),
                        "name": _s.get("name", ""),
                        "start_date": _s.get("start_date"),
                        "end_date": _s.get("end_date"),
                    }
                    semester_label = _s.get("name", "")
            except Exception:
                pass
        else:
            target_sem = await _get_act_sem(db)
            if target_sem:
                semester_label = target_sem["name"]

    target_semester_id = target_sem["id"] if target_sem else None

    # 🔧 جلب العبء التدريسي للمعلم (للحصول على weekly_hours)
    loads_query = {"teacher_id": teacher_id}
    if target_semester_id:
        loads_query["semester_id"] = target_semester_id
    teaching_loads = await db.teaching_loads.find(loads_query).to_list(500)
    loads_by_course = {tl.get("course_id"): tl.get("weekly_hours", 0) for tl in teaching_loads}
    load_id_by_course = {tl.get("course_id"): str(tl.get("_id")) for tl in teaching_loads}

    # 🔧 UNION للحصول على معرّفات المقررات
    if target_sem:
        course_ids_set = await _get_tcr_ids(db, teacher_id, target_sem)
    else:
        # include_all: ندمج كل مقررات teacher_id + كل teaching_loads
        course_ids_set = set()
        async for c in db.courses.find({"teacher_id": teacher_id}):
            course_ids_set.add(str(c["_id"]))
        for cid in loads_by_course.keys():
            if cid:
                course_ids_set.add(cid)

    if course_ids_set:
        try:
            obj_ids = [ObjectId(cid) for cid in course_ids_set]
            courses = await db.courses.find({"_id": {"$in": obj_ids}}).to_list(500)
        except Exception:
            courses = []
    else:
        courses = []

    # 🔧 حساب lecture_counts دفعةً واحدة بنفس فلتر الفصل
    course_ids_str = [str(c["_id"]) for c in courses]
    lecture_counts_map = await _get_lec_counts(db, course_ids_str, target_sem)

    result = []
    total_weekly_hours = 0
    for course in courses:
        # جلب معلومات القسم
        dept_name = ""
        faculty_name = ""
        if course.get("department_id"):
            dept = await db.departments.find_one({"_id": ObjectId(course["department_id"])})
            if dept:
                dept_name = dept.get("name", "")
                if dept.get("faculty_id"):
                    faculty = await db.faculties.find_one({"_id": ObjectId(dept["faculty_id"])})
                    faculty_name = faculty.get("name", "") if faculty else ""

        course_id_str = str(course["_id"])
        # حساب عدد الطلاب المسجلين
        students_count = await db.enrollments.count_documents({"course_id": course_id_str})
        # عدد المحاضرات (من الفلتر الموحّد)
        lectures_count = lecture_counts_map.get(course_id_str, 0)

        # 🔧 نفس منطق /api/teachers list: weekly_hours من teaching_load، fallback لـ credit_hours
        # (يشمل حالة load.weekly_hours == 0 → نأخذ credit_hours بدلاً من إظهار 0)
        load_hours = loads_by_course.get(course_id_str)
        if load_hours is None or load_hours == 0:
            weekly_hours = course.get("credit_hours", 0) or 0
        else:
            weekly_hours = load_hours
        total_weekly_hours += weekly_hours

        result.append({
            "id": course_id_str,
            "name": course["name"],
            "code": course.get("code", ""),
            "level": course.get("level", 1),
            "section": course.get("section", ""),
            "department_id": course.get("department_id"),
            "department_name": dept_name,
            "faculty_name": faculty_name,
            "students_count": students_count,
            "lectures_count": lectures_count,
            "weekly_hours": weekly_hours,
            "teaching_load_id": load_id_by_course.get(course_id_str),
            "is_active": course.get("is_active", True)
        })
    
    return {
        "teacher_id": teacher_id,
        "teacher_name": teacher.get("full_name", ""),
        "department_name": "",  # القسم الإداري
        "total_courses": len(result),
        "total_weekly_hours": total_weekly_hours,
        "semester_id": target_semester_id,
        "semester_name": semester_label,
        "courses": result
    }

@api_router.put("/teachers/{teacher_id}")
async def update_teacher(teacher_id: str, request: Request, current_user: dict = Depends(get_current_user)):
    """تحديث بيانات معلم"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_teachers"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
    if not teacher:
        raise HTTPException(status_code=404, detail="المعلم غير موجود")
    
    body = await request.json()
    
    # دعم department_ids من الفرونت اند (مصفوفة) وتحويلها إلى department_id
    if "department_ids" in body and body["department_ids"]:
        if isinstance(body["department_ids"], list) and len(body["department_ids"]) > 0:
            body["department_id"] = body["department_ids"][0]
    
    data = TeacherUpdate(**{k: v for k, v in body.items() if k in TeacherUpdate.__fields__})
    
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    
    # حفظ department_ids كمصفوفة في قاعدة البيانات
    if "department_ids" in body and isinstance(body["department_ids"], list):
        update_data["department_ids"] = body["department_ids"]
    
    # حفظ academic_title حتى لو كان فارغاً
    if "academic_title" in body:
        update_data["academic_title"] = body.get("academic_title") or ""
    
    # تحديث faculty_id إذا تغير القسم
    if "department_id" in update_data:
        update_data["faculty_id"] = await _resolve_faculty_id(update_data["department_id"])
    
    if update_data:
        await db.teachers.update_one({"_id": ObjectId(teacher_id)}, {"$set": update_data})
    
    updated = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
    return {
        "id": str(updated["_id"]),
        "teacher_id": updated.get("teacher_id", ""),
        "full_name": updated.get("full_name", ""),
        "department_id": updated.get("department_id"),
        "department_ids": updated.get("department_ids", [updated.get("department_id")] if updated.get("department_id") else []),
        "email": updated.get("email"),
        "phone": updated.get("phone"),
        "specialization": updated.get("specialization"),
        "academic_title": updated.get("academic_title"),
        "user_id": updated.get("user_id"),
        "weekly_hours": updated.get("weekly_hours", 12),
        "teaching_load": updated.get("teaching_load"),
        "is_active": updated.get("is_active", True)
    }

@api_router.get("/teachers/{teacher_id}/backup-info")
async def get_teacher_backup_info(teacher_id: str, current_user: dict = Depends(get_current_user)):
    """معلومات المعلم قبل الحذف"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_teachers"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
    if not teacher:
        raise HTTPException(status_code=404, detail="المعلم غير موجود")
    
    courses = await db.courses.find({"teacher_id": teacher_id}).to_list(100)
    course_ids = [str(c["_id"]) for c in courses]
    lectures_count = await db.lectures.count_documents({"course_id": {"$in": course_ids}}) if course_ids else 0
    attendance_count = await db.attendance.count_documents({"course_id": {"$in": course_ids}}) if course_ids else 0
    
    return {
        "teacher_name": teacher.get("full_name", ""),
        "employee_id": teacher.get("employee_id", ""),
        "courses_count": len(courses),
        "courses_names": [c.get("name", "") for c in courses],
        "lectures_count": lectures_count,
        "attendance_count": attendance_count,
        "has_user_account": bool(teacher.get("user_id")),
    }

@api_router.post("/teachers/{teacher_id}/safe-delete")
async def safe_delete_teacher(teacher_id: str, current_user: dict = Depends(get_current_user)):
    """حذف آمن للمعلم مع نسخة احتياطية"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_teachers"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
    if not teacher:
        raise HTTPException(status_code=404, detail="المعلم غير موجود")
    
    def clean_doc(doc):
        d = {}
        for k, v in doc.items():
            if k == '_id' or isinstance(v, ObjectId):
                d[k] = str(v)
            elif hasattr(v, 'isoformat'):
                d[k] = v.isoformat()
            else:
                d[k] = v
        return d
    
    # جمع نصاب التدريس والبيانات المرتبطة
    courses = await db.courses.find({"teacher_id": teacher_id}).to_list(100)
    course_ids = [str(c["_id"]) for c in courses]
    lectures = await db.lectures.find({"course_id": {"$in": course_ids}}).to_list(10000) if course_ids else []
    attendance = await db.attendance.find({"course_id": {"$in": course_ids}}).to_list(50000) if course_ids else []
    
    backup = {
        "backup_type": "teacher_backup",
        "backup_date": get_yemen_time().isoformat(),
        "teacher": clean_doc(teacher),
        "teaching_load": [clean_doc(c) for c in courses],
        "lectures": [clean_doc(l) for l in lectures],
        "attendance": [clean_doc(a) for a in attendance],
    }
    
    # إزالة ربط المعلم بالمقررات (المقررات تبقى بدون معلم)
    await db.courses.update_many({"teacher_id": teacher_id}, {"$set": {"teacher_id": None}})
    
    # حذف حساب المستخدم
    if teacher.get("user_id"):
        await db.users.delete_one({"_id": ObjectId(teacher["user_id"])})
    
    # حذف المعلم
    await db.teachers.delete_one({"_id": ObjectId(teacher_id)})
    
    # حفظ في سلة المحذوفات
    await save_to_trash("teacher", teacher.get("full_name", ""), backup, current_user.get("username", "admin"))
    
    await log_activity(current_user, "safe_delete_teacher", "teacher", teacher_id, teacher.get("full_name", ""))
    
    return {
        "message": "تم حذف المعلم بنجاح",
        "backup": backup,
        "deleted": {"courses_unlinked": len(courses), "lectures_in_backup": len(lectures), "attendance_in_backup": len(attendance)}
    }

@api_router.post("/teachers/restore")
async def restore_teacher(request: Request, current_user: dict = Depends(get_current_user)):
    """استعادة معلم من نسخة احتياطية"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_teachers"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    data = await request.json()
    if data.get("backup_type") != "teacher_backup":
        raise HTTPException(status_code=400, detail="ملف النسخة الاحتياطية غير صالح")
    
    teacher_data = data.get("teacher", {})
    old_teacher_id = teacher_data.pop("_id", None)
    teacher_data.pop("user_id", None)
    
    result = await db.teachers.insert_one(teacher_data)
    new_teacher_id = str(result.inserted_id)
    
    # إعادة ربط المقررات
    for c in data.get("teaching_load", []):
        course_code = c.get("code")
        if course_code:
            await db.courses.update_many({"code": course_code, "teacher_id": None}, {"$set": {"teacher_id": new_teacher_id}})
    
    await log_activity(current_user, "restore_teacher", "teacher", new_teacher_id, teacher_data.get("full_name", ""))
    return {"message": "تم استعادة المعلم بنجاح", "new_teacher_id": new_teacher_id}

@api_router.delete("/teachers/{teacher_id}")
async def delete_teacher(teacher_id: str, current_user: dict = Depends(get_current_user)):
    """حذف معلم"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_teachers"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
    if not teacher:
        raise HTTPException(status_code=404, detail="المعلم غير موجود")
    
    # حذف حساب المستخدم المرتبط إن وجد
    if teacher.get("user_id"):
        await db.users.delete_one({"_id": ObjectId(teacher["user_id"])})
    
    await db.teachers.delete_one({"_id": ObjectId(teacher_id)})
    await log_activity(current_user, "delete_teacher", "teacher", teacher_id, teacher.get("full_name", ""))
    return {"message": "تم حذف المعلم بنجاح"}

# ==================== Teacher Account Activation (تفعيل حساب المعلم) ====================

@api_router.post("/teachers/{teacher_id}/activate")
async def activate_teacher_account(teacher_id: str, current_user: dict = Depends(get_current_user)):
    """تفعيل حساب للمعلم - الرقم الوظيفي يكون اسم المستخدم وكلمة المرور الافتراضية"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_teachers"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بتفعيل حسابات المعلمين")
    
    teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
    if not teacher:
        raise HTTPException(status_code=404, detail="المعلم غير موجود")
    
    if teacher.get("user_id"):
        raise HTTPException(status_code=400, detail="المعلم لديه حساب مفعل مسبقاً")
    
    existing_user = await db.users.find_one({"username": teacher["teacher_id"]})
    if existing_user:
        raise HTTPException(status_code=400, detail="يوجد مستخدم بهذا الرقم الوظيفي")
    
    user_dict = {
        "username": teacher["teacher_id"],
        "password": get_password_hash(teacher["teacher_id"]),
        "full_name": teacher["full_name"],
        "role": UserRole.TEACHER,
        "email": teacher.get("email"),
        "phone": teacher.get("phone"),
        "teacher_record_id": str(teacher["_id"]),
        "must_change_password": True,
        "is_active": True,
        "created_at": get_yemen_time()
    }
    
    result = await db.users.insert_one(user_dict)
    user_id = str(result.inserted_id)
    
    await db.teachers.update_one(
        {"_id": ObjectId(teacher_id)},
        {"$set": {"user_id": user_id}}
    )
    
    return {
        "message": "تم تفعيل حساب المعلم بنجاح",
        "username": teacher["teacher_id"],
        "user_id": user_id,
        "must_change_password": True
    }

@api_router.post("/teachers/{teacher_id}/deactivate")
async def deactivate_teacher_account(teacher_id: str, current_user: dict = Depends(get_current_user)):
    """إلغاء تفعيل حساب المعلم"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_teachers"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
    if not teacher:
        raise HTTPException(status_code=404, detail="المعلم غير موجود")
    
    if not teacher.get("user_id"):
        raise HTTPException(status_code=400, detail="المعلم ليس لديه حساب مفعل")
    
    await db.users.delete_one({"_id": ObjectId(teacher["user_id"])})
    await db.teachers.update_one(
        {"_id": ObjectId(teacher_id)},
        {"$unset": {"user_id": ""}}
    )
    
    return {"message": "تم إلغاء تفعيل حساب المعلم"}

@api_router.post("/teachers/{teacher_id}/reset-password")
async def reset_teacher_password(teacher_id: str, current_user: dict = Depends(get_current_user)):
    """إعادة تعيين كلمة مرور المعلم"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_teachers"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
    if not teacher:
        raise HTTPException(status_code=404, detail="المعلم غير موجود")
    
    if not teacher.get("user_id"):
        raise HTTPException(status_code=400, detail="المعلم ليس لديه حساب مفعل")
    
    # إعادة كلمة المرور للرقم الوظيفي
    await db.users.update_one(
        {"_id": ObjectId(teacher["user_id"])},
        {"$set": {
            "password": get_password_hash(teacher["teacher_id"]),
            "must_change_password": True
        }}
    )
    
    return {
        "message": "تم إعادة تعيين كلمة المرور بنجاح",
        "new_password": teacher["teacher_id"]
    }

@api_router.post("/auth/change-password")
async def change_password(data: ChangePasswordRequest, current_user: dict = Depends(get_current_user)):
    """تغيير كلمة المرور"""
    user = await db.users.find_one({"_id": ObjectId(current_user["id"])})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # التحقق من كلمة المرور الحالية
    if not verify_password(data.current_password, user["password"]):
        raise HTTPException(status_code=400, detail="كلمة المرور الحالية غير صحيحة")
    
    # التحقق من أن كلمة المرور الجديدة مختلفة
    if data.current_password == data.new_password:
        raise HTTPException(status_code=400, detail="كلمة المرور الجديدة يجب أن تكون مختلفة")
    
    # التحقق من طول كلمة المرور
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="كلمة المرور يجب أن تكون 6 أحرف على الأقل")
    
    # تحديث كلمة المرور وإلغاء علامة تغيير كلمة المرور الإجباري
    await db.users.update_one(
        {"_id": ObjectId(current_user["id"])},
        {
            "$set": {
                "password": get_password_hash(data.new_password),
                "must_change_password": False,
                "password_changed_at": get_yemen_time()
            }
        }
    )
    
    return {"message": "تم تغيير كلمة المرور بنجاح"}

@api_router.post("/auth/force-change-password")
async def force_change_password(data: ForceChangePasswordRequest, current_user: dict = Depends(get_current_user)):
    """تغيير كلمة المرور الإجباري (عند أول دخول)"""
    user = await db.users.find_one({"_id": ObjectId(current_user["id"])})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # لا نمنع التغيير حتى لو سبق أن تم - لتجنب حالة علق التطبيق
    # if not user.get("must_change_password", False):
    #     raise HTTPException(status_code=400, detail="ليس مطلوباً منك تغيير كلمة المرور")
    
    # التحقق من طول كلمة المرور
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="كلمة المرور يجب أن تكون 6 أحرف على الأقل")
    
    # التحقق من أن كلمة المرور الجديدة مختلفة عن الرقم الجامعي
    if data.new_password == user.get("username"):
        raise HTTPException(status_code=400, detail="كلمة المرور لا يمكن أن تكون نفس الرقم الجامعي")
    
    # تحديث كلمة المرور
    await db.users.update_one(
        {"_id": ObjectId(current_user["id"])},
        {
            "$set": {
                "password": get_password_hash(data.new_password),
                "must_change_password": False,
                "password_changed_at": get_yemen_time()
            }
        }
    )
    
    return {"message": "تم تغيير كلمة المرور بنجاح", "must_change_password": False}

# ==================== Course Routes ====================


# ==================== Clone Course Section ====================

@api_router.post("/courses/{course_id}/clone-section")
async def clone_course_section(
    course_id: str,
    new_section: str = Body(..., embed=True, description="حرف الشعبة الجديدة: ب، ج، د"),
    teacher_id: Optional[str] = Body(None, embed=True, description="معلم الشعبة الجديدة (اختياري)"),
    room: Optional[str] = Body(None, embed=True, description="قاعة الشعبة الجديدة (اختياري)"),
    current_user: dict = Depends(get_current_user),
):
    """استنساخ مقرر بشعبة جديدة. يحافظ على نفس الكود الأم + يضيف لاحقة الشعبة.
    مثال: TFS201 (شعبة أ) → TFS201 (شعبة ب) - مقرر مستقل بمعلم وطلاب جدد.
    """
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح")

    new_section = (new_section or "").strip()
    if not new_section:
        raise HTTPException(status_code=400, detail="يرجى تحديد الشعبة الجديدة")
    if len(new_section) > 5:
        raise HTTPException(status_code=400, detail="حرف الشعبة طويل جداً")

    try:
        src = await db.courses.find_one({"_id": ObjectId(course_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرّف المقرر غير صحيح")
    if not src:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")

    # فحص عدم التكرار: نفس الكود الأساسي + نفس الشعبة + نفس الفصل
    # ابحث عن أي مقرر بنفس قسم + مستوى + كود + شعبة + فصل
    existing = await db.courses.find_one({
        "department_id": src.get("department_id"),
        "level": src.get("level"),
        "section": new_section,
        "semester_id": src.get("semester_id"),
        "$or": [
            {"code": src.get("code")},
            {"curriculum_course_id": src.get("curriculum_course_id")} if src.get("curriculum_course_id") else {"code": "_NEVER_MATCH_"},
        ],
        "is_active": True,
    })
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"الشعبة '{new_section}' موجودة بالفعل لهذا المقرر",
        )

    # بناء الـ doc الجديد - استنساخ مع تغيير الشعبة
    new_doc = {
        "code": src.get("code"),  # نفس الكود الأم
        "name": src.get("name"),
        "department_id": src.get("department_id"),
        "faculty_id": src.get("faculty_id"),
        "level": src.get("level"),
        "credit_hours": src.get("credit_hours", 3),
        "semester_id": src.get("semester_id"),
        "academic_year": src.get("academic_year", ""),
        "term": src.get("term"),
        "curriculum_course_id": src.get("curriculum_course_id"),
        "section": new_section,  # ← الشعبة الجديدة
        "teacher_id": teacher_id if teacher_id is not None else None,
        "room": room or "",
        "is_active": True,
        "auto_generated": False,  # ليست من توليد الخطة
        "cloned_from": str(src["_id"]),  # مرجع للأصل
        "created_at": get_yemen_time(),
        "created_by": current_user.get("id"),
    }
    result = await db.courses.insert_one(new_doc)

    # سجل النشاط
    try:
        await db.activity_logs.insert_one({
            "user_id": current_user.get("id"),
            "username": current_user.get("username"),
            "action": "clone_course_section",
            "target_type": "course",
            "target_id": str(result.inserted_id),
            "details": f"استنساخ {src.get('code')} شعبة {new_section} من {src.get('section') or 'أ'}",
            "timestamp": get_yemen_time(),
        })
    except Exception:
        pass

    return {
        "message": f"تم إنشاء شعبة '{new_section}' للمقرر {src.get('code')}",
        "id": str(result.inserted_id),
        "code": new_doc["code"],
        "section": new_doc["section"],
        "cloned_from": str(src["_id"]),
        "teacher_id": new_doc["teacher_id"],
    }


@api_router.get("/courses/{course_id}/sections")
async def get_course_sections(
    course_id: str,
    current_user: dict = Depends(get_current_user),
):
    """جلب كل الشُعب لنفس المقرر (نفس الكود + نفس الفصل)."""
    try:
        src = await db.courses.find_one({"_id": ObjectId(course_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرّف غير صحيح")
    if not src:
        raise HTTPException(status_code=404, detail="غير موجود")

    # كل المقررات بنفس الكود + الفصل + القسم + المستوى
    q = {
        "code": src.get("code"),
        "semester_id": src.get("semester_id"),
        "department_id": src.get("department_id"),
        "level": src.get("level"),
        "is_active": True,
    }
    sections = []
    async for c in db.courses.find(q):
        # عدد الطلاب
        cnt = await db.enrollments.count_documents({"course_id": str(c["_id"])})
        teacher_name = None
        if c.get("teacher_id"):
            try:
                t = await db.teachers.find_one({"_id": ObjectId(c["teacher_id"])})
                if t:
                    teacher_name = t.get("full_name")
                else:
                    u = await db.users.find_one({"_id": ObjectId(c["teacher_id"])})
                    if u:
                        teacher_name = u.get("full_name")
            except Exception:
                pass
        sections.append({
            "id": str(c["_id"]),
            "section": c.get("section") or "أ",
            "teacher_id": c.get("teacher_id"),
            "teacher_name": teacher_name,
            "room": c.get("room") or "",
            "students_count": cnt,
            "is_primary": str(c["_id"]) == course_id,
        })
    sections.sort(key=lambda s: s["section"])
    return {
        "course_code": src.get("code"),
        "course_name": src.get("name"),
        "total_sections": len(sections),
        "sections": sections,
    }



@api_router.post("/courses")
async def create_course(course: CourseCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses") and not has_permission(current_user, "add_course"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بإضافة مقرر")
    
    course_dict = course.dict()
    course_dict["created_at"] = get_yemen_time()
    course_dict["is_active"] = True
    
    # ربط تلقائي بالفصل الدراسي النشط إذا لم يتم تحديده
    if not course_dict.get("semester_id"):
        settings = await db.settings.find_one({"_id": "system_settings"})
        if settings and settings.get("current_semester_id"):
            course_dict["semester_id"] = settings["current_semester_id"]
        else:
            active_sem = await db.semesters.find_one({"status": "active"})
            if active_sem:
                course_dict["semester_id"] = str(active_sem["_id"])
    
    # تنبيه عند تعيين معلم من قسم آخر
    warning = None
    if course_dict.get("teacher_id") and course_dict.get("department_id"):
        try:
            teacher = await db.teachers.find_one({"_id": ObjectId(course_dict["teacher_id"])})
            if teacher and teacher.get("department_id") and teacher["department_id"] != course_dict["department_id"]:
                dept = await db.departments.find_one({"_id": ObjectId(teacher["department_id"])})
                teacher_dept_name = dept["name"] if dept else "غير معروف"
                warning = f"تنبيه: المعلم {teacher['full_name']} ينتمي لقسم {teacher_dept_name} وليس لقسم هذا المقرر"
        except:
            pass
    
    result = await db.courses.insert_one(course_dict)
    
    # البحث عن طلاب مطابقين (القسم + المستوى + الشعبة)
    matching_students_count = 0
    auto_enrolled_count = 0
    try:
        student_query = {
            "department_id": course_dict.get("department_id"),
            "level": course_dict.get("level"),
            "is_active": True,
        }
        if course_dict.get("section"):
            student_query["section"] = course_dict["section"]
        matching_students = await db.students.find(student_query, {"_id": 1}).to_list(10000)
        matching_students_count = len(matching_students)
        
        # تسجيل تلقائي لجميع الطلاب المطابقين
        course_id_str = str(result.inserted_id)
        for s in matching_students:
            sid = str(s["_id"])
            existing = await db.enrollments.find_one({"course_id": course_id_str, "student_id": sid})
            if not existing:
                await db.enrollments.insert_one({
                    "course_id": course_id_str,
                    "student_id": sid,
                    "enrolled_at": get_yemen_time(),
                    "enrolled_by": current_user["id"]
                })
                auto_enrolled_count += 1
    except:
        pass
    
    response = {
        "id": str(result.inserted_id),
        "name": course_dict["name"],
        "code": course_dict.get("code", ""),
        "department_id": course_dict.get("department_id", ""),
        "teacher_id": course_dict.get("teacher_id"),
        "level": course_dict.get("level", 1),
        "section": course_dict.get("section"),
        "semester": course_dict.get("semester"),
        "semester_id": course_dict.get("semester_id"),
        "academic_year": course_dict.get("academic_year"),
        "created_at": str(course_dict.get("created_at", "")),
        "is_active": course_dict.get("is_active", True),
        "matching_students_count": matching_students_count,
        "auto_enrolled_count": auto_enrolled_count,
    }
    if warning:
        response["warning"] = warning
    
    return response

@api_router.get("/courses")
async def get_courses(
    teacher_id: Optional[str] = None,
    department_id: Optional[str] = None,
    semester_id: Optional[str] = None,
    level: Optional[int] = None,
    fields: Optional[str] = None,
    all_semesters: bool = False,
    current_user: dict = Depends(get_current_user)
):
    query = {"is_active": True}
    
    # تطبيق فلتر النطاق بناءً على دور المستخدم
    scope_filter = await get_user_scope_filter(current_user, "courses")
    query.update(scope_filter)
    
    # Teachers can only see their own courses (تم تطبيقه تلقائياً في scope_filter)
    # لكن إذا جاء teacher_id من الـ query وكان المستخدم له حق الوصول:
    if current_user["role"] != UserRole.TEACHER and teacher_id:
        query["teacher_id"] = teacher_id
    
    # تطبيق فلتر القسم الإضافي
    if department_id:
        if "$in" in query.get("department_id", {}):
            if department_id in query["department_id"]["$in"]:
                query["department_id"] = department_id
        elif not query.get("department_id"):
            query["department_id"] = department_id

    # فلتر الفصل الدراسي:
    # - عند تمرير semester_id صريحاً: يُطبَّق كما هو
    # - عدا ذلك: يُفعَّل تلقائياً للمعلم/الطالب أو عند تمرير teacher_id
    #   (تطبيقات الجوال يجب أن ترى مقررات الفصل النشط فقط)
    role = current_user.get("role")
    auto_filter_active = (
        not all_semesters
        and not semester_id
        and (role in (UserRole.TEACHER, UserRole.STUDENT) or teacher_id is not None)
    )
    if semester_id:
        query["semester_id"] = semester_id
    elif auto_filter_active:
        active_sem = await db.semesters.find_one({"status": "active"})
        if active_sem:
            query["semester_id"] = str(active_sem["_id"])

    # فلتر المستوى
    if level is not None:
        query["level"] = level
    
    # تحليل ?fields= لاختصار الاستجابة
    allowed = parse_fields(fields)
    need_enrollment_count = (allowed is None) or ("students_count" in allowed)
    need_lecture_count = (allowed is None) or ("lectures_count" in allowed)
    need_teacher_name = (allowed is None) or ("teacher_name" in allowed)
    need_semester_name = (allowed is None) or ("semester_name" in allowed)

    courses = await db.courses.find(query).to_list(None)
    
    # جلب عدد الطلاب لكل مقرر دفعة واحدة (فقط إذا مطلوب)
    course_ids = [str(c["_id"]) for c in courses]
    enrollment_counts: dict = {}
    lecture_counts: dict = {}
    if course_ids and need_enrollment_count:
        pipeline = [
            {"$match": {"course_id": {"$in": course_ids}}},
            {"$group": {"_id": "$course_id", "count": {"$sum": 1}}}
        ]
        counts = await db.enrollments.aggregate(pipeline).to_list(None)
        enrollment_counts = {item["_id"]: item["count"] for item in counts}

    if course_ids and need_lecture_count:
        # 🔧 توحيد منطق الفلترة: عدد المحاضرات يجب أن يتطابق مع نطاق الفصل
        # المعروض للمقررات نفسها.
        # - إذا كان query يحتوي semester_id صريح (سواء مرّره المستخدم أو auto_filter)
        #   → نفلتر المحاضرات بنفس الفصل.
        # - إذا كان all_semesters → بلا فلتر.
        from routes._active_semester import (
            get_courses_lecture_counts as _get_lec_counts,
        )
        _target_sem = None
        _effective_sem_id = query.get("semester_id")
        if not all_semesters and _effective_sem_id:
            try:
                _s = await db.semesters.find_one({"_id": ObjectId(_effective_sem_id)})
                if _s:
                    _target_sem = {
                        "id": str(_s["_id"]),
                        "name": _s.get("name", ""),
                        "start_date": _s.get("start_date"),
                        "end_date": _s.get("end_date"),
                    }
            except Exception:
                pass
        lecture_counts = await _get_lec_counts(db, course_ids, _target_sem)
    
    # جلب أسماء المعلمين من collection المعلمين أو المستخدمين
    # 🆕 batch lookup لأسماء الأقسام والكليات
    dept_ids_set = {c.get("department_id") for c in courses if c.get("department_id")}
    dept_map: dict = {}
    fac_ids_set: set = set()
    if dept_ids_set:
        try:
            dept_obj_ids = [ObjectId(x) for x in dept_ids_set if x]
            async for d in db.departments.find(
                {"_id": {"$in": dept_obj_ids}}, {"name": 1, "faculty_id": 1}
            ):
                dept_map[str(d["_id"])] = {
                    "name": (d.get("name") or "").strip(),
                    "faculty_id": str(d.get("faculty_id") or "") or None,
                }
                if d.get("faculty_id"):
                    fac_ids_set.add(str(d["faculty_id"]))
        except Exception:
            pass
    fac_map: dict = {}
    if fac_ids_set:
        try:
            fac_obj_ids = [ObjectId(x) for x in fac_ids_set if x]
            async for f in db.faculties.find(
                {"_id": {"$in": fac_obj_ids}}, {"name": 1}
            ):
                fac_map[str(f["_id"])] = (f.get("name") or "").strip()
        except Exception:
            pass

    result = []
    for c in courses:
        teacher_name = None
        if need_teacher_name and c.get("teacher_id"):
            try:
                # أولاً: البحث في collection المعلمين (الجديد)
                teacher = await db.teachers.find_one({"_id": ObjectId(c["teacher_id"])})
                if teacher:
                    teacher_name = teacher.get("full_name")
                else:
                    # ثانياً: البحث في المستخدمين (القديم)
                    teacher = await db.users.find_one({"_id": ObjectId(c["teacher_id"])})
                    if teacher:
                        teacher_name = teacher.get("full_name")
            except:
                pass
        
        # اسم الفصل الدراسي
        semester_name = None
        if need_semester_name and c.get("semester_id"):
            try:
                sem = await db.semesters.find_one({"_id": ObjectId(c["semester_id"])})
                if sem:
                    semester_name = sem.get("name")
            except:
                pass
        
        # 🆕 اسم القسم واسم الكلية
        c_dept_id = str(c.get("department_id", "") or "")
        c_dept_info = dept_map.get(c_dept_id, {})
        c_dept_name = c_dept_info.get("name") or None
        c_fac_id = c_dept_info.get("faculty_id")
        c_fac_name = fac_map.get(c_fac_id, "") if c_fac_id else None

        result.append({
            "id": str(c["_id"]),
            "name": c.get("name", ""),
            "code": c.get("code", ""),
            "department_id": c.get("department_id", ""),
            "department_name": c_dept_name,
            "faculty_id": c_fac_id,
            "faculty_name": c_fac_name,
            "teacher_id": c.get("teacher_id"),
            "teacher_name": teacher_name,
            "level": c.get("level", 1),
            "section": c.get("section"),
            "credit_hours": c.get("credit_hours", 3),
            "room": c.get("room"),
            "semester": c.get("semester"),
            "semester_id": c.get("semester_id"),
            "semester_name": semester_name,
            "academic_year": c.get("academic_year"),
            "students_count": enrollment_counts.get(str(c["_id"]), 0),
            "lectures_count": lecture_counts.get(str(c["_id"]), 0),
            "created_at": c.get("created_at"),
            "is_active": c.get("is_active", True),
        })
    
    return apply_fields(result, allowed)

@api_router.get("/courses/{course_id}")
async def get_course(course_id: str, current_user: dict = Depends(get_current_user)):
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    # جلب اسم المعلم
    teacher_name = None
    if course.get("teacher_id"):
        try:
            teacher = await db.teachers.find_one({"_id": ObjectId(course["teacher_id"])})
            if teacher:
                teacher_name = teacher.get("full_name")
            else:
                teacher_user = await db.users.find_one({"_id": ObjectId(course["teacher_id"])})
                if teacher_user:
                    teacher_name = teacher_user.get("full_name")
        except:
            pass
    
    # جلب اسم القسم
    department_name = None
    if course.get("department_id"):
        try:
            dept = await db.departments.find_one({"_id": ObjectId(course["department_id"])})
            if dept:
                department_name = dept.get("name")
        except:
            pass
    
    # عدد الطلاب المسجلين
    students_count = await db.enrollments.count_documents({"course_id": str(course["_id"])})
    
    return {
        "id": str(course["_id"]),
        "name": course["name"],
        "code": course["code"],
        "department_id": course["department_id"],
        "department_name": department_name,
        "teacher_id": course["teacher_id"],
        "teacher_name": teacher_name,
        "level": course["level"],
        "section": course.get("section", ""),
        "credit_hours": course.get("credit_hours", 3),
        "students_count": students_count,
        "semester": course.get("semester", "الفصل الأول"),
        "semester_id": course.get("semester_id"),
        "academic_year": course.get("academic_year", "2024-2025"),
        "created_at": course["created_at"],
        "is_active": course.get("is_active", True)
    }

@api_router.get("/courses/{course_id}/lecture-stats")
async def get_course_lecture_stats(
    course_id: str,
    semester_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """إحصائيات محاضرات مقرر معين (خفيف وسريع - بدون تحميل كامل المحاضرات).

    يدعم فلتر اختياري بالفصل عبر `semester_id`.
    """
    try:
        course = await db.courses.find_one({"_id": ObjectId(course_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرف المقرر غير صالح")
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")

    # تحقق الصلاحيات: المعلم على مقرراته فقط
    if current_user["role"] == UserRole.TEACHER:
        user = await db.users.find_one({"_id": ObjectId(current_user["id"])})
        teacher_record_id = (user or {}).get("teacher_record_id") or current_user["id"]
        if str(course.get("teacher_id")) != str(teacher_record_id):
            raise HTTPException(status_code=403, detail="هذا المقرر ليس ضمن مقرراتك")
    elif current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "view_lectures") and not has_permission(current_user, "manage_lectures"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    match: dict = {"course_id": course_id}
    # 🔧 توحيد الفلترة: نطبّق فلتر فصل المقرر نفسه افتراضياً
    # (وليس الفصل النشط) كي تتطابق الأرقام مع توقع المستخدم.
    from routes._active_semester import (
        get_active_semester as _get_act_sem,
        apply_lecture_active_sem as _apply_sem,
    )
    _target_sem_doc = None
    if semester_id:
        try:
            _target_sem_doc = await db.semesters.find_one({"_id": ObjectId(semester_id)})
        except Exception:
            pass
    else:
        # افتراضياً: فصل المقرر نفسه إن وُجد، وإلا الفصل النشط
        course_sem_id = course.get("semester_id")
        if course_sem_id:
            try:
                _target_sem_doc = await db.semesters.find_one({"_id": ObjectId(course_sem_id)})
            except Exception:
                pass
        if not _target_sem_doc:
            _act = await _get_act_sem(db)
            if _act:
                _apply_sem(match, _act)
    if _target_sem_doc:
        _apply_sem(match, {
            "id": str(_target_sem_doc["_id"]),
            "name": _target_sem_doc.get("name", ""),
            "start_date": _target_sem_doc.get("start_date"),
            "end_date": _target_sem_doc.get("end_date"),
        })

    pipeline = [
        {"$match": match},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}},
    ]
    buckets = {
        LectureStatus.SCHEDULED: 0,
        LectureStatus.COMPLETED: 0,
        LectureStatus.CANCELLED: 0,
        LectureStatus.ABSENT: 0,
    }
    total = 0
    async for row in db.lectures.aggregate(pipeline):
        status = row.get("_id") or LectureStatus.SCHEDULED
        count = int(row.get("count") or 0)
        total += count
        if status in buckets:
            buckets[status] = count

    return {
        "total": total,
        "scheduled": buckets[LectureStatus.SCHEDULED],
        "completed": buckets[LectureStatus.COMPLETED],
        "cancelled": buckets[LectureStatus.CANCELLED],
        "absent": buckets[LectureStatus.ABSENT],
    }

@api_router.get("/courses/{course_id}/backup-info")
async def get_course_backup_info(course_id: str, current_user: dict = Depends(get_current_user)):
    """معلومات المقرر قبل الحذف"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses") and not has_permission(current_user, "delete_course"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    enrollments_count = await db.enrollments.count_documents({"course_id": course_id})
    lectures_count = await db.lectures.count_documents({"course_id": course_id})
    attendance_count = await db.attendance.count_documents({"course_id": course_id})
    
    return {
        "course_name": course.get("name", ""),
        "course_code": course.get("code", ""),
        "enrollments_count": enrollments_count,
        "lectures_count": lectures_count,
        "attendance_count": attendance_count,
    }

@api_router.post("/courses/{course_id}/safe-delete")
async def safe_delete_course(course_id: str, current_user: dict = Depends(get_current_user)):
    """حذف آمن للمقرر - يرجع نسخة احتياطية ثم يحذف كل شيء"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses") and not has_permission(current_user, "delete_course"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بحذف المقرر")
    
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    # 1. جمع بيانات النسخة الاحتياطية
    enrollments = await db.enrollments.find({"course_id": course_id}).to_list(10000)
    student_ids = []
    for e in enrollments:
        try:
            student_ids.append(ObjectId(e["student_id"]))
        except:
            pass
    students = await db.students.find({"_id": {"$in": student_ids}}).to_list(10000) if student_ids else []
    lectures = await db.lectures.find({"course_id": course_id}).to_list(10000)
    attendance = await db.attendance.find({"course_id": course_id}).to_list(50000)
    
    # تحويل ObjectId لـ string
    def clean_doc(doc):
        d = {k: str(v) if k == '_id' or isinstance(v, ObjectId) else v for k, v in doc.items()}
        # Convert datetime objects
        for k, v in d.items():
            if hasattr(v, 'isoformat'):
                d[k] = v.isoformat()
        return d
    
    backup = {
        "backup_type": "course_backup",
        "backup_date": get_yemen_time().isoformat(),
        "course": clean_doc(course),
        "students": [clean_doc(s) for s in students],
        "enrollments": [clean_doc(e) for e in enrollments],
        "lectures": [clean_doc(l) for l in lectures],
        "attendance": [clean_doc(a) for a in attendance],
    }
    
    # 2. إلغاء تسجيل جميع الطلاب
    await db.enrollments.delete_many({"course_id": course_id})
    
    # 3. حذف سجلات الحضور
    await db.attendance.delete_many({"course_id": course_id})
    
    # 4. حذف المحاضرات
    await db.lectures.delete_many({"course_id": course_id})
    
    # 5. حذف المقرر
    await db.courses.delete_one({"_id": ObjectId(course_id)})
    
    # حفظ في سلة المحذوفات
    await save_to_trash("course", course.get("name", ""), backup, current_user.get("username", "admin"))
    
    await log_activity(current_user, "safe_delete_course", "course", course_id, course.get("name", ""))
    
    return {
        "message": "تم حذف المقرر بنجاح",
        "backup": backup,
        "deleted": {
            "enrollments": len(enrollments),
            "lectures": len(lectures),
            "attendance": len(attendance),
        }
    }

@api_router.post("/courses/restore")
async def restore_course(request: Request, current_user: dict = Depends(get_current_user)):
    """استعادة مقرر محذوف من نسخة احتياطية"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    data = await request.json()
    
    if data.get("backup_type") != "course_backup":
        raise HTTPException(status_code=400, detail="ملف النسخة الاحتياطية غير صالح")
    
    course_data = data.get("course", {})
    if not course_data:
        raise HTTPException(status_code=400, detail="بيانات المقرر مفقودة")
    
    # 1. استعادة المقرر
    course_id = course_data.pop("_id", None)
    course_data.pop("id", None)
    result = await db.courses.insert_one(course_data)
    new_course_id = str(result.inserted_id)
    
    # 2. استعادة الطلاب (فقط غير الموجودين)
    restored_students = 0
    student_id_map = {}
    for s in data.get("students", []):
        old_id = s.pop("_id", None)
        existing = await db.students.find_one({"student_id": s.get("student_id")})
        if existing:
            student_id_map[old_id] = str(existing["_id"])
        else:
            res = await db.students.insert_one(s)
            student_id_map[old_id] = str(res.inserted_id)
            restored_students += 1
    
    # 3. استعادة التسجيلات
    restored_enrollments = 0
    for e in data.get("enrollments", []):
        e.pop("_id", None)
        e["course_id"] = new_course_id
        if e.get("student_id") in student_id_map:
            e["student_id"] = student_id_map[e["student_id"]]
        await db.enrollments.insert_one(e)
        restored_enrollments += 1
    
    # 4. استعادة المحاضرات
    restored_lectures = 0
    lecture_id_map = {}
    for l in data.get("lectures", []):
        old_id = l.pop("_id", None)
        l["course_id"] = new_course_id
        res = await db.lectures.insert_one(l)
        lecture_id_map[old_id] = str(res.inserted_id)
        restored_lectures += 1
    
    # 5. استعادة الحضور
    restored_attendance = 0
    for a in data.get("attendance", []):
        a.pop("_id", None)
        a["course_id"] = new_course_id
        if a.get("lecture_id") in lecture_id_map:
            a["lecture_id"] = lecture_id_map[a["lecture_id"]]
        if a.get("student_id") in student_id_map:
            a["student_id"] = student_id_map[a["student_id"]]
        await db.attendance.insert_one(a)
        restored_attendance += 1
    
    await log_activity(current_user, "restore_course", "course", new_course_id, course_data.get("name", ""))
    
    return {
        "message": "تم استعادة المقرر بنجاح",
        "new_course_id": new_course_id,
        "restored": {
            "students": restored_students,
            "enrollments": restored_enrollments,
            "lectures": restored_lectures,
            "attendance": restored_attendance,
        }
    }

@api_router.delete("/courses/{course_id}")
async def delete_course(course_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses") and not has_permission(current_user, "delete_course"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بحذف المقرر")

    # 🔒 ملكية القسم: لا يحذف مقرراً ليس من قسمه/كليته
    if current_user.get("role") != UserRole.ADMIN:
        course_doc = await db.courses.find_one({"_id": ObjectId(course_id)})
        if course_doc:
            user_dept = current_user.get("department_id")
            user_fac = current_user.get("faculty_id")
            c_dept = str(course_doc.get("department_id") or "")
            c_fac = str(course_doc.get("faculty_id") or "")
            if user_dept and c_dept and c_dept != str(user_dept):
                raise HTTPException(status_code=403, detail="لا يمكنك حذف مقرر لا ينتمي لقسمك")
            if not user_dept and user_fac and c_fac and c_fac != str(user_fac):
                raise HTTPException(status_code=403, detail="لا يمكنك حذف مقرر لا ينتمي لكليتك")

    # التحقق من وجود طلاب مسجلين
    enrollments_count = await db.enrollments.count_documents({"course_id": course_id})
    if enrollments_count > 0:
        raise HTTPException(status_code=400, detail=f"لا يمكن حذف المقرر - يوجد {enrollments_count} طالب مسجل. استخدم الحذف الآمن")
    
    result = await db.courses.delete_one({"_id": ObjectId(course_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    return {"message": "تم حذف المقرر بنجاح"}

# ==================== Enrollment Routes (تسجيل الطلاب في المقررات) ====================

@api_router.post("/enrollments/bulk-copy")
async def bulk_copy_students(request: Request, current_user: dict = Depends(get_current_user)):
    """نسخ طلاب إلى عدة مقررات"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_enrollments") and not has_permission(current_user, "add_enrollment"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    data = await request.json()
    student_ids = data.get("student_ids", [])
    # دعم مقرر واحد أو عدة مقررات
    target_course_ids = data.get("target_course_ids", [])
    if not target_course_ids:
        single = data.get("target_course_id")
        if single:
            target_course_ids = [single]
    if not student_ids or not target_course_ids:
        raise HTTPException(status_code=400, detail="بيانات ناقصة")
    
    total_copied = 0
    total_already = 0
    results = []
    
    for target_course_id in target_course_ids:
        target = await db.courses.find_one({"_id": ObjectId(target_course_id)})
        if not target:
            results.append({"course_id": target_course_id, "error": "غير موجود"})
            continue
        
        copied = 0
        already = 0
        for sid in student_ids:
            existing = await db.enrollments.find_one({"course_id": target_course_id, "student_id": sid})
            if existing:
                already += 1
                continue
            await db.enrollments.insert_one({
                "course_id": target_course_id,
                "student_id": sid,
                "enrolled_at": get_yemen_time(),
                "enrolled_by": current_user["id"]
            })
            copied += 1
        total_copied += copied
        total_already += already
        results.append({"course": target.get("name", ""), "copied": copied, "already": already})
    
    return {
        "message": f"تم نسخ {total_copied} طالب إلى {len(target_course_ids)} مقرر",
        "total_copied": total_copied,
        "total_already": total_already,
        "details": results
    }

@api_router.post("/enrollments/bulk-move")
async def bulk_move_students(request: Request, current_user: dict = Depends(get_current_user)):
    """نقل طلاب من مقرر إلى آخر"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_enrollments") and not has_permission(current_user, "add_enrollment"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    data = await request.json()
    student_ids = data.get("student_ids", [])
    source_course_id = data.get("source_course_id")
    # دعم مقرر واحد أو عدة مقررات
    target_course_ids = data.get("target_course_ids", [])
    if not target_course_ids:
        single = data.get("target_course_id")
        if single:
            target_course_ids = [single]
    target_course_id = target_course_ids[0] if target_course_ids else None
    if not student_ids or not source_course_id or not target_course_id:
        raise HTTPException(status_code=400, detail="بيانات ناقصة")
    target = await db.courses.find_one({"_id": ObjectId(target_course_id)})
    if not target:
        raise HTTPException(status_code=404, detail="المقرر المستهدف غير موجود")
    
    moved = 0
    already = 0
    target_section = target.get("section", "")
    for sid in student_ids:
        existing = await db.enrollments.find_one({"course_id": target_course_id, "student_id": sid})
        if existing:
            already += 1
        else:
            await db.enrollments.insert_one({
                "course_id": target_course_id,
                "student_id": sid,
                "enrolled_at": get_yemen_time(),
                "enrolled_by": current_user["id"]
            })
        await db.enrollments.delete_one({"course_id": source_course_id, "student_id": sid})
        # تحديث شعبة الطالب في سجله لتطابق المقرر المستهدف
        if target_section:
            try:
                await db.students.update_one(
                    {"_id": ObjectId(sid)},
                    {"$set": {"section": target_section}}
                )
            except:
                pass
        moved += 1
    return {"message": f"تم نقل {moved} طالب", "moved": moved, "already_enrolled": already}

@api_router.get("/enrollments/{course_id}")
async def get_course_enrollments(course_id: str, current_user: dict = Depends(get_current_user)):
    """الحصول على قائمة الطلاب المسجلين في مقرر"""
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    enrollments = await db.enrollments.find({"course_id": course_id}).to_list(10000)
    
    # Get student details
    student_ids = [ObjectId(e["student_id"]) for e in enrollments]
    students = await db.students.find({"_id": {"$in": student_ids}}).to_list(10000)
    students_map = {str(s["_id"]): s for s in students}
    
    result = []
    for enrollment in enrollments:
        student = students_map.get(enrollment["student_id"])
        if student:
            result.append({
                "enrollment_id": str(enrollment["_id"]),
                "student_id": str(student["_id"]),
                "student_number": student["student_id"],
                "full_name": student["full_name"],
                "level": student["level"],
                "section": student["section"],
                "enrolled_at": enrollment["enrolled_at"]
            })
    
    return result

@api_router.post("/courses/{course_id}/auto-enroll")
async def auto_enroll_matching_students(course_id: str, current_user: dict = Depends(get_current_user)):
    """تسجيل الطلاب المطابقين تلقائياً في المقرر بناءً على القسم والمستوى والشعبة"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses") and not has_permission(current_user, "manage_students"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    student_query = {
        "department_id": course.get("department_id"),
        "level": course.get("level"),
        "is_active": True,
    }
    if course.get("section"):
        student_query["section"] = course["section"]
    
    students = await db.students.find(student_query, {"_id": 1}).to_list(10000)
    
    enrolled = 0
    already = 0
    for s in students:
        sid = str(s["_id"])
        existing = await db.enrollments.find_one({"course_id": course_id, "student_id": sid})
        if existing:
            already += 1
            continue
        await db.enrollments.insert_one({
            "course_id": course_id,
            "student_id": sid,
            "enrolled_at": get_yemen_time(),
            "enrolled_by": current_user["id"]
        })
        enrolled += 1
    
    return {
        "message": f"تم تسجيل {enrolled} طالب في المقرر",
        "enrolled": enrolled,
        "already_enrolled": already,
        "total_matching": len(students)
    }



@api_router.get("/courses/diagnose-enrollment")
async def diagnose_enrollment(
    department_id: str,
    course_code: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تشخيص مشكلة التسجيل التلقائي"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    result = {}
    
    # 1. المقررات
    c_query = {"department_id": department_id, "is_active": True}
    if course_code:
        c_query["code"] = {"$regex": course_code, "$options": "i"}
    courses = await db.courses.find(c_query).to_list(500)
    result["courses_count"] = len(courses)
    result["courses"] = []
    
    for c in courses[:20]:
        cid = str(c["_id"])
        c_section = (c.get("section") or "").strip()
        c_level = c.get("level")
        
        # طلاب مطابقين بالضبط
        sq = {"department_id": department_id, "level": c_level, "is_active": True}
        if c_section:
            sq["section"] = c_section
        exact_count = await db.students.count_documents(sq)
        
        # طلاب بنفس المستوى (كل الشعب)
        all_level = await db.students.count_documents({"department_id": department_id, "level": c_level, "is_active": True})
        
        # تسجيلات حالية
        enroll_count = await db.enrollments.count_documents({"course_id": cid})
        
        # الشعب الموجودة لهذا المستوى
        pipeline = [
            {"$match": {"department_id": department_id, "level": c_level, "is_active": True}},
            {"$group": {"_id": "$section", "count": {"$sum": 1}}}
        ]
        sections = await db.students.aggregate(pipeline).to_list(20)
        
        result["courses"].append({
            "name": c.get("name", ""),
            "code": c.get("code", ""),
            "level": c_level,
            "section": repr(c.get("section")),
            "section_stripped": repr(c_section),
            "exact_match_students": exact_count,
            "all_level_students": all_level,
            "current_enrollments": enroll_count,
            "available_sections": [{
                "section": repr(s["_id"]),
                "count": s["count"]
            } for s in sections]
        })
    
    return result


@api_router.post("/courses/auto-enroll-all")
async def auto_enroll_all_courses(
    department_id: Optional[str] = None,
    level: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    """تسجيل تلقائي للطلاب في جميع المقررات المطابقة دفعة واحدة"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    course_query = {"is_active": True}
    if department_id:
        course_query["department_id"] = department_id
    if level:
        course_query["level"] = level
    
    courses = await db.courses.find(course_query).to_list(2000)
    
    total_enrolled = 0
    total_already = 0
    courses_processed = 0
    details = []
    
    for course in courses:
        cid = str(course["_id"])
        c_section = (course.get("section") or "").strip()
        student_query = {
            "department_id": course.get("department_id"),
            "level": course.get("level"),
            "is_active": True,
        }
        if c_section:
            # مطابقة مرنة: trim المسافات
            student_query["$or"] = [
                {"section": c_section},
                {"section": c_section + " "},
                {"section": " " + c_section},
            ]
        
        students = await db.students.find(student_query, {"_id": 1}).to_list(10000)
        if not students:
            courses_processed += 1
            continue
        
        student_ids = [str(s["_id"]) for s in students]
        
        # Batch: جلب كل التسجيلات الموجودة دفعة واحدة
        existing_enrollments = await db.enrollments.find(
            {"course_id": cid, "student_id": {"$in": student_ids}},
            {"student_id": 1}
        ).to_list(10000)
        already_enrolled_ids = {e["student_id"] for e in existing_enrollments}
        
        # فلتر الطلاب الجدد فقط
        new_ids = [sid for sid in student_ids if sid not in already_enrolled_ids]
        already = len(already_enrolled_ids)
        
        # Batch insert
        if new_ids:
            now = get_yemen_time()
            docs = [{
                "course_id": cid,
                "student_id": sid,
                "enrolled_at": now,
                "enrolled_by": current_user["id"]
            } for sid in new_ids]
            await db.enrollments.insert_many(docs)
        
        enrolled = len(new_ids)
        total_enrolled += enrolled
        total_already += already
        courses_processed += 1
        if enrolled > 0:
            details.append(f"{course.get('name', '')} ({course.get('code', '')}): +{enrolled} طالب")
    
    return {
        "message": f"تم معالجة {courses_processed} مقرر وتسجيل {total_enrolled} طالب جديد",
        "courses_processed": courses_processed,
        "total_enrolled": total_enrolled,
        "total_already_enrolled": total_already,
        "details": details,
    }


@api_router.post("/enrollments/{course_id}")
async def enroll_students(
    course_id: str, 
    data: EnrollmentCreate, 
    current_user: dict = Depends(get_current_user)
):
    """تسجيل طلاب في مقرر"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_enrollments") and not has_permission(current_user, "add_enrollment"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    course_dept = course.get("department_id", "")
    course_level = course.get("level")
    
    enrolled_count = 0
    already_enrolled = 0
    not_found = 0
    wrong_department = 0
    level_mismatch = []
    
    for student_id in data.student_ids:
        # Find student by ObjectId or student_id number
        student = None
        try:
            student = await db.students.find_one({"_id": ObjectId(student_id)})
        except:
            student = await db.students.find_one({"student_id": student_id})
        
        if not student:
            not_found += 1
            continue
        
        # منع التسجيل من قسم آخر
        if course_dept and student.get("department_id") and student["department_id"] != course_dept:
            wrong_department += 1
            continue
        
        # Check if already enrolled
        existing = await db.enrollments.find_one({
            "course_id": course_id,
            "student_id": str(student["_id"])
        })
        
        if existing:
            already_enrolled += 1
            continue
        
        # تنبيه عند اختلاف المستوى (لا يمنع)
        if course_level and student.get("level") and student["level"] != course_level:
            level_mismatch.append(f"{student.get('full_name', student.get('student_id', ''))}")
        
        # Create enrollment
        enrollment = {
            "course_id": course_id,
            "student_id": str(student["_id"]),
            "enrolled_at": get_yemen_time(),
            "enrolled_by": current_user["id"]
        }
        await db.enrollments.insert_one(enrollment)
        enrolled_count += 1
    
    message = f"تم تسجيل {enrolled_count} طالب"
    warnings = []
    if wrong_department > 0:
        warnings.append(f"تم رفض {wrong_department} طالب (من قسم آخر)")
    if level_mismatch:
        warnings.append(f"تنبيه: {len(level_mismatch)} طالب مستواهم مختلف عن المقرر ({', '.join(level_mismatch[:5])})")
    
    return {
        "message": message,
        "enrolled": enrolled_count,
        "already_enrolled": already_enrolled,
        "not_found": not_found,
        "wrong_department": wrong_department,
        "warnings": warnings
    }

@api_router.delete("/enrollments/{course_id}/{student_id}")
async def unenroll_student(
    course_id: str, 
    student_id: str, 
    current_user: dict = Depends(get_current_user)
):
    """إلغاء تسجيل طالب من مقرر"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_enrollments") and not has_permission(current_user, "delete_enrollment"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    result = await db.enrollments.delete_one({
        "course_id": course_id,
        "student_id": student_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="التسجيل غير موجود")
    
    return {"message": "تم إلغاء التسجيل بنجاح"}

@api_router.post("/enrollments/{course_id}/import")
async def import_enrollments_excel(
    course_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """استيراد طلاب إلى مقرر من ملف Excel"""
    logger.info(f"Import enrollments called: course_id={course_id}, filename={file.filename}, content_type={file.content_type}")
    
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_enrollments") and not has_permission(current_user, "add_enrollment") and not has_permission(current_user, "import_data"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    try:
        contents = await file.read()
        logger.info(f"File read successfully, size: {len(contents)} bytes")
        
        df = pd.read_excel(BytesIO(contents))
        logger.info(f"Excel parsed, columns: {list(df.columns)}, rows: {len(df)}")
        
        # Try to find student_id column
        student_id_col = None
        possible_names = ['student_id', 'رقم_الطالب', 'رقم الطالب', 'الرقم', 'id', 'ID']
        for col in df.columns:
            if col in possible_names or 'رقم' in str(col) or 'طالب' in str(col):
                student_id_col = col
                break
        
        if student_id_col is None:
            student_id_col = df.columns[0]  # Use first column
        
        enrolled_count = 0
        already_enrolled = 0
        not_found = 0
        errors = []
        
        for idx, row in df.iterrows():
            student_id_value = str(row[student_id_col]).strip()
            
            if not student_id_value or student_id_value == 'nan':
                continue
            
            # Find student
            student = await db.students.find_one({"student_id": student_id_value})
            
            if not student:
                not_found += 1
                errors.append(f"الطالب رقم {student_id_value} غير موجود")
                continue
            
            # Check if already enrolled
            existing = await db.enrollments.find_one({
                "course_id": course_id,
                "student_id": str(student["_id"])
            })
            
            if existing:
                already_enrolled += 1
                continue
            
            # Create enrollment
            enrollment = {
                "course_id": course_id,
                "student_id": str(student["_id"]),
                "enrolled_at": get_yemen_time(),
                "enrolled_by": current_user["id"]
            }
            await db.enrollments.insert_one(enrollment)
            enrolled_count += 1
        
        return {
            "message": f"تم تسجيل {enrolled_count} طالب في المقرر",
            "enrolled": enrolled_count,
            "already_enrolled": already_enrolled,
            "not_found": not_found,
            "errors": errors[:10]  # Return first 10 errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"خطأ في قراءة الملف: {str(e)}")

@api_router.get("/enrollments/{course_id}/students")
async def get_enrolled_students_for_attendance(
    course_id: str,
    fields: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """الحصول على الطلاب المسجلين للحضور"""
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    # Get enrollments
    enrollments = await db.enrollments.find({"course_id": course_id}).to_list(10000)
    
    if not enrollments:
        # If no enrollments, return empty (or could fall back to old behavior)
        return []
    
    # Get student details
    student_ids = [ObjectId(e["student_id"]) for e in enrollments]
    students = await db.students.find({
        "_id": {"$in": student_ids},
        "is_active": True
    }).to_list(10000)
    
    result = []
    for student in students:
        result.append({
            "id": str(student["_id"]),
            "student_id": student["student_id"],
            "full_name": student["full_name"],
            "department_id": student["department_id"],
            "level": student["level"],
            "section": student["section"],
            "qr_code": student.get("qr_code", "")
        })
    
    return apply_fields(result, parse_fields(fields))

class CourseUpdate(BaseModel):
    name: Optional[str] = None
    code: Optional[str] = None
    department_id: Optional[str] = None
    teacher_id: Optional[str] = None
    level: Optional[int] = None
    section: Optional[str] = None
    semester: Optional[str] = None
    academic_year: Optional[str] = None

@api_router.put("/courses/{course_id}")
async def update_course(course_id: str, data: CourseUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses") and not has_permission(current_user, "edit_course"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل المقرر")

    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")

    # 🔒 ملكية القسم: لا يعدّل مقرراً ليس من قسمه/كليته
    if current_user.get("role") != UserRole.ADMIN:
        user_dept = current_user.get("department_id")
        user_fac = current_user.get("faculty_id")
        c_dept = str(course.get("department_id") or "")
        c_fac = str(course.get("faculty_id") or "")
        if user_dept and c_dept and c_dept != str(user_dept):
            raise HTTPException(status_code=403, detail="لا يمكنك تعديل مقرر لا ينتمي لقسمك")
        if not user_dept and user_fac and c_fac and c_fac != str(user_fac):
            raise HTTPException(status_code=403, detail="لا يمكنك تعديل مقرر لا ينتمي لكليتك")

    update_data = data.dict(exclude_unset=True)
    
    # تنبيه عند تعيين معلم من قسم آخر
    warning = None
    teacher_id = update_data.get("teacher_id", course.get("teacher_id"))
    dept_id = update_data.get("department_id", course.get("department_id"))
    if teacher_id and dept_id:
        try:
            teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
            if teacher and teacher.get("department_id") and teacher["department_id"] != dept_id:
                dept = await db.departments.find_one({"_id": ObjectId(teacher["department_id"])})
                teacher_dept_name = dept["name"] if dept else "غير معروف"
                warning = f"تنبيه: المعلم {teacher['full_name']} ينتمي لقسم {teacher_dept_name} وليس لقسم هذا المقرر"
        except:
            pass
    
    if update_data:
        await db.courses.update_one(
            {"_id": ObjectId(course_id)},
            {"$set": update_data}
        )
    
    # مزامنة مع العبء التدريسي عند تغيير المعلم
    old_teacher_id = course.get("teacher_id")
    teacher_id_in_update = "teacher_id" in update_data
    new_teacher_id = update_data.get("teacher_id") if teacher_id_in_update else old_teacher_id
    if teacher_id_in_update and new_teacher_id != old_teacher_id:
        # حذف عبء المعلم القديم لهذا المقرر
        if old_teacher_id:
            await db.teaching_loads.delete_many({
                "course_id": course_id,
                "teacher_id": old_teacher_id,
            })
        # إذا المعلم الجديد ليس فارغاً، تحقق إذا عنده عبء مسبق
        if new_teacher_id:
            existing_load = await db.teaching_loads.find_one({
                "course_id": course_id,
                "teacher_id": new_teacher_id,
            })
            if not existing_load:
                # إنشاء سجل عبء تدريسي جديد بساعات المقرر المعتمدة
                credit_hours = course.get("credit_hours", 3)
                # 🔧 نُفضّل الفصل النشط دائماً حتى يظهر الإسناد فوراً في صفحة العبء التدريسي
                # (لو أخذنا semester_id من مقرر مؤرشف، لن يظهر للمستخدم)
                course_sem_id = None
                try:
                    active_sem = await db.semesters.find_one({"status": "active"})
                    if active_sem:
                        course_sem_id = str(active_sem["_id"])
                except Exception:
                    pass
                # احتياطي: إن لم يوجد فصل نشط، استخدم فصل المقرر
                if not course_sem_id:
                    course_sem_id = course.get("semester_id")
                new_load_doc = {
                    "teacher_id": new_teacher_id,
                    "course_id": course_id,
                    "weekly_hours": credit_hours,
                    "created_by": current_user["id"],
                    "created_at": datetime.now(timezone.utc),
                    "updated_at": datetime.now(timezone.utc),
                }
                if course_sem_id:
                    new_load_doc["semester_id"] = course_sem_id
                await db.teaching_loads.insert_one(new_load_doc)

    updated = await db.courses.find_one({"_id": ObjectId(course_id)})
    result = {
        "id": str(updated["_id"]),
        "name": updated["name"],
        "code": updated.get("code", ""),
        "department_id": updated.get("department_id", ""),
        "teacher_id": updated.get("teacher_id"),
        "level": updated.get("level", 1),
        "section": updated.get("section"),
        "semester": updated.get("semester"),
        "academic_year": updated.get("academic_year"),
        "created_at": updated.get("created_at"),
        "is_active": updated.get("is_active", True)
    }
    if warning:
        result["warning"] = warning
    return result

# Study Plan endpoints moved to routes/study_plans.py



@api_router.get("/reports/lesson-completion")
async def get_lesson_completion_report(
    department_id: Optional[str] = None,
    faculty_id: Optional[str] = None,
    course_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تقرير إنجاز الدروس - المخطط مقابل المنجز"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, Permission.REPORT_LESSON_COMPLETION):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # جلب المقررات
    course_query = {"is_active": True}
    if course_id:
        try:
            course_query["_id"] = ObjectId(course_id)
        except Exception:
            raise HTTPException(status_code=400, detail="معرف المقرر غير صالح")
    elif department_id:
        course_query["department_id"] = department_id
    elif faculty_id:
        depts = await db.departments.find({"faculty_id": faculty_id}).to_list(100)
        dept_ids = [str(d["_id"]) for d in depts]
        course_query["department_id"] = {"$in": dept_ids}
    
    courses = await db.courses.find(course_query).to_list(500)
    
    report = []
    for course in courses:
        cid = str(course["_id"])
        
        # جلب الخطة الدراسية (ومعها التأكيدات اليدوية)
        plan = await db.study_plans.find_one({"course_id": cid})
        planned_topics = 0
        manually_confirmed_count = 0
        confirmed_topic_ids = set()
        if plan:
            for week in plan.get("weeks", []):
                for topic in week.get("topics", []):
                    planned_topics += 1
                    if topic.get("confirmed") and topic.get("was_taught", True):
                        manually_confirmed_count += 1
                        if topic.get("id"):
                            confirmed_topic_ids.add(topic["id"])
        
        # جلب المحاضرات المكتملة مع عنوان الدرس
        completed_lectures = await db.lectures.find({
            "course_id": cid,
            "status": LectureStatus.COMPLETED
        }).sort("date", 1).to_list(500)
        
        total_lectures = await db.lectures.count_documents({"course_id": cid})
        completed_count = len(completed_lectures)
        lessons_with_title = sum(1 for l in completed_lectures if l.get("lesson_title"))

        # قائمة الدروس المنجزة فعلاً (بالعناوين والتواريخ)
        completed_lessons_list = []
        linked_topic_ids = set()
        for lec in completed_lectures:
            title = lec.get("lesson_title")
            if title:
                d = lec.get("date", "")
                if hasattr(d, "strftime"):
                    d = d.strftime("%Y-%m-%d")
                topic_id = lec.get("plan_topic_id")
                if topic_id:
                    linked_topic_ids.add(topic_id)
                completed_lessons_list.append({
                    "lesson_title": title,
                    "date": str(d),
                    "plan_topic_id": topic_id,
                })
        
        # جلب اسم المعلم
        teacher_name = "غير محدد"
        if course.get("teacher_id"):
            teacher = await db.teachers.find_one({"_id": ObjectId(course["teacher_id"])})
            if teacher:
                teacher_name = teacher["full_name"]
        
        # حساب المواضيع المُنجَزة الفعلية = (محاضرة مرتبطة) ∪ (تأكيد يدوي)
        topics_completed = len(linked_topic_ids | confirmed_topic_ids)

        # حساب نسبة الإنجاز (المواضيع المنجزة / المواضيع المخططة)
        completion_percent = 0
        if planned_topics > 0:
            completion_percent = round((topics_completed / planned_topics) * 100, 1)
        
        report.append({
            "course_id": cid,
            "course_name": course["name"],
            "course_code": course.get("code", ""),
            "teacher_name": teacher_name,
            "planned_topics": planned_topics,
            "has_plan": plan is not None and planned_topics > 0,
            "total_lectures": total_lectures,
            "completed_lectures": completed_count,
            "lessons_with_title": lessons_with_title,
            "lessons_without_title": completed_count - lessons_with_title,
            "topics_completed": topics_completed,
            "topics_linked_via_lecture": len(linked_topic_ids),
            "topics_manually_confirmed": manually_confirmed_count,
            "completion_percent": completion_percent,
            "completed_lessons": completed_lessons_list,  # قائمة الدروس المنجزة فعلاً
        })
    
    return report

@api_router.get("/export/report/lesson-completion/excel")
async def export_lesson_completion_excel(
    department_id: Optional[str] = None,
    faculty_id: Optional[str] = None,
    course_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير تقرير إنجاز الدروس كملف Excel"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, Permission.REPORT_LESSON_COMPLETION):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # جلب المقررات
    course_query = {"is_active": True}
    if course_id:
        try:
            course_query["_id"] = ObjectId(course_id)
        except Exception:
            raise HTTPException(status_code=400, detail="معرف المقرر غير صالح")
    elif department_id:
        course_query["department_id"] = department_id
    elif faculty_id:
        depts = await db.departments.find({"faculty_id": faculty_id}).to_list(100)
        dept_ids = [str(d["_id"]) for d in depts]
        course_query["department_id"] = {"$in": dept_ids}
    
    courses = await db.courses.find(course_query).to_list(500)
    
    # تصدير تفصيلي: صف لكل موضوع من الخطة + صف لكل درس منجز غير مرتبط
    detail_rows = []
    summary_rows = []

    for course in courses:
        cid = str(course["_id"])
        
        # جلب الخطة الدراسية (مع التأكيدات اليدوية)
        plan = await db.study_plans.find_one({"course_id": cid})
        
        # جلب المحاضرات المكتملة
        completed_lectures = await db.lectures.find({
            "course_id": cid,
            "status": LectureStatus.COMPLETED
        }).sort("date", 1).to_list(500)
        
        total_lectures = await db.lectures.count_documents({"course_id": cid})
        completed_count = len(completed_lectures)
        
        # خريطة: plan_topic_id -> معلومات المحاضرة المرتبطة
        lecture_by_topic = {}
        unlinked_lectures = []
        for lec in completed_lectures:
            d = lec.get("date", "")
            if hasattr(d, "strftime"):
                d = d.strftime("%Y-%m-%d")
            tid = lec.get("plan_topic_id")
            if tid:
                lecture_by_topic[tid] = {
                    "title": lec.get("lesson_title", "") or "-",
                    "date": str(d),
                }
            elif lec.get("lesson_title"):
                unlinked_lectures.append({
                    "title": lec.get("lesson_title", ""),
                    "date": str(d),
                    "lecture_id": str(lec["_id"]),
                })
        
        # اسم المعلم والقسم
        teacher_name = "غير محدد"
        dept_name = "غير محدد"
        if course.get("teacher_id"):
            teacher = await db.teachers.find_one({"_id": ObjectId(course["teacher_id"])})
            if teacher:
                teacher_name = teacher["full_name"]
        if course.get("department_id"):
            dept = await db.departments.find_one({"_id": ObjectId(course["department_id"])})
            if dept:
                dept_name = dept["name"]
        
        planned_topics = 0
        topics_completed = 0
        manually_confirmed_count = 0
        
        # ===== صفوف تفصيلية: كل موضوع من الخطة = صف واحد =====
        if plan:
            for week in plan.get("weeks", []):
                week_num = week.get("week_number", "?")
                for topic in week.get("topics", []):
                    planned_topics += 1
                    tid = topic.get("id", "")
                    planned_title = topic.get("title", "")
                    
                    # فحص المصدر
                    linked_lecture = lecture_by_topic.get(tid)
                    manually_confirmed = topic.get("confirmed") and topic.get("was_taught", True)
                    
                    if linked_lecture:
                        status = "منجز (محاضرة)"
                        actual_title = linked_lecture["title"]
                        actual_date = linked_lecture["date"]
                        topics_completed += 1
                    elif manually_confirmed:
                        status = "منجز (تأكيد يدوي)"
                        actual_title = planned_title  # الأستاذ أكّد الموضوع بالعنوان نفسه
                        actual_date = (topic.get("confirmed_date", "") or "")[:10]
                        topics_completed += 1
                        manually_confirmed_count += 1
                    else:
                        status = "غير منجز"
                        actual_title = "-"
                        actual_date = "-"
                    
                    detail_rows.append({
                        "القسم": dept_name,
                        "المقرر": course["name"],
                        "رمز المقرر": course.get("code", ""),
                        "المعلم": teacher_name,
                        "الأسبوع": week_num,
                        "الخطة (مطلوب)": planned_title,
                        "المنجز (الأستاذ)": actual_title,
                        "التاريخ": actual_date,
                        "الحالة": status,
                    })
        
        # ===== صفوف إضافية: الدروس المنجزة غير المرتبطة بأي موضوع =====
        for lec in unlinked_lectures:
            detail_rows.append({
                "القسم": dept_name,
                "المقرر": course["name"],
                "رمز المقرر": course.get("code", ""),
                "المعلم": teacher_name,
                "الأسبوع": "-",
                "الخطة (مطلوب)": "(خارج الخطة)",
                "المنجز (الأستاذ)": lec["title"],
                "التاريخ": lec["date"],
                "الحالة": "منجز خارج الخطة",
                "معرّف المحاضرة": lec.get("lecture_id", ""),
                "الإجراء المقترح": "افتح تقرير إنجاز الدروس ← اضغط المقرر ← اربط الدرس أو امسح العنوان",
            })
        
        # ===== ملخص المقرر =====
        completion_percent = round((topics_completed / planned_topics) * 100, 1) if planned_topics > 0 else 0
        summary_rows.append({
            "القسم": dept_name,
            "المقرر": course["name"],
            "رمز المقرر": course.get("code", ""),
            "المعلم": teacher_name,
            "المواضيع المخططة": planned_topics,
            "المواضيع المُنجَزة": topics_completed,
            "  - منجز بمحاضرة": len(lecture_by_topic),
            "  - مؤكَّد يدوياً": manually_confirmed_count,
            "المحاضرات الكلية": total_lectures,
            "المحاضرات المنعقدة": completed_count,
            "دروس خارج الخطة": len(unlinked_lectures),
            "نسبة الإنجاز (%)": completion_percent,
        })

    # بناء ملف Excel بشيتين: ملخص + تفصيلي
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: ملخص (صف لكل مقرر)
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, index=False, sheet_name='ملخص المقررات')
        ws = writer.sheets['ملخص المقررات']
        for col in ws.columns:
            max_length = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)
        
        # Sheet 2: تفصيلي (صف لكل موضوع)
        if detail_rows:
            df_detail = pd.DataFrame(detail_rows)
            df_detail.to_excel(writer, index=False, sheet_name='تفصيلي المواضيع')
            ws2 = writer.sheets['تفصيلي المواضيع']
            # تلوين الصفوف حسب الحالة
            from openpyxl.styles import PatternFill, Alignment
            header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
            from openpyxl.styles import Font
            header_font = Font(color='FFFFFF', bold=True)
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            blue_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            red_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
            orange_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
            
            status_col_idx = None
            for idx, cell in enumerate(ws2[1], 1):
                if cell.value == 'الحالة':
                    status_col_idx = idx
                    break
            
            if status_col_idx:
                for row_idx in range(2, ws2.max_row + 1):
                    status_val = ws2.cell(row=row_idx, column=status_col_idx).value
                    fill = None
                    if status_val == 'منجز (محاضرة)':
                        fill = green_fill
                    elif status_val == 'منجز (تأكيد يدوي)':
                        fill = blue_fill
                    elif status_val == 'غير منجز':
                        fill = red_fill
                    elif status_val == 'منجز خارج الخطة':
                        fill = orange_fill
                    if fill:
                        for col_idx in range(1, ws2.max_column + 1):
                            ws2.cell(row=row_idx, column=col_idx).fill = fill
            
            # عرض الأعمدة
            col_widths = {
                'القسم': 18, 'المقرر': 20, 'رمز المقرر': 12, 'المعلم': 22,
                'الأسبوع': 8, 'الخطة (مطلوب)': 50, 'المنجز (الأستاذ)': 50,
                'التاريخ': 12, 'الحالة': 18,
            }
            for idx, cell in enumerate(ws2[1], 1):
                name = cell.value
                if name in col_widths:
                    ws2.column_dimensions[cell.column_letter].width = col_widths[name]
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=lesson_completion_report.xlsx"}
    )

@api_router.get("/reports/lesson-completion/{course_id}/comparison")
async def get_lesson_completion_comparison(course_id: str, current_user: dict = Depends(get_current_user)):
    """مقارنة جنباً لجنب: الخطة المخططة vs الدروس المنجزة فعلاً لمقرر واحد"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, Permission.REPORT_LESSON_COMPLETION):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    try:
        course_oid = ObjectId(course_id)
    except Exception:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    course = await db.courses.find_one({"_id": course_oid})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")

    # الخطة الدراسية (نحتاج كامل بيانات الـ topic لمعرفة الـ confirmed يدوياً)
    plan = await db.study_plans.find_one({"course_id": course_id})
    plan_topics = []
    if plan:
        for week in plan.get("weeks", []):
            for topic in week.get("topics", []):
                plan_topics.append({
                    "id": topic.get("id"),
                    "week_number": week.get("week_number"),
                    "title": topic.get("title", ""),
                    # تأكيد يدوي من المعلم (عبر /confirm-topics)
                    "confirmed": bool(topic.get("confirmed")),
                    "was_taught": bool(topic.get("was_taught", True)),
                    "confirmed_date": topic.get("confirmed_date", ""),
                })

    # المحاضرات المكتملة
    completed_lectures = await db.lectures.find({
        "course_id": course_id,
        "status": LectureStatus.COMPLETED,
    }).sort("date", 1).to_list(500)

    completed_lessons = []
    matched_topic_ids = set()
    for lec in completed_lectures:
        d = lec.get("date", "")
        if hasattr(d, "strftime"):
            d = d.strftime("%Y-%m-%d")
        topic_id = lec.get("plan_topic_id")
        if topic_id:
            matched_topic_ids.add(topic_id)
        completed_lessons.append({
            "lecture_id": str(lec["_id"]),
            "date": str(d),
            "lesson_title": lec.get("lesson_title", ""),
            "plan_topic_id": topic_id,
            "notes": lec.get("notes", ""),
        })

    # بناء المقارنة: كل موضوع مخطط + هل تم إنجازه + العنوان الذي كتبه الأستاذ
    comparison = []
    for topic in plan_topics:
        matched = None
        for lesson in completed_lessons:
            if lesson.get("plan_topic_id") == topic["id"]:
                matched = lesson
                break
        # Manual confirmation fallback
        manually_confirmed = topic.get("confirmed") and topic.get("was_taught", True)
        is_completed = matched is not None or manually_confirmed
        comparison.append({
            "week_number": topic["week_number"],
            "topic_id": topic["id"],
            "planned_title": topic["title"],
            "completed": is_completed,
            "actual_title": matched["lesson_title"] if matched else (topic["title"] if manually_confirmed else None),
            "actual_date": matched["date"] if matched else (topic.get("confirmed_date", "")[:10] if manually_confirmed else None),
            "source": "lecture" if matched else ("manual_confirm" if manually_confirmed else None),
            "lecture_id": matched.get("lecture_id") if matched else None,
        })

    # الدروس المنجزة التي لا ترتبط بأي موضوع من الخطة
    unlinked_lessons = [l for l in completed_lessons if not l.get("plan_topic_id")]

    # احسب أيضاً "مؤكَّد يدوياً" — المواضيع التي أُكدت يدوياً بدون محاضرة مرتبطة
    manually_confirmed_count = sum(
        1 for t in plan_topics
        if t.get("confirmed") and t.get("was_taught", True) and t["id"] not in matched_topic_ids
    )

    teacher_name = "غير محدد"
    if course.get("teacher_id"):
        teacher = await db.teachers.find_one({"_id": ObjectId(course["teacher_id"])})
        if teacher:
            teacher_name = teacher["full_name"]

    return {
        "course_id": course_id,
        "course_name": course["name"],
        "course_code": course.get("code", ""),
        "teacher_name": teacher_name,
        "total_planned": len(plan_topics),
        "total_completed": len(completed_lessons),
        "matched_count": len(matched_topic_ids),
        "manually_confirmed_count": manually_confirmed_count,
        "unmatched_count": len(unlinked_lessons),
        "comparison": comparison,
        "unlinked_lessons": unlinked_lessons,
    }


@api_router.get("/admin/backfill-sections/preview")
async def preview_backfill_sections(current_user: dict = Depends(get_current_user)):
    """فحص المقررات التي اسمها يحتوي شعبة بين قوسين لكن حقل section فارغ.
    + يكتشف المقررات بدون شعبة في القاعدة لكن لها أخوات بنفس الاسم/المستوى/القسم بشُعب أ ب ج → يقترح الحرف التالي."""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="للمدير فقط")
    import re
    pattern = re.compile(r"\(([^)]+)\)\s*$")
    SECTION_LETTERS = ["أ", "ب", "ج", "د", "هـ", "و", "ز", "ح", "ط", "ي"]

    candidates = []
    total_courses = 0
    # اجمع أولاً كل المقررات لتحليل الأخوة
    all_courses = await db.courses.find({}).to_list(5000)
    total_courses = len(all_courses)

    # خريطة: (department_id, level, base_name) -> {existing_sections: set, courses_no_section: list}
    siblings_map: dict = {}

    def base_name_of(name: str) -> str:
        # أزل اللواحق "(...)" من الاسم
        return pattern.sub("", name).strip()

    for c in all_courses:
        name = c.get("name", "")
        section_val = (c.get("section") or "").strip()
        key = (
            str(c.get("department_id", "")),
            int(c.get("level", 1) or 1),
            base_name_of(name),
        )
        if key not in siblings_map:
            siblings_map[key] = {"existing_sections": set(), "courses_no_section": []}
        if section_val:
            siblings_map[key]["existing_sections"].add(section_val)
        else:
            siblings_map[key]["courses_no_section"].append(c)

    # الآن نمر على المقررات بدون شعبة ونقترح
    for c in all_courses:
        section_val = (c.get("section") or "").strip()
        if section_val:
            continue
        name = c.get("name", "")
        # طريقة 1: استخراج من قوسين في الاسم
        m = pattern.search(name)
        suggested = None
        reason = None
        if m:
            extracted = m.group(1).strip()
            if 1 <= len(extracted) <= 10:
                suggested = extracted
                reason = "من قوسين الاسم"

        # طريقة 2: اقتراح من الأخوة (نفس الاسم بدون قوسين/المستوى/القسم)
        if not suggested:
            key = (
                str(c.get("department_id", "")),
                int(c.get("level", 1) or 1),
                base_name_of(name),
            )
            sibs = siblings_map.get(key, {})
            existing = sibs.get("existing_sections", set())
            if existing:
                # اقترح أول حرف غير مستخدم
                for letter in SECTION_LETTERS:
                    if letter not in existing:
                        suggested = letter
                        reason = f"المقررات الأخرى لها شُعب: {', '.join(sorted(existing))}"
                        break

        if suggested:
            candidates.append({
                "course_id": str(c["_id"]),
                "current_name": name,
                "extracted_section": suggested,
                "reason": reason,
            })

    return {
        "total_courses": total_courses,
        "candidates_count": len(candidates),
        "candidates": candidates[:50],
        "truncated": len(candidates) > 50,
    }


@api_router.post("/admin/backfill-sections")
async def backfill_sections(current_user: dict = Depends(get_current_user)):
    """تعبئة الشعبة من اسم المقرر أو من أخوته.
    - الأولوية للقوسين في الاسم
    - وإلا: اقتراح الحرف الناقص بناءً على أخوة المقرر"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="للمدير فقط")
    import re
    pattern = re.compile(r"\(([^)]+)\)\s*$")
    SECTION_LETTERS = ["أ", "ب", "ج", "د", "هـ", "و", "ز", "ح", "ط", "ي"]

    all_courses = await db.courses.find({}).to_list(5000)
    siblings_map: dict = {}

    def base_name_of(name: str) -> str:
        return pattern.sub("", name).strip()

    for c in all_courses:
        section_val = (c.get("section") or "").strip()
        key = (
            str(c.get("department_id", "")),
            int(c.get("level", 1) or 1),
            base_name_of(c.get("name", "")),
        )
        if key not in siblings_map:
            siblings_map[key] = set()
        if section_val:
            siblings_map[key].add(section_val)

    updated = 0
    for c in all_courses:
        if (c.get("section") or "").strip():
            continue
        name = c.get("name", "")
        suggested = None
        m = pattern.search(name)
        if m:
            extracted = m.group(1).strip()
            if 1 <= len(extracted) <= 10:
                suggested = extracted
        if not suggested:
            key = (
                str(c.get("department_id", "")),
                int(c.get("level", 1) or 1),
                base_name_of(name),
            )
            existing = siblings_map.get(key, set())
            if existing:
                for letter in SECTION_LETTERS:
                    if letter not in existing:
                        suggested = letter
                        # احجز الحرف لتجنب تكراره في مقررات أخوة بنفس الـkey
                        existing.add(letter)
                        break
        if suggested:
            await db.courses.update_one(
                {"_id": c["_id"]},
                {"$set": {"section": suggested}}
            )
            updated += 1

    await log_activity(
        user=current_user,
        action="backfill_sections",
        entity_type="course",
        entity_id=None,
        entity_name="Backfill sections (smart)",
        details={"updated": updated},
    )
    return {
        "message": f"تم تحديث الشعبة في {updated} مقرر",
        "updated": updated,
    }


# ==================== Lecture Routes (المحاضرات/الحصص) ====================

@api_router.get("/lectures/today")
async def get_today_lectures(
    fields: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """الحصول على محاضرات اليوم للمعلم"""
    today = get_yemen_time().strftime("%Y-%m-%d")
    
    # جلب مقررات المعلم
    course_query = {"is_active": True}
    
    if current_user["role"] == UserRole.TEACHER:
        # البحث عن teacher_record_id في حساب المستخدم
        user = await db.users.find_one({"_id": ObjectId(current_user["id"])})
        if user and user.get("teacher_record_id"):
            course_query["teacher_id"] = user["teacher_record_id"]
        else:
            course_query["teacher_id"] = current_user["id"]
    elif current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "view_lectures") and not has_permission(current_user, "manage_lectures"):
        # للمستخدمين غير المعلمين والمدير
        return []
    
    courses = await db.courses.find(course_query).to_list(100)
    course_ids = [str(c["_id"]) for c in courses]
    course_map = {str(c["_id"]): c for c in courses}
    
    if not course_ids:
        return []
    
    # جلب محاضرات اليوم لهذه المقررات
    lectures = await db.lectures.find({
        "course_id": {"$in": course_ids},
        "date": today
    }).sort("start_time", 1).to_list(100)
    
    # تحديث تلقائي: المحاضرات المجدولة التي انتهى وقتها بدون تحضير → غائب
    now = get_yemen_time()
    for lecture in lectures:
        if lecture.get("status") == LectureStatus.SCHEDULED and not lecture.get("status_override"):
            try:
                lecture_end = datetime.strptime(f"{lecture['date']} {lecture['end_time']}", "%Y-%m-%d %H:%M")
                lecture_end = lecture_end.replace(tzinfo=YEMEN_TIMEZONE)
                lecture_start = datetime.strptime(f"{lecture['date']} {lecture['start_time']}", "%Y-%m-%d %H:%M")
                lecture_start = lecture_start.replace(tzinfo=YEMEN_TIMEZONE)
                duration = (lecture_end - lecture_start).total_seconds() / 60
                allowed_end = lecture_end + timedelta(minutes=duration)
                if now > allowed_end:
                    await db.lectures.update_one(
                        {"_id": lecture["_id"]},
                        {"$set": {"status": LectureStatus.ABSENT}}
                    )
                    lecture["status"] = LectureStatus.ABSENT
            except:
                pass

    result = []
    allowed = parse_fields(fields)
    need_attendance = (allowed is None) or ("attendance_count" in allowed) or ("total_enrolled" in allowed)

    # 🆕 batch lookup لأسماء الأقسام والكليات
    _dept_ids = {course_map[cid].get("department_id") for cid in course_map if course_map[cid].get("department_id")}
    _dept_map: dict = {}
    _fac_ids: set = set()
    if _dept_ids:
        try:
            _oids = [ObjectId(x) for x in _dept_ids if x]
            async for d in db.departments.find({"_id": {"$in": _oids}}, {"name": 1, "faculty_id": 1}):
                _dept_map[str(d["_id"])] = {"name": (d.get("name") or "").strip(), "faculty_id": str(d.get("faculty_id") or "") or None}
                if d.get("faculty_id"): _fac_ids.add(str(d["faculty_id"]))
        except Exception:
            pass
    _fac_map: dict = {}
    if _fac_ids:
        try:
            _foids = [ObjectId(x) for x in _fac_ids if x]
            async for f in db.faculties.find({"_id": {"$in": _foids}}, {"name": 1}):
                _fac_map[str(f["_id"])] = (f.get("name") or "").strip()
        except Exception:
            pass

    for lecture in lectures:
        course = course_map.get(lecture["course_id"], {})
        # حساب عدد الطلاب الحاضرين (فقط إذا مطلوب)
        attendance_count = 0
        total_enrolled = 0
        if need_attendance:
            attendance_count = await db.attendance.count_documents({
                "lecture_id": str(lecture["_id"]),
                "status": "present"
            })
            total_enrolled = await db.enrollments.count_documents({"course_id": lecture["course_id"]})

        # 🆕 اسم القسم والكلية
        _cdept_id = str(course.get("department_id", "") or "")
        _cdept_info = _dept_map.get(_cdept_id, {})
        _cdept_name = _cdept_info.get("name") or None
        _cfac_id = _cdept_info.get("faculty_id")
        _cfac_name = _fac_map.get(_cfac_id, "") if _cfac_id else None

        result.append({
            "id": str(lecture["_id"]),
            "course_id": lecture["course_id"],
            "course_name": course.get("name", ""),
            "course_code": course.get("code", ""),
            "section": course.get("section", ""),
            "department_id": _cdept_id or None,
            "department_name": _cdept_name,
            "faculty_id": _cfac_id,
            "faculty_name": _cfac_name,
            "date": lecture["date"],
            "start_time": lecture["start_time"],
            "end_time": lecture["end_time"],
            "room": lecture.get("room", ""),
            "status": lecture.get("status", LectureStatus.SCHEDULED),
            "notes": lecture.get("notes", ""),
            "attendance_count": attendance_count,
            "total_enrolled": total_enrolled,
            "created_at": lecture["created_at"]
        })

    return apply_fields(result, allowed)


@api_router.get("/lectures/all-schedule")
async def get_all_schedule_lectures(
    date: str = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب محاضرات يوم محدد فقط - للجدول الأسبوعي"""
    
    # إذا لم يُحدد تاريخ، نستخدم اليوم
    if not date:
        now = get_yemen_time()
        date = now.strftime("%Y-%m-%d")
    
    course_query = {"is_active": True}
    scope_filter = await get_user_scope_filter(current_user)
    if scope_filter:
        course_query.update(scope_filter)
    
    if current_user["role"] == UserRole.TEACHER:
        user = await db.users.find_one({"_id": ObjectId(current_user["id"])})
        if user and user.get("teacher_record_id"):
            course_query["teacher_id"] = user["teacher_record_id"]
        else:
            course_query["teacher_id"] = current_user["id"]
    
    courses = await db.courses.find(course_query).to_list(500)
    course_ids = [str(c["_id"]) for c in courses]
    course_map = {str(c["_id"]): c for c in courses}
    
    if not course_ids:
        return {"lectures": [], "date": date}
    
    # جلب محاضرات اليوم المحدد فقط
    lectures = await db.lectures.find({
        "course_id": {"$in": course_ids},
        "date": date
    }).sort("start_time", 1).to_list(200)
    
    # جلب أسماء المعلمين دفعة واحدة
    teacher_ids = set()
    for c in courses:
        if c.get("teacher_id") and ObjectId.is_valid(str(c["teacher_id"])):
            teacher_ids.add(str(c["teacher_id"]))
    
    teacher_map = {}
    if teacher_ids:
        teachers = await db.teachers.find({"_id": {"$in": [ObjectId(tid) for tid in teacher_ids]}}).to_list(200)
        teacher_map = {str(t["_id"]): t.get("full_name", "") for t in teachers}
    
    result = []
    for lecture in lectures:
        course = course_map.get(lecture["course_id"], {})
        result.append({
            "id": str(lecture["_id"]),
            "course_id": lecture["course_id"],
            "course_name": course.get("name", ""),
            "course_code": course.get("code", ""),
            "section": course.get("section", ""),
            "date": lecture["date"],
            "day": lecture.get("day", ""),
            "start_time": lecture["start_time"],
            "end_time": lecture["end_time"],
            "room": lecture.get("room", ""),
            "status": lecture.get("status", LectureStatus.SCHEDULED),
            "teacher_name": teacher_map.get(str(course.get("teacher_id", "")), ""),
            "created_at": lecture.get("created_at", "")
        })
    
    return {
        "lectures": result,
        "date": date,
        "total": len(result),
    }


@api_router.get("/lectures/month/{year}/{month}")
async def get_month_lectures(
    year: int,
    month: int,
    fields: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """الحصول على محاضرات شهر معين للمعلم - يرجع التواريخ التي فيها محاضرات"""
    
    # جلب مقررات المعلم
    course_query = {"is_active": True}
    
    if current_user["role"] == UserRole.TEACHER:
        user = await db.users.find_one({"_id": ObjectId(current_user["id"])})
        if user and user.get("teacher_record_id"):
            course_query["teacher_id"] = user["teacher_record_id"]
        else:
            course_query["teacher_id"] = current_user["id"]
    elif current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "view_lectures") and not has_permission(current_user, "manage_lectures"):
        return {"dates": [], "lectures": []}
    
    courses = await db.courses.find(course_query).to_list(100)
    course_ids = [str(c["_id"]) for c in courses]
    course_map = {str(c["_id"]): c for c in courses}
    
    if not course_ids:
        return {"dates": [], "lectures": []}
    
    # حساب بداية ونهاية الشهر
    start_date = f"{year:04d}-{month:02d}-01"
    if month == 12:
        end_date = f"{year+1:04d}-01-01"
    else:
        end_date = f"{year:04d}-{month+1:02d}-01"
    
    # جلب محاضرات الشهر
    lectures = await db.lectures.find({
        "course_id": {"$in": course_ids},
        "date": {"$gte": start_date, "$lt": end_date}
    }).sort("date", 1).to_list(500)
    
    # تحديث تلقائي: المحاضرات المجدولة التي انتهى وقتها بدون تحضير → غائب
    now = get_yemen_time()
    for lecture in lectures:
        if lecture.get("status") == LectureStatus.SCHEDULED and not lecture.get("status_override"):
            try:
                lecture_end = datetime.strptime(f"{lecture['date']} {lecture['end_time']}", "%Y-%m-%d %H:%M")
                lecture_end = lecture_end.replace(tzinfo=YEMEN_TIMEZONE)
                lecture_start = datetime.strptime(f"{lecture['date']} {lecture['start_time']}", "%Y-%m-%d %H:%M")
                lecture_start = lecture_start.replace(tzinfo=YEMEN_TIMEZONE)
                duration = (lecture_end - lecture_start).total_seconds() / 60
                allowed_end = lecture_end + timedelta(minutes=duration)
                if now > allowed_end:
                    await db.lectures.update_one(
                        {"_id": lecture["_id"]},
                        {"$set": {"status": LectureStatus.ABSENT}}
                    )
                    lecture["status"] = LectureStatus.ABSENT
            except:
                pass

    # تجميع التواريخ والمحاضرات
    dates_with_lectures = {}
    lectures_list = []

    # 🆕 batch lookup لأسماء الأقسام والكليات
    _dept_ids = {c.get("department_id") for c in course_map.values() if c.get("department_id")}
    _dept_map: dict = {}
    _fac_ids: set = set()
    if _dept_ids:
        try:
            _oids = [ObjectId(x) for x in _dept_ids if x]
            async for d in db.departments.find({"_id": {"$in": _oids}}, {"name": 1, "faculty_id": 1}):
                _dept_map[str(d["_id"])] = {"name": (d.get("name") or "").strip(), "faculty_id": str(d.get("faculty_id") or "") or None}
                if d.get("faculty_id"): _fac_ids.add(str(d["faculty_id"]))
        except Exception:
            pass
    _fac_map: dict = {}
    if _fac_ids:
        try:
            _foids = [ObjectId(x) for x in _fac_ids if x]
            async for f in db.faculties.find({"_id": {"$in": _foids}}, {"name": 1}):
                _fac_map[str(f["_id"])] = (f.get("name") or "").strip()
        except Exception:
            pass

    for lecture in lectures:
        date = lecture["date"]
        course = course_map.get(lecture["course_id"], {})
        _cdept_id = str(course.get("department_id", "") or "")
        _cdept_info = _dept_map.get(_cdept_id, {})
        _cdept_name = _cdept_info.get("name") or None
        _cfac_id = _cdept_info.get("faculty_id")
        _cfac_name = _fac_map.get(_cfac_id, "") if _cfac_id else None

        if date not in dates_with_lectures:
            dates_with_lectures[date] = []

        lecture_info = {
            "id": str(lecture["_id"]),
            "course_id": lecture["course_id"],
            "course_name": course.get("name", ""),
            "course_code": course.get("code", ""),
            "section": course.get("section", ""),
            "department_id": _cdept_id or None,
            "department_name": _cdept_name,
            "faculty_id": _cfac_id,
            "faculty_name": _cfac_name,
            "date": date,
            "start_time": lecture["start_time"],
            "end_time": lecture["end_time"],
            "room": lecture.get("room", ""),
            "status": lecture.get("status", LectureStatus.SCHEDULED),
        }

        dates_with_lectures[date].append(lecture_info)
        lectures_list.append(lecture_info)
    
    return {
        "dates": list(dates_with_lectures.keys()),
        "lectures_by_date": dates_with_lectures,
        "lectures": apply_fields(lectures_list, parse_fields(fields)),
        "total_lectures": len(lectures_list)
    }

@api_router.get("/lectures/{course_id}")
async def get_course_lectures(
    course_id: str,
    status: Optional[str] = None,
    page: int = 1,
    per_page: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على محاضرات مقرر مع تقسيم صفحات"""
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    query = {"course_id": course_id}
    if status:
        query["status"] = status
    
    # عدد المحاضرات الإجمالي (مع الفلتر)
    total = await db.lectures.count_documents(query)
    
    # 🔧 إحصائيات حسب فصل المقرر نفسه (وليس الفصل النشط)
    # كي تكون متطابقة مع /courses/{id}/full-details و /lecture-stats
    base_query = {"course_id": course_id}
    course_sem_id = course.get("semester_id")
    target_sem = None
    if course_sem_id:
        try:
            _s = await db.semesters.find_one({"_id": ObjectId(course_sem_id)})
            if _s:
                target_sem = {
                    "id": str(_s["_id"]),
                    "name": _s.get("name", ""),
                    "start_date": _s.get("start_date"),
                    "end_date": _s.get("end_date"),
                }
        except Exception:
            pass
    if not target_sem:
        target_sem = await get_active_semester_with_dates(db)
    if target_sem:
        sem_id = target_sem["id"]
        sd = target_sem.get("start_date")
        ed = target_sem.get("end_date")
        date_range = {}
        if sd:
            date_range["$gte"] = sd
        if ed:
            date_range["$lte"] = ed
        or_clauses = [{"semester_id": sem_id}]
        if date_range:
            or_clauses.append({
                "$and": [
                    {"$or": [
                        {"semester_id": {"$exists": False}},
                        {"semester_id": None},
                        {"semester_id": ""},
                    ]},
                    {"date": date_range},
                ]
            })
        base_query["$or"] = or_clauses
        # كذلك نطبق نفس الفلتر على query (لقائمة المحاضرات المعروضة)
        query["$or"] = or_clauses

    stats = {
        "total": await db.lectures.count_documents(base_query),
        "scheduled": await db.lectures.count_documents({**base_query, "status": "scheduled"}),
        "completed": await db.lectures.count_documents({**base_query, "status": "completed"}),
        "cancelled": await db.lectures.count_documents({**base_query, "status": "cancelled"}),
        "absent": await db.lectures.count_documents({**base_query, "status": "absent"}),
    }
    if target_sem:
        stats["semester_id"] = target_sem["id"]
        stats["semester_name"] = target_sem.get("name", "")
        stats["semester_start_date"] = target_sem.get("start_date")
        stats["semester_end_date"] = target_sem.get("end_date")

    # نُحدّث total ليتوافق مع query (الذي قد يحوي فلتر حالة + فصل)
    total = await db.lectures.count_documents(query)
    
    # تحديث جماعي: المحاضرات المجدولة التي انتهى وقتها → غائب
    now = get_yemen_time()
    now_str = now.strftime("%Y-%m-%d %H:%M")
    today_str = now.strftime("%Y-%m-%d")
    
    # تحديث المحاضرات القديمة دفعة واحدة (التي تاريخها قبل اليوم)
    await db.lectures.update_many(
        {
            "course_id": course_id,
            "status": LectureStatus.SCHEDULED,
            "status_override": {"$ne": True},
            "date": {"$lt": today_str}
        },
        {"$set": {"status": LectureStatus.ABSENT}}
    )
    
    # تقسيم الصفحات
    skip = (page - 1) * per_page
    lectures = await db.lectures.find(query).sort("date", 1).skip(skip).limit(per_page).to_list(per_page)
    
    # تحديث محاضرات اليوم فقط (عدد قليل)
    for lecture in lectures:
        if lecture.get("status") == LectureStatus.SCHEDULED and not lecture.get("status_override") and lecture.get("date") == today_str:
            try:
                lecture_end = datetime.strptime(f"{lecture['date']} {lecture['end_time']}", "%Y-%m-%d %H:%M")
                lecture_end = lecture_end.replace(tzinfo=YEMEN_TIMEZONE)
                lecture_start = datetime.strptime(f"{lecture['date']} {lecture['start_time']}", "%Y-%m-%d %H:%M")
                lecture_start = lecture_start.replace(tzinfo=YEMEN_TIMEZONE)
                duration = (lecture_end - lecture_start).total_seconds() / 60
                allowed_end = lecture_end + timedelta(minutes=duration)
                if now > allowed_end:
                    await db.lectures.update_one(
                        {"_id": lecture["_id"]},
                        {"$set": {"status": LectureStatus.ABSENT}}
                    )
                    lecture["status"] = LectureStatus.ABSENT
            except:
                pass
    
    result = []
    for lecture in lectures:
        result.append({
            "id": str(lecture["_id"]),
            "course_id": lecture["course_id"],
            "date": lecture["date"],
            "start_time": lecture["start_time"],
            "end_time": lecture["end_time"],
            "room": lecture.get("room", ""),
            "status": lecture.get("status", LectureStatus.SCHEDULED),
            "notes": lecture.get("notes", ""),
            "created_at": lecture["created_at"],
            # ملاحظات الإلغاء وإعادة الجدولة
            "original_date": lecture.get("original_date"),
            "last_rescheduled_from": lecture.get("last_rescheduled_from"),
            "last_rescheduled_at": lecture.get("last_rescheduled_at"),
            "rescheduled_by_name": lecture.get("rescheduled_by_name"),
            "cancellation_reason": lecture.get("cancellation_reason"),
            "cancelled_at": lecture.get("cancelled_at"),
            "cancelled_by_name": lecture.get("cancelled_by_name"),
        })
    
    return {
        "lectures": result,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page,
        "stats": stats
    }

async def notify_lecture_created(course: dict, date: str, start_time: str, end_time: str):
    """إرسال تنبيه تلقائي عند إنشاء محاضرة جديدة - للمعلم وطلاب المقرر"""
    try:
        course_name = course.get("name", "")
        course_id = str(course["_id"])
        title = f"محاضرة جديدة - {course_name}"
        message = f"تم إنشاء محاضرة جديدة لمقرر {course_name} بتاريخ {date} من {start_time} إلى {end_time}. يرجى مراجعة التطبيق للاطلاع على التحديثات."
        
        target_user_ids = []
        
        # 1. تنبيه المعلم
        teacher_id = course.get("teacher_id")
        if teacher_id:
            teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
            if teacher and teacher.get("user_id"):
                target_user_ids.append(teacher["user_id"])
        
        # 2. تنبيه طلاب المقرر
        enrollments = await db.enrollments.find({"course_id": course_id}).to_list(5000)
        student_ids = [e["student_id"] for e in enrollments]
        if student_ids:
            students = await db.students.find(
                {"_id": {"$in": [ObjectId(sid) for sid in student_ids]}}
            ).to_list(5000)
            for s in students:
                if s.get("user_id"):
                    target_user_ids.append(s["user_id"])
        
        # حفظ الإشعارات داخل التطبيق
        if target_user_ids:
            in_app = [{
                "user_id": uid,
                "title": title,
                "message": message,
                "type": "reminder",
                "course_id": course_id,
                "course_name": course_name,
                "is_read": False,
                "created_at": get_yemen_time().isoformat(),
            } for uid in set(target_user_ids)]
            await db.notifications.insert_many(in_app)
        
        # إرسال push notifications
        from services.firebase_service import send_notification_to_many
        if target_user_ids:
            tokens_docs = await db.fcm_tokens.find(
                {"user_id": {"$in": list(set(target_user_ids))}}
            ).to_list(5000)
            tokens = [doc["token"] for doc in tokens_docs if doc.get("token")]
            if tokens:
                await send_notification_to_many(tokens, title, message)
        
        logging.info(f"تم إرسال تنبيهات محاضرة جديدة: {course_name} إلى {len(set(target_user_ids))} مستخدم")
    except Exception as e:
        logging.error(f"خطأ في إرسال تنبيه المحاضرة: {str(e)}")


async def check_teacher_lecture_conflict(course_id: str, date: str, start_time: str, end_time: str, exclude_lecture_id: str = None, allow_same_course: bool = False):
    """فحص تعارض محاضرات الأستاذ - هل لديه محاضرة أخرى في نفس الوقت؟
    allow_same_course: إذا True يرجع تحذير بدل خطأ للمقررات التي لها نفس الاسم الأساسي (شعب مختلفة)
    """
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        return None
    course_name = course.get("name", "")
    # استخراج الاسم الأساسي بدون حرف الشعبة
    import re
    base_name = re.sub(r'\s*\([أ-ي]\)\s*$', '', course_name).strip()

    # 🔒 (1) فحص تعارض المقرر نفسه — خطأ صارم دائماً (لا يمكن تجاوزه بـ force)
    # يعمل حتى لو لم يكن للمقرر أستاذ مرتبط
    same_course_query = {
        "course_id": course_id,
        "date": date,
        "status": {"$ne": LectureStatus.CANCELLED},
    }
    if exclude_lecture_id:
        same_course_query["_id"] = {"$ne": ObjectId(exclude_lecture_id)}
    for lec in await db.lectures.find(same_course_query).to_list(1000):
        ex_start = lec.get("start_time", "")
        ex_end = lec.get("end_time", "")
        if ex_start and ex_end and start_time and end_time:
            if start_time < ex_end and end_time > ex_start:
                return {
                    "type": "error",
                    "message": f"يوجد تعارض: المقرر \"{course_name}\" لديه محاضرة أخرى بنفس التوقيت يوم {date} من {ex_start} إلى {ex_end} — لا يمكن إنشاء محاضرتين متداخلتين لنفس المقرر"
                }

    if not course.get("teacher_id"):
        return None  # لا يوجد أستاذ مرتبط — اكتفينا بفحص المقرر نفسه

    teacher_id = course["teacher_id"]
    
    # جلب جميع مقررات هذا الأستاذ
    teacher_courses = await db.courses.find({"teacher_id": teacher_id}).to_list(1000)
    course_ids = [str(c["_id"]) for c in teacher_courses]
    
    # البحث عن محاضرات في نفس التاريخ (غير ملغاة)
    query = {
        "course_id": {"$in": course_ids},
        "date": date,
        "status": {"$ne": LectureStatus.CANCELLED},
    }
    if exclude_lecture_id:
        query["_id"] = {"$ne": ObjectId(exclude_lecture_id)}
    
    existing_lectures = await db.lectures.find(query).to_list(1000)
    
    for lec in existing_lectures:
        ex_start = lec.get("start_time", "")
        ex_end = lec.get("end_time", "")
        if ex_start and ex_end and start_time and end_time:
            # فحص التداخل الزمني
            if start_time < ex_end and end_time > ex_start:
                conflict_course = next((c for c in teacher_courses if str(c["_id"]) == lec["course_id"]), None)
                conflict_name = conflict_course.get("name", "") if conflict_course else ""
                conflict_base = re.sub(r'\s*\([أ-ي]\)\s*$', '', conflict_name).strip()
                
                # إذا كان التعارض من نفس المقرر (شعب مختلفة)
                if allow_same_course and base_name == conflict_base:
                    return {"type": "warning", "message": f"تنبيه: يوجد محاضرة لشعبة أخرى من نفس المقرر \"{conflict_name}\" في نفس الوقت. هل تريد المتابعة؟"}
                
                return {"type": "error", "message": f"يوجد تعارض: الأستاذ لديه محاضرة في مقرر \"{conflict_name}\" يوم {date} من {ex_start} إلى {ex_end}"}
    
    return None


async def check_room_lecture_conflict(course_id: str, room: str, date: str, start_time: str, end_time: str, exclude_lecture_id: str = None):
    """🏛️ فحص تعارض القاعة - هل القاعة محجوزة لمحاضرة أخرى في نفس الوقت؟
    - معلم مختلف أو مقرر مختلف → خطأ صارم (لا يمكن تجاوزه)
    - نفس المعلم وشعبة أخرى من نفس المقرر الأساسي → تحذير قابل للتجاوز (دمج شعبتين في قاعة واحدة)
    """
    room = (room or "").strip()
    if not room:
        return None
    import re
    my_course = await db.courses.find_one({"_id": ObjectId(course_id)}) if course_id else None
    my_teacher = my_course.get("teacher_id") if my_course else None
    my_base = re.sub(r'\s*\([أ-ي]\)\s*$', '', (my_course or {}).get("name", "")).strip()

    query = {
        "room": room,
        "date": date,
        "status": {"$ne": LectureStatus.CANCELLED},
    }
    if exclude_lecture_id:
        query["_id"] = {"$ne": ObjectId(exclude_lecture_id)}
    for lec in await db.lectures.find(query).to_list(1000):
        # نفس المقرر: يُغطى بفحص تعارض المقرر نفسه (رسالة أوضح هناك)
        if lec.get("course_id") == course_id:
            continue
        ex_start = lec.get("start_time", "")
        ex_end = lec.get("end_time", "")
        if not (ex_start and ex_end and start_time and end_time):
            continue
        if not (start_time < ex_end and end_time > ex_start):
            continue
        other = None
        if lec.get("course_id"):
            try:
                other = await db.courses.find_one({"_id": ObjectId(lec["course_id"])})
            except Exception:
                other = None
        oname = (other or {}).get("name", "مقرر آخر")
        obase = re.sub(r'\s*\([أ-ي]\)\s*$', '', oname).strip()
        msg = f"تعارض قاعة: القاعة \"{room}\" محجوزة يوم {date} من {ex_start} إلى {ex_end} لمقرر \"{oname}\""
        # دمج شعبتين لنفس المعلم ونفس المقرر الأساسي في نفس القاعة → تحذير قابل للتجاوز
        if my_teacher and other and other.get("teacher_id") == my_teacher and my_base and my_base == obase:
            return {"type": "warning", "message": msg + " — شعبة أخرى لنفس المعلم، يمكنك المتابعة لدمج الشعبتين"}
        return {"type": "error", "message": msg}
    return None



@api_router.post("/lectures")
async def create_lecture(
    data: LectureCreate,
    current_user: dict = Depends(get_current_user)
):
    """إنشاء محاضرة جديدة"""
    if current_user["role"] != UserRole.ADMIN and current_user["role"] != UserRole.TEACHER and not has_permission(current_user, "manage_lectures") and not has_permission(current_user, "add_lecture"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # التحقق من أن وقت النهاية بعد وقت البداية
    if data.start_time and data.end_time and data.end_time <= data.start_time:
        raise HTTPException(status_code=400, detail="وقت النهاية يجب أن يكون بعد وقت البداية")
    
    course = await db.courses.find_one({"_id": ObjectId(data.course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    # منع إنشاء محاضرة بتاريخ ماضي
    from datetime import datetime
    now = get_yemen_time()
    try:
        lecture_date = datetime.strptime(data.date, "%Y-%m-%d")
        today = now.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)
        if lecture_date < today:
            raise HTTPException(
                status_code=400, 
                detail="لا يمكن إنشاء محاضرة بتاريخ ماضي"
            )
    except ValueError:
        raise HTTPException(status_code=400, detail="صيغة التاريخ غير صحيحة")
    
    # فحص تعارض المحاضرات مع نفس الأستاذ
    conflict = await check_teacher_lecture_conflict(data.course_id, data.date, data.start_time, data.end_time, allow_same_course=True)
    if conflict:
        if conflict["type"] == "error":
            raise HTTPException(status_code=400, detail=conflict["message"])
        elif conflict["type"] == "warning" and not data.force:
            raise HTTPException(status_code=409, detail=conflict["message"])
    
    # 🏛️ فحص تعارض القاعة
    room_conflict = await check_room_lecture_conflict(data.course_id, data.room or "", data.date, data.start_time, data.end_time)
    if room_conflict:
        if room_conflict["type"] == "error":
            raise HTTPException(status_code=400, detail=room_conflict["message"])
        elif room_conflict["type"] == "warning" and not data.force:
            raise HTTPException(status_code=409, detail=room_conflict["message"])
    
    lecture = {
        "course_id": data.course_id,
        "date": data.date,
        "start_time": data.start_time,
        "end_time": data.end_time,
        "room": data.room or "",
        "status": LectureStatus.SCHEDULED,
        "notes": data.notes or "",
        "created_at": get_yemen_time(),
        "created_by": current_user["id"]
    }

    # ربط المحاضرة الجديدة بالفصل الدراسي المُفعَّل
    active_sem_for_lec = await get_active_semester_with_dates(db)
    if active_sem_for_lec:
        lecture["semester_id"] = active_sem_for_lec["id"]
        lecture["semester_name"] = active_sem_for_lec["name"]

    try:
        result = await db.lectures.insert_one(lecture)
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail="يوجد محاضرة مطابقة لنفس المقرر في نفس التاريخ ووقت البداية — تم رفض الإنشاء (حماية قاعدة البيانات)")
    
    # إرسال تنبيه تلقائي للمعلم وطلاب المقرر
    await notify_lecture_created(course, data.date, data.start_time, data.end_time)
    
    return {
        "id": str(result.inserted_id),
        "message": "تم إنشاء المحاضرة بنجاح"
    }

@api_router.put("/lectures/{lecture_id}/status")
async def update_lecture_status(lecture_id: str, request: Request, current_user: dict = Depends(get_current_user)):
    """تغيير حالة المحاضرة (يتطلب صلاحية override_lecture_status)"""
    if not has_permission(current_user, Permission.OVERRIDE_LECTURE_STATUS):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية تغيير حالة المحاضرة")
    
    data = await request.json()
    new_status = data.get("status")
    cancellation_reason = (data.get("cancellation_reason") or "").strip()
    valid = [LectureStatus.SCHEDULED, LectureStatus.COMPLETED, LectureStatus.CANCELLED, LectureStatus.ABSENT]
    if new_status not in valid:
        raise HTTPException(status_code=400, detail=f"حالة غير صالحة. الحالات المتاحة: {valid}")
    
    lecture = await db.lectures.find_one({"_id": ObjectId(lecture_id)})
    if not lecture:
        raise HTTPException(status_code=404, detail="المحاضرة غير موجودة")
    
    update_fields = {"status": new_status, "status_override": True}

    # عند الإلغاء أو غياب الأستاذ، احفظ سبب الإلغاء (إن وُجد)
    if new_status in (LectureStatus.CANCELLED, LectureStatus.ABSENT):
        if cancellation_reason:
            update_fields["cancellation_reason"] = cancellation_reason
        update_fields["cancelled_at"] = get_yemen_time()
        update_fields["cancelled_by"] = current_user["id"]
        update_fields["cancelled_by_name"] = current_user.get("full_name", "")
    elif new_status == LectureStatus.SCHEDULED:
        # عند إعادة المحاضرة لمجدولة، أزل بيانات الإلغاء
        update_fields["cancellation_reason"] = ""
        update_fields["cancelled_at"] = None
        update_fields["cancelled_by"] = None
        update_fields["cancelled_by_name"] = ""

    await db.lectures.update_one({"_id": ObjectId(lecture_id)}, {"$set": update_fields})
    
    status_names = {"scheduled": "مجدولة", "completed": "منعقدة", "cancelled": "ملغاة", "absent": "غائب"}
    await log_activity(
        current_user, "override_lecture_status", "lecture", lecture_id, None,
        {
            "old_status": lecture.get("status", ""),
            "new_status": new_status,
            "cancellation_reason": cancellation_reason or None,
        },
    )
    return {"message": f"تم تغيير حالة المحاضرة إلى: {status_names.get(new_status, new_status)}"}

@api_router.post("/lectures/generate")
async def generate_semester_lectures(
    data: GenerateLecturesRequest,
    current_user: dict = Depends(get_current_user)
):
    """توليد محاضرات الفصل الدراسي تلقائياً"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_lectures") and not has_permission(current_user, "manage_courses") and not has_permission(current_user, "generate_lectures"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # التحقق من أن وقت النهاية بعد وقت البداية
    if data.start_time and data.end_time and data.end_time <= data.start_time:
        raise HTTPException(status_code=400, detail="وقت النهاية يجب أن يكون بعد وقت البداية")
    
    course = await db.courses.find_one({"_id": ObjectId(data.course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    from datetime import timedelta
    
    start = datetime.strptime(data.start_date, "%Y-%m-%d")
    end = datetime.strptime(data.end_date, "%Y-%m-%d")
    
    # أيام الأسبوع: 0=السبت، 1=الأحد، 2=الاثنين، ...
    day_map = {0: 5, 1: 6, 2: 0, 3: 1, 4: 2, 5: 3, 6: 4}  # تحويل من نظامنا إلى Python weekday
    target_weekday = day_map.get(data.day_of_week, data.day_of_week)
    
    # البحث عن أول يوم مطابق
    current = start
    while current.weekday() != target_weekday:
        current += timedelta(days=1)
    
    lectures_created = 0
    conflicts_skipped = 0
    
    while current <= end:
        date_str = current.strftime("%Y-%m-%d")
        # فحص التعارض (نفس المقرر أو نفس الأستاذ أو القاعة)
        conflict = await check_teacher_lecture_conflict(data.course_id, date_str, data.start_time, data.end_time, allow_same_course=True)
        if conflict and conflict["type"] == "error":
            conflicts_skipped += 1
            current += timedelta(days=7)
            continue
        room_conflict = await check_room_lecture_conflict(data.course_id, data.room or "", date_str, data.start_time, data.end_time)
        if room_conflict and room_conflict["type"] == "error":
            conflicts_skipped += 1
            current += timedelta(days=7)
            continue
        lecture = {
            "course_id": data.course_id,
            "date": date_str,
            "start_time": data.start_time,
            "end_time": data.end_time,
            "room": data.room or "",
            "status": LectureStatus.SCHEDULED,
            "notes": "",
            "created_at": get_yemen_time(),
            "created_by": current_user["id"]
        }
        try:
            await db.lectures.insert_one(lecture)
            lectures_created += 1
        except DuplicateKeyError:
            conflicts_skipped += 1
        current += timedelta(days=7)
    
    return {
        "message": f"تم إنشاء {lectures_created} محاضرة للفصل الدراسي" + (f" (تم تخطي {conflicts_skipped} بسبب تعارض)" if conflicts_skipped > 0 else ""),
        "count": lectures_created,
        "conflicts_skipped": conflicts_skipped
    }

class DaySlot(BaseModel):
    start_time: str
    end_time: str

class DayScheduleConfig(BaseModel):
    day: str
    slots: List[DaySlot]

class GenerateSemesterRequest(BaseModel):
    course_id: str
    room: str
    schedule: List[DayScheduleConfig]
    start_date: str
    end_date: str

@api_router.post("/lectures/generate-semester")
async def generate_semester_lectures_advanced(
    data: GenerateSemesterRequest,
    current_user: dict = Depends(get_current_user)
):
    """توليد محاضرات الفصل الدراسي المتقدم - دعم أيام متعددة وأوقات متعددة"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_lectures") and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    course = await db.courses.find_one({"_id": ObjectId(data.course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    from datetime import timedelta
    
    try:
        start = datetime.strptime(data.start_date, "%Y-%m-%d")
        end = datetime.strptime(data.end_date, "%Y-%m-%d")
    except:
        raise HTTPException(status_code=400, detail="صيغة التاريخ غير صحيحة")
    
    # تحويل أسماء الأيام إلى أرقام Python weekday
    day_name_to_weekday = {
        'saturday': 5,
        'sunday': 6,
        'monday': 0,
        'tuesday': 1,
        'wednesday': 2,
        'thursday': 3,
        'friday': 4,
    }
    
    lectures_created = 0
    
    for day_config in data.schedule:
        target_weekday = day_name_to_weekday.get(day_config.day.lower())
        if target_weekday is None:
            continue
        
        # التحقق من أوقات الفترات
        for slot in day_config.slots:
            if slot.start_time and slot.end_time and slot.end_time <= slot.start_time:
                raise HTTPException(status_code=400, detail=f"وقت النهاية يجب أن يكون بعد وقت البداية ({slot.start_time} - {slot.end_time})")
        
        # البحث عن أول يوم مطابق
        current = start
        while current.weekday() != target_weekday:
            current += timedelta(days=1)
        
        # توليد المحاضرات لكل أسبوع
        conflicts_skipped = 0
        while current <= end:
            # إضافة محاضرة لكل فترة زمنية في هذا اليوم
            for slot in day_config.slots:
                # فحص تعارض المحاضرات مع نفس الأستاذ
                date_str = current.strftime("%Y-%m-%d")
                conflict = await check_teacher_lecture_conflict(data.course_id, date_str, slot.start_time, slot.end_time, allow_same_course=True)
                if conflict and conflict["type"] == "error":
                    conflicts_skipped += 1
                    continue
                room_conflict = await check_room_lecture_conflict(data.course_id, data.room or "", date_str, slot.start_time, slot.end_time)
                if room_conflict and room_conflict["type"] == "error":
                    conflicts_skipped += 1
                    continue
                
                lecture = {
                    "course_id": data.course_id,
                    "date": date_str,
                    "start_time": slot.start_time,
                    "end_time": slot.end_time,
                    "room": data.room,
                    "status": LectureStatus.SCHEDULED,
                    "notes": "",
                    "created_at": get_yemen_time(),
                    "created_by": current_user["id"]
                }
                try:
                    await db.lectures.insert_one(lecture)
                    lectures_created += 1
                except DuplicateKeyError:
                    conflicts_skipped += 1
            
            current += timedelta(days=7)
    
    # إرسال تنبيه عند إنشاء محاضرات الفصل
    if lectures_created > 0:
        await notify_lecture_created(course, data.start_date if hasattr(data, 'start_date') else "", "", "")

    return {
        "message": f"تم إنشاء {lectures_created} محاضرة للفصل الدراسي" + (f" (تم تخطي {conflicts_skipped} بسبب تعارض)" if conflicts_skipped > 0 else ""),
        "lectures_created": lectures_created,
        "conflicts_skipped": conflicts_skipped
    }

@api_router.put("/lectures/{lecture_id}")
async def update_lecture(
    lecture_id: str,
    data: LectureUpdate,
    current_user: dict = Depends(get_current_user)
):
    """تعديل محاضرة"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.TEACHER]:
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    lecture = await db.lectures.find_one({"_id": ObjectId(lecture_id)})
    if not lecture:
        raise HTTPException(status_code=404, detail="المحاضرة غير موجودة")
    
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    
    # التحقق من أن وقت النهاية بعد وقت البداية
    new_start = update_data.get("start_time", lecture.get("start_time", ""))
    new_end = update_data.get("end_time", lecture.get("end_time", ""))
    if new_start and new_end and new_end <= new_start:
        raise HTTPException(status_code=400, detail="وقت النهاية يجب أن يكون بعد وقت البداية")

    # 🔒 فحص التعارض عند تغيير التاريخ أو الوقت أو القاعة (نفس المقرر أو نفس الأستاذ أو القاعة)
    if any(k in update_data for k in ("date", "start_time", "end_time", "room")):
        eff_date = update_data.get("date") or lecture.get("date", "")
        conflict = await check_teacher_lecture_conflict(
            lecture.get("course_id", ""), eff_date, new_start, new_end,
            exclude_lecture_id=lecture_id, allow_same_course=True
        )
        if conflict and conflict["type"] == "error":
            raise HTTPException(status_code=400, detail=conflict["message"])
        eff_room = update_data.get("room") if "room" in update_data else lecture.get("room", "")
        room_conflict = await check_room_lecture_conflict(
            lecture.get("course_id", ""), eff_room or "", eff_date, new_start, new_end,
            exclude_lecture_id=lecture_id
        )
        if room_conflict and room_conflict["type"] == "error":
            raise HTTPException(status_code=400, detail=room_conflict["message"])

    # عند إعادة الجدولة (تغيير التاريخ): احفظ التاريخ الأصلي
    new_date = update_data.get("date")
    old_date = lecture.get("date")
    if new_date and old_date and new_date != old_date:
        # نحفظ أول تاريخ أصلي فقط (لا نُحدّث original_date في كل تعديل)
        if not lecture.get("original_date"):
            update_data["original_date"] = old_date
        update_data["last_rescheduled_at"] = get_yemen_time()
        update_data["last_rescheduled_from"] = old_date
        update_data["rescheduled_by_name"] = current_user.get("full_name", "")

    # عند تغيير الحالة لـ cancelled، احفظ السبب
    new_status = update_data.get("status")
    if new_status in (LectureStatus.CANCELLED, LectureStatus.ABSENT):
        if update_data.get("cancellation_reason"):
            update_data["cancelled_at"] = get_yemen_time()
            update_data["cancelled_by"] = current_user["id"]
            update_data["cancelled_by_name"] = current_user.get("full_name", "")

    # 🔧 إصلاح "ghost completion": عند مسح عنوان الدرس، نُلغي إنجاز المحاضرة
    # حتى لا يبقى محسوباً ضمن نسبة الإنجاز رغم عدم وجود درس فعلاً
    if "lesson_title" in update_data and not (update_data.get("lesson_title") or "").strip():
        old_topic_id = lecture.get("plan_topic_id")
        # إعادة الحالة إلى scheduled إن لم تتغيّر صراحةً لشيء آخر
        if "status" not in update_data and lecture.get("status") == LectureStatus.COMPLETED:
            update_data["status"] = LectureStatus.SCHEDULED
        update_data["lesson_title"] = ""
        update_data["plan_topic_id"] = None
        # إلغاء التأكيد اليدوي على الموضوع المرتبط (إن وُجد)
        if old_topic_id:
            await db.study_plans.update_one(
                {"course_id": lecture.get("course_id"), "topics.id": old_topic_id},
                {"$set": {
                    "topics.$.confirmed": False,
                    "topics.$.confirmed_date": None,
                    "topics.$.confirmed_by": None,
                }}
            )

    if update_data:
        try:
            await db.lectures.update_one(
                {"_id": ObjectId(lecture_id)},
                {"$set": update_data}
            )
        except DuplicateKeyError:
            raise HTTPException(status_code=409, detail="يوجد محاضرة مطابقة لنفس المقرر في نفس التاريخ ووقت البداية — تم رفض التعديل (حماية قاعدة البيانات)")
    
    return {"message": "تم تحديث المحاضرة بنجاح"}

@api_router.delete("/lectures/{lecture_id}")
async def delete_lecture(
    lecture_id: str,
    current_user: dict = Depends(get_current_user)
):
    """حذف محاضرة"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_lectures") and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # 🔧 جلب المحاضرة قبل حذفها لإلغاء التأكيد اليدوي على الموضوع المرتبط
    lecture = await db.lectures.find_one({"_id": ObjectId(lecture_id)})
    if not lecture:
        raise HTTPException(status_code=404, detail="المحاضرة غير موجودة")
    
    topic_id = lecture.get("plan_topic_id")
    course_id = lecture.get("course_id")

    result = await db.lectures.delete_one({"_id": ObjectId(lecture_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="المحاضرة غير موجودة")
    
    # حذف سجلات الحضور المرتبطة
    await db.attendance.delete_many({"lecture_id": lecture_id})

    # إلغاء التأكيد اليدوي على الموضوع المرتبط (لمنع ghost completion)
    if topic_id and course_id:
        # نتأكد ألا توجد محاضرة أخرى مكتملة تربط بنفس الموضوع
        other = await db.lectures.find_one({
            "course_id": course_id,
            "plan_topic_id": topic_id,
            "status": LectureStatus.COMPLETED,
        })
        if not other:
            await db.study_plans.update_one(
                {"course_id": course_id, "topics.id": topic_id},
                {"$set": {
                    "topics.$.confirmed": False,
                    "topics.$.confirmed_date": None,
                    "topics.$.confirmed_by": None,
                }}
            )

    return {"message": "تم حذف المحاضرة وسجلات حضورها بنجاح"}


@api_router.get("/admin/cleanup-orphan-attendance/preview")
async def preview_orphan_attendance(current_user: dict = Depends(get_current_user)):
    """معاينة عدد السجلات اليتيمة قبل الحذف (آمن - لا يُغيّر شيئاً)"""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="هذه العملية لمدير النظام فقط")

    # IDs المحاضرات النشطة (سيتم الإبقاء على سجلاتها)
    active_lecture_ids = set()
    # خريطة: (course_id, date_str) -> True إذا وجدت محاضرة نشطة
    active_course_date_map = set()
    async for lec in db.lectures.find(
        {"status": {"$in": ACTIVE_LECTURE_STATUSES}},
        {"_id": 1, "course_id": 1, "date": 1},
    ):
        active_lecture_ids.add(str(lec["_id"]))
        cid = str(lec.get("course_id", ""))
        d = lec.get("date", "")
        if cid and d:
            active_course_date_map.add((cid, str(d)))

    # IDs المحاضرات الموجودة كلياً (لتمييز ملغاة من محذوفة)
    all_lecture_ids = set()
    cancelled_lecture_ids = set()
    async for lec in db.lectures.find({}, {"_id": 1, "status": 1}):
        lid = str(lec["_id"])
        all_lecture_ids.add(lid)
        if lec.get("status") not in ACTIVE_LECTURE_STATUSES:
            cancelled_lecture_ids.add(lid)

    total_attendance = await db.attendance.count_documents({})
    deleted_lecture_orphans = 0
    cancelled_lecture_orphans = 0
    no_lecture_id_total = 0
    no_lecture_id_orphans = 0  # بلا lecture_id ولا توجد محاضرة نشطة في تاريخه
    no_lecture_id_valid = 0  # بلا lecture_id لكن توجد محاضرة نشطة في نفس التاريخ

    async for r in db.attendance.find({}, {"_id": 1, "lecture_id": 1, "course_id": 1, "date": 1}):
        lec_id = r.get("lecture_id")
        if not lec_id:
            no_lecture_id_total += 1
            cid = str(r.get("course_id", ""))
            d = r.get("date", "")
            # date في MongoDB قد يكون datetime - نحاول مطابقتها بسلسلة
            if hasattr(d, "strftime"):
                d_str = d.strftime("%Y-%m-%d")
            else:
                d_str = str(d)[:10] if d else ""
            if cid and d_str and (cid, d_str) in active_course_date_map:
                no_lecture_id_valid += 1
            else:
                no_lecture_id_orphans += 1
            continue
        sid = str(lec_id)
        if sid in cancelled_lecture_ids:
            cancelled_lecture_orphans += 1
        elif sid not in all_lecture_ids:
            deleted_lecture_orphans += 1

    total_orphans = deleted_lecture_orphans + cancelled_lecture_orphans + no_lecture_id_orphans

    return {
        "total_attendance": total_attendance,
        "orphans_count": total_orphans,
        "deleted_lecture_orphans": deleted_lecture_orphans,
        "cancelled_lecture_orphans": cancelled_lecture_orphans,
        "records_without_lecture_id": no_lecture_id_total,
        "no_lecture_id_orphans": no_lecture_id_orphans,
        "no_lecture_id_valid": no_lecture_id_valid,
        "active_lectures": len(active_lecture_ids),
        "needs_cleanup": total_orphans > 0,
    }


@api_router.post("/admin/cleanup-orphan-attendance")
async def cleanup_orphan_attendance(current_user: dict = Depends(get_current_user)):
    """
    تنظيف سجلات الحضور اليتيمة:
    1. سجلات لمحاضرات محذوفة كلياً
    2. سجلات لمحاضرات ملغاة (cancelled / absent)
    3. سجلات بدون lecture_id لا توجد محاضرة نشطة لها بنفس course_id + date
    صلاحية المدير فقط.
    """
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="هذه العملية لمدير النظام فقط")

    # IDs المحاضرات النشطة + خريطة (course_id, date)
    active_lecture_ids = set()
    active_course_date_map = set()
    async for lec in db.lectures.find(
        {"status": {"$in": ACTIVE_LECTURE_STATUSES}},
        {"_id": 1, "course_id": 1, "date": 1},
    ):
        active_lecture_ids.add(str(lec["_id"]))
        cid = str(lec.get("course_id", ""))
        d = lec.get("date", "")
        if cid and d:
            active_course_date_map.add((cid, str(d)))

    total_attendance = await db.attendance.count_documents({})

    orphan_ids = []
    async for r in db.attendance.find({}, {"_id": 1, "lecture_id": 1, "course_id": 1, "date": 1}):
        lec_id = r.get("lecture_id")
        if lec_id:
            # حذف إذا lecture_id لا يطابق محاضرة نشطة (محذوفة أو ملغاة)
            if str(lec_id) not in active_lecture_ids:
                orphan_ids.append(r["_id"])
            continue

        # سجل بدون lecture_id - تحقق من التطابق بـ course_id + date
        cid = str(r.get("course_id", ""))
        d = r.get("date", "")
        if hasattr(d, "strftime"):
            d_str = d.strftime("%Y-%m-%d")
        else:
            d_str = str(d)[:10] if d else ""

        # إذا لا توجد محاضرة نشطة في نفس التاريخ والمقرر → يتيم
        if not (cid and d_str and (cid, d_str) in active_course_date_map):
            orphan_ids.append(r["_id"])

    deleted = 0
    if orphan_ids:
        batch_size = 1000
        for i in range(0, len(orphan_ids), batch_size):
            batch = orphan_ids[i:i + batch_size]
            res = await db.attendance.delete_many({"_id": {"$in": batch}})
            deleted += res.deleted_count

    await log_activity(
        user=current_user,
        action="cleanup_orphan_attendance",
        entity_type="attendance",
        entity_id=None,
        entity_name="Orphan attendance cleanup",
        details={
            "total_before": total_attendance,
            "orphans_deleted": deleted,
        },
    )

    return {
        "message": f"تم حذف {deleted} سجل حضور يتيم",
        "total_attendance_before": total_attendance,
        "orphans_deleted": deleted,
        "active_lectures": len(active_lecture_ids),
    }


@api_router.get("/admin/cleanup-ghost-completions/preview")
async def preview_ghost_completions(current_user: dict = Depends(get_current_user)):
    """معاينة الإنجازات الوهمية في الخطط الدراسية قبل التنظيف.

    تشمل:
    1) محاضرات بحالة COMPLETED لكن عنوان الدرس فارغ (إنجاز كاذب).
    2) محاضرات بـ plan_topic_id لكن العنوان فارغ (ربط يتيم).
    3) مواضيع مؤكدة يدوياً (confirmed=true) لا توجد لها أي محاضرة مكتملة بعنوان فعلي
       تربطها بنفس course_id (تأكيد يتيم).
    """
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="هذه العملية لمدير النظام فقط")

    # 1) محاضرات COMPLETED بدون عنوان فعلي
    ghost_completed_lectures = await db.lectures.count_documents({
        "status": LectureStatus.COMPLETED,
        "$or": [
            {"lesson_title": {"$exists": False}},
            {"lesson_title": None},
            {"lesson_title": ""},
        ],
    })

    # 2) محاضرات لها plan_topic_id لكن لا عنوان درس
    orphan_topic_links = await db.lectures.count_documents({
        "plan_topic_id": {"$ne": None},
        "$or": [
            {"lesson_title": {"$exists": False}},
            {"lesson_title": None},
            {"lesson_title": ""},
        ],
    })

    # 3) تأكيدات يدوية يتيمة (لا توجد محاضرة مكتملة فعلية تربطها)
    orphan_confirmations = 0
    async for plan in db.study_plans.find({}, {"_id": 1, "course_id": 1, "topics": 1}):
        cid = plan.get("course_id")
        if not cid:
            continue
        for topic in plan.get("topics", []) or []:
            if not topic.get("confirmed"):
                continue
            tid = topic.get("id")
            if not tid:
                continue
            # هل توجد محاضرة مكتملة بعنوان فعلي مرتبطة بهذا الموضوع؟
            real_lec = await db.lectures.find_one({
                "course_id": cid,
                "plan_topic_id": tid,
                "status": LectureStatus.COMPLETED,
                "lesson_title": {"$nin": [None, ""]},
            })
            if not real_lec:
                orphan_confirmations += 1

    return {
        "ghost_completed_lectures": ghost_completed_lectures,
        "orphan_topic_links": orphan_topic_links,
        "orphan_manual_confirmations": orphan_confirmations,
        "total_to_fix": ghost_completed_lectures + orphan_topic_links + orphan_confirmations,
    }


@api_router.post("/admin/cleanup-ghost-completions")
async def cleanup_ghost_completions(current_user: dict = Depends(get_current_user)):
    """تنظيف الإنجازات الوهمية وإعادة المحاضرات والمواضيع المؤكدة إلى وضعها الصحيح."""
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="هذه العملية لمدير النظام فقط")

    empty_title_query = {
        "$or": [
            {"lesson_title": {"$exists": False}},
            {"lesson_title": None},
            {"lesson_title": ""},
        ],
    }

    # 1) إعادة حالة المحاضرات COMPLETED بدون عنوان إلى SCHEDULED
    res1 = await db.lectures.update_many(
        {"status": LectureStatus.COMPLETED, **empty_title_query},
        {"$set": {"status": LectureStatus.SCHEDULED, "plan_topic_id": None}},
    )

    # 2) إزالة plan_topic_id من أي محاضرة بدون عنوان فعلي
    res2 = await db.lectures.update_many(
        {"plan_topic_id": {"$ne": None}, **empty_title_query},
        {"$set": {"plan_topic_id": None}},
    )

    # ملاحظة: لا نلمس التأكيدات اليدوية تلقائياً لأنها قد تكون مشروعة
    # (تأكيد المعلم بشكل مستقل عن وجود محاضرة). تُعرض فقط للمعاينة كإشعار للمسؤول.
    unconfirmed_count = 0

    await log_activity(
        user=current_user,
        action="cleanup_ghost_completions",
        entity_type="study_plan",
        entity_id=None,
        entity_name="Ghost completions cleanup",
        details={
            "lectures_reverted_to_scheduled": res1.modified_count,
            "lecture_topic_links_cleared": res2.modified_count,
            "manual_confirmations_unconfirmed": unconfirmed_count,
        },
    )

    return {
        "message": "تم تنظيف الإنجازات الوهمية بنجاح",
        "lectures_reverted_to_scheduled": res1.modified_count,
        "lecture_topic_links_cleared": res2.modified_count,
        "manual_confirmations_unconfirmed": unconfirmed_count,
    }


@api_router.put("/lectures/{lecture_id}/reschedule")
async def reschedule_lecture(
    lecture_id: str,
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """إعادة جدولة محاضرة لم يحضرها المعلم"""
    # التحقق من الصلاحية
    is_admin = current_user["role"] == UserRole.ADMIN
    if not is_admin and not has_permission(current_user, "reschedule_lecture") and not has_permission(current_user, "manage_lectures"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بإعادة جدولة المحاضرات")
    
    lecture = await db.lectures.find_one({"_id": ObjectId(lecture_id)})
    if not lecture:
        raise HTTPException(status_code=404, detail="المحاضرة غير موجودة")
    
    # التحقق أن المحاضرة لم يحضرها المعلم (scheduled أو absent فقط)
    if lecture.get("status") == LectureStatus.COMPLETED:
        raise HTTPException(status_code=400, detail="لا يمكن إعادة جدولة محاضرة تم تسجيل حضورها")
    
    body = await request.json()
    new_date = body.get("date")
    new_start_time = body.get("start_time")
    new_end_time = body.get("end_time")
    
    if not new_date:
        raise HTTPException(status_code=400, detail="يرجى تحديد التاريخ الجديد")
    
    # التحقق من أن التاريخ الجديد مستقبلي
    from datetime import datetime
    now = get_yemen_time()
    try:
        new_lecture_date = datetime.strptime(new_date, "%Y-%m-%d")
        today = now.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)
        if new_lecture_date < today:
            raise HTTPException(status_code=400, detail="يجب اختيار تاريخ مستقبلي")
    except ValueError:
        raise HTTPException(status_code=400, detail="صيغة التاريخ غير صحيحة")
    
    # التحقق من الأوقات
    start_time = new_start_time or lecture.get("start_time", "")
    end_time = new_end_time or lecture.get("end_time", "")
    if start_time and end_time and end_time <= start_time:
        raise HTTPException(status_code=400, detail="وقت النهاية يجب أن يكون بعد وقت البداية")
    
    # فحص تعارض المحاضرات مع نفس الأستاذ
    conflict = await check_teacher_lecture_conflict(
        lecture.get("course_id", ""), new_date, start_time, end_time, exclude_lecture_id=lecture_id, allow_same_course=True
    )
    if conflict and conflict["type"] == "error":
        raise HTTPException(status_code=400, detail=conflict["message"])
    
    # 🏛️ فحص تعارض القاعة في الموعد الجديد
    room_conflict = await check_room_lecture_conflict(
        lecture.get("course_id", ""), lecture.get("room", "") or "", new_date, start_time, end_time,
        exclude_lecture_id=lecture_id
    )
    if room_conflict and room_conflict["type"] == "error":
        raise HTTPException(status_code=400, detail=room_conflict["message"])
    
    # حفظ التاريخ القديم للإشعار
    old_date = lecture.get("date", "")
    old_start = lecture.get("start_time", "")
    if hasattr(old_date, 'strftime'):
        old_date = old_date.strftime("%Y-%m-%d")
    
    # تحديث المحاضرة
    update_data = {
        "date": new_date,
        "status": LectureStatus.SCHEDULED,
        "rescheduled": True,
        "rescheduled_at": get_yemen_time().isoformat(),
        "rescheduled_by": current_user["id"],
        # حقول جديدة لعرض الملاحظة في بطاقة المحاضرة
        "last_rescheduled_from": old_date,
        "last_rescheduled_at": get_yemen_time(),
        "rescheduled_by_name": current_user.get("full_name", ""),
    }
    # احفظ التاريخ الأصلي مرة واحدة فقط (عند أول إعادة جدولة)
    if not lecture.get("original_date"):
        update_data["original_date"] = old_date
    if new_start_time:
        update_data["start_time"] = new_start_time
    if new_end_time:
        update_data["end_time"] = new_end_time
    
    try:
        await db.lectures.update_one({"_id": ObjectId(lecture_id)}, {"$set": update_data})
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail="يوجد محاضرة مطابقة لنفس المقرر في نفس التاريخ ووقت البداية — تم رفض إعادة الجدولة (حماية قاعدة البيانات)")
    
    # إرسال إشعار للمعلم والطلاب
    try:
        course = await db.courses.find_one({"_id": ObjectId(lecture.get("course_id", ""))})
        if course:
            course_name = course.get("name", "")
            title = f"إعادة جدولة محاضرة - {course_name}"
            message = f"تم إعادة جدولة محاضرة {course_name} من {old_date} ({old_start}) إلى {new_date} ({start_time})"
            
            target_user_ids = []
            
            # المعلم
            teacher_id = course.get("teacher_id")
            if teacher_id:
                teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
                if teacher and teacher.get("user_id"):
                    target_user_ids.append(teacher["user_id"])
            
            # الطلاب
            enrollments = await db.enrollments.find({"course_id": str(course["_id"])}).to_list(5000)
            student_ids = [e.get("student_id") for e in enrollments if e.get("student_id")]
            if student_ids:
                students = await db.students.find(
                    {"_id": {"$in": [ObjectId(sid) for sid in student_ids]}}
                ).to_list(5000)
                for s in students:
                    if s.get("user_id"):
                        target_user_ids.append(s["user_id"])
            
            # حفظ الإشعارات
            if target_user_ids:
                in_app = [{
                    "user_id": uid,
                    "title": title,
                    "message": message,
                    "type": "reschedule",
                    "course_id": str(course["_id"]),
                    "course_name": course_name,
                    "is_read": False,
                    "created_at": get_yemen_time().isoformat(),
                } for uid in set(target_user_ids)]
                await db.notifications.insert_many(in_app)
            
            # إرسال push
            from services.firebase_service import send_notification_to_many
            if target_user_ids:
                tokens_docs = await db.fcm_tokens.find(
                    {"user_id": {"$in": list(set(target_user_ids))}}
                ).to_list(5000)
                tokens = [doc["token"] for doc in tokens_docs if doc.get("token")]
                if tokens:
                    await send_notification_to_many(tokens, title, message)
            
            logger.info(f"تم إرسال إشعار إعادة جدولة: {course_name} إلى {len(set(target_user_ids))} مستخدم")
    except Exception as e:
        logger.error(f"خطأ في إرسال إشعار إعادة الجدولة: {e}")
    
    return {"message": f"تم إعادة جدولة المحاضرة إلى {new_date}"}



@api_router.get("/lectures/{lecture_id}/details")
async def get_lecture_details(
    lecture_id: str,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على تفاصيل محاضرة مع الطلاب المسجلين"""
    lecture = await db.lectures.find_one({"_id": ObjectId(lecture_id)})
    if not lecture:
        raise HTTPException(status_code=404, detail="المحاضرة غير موجودة")
    
    course = await db.courses.find_one({"_id": ObjectId(lecture["course_id"])})
    
    # الطلاب المسجلين في المقرر
    enrollments = await db.enrollments.find({"course_id": lecture["course_id"]}).to_list(10000)
    valid_student_ids = []
    for e in enrollments:
        try:
            valid_student_ids.append(ObjectId(e["student_id"]))
        except:
            pass
    students = await db.students.find({"_id": {"$in": valid_student_ids}}).to_list(10000)
    
    # سجلات الحضور لهذه المحاضرة
    attendance_records = await db.attendance.find({"lecture_id": lecture_id}).to_list(10000)
    attendance_map = {r["student_id"]: r["status"] for r in attendance_records}
    
    students_with_attendance = []
    for student in students:
        students_with_attendance.append({
            "id": str(student["_id"]),
            "student_id": student["student_id"],
            "full_name": student["full_name"],
            "attendance_status": attendance_map.get(str(student["_id"]), None)
        })
    
    return {
        "lecture": {
            "id": str(lecture["_id"]),
            "date": lecture["date"],
            "start_time": lecture["start_time"],
            "end_time": lecture["end_time"],
            "room": lecture.get("room", ""),
            "status": lecture.get("status", LectureStatus.SCHEDULED),
            "notes": lecture.get("notes", ""),
            "lesson_title": lecture.get("lesson_title", ""),
            "plan_topic_id": lecture.get("plan_topic_id", ""),
            "attendance_started_at": lecture.get("attendance_started_at", ""),
        },
        "course": {
            "id": str(course["_id"]),
            "name": course["name"],
            "code": course["code"]
        },
        "students": students_with_attendance,
        "attendance_recorded": len(attendance_records) > 0,
        "attendance_status": get_lecture_attendance_status(lecture, *await _get_faculty_attendance_settings(course))
    }

async def _get_faculty_attendance_settings(course: dict) -> tuple:
    """Helper to get attendance settings from faculty via course → department → faculty"""
    attendance_duration = 15
    max_delay = 30
    attendance_edit_minutes = 60
    try:
        # أولاً: جلب من إعدادات النظام
        settings = await db.settings.find_one({"_id": "system_settings"})
        if settings:
            attendance_edit_minutes = settings.get("attendance_edit_minutes", 60)
        # ثانياً: جلب من إعدادات الكلية (تتجاوز إعدادات النظام لـ duration و delay)
        if course.get("department_id"):
            dept = await db.departments.find_one({"_id": ObjectId(course["department_id"])})
            if dept and dept.get("faculty_id"):
                faculty = await db.faculties.find_one({"_id": ObjectId(dept["faculty_id"])})
                if faculty:
                    attendance_duration = faculty.get("attendance_duration_minutes", 15)
                    max_delay = faculty.get("max_attendance_delay_minutes", 30)
    except:
        pass
    return (attendance_duration, max_delay, attendance_edit_minutes)

async def _resolve_faculty_id(department_id: str) -> Optional[str]:
    """جلب faculty_id من القسم تلقائياً"""
    if not department_id:
        return None
    try:
        dept = await db.departments.find_one({"_id": ObjectId(department_id)})
        if dept:
            return dept.get("faculty_id")
    except:
        pass
    return None

@api_router.get("/lectures/{lecture_id}/pdf")
async def export_lecture_attendance_pdf(
    lecture_id: str,
    current_user: dict = Depends(get_current_user)
):
    """تصدير تقرير حضور المحاضرة كملف PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from bidi.algorithm import get_display
    import arabic_reshaper

    try:
        oid = ObjectId(lecture_id)
    except Exception:
        raise HTTPException(status_code=400, detail="معرف المحاضرة غير صالح")

    lecture = await db.lectures.find_one({"_id": oid})
    if not lecture:
        raise HTTPException(status_code=404, detail="المحاضرة غير موجودة")

    course = await db.courses.find_one({"_id": ObjectId(lecture["course_id"])})
    enrollments = await db.enrollments.find({"course_id": lecture["course_id"]}).to_list(10000)
    student_ids = [ObjectId(e["student_id"]) for e in enrollments]
    students = await db.students.find({"_id": {"$in": student_ids}}).to_list(10000)

    attendance_records = await db.attendance.find({"lecture_id": lecture_id}).to_list(10000)
    attendance_map = {r["student_id"]: r["status"] for r in attendance_records}
    
    # وقت بدء التحضير
    attendance_time = None
    if attendance_records:
        times = [r.get("recorded_at") or r.get("created_at") for r in attendance_records if r.get("recorded_at") or r.get("created_at")]
        if times:
            earliest = min(times)
            if hasattr(earliest, 'strftime'):
                # تحويل من UTC إلى توقيت اليمن
                if earliest.tzinfo is None:
                    earliest = earliest.replace(tzinfo=timezone.utc)
                yemen_time = earliest.astimezone(YEMEN_TIMEZONE)
                attendance_time = yemen_time.strftime('%H:%M')
    
    # المعلم المحضّر
    teacher_name = ""
    if course and course.get("teacher_id"):
        try:
            teacher = await db.teachers.find_one({"_id": ObjectId(course["teacher_id"])})
            if teacher:
                teacher_name = teacher.get("full_name", "")
        except:
            pass

    # جلب بيانات القسم والكلية
    department = None
    if course and course.get("department_id"):
        try:
            department = await db.departments.find_one({"_id": ObjectId(course["department_id"])})
        except:
            pass

    # Register Arabic font
    font_path = Path(__file__).parent / "fonts" / "Amiri-Regular.ttf"
    if font_path.exists():
        pdfmetrics.registerFont(TTFont('Amiri', str(font_path)))
        arabic_font = 'Amiri'
    else:
        arabic_font = 'Helvetica'

    def draw_arabic(c, text, x, y, font_name=arabic_font, font_size=12):
        reshaped = arabic_reshaper.reshape(text)
        bidi_text = get_display(reshaped)
        c.setFont(font_name, font_size)
        c.drawRightString(x, y, bidi_text)

    buffer = BytesIO()
    width, height = A4
    c = canvas.Canvas(buffer, pagesize=A4)

    # Header
    y = height - 40
    draw_arabic(c, "جامعة الأحقاف - تقرير الحضور", width - 30, y, font_size=18)
    y -= 8
    c.setStrokeColor(colors.HexColor("#1565c0"))
    c.setLineWidth(2)
    c.line(30, y, width - 30, y)

    # Course & Lecture info
    y -= 28
    course_name = course["name"] if course else "غير معروف"
    course_code = course.get("code", "") if course else ""
    draw_arabic(c, f"المقرر: {course_name} ({course_code})", width - 30, y, font_size=13)
    y -= 22
    
    # المستوى والشعبة
    level = course.get("level", "") if course else ""
    section = course.get("section", "") if course else ""
    level_section = ""
    if level:
        level_section += f"المستوى: {level}"
    if section:
        level_section += f"          الشعبة: {section}"
    if level_section:
        draw_arabic(c, level_section, width - 30, y, font_size=12)
        y -= 22
    
    # القسم والكلية
    dept_name = department.get("name", "") if department else ""
    college_name = department.get("college", "") if department else ""
    dept_info = ""
    if dept_name:
        dept_info += f"القسم: {dept_name}"
    if college_name:
        dept_info += f"          الكلية: {college_name}"
    if dept_info:
        draw_arabic(c, dept_info, width - 30, y, font_size=12)
        y -= 22
    
    draw_arabic(c, f"التاريخ: {lecture['date']}          الوقت: {lecture['start_time']} - {lecture['end_time']}", width - 30, y, font_size=12)
    y -= 20
    room = lecture.get("room", "")
    status_ar = {"scheduled": "مجدولة", "completed": "منعقدة", "cancelled": "ملغاة", "absent": "غائب"}.get(lecture.get("status", ""), lecture.get("status", ""))
    draw_arabic(c, f"القاعة: {room}          الحالة: {status_ar}", width - 30, y, font_size=12)

    # Stats
    present = sum(1 for s in students if attendance_map.get(str(s["_id"])) == "present")
    absent = sum(1 for s in students if attendance_map.get(str(s["_id"])) == "absent")
    excused = sum(1 for s in students if attendance_map.get(str(s["_id"])) == "excused")
    total = len(students)

    y -= 30
    c.setFillColor(colors.HexColor("#f5f5f5"))
    c.roundRect(30, y - 20, width - 60, 35, 5, fill=1, stroke=0)
    c.setFillColor(colors.black)
    draw_arabic(c, f"الإجمالي: {total}     حاضر: {present}     غائب: {absent}     معذور: {excused}", width - 40, y, font_size=11)

    # Table header
    y -= 45
    col_status_x = 80
    col_id_x = 200
    col_name_x = width - 40
    col_num_x = width - 20

    c.setFillColor(colors.HexColor("#1565c0"))
    c.roundRect(30, y - 8, width - 60, 25, 3, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont(arabic_font, 11)

    reshaped = arabic_reshaper.reshape("الحالة")
    c.drawString(col_status_x - 20, y, get_display(reshaped))
    reshaped = arabic_reshaper.reshape("الرقم الأكاديمي")
    c.drawString(col_id_x - 30, y, get_display(reshaped))
    reshaped = arabic_reshaper.reshape("اسم الطالب")
    c.drawRightString(col_name_x, y, get_display(reshaped))
    reshaped = arabic_reshaper.reshape("#")
    c.drawRightString(col_num_x, y, get_display(reshaped))

    # Table rows
    y -= 28
    status_labels = {"present": "حاضر", "absent": "غائب", "excused": "معذور"}
    status_colors = {"present": "#4caf50", "absent": "#f44336", "excused": "#ff9800"}

    for i, student in enumerate(students):
        if y < 50:
            c.showPage()
            y = height - 40
            draw_arabic(c, "جامعة الأحقاف - تقرير الحضور (تابع)", width - 30, y, font_size=14)
            y -= 30

        # Alternating row background
        if i % 2 == 0:
            c.setFillColor(colors.HexColor("#fafafa"))
            c.rect(30, y - 6, width - 60, 22, fill=1, stroke=0)

        c.setFillColor(colors.black)

        # Row number
        c.setFont(arabic_font, 11)
        reshaped = arabic_reshaper.reshape(str(i + 1))
        c.drawRightString(col_num_x, y, get_display(reshaped))

        # Student name
        draw_arabic(c, student["full_name"], col_name_x, y, font_size=11)

        # Student ID
        sid = student.get("student_id", "")
        c.setFont(arabic_font, 10)
        c.drawString(col_id_x - 30, y, str(sid))

        # Status with color
        att_status = attendance_map.get(str(student["_id"]), None)
        label = status_labels.get(att_status, "---")
        color = status_colors.get(att_status, "#999")
        c.setFillColor(colors.HexColor(color))
        reshaped = arabic_reshaper.reshape(label)
        c.drawString(col_status_x - 15, y, get_display(reshaped))
        c.setFillColor(colors.black)

        y -= 24

    # Footer - التوقيع
    y -= 20
    c.setStrokeColor(colors.HexColor("#e0e0e0"))
    c.setLineWidth(0.5)
    c.line(30, y, width - 30, y)
    y -= 18
    
    # اسم المحضّر ووقت التحضير
    if teacher_name:
        draw_arabic(c, f"المحضّر: {teacher_name}", width - 30, y, font_size=10)
        y -= 16
    if attendance_time:
        draw_arabic(c, f"وقت بدء التحضير: {attendance_time}", width - 30, y, font_size=10)
        y -= 16
    
    now = get_yemen_time()
    draw_arabic(c, f"تم إنشاء التقرير في: {now.strftime('%Y-%m-%d %H:%M')} (توقيت اليمن)", width - 30, y, font_size=9)
    
    # خط التوقيع
    y -= 30
    c.setStrokeColor(colors.HexColor("#333"))
    c.setLineWidth(0.8)
    c.line(width - 200, y, width - 30, y)
    y -= 14
    draw_arabic(c, "توقيع المحضّر", width - 30, y, font_size=9)

    c.save()
    buffer.seek(0)

    filename = f"{course_name}_{lecture['date']}.pdf"
    # URL encode for HTTP header (Arabic characters)
    from urllib.parse import quote
    encoded_filename = quote(filename)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


def get_lecture_attendance_status(lecture: dict, attendance_duration: int = 15, max_delay: int = 30, attendance_edit_minutes: int = 60) -> dict:
    """حساب حالة التحضير للمحاضرة"""
    now = get_yemen_time()
    
    try:
        lecture_start = datetime.strptime(f"{lecture['date']} {lecture['start_time']}", "%Y-%m-%d %H:%M")
        lecture_end = datetime.strptime(f"{lecture['date']} {lecture['end_time']}", "%Y-%m-%d %H:%M")
        lecture_start = lecture_start.replace(tzinfo=YEMEN_TIMEZONE)
        lecture_end = lecture_end.replace(tzinfo=YEMEN_TIMEZONE)
    except:
        return {
            "can_take_attendance": False,
            "reason": "خطأ في تنسيق التاريخ أو الوقت",
            "status": "error"
        }
    
    # لا يمكن التحضير قبل بداية المحاضرة
    if now < lecture_start:
        minutes_until = int((lecture_start - now).total_seconds() / 60)
        return {
            "can_take_attendance": False,
            "reason": f"لم يحن وقت المحاضرة بعد. تبدأ بعد {minutes_until} دقيقة",
            "status": "not_started",
            "starts_at": lecture_start.strftime("%H:%M")
        }
    
    # الحد الأقصى للتأخير: لا يمكن فتح التحضير بعد max_delay دقيقة من بداية المحاضرة
    max_open_time = lecture_start + timedelta(minutes=max_delay)
    
    # هل بدأ التحضير فعلاً؟
    attendance_started_at = lecture.get("attendance_started_at")
    
    if attendance_started_at:
        # التحضير بدأ - المؤقت يحسب من وقت بدء التحضير الفعلي
        if isinstance(attendance_started_at, str):
            started = datetime.fromisoformat(attendance_started_at)
            if started.tzinfo is None:
                started = started.replace(tzinfo=YEMEN_TIMEZONE)
        else:
            started = attendance_started_at
            if started.tzinfo is None:
                started = started.replace(tzinfo=YEMEN_TIMEZONE)
        
        attendance_deadline = started + timedelta(minutes=attendance_duration)
        is_update = lecture.get("status") == LectureStatus.COMPLETED
        teacher_delay = int((started - lecture_start).total_seconds() / 60)
        
        if now > attendance_deadline:
            # حساب مهلة التعديل من بدء التحضير
            edit_deadline = started + timedelta(minutes=attendance_edit_minutes)
            # إذا تم التحضير ولا تزال مهلة التعديل سارية، يُسمح بالتعديل
            if is_update and now <= edit_deadline:
                time_remaining = int((edit_deadline - now).total_seconds() / 60)
                return {
                    "can_take_attendance": True,
                    "reason": f"يمكن تعديل الحضور (متبقي {time_remaining} دقيقة)",
                    "status": "available",
                    "minutes_remaining": time_remaining,
                    "deadline": edit_deadline.strftime("%H:%M"),
                    "teacher_delay_minutes": teacher_delay,
                    "attendance_started_at": started.strftime("%H:%M"),
                }
            if is_update:
                return {
                    "can_take_attendance": False,
                    "reason": "تم التحضير وانتهت مدة التعديل",
                    "status": "completed"
                }
            return {
                "can_take_attendance": False,
                "reason": f"انتهت مدة التحضير ({attendance_duration} دقيقة من بدء التحضير)",
                "status": "expired"
            }
        
        time_remaining = int((attendance_deadline - now).total_seconds() / 60)
        return {
            "can_take_attendance": True,
            "reason": "يمكن تعديل الحضور" if is_update else "يمكن التحضير الآن",
            "status": "available",
            "minutes_remaining": time_remaining,
            "deadline": attendance_deadline.strftime("%H:%M"),
            "teacher_delay_minutes": teacher_delay,
            "attendance_started_at": started.strftime("%H:%M"),
        }
    else:
        # التحضير لم يبدأ بعد
        if now > max_open_time:
            return {
                "can_take_attendance": False,
                "reason": f"انتهى وقت فتح التحضير (الحد الأقصى للتأخير {max_delay} دقيقة)",
                "status": "expired"
            }
        
        time_remaining = int((max_open_time - now).total_seconds() / 60)
        return {
            "can_take_attendance": True,
            "reason": f"يمكن التحضير الآن (متبقي {time_remaining} دقيقة لفتح التحضير)",
            "status": "available",
            "minutes_remaining": time_remaining,
            "deadline": max_open_time.strftime("%H:%M"),
        }

@api_router.get("/lectures/{lecture_id}/attendance-status")
async def get_lecture_attendance_status_api(
    lecture_id: str,
    current_user: dict = Depends(get_current_user)
):
    """التحقق من حالة التحضير للمحاضرة"""
    lecture = await db.lectures.find_one({"_id": ObjectId(lecture_id)})
    if not lecture:
        raise HTTPException(status_code=404, detail="المحاضرة غير موجودة")
    
    # جلب إعدادات الكلية عبر المقرر → القسم → الكلية
    attendance_duration = 15
    max_delay = 30
    attendance_edit_minutes = 60
    try:
        course = await db.courses.find_one({"_id": ObjectId(lecture["course_id"])})
        if course:
            attendance_duration, max_delay, attendance_edit_minutes = await _get_faculty_attendance_settings(course)
    except:
        pass
    
    return get_lecture_attendance_status(lecture, attendance_duration, max_delay, attendance_edit_minutes)


@api_router.get("/attendance/lecture/{lecture_id}")
async def get_lecture_attendance(
    lecture_id: str,
    current_user: dict = Depends(get_current_user)
):
    """جلب سجلات الحضور لمحاضرة معينة"""
    lecture = await db.lectures.find_one({"_id": ObjectId(lecture_id)})
    if not lecture:
        raise HTTPException(status_code=404, detail="المحاضرة غير موجودة")
    
    records = await db.attendance.find({"lecture_id": lecture_id}).to_list(1000)
    
    result = []
    for r in records:
        student = await db.students.find_one({"_id": ObjectId(r["student_id"])})
        result.append({
            "id": str(r["_id"]),
            "lecture_id": r["lecture_id"],
            "student_id": r["student_id"],
            "student_name": student["full_name"] if student else "غير معروف",
            "student_number": student.get("student_id", "") if student else "",
            "status": r["status"],
            "date": r["date"].isoformat() if isinstance(r["date"], datetime) else r["date"],
            "method": r["method"],
            "notes": r.get("notes")
        })
    
    return result


# ==================== Attendance Routes ====================

@api_router.post("/attendance/session")
async def record_attendance_session(
    session: AttendanceSessionCreate,
    current_user: dict = Depends(get_current_user)
):
    """تسجيل حضور جماعي لمحاضرة"""
    can_record = current_user["role"] in [UserRole.ADMIN, UserRole.TEACHER]
    has_edit_perm = has_permission(current_user, "edit_attendance")
    has_manage_perm = has_permission(current_user, "manage_attendance")
    has_record_perm = has_permission(current_user, "record_attendance")
    
    if not (can_record or has_edit_perm or has_manage_perm or has_record_perm):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # Verify lecture exists
    lecture = await db.lectures.find_one({"_id": ObjectId(session.lecture_id)})
    if not lecture:
        raise HTTPException(status_code=404, detail="المحاضرة غير موجودة")
    
    course = await db.courses.find_one({"_id": ObjectId(lecture["course_id"])})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    # Check if teacher owns this course
    if current_user["role"] == UserRole.TEACHER:
        # البحث عن teacher_record_id من حساب المستخدم
        user_doc = await db.users.find_one({"_id": ObjectId(current_user["id"])})
        teacher_record_id = user_doc.get("teacher_record_id") if user_doc else None
        # المقارنة بكلا المعرفين: user_id و teacher_record_id
        if course["teacher_id"] != current_user["id"] and course["teacher_id"] != teacher_record_id:
            raise HTTPException(status_code=403, detail="غير مصرح لك بتسجيل حضور هذا المقرر")
    
    # === التحقق من قواعد التحضير ===
    now = get_yemen_time()
    
    # جلب إعدادات الكلية
    attendance_duration, max_delay, attendance_edit_minutes = await _get_faculty_attendance_settings(course)
    
    # إذا كان هذا تسجيل أوفلاين، استخدم وقت التسجيل الأصلي للتحقق
    check_time = now
    is_offline_sync = False
    if session.offline_recorded_at:
        try:
            offline_time = datetime.fromisoformat(session.offline_recorded_at.replace('Z', '+00:00'))
            if offline_time.tzinfo is None:
                offline_time = offline_time.replace(tzinfo=YEMEN_TIMEZONE)
            else:
                offline_time = offline_time.astimezone(YEMEN_TIMEZONE)
            # التحقق أن الوقت معقول (خلال آخر 48 ساعة)
            if (now - offline_time).total_seconds() < 48 * 3600:
                check_time = offline_time
                is_offline_sync = True
        except:
            pass
    
    lecture_date = datetime.strptime(lecture["date"], "%Y-%m-%d")
    lecture_start = datetime.strptime(f"{lecture['date']} {lecture['start_time']}", "%Y-%m-%d %H:%M")
    lecture_end = datetime.strptime(f"{lecture['date']} {lecture['end_time']}", "%Y-%m-%d %H:%M")
    # توحيد التوقيت
    lecture_start = lecture_start.replace(tzinfo=YEMEN_TIMEZONE)
    lecture_end = lecture_end.replace(tzinfo=YEMEN_TIMEZONE)
    
    # التحقق: لا يُسمح بالتحضير قبل وقت بداية المحاضرة
    if check_time < lecture_start and not is_offline_sync and not has_edit_perm:
        raise HTTPException(
            status_code=400,
            detail=f"لم يحن وقت المحاضرة بعد. تبدأ الساعة {lecture_start.strftime('%H:%M')}"
        )
    
    # هل بدأ التحضير مسبقاً؟
    attendance_started_at = lecture.get("attendance_started_at")
    
    if attendance_started_at:
        # التحضير بدأ - المؤقت من وقت البدء الفعلي
        if isinstance(attendance_started_at, str):
            started = datetime.fromisoformat(attendance_started_at)
            if started.tzinfo is None:
                started = started.replace(tzinfo=YEMEN_TIMEZONE)
        else:
            started = attendance_started_at
            if started.tzinfo is None:
                started = started.replace(tzinfo=YEMEN_TIMEZONE)
        
        attendance_deadline = started + timedelta(minutes=attendance_duration)
        # إذا تم التحضير، استخدام مهلة التعديل (attendance_edit_minutes) من الإعدادات
        is_completed = lecture.get("status") == LectureStatus.COMPLETED
        edit_deadline = started + timedelta(minutes=attendance_edit_minutes)
        effective_deadline = edit_deadline if is_completed else attendance_deadline
        is_bypass_edit = (
            check_time > effective_deadline
            and not is_offline_sync
            and has_edit_perm
            and is_completed
        )
        if check_time > effective_deadline and not is_offline_sync and not has_edit_perm:
            raise HTTPException(
                status_code=400,
                detail=f"انتهت مدة التحضير ({attendance_duration} دقيقة)" if not is_completed else f"تم التحضير وانتهت مدة التعديل ({attendance_edit_minutes} دقيقة)"
            )
        # 🆕 مسار اعتماد العميد: تعديل خارج المهلة من مستخدم غير ADMIN/DEAN
        if is_bypass_edit and current_user["role"] not in [UserRole.ADMIN, UserRole.DEAN]:
            diff_result = await create_change_requests_for_diff(
                lecture, course, session.records, current_user, reason=session.notes
            )
            return {
                "status": "pending_approval",
                "message": f"تم إرسال {diff_result['created']} تعديل لاعتماد العميد",
                "created": diff_result["created"],
                "skipped_no_change": diff_result["skipped"],
                "request_ids": diff_result["request_ids"],
            }
    else:
        # التحضير لم يبدأ بعد - التحقق من الحد الأقصى للتأخير
        max_open_time = lecture_start + timedelta(minutes=max_delay)
        if check_time > max_open_time and not is_offline_sync and not has_edit_perm:
            raise HTTPException(
                status_code=400,
                detail=f"انتهى وقت فتح التحضير (الحد الأقصى للتأخير {max_delay} دقيقة)"
            )
        # تسجيل وقت بدء التحضير الفعلي
        await db.lectures.update_one(
            {"_id": ObjectId(session.lecture_id)},
            {"$set": {"attendance_started_at": check_time.isoformat()}}
        )
    
    # === نهاية التحقق من القواعد ===
    
    # Delete existing attendance for this lecture (فقط إذا لم تكن مكتملة)
    await db.attendance.delete_many({"lecture_id": session.lecture_id})
    
    attendance_records = []
    for record in session.records:
        att_record = {
            "lecture_id": session.lecture_id,
            "course_id": lecture["course_id"],
            "student_id": record.student_id,
            "status": record.status,
            "date": get_yemen_time(),
            "recorded_by": current_user["id"],
            "method": "manual",
            "notes": session.notes,
            "created_at": get_yemen_time()
        }
        attendance_records.append(att_record)
    
    if attendance_records:
        await db.attendance.insert_many(attendance_records)
        
        # التحقق من نسب الغياب وإنشاء إشعارات للطلاب الغائبين
        for record in session.records:
            if record.status == AttendanceStatus.ABSENT:
                await check_and_create_absence_notifications(record.student_id, lecture["course_id"])
    
    # Update lecture status to completed
    lecture_update = {"status": LectureStatus.COMPLETED}
    if session.lesson_title:
        lecture_update["lesson_title"] = session.lesson_title
    if session.plan_topic_id:
        lecture_update["plan_topic_id"] = session.plan_topic_id
    
    await db.lectures.update_one(
        {"_id": ObjectId(session.lecture_id)},
        {"$set": lecture_update}
    )
    
    return {"message": f"تم تسجيل حضور {len(attendance_records)} طالب بنجاح"}

    # TODO: Send push notifications after attendance (background task)
    # This will be triggered when we have enough FCM tokens registered

@api_router.post("/attendance/single")
async def record_single_attendance(
    data: SingleAttendanceCreate,
    current_user: dict = Depends(get_current_user)
):
    """تسجيل حضور طالب واحد (QR)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.TEACHER]:
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # Verify lecture exists
    lecture = await db.lectures.find_one({"_id": ObjectId(data.lecture_id)})
    if not lecture:
        raise HTTPException(status_code=404, detail="المحاضرة غير موجودة")
    
    # Verify student exists
    student = await db.students.find_one({"_id": ObjectId(data.student_id)})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    # Check if already recorded for this lecture
    existing = await db.attendance.find_one({
        "lecture_id": data.lecture_id,
        "student_id": data.student_id
    })
    
    if existing:
        raise HTTPException(status_code=400, detail="تم تسجيل حضور هذا الطالب مسبقاً")
    
    att_record = {
        "lecture_id": data.lecture_id,
        "course_id": lecture["course_id"],
        "student_id": data.student_id,
        "status": data.status,
        "date": get_yemen_time(),
        "recorded_by": current_user["id"],
        "method": data.method,
        "notes": data.notes,
        "created_at": get_yemen_time()
    }
    
    await db.attendance.insert_one(att_record)
    
    return {"message": "تم تسجيل الحضور بنجاح", "student_name": student["full_name"]}

@api_router.get("/attendance/course/{course_id}")
async def get_course_attendance(
    course_id: str,
    date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    # جلب المحاضرات النشطة فقط
    active_lecture_ids = await get_active_lecture_ids(course_id)
    
    query = {"course_id": course_id}
    
    # فلتر: فقط سجلات الحضور المرتبطة بمحاضرات نشطة أو بدون lecture_id
    if active_lecture_ids:
        active_ids_str = [str(lid) for lid in active_lecture_ids]
        query["$or"] = [
            {"lecture_id": {"$in": active_ids_str}},
            {"lecture_id": None},
            {"lecture_id": {"$exists": False}},
        ]
    
    if date:
        date_obj = datetime.fromisoformat(date)
        query["date"] = {
            "$gte": date_obj.replace(hour=0, minute=0, second=0),
            "$lt": date_obj.replace(hour=23, minute=59, second=59)
        }
    
    records = await db.attendance.find(query).sort("date", -1).to_list(1000)
    
    # Get student details
    result = []
    for r in records:
        student = await db.students.find_one({"_id": ObjectId(r["student_id"])})
        result.append({
            "id": str(r["_id"]),
            "course_id": r["course_id"],
            "student_id": r["student_id"],
            "student_name": student["full_name"] if student else "غير معروف",
            "student_number": student["student_id"] if student else "",
            "status": r["status"],
            "date": r["date"].isoformat(),
            "method": r["method"],
            "notes": r.get("notes")
        })
    
    return result


@api_router.put("/attendance/{record_id}/status")
async def update_attendance_status(
    record_id: str,
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """تعديل حالة حضور طالب - يتطلب صلاحية edit_attendance"""
    user_permissions = current_user.get("permissions", [])
    # 🔧 admin يُسمح له مباشرة (feature spec: ADMIN bypass كامل)
    # ملاحظة: edit_attendance مصنّف TEACHER_ONLY_PERMISSIONS فلا يمنحه has_permission لـ admin تلقائياً
    if current_user["role"] == UserRole.ADMIN:
        has_edit_perm = True
    else:
        has_edit_perm = has_permission(current_user, "edit_attendance")
    
    if not has_edit_perm:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية تعديل الحضور")
    
    body = await request.json()
    new_status = body.get("status")
    reason = body.get("reason", "")
    
    if new_status not in ["present", "absent", "late", "excused"]:
        raise HTTPException(status_code=400, detail="حالة غير صالحة")
    
    record = await db.attendance.find_one({"_id": ObjectId(record_id)})
    if not record:
        raise HTTPException(status_code=404, detail="سجل الحضور غير موجود")
    
    old_status = record.get("status")
    
    # 🆕 مسار اعتماد العميد: أي مستخدم يعدّل عبر هذا الـ endpoint (bypass) وليس ADMIN/DEAN → طلب اعتماد
    if current_user["role"] not in [UserRole.ADMIN, UserRole.DEAN] and old_status != new_status:
        lecture = await db.lectures.find_one({"_id": ObjectId(record["lecture_id"])})
        course = await db.courses.find_one({"_id": ObjectId(record["course_id"])}) if record.get("course_id") else None
        if lecture and course:
            # صياغة تنسيق يتوافق مع helper (يستخدم .student_id و .status)
            class _R:
                def __init__(self, sid, st):
                    self.student_id = sid
                    self.status = st
            diff_result = await create_change_requests_for_diff(
                lecture, course, [_R(record["student_id"], new_status)], current_user, reason=reason or None
            )
            return {
                "status": "pending_approval",
                "message": "تم إرسال طلب التعديل لاعتماد العميد",
                "request_ids": diff_result["request_ids"],
            }
    
    await db.attendance.update_one(
        {"_id": ObjectId(record_id)},
        {"$set": {
            "status": new_status,
            "edited_by": current_user.get("sub", current_user.get("user_id")),
            "edited_at": get_yemen_time(),
            "edit_reason": reason,
            "original_status": old_status if not record.get("original_status") else record.get("original_status"),
        }}
    )
    
    # تسجيل النشاط
    student = None
    try:
        student = await db.students.find_one({"_id": ObjectId(record["student_id"])})
    except Exception:
        student = await db.students.find_one({"student_id": record["student_id"]})
    student_name = student["full_name"] if student else record.get("student_id", "غير معروف")
    await log_activity(
        current_user,
        "update_attendance",
        entity_type="attendance",
        entity_id=record_id,
        entity_name=student_name,
        details={"old_status": old_status, "new_status": new_status, "reason": reason}
    )
    
    return {"message": f"تم تعديل الحالة إلى {new_status}", "old_status": old_status, "new_status": new_status}


@api_router.get("/attendance/student/{student_id}")
async def get_student_attendance(
    student_id: str,
    course_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"student_id": student_id}
    if course_id:
        query["course_id"] = course_id

    records = await db.attendance.find(query).sort("date", -1).to_list(1000)

    # فلترة سجلات الحضور لاستبعاد المحاضرات المحذوفة/الملغاة
    active_lecture_ids = await get_active_lecture_ids(course_id)
    active_ids_str = {str(lid) for lid in active_lecture_ids}

    # خريطة (course_id, date) -> active، لمعالجة السجلات بدون lecture_id
    active_course_date = set()
    lec_query = {"status": {"$in": ACTIVE_LECTURE_STATUSES}}
    if course_id:
        lec_query["course_id"] = course_id
    async for lec in db.lectures.find(lec_query, {"course_id": 1, "date": 1}):
        cid = str(lec.get("course_id", ""))
        d = lec.get("date", "")
        if hasattr(d, "strftime"):
            d_str = d.strftime("%Y-%m-%d")
        else:
            d_str = str(d)[:10] if d else ""
        if cid and d_str:
            active_course_date.add((cid, d_str))

    def _is_valid(r):
        lid = r.get("lecture_id")
        if lid:
            return str(lid) in active_ids_str
        # سجل بدون lecture_id - تحقق من تطابق course_id + date مع محاضرة نشطة
        cid = str(r.get("course_id", ""))
        d = r.get("date", "")
        if hasattr(d, "strftime"):
            d_str = d.strftime("%Y-%m-%d")
        else:
            d_str = str(d)[:10] if d else ""
        return bool(cid and d_str and (cid, d_str) in active_course_date)

    records = [r for r in records if _is_valid(r)]

    result = []
    for r in records:
        course = await db.courses.find_one({"_id": ObjectId(r["course_id"])})
        
        # جلب start_time و end_time من جدول المحاضرات
        start_time = None
        end_time = None
        lecture_query = {"course_id": r["course_id"]}
        if isinstance(r["date"], str):
            lecture_query["date"] = r["date"][:10]
        else:
            lecture_query["date"] = r["date"].strftime("%Y-%m-%d")
        lecture = await db.lectures.find_one(lecture_query)
        if lecture:
            start_time = lecture.get("start_time")
            end_time = lecture.get("end_time")
        
        result.append({
            "id": str(r["_id"]),
            "course_id": r["course_id"],
            "course_name": course["name"] if course else "غير معروف",
            "status": r["status"],
            "date": r["date"].isoformat() if hasattr(r["date"], 'isoformat') else r["date"],
            "start_time": start_time,
            "end_time": end_time,
            "method": r["method"]
        })
    
    return result

@api_router.get("/attendance/stats/student/{student_id}", response_model=AttendanceStats)
async def get_student_stats(
    student_id: str,
    course_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"student_id": student_id}
    if course_id:
        query["course_id"] = course_id
    
    records = await db.attendance.find(query).to_list(1000)
    
    # فلترة السجلات لاستبعاد المحاضرات الملغاة
    records = await filter_attendance_by_active_lectures(records, course_id)
    
    total = len(records)
    present = sum(1 for r in records if r["status"] == AttendanceStatus.PRESENT)
    absent = sum(1 for r in records if r["status"] == AttendanceStatus.ABSENT)
    late = sum(1 for r in records if r["status"] == AttendanceStatus.LATE)
    excused = sum(1 for r in records if r["status"] == AttendanceStatus.EXCUSED)
    
    rate = (present + late * 0.5) / total * 100 if total > 0 else 0
    
    return {
        "total_sessions": total,
        "present_count": present,
        "absent_count": absent,
        "late_count": late,
        "excused_count": excused,
        "attendance_rate": round(rate, 2)
    }

@api_router.get("/attendance/stats/course/{course_id}")
async def get_course_stats(course_id: str, current_user: dict = Depends(get_current_user)):
    # Get all attendance records for this course
    records = await db.attendance.find({"course_id": course_id}).to_list(10000)
    
    # فلترة السجلات لاستبعاد المحاضرات الملغاة
    records = await filter_attendance_by_active_lectures(records, course_id)
    
    # Group by date to get unique sessions (only from active lectures)
    sessions = set()
    for r in records:
        sessions.add(r["date"].strftime("%Y-%m-%d"))
    
    # Get course details
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    # Get students in this course's level and section
    students = await db.students.find({
        "department_id": course["department_id"],
        "level": course["level"],
        "section": course["section"]
    }).to_list(500)
    
    # Calculate stats per student (only from active lectures)
    student_stats = []
    for student in students:
        student_records = [r for r in records if r["student_id"] == str(student["_id"])]
        total = len(student_records)
        present = sum(1 for r in student_records if r["status"] == AttendanceStatus.PRESENT)
        absent = sum(1 for r in student_records if r["status"] == AttendanceStatus.ABSENT)
        late = sum(1 for r in student_records if r["status"] == AttendanceStatus.LATE)
        
        rate = (present + late * 0.5) / total * 100 if total > 0 else 0
        
        student_stats.append({
            "student_id": str(student["_id"]),
            "student_number": student["student_id"],
            "student_name": student["full_name"],
            "total_sessions": total,
            "present_count": present,
            "absent_count": absent,
            "late_count": late,
            "attendance_rate": round(rate, 2)
        })
    
    return {
        "course_id": course_id,
        "course_name": course["name"],
        "total_sessions": len(sessions),
        "total_students": len(students),
        "student_stats": student_stats
    }

# ==================== Offline Sync Routes ====================

@api_router.post("/sync/attendance")
async def sync_offline_attendance(
    data: OfflineSyncData,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.TEACHER]:
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    synced = 0
    errors = []
    
    for record in data.attendance_records:
        try:
            # Check if already synced (using local_id if provided)
            if record.get("local_id"):
                existing = await db.attendance.find_one({"local_id": record["local_id"]})
                if existing:
                    continue
            
            att_record = {
                "course_id": record["course_id"],
                "student_id": record["student_id"],
                "status": record.get("status", AttendanceStatus.PRESENT),
                "date": datetime.fromisoformat(record["date"]) if record.get("date") else get_yemen_time(),
                "recorded_by": current_user["id"],
                "method": record.get("method", "manual"),
                "notes": record.get("notes"),
                "local_id": record.get("local_id"),
                "created_at": get_yemen_time(),
                "synced_at": get_yemen_time()
            }
            
            await db.attendance.insert_one(att_record)
            synced += 1
        except Exception as e:
            errors.append(str(e))
    
    return {
        "synced": synced,
        "errors": errors,
        "message": f"تمت مزامنة {synced} سجل بنجاح"
    }

# ==================== Reports Routes ====================

@api_router.get("/reports/department/{dept_id}")
async def get_department_report(
    dept_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "view_reports"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # Get department
    dept = await db.departments.find_one({"_id": ObjectId(dept_id)})
    if not dept:
        raise HTTPException(status_code=404, detail="القسم غير موجود")
    
    # Get all courses in department
    courses = await db.courses.find({"department_id": dept_id}).to_list(100)
    course_ids = [str(c["_id"]) for c in courses]
    
    # جلب IDs المحاضرات الفعّالة فقط لجميع المقررات
    all_active_lecture_ids = set()
    for cid in course_ids:
        active_ids = await get_active_lecture_ids(cid)
        all_active_lecture_ids.update(active_ids)
    
    # Build query
    query = {"course_id": {"$in": course_ids}}
    if start_date:
        query["date"] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = datetime.fromisoformat(end_date)
        else:
            query["date"] = {"$lte": datetime.fromisoformat(end_date)}
    
    records = await db.attendance.find(query).to_list(10000)
    
    # فلترة السجلات لاستبعاد المحاضرات الملغاة
    records = [r for r in records if r.get("lecture_id") in all_active_lecture_ids]
    
    total = len(records)
    present = sum(1 for r in records if r["status"] == AttendanceStatus.PRESENT)
    absent = sum(1 for r in records if r["status"] == AttendanceStatus.ABSENT)
    late = sum(1 for r in records if r["status"] == AttendanceStatus.LATE)
    
    return {
        "department": dept["name"],
        "total_courses": len(courses),
        "total_records": total,
        "present_count": present,
        "absent_count": absent,
        "late_count": late,
        "attendance_rate": round((present + late * 0.5) / total * 100, 2) if total > 0 else 0
    }

@api_router.get("/reports/summary")
async def get_summary_report(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "view_reports"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # Get counts
    total_students = await db.students.count_documents({"is_active": True})
    total_teachers = await db.users.count_documents({"role": UserRole.TEACHER, "is_active": True})
    total_courses = await db.courses.count_documents({"is_active": True})
    total_departments = await db.departments.count_documents({})
    
    # Get today's attendance (only from active lectures)
    today_start = get_yemen_time().replace(hour=0, minute=0, second=0, microsecond=0)
    today_records = await db.attendance.find({"date": {"$gte": today_start}}).to_list(10000)
    
    # جلب جميع المحاضرات الفعّالة
    all_active_lecture_ids = await get_active_lecture_ids()
    
    # فلترة سجلات اليوم لاستبعاد المحاضرات الملغاة
    today_records = [r for r in today_records if r.get("lecture_id") in all_active_lecture_ids]
    
    today_present = sum(1 for r in today_records if r["status"] == AttendanceStatus.PRESENT)
    today_absent = sum(1 for r in today_records if r["status"] == AttendanceStatus.ABSENT)
    
    return {
        "total_students": total_students,
        "total_teachers": total_teachers,
        "total_courses": total_courses,
        "total_departments": total_departments,
        "today_attendance": {
            "total": len(today_records),
            "present": today_present,
            "absent": today_absent,
            "rate": round(today_present / len(today_records) * 100, 2) if today_records else 0
        }
    }

# ==================== Advanced Reports ====================

@api_router.get("/reports/attendance-overview")
async def get_attendance_overview_report(
    department_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تقرير الحضور الشامل - نسب الحضور لجميع المقررات"""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # بناء query للمقررات
    course_query = {"is_active": True}
    if department_id:
        course_query["department_id"] = department_id
    
    courses = await db.courses.find(course_query).to_list(100)
    
    result = []
    for course in courses:
        course_id = str(course["_id"])
        
        # جلب المحاضرات الفعالة
        active_lecture_ids = await get_active_lecture_ids(course_id)
        
        # بناء query للحضور
        attendance_query = {"course_id": course_id, "lecture_id": {"$in": list(active_lecture_ids)}}
        if start_date:
            attendance_query["date"] = {"$gte": datetime.fromisoformat(start_date)}
        if end_date:
            if "date" in attendance_query:
                attendance_query["date"]["$lte"] = datetime.fromisoformat(end_date)
            else:
                attendance_query["date"] = {"$lte": datetime.fromisoformat(end_date)}
        
        records = await db.attendance.find(attendance_query).to_list(10000)
        
        total = len(records)
        present = sum(1 for r in records if r.get("status") == AttendanceStatus.PRESENT)
        absent = sum(1 for r in records if r.get("status") == AttendanceStatus.ABSENT)
        late = sum(1 for r in records if r.get("status") == AttendanceStatus.LATE)
        
        # جلب اسم المعلم
        teacher_name = None
        if course.get("teacher_id"):
            try:
                teacher = await db.teachers.find_one({"_id": ObjectId(course["teacher_id"])})
                if teacher:
                    teacher_name = teacher.get("full_name")
            except Exception:
                pass
        
        result.append({
            "course_id": course_id,
            "course_name": course.get("name", ""),
            "course_code": course.get("code", ""),
            "teacher_name": teacher_name,
            "total_records": total,
            "present_count": present,
            "absent_count": absent,
            "late_count": late,
            "attendance_rate": round((present + late * 0.5) / total * 100, 2) if total > 0 else 0
        })
    
    return {
        "courses": result,
        "summary": {
            "total_courses": len(result),
            "avg_attendance_rate": round(sum(c["attendance_rate"] for c in result) / len(result), 2) if result else 0
        }
    }

@api_router.get("/reports/absent-students")
async def get_absent_students_report(
    department_id: Optional[str] = None,
    course_id: Optional[str] = None,
    min_absence_rate: float = 25.0,
    current_user: dict = Depends(get_current_user)
):
    """تقرير الطلاب المتغيبين - الذين تجاوزوا نسبة غياب معينة"""
    if not has_permission(current_user, Permission.VIEW_REPORTS) and not has_permission(current_user, Permission.REPORT_ABSENT_STUDENTS):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # جلب المقررات
    course_query = {"is_active": True}
    if department_id:
        course_query["department_id"] = department_id
    if course_id:
        course_query["_id"] = ObjectId(course_id)
    
    # المعلم يرى مقرراته فقط
    if current_user["role"] == UserRole.TEACHER:
        teacher_record = await db.teachers.find_one({"user_id": current_user["id"]})
        if teacher_record:
            course_query["teacher_id"] = str(teacher_record["_id"])
        else:
            return {"students": [], "total_count": 0, "min_absence_rate_filter": min_absence_rate}
    
    courses = await db.courses.find(course_query).to_list(100)

    # جلب جميع الأقسام مرة واحدة لربط الاسم
    dept_ids = list({c.get("department_id") for c in courses if c.get("department_id")})
    depts_map = {}
    if dept_ids:
        async for d in db.departments.find({"_id": {"$in": [ObjectId(x) for x in dept_ids]}}):
            depts_map[str(d["_id"])] = d.get("name", "")

    result = []
    for course in courses:
        cid = str(course["_id"])
        active_lecture_ids = await get_active_lecture_ids(cid)
        total_lectures = len(active_lecture_ids)
        
        if total_lectures == 0:
            continue
        
        # جلب الطلاب المسجلين
        enrollments = await db.enrollments.find({"course_id": cid}).to_list(1000)
        
        for enrollment in enrollments:
            student_id = enrollment.get("student_id")
            if not student_id:
                continue
            
            # حساب الغياب
            attendance_records = await db.attendance.find({
                "student_id": student_id,
                "course_id": cid,
                "lecture_id": {"$in": list(active_lecture_ids)}
            }).to_list(1000)
            
            absent_count = sum(1 for r in attendance_records if r.get("status") == AttendanceStatus.ABSENT)
            late_count = sum(1 for r in attendance_records if r.get("status") == AttendanceStatus.LATE)
            present_count = sum(1 for r in attendance_records if r.get("status") == AttendanceStatus.PRESENT)
            absence_rate = (absent_count / total_lectures) * 100

            if absence_rate >= min_absence_rate:
                try:
                    student = await db.students.find_one({"_id": ObjectId(student_id)})
                except Exception:
                    continue
                if student:
                    # آخر تاريخ غياب
                    last_absent_date = ""
                    for r in sorted(
                        [x for x in attendance_records if x.get("status") == AttendanceStatus.ABSENT],
                        key=lambda x: x.get("date", ""),
                        reverse=True,
                    ):
                        last_absent_date = str(r.get("date", ""))
                        break

                    dept_name = depts_map.get(course.get("department_id", ""), "")
                    presence_rate = (present_count / total_lectures) * 100 if total_lectures else 0

                    result.append({
                        "student_id": student.get("student_id", ""),
                        "student_name": student.get("full_name", ""),
                        "student_phone": student.get("phone", ""),
                        "department_name": dept_name,
                        "level": course.get("level", student.get("level", "")),
                        "section": course.get("section", student.get("section", "")),
                        "course_name": course.get("name", ""),
                        "course_code": course.get("code", ""),
                        "total_lectures": total_lectures,
                        "absent_count": absent_count,
                        "late_count": late_count,
                        "present_count": present_count,
                        "absence_rate": round(absence_rate, 2),
                        "presence_rate": round(presence_rate, 2),
                        "last_absent_date": last_absent_date,
                    })
    
    # ترتيب حسب نسبة الغياب (الأعلى أولاً)
    result.sort(key=lambda x: x["absence_rate"], reverse=True)
    
    return {
        "students": result,
        "total_count": len(result),
        "min_absence_rate_filter": min_absence_rate
    }

@api_router.get("/reports/student/{student_id}")
async def get_student_attendance_report(
    student_id: str,
    current_user: dict = Depends(get_current_user)
):
    """تقرير حضور طالب واحد في جميع مقرراته"""
    if not has_permission(current_user, Permission.VIEW_REPORTS) and not has_permission(current_user, Permission.VIEW_ATTENDANCE):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # جلب الطالب
    student = await db.students.find_one({"_id": ObjectId(student_id)})
    if not student:
        # محاولة البحث برقم القيد
        student = await db.students.find_one({"student_id": student_id})
    
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    # جلب المقررات المسجل فيها
    enrollments = await db.enrollments.find({"student_id": str(student["_id"])}).to_list(100)
    
    courses_data = []
    total_present = 0
    total_absent = 0
    total_late = 0
    total_lectures = 0
    
    for enrollment in enrollments:
        try:
            course = await db.courses.find_one({"_id": ObjectId(enrollment.get("course_id", ""))})
        except Exception:
            continue
        if not course:
            continue
        
        course_id = str(course["_id"])
        active_lecture_ids = await get_active_lecture_ids(course_id)
        
        # جلب سجلات الحضور
        records = await db.attendance.find({
            "student_id": str(student["_id"]),
            "course_id": course_id,
            "lecture_id": {"$in": list(active_lecture_ids)}
        }).to_list(1000)
        
        present = sum(1 for r in records if r.get("status") == AttendanceStatus.PRESENT)
        absent = sum(1 for r in records if r.get("status") == AttendanceStatus.ABSENT)
        late = sum(1 for r in records if r.get("status") == AttendanceStatus.LATE)
        
        total_present += present
        total_absent += absent
        total_late += late
        total_lectures += len(active_lecture_ids)
        
        courses_data.append({
            "course_name": course.get("name", ""),
            "course_code": course.get("code", ""),
            "total_lectures": len(active_lecture_ids),
            "present": present,
            "absent": absent,
            "late": late,
            "attendance_rate": round((present + late * 0.5) / len(active_lecture_ids) * 100, 2) if active_lecture_ids else 0
        })
    
    return {
        "student": {
            "id": str(student["_id"]),
            "student_id": student.get("student_id"),
            "full_name": student.get("full_name"),
            "department_id": student.get("department_id"),
            "level": student.get("level"),
            "section": student.get("section")
        },
        "courses": courses_data,
        "summary": {
            "total_courses": len(courses_data),
            "total_lectures": total_lectures,
            "total_present": total_present,
            "total_absent": total_absent,
            "total_late": total_late,
            "overall_attendance_rate": round((total_present + total_late * 0.5) / total_lectures * 100, 2) if total_lectures > 0 else 0
        }
    }

@api_router.get("/reports/daily")
async def get_daily_report(
    date: Optional[str] = None,
    department_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تقرير يومي - ملخص الحضور ليوم محدد"""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # تحديد التاريخ
    if date:
        report_date = datetime.fromisoformat(date)
    else:
        report_date = get_yemen_time()
    
    day_start = report_date.replace(hour=0, minute=0, second=0, microsecond=0)
    day_end = day_start + timedelta(days=1)
    
    # جلب المقررات
    course_query = {"is_active": True}
    if department_id:
        course_query["department_id"] = department_id
    courses = await db.courses.find(course_query).to_list(100)
    course_ids = [str(c["_id"]) for c in courses]
    
    # جلب المحاضرات في هذا اليوم
    lectures = await db.lectures.find({
        "course_id": {"$in": course_ids},
        "date": {"$gte": day_start, "$lt": day_end},
        "is_cancelled": {"$ne": True}
    }).to_list(100)
    
    lectures_data = []
    total_present = 0
    total_absent = 0
    total_late = 0
    
    for lecture in lectures:
        course = next((c for c in courses if str(c["_id"]) == lecture["course_id"]), None)
        if not course:
            continue
        
        # جلب سجلات الحضور
        records = await db.attendance.find({
            "lecture_id": str(lecture["_id"])
        }).to_list(1000)
        
        present = sum(1 for r in records if r.get("status") == AttendanceStatus.PRESENT)
        absent = sum(1 for r in records if r.get("status") == AttendanceStatus.ABSENT)
        late = sum(1 for r in records if r.get("status") == AttendanceStatus.LATE)
        
        total_present += present
        total_absent += absent
        total_late += late
        
        lectures_data.append({
            "lecture_id": str(lecture["_id"]),
            "course_name": course.get("name", ""),
            "course_code": course.get("code", ""),
            "start_time": lecture.get("start_time", ""),
            "end_time": lecture.get("end_time", ""),
            "present": present,
            "absent": absent,
            "late": late,
            "total_students": len(records),
            "attendance_rate": round((present + late * 0.5) / len(records) * 100, 2) if records else 0
        })
    
    total_records = total_present + total_absent + total_late
    
    return {
        "date": day_start.isoformat(),
        "lectures": lectures_data,
        "summary": {
            "total_lectures": len(lectures_data),
            "total_students_recorded": total_records,
            "total_present": total_present,
            "total_absent": total_absent,
            "total_late": total_late,
            "overall_attendance_rate": round((total_present + total_late * 0.5) / total_records * 100, 2) if total_records > 0 else 0
        }
    }

@api_router.get("/reports/warnings")
async def get_warnings_report(
    department_id: Optional[str] = None,
    warning_threshold: float = 25.0,
    deprivation_threshold: float = 40.0,
    current_user: dict = Depends(get_current_user)
):
    """تقرير الإنذارات - الطلاب المعرضين للحرمان"""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # جلب المقررات
    course_query = {"is_active": True}
    if department_id:
        course_query["department_id"] = department_id
    courses = await db.courses.find(course_query).to_list(100)
    
    warnings = []
    deprivations = []
    
    for course in courses:
        cid = str(course["_id"])
        active_lecture_ids = await get_active_lecture_ids(cid)
        total_lectures = len(active_lecture_ids)
        
        if total_lectures == 0:
            continue
        
        # جلب التسجيلات
        enrollments = await db.enrollments.find({"course_id": cid}).to_list(1000)
        
        for enrollment in enrollments:
            student_id = enrollment.get("student_id")
            if not student_id:
                continue
            
            # حساب الغياب
            records = await db.attendance.find({
                "student_id": student_id,
                "course_id": cid,
                "lecture_id": {"$in": list(active_lecture_ids)}
            }).to_list(1000)
            
            absent_count = sum(1 for r in records if r.get("status") == AttendanceStatus.ABSENT)
            absence_rate = (absent_count / total_lectures) * 100
            
            if absence_rate >= warning_threshold:
                try:
                    student = await db.students.find_one({"_id": ObjectId(student_id)})
                except Exception:
                    continue
                if student:
                    student_data = {
                        "student_id": student.get("student_id", ""),
                        "student_name": student.get("full_name", ""),
                        "course_name": course.get("name", ""),
                        "course_code": course.get("code", ""),
                        "total_lectures": total_lectures,
                        "absent_count": absent_count,
                        "absence_rate": round(absence_rate, 2),
                        "remaining_allowed": max(0, int((deprivation_threshold / 100) * total_lectures) - absent_count)
                    }
                    
                    if absence_rate >= deprivation_threshold:
                        student_data["status"] = "محروم"
                        deprivations.append(student_data)
                    else:
                        student_data["status"] = "إنذار"
                        warnings.append(student_data)
    
    return {
        "warnings": warnings,
        "deprivations": deprivations,
        "summary": {
            "total_warnings": len(warnings),
            "total_deprivations": len(deprivations),
            "warning_threshold": warning_threshold,
            "deprivation_threshold": deprivation_threshold
        }
    }

@api_router.get("/reports/course/{course_id}/detailed")
async def get_course_detailed_report(
    course_id: str,
    current_user: dict = Depends(get_current_user)
):
    """تقرير المقرر التفصيلي - تحليل كامل لمقرر معين"""
    if not has_permission(current_user, Permission.VIEW_REPORTS) and not has_permission(current_user, Permission.REPORT_COURSE):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    # المعلم يرى مقرراته فقط
    if current_user["role"] == UserRole.TEACHER:
        teacher_record = await db.teachers.find_one({"user_id": current_user["id"]})
        if not teacher_record or course.get("teacher_id") != str(teacher_record["_id"]):
            raise HTTPException(status_code=403, detail="هذا المقرر ليس من مقرراتك")
    
    # جلب المعلم
    teacher_name = None
    if course.get("teacher_id"):
        try:
            teacher = await db.teachers.find_one({"_id": ObjectId(course["teacher_id"])})
            if teacher:
                teacher_name = teacher.get("full_name")
        except Exception:
            pass
    
    # جلب المحاضرات
    active_lecture_ids = await get_active_lecture_ids(course_id)
    lectures = await db.lectures.find({
        "_id": {"$in": [ObjectId(lid) for lid in active_lecture_ids]}
    }).sort("date", 1).to_list(100)
    
    # جلب التسجيلات
    enrollments = await db.enrollments.find({"course_id": course_id}).to_list(1000)
    
    # بيانات الطلاب
    students_data = []
    for enrollment in enrollments:
        try:
            sid = enrollment.get("student_id")
            if not sid:
                continue
            student = await db.students.find_one({"_id": ObjectId(sid)})
            if not student:
                continue
            
            records = await db.attendance.find({
                "student_id": str(student["_id"]),
                "course_id": course_id,
                "lecture_id": {"$in": list(active_lecture_ids)}
            }).to_list(1000)
            
            present = sum(1 for r in records if r.get("status") == AttendanceStatus.PRESENT)
            absent = sum(1 for r in records if r.get("status") == AttendanceStatus.ABSENT)
            late = sum(1 for r in records if r.get("status") == AttendanceStatus.LATE)
            
            students_data.append({
                "student_id": student.get("student_id", ""),
                "student_name": student.get("full_name", ""),
                "present": present,
                "absent": absent,
                "late": late,
                "attendance_rate": round((present + late * 0.5) / len(active_lecture_ids) * 100, 2) if active_lecture_ids else 0
            })
        except Exception as e:
            logger.warning(f"Skipping enrollment {enrollment.get('_id')}: {e}")
            continue
    
    # ترتيب حسب نسبة الحضور
    students_data.sort(key=lambda x: x["attendance_rate"], reverse=True)
    
    # بيانات المحاضرات
    lectures_data = []
    for lecture in lectures:
        try:
            records = await db.attendance.find({"lecture_id": str(lecture["_id"])}).to_list(1000)
            present = sum(1 for r in records if r.get("status") == AttendanceStatus.PRESENT)
            
            date_val = lecture.get("date")
            if date_val and hasattr(date_val, 'isoformat'):
                date_str = date_val.isoformat()
            elif date_val:
                date_str = str(date_val)
            else:
                date_str = ""
            
            lectures_data.append({
                "date": date_str,
                "start_time": lecture.get("start_time", ""),
                "present_count": present,
                "total_students": len(records),
                "attendance_rate": round(present / len(records) * 100, 2) if records else 0
            })
        except Exception as e:
            logger.warning(f"Skipping lecture {lecture.get('_id')}: {e}")
            continue
    
    return {
        "course": {
            "id": course_id,
            "name": course.get("name", ""),
            "code": course.get("code", ""),
            "teacher_name": teacher_name,
            "level": course.get("level"),
            "section": course.get("section", "")
        },
        "students": students_data,
        "lectures": lectures_data,
        "summary": {
            "total_students": len(students_data),
            "total_lectures": len(lectures_data),
            "avg_attendance_rate": round(sum(s["attendance_rate"] for s in students_data) / len(students_data), 2) if students_data else 0
        }
    }

@api_router.get("/reports/teacher-summary")
async def get_teacher_summary_report(
    teacher_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تقرير ملخص المعلم - جميع مقرراته مع نسب الحضور"""
    # المعلم يرى بياناته فقط
    if current_user["role"] == UserRole.TEACHER:
        teacher_record = await db.teachers.find_one({"user_id": current_user["id"]})
        if teacher_record:
            teacher_id = str(teacher_record["_id"])
        else:
            raise HTTPException(status_code=404, detail="لم يتم العثور على سجل المعلم")
    elif not has_permission(current_user, Permission.VIEW_REPORTS) and not has_permission(current_user, Permission.REPORT_TEACHER_WORKLOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # جلب بيانات المعلم
    if teacher_id:
        teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
    else:
        raise HTTPException(status_code=400, detail="يجب تحديد المعلم")
    
    if not teacher:
        raise HTTPException(status_code=404, detail="المعلم غير موجود")
    
    # جلب مقررات المعلم
    courses = await db.courses.find({"teacher_id": teacher_id, "is_active": True}).to_list(50)
    
    courses_summary = []
    total_students = 0
    total_lectures_count = 0
    total_present = 0
    total_absent = 0
    total_late = 0
    total_records = 0
    
    for course in courses:
        cid = str(course["_id"])
        
        # المحاضرات الفعالة
        active_lecture_ids = await get_active_lecture_ids(cid)
        lectures_count = len(active_lecture_ids)
        
        # المحاضرات المنعقدة فعلياً
        held_lectures = await db.lectures.count_documents({
            "_id": {"$in": [ObjectId(lid) for lid in active_lecture_ids]},
            "status": {"$in": ["held", "completed"]}
        }) if active_lecture_ids else 0
        
        # عدد الطلاب المسجلين
        enrollments = await db.enrollments.find({"course_id": cid}).to_list(1000)
        students_count = len(enrollments)
        total_students += students_count
        total_lectures_count += lectures_count
        
        # إحصائيات الحضور
        attendance_records = await db.attendance.find({
            "course_id": cid,
            "lecture_id": {"$in": list(active_lecture_ids)}
        }).to_list(10000) if active_lecture_ids else []
        
        present = sum(1 for r in attendance_records if r["status"] == AttendanceStatus.PRESENT)
        absent = sum(1 for r in attendance_records if r["status"] == AttendanceStatus.ABSENT)
        late = sum(1 for r in attendance_records if r["status"] == AttendanceStatus.LATE)
        course_total = len(attendance_records)
        
        total_present += present
        total_absent += absent
        total_late += late
        total_records += course_total
        
        attendance_rate = round((present + late * 0.5) / course_total * 100, 1) if course_total > 0 else 0
        
        # جلب اسم القسم
        dept_name = ""
        if course.get("department_id"):
            dept = await db.departments.find_one({"_id": ObjectId(course["department_id"])})
            if dept:
                dept_name = dept.get("name", "")
        
        courses_summary.append({
            "course_id": cid,
            "course_name": course["name"],
            "course_code": course.get("code", ""),
            "department_name": dept_name,
            "level": course.get("level", ""),
            "section": course.get("section", ""),
            "students_count": students_count,
            "total_lectures": lectures_count,
            "held_lectures": held_lectures,
            "present_count": present,
            "absent_count": absent,
            "late_count": late,
            "attendance_rate": attendance_rate
        })
    
    # ترتيب حسب نسبة الحضور
    courses_summary.sort(key=lambda x: x["attendance_rate"], reverse=True)
    
    overall_rate = round((total_present + total_late * 0.5) / total_records * 100, 1) if total_records > 0 else 0
    
    return {
        "teacher": {
            "id": teacher_id,
            "full_name": teacher.get("full_name", ""),
            "teacher_id": teacher.get("teacher_id", ""),
            "phone": teacher.get("phone", ""),
            "email": teacher.get("email", "")
        },
        "courses": courses_summary,
        "summary": {
            "total_courses": len(courses_summary),
            "total_students": total_students,
            "total_lectures": total_lectures_count,
            "total_present": total_present,
            "total_absent": total_absent,
            "total_late": total_late,
            "overall_attendance_rate": overall_rate
        }
    }

@api_router.get("/export/report/teacher-summary/excel")
async def export_teacher_summary_excel(
    teacher_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير تقرير ملخص المعلم إلى Excel"""
    if not has_permission(current_user, Permission.EXPORT_REPORTS):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    report = await get_teacher_summary_report(teacher_id, current_user)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # ورقة الملخص
        summary_data = [{
            "المعلم": report["teacher"]["full_name"],
            "الرقم الوظيفي": report["teacher"]["teacher_id"],
            "عدد المقررات": report["summary"]["total_courses"],
            "إجمالي الطلاب": report["summary"]["total_students"],
            "إجمالي المحاضرات": report["summary"]["total_lectures"],
            "نسبة الحضور العامة %": report["summary"]["overall_attendance_rate"]
        }]
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="الملخص", index=False)
        
        # ورقة المقررات
        courses_data = []
        for idx, c in enumerate(report["courses"], 1):
            courses_data.append({
                "#": idx,
                "المقرر": c["course_name"],
                "الرمز": c["course_code"],
                "القسم": c["department_name"],
                "المستوى": c["level"],
                "الشعبة": c["section"],
                "الطلاب": c["students_count"],
                "المحاضرات": c["total_lectures"],
                "المنعقدة": c["held_lectures"],
                "حاضر": c["present_count"],
                "غائب": c["absent_count"],
                "متأخر": c["late_count"],
                "نسبة الحضور %": c["attendance_rate"]
            })
        if courses_data:
            pd.DataFrame(courses_data).to_excel(writer, sheet_name="المقررات", index=False)
    
    output.seek(0)
    teacher_name = report["teacher"]["full_name"].replace(" ", "_")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=teacher_summary_{teacher_name}.xlsx"}
    )




@api_router.get("/reports/teacher-delays")
async def get_teacher_delays_report(
    start_date: str = None,
    end_date: str = None,
    department_id: str = None,
    faculty_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    """تقرير تأخر المعلمين في بدء التحضير"""
    if not has_permission(current_user, "view_reports"):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    # فلتر المحاضرات التي بدأ فيها التحضير
    query = {"attendance_started_at": {"$exists": True, "$ne": None}}
    
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    lectures = await db.lectures.find(query).to_list(1000)
    
    # تجميع التأخيرات حسب المعلم
    teacher_delays = {}
    
    for lec in lectures:
        course = await db.courses.find_one({"_id": ObjectId(lec["course_id"])})
        if not course:
            continue
        
        # فلتر القسم/الكلية
        if department_id and course.get("department_id") != department_id:
            continue
        if faculty_id:
            dept = await db.departments.find_one({"_id": ObjectId(course.get("department_id", ""))})
            if not dept or dept.get("faculty_id") != faculty_id:
                continue
        
        teacher_id = course.get("teacher_id")
        if not teacher_id:
            continue
        
        # حساب التأخير
        try:
            lecture_start = datetime.strptime(f"{lec['date']}T{lec['start_time']}", "%Y-%m-%dT%H:%M")
            started_at = lec["attendance_started_at"]
            if isinstance(started_at, str):
                started_at = datetime.fromisoformat(started_at)
            
            # إزالة timezone للمقارنة العادلة
            if hasattr(started_at, 'tzinfo') and started_at.tzinfo is not None:
                started_at = started_at.replace(tzinfo=None)
            
            # تجاهل إذا كان التحضير في يوم مختلف (تحضير متأخر جداً)
            if started_at.date() != lecture_start.date():
                continue
            
            delay_minutes = int((started_at - lecture_start).total_seconds() / 60)
            if delay_minutes < 0:
                delay_minutes = 0
        except Exception:
            continue
        
        if teacher_id not in teacher_delays:
            # جلب بيانات المعلم
            teacher = None
            try:
                teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
                if not teacher:
                    teacher = await db.users.find_one({"_id": ObjectId(teacher_id)})
            except Exception:
                pass
            
            teacher_delays[teacher_id] = {
                "teacher_id": teacher_id,
                "teacher_name": teacher["full_name"] if teacher else "غير معروف",
                "employee_id": teacher.get("teacher_id", "") if teacher else "",
                "department": course.get("department_id", ""),
                "total_lectures": 0,
                "delayed_lectures": 0,
                "total_delay_minutes": 0,
                "max_delay_minutes": 0,
                "delays": []
            }
        
        entry = teacher_delays[teacher_id]
        entry["total_lectures"] += 1
        
        if delay_minutes > 0:
            entry["delayed_lectures"] += 1
            entry["total_delay_minutes"] += delay_minutes
            if delay_minutes > entry["max_delay_minutes"]:
                entry["max_delay_minutes"] = delay_minutes
            entry["delays"].append({
                "date": lec["date"],
                "course_name": course["name"],
                "start_time": lec["start_time"],
                "started_at": started_at.strftime("%H:%M") if hasattr(started_at, 'strftime') else str(started_at),
                "delay_minutes": delay_minutes
            })
    
    # ترتيب حسب إجمالي التأخير
    result = sorted(teacher_delays.values(), key=lambda x: x["total_delay_minutes"], reverse=True)
    
    # حساب المتوسط
    for t in result:
        t["avg_delay_minutes"] = round(t["total_delay_minutes"] / t["delayed_lectures"], 1) if t["delayed_lectures"] > 0 else 0
    
    return {
        "teachers": result,
        "summary": {
            "total_teachers": len(result),
            "total_delayed_teachers": len([t for t in result if t["delayed_lectures"] > 0]),
            "total_delay_incidents": sum(t["delayed_lectures"] for t in result),
        }
    }


@api_router.get("/reports/teacher-delays/export")
async def export_teacher_delays_report(
    start_date: str = None,
    end_date: str = None,
    department_id: str = None,
    faculty_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير تقرير تأخر المعلمين إلى Excel"""
    if not has_permission(current_user, "export_reports"):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    # استخدام نفس البيانات
    report = await get_teacher_delays_report(start_date, end_date, department_id, faculty_id, current_user)
    
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تأخر المعلمين"
    ws.sheet_view.rightToLeft = True
    
    # العنوان
    ws.merge_cells('A1:G1')
    ws['A1'] = 'تقرير تأخر المعلمين في بدء التحضير'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # ملخص
    ws['A3'] = f"إجمالي المعلمين: {report['summary']['total_teachers']}"
    ws['D3'] = f"حالات التأخر: {report['summary']['total_delay_incidents']}"
    
    # رؤوس الأعمدة
    headers = ['المعلم', 'الرقم الوظيفي', 'إجمالي المحاضرات', 'محاضرات متأخرة', 'إجمالي التأخير (دقيقة)', 'متوسط التأخير', 'أقصى تأخير']
    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # البيانات
    for row, t in enumerate(report['teachers'], 6):
        ws.cell(row=row, column=1, value=t['teacher_name'])
        ws.cell(row=row, column=2, value=t['employee_id'])
        ws.cell(row=row, column=3, value=t['total_lectures'])
        ws.cell(row=row, column=4, value=t['delayed_lectures'])
        ws.cell(row=row, column=5, value=t['total_delay_minutes'])
        ws.cell(row=row, column=6, value=t['avg_delay_minutes'])
        ws.cell(row=row, column=7, value=t['max_delay_minutes'])
        
        # تلوين المتأخرين
        if t['delayed_lectures'] > 0:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    
    # تفاصيل التأخيرات
    if report['teachers']:
        detail_row = len(report['teachers']) + 8
        ws.merge_cells(f'A{detail_row}:G{detail_row}')
        ws.cell(row=detail_row, column=1, value='تفاصيل التأخيرات')
        ws.cell(row=detail_row, column=1).font = Font(bold=True, size=12)
        
        detail_row += 1
        detail_headers = ['المعلم', 'المقرر', 'التاريخ', 'وقت البدء المحدد', 'وقت البدء الفعلي', 'التأخير (دقيقة)']
        for col, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=detail_row, column=col, value=h)
            cell.fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        detail_row += 1
        for t in report['teachers']:
            for d in t['delays']:
                ws.cell(row=detail_row, column=1, value=t['teacher_name'])
                ws.cell(row=detail_row, column=2, value=d['course_name'])
                ws.cell(row=detail_row, column=3, value=d['date'])
                ws.cell(row=detail_row, column=4, value=d['start_time'])
                ws.cell(row=detail_row, column=5, value=d['started_at'])
                ws.cell(row=detail_row, column=6, value=d['delay_minutes'])
                detail_row += 1
    
    # ضبط عرض الأعمدة
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=teacher_delays_report.xlsx"}
    )


@api_router.get("/reports/teacher-workload")
async def get_teacher_workload_report(
    teacher_id: Optional[str] = None,
    start_date: str = None,
    end_date: str = None,
    current_user: dict = Depends(get_current_user)
):
    """تقرير نصاب المدرس - كم نصابه، كم درسها فعلياً، كم ساعات زائدة"""
    # السماح للمعلم بالوصول لتقريره الخاص
    if current_user["role"] == UserRole.TEACHER:
        # المعلم يرى تقريره فقط
        teacher_record = await db.teachers.find_one({"user_id": current_user["id"]})
        if teacher_record:
            teacher_id = str(teacher_record["_id"])
        else:
            raise HTTPException(status_code=404, detail="لم يتم العثور على سجل المعلم")
    elif not has_permission(current_user, Permission.VIEW_REPORTS) and not has_permission(current_user, Permission.REPORT_TEACHER_WORKLOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # تحديد الفترة
    if not start_date or not end_date:
        # افتراضي: الشهر الحالي
        today = get_yemen_time()
        start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end = (start + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
    else:
        start = datetime.fromisoformat(start_date)
        end = datetime.fromisoformat(end_date)
    
    # جلب المعلمين
    teacher_query = {"is_active": {"$ne": False}}
    if teacher_id:
        teacher_query["_id"] = ObjectId(teacher_id)
    
    teachers = await db.teachers.find(teacher_query).to_list(100)
    
    result = []
    for teacher in teachers:
        tid = str(teacher["_id"])
        
        # النصاب الأسبوعي (افتراضي 12 ساعة)
        weekly_hours = teacher.get("weekly_hours", 12)
        
        # حساب عدد الأسابيع في الفترة (أسابيع كاملة فقط)
        total_days = (end - start).days + 1
        total_weeks = total_days // 7
        
        # الساعات المطلوبة = النصاب الأسبوعي × عدد الأسابيع
        required_hours = round(weekly_hours * total_weeks, 2)
        
        # جلب المقررات
        courses = await db.courses.find({"teacher_id": tid, "is_active": True}).to_list(50)
        
        total_scheduled_hours = 0
        total_actual_hours = 0
        courses_data = []
        
        for course in courses:
            cid = str(course["_id"])
            
            # المحاضرات المجدولة في الفترة
            # تحويل التواريخ لمقارنة النصوص
            start_str = start.strftime("%Y-%m-%d")
            end_str = end.strftime("%Y-%m-%d")

            # المحاضرات المجدولة في الفترة (دعم date كنص أو datetime)
            scheduled_lectures = await db.lectures.find({
                "course_id": cid,
                "$or": [
                    {"date": {"$gte": start_str, "$lte": end_str}},
                    {"date": {"$gte": start, "$lte": end}}
                ],
                "is_cancelled": {"$ne": True}
            }).to_list(500)
            
            # المحاضرات التي تم تسجيل حضور فيها (المنفذة فعلياً)
            executed_lectures = []
            for lecture in scheduled_lectures:
                attendance_count = await db.attendance.count_documents({"lecture_id": str(lecture["_id"])})
                if attendance_count > 0:
                    executed_lectures.append(lecture)
            
            # حساب الساعات
            scheduled_hours = 0
            actual_hours = 0
            
            for lecture in scheduled_lectures:
                try:
                    start_time = datetime.strptime(lecture.get("start_time", "00:00"), "%H:%M")
                    end_time = datetime.strptime(lecture.get("end_time", "00:00"), "%H:%M")
                    duration = (end_time - start_time).seconds / 3600
                    scheduled_hours += duration
                except:
                    scheduled_hours += 1
            
            for lecture in executed_lectures:
                try:
                    start_time = datetime.strptime(lecture.get("start_time", "00:00"), "%H:%M")
                    end_time = datetime.strptime(lecture.get("end_time", "00:00"), "%H:%M")
                    duration = (end_time - start_time).seconds / 3600
                    actual_hours += duration
                except:
                    actual_hours += 1
            
            total_scheduled_hours += scheduled_hours
            total_actual_hours += actual_hours
            
            courses_data.append({
                "course_name": course["name"],
                "course_code": course.get("code", ""),
                "scheduled_lectures": len(scheduled_lectures),
                "executed_lectures": len(executed_lectures),
                "scheduled_hours": round(scheduled_hours, 2),
                "actual_hours": round(actual_hours, 2)
            })
        
        # الفرق = الساعات المنفذة - الساعات المطلوبة (حسب النصاب الأسبوعي)
        difference_hours = round(total_actual_hours - required_hours, 2)
        
        result.append({
            "teacher_id": teacher.get("teacher_id", ""),
            "teacher_name": teacher.get("full_name", ""),
            "department_id": teacher.get("department_id"),
            "weekly_hours": weekly_hours,
            "courses": courses_data,
            "summary": {
                "total_courses": len(courses_data),
                "weekly_hours": weekly_hours,
                "total_weeks": round(total_weeks, 1),
                "required_hours": required_hours,
                "total_scheduled_hours": round(total_scheduled_hours, 2),
                "total_actual_hours": round(total_actual_hours, 2),
                "difference_hours": difference_hours,
                "completion_rate": round((total_actual_hours / required_hours) * 100, 2) if required_hours > 0 else 0
            }
        })
    
    return {
        "period": {
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
            "total_weeks": round((end - start).days / 7, 1)
        },
        "teachers": result,
        "summary": {
            "total_teachers": len(result),
            "total_required_hours": round(sum(t["summary"]["required_hours"] for t in result), 2),
            "total_scheduled_hours": round(sum(t["summary"]["total_scheduled_hours"] for t in result), 2),
            "total_actual_hours": round(sum(t["summary"]["total_actual_hours"] for t in result), 2),
            "total_difference_hours": round(sum(t["summary"]["difference_hours"] for t in result), 2)
        }
    }

# ==================== Reports Export (PDF & Excel) ====================

@api_router.get("/export/report/warnings/excel")
async def export_warnings_excel(
    department_id: Optional[str] = None,
    warning_threshold: float = 25.0,
    deprivation_threshold: float = 40.0,
    current_user: dict = Depends(get_current_user)
):
    """تصدير تقرير الإنذارات إلى Excel"""
    if not has_permission(current_user, Permission.EXPORT_REPORTS):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # جلب البيانات
    report = await get_warnings_report(department_id, warning_threshold, deprivation_threshold, current_user)
    
    # إعداد البيانات
    all_data = []
    for item in report["warnings"]:
        item["الحالة"] = "إنذار"
        all_data.append(item)
    for item in report["deprivations"]:
        item["الحالة"] = "محروم"
        all_data.append(item)
    
    if not all_data:
        all_data = [{"ملاحظة": "لا توجد بيانات"}]
    
    # تحويل للـ DataFrame
    df = pd.DataFrame(all_data)
    df = df.rename(columns={
        "student_id": "رقم القيد",
        "student_name": "اسم الطالب",
        "course_name": "المقرر",
        "course_code": "رمز المقرر",
        "total_lectures": "المحاضرات",
        "absent_count": "الغياب",
        "absence_rate": "نسبة الغياب %",
        "remaining_allowed": "المتبقي المسموح",
        "status": "الحالة"
    })
    
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=warnings_report.xlsx"}
    )

@api_router.get("/export/report/absent-students/excel")
async def export_absent_students_excel(
    department_id: Optional[str] = None,
    course_id: Optional[str] = None,
    min_absence_rate: float = 25.0,
    current_user: dict = Depends(get_current_user)
):
    """تصدير تقرير الطلاب المتغيبين إلى Excel"""
    if not has_permission(current_user, Permission.EXPORT_REPORTS):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    report = await get_absent_students_report(department_id, course_id, min_absence_rate, current_user)
    
    if not report["students"]:
        data = [{"ملاحظة": "لا توجد بيانات"}]
    else:
        data = report["students"]
    
    df = pd.DataFrame(data)
    df = df.rename(columns={
        "student_id": "رقم القيد",
        "student_name": "اسم الطالب",
        "course_name": "المقرر",
        "course_code": "رمز المقرر",
        "total_lectures": "المحاضرات",
        "absent_count": "الغياب",
        "absence_rate": "نسبة الغياب %"
    })
    
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=absent_students.xlsx"}
    )

@api_router.get("/export/report/teacher-workload/excel")
async def export_teacher_workload_excel(
    teacher_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    department_id: Optional[str] = None,
    monthly: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """تصدير تقرير نصاب المدرسين إلى Excel
    - إذا monthly=true والفترة تشمل أكثر من شهر، يُصدّر كل شهر في sheet منفصل
    - يدعم فلتر القسم
    """
    if not has_permission(current_user, Permission.EXPORT_REPORTS):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    output = BytesIO()

    def teacher_rows(teachers_list, dept_filter):
        rows = []
        for teacher in teachers_list:
            if dept_filter and teacher.get("department_id") != dept_filter:
                continue
            for course in teacher["courses"]:
                rows.append({
                    "الرقم الوظيفي": teacher.get("teacher_id", ""),
                    "اسم المدرس": teacher.get("teacher_name", ""),
                    "المقرر": course["course_name"],
                    "رمز المقرر": course.get("course_code", ""),
                    "المحاضرات المجدولة": course["scheduled_lectures"],
                    "المحاضرات المنفذة": course["executed_lectures"],
                    "الساعات المجدولة": course["scheduled_hours"],
                    "الساعات المنفذة": course["actual_hours"],
                })
            s = teacher.get("summary", {})
            rows.append({
                "الرقم الوظيفي": teacher.get("teacher_id", ""),
                "اسم المدرس": teacher.get("teacher_name", ""),
                "المقرر": "*** الإجمالي ***",
                "رمز المقرر": "",
                "المحاضرات المجدولة": "",
                "المحاضرات المنفذة": "",
                "الساعات المجدولة": s.get("total_scheduled_hours", 0),
                "الساعات المنفذة": s.get("total_actual_hours", 0),
                "النصاب المطلوب": s.get("required_hours", 0),
                "الفرق": s.get("difference_hours", 0),
                "نسبة الإنجاز %": s.get("completion_rate", 0),
            })
        return rows

    # تحديد ما إذا كانت الفترة متعددة الأشهر
    months_to_export = []
    if monthly and start_date and end_date:
        try:
            from datetime import datetime as _dt
            s = _dt.fromisoformat(start_date)
            e = _dt.fromisoformat(end_date)
            cur = _dt(s.year, s.month, 1)
            while cur <= e:
                month_start = cur
                next_month = _dt(cur.year + (1 if cur.month == 12 else 0), 1 if cur.month == 12 else cur.month + 1, 1)
                month_end = next_month - timedelta(days=1)
                # حد بداية الشهر بـ start_date و نهايته بـ end_date
                actual_start = max(month_start, s)
                actual_end = min(month_end, e)
                months_to_export.append((actual_start.strftime("%Y-%m-%d"), actual_end.strftime("%Y-%m-%d"), cur.strftime("%Y-%m")))
                cur = next_month
        except Exception:
            months_to_export = []

    if not months_to_export:
        # Sheet واحد للفترة كاملة
        report = await get_teacher_workload_report(teacher_id, start_date, end_date, current_user)
        data = teacher_rows(report["teachers"], department_id)
        if not data:
            data = [{"ملاحظة": "لا توجد بيانات"}]
        df = pd.DataFrame(data)
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="نصاب المدرسين")
    else:
        # عدة Sheets - واحد لكل شهر
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for s_str, e_str, label in months_to_export:
                report = await get_teacher_workload_report(teacher_id, s_str, e_str, current_user)
                data = teacher_rows(report["teachers"], department_id)
                if not data:
                    data = [{"ملاحظة": "لا توجد بيانات لهذا الشهر"}]
                df = pd.DataFrame(data)
                # اسم Sheet: YYYY-MM
                df.to_excel(writer, index=False, sheet_name=label[:31])

    output.seek(0)
    filename = "teacher_workload_monthly.xlsx" if months_to_export else "teacher_workload.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/export/report/daily/excel")
async def export_daily_excel(
    date: Optional[str] = None,
    department_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير التقرير اليومي إلى Excel"""
    if not has_permission(current_user, Permission.EXPORT_REPORTS):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    report = await get_daily_report(date, department_id, current_user)
    
    data = []
    for lecture in report["lectures"]:
        data.append({
            "المقرر": lecture["course_name"],
            "رمز المقرر": lecture["course_code"],
            "وقت البداية": lecture["start_time"],
            "وقت النهاية": lecture["end_time"],
            "حاضر": lecture["present"],
            "غائب": lecture["absent"],
            "متأخر": lecture["late"],
            "الإجمالي": lecture["total_students"],
            "نسبة الحضور %": lecture["attendance_rate"]
        })
    
    if not data:
        data = [{"ملاحظة": "لا توجد محاضرات في هذا اليوم"}]
    
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    
    report_date = date or get_yemen_time().strftime("%Y-%m-%d")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=daily_report_{report_date}.xlsx"}
    )

@api_router.get("/export/report/student/{student_id}/excel")
async def export_student_report_excel(
    student_id: str,
    current_user: dict = Depends(get_current_user)
):
    """تصدير تقرير طالب إلى Excel"""
    if not has_permission(current_user, Permission.EXPORT_REPORTS):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    report = await get_student_attendance_report(student_id, current_user)
    
    # بيانات الطالب
    student_info = report["student"]
    
    data = []
    for course in report["courses"]:
        data.append({
            "المقرر": course["course_name"],
            "رمز المقرر": course["course_code"],
            "المحاضرات": course["total_lectures"],
            "حاضر": course["present"],
            "غائب": course["absent"],
            "متأخر": course["late"],
            "نسبة الحضور %": course["attendance_rate"]
        })
    
    # إضافة صف الملخص
    data.append({
        "المقرر": "*** الإجمالي ***",
        "رمز المقرر": "",
        "المحاضرات": report["summary"]["total_lectures"],
        "حاضر": report["summary"]["total_present"],
        "غائب": report["summary"]["total_absent"],
        "متأخر": report["summary"]["total_late"],
        "نسبة الحضور %": report["summary"]["overall_attendance_rate"]
    })
    
    df = pd.DataFrame(data)
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # ورقة معلومات الطالب
        info_df = pd.DataFrame([{
            "رقم القيد": student_info["student_id"],
            "الاسم": student_info["full_name"],
            "المستوى": student_info["level"],
            "الشعبة": student_info.get("section", "")
        }])
        info_df.to_excel(writer, sheet_name="معلومات الطالب", index=False)
        # ورقة التقرير
        df.to_excel(writer, sheet_name="تفاصيل الحضور", index=False)
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=student_{student_info['student_id']}_report.xlsx"}
    )


@api_router.get("/export/report/student/{student_id}/pdf")
async def export_student_report_pdf(
    student_id: str,
    current_user: dict = Depends(get_current_user)
):
    """تصدير تقرير طالب إلى PDF"""
    if not has_permission(current_user, Permission.EXPORT_REPORTS):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    report = await get_student_attendance_report(student_id, current_user)
    student_info = report["student"]
    summary = report["summary"]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    elements = []
    add_pdf_header(
        elements,
        title=f"تقرير حضور الطالب",
        subtitle=f"{student_info.get('full_name', '')} - رقم القيد: {student_info.get('student_id', '')}",
    )

    # بطاقة معلومات الطالب
    info_data = [
        [arabic_text("القيمة"), arabic_text("البيان")],
        [arabic_text(student_info.get("full_name", "")), arabic_text("الاسم")],
        [arabic_text(str(student_info.get("student_id", ""))), arabic_text("رقم القيد")],
        [arabic_text(f"م{student_info.get('level', '')}"), arabic_text("المستوى")],
        [arabic_text(student_info.get("section", "") or "—"), arabic_text("الشعبة")],
        [arabic_text(f"{summary.get('overall_attendance_rate', 0)}%"), arabic_text("نسبة الحضور الإجمالية")],
    ]
    info_table = Table(info_data, colWidths=[260, 200])
    info_table.setStyle(get_pro_table_style(name_cols=[0, 1]))
    elements.append(info_table)
    elements.append(Spacer(1, 14))

    # جدول المقررات
    elements.append(Paragraph(
        arabic_text("تفاصيل الحضور حسب المقرر"),
        ParagraphStyle('SH', fontName=ARABIC_FONT, fontSize=12, alignment=TA_CENTER, spaceAfter=6, textColor=colors.HexColor('#1565c0'))
    ))

    courses_data = [[
        arabic_text("نسبة %"),
        arabic_text("متأخر"),
        arabic_text("غائب"),
        arabic_text("حاضر"),
        arabic_text("المحاضرات"),
        arabic_text("الكود"),
        arabic_text("المقرر"),
    ]]
    for course in report["courses"]:
        courses_data.append([
            arabic_text(f"{course.get('attendance_rate', 0)}%"),
            arabic_text(str(course.get("late", 0))),
            arabic_text(str(course.get("absent", 0))),
            arabic_text(str(course.get("present", 0))),
            arabic_text(str(course.get("total_lectures", 0))),
            arabic_text(course.get("course_code", "")),
            arabic_text(course.get("course_name", "")),
        ])
    # صف الإجمالي
    courses_data.append([
        arabic_text(f"{summary.get('overall_attendance_rate', 0)}%"),
        arabic_text(str(summary.get("total_late", 0))),
        arabic_text(str(summary.get("total_absent", 0))),
        arabic_text(str(summary.get("total_present", 0))),
        arabic_text(str(summary.get("total_lectures", 0))),
        arabic_text(""),
        arabic_text("الإجمالي"),
    ])
    courses_table = Table(courses_data, colWidths=[55, 50, 50, 50, 65, 70, 180], repeatRows=1)
    courses_table.setStyle(get_pro_table_style(num_cols=[0, 1, 2, 3, 4], name_cols=[6]))
    elements.append(courses_table)

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=student_{student_info.get('student_id', student_id)}_report.pdf"}
    )
async def export_course_report_excel(
    course_id: str,
    current_user: dict = Depends(get_current_user)
):
    """تصدير تقرير مقرر إلى Excel"""
    if not has_permission(current_user, Permission.EXPORT_REPORTS):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    report = await get_course_detailed_report(course_id, current_user)
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # ورقة الطلاب
        students_data = []
        for idx, student in enumerate(report["students"], 1):
            students_data.append({
                "الترتيب": idx,
                "رقم القيد": student["student_id"],
                "اسم الطالب": student["student_name"],
                "حاضر": student["present"],
                "غائب": student["absent"],
                "متأخر": student["late"],
                "نسبة الحضور %": student["attendance_rate"]
            })
        
        if students_data:
            pd.DataFrame(students_data).to_excel(writer, sheet_name="الطلاب", index=False)
        
        # ورقة المحاضرات
        lectures_data = []
        for lecture in report["lectures"]:
            lectures_data.append({
                "التاريخ": lecture["date"],
                "الوقت": lecture["start_time"],
                "عدد الحاضرين": lecture["present_count"],
                "الإجمالي": lecture["total_students"],
                "نسبة الحضور %": lecture["attendance_rate"]
            })
        
        if lectures_data:
            pd.DataFrame(lectures_data).to_excel(writer, sheet_name="المحاضرات", index=False)
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=course_{report['course']['code']}_report.xlsx"}
    )

@api_router.get("/export/report/attendance-overview/excel")
async def export_attendance_overview_excel(
    department_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير تقرير الحضور الشامل إلى Excel"""
    if not has_permission(current_user, Permission.EXPORT_REPORTS):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    report = await get_attendance_overview_report(department_id, None, None, current_user)
    
    data = []
    for course in report["courses"]:
        data.append({
            "المقرر": course["course_name"],
            "رمز المقرر": course["course_code"],
            "المدرس": course.get("teacher_name", ""),
            "إجمالي السجلات": course["total_records"],
            "حاضر": course["present_count"],
            "غائب": course["absent_count"],
            "متأخر": course["late_count"],
            "نسبة الحضور %": course["attendance_rate"]
        })
    
    if not data:
        data = [{"ملاحظة": "لا توجد بيانات"}]
    
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=attendance_overview.xlsx"}
    )

# ==================== Semester Report PDF ====================

@api_router.get("/export/semester-report/pdf")
async def export_semester_report_pdf(
    course_id: Optional[str] = None,
    department_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تقرير PDF شامل للفصل الدراسي - جميع المقررات أو مقرر محدد"""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from bidi.algorithm import get_display
    import arabic_reshaper
    
    # Arabic font
    font_path = Path(__file__).parent / "fonts" / "Amiri-Regular.ttf"
    if font_path.exists():
        try:
            pdfmetrics.registerFont(TTFont('Amiri', str(font_path)))
        except:
            pass
        arabic_font = 'Amiri'
    else:
        arabic_font = 'Helvetica'
    
    def draw_arabic(c, text, x, y, font_name=arabic_font, font_size=12):
        reshaped = arabic_reshaper.reshape(str(text))
        bidi_text = get_display(reshaped)
        c.setFont(font_name, font_size)
        c.drawRightString(x, y, bidi_text)
    
    def draw_arabic_left(c, text, x, y, font_name=arabic_font, font_size=12):
        reshaped = arabic_reshaper.reshape(str(text))
        bidi_text = get_display(reshaped)
        c.setFont(font_name, font_size)
        c.drawString(x, y, bidi_text)
    
    # جلب المقررات
    course_query = {"is_active": True}
    if course_id:
        course_query["_id"] = ObjectId(course_id)
    if department_id:
        course_query["department_id"] = department_id
    
    courses_list = await db.courses.find(course_query).to_list(200)
    if not courses_list:
        raise HTTPException(status_code=404, detail="لا توجد مقررات")
    
    # جلب بيانات الأقسام والمعلمين
    dept_cache = {}
    teacher_cache = {}
    
    buffer = BytesIO()
    width, height = A4
    c_pdf = canvas.Canvas(buffer, pagesize=A4)
    
    now = get_yemen_time()
    
    # =================== صفحة الغلاف ===================
    y = height - 80
    draw_arabic(c_pdf, "جامعة الأحقاف", width - 30, y, font_size=24)
    y -= 35
    draw_arabic(c_pdf, "تقرير الفصل الدراسي الشامل", width - 30, y, font_size=20)
    y -= 10
    c_pdf.setStrokeColor(colors.HexColor("#1565c0"))
    c_pdf.setLineWidth(2)
    c_pdf.line(30, y, width - 30, y)
    y -= 30
    
    draw_arabic(c_pdf, f"تاريخ التقرير: {now.strftime('%Y-%m-%d %H:%M')}", width - 30, y, font_size=12)
    y -= 22
    draw_arabic(c_pdf, f"عدد المقررات: {len(courses_list)}", width - 30, y, font_size=12)
    y -= 22
    
    if department_id:
        dept = await db.departments.find_one({"_id": ObjectId(department_id)})
        if dept:
            draw_arabic(c_pdf, f"القسم: {dept.get('name', '')}", width - 30, y, font_size=12)
            y -= 22
            if dept.get("college"):
                draw_arabic(c_pdf, f"الكلية: {dept['college']}", width - 30, y, font_size=12)
                y -= 22
    
    # ملخص عام
    y -= 20
    total_lectures_all = 0
    total_completed_all = 0
    total_absent_all = 0
    total_cancelled_all = 0
    total_students_all = set()
    
    for course in courses_list:
        cid = str(course["_id"])
        lectures = await db.lectures.find({"course_id": cid}).to_list(500)
        total_lectures_all += len(lectures)
        total_completed_all += sum(1 for l in lectures if l.get("status") == LectureStatus.COMPLETED)
        total_absent_all += sum(1 for l in lectures if l.get("status") == LectureStatus.ABSENT)
        total_cancelled_all += sum(1 for l in lectures if l.get("status") == LectureStatus.CANCELLED)
        enrollments = await db.enrollments.find({"course_id": cid}).to_list(1000)
        for e in enrollments:
            total_students_all.add(e["student_id"])
    
    c_pdf.setFillColor(colors.HexColor("#e3f2fd"))
    c_pdf.roundRect(30, y - 80, width - 60, 90, 8, fill=1, stroke=0)
    c_pdf.setFillColor(colors.black)
    y -= 15
    draw_arabic(c_pdf, "ملخص عام", width - 40, y, font_size=14)
    y -= 22
    draw_arabic(c_pdf, f"إجمالي المحاضرات: {total_lectures_all}          المنعقدة: {total_completed_all}          الغائب: {total_absent_all}          الملغاة: {total_cancelled_all}", width - 40, y, font_size=11)
    y -= 22
    scheduled_all = total_lectures_all - total_completed_all - total_absent_all - total_cancelled_all
    draw_arabic(c_pdf, f"المجدولة: {scheduled_all}          إجمالي الطلاب: {len(total_students_all)}", width - 40, y, font_size=11)
    
    c_pdf.showPage()
    
    # =================== تفاصيل كل مقرر ===================
    for course_idx, course in enumerate(courses_list):
        cid = str(course["_id"])
        
        # جلب المعلم
        teacher_name = ""
        if course.get("teacher_id"):
            if course["teacher_id"] not in teacher_cache:
                try:
                    t = await db.teachers.find_one({"_id": ObjectId(course["teacher_id"])})
                    teacher_cache[course["teacher_id"]] = t.get("full_name", "") if t else ""
                except:
                    teacher_cache[course["teacher_id"]] = ""
            teacher_name = teacher_cache[course["teacher_id"]]
        
        # جلب القسم
        dept_name = ""
        college_name = ""
        if course.get("department_id"):
            if course["department_id"] not in dept_cache:
                try:
                    d = await db.departments.find_one({"_id": ObjectId(course["department_id"])})
                    dept_cache[course["department_id"]] = d if d else {}
                except:
                    dept_cache[course["department_id"]] = {}
            dept_obj = dept_cache[course["department_id"]]
            dept_name = dept_obj.get("name", "")
            college_name = dept_obj.get("college", "")
        
        # جلب المحاضرات
        lectures = await db.lectures.find({"course_id": cid}).sort("date", 1).to_list(500)
        active_lectures = [l for l in lectures if l.get("status") != LectureStatus.CANCELLED]
        
        # إحصائيات المحاضرات
        completed_count = sum(1 for l in lectures if l.get("status") == LectureStatus.COMPLETED)
        absent_count = sum(1 for l in lectures if l.get("status") == LectureStatus.ABSENT)
        cancelled_count = sum(1 for l in lectures if l.get("status") == LectureStatus.CANCELLED)
        scheduled_count = sum(1 for l in lectures if l.get("status") == LectureStatus.SCHEDULED)
        
        # جلب الطلاب
        enrollments = await db.enrollments.find({"course_id": cid}).to_list(1000)
        student_ids = [ObjectId(e["student_id"]) for e in enrollments]
        students = await db.students.find({"_id": {"$in": student_ids}}).to_list(1000) if student_ids else []
        
        active_lecture_ids = {str(l["_id"]) for l in active_lectures}
        
        # =================== رأس صفحة المقرر ===================
        y = height - 40
        c_pdf.setFillColor(colors.HexColor("#1565c0"))
        c_pdf.roundRect(30, y - 15, width - 60, 35, 5, fill=1, stroke=0)
        c_pdf.setFillColor(colors.white)
        draw_arabic(c_pdf, f"المقرر {course_idx + 1}/{len(courses_list)}: {course['name']} ({course.get('code', '')})", width - 40, y, font_size=14)
        c_pdf.setFillColor(colors.black)
        
        y -= 40
        if teacher_name:
            draw_arabic(c_pdf, f"المعلم: {teacher_name}", width - 30, y, font_size=11)
            y -= 20
        
        info_parts = []
        if course.get("level"):
            info_parts.append(f"المستوى: {course['level']}")
        if course.get("section"):
            info_parts.append(f"الشعبة: {course['section']}")
        if dept_name:
            info_parts.append(f"القسم: {dept_name}")
        if college_name:
            info_parts.append(f"الكلية: {college_name}")
        if info_parts:
            draw_arabic(c_pdf, "          ".join(info_parts), width - 30, y, font_size=10)
            y -= 20
        
        # إحصائيات المحاضرات
        y -= 5
        c_pdf.setFillColor(colors.HexColor("#f5f5f5"))
        c_pdf.roundRect(30, y - 20, width - 60, 35, 5, fill=1, stroke=0)
        c_pdf.setFillColor(colors.black)
        draw_arabic(c_pdf, f"المحاضرات: {len(lectures)}     منعقدة: {completed_count}     غائب: {absent_count}     ملغاة: {cancelled_count}     مجدولة: {scheduled_count}     الطلاب: {len(students)}", width - 40, y, font_size=10)
        
        # =================== جدول الطلاب ===================
        y -= 45
        if students:
            # حساب حضور كل طالب
            students_data = []
            for student in students:
                sid = str(student["_id"])
                records = await db.attendance.find({
                    "student_id": sid,
                    "course_id": cid,
                    "lecture_id": {"$in": list(active_lecture_ids)}
                }).to_list(1000)
                
                present = sum(1 for r in records if r["status"] == AttendanceStatus.PRESENT)
                absent_s = sum(1 for r in records if r["status"] == AttendanceStatus.ABSENT)
                excused = sum(1 for r in records if r["status"] == AttendanceStatus.EXCUSED)
                late = sum(1 for r in records if r.get("status") == AttendanceStatus.LATE)
                total_active = len(active_lecture_ids) if active_lecture_ids else 1
                rate = round((present + late * 0.5) / total_active * 100, 1) if total_active > 0 else 0
                
                students_data.append({
                    "name": student.get("full_name", ""),
                    "student_id": student.get("student_id", ""),
                    "present": present,
                    "absent": absent_s,
                    "excused": excused,
                    "late": late,
                    "rate": rate
                })
            
            students_data.sort(key=lambda x: x["rate"], reverse=True)
            
            # Table header
            c_pdf.setFillColor(colors.HexColor("#1565c0"))
            c_pdf.roundRect(30, y - 8, width - 60, 25, 3, fill=1, stroke=0)
            c_pdf.setFillColor(colors.white)
            
            col_rate = 50
            col_late = 95
            col_excused = 140
            col_absent = 185
            col_present = 235
            col_sid = 310
            col_name = width - 40
            col_num = width - 20
            
            c_pdf.setFont(arabic_font, 9)
            for label, x in [("النسبة", col_rate), ("متأخر", col_late), ("معذور", col_excused), ("غائب", col_absent), ("حاضر", col_present)]:
                reshaped = arabic_reshaper.reshape(label)
                c_pdf.drawString(x - 15, y, get_display(reshaped))
            
            reshaped = arabic_reshaper.reshape("الرقم")
            c_pdf.drawString(col_sid - 15, y, get_display(reshaped))
            reshaped = arabic_reshaper.reshape("الاسم")
            c_pdf.drawRightString(col_name, y, get_display(reshaped))
            reshaped = arabic_reshaper.reshape("#")
            c_pdf.drawRightString(col_num, y, get_display(reshaped))
            
            y -= 26
            c_pdf.setFillColor(colors.black)
            
            for i, sd in enumerate(students_data):
                if y < 60:
                    c_pdf.showPage()
                    y = height - 40
                    # إعادة رسم العنوان
                    draw_arabic(c_pdf, f"{course['name']} ({course.get('code', '')}) - تابع", width - 30, y, font_size=12)
                    y -= 30
                
                if i % 2 == 0:
                    c_pdf.setFillColor(colors.HexColor("#fafafa"))
                    c_pdf.rect(30, y - 6, width - 60, 20, fill=1, stroke=0)
                    c_pdf.setFillColor(colors.black)
                
                c_pdf.setFont(arabic_font, 9)
                
                # Row number
                reshaped = arabic_reshaper.reshape(str(i + 1))
                c_pdf.drawRightString(col_num, y, get_display(reshaped))
                
                # Student name
                draw_arabic(c_pdf, sd["name"], col_name, y, font_size=9)
                
                # Student ID
                c_pdf.drawString(col_sid - 15, y, str(sd["student_id"]))
                
                # Counts
                c_pdf.setFont(arabic_font, 9)
                c_pdf.drawString(col_present, y, str(sd["present"]))
                c_pdf.drawString(col_absent, y, str(sd["absent"]))
                c_pdf.drawString(col_excused, y, str(sd["excused"]))
                c_pdf.drawString(col_late, y, str(sd["late"]))
                
                # Rate with color
                rate_color = "#4caf50" if sd["rate"] >= 75 else "#ff9800" if sd["rate"] >= 50 else "#f44336"
                c_pdf.setFillColor(colors.HexColor(rate_color))
                c_pdf.drawString(col_rate - 5, y, f"{sd['rate']}%")
                c_pdf.setFillColor(colors.black)
                
                y -= 20
        else:
            draw_arabic(c_pdf, "لا يوجد طلاب مسجلين في هذا المقرر", width - 30, y, font_size=11)
            y -= 20
        
        # فاصل بين المقررات
        if course_idx < len(courses_list) - 1:
            c_pdf.showPage()
    
    # =================== الصفحة الأخيرة - التوقيع ===================
    y -= 30
    if y < 100:
        c_pdf.showPage()
        y = height - 60
    
    c_pdf.setStrokeColor(colors.HexColor("#e0e0e0"))
    c_pdf.setLineWidth(0.5)
    c_pdf.line(30, y, width - 30, y)
    y -= 20
    draw_arabic(c_pdf, f"تم إنشاء التقرير في: {now.strftime('%Y-%m-%d %H:%M')} (توقيت اليمن)", width - 30, y, font_size=9)
    y -= 30
    c_pdf.setStrokeColor(colors.HexColor("#333"))
    c_pdf.setLineWidth(0.8)
    c_pdf.line(width - 200, y, width - 30, y)
    y -= 14
    draw_arabic(c_pdf, "توقيع المسؤول", width - 30, y, font_size=9)
    
    c_pdf.save()
    buffer.seek(0)
    
    filename = f"تقرير_الفصل_{now.strftime('%Y%m%d')}.pdf"
    from urllib.parse import quote
    encoded_filename = quote(filename)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )

# ==================== Initialize Admin ====================

@api_router.post("/init-admin")
async def init_admin():
    """Create initial admin user if not exists"""
    existing = await db.users.find_one({"role": UserRole.ADMIN})
    if existing:
        return {"message": "المشرف موجود مسبقاً"}
    
    admin_data = {
        "username": "admin",
        "password": get_password_hash("admin123"),
        "full_name": "مدير النظام",
        "role": UserRole.ADMIN,
        "email": "admin@sharia.edu",
        "created_at": get_yemen_time(),
        "is_active": True
    }
    
    await db.users.insert_one(admin_data)
    return {"message": "تم إنشاء حساب المشرف بنجاح", "username": "admin", "password": "admin123"}

@api_router.post("/reset-admin")
async def reset_admin():
    """Reset admin user - creates new or updates existing"""
    try:
        # Delete existing admin
        await db.users.delete_many({"username": "admin"})
        
        # Create fresh admin - use dynamic hash
        admin_password = "admin123"
        hashed = get_password_hash(admin_password)
        
        admin_data = {
            "username": "admin",
            "password": hashed,
            "full_name": "مدير النظام",
            "role": "admin",
            "email": "admin@sharia.edu",
            "created_at": get_yemen_time(),
            "is_active": True,
            "permissions": [
                "manage_users", "manage_departments", "manage_courses",
                "manage_students", "view_attendance",
                "view_reports", "export_reports",
                "import_data", "view_lectures", "view_courses",
                "manage_faculties", "manage_teachers", "manage_enrollments",
                "view_statistics", "manage_roles", "manage_semesters"
            ]
        }
        
        await db.users.insert_one(admin_data)
        return {"message": "تم إعادة إنشاء حساب المشرف بنجاح", "username": "admin", "password": "admin123"}
    except Exception as e:
        logger.error(f"Error resetting admin: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== Schedule Routes ====================

class ScheduleBase(BaseModel):
    course_id: str
    day: str  # saturday, sunday, monday, tuesday, wednesday, thursday, friday
    start_time: str  # HH:MM
    end_time: str
    room: str

class ScheduleCreate(ScheduleBase):
    pass

class ScheduleResponse(ScheduleBase):
    id: str
    created_at: datetime

@api_router.post("/schedule", response_model=ScheduleResponse)
async def create_schedule(schedule: ScheduleCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    schedule_dict = schedule.dict()
    schedule_dict["created_at"] = get_yemen_time()
    
    result = await db.schedule.insert_one(schedule_dict)
    schedule_dict["id"] = str(result.inserted_id)
    
    return schedule_dict

@api_router.get("/schedule", response_model=List[ScheduleResponse])
async def get_schedule(
    day: Optional[str] = None,
    teacher_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    
    if day:
        query["day"] = day
    
    # Filter by teacher's courses if teacher
    if current_user["role"] == UserRole.TEACHER:
        # Get teacher's courses
        courses = await db.courses.find({"teacher_id": current_user["id"]}).to_list(100)
        course_ids = [str(c["_id"]) for c in courses]
        query["course_id"] = {"$in": course_ids}
    elif teacher_id:
        courses = await db.courses.find({"teacher_id": teacher_id}).to_list(100)
        course_ids = [str(c["_id"]) for c in courses]
        query["course_id"] = {"$in": course_ids}
    
    schedules = await db.schedule.find(query).to_list(500)
    return [{
        "id": str(s["_id"]),
        "course_id": s["course_id"],
        "day": s["day"],
        "start_time": s["start_time"],
        "end_time": s["end_time"],
        "room": s["room"],
        "created_at": s["created_at"]
    } for s in schedules]

@api_router.put("/schedule/{schedule_id}", response_model=ScheduleResponse)
async def update_schedule(
    schedule_id: str,
    schedule: ScheduleCreate,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    existing = await db.schedule.find_one({"_id": ObjectId(schedule_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="الموعد غير موجود")
    
    schedule_dict = schedule.dict()
    await db.schedule.update_one(
        {"_id": ObjectId(schedule_id)},
        {"$set": schedule_dict}
    )
    
    return {
        "id": schedule_id,
        **schedule_dict,
        "created_at": existing["created_at"]
    }

@api_router.delete("/schedule/{schedule_id}")
async def delete_schedule(schedule_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    result = await db.schedule.delete_one({"_id": ObjectId(schedule_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="الموعد غير موجود")
    
    return {"message": "تم حذف الموعد بنجاح"}

@api_router.get("/schedule/today")
async def get_today_schedule(current_user: dict = Depends(get_current_user)):
    """Get today's schedule based on current day"""
    days_map = {
        0: 'monday', 1: 'tuesday', 2: 'wednesday', 
        3: 'thursday', 4: 'friday', 5: 'saturday', 6: 'sunday'
    }
    today = days_map[get_yemen_time().weekday()]
    
    query = {"day": today}
    
    if current_user["role"] == UserRole.TEACHER:
        courses = await db.courses.find({"teacher_id": current_user["id"]}).to_list(100)
        course_ids = [str(c["_id"]) for c in courses]
        query["course_id"] = {"$in": course_ids}
    
    schedules = await db.schedule.find(query).sort("start_time", 1).to_list(100)
    
    result = []
    for s in schedules:
        course = await db.courses.find_one({"_id": ObjectId(s["course_id"])})
        result.append({
            "id": str(s["_id"]),
            "course_id": s["course_id"],
            "course_name": course["name"] if course else "غير معروف",
            "course_code": course["code"] if course else "",
            "day": s["day"],
            "start_time": s["start_time"],
            "end_time": s["end_time"],
            "room": s["room"],
            "level": course["level"] if course else 0,
            "section": course["section"] if course else ""
        })
    
    return result

# ==================== Import/Export Routes ====================

@api_router.post("/students/import-preview/{course_id}")
async def preview_import_students(
    course_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Preview students from Excel file before importing"""
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        # تنظيف أسماء الأعمدة من المسافات الزائدة
        df.columns = df.columns.str.strip()
        
        column_mapping = {
            'رقم القيد': 'student_id', 'الرقم الجامعي': 'student_id', 'رقم الطالب': 'student_id',
            'اسم الطالب': 'full_name', 'الاسم': 'full_name', 'الاسم الكامل': 'full_name',
        }
        df = df.rename(columns=column_mapping)
        
        total = len(df)
        names = []
        if 'full_name' in df.columns:
            names = df['full_name'].dropna().head(10).tolist()
        
        return {"total": total, "sample_names": names, "columns": list(df.columns)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"فشل قراءة الملف: {str(e)}")


@api_router.post("/students/import/{course_id}")
async def import_students_to_course(
    course_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Import students from Excel file and enroll them in a course"""
    if not has_permission(current_user, Permission.MANAGE_STUDENTS) and not has_permission(current_user, Permission.IMPORT_DATA):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # Get course info
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        logger.info(f"Import to course: columns={list(df.columns)}, rows={len(df)}")
        
        # تنظيف أسماء الأعمدة من المسافات الزائدة
        df.columns = df.columns.str.strip()
        
        # دعم أسماء الأعمدة بالعربي والإنجليزي
        column_mapping = {
            'رقم القيد': 'student_id',
            'الرقم الجامعي': 'student_id',
            'رقم الطالب': 'student_id',
            'اسم الطالب': 'full_name',
            'الاسم': 'full_name',
            'الاسم الكامل': 'full_name',
            'الهاتف': 'phone',
            'رقم الهاتف': 'phone',
            'الجوال': 'phone',
            'البريد': 'email',
            'البريد الإلكتروني': 'email',
        }
        
        # إعادة تسمية الأعمدة
        df = df.rename(columns=column_mapping)
        logger.info(f"Columns after mapping: {list(df.columns)}")
        
        if 'student_id' not in df.columns:
            raise HTTPException(status_code=400, detail="العمود المطلوب غير موجود: رقم الطالب (student_id)")
        if 'full_name' not in df.columns:
            raise HTTPException(status_code=400, detail="العمود المطلوب غير موجود: اسم الطالب (full_name)")
        
        imported = 0
        enrolled = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                student_id_str = str(row['student_id'])
                
                # Check if student exists
                existing = await db.students.find_one({"student_id": student_id_str})
                
                if not existing:
                    # Create new student
                    student_data = {
                        "student_id": student_id_str,
                        "full_name": str(row['full_name']),
                        "department_id": course.get("department_id", ""),
                        "level": course.get("level", 1),
                        "section": course.get("section", ""),
                        "phone": str(row.get('phone', '')) if pd.notna(row.get('phone')) else None,
                        "email": str(row.get('email', '')) if pd.notna(row.get('email')) else None,
                        "qr_code": generate_qr_code(student_id_str),
                        "created_at": get_yemen_time(),
                        "is_active": True,
                        "user_id": None
                    }
                    result = await db.students.insert_one(student_data)
                    student_db_id = str(result.inserted_id)
                    imported += 1
                else:
                    student_db_id = str(existing["_id"])
                
                # Check if already enrolled
                existing_enrollment = await db.enrollments.find_one({
                    "student_id": student_db_id,
                    "course_id": course_id
                })
                
                if not existing_enrollment:
                    # Enroll student in course
                    await db.enrollments.insert_one({
                        "student_id": student_db_id,
                        "course_id": course_id,
                        "enrolled_at": get_yemen_time(),
                        "is_active": True
                    })
                    enrolled += 1
                    
            except Exception as e:
                errors.append(f"خطأ في الصف {index + 2}: {str(e)}")
        
        return {
            "message": f"تم استيراد {imported} طالب جديد وتسجيل {enrolled} طالب في المقرر",
            "imported": imported,
            "enrolled": enrolled,
            "errors": errors[:10]
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"فشل في قراءة الملف: {str(e)}")

@api_router.post("/import/students")
async def import_students_from_excel(
    file: UploadFile = File(...),
    department_id: Optional[str] = None,
    level: Optional[int] = None,
    section: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Import students from Excel file"""
    logger.info(f"Import students called: department_id={department_id}, level={level}, section={section}, file={file.filename}")
    
    # التحقق من الصلاحيات
    if not has_permission(current_user, Permission.MANAGE_STUDENTS) and not has_permission(current_user, Permission.IMPORT_DATA):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    if not department_id:
        raise HTTPException(status_code=400, detail="يجب تحديد القسم")
    
    try:
        contents = await file.read()
        logger.info(f"File read successfully, size: {len(contents)} bytes")
        
        df = pd.read_excel(BytesIO(contents))
        logger.info(f"Excel parsed, columns: {list(df.columns)}, rows: {len(df)}")
        
        # دعم أسماء الأعمدة بالعربي والإنجليزي
        # تحويل أسماء الأعمدة العربية للإنجليزية
        column_mapping = {
            'رقم القيد': 'student_id',
            'الرقم الجامعي': 'student_id',
            'رقم الطالب': 'student_id',
            'اسم الطالب': 'full_name',
            'الاسم': 'full_name',
            'الاسم الكامل': 'full_name',
            'المستوى': 'level',
            'الشعبة': 'section',
            'القسم': 'section',
            'الهاتف': 'phone',
            'رقم الهاتف': 'phone',
            'الجوال': 'phone',
            'البريد': 'email',
            'البريد الإلكتروني': 'email',
            'الإيميل': 'email',
            # إضافة "رقم الطالب" كبديل لـ "رقم القيد"
            'رقم الطالب': 'student_id',
            # حقول الرقم المرجعي
            'البرنامج': 'program_code',
            'رمز البرنامج': 'program_code',
            'سنة الالتحاق': 'enrollment_year',
            'العام': 'enrollment_year',
        }
        
        # إعادة تسمية الأعمدة
        df = df.rename(columns=column_mapping)
        logger.info(f"Columns after mapping: {list(df.columns)}")
        
        # student_id and full_name are required, level can come from parameter
        if 'student_id' not in df.columns:
            raise HTTPException(status_code=400, detail="العمود المطلوب غير موجود: رقم القيد (student_id)")
        if 'full_name' not in df.columns:
            raise HTTPException(status_code=400, detail="العمود المطلوب غير موجود: اسم الطالب (full_name)")
        
        # If level not in Excel and not provided as parameter, error
        if 'level' not in df.columns and not level:
            raise HTTPException(status_code=400, detail="يجب تحديد المستوى أو تضمينه في ملف Excel")
        
        imported = 0
        imported_ids = []
        errors = []
        
        for index, row in df.iterrows():
            try:
                # ✅ السماح بتكرار رقم القيد فقط لو القسم مختلف
                existing = await db.students.find_one({
                    "student_id": str(row['student_id']),
                    "department_id": department_id,
                })
                if existing:
                    errors.append(f"الطالب {row['student_id']} موجود مسبقاً في نفس القسم")
                    continue
                
                # Use level from parameter first, then from Excel
                student_level = level if level else int(row.get('level', 1))
                
                # Use section from parameter first, then from Excel, or empty (section is optional)
                student_section = ""
                if section:
                    student_section = section
                elif 'section' in df.columns and pd.notna(row.get('section')):
                    student_section = str(row.get('section', ''))
                
                student_data = {
                    "student_id": str(row['student_id']),
                    "full_name": str(row['full_name']),
                    "department_id": department_id,
                    "faculty_id": await _resolve_faculty_id(department_id),
                    "level": student_level,
                    "section": student_section,
                    "phone": str(row.get('phone', '')) if pd.notna(row.get('phone')) else None,
                    "email": str(row.get('email', '')) if pd.notna(row.get('email')) else None,
                    "qr_code": generate_qr_code(str(row['student_id'])),
                    "created_at": get_yemen_time(),
                    "is_active": True,
                    "user_id": None
                }

                # برنامج الطالب: من Excel، أو من القسم، أو يُترك فارغاً
                if 'program_code' in df.columns and pd.notna(row.get('program_code')):
                    pc = str(row.get('program_code')).strip().upper()
                    if pc in ("B", "M", "D", "E", "P"):
                        student_data["program_code"] = pc
                if not student_data.get("program_code"):
                    try:
                        dept = await db.departments.find_one({"_id": ObjectId(department_id)})
                        if dept and dept.get("default_program_code"):
                            student_data["program_code"] = dept["default_program_code"]
                    except Exception:
                        pass

                # سنة الالتحاق: من Excel، أو محسوبة من المستوى
                if 'enrollment_year' in df.columns and pd.notna(row.get('enrollment_year')):
                    ey_raw = str(row.get('enrollment_year')).strip()
                    # 2025 → 25
                    if len(ey_raw) == 4 and ey_raw.isdigit():
                        student_data["enrollment_year"] = ey_raw[-2:]
                    elif len(ey_raw) == 2 and ey_raw.isdigit():
                        student_data["enrollment_year"] = ey_raw
                if not student_data.get("enrollment_year"):
                    try:
                        from routes.admin_tools import compute_enrollment_year_from_level, generate_reference_for_new_student
                        ey = await compute_enrollment_year_from_level(db, student_level)
                        if ey:
                            student_data["enrollment_year"] = ey
                    except Exception:
                        pass

                # توليد الرقم المرجعي
                try:
                    from routes.admin_tools import generate_reference_for_new_student
                    ref = await generate_reference_for_new_student(db, student_data)
                    if ref:
                        student_data["reference_number"] = ref
                except Exception:
                    pass

                result = await db.students.insert_one(student_data)
                imported_ids.append(str(result.inserted_id))
                imported += 1
            except Exception as e:
                errors.append(f"خطأ في الصف {index + 2}: {str(e)}")
        
        # التسجيل التلقائي في المقررات المطابقة (batch)
        enrolled_courses = 0
        enrolled_students_total = 0
        if imported_ids and department_id:
            all_courses = await db.courses.find({
                "department_id": department_id, "is_active": True
            }).to_list(500)
            
            # جلب بيانات الطلاب المستوردين دفعة واحدة
            imported_students = await db.students.find(
                {"_id": {"$in": [ObjectId(sid) for sid in imported_ids]}},
                {"_id": 1, "level": 1, "section": 1}
            ).to_list(10000)
            
            for mc in all_courses:
                cid = str(mc["_id"])
                mc_level = mc.get("level")
                mc_section = mc.get("section", "")
                
                matching_sids = []
                for st in imported_students:
                    if mc_level and st.get("level") != mc_level:
                        continue
                    if mc_section and st.get("section", "") != mc_section:
                        continue
                    matching_sids.append(str(st["_id"]))
                
                if not matching_sids:
                    continue
                
                existing = await db.enrollments.find(
                    {"course_id": cid, "student_id": {"$in": matching_sids}},
                    {"student_id": 1}
                ).to_list(10000)
                already_set = {e["student_id"] for e in existing}
                new_sids = [s for s in matching_sids if s not in already_set]
                
                if new_sids:
                    now = get_yemen_time()
                    await db.enrollments.insert_many([{
                        "course_id": cid, "student_id": s,
                        "enrolled_at": now, "enrolled_by": current_user["id"]
                    } for s in new_sids])
                    enrolled_students_total += len(new_sids)
        
        return {
            "message": f"تم استيراد {imported} طالب بنجاح",
            "imported": imported,
            "imported_ids": imported_ids,
            "enrolled_courses": enrolled_courses,
            "enrolled_courses_msg": f"تم تسجيلهم تلقائياً في {enrolled_courses} مقرر" if enrolled_courses > 0 else "",
            "errors": errors[:10]  # Return first 10 errors only
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Import error: {str(e)}")
        raise HTTPException(status_code=400, detail=f"خطأ في قراءة الملف: {str(e)}")


@api_router.get("/template/teachers")
async def get_teachers_template(current_user: dict = Depends(get_current_user)):
    """Download teachers import template"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_teachers") and not has_permission(current_user, "import_data"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    data = {
        "الرقم الوظيفي": ["T001", "T002", "T003"],
        "اسم المعلم": ["د. محمد أحمد", "د. علي سالم", "أ. خالد عمر"],
        "النصاب الأسبوعي": [12, 14, 10],
        "التخصص": ["فقه", "شريعة", "لغة عربية"],
        "الهاتف": ["777123456", "777654321", ""],
        "البريد الإلكتروني": ["m.ahmed@univ.edu", "", ""],
    }
    
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=teachers_template.xlsx"}
    )

# ==================== استيراد المحاضرات من Excel ====================

@api_router.get("/template/lectures")
async def get_lectures_template(current_user: dict = Depends(get_current_user)):
    """تحميل نموذج استيراد المحاضرات"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_lectures") and not has_permission(current_user, "import_data"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    data = {
        "رمز المقرر": ["FQH101", "FQH101", "ARB102"],
        "الشعبة": ["A", "A", "B"],
        "اليوم": ["السبت", "الإثنين", "الأحد"],
        "وقت البداية": ["08:00", "10:00", "08:00"],
        "وقت النهاية": ["09:30", "11:30", "09:30"],
        "القاعة": ["قاعة 101", "قاعة 203", "قاعة 105"],
        "تاريخ البداية": ["2026-02-01", "2026-02-01", "2026-02-01"],
        "تاريخ النهاية": ["2026-05-30", "2026-05-30", "2026-05-30"],
    }
    
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=lectures_template.xlsx"}
    )


@api_router.post("/import/lectures")
async def import_lectures_from_excel(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """استيراد محاضرات من ملف Excel - يولّد محاضرات أسبوعية تلقائياً"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_lectures") and not has_permission(current_user, "import_data"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        column_mapping = {
            'رمز المقرر': 'course_code',
            'الرمز': 'course_code',
            'الشعبة': 'section',
            'شعبة': 'section',
            'Section': 'section',
            'اليوم': 'day',
            'وقت البداية': 'start_time',
            'البداية': 'start_time',
            'وقت النهاية': 'end_time',
            'النهاية': 'end_time',
            'القاعة': 'room',
            'تاريخ البداية': 'start_date',
            'بداية الفصل': 'start_date',
            'تاريخ النهاية': 'end_date',
            'نهاية الفصل': 'end_date',
        }
        
        df = df.rename(columns=column_mapping)
        
        required_cols = ['course_code', 'day', 'start_time', 'end_time', 'start_date', 'end_date']
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            col_names = {'course_code': 'رمز المقرر', 'day': 'اليوم', 'start_time': 'وقت البداية', 'end_time': 'وقت النهاية', 'start_date': 'تاريخ البداية', 'end_date': 'تاريخ النهاية'}
            missing_ar = [col_names.get(c, c) for c in missing]
            raise HTTPException(status_code=400, detail=f"أعمدة مطلوبة غير موجودة: {', '.join(missing_ar)}")
        
        # تحويل أسماء الأيام العربية إلى أرقام Python weekday
        day_name_map = {
            'السبت': 5, 'سبت': 5,
            'الأحد': 6, 'أحد': 6, 'الاحد': 6, 'احد': 6,
            'الإثنين': 0, 'الاثنين': 0, 'إثنين': 0, 'اثنين': 0,
            'الثلاثاء': 1, 'ثلاثاء': 1,
            'الأربعاء': 2, 'الاربعاء': 2, 'أربعاء': 2, 'اربعاء': 2,
            'الخميس': 3, 'خميس': 3,
            'الجمعة': 4, 'جمعة': 4,
            'saturday': 5, 'sunday': 6, 'monday': 0,
            'tuesday': 1, 'wednesday': 2, 'thursday': 3, 'friday': 4,
        }
        
        total_lectures = 0
        total_conflicts = 0
        errors = []
        courses_processed = set()
        
        from datetime import timedelta
        
        for index, row in df.iterrows():
            row_num = index + 2  # +2 لأن Excel يبدأ من 1 + header
            try:
                course_code = str(row.get('course_code', '')).strip()
                section = str(row.get('section', '')).strip() if pd.notna(row.get('section')) else ''
                day_name = str(row.get('day', '')).strip()
                start_time = str(row.get('start_time', '')).strip()
                end_time = str(row.get('end_time', '')).strip()
                room = str(row.get('room', '')).strip() if pd.notna(row.get('room')) else ''
                start_date_str = str(row.get('start_date', '')).strip()
                end_date_str = str(row.get('end_date', '')).strip()
                
                if not course_code:
                    errors.append(f"سطر {row_num}: رمز المقرر فارغ")
                    continue
                
                # البحث عن المقرر مع مراعاة الشعبة
                course_query = {"code": course_code}
                if section:
                    course_query["section"] = section
                course = await db.courses.find_one(course_query)
                if not course and section:
                    errors.append(f"سطر {row_num}: رمز المقرر '{course_code}' شعبة '{section}' غير موجود في النظام")
                    continue
                elif not course:
                    errors.append(f"سطر {row_num}: رمز المقرر '{course_code}' غير موجود في النظام")
                    continue
                
                course_id = str(course["_id"])
                courses_processed.add(course_code)
                
                # التحقق من اليوم
                target_weekday = day_name_map.get(day_name)
                if target_weekday is None:
                    errors.append(f"سطر {row_num}: اسم اليوم '{day_name}' غير معروف")
                    continue
                
                # تنظيف الأوقات (تحويل من HH:MM:SS إلى HH:MM)
                if len(start_time) > 5:
                    start_time = start_time[:5]
                if len(end_time) > 5:
                    end_time = end_time[:5]
                
                if not start_time or not end_time:
                    errors.append(f"سطر {row_num}: وقت البداية أو النهاية فارغ")
                    continue
                
                if end_time <= start_time:
                    errors.append(f"سطر {row_num}: وقت النهاية ({end_time}) يجب أن يكون بعد وقت البداية ({start_time})")
                    continue
                
                # تحليل التواريخ
                try:
                    # دعم تنسيقات متعددة
                    start_date_str = start_date_str.split(' ')[0]  # إزالة الوقت إن وجد
                    end_date_str = end_date_str.split(' ')[0]
                    start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
                    end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
                except:
                    errors.append(f"سطر {row_num}: صيغة التاريخ غير صحيحة (المتوقع: YYYY-MM-DD)")
                    continue
                
                if end_date <= start_date:
                    errors.append(f"سطر {row_num}: تاريخ النهاية يجب أن يكون بعد تاريخ البداية")
                    continue
                
                # البحث عن أول يوم مطابق
                current = start_date
                while current.weekday() != target_weekday:
                    current += timedelta(days=1)
                
                # توليد محاضرة لكل أسبوع
                row_lectures = 0
                row_conflicts = 0
                while current <= end_date:
                    date_str = current.strftime("%Y-%m-%d")
                    
                    # فحص التعارض
                    conflict = await check_teacher_lecture_conflict(course_id, date_str, start_time, end_time, allow_same_course=True)
                    if conflict and conflict["type"] == "error":
                        row_conflicts += 1
                        total_conflicts += 1
                        current += timedelta(days=7)
                        continue
                    room_conflict = await check_room_lecture_conflict(course_id, room or "", date_str, start_time, end_time)
                    if room_conflict and room_conflict["type"] == "error":
                        row_conflicts += 1
                        total_conflicts += 1
                        current += timedelta(days=7)
                        continue
                    
                    lecture = {
                        "course_id": course_id,
                        "date": date_str,
                        "start_time": start_time,
                        "end_time": end_time,
                        "room": room,
                        "status": LectureStatus.SCHEDULED,
                        "notes": f"تم الاستيراد من Excel",
                        "created_at": get_yemen_time(),
                        "created_by": current_user["id"]
                    }
                    await db.lectures.insert_one(lecture)
                    row_lectures += 1
                    total_lectures += 1
                    
                    current += timedelta(days=7)
                
                if row_conflicts > 0:
                    errors.append(f"سطر {row_num} ({course_code} - {day_name}): تم تخطي {row_conflicts} محاضرة بسبب تعارض")
                
                # إرسال إشعار للمدرس بعد استيراد محاضرات مقرره
                if row_lectures > 0 and course:
                    await notify_lecture_created(course, "", "", "")
                    
            except Exception as e:
                errors.append(f"سطر {row_num}: {str(e)}")
        
        return {
            "message": f"تم إنشاء {total_lectures} محاضرة لـ {len(courses_processed)} مقرر",
            "imported": total_lectures,
            "courses_count": len(courses_processed),
            "conflicts_skipped": total_conflicts,
            "errors": errors[:20],
            "total_errors": len(errors)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Lectures import error: {str(e)}")
        raise HTTPException(status_code=400, detail=f"خطأ في قراءة الملف: {str(e)}")

@api_router.get("/template/courses")
async def get_courses_template(current_user: dict = Depends(get_current_user)):
    """Download courses import template"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses") and not has_permission(current_user, "import_data"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    data = {
        "اسم المقرر": ["الفقه الإسلامي", "أصول التفسير", "النحو والصرف"],
        "رمز المقرر": ["FQH101", "TFS201", "ARB102"],
        "عدد الساعات": [3, 2, 3],
        "المستوى": [1, 2, 1],
        "عدد الشعب": [1, 2, 1],
        "الوصف": ["مقدمة في الفقه", "أصول التفسير القرآني", ""],
    }
    
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=courses_template.xlsx"}
    )

@api_router.post("/import/courses")
async def import_courses_from_excel(
    file: UploadFile = File(...),
    department_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """استيراد مقررات من ملف Excel"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses") and not has_permission(current_user, "import_data"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    if not department_id:
        raise HTTPException(status_code=400, detail="يجب تحديد القسم")
    
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        column_mapping = {
            'اسم المقرر': 'name',
            'المقرر': 'name',
            'رمز المقرر': 'code',
            'الرمز': 'code',
            'عدد الساعات': 'credit_hours',
            'الساعات': 'credit_hours',
            'ساعات': 'credit_hours',
            'المستوى': 'level',
            'عدد الشعب': 'sections_count',
            'الشعب': 'sections_count',
            'الوصف': 'description',
            'وصف المقرر': 'description',
        }
        
        df = df.rename(columns=column_mapping)
        
        if 'name' not in df.columns:
            raise HTTPException(status_code=400, detail="العمود المطلوب غير موجود: اسم المقرر")
        if 'code' not in df.columns:
            raise HTTPException(status_code=400, detail="العمود المطلوب غير موجود: رمز المقرر")
        
        # الحصول على faculty_id من القسم
        dept = await db.departments.find_one({"_id": ObjectId(department_id)})
        faculty_id = dept.get("faculty_id") if dept else None
        
        # الحصول على الفصل الدراسي النشط
        active_semester = await db.semesters.find_one({"is_active": True})
        semester_id = str(active_semester["_id"]) if active_semester else None
        
        imported = 0
        errors = []
        section_letters = ["أ", "ب", "ج", "د", "هـ", "و", "ز", "ح", "ط", "ي"]
        
        for index, row in df.iterrows():
            try:
                name = str(row['name']).strip()
                code = str(row['code']).strip()
                
                if not name or not code or name == 'nan' or code == 'nan':
                    errors.append(f"السطر {index + 2}: بيانات فارغة")
                    continue
                
                # التحقق من عدم التكرار
                existing = await db.courses.find_one({"code": code, "department_id": department_id})
                if existing:
                    errors.append(f"المقرر {code} موجود مسبقاً")
                    continue
                
                credit_hours = 3
                try:
                    credit_hours = int(row.get('credit_hours', 3))
                except (ValueError, TypeError):
                    credit_hours = 3
                
                level = 1
                try:
                    level = int(row.get('level', 1))
                except (ValueError, TypeError):
                    level = 1
                
                sections_count = 1
                try:
                    sections_count = int(row.get('sections_count', 1))
                except (ValueError, TypeError):
                    sections_count = 1
                sections_count = min(max(sections_count, 1), 10)
                
                description = ""
                if 'description' in df.columns and pd.notna(row.get('description')):
                    description = str(row['description']).strip()
                
                # إنشاء الشعب
                for si in range(sections_count):
                    section_name = section_letters[si] if sections_count > 1 else ""
                    course_name = f"{name} ({section_name})" if section_name else name
                    course_code = f"{code}-{section_name}" if section_name else code
                    
                    course_doc = {
                        "name": course_name,
                        "code": course_code,
                        "section": section_name,
                        "credit_hours": credit_hours,
                        "level": level,
                        "description": description,
                        "department_id": department_id,
                        "faculty_id": faculty_id,
                        "teacher_id": None,
                        "semester_id": semester_id,
                        "syllabus_items": [],
                        "is_active": True,
                        "is_deleted": False,
                        "created_at": get_yemen_time(),
                    }
                    
                    await db.courses.insert_one(course_doc)
                    imported += 1
                
            except Exception as e:
                errors.append(f"السطر {index + 2}: {str(e)}")
        
        await log_activity(
            user=current_user,
            action="import_courses",
            entity_type="course",
            entity_id="bulk",
            entity_name=f"استيراد {imported} مقرر"
        )
        
        return {
            "message": f"تم استيراد {imported} مقرر بنجاح",
            "imported": imported,
            "errors": errors[:20],
            "total_errors": len(errors)
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Course import error: {e}")
        raise HTTPException(status_code=500, detail=f"خطأ في معالجة الملف: {str(e)}")

@api_router.post("/import/teachers")
async def import_teachers_from_excel(
    file: UploadFile = File(...),
    department_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Import teachers from Excel file and auto-activate their accounts"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_teachers") and not has_permission(current_user, "import_data"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    if not department_id:
        raise HTTPException(status_code=400, detail="يجب تحديد القسم")
    
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        column_mapping = {
            'الرقم الوظيفي': 'teacher_id',
            'رقم الموظف': 'teacher_id',
            'رقم المعلم': 'teacher_id',
            'اسم المعلم': 'full_name',
            'الاسم': 'full_name',
            'الاسم الكامل': 'full_name',
            'النصاب الأسبوعي': 'weekly_hours',
            'النصاب': 'weekly_hours',
            'الساعات': 'weekly_hours',
            'التخصص': 'specialization',
            'الهاتف': 'phone',
            'رقم الهاتف': 'phone',
            'البريد': 'email',
            'البريد الإلكتروني': 'email',
        }
        
        df = df.rename(columns=column_mapping)
        
        if 'teacher_id' not in df.columns:
            raise HTTPException(status_code=400, detail="العمود المطلوب غير موجود: الرقم الوظيفي")
        if 'full_name' not in df.columns:
            raise HTTPException(status_code=400, detail="العمود المطلوب غير موجود: اسم المعلم")
        if 'weekly_hours' not in df.columns:
            raise HTTPException(status_code=400, detail="العمود المطلوب غير موجود: النصاب الأسبوعي")
        
        imported = 0
        activated = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                tid = str(row['teacher_id']).strip()
                existing = await db.teachers.find_one({"teacher_id": tid})
                if existing:
                    errors.append(f"المعلم {tid} موجود مسبقاً")
                    continue
                
                weekly_h = 12
                try:
                    weekly_h = int(row['weekly_hours'])
                except (ValueError, TypeError):
                    weekly_h = 12
                
                teacher_data = {
                    "teacher_id": tid,
                    "full_name": str(row['full_name']).strip(),
                    "department_id": department_id,
                    "faculty_id": await _resolve_faculty_id(department_id),
                    "weekly_hours": weekly_h,
                    "specialization": str(row.get('specialization', '')).strip() if pd.notna(row.get('specialization')) else None,
                    "phone": str(row.get('phone', '')).strip() if pd.notna(row.get('phone')) else None,
                    "email": str(row.get('email', '')).strip() if pd.notna(row.get('email')) else None,
                    "academic_title": None,
                    "teaching_load": None,
                    "created_at": get_yemen_time(),
                    "is_active": True,
                    "user_id": None,
                }
                
                result = await db.teachers.insert_one(teacher_data)
                teacher_obj_id = result.inserted_id
                imported += 1
                
                # Auto-activate account: username = teacher_id, password = teacher_id
                existing_user = await db.users.find_one({"username": tid})
                if not existing_user:
                    user_dict = {
                        "username": tid,
                        "password": get_password_hash(tid),
                        "full_name": str(row['full_name']).strip(),
                        "role": UserRole.TEACHER,
                        "email": teacher_data.get("email"),
                        "phone": teacher_data.get("phone"),
                        "teacher_record_id": str(teacher_obj_id),
                        "must_change_password": True,
                        "is_active": True,
                        "created_at": get_yemen_time(),
                    }
                    user_result = await db.users.insert_one(user_dict)
                    await db.teachers.update_one(
                        {"_id": teacher_obj_id},
                        {"$set": {"user_id": str(user_result.inserted_id)}}
                    )
                    activated += 1
                    
            except Exception as e:
                errors.append(f"خطأ في الصف {index + 2}: {str(e)}")
        
        return {
            "message": f"تم استيراد {imported} معلم وتفعيل {activated} حساب",
            "imported": imported,
            "activated": activated,
            "errors": errors[:10]
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Teacher import error: {str(e)}")
        raise HTTPException(status_code=400, detail=f"خطأ في قراءة الملف: {str(e)}")


@api_router.get("/export/students")
async def export_students_to_excel(
    department_id: Optional[str] = None,
    level: Optional[int] = None,
    section: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export students to Excel file - يدعم التصفية بالقسم/المستوى/الشعبة"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students") and not has_permission(current_user, "export_reports"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    query: dict = {"is_active": True}
    if department_id:
        query["department_id"] = department_id
    if level:
        query["level"] = level
    if section:
        query["section"] = section
    
    students = await db.students.find(query).sort([("level", 1), ("section", 1), ("full_name", 1)]).to_list(10000)
    departments = await db.departments.find().to_list(200)
    dept_map = {str(d["_id"]): d for d in departments}
    faculties = await db.faculties.find().to_list(200)
    fac_map = {str(f["_id"]): f.get("name", "") for f in faculties}
    
    data = []
    for s in students:
        dept = dept_map.get(s.get("department_id", ""), {})
        dept_name = dept.get("name", "غير محدد")
        # الأولوية لـ faculty_id الموجود في سجل الطالب (لأن الرقم المرجعي مولّد بناءً عليه)
        # في حال عدم وجوده، نستخدم faculty_id الخاص بالقسم
        student_fac_id = s.get("faculty_id") or (dept.get("faculty_id") if dept else None)
        fac_name = fac_map.get(student_fac_id or "", "")
        # الكلية الأكاديمية للقسم (قد تختلف عن كلية قيد الطالب)
        dept_fac_id = dept.get("faculty_id") if dept else None
        dept_fac_name = fac_map.get(dept_fac_id or "", "") if dept_fac_id else ""
        data.append({
            "الرقم المرجعي": s.get("reference_number") or "",
            "رقم الطالب": s.get("student_id", ""),
            "الاسم الكامل": s.get("full_name", ""),
            "الكلية": fac_name,
            "كلية القسم": dept_fac_name if dept_fac_name and dept_fac_name != fac_name else "",
            "القسم": dept_name,
            "المستوى": s.get("level", ""),
            "الشعبة": s.get("section", ""),
            "البرنامج": s.get("program_code", ""),
            "سنة الالتحاق": s.get("enrollment_year", ""),
            "الهاتف": s.get("phone", ""),
            "البريد": s.get("email", ""),
            "نشط": "نعم" if s.get("is_active", True) else "لا",
        })
    
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    
    # اسم ملف ديناميكي حسب الفلاتر
    filename_parts = ["students"]
    if department_id and department_id in dept_map:
        filename_parts.append(dept_map[department_id].get("code", "dept"))
    if level:
        filename_parts.append(f"L{level}")
    if section:
        filename_parts.append(section)
    filename = "_".join(filename_parts) + ".xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/export/attendance/{course_id}")
async def export_course_attendance(
    course_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export course attendance to Excel"""
    # Get course details
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    # Get attendance records
    query = {"course_id": course_id}
    if start_date:
        query["date"] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = datetime.fromisoformat(end_date)
        else:
            query["date"] = {"$lte": datetime.fromisoformat(end_date)}
    
    records = await db.attendance.find(query).sort("date", 1).to_list(50000)
    
    # Get students
    students = await db.students.find({
        "department_id": course["department_id"],
        "level": course["level"],
        "section": course["section"]
    }).to_list(500)
    student_map = {str(s["_id"]): s for s in students}
    
    # Group by date
    dates = sorted(set(r["date"].strftime("%Y-%m-%d") for r in records))
    
    data = []
    for student in students:
        row = {
            "رقم الطالب": student["student_id"],
            "اسم الطالب": student["full_name"]
        }
        
        student_records = [r for r in records if r["student_id"] == str(student["_id"])]
        
        total_present = 0
        total_absent = 0
        total_late = 0
        
        for date in dates:
            record = next((r for r in student_records if r["date"].strftime("%Y-%m-%d") == date), None)
            if record:
                status = record["status"]
                if status == "present":
                    row[date] = "حاضر"
                    total_present += 1
                elif status == "absent":
                    row[date] = "غائب"
                    total_absent += 1
                elif status == "late":
                    row[date] = "متأخر"
                    total_late += 1
                elif status == "excused":
                    row[date] = "معذور"
            else:
                row[date] = "-"
        
        total = total_present + total_absent + total_late
        rate = (total_present + total_late * 0.5) / total * 100 if total > 0 else 0
        
        row["إجمالي الحضور"] = total_present
        row["إجمالي الغياب"] = total_absent
        row["إجمالي التأخير"] = total_late
        row["نسبة الحضور %"] = round(rate, 2)
        
        data.append(row)
    
    df = pd.DataFrame(data)
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='سجل الحضور', index=False)
        
        # Add summary sheet
        summary_data = {
            "المقرر": [course["name"]],
            "الرمز": [course["code"]],
            "المستوى": [course["level"]],
            "الشعبة": [course["section"]],
            "إجمالي الطلاب": [len(students)],
            "عدد المحاضرات": [len(dates)]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='ملخص', index=False)
    
    output.seek(0)
    
    filename = f"attendance_{course['code']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/export/report/{dept_id}")
async def export_department_report(
    dept_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Export comprehensive department report"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "view_reports") and not has_permission(current_user, "export_reports"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    dept = await db.departments.find_one({"_id": ObjectId(dept_id)})
    if not dept:
        raise HTTPException(status_code=404, detail="القسم غير موجود")
    
    # Get all courses in department
    courses = await db.courses.find({"department_id": dept_id}).to_list(100)
    
    # Get all students in department
    students = await db.students.find({"department_id": dept_id}).to_list(10000)
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Students sheet
        students_data = []
        for s in students:
            # Get attendance stats
            records = await db.attendance.find({"student_id": str(s["_id"])}).to_list(10000)
            total = len(records)
            present = sum(1 for r in records if r["status"] == "present")
            absent = sum(1 for r in records if r["status"] == "absent")
            rate = (present / total * 100) if total > 0 else 0
            
            students_data.append({
                "رقم الطالب": s["student_id"],
                "الاسم": s["full_name"],
                "المستوى": s["level"],
                "الشعبة": s["section"],
                "إجمالي المحاضرات": total,
                "الحضور": present,
                "الغياب": absent,
                "نسبة الحضور %": round(rate, 2)
            })
        
        if students_data:
            pd.DataFrame(students_data).to_excel(writer, sheet_name='الطلاب', index=False)
        
        # Courses sheet
        courses_data = []
        for c in courses:
            records = await db.attendance.find({"course_id": str(c["_id"])}).to_list(10000)
            sessions = len(set(r["date"].strftime("%Y-%m-%d") for r in records))
            
            courses_data.append({
                "رمز المقرر": c["code"],
                "اسم المقرر": c["name"],
                "المستوى": c["level"],
                "الشعبة": c["section"],
                "عدد المحاضرات": sessions
            })
        
        if courses_data:
            pd.DataFrame(courses_data).to_excel(writer, sheet_name='المقررات', index=False)
        
        # Summary sheet
        summary = {
            "القسم": [dept["name"]],
            "إجمالي الطلاب": [len(students)],
            "إجمالي المقررات": [len(courses)],
            "تاريخ التقرير": [datetime.now().strftime("%Y-%m-%d %H:%M")]
        }
        pd.DataFrame(summary).to_excel(writer, sheet_name='ملخص', index=False)
    
    output.seek(0)
    
    filename = f"report_{dept['code']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/template/students")
async def get_students_template(current_user: dict = Depends(get_current_user)):
    """Download students import template - يحتوي فقط على الحقول المطلوبة"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, Permission.IMPORT_DATA):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # الحقول: رقم القيد، اسم الطالب، البرنامج (اختياري)، سنة الالتحاق (اختياري)
    data = {
        "رقم القيد": ["12345", "12346", "12347"],
        "اسم الطالب": ["محمد أحمد علي", "أحمد محمد سعيد", "خالد سالم"],
        "البرنامج": ["B", "B", "M"],
        "سنة الالتحاق": ["25", "25", "24"],
    }
    
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students_template.xlsx"}
    )

# ==================== PDF Export Routes ====================

# Register Arabic font
import os
from pathlib import Path as _Path

# مسار ديناميكي — يعمل محلياً وعلى الإنتاج (Railway/Docker)
_CANDIDATE_FONT_PATHS = [
    str(_Path(__file__).parent / "fonts" / "Amiri-Regular.ttf"),
    "/app/backend/backend/fonts/Amiri-Regular.ttf",
    "/app/backend/fonts/Amiri-Regular.ttf",
]
FONT_PATH = next((p for p in _CANDIDATE_FONT_PATHS if os.path.exists(p)), None)

if FONT_PATH:
    try:
        pdfmetrics.registerFont(TTFont('Arabic', FONT_PATH))
        ARABIC_FONT = 'Arabic'
        logger.info(f"PDF Arabic font loaded from: {FONT_PATH}")
    except Exception as e:
        logger.error(f"Failed to register Arabic font: {e}")
        ARABIC_FONT = 'Helvetica'
else:
    # محاولة أخيرة: FreeSans (غالباً لا يدعم العربية - للسلامة فقط)
    try:
        pdfmetrics.registerFont(TTFont('Arabic', '/usr/share/fonts/truetype/freefont/FreeSans.ttf'))
        ARABIC_FONT = 'Arabic'
        logger.warning("Using FreeSans fallback — Arabic may not render correctly!")
    except Exception:
        ARABIC_FONT = 'Helvetica'
        logger.error("No Arabic font found — PDFs will be garbled!")

def arabic_text(text):
    """Reshape Arabic text for PDF display"""
    if not text:
        return ""
    try:
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)
    except:
        return str(text)


def get_pro_table_style(extra: list = None, num_cols: list = None, name_cols: list = None):
    """
    أسلوب موحّد محترف للجداول في PDFs
    - num_cols: قائمة بالأعمدة الرقمية (محاذاة وسط)
    - name_cols: قائمة بأعمدة الأسماء العربية (محاذاة يمين)
    """
    style = [
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565c0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 9),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        # Body
        ('FONTNAME', (0, 0), (-1, -1), ARABIC_FONT),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        # Borders (soft gray)
        ('LINEBELOW', (0, 0), (-1, 0), 1.2, colors.HexColor('#0d47a1')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cfd8dc')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#90a4ae')),
        # Alternating rows
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f9ff')]),
        # Padding
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]
    # Right-align name columns (Arabic)
    if name_cols:
        for col in name_cols:
            style.append(('ALIGN', (col, 1), (col, -1), 'RIGHT'))
    if extra:
        style.extend(extra)
    return TableStyle(style)


def add_pdf_header(elements, title: str, subtitle: str = None, date_range: str = None):
    """رأس PDF موحَّد: عنوان + معلومات الجامعة + الفترة الزمنية"""
    title_style = ParagraphStyle('Title', fontName=ARABIC_FONT, fontSize=17, alignment=TA_CENTER, spaceAfter=4, textColor=colors.HexColor('#1565c0'))
    elements.append(Paragraph(arabic_text("جامعة الأحقاف"), ParagraphStyle('Univ', fontName=ARABIC_FONT, fontSize=13, alignment=TA_CENTER, spaceAfter=2, textColor=colors.HexColor('#666666'))))
    elements.append(Paragraph(arabic_text(title), title_style))
    if subtitle:
        elements.append(Paragraph(arabic_text(subtitle), ParagraphStyle('Sub', fontName=ARABIC_FONT, fontSize=11, alignment=TA_CENTER, textColor=colors.HexColor('#555'))))
    if date_range:
        elements.append(Paragraph(arabic_text(date_range), ParagraphStyle('DR', fontName=ARABIC_FONT, fontSize=10, alignment=TA_CENTER, textColor=colors.HexColor('#1565c0'))))
    elements.append(Paragraph(arabic_text(f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"),
                              ParagraphStyle('Date', fontName=ARABIC_FONT, fontSize=9, alignment=TA_CENTER, textColor=colors.HexColor('#888'))))
    elements.append(Spacer(1, 14))

@api_router.get("/export/students/pdf")
async def export_students_pdf(
    department_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export students list to PDF"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_students") and not has_permission(current_user, "export_reports"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    query = {"is_active": True}
    if department_id:
        query["department_id"] = department_id
    
    students = await db.students.find(query).to_list(10000)
    departments = await db.departments.find().to_list(100)
    dept_map = {str(d["_id"]): d["name"] for d in departments}
    
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []

    add_pdf_header(elements, "كشف الطلاب", subtitle=(arabic_text(dept_map.get(department_id, "")) if department_id else None))

    # Table header (RTL - from right to left)
    headers = [
        arabic_text("الشعبة"),
        arabic_text("المستوى"),
        arabic_text("القسم"),
        arabic_text("الاسم"),
        arabic_text("رقم الطالب"),
        arabic_text("م"),
    ]
    
    # Table data
    data = [headers]
    for i, s in enumerate(students, 1):
        row = [
            str(i),
            s["student_id"],
            arabic_text(s["full_name"]),
            arabic_text(dept_map.get(s["department_id"], "غير محدد")),
            str(s["level"]),
            s.get("section", "") or "",
        ]
        data.append(row)
    
    # Create table - col widths in RTL order: [م, رقم, الاسم, القسم, مستوى, شعبة]
    # Note: data is built reversed in RTL fashion - last col in array is leftmost
    table = Table(data, colWidths=[40, 70, 110, 95, 50, 50], repeatRows=1)
    # Name col index in array = 2 (الاسم), dept col = 3 (القسم)
    table.setStyle(get_pro_table_style(name_cols=[2, 3]))
    
    elements.append(table)
    
    # Summary
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(arabic_text(f"إجمالي الطلاب: {len(students)}"), 
                             ParagraphStyle('Summary', fontName=ARABIC_FONT, fontSize=12, alignment=TA_RIGHT)))
    
    doc.build(elements)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=students_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )

@api_router.get("/export/attendance/{course_id}/pdf")
async def export_attendance_pdf(
    course_id: str,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export course attendance to PDF (optional date range: YYYY-MM-DD)"""
    course = await db.courses.find_one({"_id": ObjectId(course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    
    # Get attendance records with optional date filter
    att_query = {"course_id": course_id}
    if date_from or date_to:
        date_filter = {}
        if date_from:
            try:
                date_filter["$gte"] = datetime.strptime(date_from, "%Y-%m-%d")
            except ValueError:
                pass
        if date_to:
            try:
                # include full day
                dt_to = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                date_filter["$lte"] = dt_to
            except ValueError:
                pass
        if date_filter:
            att_query["date"] = date_filter

    records = await db.attendance.find(att_query).sort("date", 1).to_list(50000)
    
    # Get enrolled students
    enrollments = await db.enrollments.find({"course_id": course_id}).to_list(10000)
    student_ids = [ObjectId(e["student_id"]) for e in enrollments]
    students = await db.students.find({"_id": {"$in": student_ids}}).to_list(500)
    
    # Group by date
    dates = sorted(set(r["date"].strftime("%Y-%m-%d") for r in records)) if records else []
    
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
    
    elements = []
    
    # Title
    title_style = ParagraphStyle('Title', fontName=ARABIC_FONT, fontSize=16, alignment=TA_CENTER, spaceAfter=10)
    elements.append(Paragraph(arabic_text("سجل الحضور"), title_style))
    elements.append(Paragraph(arabic_text(f"المقرر: {course['name']} ({course['code']})"), 
                             ParagraphStyle('Subtitle', fontName=ARABIC_FONT, fontSize=12, alignment=TA_CENTER)))
    section_text = course.get('section', '') or ''
    elements.append(Paragraph(arabic_text(f"المستوى {course['level']} - الشعبة {section_text}"), 
                             ParagraphStyle('Info', fontName=ARABIC_FONT, fontSize=10, alignment=TA_CENTER)))

    # Date range subtitle (if provided)
    if date_from or date_to:
        df = date_from or "البداية"
        dt = date_to or "النهاية"
        elements.append(Paragraph(
            arabic_text(f"الفترة: من {df} إلى {dt}"),
            ParagraphStyle('DateRange', fontName=ARABIC_FONT, fontSize=10, alignment=TA_CENTER, textColor=colors.HexColor('#1565c0'))
        ))
    elements.append(Spacer(1, 15))
    
    # Table header (RTL - from right to left)
    # Show ALL dates in range (if range is set), else last 10
    display_dates = dates if (date_from or date_to) else (dates[-10:] if len(dates) > 10 else dates)

    # Legend - شرح الرموز
    legend_style = ParagraphStyle('Legend', fontName=ARABIC_FONT, fontSize=10, alignment=TA_RIGHT, spaceAfter=6)
    elements.append(Paragraph(
        arabic_text("الرموز:  ح = حاضر    غ = غائب    ت = متأخر    ع = معذور    — = لا توجد محاضرة"),
        legend_style
    ))
    elements.append(Spacer(1, 8))
    
    # Build headers RTL
    headers = [arabic_text("%"), arabic_text("غياب"), arabic_text("حضور")]
    for d in reversed(display_dates):
        headers.append(d[5:])  # Show MM-DD only
    headers.extend([arabic_text("الاسم"), arabic_text("رقم الطالب"), arabic_text("م")])
    
    data = [headers]
    
    for i, student in enumerate(students, 1):
        student_records = [r for r in records if r["student_id"] == str(student["_id"])]
        
        total_present = 0
        total_absent = 0
        
        # Build date columns in reverse order for RTL
        date_cols = []
        for date in reversed(display_dates):
            record = next((r for r in student_records if r["date"].strftime("%Y-%m-%d") == date), None)
            if record:
                if record["status"] == "present":
                    date_cols.append(arabic_text("ح"))
                    total_present += 1
                elif record["status"] == "absent":
                    date_cols.append(arabic_text("غ"))
                    total_absent += 1
                elif record["status"] == "excused":
                    date_cols.append(arabic_text("ع"))
                    total_present += 0.5  # Count excused as half present
                elif record["status"] == "late":
                    date_cols.append(arabic_text("ت"))
                    total_present += 1
                else:
                    date_cols.append("-")
            else:
                date_cols.append("-")
        
        total = total_present + total_absent
        rate = (total_present / total * 100) if total > 0 else 0
        
        # Build row RTL: % | absent | present | dates... | name | student_id | #
        row = [f"{rate:.0f}", str(int(total_absent)), str(int(total_present))]
        row.extend(date_cols)
        row.extend([arabic_text(student["full_name"]), student["student_id"], str(i)])
        data.append(row)
    
    # Calculate column widths (RTL)
    col_widths = [30, 35, 35] + [28] * len(display_dates) + [90, 55, 25]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table_style = [
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565c0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 9),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        # All cells
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), ARABIC_FONT),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        # Name column (3rd from end) محاذاة يمين
        ('ALIGN', (-3, 1), (-3, -1), 'RIGHT'),
        # Borders - softer
        ('LINEBELOW', (0, 0), (-1, 0), 1.2, colors.HexColor('#0d47a1')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cfd8dc')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#90a4ae')),
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f9ff')]),
        # Padding
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
    ]

    # Color-code attendance rates (% column = first col, 0)
    for row_idx in range(1, len(data)):
        try:
            rate_val = int(data[row_idx][0])
        except (ValueError, IndexError):
            continue
        if rate_val >= 75:
            table_style.append(('BACKGROUND', (0, row_idx), (0, row_idx), colors.HexColor('#c8e6c9')))
            table_style.append(('TEXTCOLOR', (0, row_idx), (0, row_idx), colors.HexColor('#1b5e20')))
        elif rate_val >= 50:
            table_style.append(('BACKGROUND', (0, row_idx), (0, row_idx), colors.HexColor('#fff9c4')))
            table_style.append(('TEXTCOLOR', (0, row_idx), (0, row_idx), colors.HexColor('#f57f17')))
        else:
            table_style.append(('BACKGROUND', (0, row_idx), (0, row_idx), colors.HexColor('#ffcdd2')))
            table_style.append(('TEXTCOLOR', (0, row_idx), (0, row_idx), colors.HexColor('#b71c1c')))

    table.setStyle(TableStyle(table_style))

    elements.append(table)

    # Summary box
    elements.append(Spacer(1, 18))
    summary_style = ParagraphStyle('Summary', fontName=ARABIC_FONT, fontSize=11, alignment=TA_RIGHT, textColor=colors.HexColor('#1565c0'), spaceAfter=4)
    elements.append(Paragraph(arabic_text(f"إجمالي الطلاب: {len(students)}    |    عدد المحاضرات: {len(dates)}    |    تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"),
                             summary_style))
    
    doc.build(elements)
    output.seek(0)
    
    filename = f"attendance_{course['code']}_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/export/report/{dept_id}/pdf")
async def export_department_report_pdf(
    dept_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Export department report to PDF"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "view_reports") and not has_permission(current_user, "export_reports"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    dept = await db.departments.find_one({"_id": ObjectId(dept_id)})
    if not dept:
        raise HTTPException(status_code=404, detail="القسم غير موجود")
    
    students = await db.students.find({"department_id": dept_id}).to_list(10000)
    courses = await db.courses.find({"department_id": dept_id}).to_list(100)
    
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    
    # Title
    title_style = ParagraphStyle('Title', fontName=ARABIC_FONT, fontSize=18, alignment=TA_CENTER, spaceAfter=10)
    elements.append(Paragraph(arabic_text("تقرير القسم"), title_style))
    elements.append(Paragraph(arabic_text(dept["name"]), 
                             ParagraphStyle('DeptName', fontName=ARABIC_FONT, fontSize=14, alignment=TA_CENTER, spaceAfter=20)))
    
    # Summary stats
    elements.append(Paragraph(arabic_text(f"إجمالي الطلاب: {len(students)}"), 
                             ParagraphStyle('Stat', fontName=ARABIC_FONT, fontSize=12, alignment=TA_RIGHT)))
    elements.append(Paragraph(arabic_text(f"إجمالي المقررات: {len(courses)}"), 
                             ParagraphStyle('Stat', fontName=ARABIC_FONT, fontSize=12, alignment=TA_RIGHT)))
    elements.append(Spacer(1, 20))
    
    # Students table
    elements.append(Paragraph(arabic_text("قائمة الطلاب"), 
                             ParagraphStyle('SectionTitle', fontName=ARABIC_FONT, fontSize=14, alignment=TA_CENTER, spaceAfter=10)))
    
    # RTL headers - from right to left
    headers = [arabic_text("نسبة الحضور"), arabic_text("الشعبة"), arabic_text("المستوى"),
               arabic_text("الاسم"), arabic_text("رقم الطالب"), arabic_text("م")]
    
    data = [headers]
    for i, s in enumerate(students[:50], 1):  # Limit to 50 for PDF
        # Calculate attendance rate
        records = await db.attendance.find({"student_id": str(s["_id"])}).to_list(10000)
        total = len(records)
        present = sum(1 for r in records if r["status"] == "present")
        rate = (present / total * 100) if total > 0 else 0
        
        # RTL row - from right to left
        row = [
            f"{rate:.1f}%",
            s.get("section", "") or "",
            str(s["level"]),
            arabic_text(s["full_name"]),
            s["student_id"],
            str(i),
        ]
        data.append(row)
    
    table = Table(data, colWidths=[60, 50, 50, 140, 75, 35], repeatRows=1)
    # Add color-coding for attendance rate (col 0 = first in array = % column)
    extra_styles = []
    for row_idx in range(1, len(data)):
        try:
            rate_val = float(data[row_idx][0].replace('%', ''))
        except Exception:
            continue
        if rate_val >= 75:
            extra_styles.append(('BACKGROUND', (0, row_idx), (0, row_idx), colors.HexColor('#c8e6c9')))
            extra_styles.append(('TEXTCOLOR', (0, row_idx), (0, row_idx), colors.HexColor('#1b5e20')))
        elif rate_val >= 50:
            extra_styles.append(('BACKGROUND', (0, row_idx), (0, row_idx), colors.HexColor('#fff9c4')))
            extra_styles.append(('TEXTCOLOR', (0, row_idx), (0, row_idx), colors.HexColor('#f57f17')))
        else:
            extra_styles.append(('BACKGROUND', (0, row_idx), (0, row_idx), colors.HexColor('#ffcdd2')))
            extra_styles.append(('TEXTCOLOR', (0, row_idx), (0, row_idx), colors.HexColor('#b71c1c')))
    # name col index in array = 3 (الاسم)
    table.setStyle(get_pro_table_style(extra=extra_styles, name_cols=[3]))
    
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(arabic_text(f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"), 
                             ParagraphStyle('Footer', fontName=ARABIC_FONT, fontSize=9, alignment=TA_CENTER)))
    
    doc.build(elements)
    output.seek(0)
    
    filename = f"report_{dept['code']}_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== Semesters Routes (إدارة الفصول الدراسية) ====================

@api_router.get("/semesters")
async def get_all_semesters(
    academic_year: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """الحصول على جميع الفصول الدراسية"""
    query = {}
    if academic_year:
        query["academic_year"] = academic_year
    if status:
        query["status"] = status
    
    semesters = await db.semesters.find(query).sort("created_at", -1).to_list(100)
    
    result = []
    for sem in semesters:
        # حساب الإحصائيات
        courses_count = await db.courses.count_documents({"semester_id": str(sem["_id"])})
        
        result.append({
            "id": str(sem["_id"]),
            "name": sem["name"],
            "academic_year": sem["academic_year"],
            "start_date": sem.get("start_date"),
            "end_date": sem.get("end_date"),
            "status": sem.get("status", SemesterStatus.UPCOMING),
            "term": sem.get("term"),
            "courses_count": courses_count,
            "created_at": sem.get("created_at", get_yemen_time()),
            "closed_at": sem.get("closed_at"),
            "archived_at": sem.get("archived_at"),
        })
    
    return result

@api_router.post("/semesters")
async def create_semester(data: SemesterCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء فصل دراسي جديد"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # التحقق من عدم وجود فصل بنفس الاسم في نفس السنة
    existing = await db.semesters.find_one({
        "name": data.name,
        "academic_year": data.academic_year
    })
    if existing:
        raise HTTPException(status_code=400, detail="يوجد فصل بهذا الاسم في هذه السنة")
    
    semester_dict = data.dict()
    semester_dict["created_at"] = get_yemen_time()
    semester_dict["created_by"] = current_user["id"]
    
    result = await db.semesters.insert_one(semester_dict)
    
    return {
        "id": str(result.inserted_id),
        "message": "تم إنشاء الفصل الدراسي بنجاح"
    }

@api_router.get("/semesters/current")
async def get_current_semester(current_user: dict = Depends(get_current_user)):
    """الحصول على الفصل الدراسي الحالي (النشط)"""
    semester = await db.semesters.find_one({"status": SemesterStatus.ACTIVE})
    
    if not semester:
        # إذا لم يوجد فصل نشط، نحاول الحصول عليه من الإعدادات
        settings = await db.settings.find_one({"_id": "system_settings"})
        if settings and settings.get("current_semester_id"):
            semester = await db.semesters.find_one({"_id": ObjectId(settings["current_semester_id"])})
    
    if not semester:
        return None
    
    return {
        "id": str(semester["_id"]),
        "name": semester["name"],
        "academic_year": semester["academic_year"],
        "start_date": semester.get("start_date"),
        "end_date": semester.get("end_date"),
        "status": semester.get("status"),
        "term": semester.get("term"),
    }

@api_router.put("/semesters/{semester_id}")
async def update_semester(semester_id: str, data: SemesterUpdate, current_user: dict = Depends(get_current_user)):
    """تحديث فصل دراسي"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    semester = await db.semesters.find_one({"_id": ObjectId(semester_id)})
    if not semester:
        raise HTTPException(status_code=404, detail="الفصل غير موجود")
    
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    if update_data:
        update_data["updated_at"] = get_yemen_time()
        await db.semesters.update_one({"_id": ObjectId(semester_id)}, {"$set": update_data})
        
        # إذا كان الفصل نشطاً، قم بتحديث الإعدادات أيضاً
        if semester.get("status") == SemesterStatus.ACTIVE:
            settings_update = {"updated_at": get_yemen_time()}
            
            if data.name:
                settings_update["current_semester"] = data.name
            if data.academic_year:
                settings_update["academic_year"] = data.academic_year
            if data.start_date:
                settings_update["semester_start_date"] = data.start_date
            if data.end_date:
                settings_update["semester_end_date"] = data.end_date
            
            if len(settings_update) > 1:  # أكثر من updated_at فقط
                await db.settings.update_one(
                    {"_id": "system_settings"},
                    {"$set": settings_update}
                )
    
    return {"message": "تم تحديث الفصل الدراسي بنجاح"}

@api_router.post("/semesters/{semester_id}/activate")
async def activate_semester(semester_id: str, auto_generate_from_curriculum: bool = False, current_user: dict = Depends(get_current_user)):
    """تفعيل فصل دراسي (جعله الفصل الحالي)
    - إذا auto_generate_from_curriculum=true: يُولّد المقررات تلقائياً من الخطة الدراسية
    """
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    semester = await db.semesters.find_one({"_id": ObjectId(semester_id)})
    if not semester:
        raise HTTPException(status_code=404, detail="الفصل غير موجود")
    
    if semester.get("status") == SemesterStatus.ARCHIVED:
        raise HTTPException(status_code=400, detail="لا يمكن تفعيل فصل مؤرشف")
    
    # إلغاء تفعيل الفصل الحالي
    await db.semesters.update_many(
        {"status": SemesterStatus.ACTIVE},
        {"$set": {"status": SemesterStatus.CLOSED, "closed_at": get_yemen_time()}}
    )
    
    # تفعيل الفصل الجديد
    await db.semesters.update_one(
        {"_id": ObjectId(semester_id)},
        {"$set": {"status": SemesterStatus.ACTIVE}}
    )
    
    # تحديث الإعدادات - تضمين تواريخ الفصل الدراسي
    settings_update = {
        "current_semester_id": semester_id,
        "current_semester": semester["name"],
        "academic_year": semester["academic_year"],
        "updated_at": get_yemen_time()
    }
    
    # إضافة تواريخ الفصل إذا كانت موجودة
    if semester.get("start_date"):
        settings_update["semester_start_date"] = semester["start_date"]
    if semester.get("end_date"):
        settings_update["semester_end_date"] = semester["end_date"]
    
    await db.settings.update_one(
        {"_id": "system_settings"},
        {"$set": settings_update},
        upsert=True
    )
    
    # توليد تلقائي للمقررات من الخطة الدراسية (Layer 3 من Layer 1+2)
    auto_gen_result = None
    if auto_generate_from_curriculum:
        try:
            # تحديد term: أولوية: semester.term (من الإعدادات) → استنتاج من الاسم (fallback)
            auto_term = semester.get("term")
            if auto_term is None:
                sem_name = (semester.get("name") or "").strip()
                if "الأول" in sem_name or "الاول" in sem_name or "first" in sem_name.lower():
                    auto_term = 1
                elif "الثاني" in sem_name or "ثاني" in sem_name or "second" in sem_name.lower():
                    auto_term = 2
                elif "صيف" in sem_name or "summer" in sem_name.lower():
                    auto_term = 3
                # حفظ القيمة المستنتجة للمرات القادمة
                if auto_term is not None:
                    try:
                        await db.semesters.update_one(
                            {"_id": ObjectId(semester_id)},
                            {"$set": {"term": auto_term}}
                        )
                    except Exception:
                        pass

            curr_filter = {"is_active": {"$ne": False}}
            # الصيفي (3) لا يفلتر - يأخذ كل المقررات المعتمدة
            if auto_term is not None and auto_term in (1, 2):
                curr_filter["term"] = auto_term

            total_curriculum = await db.curriculum_courses.count_documents(curr_filter)
            if total_curriculum > 0:
                assignments_map = {}
                async for a in db.teacher_assignments.find({"is_active": {"$ne": False}}):
                    ccid = a.get("curriculum_course_id")
                    if ccid and ccid not in assignments_map:
                        assignments_map[ccid] = a.get("teacher_id")
                created = 0
                skipped = 0
                async for cc in db.curriculum_courses.find(curr_filter):
                    cc_id = str(cc["_id"])
                    exists = await db.courses.find_one({
                        "semester_id": semester_id,
                        "curriculum_course_id": cc_id,
                    })
                    if exists:
                        skipped += 1
                        continue
                    doc = {
                        "curriculum_course_id": cc_id,
                        "semester_id": semester_id,
                        "code": cc.get("code"),
                        "name": cc.get("name"),
                        "credit_hours": cc.get("credit_hours", 3),
                        "department_id": cc.get("department_id"),
                        "faculty_id": cc.get("faculty_id"),
                        "level": cc.get("level"),
                        "term": cc.get("term"),
                        "section": "أ",
                        "room": "",
                        "teacher_id": assignments_map.get(cc_id),
                        "created_at": get_yemen_time(),
                        "created_by": current_user.get("id"),
                        "auto_generated": True,
                    }
                    await db.courses.insert_one(doc)
                    created += 1
                auto_gen_result = {"created": created, "skipped": skipped, "total_curriculum": total_curriculum}
        except Exception as e:
            logger.warning(f"auto_generate_from_curriculum failed: {e}")
    
    msg = f"تم تفعيل الفصل '{semester['name']}' بنجاح"
    if auto_gen_result:
        msg += f" • تم توليد {auto_gen_result['created']} مقرر من الخطة"
    return {"message": msg, "auto_generated": auto_gen_result}

@api_router.post("/semesters/{semester_id}/close")
async def close_semester(semester_id: str, current_user: dict = Depends(get_current_user)):
    """إغلاق فصل دراسي"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    semester = await db.semesters.find_one({"_id": ObjectId(semester_id)})
    if not semester:
        raise HTTPException(status_code=404, detail="الفصل غير موجود")
    
    if semester.get("status") == SemesterStatus.ARCHIVED:
        raise HTTPException(status_code=400, detail="الفصل مؤرشف مسبقاً")
    
    await db.semesters.update_one(
        {"_id": ObjectId(semester_id)},
        {"$set": {"status": SemesterStatus.CLOSED, "closed_at": get_yemen_time()}}
    )
    
    return {"message": "تم إغلاق الفصل الدراسي بنجاح"}

@api_router.post("/semesters/{semester_id}/archive")
async def archive_semester(semester_id: str, current_user: dict = Depends(get_current_user)):
    """أرشفة فصل دراسي بشكل شامل:
    - يحفظ snapshot كامل للعمليات (مقررات، محاضرات، حضور، تسجيلات، خطط)
    - مع denormalized snapshot لأسماء الطلاب والمعلمين والأقسام وقت الأرشفة
    - يحذف العمليات التشغيلية فقط بعد النسخ (لا يلمس students/teachers/departments/faculties)
    """
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    semester = await db.semesters.find_one({"_id": ObjectId(semester_id)})
    if not semester:
        raise HTTPException(status_code=404, detail="الفصل غير موجود")

    if semester.get("status") == SemesterStatus.ACTIVE:
        raise HTTPException(status_code=400, detail="لا يمكن أرشفة فصل نشط. يرجى إغلاقه أولاً")

    if semester.get("status") == SemesterStatus.ARCHIVED:
        raise HTTPException(status_code=400, detail="الفصل مؤرشف مسبقاً")

    # ============== جمع كل البيانات التشغيلية ==============
    courses_live = await db.courses.find({"semester_id": semester_id}).to_list(5000)
    course_ids = [str(c["_id"]) for c in courses_live]

    if not course_ids:
        # فصل بلا مقررات - أرشفة فارغة
        archive_record = {
            "semester_id": semester_id,
            "semester_name": semester["name"],
            "academic_year": semester["academic_year"],
            "semester_start": semester.get("start_date"),
            "semester_end": semester.get("end_date"),
            "archived_at": get_yemen_time(),
            "archived_by": current_user["id"],
            "archived_by_name": current_user.get("full_name") or current_user.get("username"),
            "summary": {"total_courses": 0, "total_students": 0, "total_teachers": 0,
                        "total_lectures": 0, "completed_lectures": 0,
                        "total_attendance_records": 0, "overall_attendance_rate": 0},
            "courses": [], "lectures": [], "attendance": [], "enrollments": [], "study_plans": [],
            "students_snapshot": [], "teachers_snapshot": [],
            "departments_snapshot": [], "faculties_snapshot": [],
        }
        await db.semester_archives.insert_one(archive_record)
        await db.semesters.update_one(
            {"_id": ObjectId(semester_id)},
            {"$set": {"status": SemesterStatus.ARCHIVED, "archived_at": get_yemen_time(),
                      "archive_stats": archive_record["summary"]}}
        )
        return {"message": "تم أرشفة الفصل (لا توجد بيانات تشغيلية)", "summary": archive_record["summary"]}

    lectures_live = await db.lectures.find({"semester_id": semester_id}).to_list(50000)
    if not lectures_live:
        lectures_live = await db.lectures.find({"course_id": {"$in": course_ids}}).to_list(50000)
    attendance_live = await db.attendance.find({"course_id": {"$in": course_ids}}).to_list(500000)
    enrollments_live = await db.enrollments.find({"course_id": {"$in": course_ids}}).to_list(50000)
    study_plans_live = await db.study_plans.find({"course_id": {"$in": course_ids}}).to_list(5000)

    # ============== جمع snapshot للأسماء (denormalized) ==============
    teacher_ids = list({c.get("teacher_id") for c in courses_live if c.get("teacher_id")})
    student_ids = list({e.get("student_id") for e in enrollments_live if e.get("student_id")})
    dept_ids = list({c.get("department_id") for c in courses_live if c.get("department_id")})

    teachers_map, students_map, depts_map, faculties_map = {}, {}, {}, {}
    if teacher_ids:
        try:
            tobj = [ObjectId(t) for t in teacher_ids if t]
            async for t in db.teachers.find({"_id": {"$in": tobj}}):
                teachers_map[str(t["_id"])] = {
                    "id": str(t["_id"]), "full_name": t.get("full_name", ""),
                    "teacher_id": t.get("teacher_id", ""),
                    "academic_title": t.get("academic_title"),
                    "specialization": t.get("specialization"),
                }
        except Exception:
            pass
    if student_ids:
        try:
            sobj = [ObjectId(s) for s in student_ids if s]
            async for s in db.students.find({"_id": {"$in": sobj}}):
                students_map[str(s["_id"])] = {
                    "id": str(s["_id"]), "full_name": s.get("full_name", ""),
                    "student_id": s.get("student_id", ""),
                    "reference_number": s.get("reference_number"),
                    "department_id": s.get("department_id"),
                    "level": s.get("level"), "section": s.get("section"),
                }
        except Exception:
            pass
    if dept_ids:
        try:
            dobj = [ObjectId(d) for d in dept_ids if d]
            async for d in db.departments.find({"_id": {"$in": dobj}}):
                depts_map[str(d["_id"])] = {
                    "id": str(d["_id"]), "name": d.get("name", ""), "code": d.get("code", ""),
                    "faculty_id": d.get("faculty_id"),
                }
                if d.get("faculty_id"):
                    try:
                        f = await db.faculties.find_one({"_id": ObjectId(d["faculty_id"])})
                        if f:
                            faculties_map[str(f["_id"])] = {
                                "id": str(f["_id"]), "name": f.get("name", ""), "code": f.get("code", ""),
                            }
                    except Exception:
                        pass
        except Exception:
            pass

    # ============== إثراء المقررات بالأسماء + الإحصائيات ==============
    courses_enriched = []
    completed_lec_total = 0
    for c in courses_live:
        cid = str(c["_id"])
        c_lecs = [l for l in lectures_live if str(l.get("course_id")) == cid]
        c_completed = sum(1 for l in c_lecs if l.get("status") == "completed")
        completed_lec_total += c_completed
        c_students = sum(1 for e in enrollments_live if str(e.get("course_id")) == cid)
        t_info = teachers_map.get(str(c.get("teacher_id"))) if c.get("teacher_id") else None
        d_info = depts_map.get(str(c.get("department_id"))) if c.get("department_id") else None
        courses_enriched.append({
            "id": cid,
            "name": c.get("name", ""), "code": c.get("code", ""),
            "level": c.get("level"), "section": c.get("section", ""),
            "credit_hours": c.get("credit_hours", 3),
            "room": c.get("room", ""),
            "teacher_id": c.get("teacher_id"),
            "teacher_name": (t_info or {}).get("full_name"),
            "department_id": c.get("department_id"),
            "department_name": (d_info or {}).get("name"),
            "students_count": c_students,
            "lectures_total": len(c_lecs),
            "lectures_completed": c_completed,
            "completion_pct": round((c_completed / len(c_lecs) * 100) if c_lecs else 0, 1),
        })

    # ============== الملخص العام ==============
    overall_rate = 0.0
    if attendance_live:
        present_count = sum(1 for a in attendance_live if a.get("status") in ("present", "late"))
        overall_rate = round((present_count / len(attendance_live)) * 100, 2)

    summary = {
        "total_courses": len(courses_live),
        "total_students": len(student_ids),
        "total_teachers": len(teacher_ids),
        "total_lectures": len(lectures_live),
        "completed_lectures": completed_lec_total,
        "total_attendance_records": len(attendance_live),
        "overall_attendance_rate": overall_rate,
    }

    # ============== تحويل ObjectId لـ str وتنظيف ==============
    def _clean(items, drop_id=True):
        result = []
        for x in items:
            d = dict(x)
            d["_archive_id"] = str(d.pop("_id"))
            for k, v in list(d.items()):
                if isinstance(v, ObjectId):
                    d[k] = str(v)
            result.append(d)
        return result

    archive_record = {
        "semester_id": semester_id,
        "semester_name": semester["name"],
        "academic_year": semester["academic_year"],
        "semester_start": semester.get("start_date"),
        "semester_end": semester.get("end_date"),
        "archived_at": get_yemen_time(),
        "archived_by": current_user["id"],
        "archived_by_name": current_user.get("full_name") or current_user.get("username"),
        "summary": summary,
        # snapshots مُحسَّبة جاهزة للعرض
        "courses": courses_enriched,
        # raw data للتفاصيل والاستعادة
        "courses_raw": _clean(courses_live),
        "lectures": _clean(lectures_live),
        "attendance": _clean(attendance_live),
        "enrollments": _clean(enrollments_live),
        "study_plans": _clean(study_plans_live),
        # snapshots للأسماء (denormalized لحماية الأرشيف من تغييرات لاحقة)
        "students_snapshot": list(students_map.values()),
        "teachers_snapshot": list(teachers_map.values()),
        "departments_snapshot": list(depts_map.values()),
        "faculties_snapshot": list(faculties_map.values()),
    }

    # ============== الإدراج في الأرشيف ==============
    insert_result = await db.semester_archives.insert_one(archive_record)
    archive_doc_id = insert_result.inserted_id

    # ============== ✅ التحقق من سلامة الأرشيف قبل الحذف ==============
    saved = await db.semester_archives.find_one({"_id": archive_doc_id})
    if not saved:
        raise HTTPException(status_code=500, detail="فشل حفظ الأرشيف. لم يُحذف شيء.")

    expected = {
        "courses": len(courses_live),
        "lectures": len(lectures_live),
        "attendance": len(attendance_live),
        "enrollments": len(enrollments_live),
        "study_plans": len(study_plans_live),
    }
    actual = {
        "courses": len(saved.get("courses", [])),
        "lectures": len(saved.get("lectures", [])),
        "attendance": len(saved.get("attendance", [])),
        "enrollments": len(saved.get("enrollments", [])),
        "study_plans": len(saved.get("study_plans", [])),
    }
    mismatches = [k for k in expected if expected[k] != actual[k]]
    if mismatches:
        # rollback: نحذف الأرشيف الفاشل ونرفض الأرشفة
        await db.semester_archives.delete_one({"_id": archive_doc_id})
        raise HTTPException(
            status_code=500,
            detail=f"فشل التحقق من سلامة الأرشيف ({', '.join(mismatches)}). "
                   f"تم إلغاء العملية ولم يُحذف شيء من البيانات الحية."
        )

    # ============== حذف العمليات التشغيلية فقط ==============
    # (لا نحذف: students, teachers, departments, faculties, users)
    del_courses = await db.courses.delete_many({"semester_id": semester_id})
    del_lectures_a = await db.lectures.delete_many({"semester_id": semester_id})
    del_lectures_b = await db.lectures.delete_many({"course_id": {"$in": course_ids}})
    del_attendance = await db.attendance.delete_many({"course_id": {"$in": course_ids}})
    del_enrollments = await db.enrollments.delete_many({"course_id": {"$in": course_ids}})
    del_plans = await db.study_plans.delete_many({"course_id": {"$in": course_ids}})

    # ============== تحديث حالة الفصل ==============
    await db.semesters.update_one(
        {"_id": ObjectId(semester_id)},
        {"$set": {
            "status": SemesterStatus.ARCHIVED,
            "archived_at": get_yemen_time(),
            "archive_stats": summary,
        }}
    )

    return {
        "message": f"تم أرشفة الفصل '{semester['name']}' بنجاح (تم التحقق من سلامة النسخ)",
        "summary": summary,
        "verification": {"expected": expected, "actual": actual, "verified": True},
        "deleted": {
            "courses": del_courses.deleted_count,
            "lectures": del_lectures_a.deleted_count + del_lectures_b.deleted_count,
            "attendance": del_attendance.deleted_count,
            "enrollments": del_enrollments.deleted_count,
            "study_plans": del_plans.deleted_count,
        },
        "restore_hint": f"يمكنك التراجع عبر: POST /api/semesters/{semester_id}/restore",
    }


@api_router.post("/semesters/{semester_id}/restore")
async def restore_semester_from_archive(
    semester_id: str,
    current_user: dict = Depends(get_current_user),
):
    """استعادة فصل من الأرشيف:
    - يُعيد كل البيانات التشغيلية (courses, lectures, attendance, enrollments, study_plans)
      من collection `semester_archives` إلى الـ collections الحية.
    - يُغيّر حالة الفصل من ARCHIVED إلى CLOSED.
    - يحذف وثيقة الأرشيف بعد نجاح الاستعادة.
    - الأسماء (students, teachers, departments, faculties) لم تُلمس أصلاً.
    صلاحية مطلوبة: admin أو manage_courses.
    """
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بالاستعادة")

    try:
        sem_oid = ObjectId(semester_id)
    except Exception:
        raise HTTPException(status_code=400, detail="معرّف الفصل غير صحيح")

    semester = await db.semesters.find_one({"_id": sem_oid})
    if not semester:
        raise HTTPException(status_code=404, detail="الفصل غير موجود")

    if semester.get("status") != SemesterStatus.ARCHIVED:
        raise HTTPException(status_code=400, detail="الفصل ليس مؤرشفاً")

    archive = await db.semester_archives.find_one({"semester_id": semester_id})
    if not archive:
        raise HTTPException(status_code=500, detail="وثيقة الأرشيف غير موجودة! لا يمكن الاستعادة.")

    # ============== رفض إذا كانت البيانات الحية تحوي شيئاً للفصل ==============
    # (حماية إضافية: قد يكون أحد بدأ يستخدم الفصل من جديد بين الأرشفة والاستعادة)
    live_courses = await db.courses.count_documents({"semester_id": semester_id})
    if live_courses > 0:
        raise HTTPException(
            status_code=409,
            detail=f"البيانات الحية تحوي {live_courses} مقرراً للفصل بالفعل. "
                   f"الاستعادة قد تنتج تكراراً. يرجى مراجعة البيانات أولاً."
        )

    # ============== تحويل _archive_id إلى _id ==============
    def _restore_ids(items):
        result = []
        for x in items:
            d = dict(x)
            if "_archive_id" in d:
                try:
                    d["_id"] = ObjectId(d.pop("_archive_id"))
                except Exception:
                    d.pop("_archive_id", None)
            # تحويل الحقول التي تشبه ObjectId
            for k in ("teacher_id", "department_id", "faculty_id", "course_id",
                      "student_id", "semester_id", "lecture_id"):
                if k in d and isinstance(d[k], str) and len(d[k]) == 24:
                    # نتركها string كما كانت في النظام
                    pass
            result.append(d)
        return result

    courses_to_restore = _restore_ids(archive.get("courses_raw", []))
    lectures_to_restore = _restore_ids(archive.get("lectures", []))
    attendance_to_restore = _restore_ids(archive.get("attendance", []))
    enrollments_to_restore = _restore_ids(archive.get("enrollments", []))
    plans_to_restore = _restore_ids(archive.get("study_plans", []))

    # ============== إعادة الإدراج ==============
    inserted = {"courses": 0, "lectures": 0, "attendance": 0, "enrollments": 0, "study_plans": 0}
    if courses_to_restore:
        r = await db.courses.insert_many(courses_to_restore, ordered=False)
        inserted["courses"] = len(r.inserted_ids)
    if lectures_to_restore:
        r = await db.lectures.insert_many(lectures_to_restore, ordered=False)
        inserted["lectures"] = len(r.inserted_ids)
    if attendance_to_restore:
        r = await db.attendance.insert_many(attendance_to_restore, ordered=False)
        inserted["attendance"] = len(r.inserted_ids)
    if enrollments_to_restore:
        r = await db.enrollments.insert_many(enrollments_to_restore, ordered=False)
        inserted["enrollments"] = len(r.inserted_ids)
    if plans_to_restore:
        r = await db.study_plans.insert_many(plans_to_restore, ordered=False)
        inserted["study_plans"] = len(r.inserted_ids)

    # ============== تحديث حالة الفصل إلى CLOSED ==============
    await db.semesters.update_one(
        {"_id": sem_oid},
        {"$set": {"status": SemesterStatus.CLOSED},
         "$unset": {"archived_at": "", "archive_stats": ""}}
    )

    # ============== حذف وثيقة الأرشيف ==============
    await db.semester_archives.delete_one({"semester_id": semester_id})

    return {
        "message": f"تم استعادة الفصل '{semester['name']}' من الأرشيف بنجاح",
        "restored": inserted,
        "new_status": "closed",
    }


@api_router.get("/semesters/{semester_id}/stats")
async def get_semester_stats(semester_id: str, current_user: dict = Depends(get_current_user)):
    """الحصول على إحصائيات فصل دراسي"""
    semester = await db.semesters.find_one({"_id": ObjectId(semester_id)})
    if not semester:
        raise HTTPException(status_code=404, detail="الفصل غير موجود")
    
    # جمع الإحصائيات
    courses = await db.courses.find({"semester_id": semester_id}).to_list(1000)
    course_ids = [str(c["_id"]) for c in courses]
    
    attendance_count = await db.attendance.count_documents({"course_id": {"$in": course_ids}})
    lectures_count = await db.lectures.count_documents({"course_id": {"$in": course_ids}})
    
    # جمع الطلاب المسجلين
    enrollments = await db.enrollments.find({"course_id": {"$in": course_ids}}).to_list(10000)
    unique_students = set([e["student_id"] for e in enrollments])
    
    return {
        "semester_id": semester_id,
        "semester_name": semester["name"],
        "academic_year": semester["academic_year"],
        "status": semester.get("status"),
        "courses_count": len(courses),
        "students_count": len(unique_students),
        "lectures_count": lectures_count,
        "attendance_records": attendance_count,
    }

@api_router.delete("/semesters/{semester_id}")
async def delete_semester(semester_id: str, current_user: dict = Depends(get_current_user)):
    """حذف فصل دراسي (فقط إذا لم يكن له مقررات)"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    semester = await db.semesters.find_one({"_id": ObjectId(semester_id)})
    if not semester:
        raise HTTPException(status_code=404, detail="الفصل غير موجود")
    
    # التحقق من عدم وجود مقررات مرتبطة
    courses_count = await db.courses.count_documents({"semester_id": semester_id})
    if courses_count > 0:
        raise HTTPException(status_code=400, detail=f"لا يمكن حذف الفصل، يوجد {courses_count} مقرر مرتبط به")
    
    await db.semesters.delete_one({"_id": ObjectId(semester_id)})
    
    return {"message": "تم حذف الفصل الدراسي بنجاح"}

# ==================== Settings Routes (الإعدادات العامة) ====================

@api_router.get("/settings")
async def get_settings(current_user: dict = Depends(get_current_user)):
    """الحصول على إعدادات النظام"""
    settings = await db.settings.find_one({"_id": "system_settings"})
    if not settings:
        # إنشاء إعدادات افتراضية
        default_settings = SystemSettings().dict()
        default_settings["_id"] = "system_settings"
        default_settings["created_at"] = get_yemen_time()
        default_settings["updated_at"] = get_yemen_time()
        # إضافة سنوات افتراضية
        current_year = datetime.now().year
        default_settings["academic_years"] = [f"{current_year-1}-{current_year}", f"{current_year}-{current_year+1}"]
        await db.settings.insert_one(default_settings)
        settings = default_settings
    
    # جلب تواريخ الفصل من الفصل النشط إذا لم تكن في الإعدادات
    semester_start = settings.get("semester_start_date")
    semester_end = settings.get("semester_end_date")
    current_semester_id = settings.get("current_semester_id")
    current_semester_name = settings.get("current_semester", "الفصل الأول")
    academic_year = settings.get("academic_year", "2024-2025")
    
    # محاولة جلب البيانات من الفصل النشط إذا كانت ناقصة
    if not semester_start or not semester_end or not current_semester_id:
        active_semester = await db.semesters.find_one({"status": SemesterStatus.ACTIVE})
        if not active_semester:
            # محاولة البحث عن أي فصل غير مؤرشف
            active_semester = await db.semesters.find_one({"status": {"$ne": "archived"}})
        
        if active_semester:
            semester_start = semester_start or active_semester.get("start_date")
            semester_end = semester_end or active_semester.get("end_date")
            
            # تحديث الإعدادات تلقائياً
            update_fields = {}
            if not current_semester_id:
                current_semester_id = str(active_semester["_id"])
                update_fields["current_semester_id"] = current_semester_id
            if not settings.get("current_semester"):
                current_semester_name = active_semester.get("name", current_semester_name)
                update_fields["current_semester"] = current_semester_name
            if active_semester.get("academic_year") and not settings.get("academic_year"):
                academic_year = active_semester["academic_year"]
                update_fields["academic_year"] = academic_year
            if active_semester.get("start_date") and not settings.get("semester_start_date"):
                update_fields["semester_start_date"] = active_semester["start_date"]
            if active_semester.get("end_date") and not settings.get("semester_end_date"):
                update_fields["semester_end_date"] = active_semester["end_date"]
            
            if update_fields:
                update_fields["updated_at"] = get_yemen_time()
                await db.settings.update_one(
                    {"_id": "system_settings"},
                    {"$set": update_fields},
                    upsert=True
                )
    
    return {
        "college_name": settings.get("college_name", "كلية الشريعة والقانون"),
        "college_name_en": settings.get("college_name_en", "Faculty of Sharia and Law"),
        "academic_year": academic_year,
        "current_semester": current_semester_name,
        "current_semester_id": current_semester_id,
        "semester_start_date": semester_start,
        "semester_end_date": semester_end,
        "levels_count": settings.get("levels_count", 5),
        "sections": settings.get("sections", ["أ", "ب", "ج"]),
        "attendance_late_minutes": settings.get("attendance_late_minutes", 15),
        "attendance_edit_minutes": settings.get("attendance_edit_minutes", 60),
        "max_absence_percent": settings.get("max_absence_percent", 25.0),
        "logo_url": settings.get("logo_url"),
        "primary_color": settings.get("primary_color", "#1565c0"),
        "secondary_color": settings.get("secondary_color", "#ff9800"),
        "academic_years": settings.get("academic_years", []),
        "updated_at": settings.get("updated_at"),
    }

@api_router.put("/settings")
async def update_settings(data: SettingsUpdate, current_user: dict = Depends(get_current_user)):
    """تحديث إعدادات النظام - للمدير فقط"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل الإعدادات")
    
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    update_data["updated_at"] = get_yemen_time()
    
    await db.settings.update_one(
        {"_id": "system_settings"},
        {"$set": update_data},
        upsert=True
    )
    
    return {"message": "تم تحديث الإعدادات بنجاح"}

@api_router.post("/settings/academic-years")
async def add_academic_year(year: str = Body(..., embed=True), current_user: dict = Depends(get_current_user)):
    """إضافة سنة أكاديمية جديدة"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # التحقق من صيغة السنة
    import re
    if not re.match(r'^\d{4}-\d{4}$', year):
        raise HTTPException(status_code=400, detail="صيغة السنة غير صحيحة. استخدم الصيغة: YYYY-YYYY")
    
    await db.settings.update_one(
        {"_id": "system_settings"},
        {"$addToSet": {"academic_years": year}},
        upsert=True
    )
    
    return {"message": f"تم إضافة السنة الأكاديمية {year} بنجاح"}

@api_router.delete("/settings/academic-years/{year}")
async def delete_academic_year(year: str, current_user: dict = Depends(get_current_user)):
    """حذف سنة أكاديمية"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    await db.settings.update_one(
        {"_id": "system_settings"},
        {"$pull": {"academic_years": year}}
    )
    
    return {"message": f"تم حذف السنة الأكاديمية {year}"}

@api_router.get("/settings/academic-years")
async def get_academic_years():
    """الحصول على قائمة السنوات الأكاديمية المتاحة"""
    settings = await db.settings.find_one({"_id": "system_settings"})
    if settings and settings.get("academic_years"):
        return {"years": sorted(settings["academic_years"], reverse=True)}
    
    # إرجاع سنوات افتراضية
    current_year = datetime.now().year
    years = [f"{current_year-1}-{current_year}", f"{current_year}-{current_year+1}"]
    return {"years": years}

@api_router.get("/settings/semesters")
async def get_semesters():
    """الحصول على قائمة الفصول الدراسية"""
    return {"semesters": ["الفصل الأول", "الفصل الثاني", "الفصل الصيفي"]}

@api_router.get("/my-scope")
async def get_my_scope(current_user: dict = Depends(get_current_user)):
    """
    الحصول على نطاق صلاحيات المستخدم الحالي
    يحدد ما يمكنه الوصول إليه (جامعة، كليات، أقسام، مقررات)
    """
    user = await db.users.find_one({"_id": ObjectId(current_user["id"])})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    scope = {
        "level": "none",  # university, faculty, department, course, none
        "university_access": False,
        "faculties": [],  # قائمة الكليات المسموح بها
        "departments": [],  # قائمة الأقسام المسموح بها
        "courses": [],  # قائمة المقررات المسموح بها
        "can_manage_settings": False,  # هل يمكنه إدارة الإعدادات العامة
    }
    
    # المدير له صلاحية كاملة
    if current_user["role"] == UserRole.ADMIN:
        scope["level"] = "university"
        scope["university_access"] = True
        scope["can_manage_settings"] = True
        # جلب كل الكليات
        faculties = await db.faculties.find({}).to_list(None)
        scope["faculties"] = [{"id": str(f["_id"]), "name": f["name"]} for f in faculties]
        return scope
    
    # التحقق من مستوى الصلاحية المحدد للمستخدم
    permission_level = user.get("permission_level", "")
    
    # التحقق من الكليات المخصصة
    faculty_ids = user.get("faculty_ids", [])
    if not faculty_ids and user.get("faculty_id"):
        faculty_ids = [user.get("faculty_id")]
    
    # التحقق من الأقسام المخصصة
    department_ids = user.get("department_ids", [])
    if not department_ids and user.get("department_id"):
        department_ids = [user.get("department_id")]
    
    # التحقق من المقررات المخصصة
    course_ids = user.get("course_ids", [])
    
    # استنتاج مستوى الصلاحية إذا لم يكن محدداً
    # الأولوية للأكثر تحديداً (قسم > كلية > جامعة)
    if not permission_level:
        if department_ids:
            permission_level = "department"
        elif faculty_ids:
            permission_level = "faculty"
    
    # تحديد مستوى الصلاحية بناءً على البيانات
    if permission_level == "university" or (user.get("permissions") and "manage_university" in user.get("permissions", [])):
        scope["level"] = "university"
        scope["university_access"] = True
        scope["can_manage_settings"] = True
        faculties = await db.faculties.find({}).to_list(None)
        scope["faculties"] = [{"id": str(f["_id"]), "name": f["name"]} for f in faculties]
    elif permission_level == "department" or (department_ids and not (permission_level == "faculty")):
        # مستوى القسم - الأكثر تحديداً
        scope["level"] = "department"
        # جلب الأقسام المحددة فقط
        for did in department_ids:
            try:
                dept = await db.departments.find_one({"_id": ObjectId(did)})
                if dept:
                    scope["departments"].append({
                        "id": str(dept["_id"]), 
                        "name": dept["name"],
                        "faculty_id": dept.get("faculty_id")
                    })
                    # إضافة الكلية التابع لها القسم (للعرض فقط، ليس للصلاحية الكاملة)
                    if dept.get("faculty_id"):
                        faculty = await db.faculties.find_one({"_id": ObjectId(dept["faculty_id"])})
                        if faculty and not any(f["id"] == str(faculty["_id"]) for f in scope["faculties"]):
                            scope["faculties"].append({"id": str(faculty["_id"]), "name": faculty["name"]})
            except:
                pass
    elif permission_level == "faculty" or faculty_ids:
        # مستوى الكلية
        scope["level"] = "faculty"
        # جلب الكليات المحددة فقط
        for fid in faculty_ids:
            try:
                faculty = await db.faculties.find_one({"_id": ObjectId(fid)})
                if faculty:
                    scope["faculties"].append({"id": str(faculty["_id"]), "name": faculty["name"]})
            except:
                pass
        # جلب الأقسام التابعة للكليات المحددة
        if scope["faculties"]:
            departments = await db.departments.find({"faculty_id": {"$in": faculty_ids}}).to_list(None)
            scope["departments"] = [{"id": str(d["_id"]), "name": d["name"], "faculty_id": d.get("faculty_id")} for d in departments]
    elif course_ids:
        scope["level"] = "course"
        # جلب المقررات المحددة فقط
        for cid in course_ids:
            try:
                course = await db.courses.find_one({"_id": ObjectId(cid)})
                if course:
                    scope["courses"].append({"id": str(course["_id"]), "name": course["name"]})
            except:
                pass
    
    return scope

@api_router.get("/my-institution")
async def get_my_institution(current_user: dict = Depends(get_current_user)):
    """
    الحصول على بيانات المؤسسة الخاصة بالمستخدم الحالي
    - المدير: يحصل على بيانات الجامعة
    - المستخدمين الآخرين: يحصلون على بيانات كليتهم
    """
    user = await db.users.find_one({"_id": ObjectId(current_user["id"])})
    
    if current_user["role"] == UserRole.ADMIN:
        # المدير يرى بيانات الجامعة
        university = await db.university.find_one({})
        if university:
            return {
                "type": "university",
                "id": str(university["_id"]),
                "name": university.get("name", "جامعة الأحقاف"),
                "name_en": university.get("name_en", "Al-Ahgaff University"),
                "code": university.get("code", "AHGAFF"),
                "description": university.get("description"),
                "logo_url": university.get("logo_url"),
                "address": university.get("address"),
                "phone": university.get("phone"),
                "email": university.get("email"),
                "website": university.get("website"),
            }
        else:
            # إذا لم توجد جامعة، ارجع الإعدادات العامة
            settings = await db.settings.find_one({"_id": "system_settings"})
            return {
                "type": "settings",
                "name": settings.get("college_name", "جامعة الأحقاف") if settings else "جامعة الأحقاف",
                "name_en": settings.get("college_name_en", "Al-Ahgaff University") if settings else "Al-Ahgaff University",
            }
    else:
        # المستخدمون الآخرون يرون بيانات كليتهم
        faculty_id = user.get("faculty_id") if user else None
        
        if faculty_id:
            faculty = await db.faculties.find_one({"_id": ObjectId(faculty_id)})
            if faculty:
                return {
                    "type": "faculty",
                    "id": str(faculty["_id"]),
                    "name": faculty.get("name"),
                    "name_en": faculty.get("name_en"),
                    "code": faculty.get("code"),
                    "description": faculty.get("description"),
                    "levels_count": faculty.get("levels_count", 5),
                    "sections": faculty.get("sections", ["أ", "ب", "ج"]),
                    "attendance_late_minutes": faculty.get("attendance_late_minutes", 15),
                    "max_absence_percent": faculty.get("max_absence_percent", 25),
                }
        
        # إذا لم يكن مرتبطاً بكلية، ارجع الإعدادات العامة
        settings = await db.settings.find_one({"_id": "system_settings"})
        return {
            "type": "settings",
            "name": settings.get("college_name", "النظام") if settings else "النظام",
            "name_en": settings.get("college_name_en") if settings else None,
            "levels_count": settings.get("levels_count", 5) if settings else 5,
            "sections": settings.get("sections", ["أ", "ب", "ج"]) if settings else ["أ", "ب", "ج"],
            "attendance_late_minutes": settings.get("attendance_late_minutes", 15) if settings else 15,
            "max_absence_percent": settings.get("max_absence_percent", 25) if settings else 25,
        }

@api_router.put("/my-institution")
async def update_my_institution(data: dict = Body(...), current_user: dict = Depends(get_current_user)):
    """
    تحديث بيانات المؤسسة الخاصة بالمستخدم
    - المدير: يمكنه تحديث بيانات الجامعة
    - مستخدم لديه صلاحية: يمكنه تحديث بيانات كليته
    """
    user = await db.users.find_one({"_id": ObjectId(current_user["id"])})
    
    if current_user["role"] == UserRole.ADMIN:
        # المدير يحدث بيانات الجامعة
        update_data = {k: v for k, v in data.items() if v is not None and k != "type"}
        update_data["updated_at"] = get_yemen_time()
        
        await db.university.update_one(
            {},
            {"$set": update_data},
            upsert=True
        )
        return {"message": "تم تحديث بيانات الجامعة بنجاح"}
    else:
        # التحقق من صلاحية التحديث
        faculty_id = user.get("faculty_id") if user else None
        if not faculty_id:
            raise HTTPException(status_code=403, detail="أنت غير مرتبط بكلية")
        
        # تحديث بيانات الكلية (الإعدادات الخاصة فقط)
        allowed_fields = ["levels_count", "sections", "attendance_late_minutes", "max_absence_percent", "description"]
        update_data = {k: v for k, v in data.items() if k in allowed_fields and v is not None}
        
        if update_data:
            update_data["updated_at"] = get_yemen_time()
            await db.faculties.update_one(
                {"_id": ObjectId(faculty_id)},
                {"$set": update_data}
            )
        
        return {"message": "تم تحديث إعدادات الكلية بنجاح"}

@api_router.get("/")
async def root():
    return {"message": "نظام حضور كلية الشريعة والقانون", "version": "1.0"}

# ==================== APIs إدارة الصلاحيات المتقدمة ====================

@api_router.get("/permissions/available")
async def get_available_permissions(current_user: dict = Depends(get_current_user)):
    """الحصول على قائمة الصلاحيات المتاحة (تستبعد الصلاحيات المخفية hidden)"""
    visible_perms = [p for p in ALL_PERMISSIONS if not p.get("hidden", False)]
    return {
        "permissions": visible_perms,
        "scope_types": [
            {"key": "global", "label": "عامة (كل النظام)"},
            {"key": "department", "label": "قسم معين"},
            {"key": "course", "label": "مقرر معين"},
        ],
        "roles": [
            {"key": UserRole.ADMIN, "label": "مدير النظام"},
            {"key": UserRole.TEACHER, "label": "معلم"},
            {"key": UserRole.EMPLOYEE, "label": "موظف"},
            {"key": UserRole.STUDENT, "label": "طالب"},
        ]
    }

@api_router.get("/users/{user_id}/permissions")
async def get_user_permissions_endpoint_v2(user_id: str, current_user: dict = Depends(get_current_user)):
    """الحصول على صلاحيات مستخدم معين"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_users"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بعرض صلاحيات المستخدمين")
    
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # جلب الصلاحيات المخصصة من جدول user_permissions
    user_perms = await db.user_permissions.find({"user_id": user_id}).to_list(100)
    
    scopes = []
    for perm in user_perms:
        scopes.append({
            "id": str(perm["_id"]),
            "scope_type": perm["scope_type"],
            "scope_id": perm.get("scope_id"),
            "scope_name": perm.get("scope_name", ""),
            "permissions": perm["permissions"]
        })
    
    # إذا لم توجد صلاحيات مخصصة، نرجع الصلاحيات الافتراضية للدور
    if not scopes:
        default_perms = DEFAULT_PERMISSIONS.get(user["role"], [])
        scopes.append({
            "id": None,
            "scope_type": "global",
            "scope_id": None,
            "scope_name": "افتراضي (حسب الدور)",
            "permissions": default_perms
        })
    
    return {
        "user_id": user_id,
        "user_name": user["full_name"],
        "role": user["role"],
        "scopes": scopes
    }

@api_router.post("/users/{user_id}/permissions")
async def add_user_permission(
    user_id: str, 
    permission_data: UserPermissionScope,
    current_user: dict = Depends(get_current_user)
):
    """إضافة صلاحية لمستخدم"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_users"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل الصلاحيات")
    
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # الحصول على اسم النطاق
    scope_name = ""
    if permission_data.scope_type == "department" and permission_data.scope_id:
        dept = await db.departments.find_one({"_id": ObjectId(permission_data.scope_id)})
        scope_name = dept["name"] if dept else ""
    elif permission_data.scope_type == "course" and permission_data.scope_id:
        course = await db.courses.find_one({"_id": ObjectId(permission_data.scope_id)})
        scope_name = course["name"] if course else ""
    
    # إنشاء سجل الصلاحية
    perm_doc = {
        "user_id": user_id,
        "scope_type": permission_data.scope_type,
        "scope_id": permission_data.scope_id,
        "scope_name": scope_name,
        "permissions": permission_data.permissions,
        "created_at": get_yemen_time(),
        "created_by": current_user["id"]
    }
    
    # التحقق من عدم وجود تكرار
    existing = await db.user_permissions.find_one({
        "user_id": user_id,
        "scope_type": permission_data.scope_type,
        "scope_id": permission_data.scope_id
    })
    
    if existing:
        # تحديث الصلاحيات الموجودة
        await db.user_permissions.update_one(
            {"_id": existing["_id"]},
            {"$set": {"permissions": permission_data.permissions, "updated_at": get_yemen_time()}}
        )
        return {"message": "تم تحديث الصلاحيات بنجاح", "id": str(existing["_id"])}
    else:
        result = await db.user_permissions.insert_one(perm_doc)
        return {"message": "تمت إضافة الصلاحية بنجاح", "id": str(result.inserted_id)}

@api_router.delete("/users/{user_id}/permissions/{permission_id}")
async def delete_user_permission(
    user_id: str, 
    permission_id: str,
    current_user: dict = Depends(get_current_user)
):
    """حذف صلاحية من مستخدم"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_users"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بحذف الصلاحيات")
    
    result = await db.user_permissions.delete_one({
        "_id": ObjectId(permission_id),
        "user_id": user_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="الصلاحية غير موجودة")
    
    return {"message": "تم حذف الصلاحية بنجاح"}

@api_router.put("/users/{user_id}/permissions")
async def update_all_user_permissions(
    user_id: str,
    permissions_data: UserPermissionsUpdate,
    current_user: dict = Depends(get_current_user)
):
    """تحديث جميع صلاحيات المستخدم"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_users"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل الصلاحيات")
    
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # حذف الصلاحيات القديمة
    await db.user_permissions.delete_many({"user_id": user_id})
    
    # إضافة الصلاحيات الجديدة
    for scope in permissions_data.scopes:
        scope_name = ""
        if scope.scope_type == "department" and scope.scope_id:
            dept = await db.departments.find_one({"_id": ObjectId(scope.scope_id)})
            scope_name = dept["name"] if dept else ""
        elif scope.scope_type == "course" and scope.scope_id:
            course = await db.courses.find_one({"_id": ObjectId(scope.scope_id)})
            scope_name = course["name"] if course else ""
        
        perm_doc = {
            "user_id": user_id,
            "scope_type": scope.scope_type,
            "scope_id": scope.scope_id,
            "scope_name": scope_name,
            "permissions": scope.permissions,
            "created_at": get_yemen_time(),
            "created_by": current_user["id"]
        }
        await db.user_permissions.insert_one(perm_doc)
    
    return {"message": "تم تحديث الصلاحيات بنجاح"}

# دالة مساعدة للتحقق من صلاحية المستخدم على نطاق معين
async def check_user_permission(user_id: str, permission: str, scope_type: str = None, scope_id: str = None) -> bool:
    """التحقق من صلاحية المستخدم"""
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        return False
    
    # المدير لديه كل الصلاحيات
    if user["role"] == UserRole.ADMIN:
        return True
    
    # البحث عن صلاحيات مخصصة
    query = {"user_id": user_id}
    user_perms = await db.user_permissions.find(query).to_list(100)
    
    if user_perms:
        for perm in user_perms:
            # صلاحية عامة
            if perm["scope_type"] == "global" and permission in perm["permissions"]:
                return True
            # صلاحية على نطاق معين
            if scope_type and scope_id:
                if perm["scope_type"] == scope_type and perm.get("scope_id") == scope_id:
                    if permission in perm["permissions"]:
                        return True
    else:
        # استخدام الصلاحيات الافتراضية للدور
        default_perms = DEFAULT_PERMISSIONS.get(user["role"], [])
        return permission in default_perms
    
    return False

# ==================== University APIs (واجهات الجامعة) ====================

@api_router.get("/university")
async def get_university(current_user: dict = Depends(get_current_user)):
    """جلب بيانات الجامعة"""
    cached = get_cached("university")
    if cached:
        return cached
    
    university = await db.university.find_one()
    if not university:
        return None
    
    faculties_count = await db.faculties.count_documents({})
    
    result = {
        "id": str(university["_id"]),
        "name": university.get("name", ""),
        "code": university.get("code", ""),
        "short_code": university.get("short_code"),
        "description": university.get("description", ""),
        "logo_url": university.get("logo_url"),
        "address": university.get("address"),
        "phone": university.get("phone"),
        "email": university.get("email"),
        "website": university.get("website"),
        "faculties_count": faculties_count,
        "created_at": university.get("created_at", get_yemen_time())
    }
    set_cached("university", result)
    return result

@api_router.post("/university")
async def create_or_update_university(
    university_data: UniversityCreate,
    current_user: dict = Depends(get_current_user)
):
    """إنشاء أو تحديث بيانات الجامعة"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_faculties"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل بيانات الجامعة")
    
    existing = await db.university.find_one()
    
    university_doc = {
        "name": university_data.name,
        "code": university_data.code,
        "short_code": university_data.short_code,
        "description": university_data.description,
        "logo_url": university_data.logo_url,
        "address": university_data.address,
        "phone": university_data.phone,
        "email": university_data.email,
        "website": university_data.website,
        "updated_at": get_yemen_time()
    }
    
    if existing:
        # حفظ logo_url القديم إذا لم يُرسل جديد
        if not university_data.logo_url and existing.get("logo_url"):
            university_doc["logo_url"] = existing["logo_url"]
        await db.university.update_one(
            {"_id": existing["_id"]},
            {"$set": university_doc}
        )
        university_id = str(existing["_id"])
        action = "update_university"
    else:
        university_doc["created_at"] = get_yemen_time()
        result = await db.university.insert_one(university_doc)
        university_id = str(result.inserted_id)
        action = "create_university"
    
    # تسجيل النشاط
    await log_activity(
        user=current_user,
        action=action,
        entity_type="university",
        entity_id=university_id,
        entity_name=university_data.name
    )
    
    clear_cache("university")
    return {"message": "تم حفظ بيانات الجامعة بنجاح", "id": university_id}


# ==================== File Upload/Download APIs ====================

ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/gif", "image/webp"}
MAX_IMAGE_SIZE = 5 * 1024 * 1024  # 5MB

@api_router.post("/upload/image")
async def upload_image(
    file: UploadFile = File(...),
    folder: str = "logos",
    current_user: dict = Depends(get_current_user)
):
    """رفع صورة"""
    if file.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(status_code=400, detail="نوع الملف غير مدعوم. الأنواع المدعومة: JPEG, PNG, GIF, WebP")
    
    data = await file.read()
    if len(data) > MAX_IMAGE_SIZE:
        raise HTTPException(status_code=400, detail="حجم الملف يتجاوز الحد الأقصى (5MB)")
    
    try:
        from services.storage_service import upload_file
        result = upload_file(data, file.filename, file.content_type, folder)
        
        # حفظ مرجع الملف في قاعدة البيانات
        file_doc = {
            "file_id": result["file_id"],
            "storage_path": result["storage_path"],
            "original_filename": result["original_filename"],
            "content_type": result["content_type"],
            "size": result["size"],
            "folder": folder,
            "uploaded_by": current_user["id"],
            "is_deleted": False,
            "created_at": get_yemen_time().isoformat()
        }
        await db.files.insert_one(file_doc)
        
        return {
            "storage_path": result["storage_path"],
            "file_id": result["file_id"],
            "message": "تم رفع الصورة بنجاح"
        }
    except Exception as e:
        logging.error(f"Upload error: {e}")
        raise HTTPException(status_code=500, detail="فشل رفع الصورة")


@api_router.get("/files/{path:path}")
async def serve_file(path: str, auth: str = Query(None)):
    """تحميل ملف من التخزين"""
    # التحقق من التوكن عبر query param
    if not auth:
        raise HTTPException(status_code=401, detail="مطلوب مصادقة")
    try:
        payload = jwt.decode(auth, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status_code=401, detail="توكن غير صالح")
    except JWTError:
        raise HTTPException(status_code=401, detail="توكن غير صالح")
    
    try:
        from services.storage_service import get_object
        data, content_type = get_object(path)
        
        record = await db.files.find_one({"storage_path": path, "is_deleted": False})
        if record:
            content_type = record.get("content_type", content_type)
        
        return Response(content=data, media_type=content_type)
    except Exception as e:
        logging.error(f"File serve error: {e}")
        raise HTTPException(status_code=404, detail="الملف غير موجود")

# ==================== Faculty APIs (واجهات الكليات) ====================

@api_router.get("/faculties")
async def get_faculties(current_user: dict = Depends(get_current_user)):
    """جلب قائمة الكليات — مع تطبيق فلتر النطاق حسب دور المستخدم.

    🔒 RBAC: كل مستخدم يرى فقط كليته (أو الكليات المرتبطة بأقسامه).
    - admin: كل الكليات
    - dean / registrar / registration_manager: كليته فقط (faculty_id)
    - department_head: كلية قسمه (lookup عبر department.faculty_id)
    - custom role: كليته (إن وُجد faculty_id) أو كليات أقسامه
    - teacher / student: كل الكليات (للقراءة فقط في dropdowns العامة)
    """
    query: dict = {}
    role = current_user.get("role", "")

    if role != UserRole.ADMIN:
        # نحتاج faculty_id (أو نشتقّه من أقسام المستخدم)
        user_data = await db.users.find_one({"_id": ObjectId(current_user["id"])})
        if user_data:
            allowed_faculty_ids: set = set()
            # 1) المستخدم لديه faculty_id صريح (عميد/مسجِّل/custom-faculty)
            if user_data.get("faculty_id"):
                allowed_faculty_ids.add(user_data["faculty_id"])
            # 2) المستخدم لديه أقسام → نشتقّ كلياتها
            dept_ids = user_data.get("department_ids") or []
            if not dept_ids and user_data.get("department_id"):
                dept_ids = [user_data["department_id"]]
            if dept_ids:
                try:
                    obj_ids = [ObjectId(d) for d in dept_ids if d]
                    async for d in db.departments.find({"_id": {"$in": obj_ids}}, {"faculty_id": 1}):
                        if d.get("faculty_id"):
                            allowed_faculty_ids.add(d["faculty_id"])
                except Exception:
                    pass

            # تطبيق الفلتر فقط إن كان المستخدم في دور مقيَّد بكلية
            scoped_roles = {"dean", "department_head", "registration_manager", "registrar", "custom"}
            if role in scoped_roles and allowed_faculty_ids:
                try:
                    query["_id"] = {"$in": [ObjectId(fid) for fid in allowed_faculty_ids]}
                except Exception:
                    pass
            elif role in scoped_roles and not allowed_faculty_ids:
                # دور مقيَّد لكن بلا بيانات → لا يرى شيئاً (بدلاً من رؤية الكل)
                return []

    faculties = await db.faculties.find(query).to_list(100)
    result = []
    
    for faculty in faculties:
        departments_count = await db.departments.count_documents({"faculty_id": str(faculty["_id"])})
        dean_name = None
        if faculty.get("dean_id"):
            dean = await db.users.find_one({"_id": ObjectId(faculty["dean_id"])})
            dean_name = dean.get("full_name") if dean else None
        
        result.append({
            "id": str(faculty["_id"]),
            "name": faculty["name"],
            "code": faculty["code"],
            "numeric_code": faculty.get("numeric_code"),
            "description": faculty.get("description", ""),
            "dean_id": faculty.get("dean_id"),
            "dean_name": dean_name,
            "departments_count": departments_count,
            "created_at": faculty.get("created_at", get_yemen_time())
        })
    
    return result

@api_router.get("/faculties/{faculty_id}")
async def get_faculty(faculty_id: str, current_user: dict = Depends(get_current_user)):
    """جلب بيانات كلية محددة مع إعداداتها"""
    faculty = await db.faculties.find_one({"_id": ObjectId(faculty_id)})
    if not faculty:
        raise HTTPException(status_code=404, detail="الكلية غير موجودة")
    
    departments_count = await db.departments.count_documents({"faculty_id": faculty_id})
    dean_name = None
    if faculty.get("dean_id"):
        dean = await db.users.find_one({"_id": ObjectId(faculty["dean_id"])})
        dean_name = dean.get("full_name") if dean else None
    
    return {
        "id": str(faculty["_id"]),
        "name": faculty["name"],
        "code": faculty["code"],
        "description": faculty.get("description", ""),
        "dean_id": faculty.get("dean_id"),
        "dean_name": dean_name,
        "departments_count": departments_count,
        "created_at": faculty.get("created_at", get_yemen_time()),
        # إعدادات الكلية
        "levels_count": faculty.get("levels_count", 5),
        "sections": faculty.get("sections", ["أ", "ب", "ج"]),
        "attendance_late_minutes": faculty.get("attendance_late_minutes", 15),
        "max_absence_percent": faculty.get("max_absence_percent", 25),
        "primary_color": faculty.get("primary_color", "#1565c0"),
        "secondary_color": faculty.get("secondary_color", "#ff9800"),
        # معلومات التواصل
        "phone": faculty.get("phone", ""),
        "whatsapp": faculty.get("whatsapp", ""),
        "email": faculty.get("email", ""),
    }

@api_router.put("/faculties/{faculty_id}/settings")
async def update_faculty_settings(
    faculty_id: str,
    data: dict = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """تحديث إعدادات كلية"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_faculties"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل إعدادات الكليات")
    
    faculty = await db.faculties.find_one({"_id": ObjectId(faculty_id)})
    if not faculty:
        raise HTTPException(status_code=404, detail="الكلية غير موجودة")
    
    # الحقول المسموح تحديثها
    allowed_fields = [
        "levels_count", "sections", "attendance_late_minutes", "max_absence_percent",
        "primary_color", "secondary_color", "phone", "whatsapp", "email",
        "attendance_duration_minutes", "max_attendance_delay_minutes"
    ]
    
    update_data = {k: v for k, v in data.items() if k in allowed_fields and v is not None}
    
    if update_data:
        update_data["updated_at"] = get_yemen_time()
        await db.faculties.update_one(
            {"_id": ObjectId(faculty_id)},
            {"$set": update_data}
        )
    
    # تسجيل النشاط
    await log_activity(
        user=current_user,
        action="update_faculty_settings",
        entity_type="faculty",
        entity_id=faculty_id,
        entity_name=faculty["name"]
    )
    
    return {"message": "تم تحديث إعدادات الكلية بنجاح"}

@api_router.post("/faculties")
async def create_faculty(
    faculty_data: FacultyCreate,
    current_user: dict = Depends(get_current_user)
):
    """إنشاء كلية جديدة"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_faculties"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بإنشاء كليات")
    
    # التحقق من عدم وجود كلية بنفس الكود
    existing = await db.faculties.find_one({"code": faculty_data.code})
    if existing:
        raise HTTPException(status_code=400, detail="يوجد كلية بنفس الكود")
    
    faculty_doc = {
        "name": faculty_data.name,
        "code": faculty_data.code,
        "description": faculty_data.description,
        "dean_id": faculty_data.dean_id,
        "created_at": get_yemen_time()
    }
    
    result = await db.faculties.insert_one(faculty_doc)
    faculty_id = str(result.inserted_id)
    
    # تسجيل النشاط
    await log_activity(
        user=current_user,
        action="create_faculty",
        entity_type="faculty",
        entity_id=faculty_id,
        entity_name=faculty_data.name
    )
    
    return {"message": "تم إنشاء الكلية بنجاح", "id": faculty_id}

@api_router.put("/faculties/{faculty_id}")
async def update_faculty(
    faculty_id: str,
    faculty_data: FacultyUpdate,
    current_user: dict = Depends(get_current_user)
):
    """تحديث بيانات كلية"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_faculties"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل الكليات")
    
    faculty = await db.faculties.find_one({"_id": ObjectId(faculty_id)})
    if not faculty:
        raise HTTPException(status_code=404, detail="الكلية غير موجودة")
    
    update_data = {k: v for k, v in faculty_data.dict().items() if v is not None}
    update_data["updated_at"] = get_yemen_time()
    
    await db.faculties.update_one(
        {"_id": ObjectId(faculty_id)},
        {"$set": update_data}
    )
    
    # تسجيل النشاط
    await log_activity(
        user=current_user,
        action="update_faculty",
        entity_type="faculty",
        entity_id=faculty_id,
        entity_name=faculty.get("name")
    )
    
    return {"message": "تم تحديث الكلية بنجاح"}

@api_router.delete("/faculties/{faculty_id}")
async def delete_faculty(
    faculty_id: str,
    current_user: dict = Depends(get_current_user)
):
    """حذف كلية"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_faculties"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بحذف الكليات")
    
    faculty = await db.faculties.find_one({"_id": ObjectId(faculty_id)})
    if not faculty:
        raise HTTPException(status_code=404, detail="الكلية غير موجودة")
    
    # التحقق من عدم وجود أقسام مرتبطة
    departments_count = await db.departments.count_documents({"faculty_id": faculty_id})
    if departments_count > 0:
        raise HTTPException(status_code=400, detail=f"لا يمكن حذف الكلية - يوجد {departments_count} قسم مرتبط بها")
    
    await db.faculties.delete_one({"_id": ObjectId(faculty_id)})
    
    # تسجيل النشاط
    await log_activity(
        user=current_user,
        action="delete_faculty",
        entity_type="faculty",
        entity_id=faculty_id,
        entity_name=faculty.get("name")
    )
    
    return {"message": "تم حذف الكلية بنجاح"}


@api_router.post("/admin/fix-courses-semester")
async def fix_courses_without_semester(current_user: dict = Depends(get_current_user)):
    """إصلاح المقررات التي ليس لها فصل دراسي بربطها بالفصل النشط"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    settings = await db.settings.find_one({"_id": "system_settings"})
    semester_id = settings.get("current_semester_id") if settings else None
    
    if not semester_id:
        active_sem = await db.semesters.find_one({"status": "active"})
        if active_sem:
            semester_id = str(active_sem["_id"])
    
    if not semester_id:
        raise HTTPException(status_code=400, detail="لا يوجد فصل دراسي نشط")
    
    result = await db.courses.update_many(
        {"$or": [{"semester_id": None}, {"semester_id": ""}, {"semester_id": {"$exists": False}}]},
        {"$set": {"semester_id": semester_id}}
    )
    
    return {
        "message": f"تم إصلاح {result.modified_count} مقرر وربطها بالفصل الدراسي النشط",
        "fixed": result.modified_count
    }

@api_router.post("/admin/fix-custom-roles")
async def fix_custom_roles(current_user: dict = Depends(get_current_user)):
    """إصلاح المستخدمين الذين لديهم role فارغ وتحويلهم إلى custom"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # إصلاح المستخدمين بدور فارغ
    users_result = await db.users.update_many(
        {"role": ""},
        {"$set": {"role": "custom"}}
    )
    
    return {
        "message": f"تم إصلاح {users_result.modified_count} مستخدم (role فارغ → custom)",
        "users_fixed": users_result.modified_count
    }


@api_router.post("/admin/cleanup-duplicate-roles")
async def cleanup_duplicate_roles(current_user: dict = Depends(get_current_user)):
    """حذف الأدوار المكررة - الإبقاء على الأقدم فقط"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # جمع كل الأدوار مع system_key
    all_roles = await db.roles.find().to_list(1000)
    
    # تجميع حسب الاسم
    name_groups = {}
    for role in all_roles:
        name = role.get("name", "")
        if name not in name_groups:
            name_groups[name] = []
        name_groups[name].append(role)
    
    deleted = 0
    details = []
    for name, roles in name_groups.items():
        if len(roles) > 1:
            # الإبقاء على الأقدم (أو الذي لديه system_key)
            roles.sort(key=lambda r: (
                0 if r.get("system_key") and r["system_key"] not in [None, "", "custom"] else 1,
                r.get("created_at", datetime.min)
            ))
            keep = roles[0]
            for dup in roles[1:]:
                # نقل المستخدمين من الدور المكرر إلى الأصلي
                await db.users.update_many(
                    {"role_id": str(dup["_id"])},
                    {"$set": {"role_id": str(keep["_id"])}}
                )
                await db.roles.delete_one({"_id": dup["_id"]})
                deleted += 1
                details.append(f"حذف مكرر: {name} (ID: {dup['_id']})")
    
    return {
        "message": f"تم حذف {deleted} دور مكرر",
        "deleted": deleted,
        "details": details
    }




@api_router.post("/admin/fix-faculty-ids")
async def fix_faculty_ids(current_user: dict = Depends(get_current_user)):
    """إصلاح الطلاب والمعلمين الذين ليس لديهم faculty_id"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # بناء خريطة القسم → الكلية
    departments = await db.departments.find().to_list(100)
    dept_faculty_map = {}
    for dept in departments:
        dept_faculty_map[str(dept["_id"])] = dept.get("faculty_id")
    
    # إصلاح الطلاب
    students_fixed = 0
    students_cursor = db.students.find({"$or": [{"faculty_id": None}, {"faculty_id": ""}, {"faculty_id": {"$exists": False}}]})
    async for student in students_cursor:
        dept_id = student.get("department_id")
        if dept_id and dept_id in dept_faculty_map:
            await db.students.update_one(
                {"_id": student["_id"]},
                {"$set": {"faculty_id": dept_faculty_map[dept_id]}}
            )
            students_fixed += 1
    
    # إصلاح المعلمين
    teachers_fixed = 0
    teachers_cursor = db.teachers.find({"$or": [{"faculty_id": None}, {"faculty_id": ""}, {"faculty_id": {"$exists": False}}]})
    async for teacher in teachers_cursor:
        dept_id = teacher.get("department_id")
        if dept_id and dept_id in dept_faculty_map:
            await db.teachers.update_one(
                {"_id": teacher["_id"]},
                {"$set": {"faculty_id": dept_faculty_map[dept_id]}}
            )
            teachers_fixed += 1
    
    return {
        "message": f"تم إصلاح {students_fixed} طالب و {teachers_fixed} معلم",
        "students_fixed": students_fixed,
        "teachers_fixed": teachers_fixed
    }

@api_router.post("/admin/fix-student-sections")
async def fix_student_sections(current_user: dict = Depends(get_current_user)):
    """إصلاح شعب الطلاب بناءً على تسجيلاتهم الحالية"""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # بناء خريطة المقررات (course_id → section)
    courses = await db.courses.find({"section": {"$exists": True, "$ne": ""}}).to_list(10000)
    course_section_map = {str(c["_id"]): c["section"] for c in courses}
    
    # جلب كل التسجيلات
    enrollments = await db.enrollments.find().to_list(100000)
    
    # بناء خريطة: student_id → أحدث شعبة (بناءً على آخر تسجيل)
    student_latest = {}
    for e in enrollments:
        sid = e["student_id"]
        cid = e["course_id"]
        section = course_section_map.get(cid)
        if section:
            enrolled_at = e.get("enrolled_at", "")
            if sid not in student_latest or str(enrolled_at) > str(student_latest[sid][1]):
                student_latest[sid] = (section, enrolled_at)
    
    # تحديث شعب الطلاب
    fixed = 0
    for sid, (section, _) in student_latest.items():
        try:
            result = await db.students.update_one(
                {"_id": ObjectId(sid), "section": {"$ne": section}},
                {"$set": {"section": section}}
            )
            if result.modified_count > 0:
                fixed += 1
        except:
            pass
    
    return {
        "message": f"تم تحديث شعبة {fixed} طالب",
        "students_fixed": fixed,
        "total_checked": len(student_latest)
    }



# ==================== Activity Logs APIs (واجهات سجل الأنشطة) ====================

@api_router.get("/activity-logs")
async def get_activity_logs(
    page: int = 1,
    limit: int = 50,
    user_id: Optional[str] = None,
    action: Optional[str] = None,
    entity_type: Optional[str] = None,
    faculty_id: Optional[str] = None,
    department_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب سجلات الأنشطة"""
    # التحقق من الصلاحية
    if current_user["role"] not in [UserRole.ADMIN, "dean", "department_head"]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بعرض سجلات الأنشطة")
    
    query = {}
    
    # فلترة حسب الصلاحيات
    if current_user["role"] == "dean" and current_user.get("faculty_id"):
        query["faculty_id"] = current_user["faculty_id"]
    elif current_user["role"] == "department_head" and current_user.get("department_id"):
        query["department_id"] = current_user["department_id"]
    
    # فلترة إضافية
    if user_id:
        query["user_id"] = user_id
    if action:
        query["action"] = action
    if entity_type:
        query["entity_type"] = entity_type
    if faculty_id and current_user["role"] == UserRole.ADMIN:
        query["faculty_id"] = faculty_id
    if department_id:
        query["department_id"] = department_id
    
    # فلترة بالتاريخ
    if from_date or to_date:
        query["timestamp"] = {}
        if from_date:
            query["timestamp"]["$gte"] = datetime.fromisoformat(from_date)
        if to_date:
            query["timestamp"]["$lte"] = datetime.fromisoformat(to_date)
    
    # حساب العدد الكلي
    total = await db.activity_logs.count_documents(query)
    
    # جلب السجلات
    skip = (page - 1) * limit
    logs = await db.activity_logs.find(query).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    
    result = []
    for log in logs:
        result.append({
            "id": str(log["_id"]),
            "user_id": log.get("user_id"),
            "username": log.get("username"),
            "user_role": log.get("user_role"),
            "action": log.get("action"),
            "action_ar": log.get("action_ar"),
            "entity_type": log.get("entity_type"),
            "entity_id": log.get("entity_id"),
            "entity_name": log.get("entity_name"),
            "details": log.get("details"),
            "ip_address": log.get("ip_address"),
            "faculty_id": log.get("faculty_id"),
            "department_id": log.get("department_id"),
            "timestamp": log.get("timestamp")
        })
    
    return {
        "logs": result,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }

@api_router.get("/activity-logs/stats")
async def get_activity_logs_stats(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """إحصائيات سجلات الأنشطة"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "view_reports"):
        raise HTTPException(status_code=403, detail="غير مصرح لك بعرض الإحصائيات")
    
    query = {}
    if from_date or to_date:
        query["timestamp"] = {}
        if from_date:
            query["timestamp"]["$gte"] = datetime.fromisoformat(from_date)
        if to_date:
            query["timestamp"]["$lte"] = datetime.fromisoformat(to_date)
    
    # إحصائيات حسب نوع النشاط
    pipeline = [
        {"$match": query},
        {"$group": {"_id": "$action", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    
    by_action = await db.activity_logs.aggregate(pipeline).to_list(100)
    
    # إحصائيات حسب المستخدم
    pipeline = [
        {"$match": query},
        {"$group": {"_id": "$username", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    
    by_user = await db.activity_logs.aggregate(pipeline).to_list(10)
    
    # إحصائيات حسب اليوم
    pipeline = [
        {"$match": query},
        {
            "$group": {
                "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$timestamp"}},
                "count": {"$sum": 1}
            }
        },
        {"$sort": {"_id": -1}},
        {"$limit": 30}
    ]
    
    by_day = await db.activity_logs.aggregate(pipeline).to_list(30)
    
    total = await db.activity_logs.count_documents(query)
    
    return {
        "total": total,
        "by_action": [{"action": a["_id"], "action_ar": ACTION_TRANSLATIONS.get(a["_id"], a["_id"]), "count": a["count"]} for a in by_action],
        "by_user": [{"username": u["_id"], "count": u["count"]} for u in by_user],
        "by_day": [{"date": d["_id"], "count": d["count"]} for d in by_day]
    }

@api_router.post("/activity-logs/record-view")
async def record_page_view(
    page_name: str = Body(...),
    page_path: str = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """تسجيل مشاهدة صفحة"""
    await log_activity(
        user=current_user,
        action="view_page",
        entity_type="page",
        entity_name=page_name,
        details={"path": page_path}
    )
    return {"message": "تم التسجيل"}

# ==================== Trash Management (سلة المحذوفات) ====================

@api_router.get("/trash")
async def get_trash_items(current_user: dict = Depends(get_current_user)):
    """عرض جميع العناصر في سلة المحذوفات"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses") and not has_permission(current_user, "manage_students") and not has_permission(current_user, "manage_teachers"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    now = get_yemen_time()
    now_naive = now.replace(tzinfo=None) if now.tzinfo else now
    # حذف العناصر المنتهية الصلاحية
    await db.trash.delete_many({"expires_at": {"$lt": now_naive}})
    
    items = await db.trash.find().sort("deleted_at", -1).to_list(500)
    result = []
    for item in items:
        exp = item["expires_at"]
        now_naive = now.replace(tzinfo=None) if now.tzinfo else now
        exp_naive = exp.replace(tzinfo=None) if hasattr(exp, 'tzinfo') and exp.tzinfo else exp
        days_remaining = (exp_naive - now_naive).days
        result.append({
            "id": str(item["_id"]),
            "item_type": item["item_type"],
            "item_name": item["item_name"],
            "deleted_by": item.get("deleted_by", ""),
            "deleted_at": item["deleted_at"].isoformat() if hasattr(item["deleted_at"], 'isoformat') else str(item["deleted_at"]),
            "expires_at": exp.isoformat() if hasattr(exp, 'isoformat') else str(exp),
            "days_remaining": max(0, days_remaining),
        })
    return result

@api_router.post("/trash/{trash_id}/restore")
async def restore_trash_item(trash_id: str, current_user: dict = Depends(get_current_user)):
    """استعادة عنصر من سلة المحذوفات"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses") and not has_permission(current_user, "manage_students") and not has_permission(current_user, "manage_teachers"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    trash_item = await db.trash.find_one({"_id": ObjectId(trash_id)})
    if not trash_item:
        raise HTTPException(status_code=404, detail="العنصر غير موجود في سلة المحذوفات")
    
    result = await restore_from_trash_helper(trash_item["backup_data"])
    
    # حذف من السلة بعد الاستعادة
    await db.trash.delete_one({"_id": ObjectId(trash_id)})
    
    return result

@api_router.delete("/trash/{trash_id}")
async def permanent_delete_trash_item(trash_id: str, current_user: dict = Depends(get_current_user)):
    """حذف نهائي لعنصر من سلة المحذوفات"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses") and not has_permission(current_user, "manage_students") and not has_permission(current_user, "manage_teachers"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    result = await db.trash.delete_one({"_id": ObjectId(trash_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="العنصر غير موجود")
    
    return {"message": "تم الحذف النهائي بنجاح"}

@api_router.delete("/trash")
async def clear_all_trash(current_user: dict = Depends(get_current_user)):
    """تفريغ سلة المحذوفات بالكامل"""
    if current_user["role"] != UserRole.ADMIN and not has_permission(current_user, "manage_courses") and not has_permission(current_user, "manage_students") and not has_permission(current_user, "manage_teachers"):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    result = await db.trash.delete_many({})
    return {"message": f"تم تفريغ سلة المحذوفات ({result.deleted_count} عنصر)"}


@api_router.get("/admin/diagnose-roles")
async def diagnose_roles(current_user: dict = Depends(get_current_user)):
    """تشخيص شامل: يُظهر كل الأدوار، عدد المستخدمين، وأي تكرارات أو عدم تطابق."""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح لك — admin فقط")
    all_roles = await db.roles.find().to_list(1000)
    all_users = await db.users.find().to_list(2000)
    roles_summary = []
    role_name_count = {}
    for r in all_roles:
        rid = str(r["_id"])
        name = (r.get("name") or "").strip()
        sys_key = (r.get("system_key") or "").strip()
        role_name_count[name] = role_name_count.get(name, 0) + 1
        users_by_id = [u for u in all_users if str(u.get("role_id") or "") == rid]
        users_by_role = [u for u in all_users if sys_key and u.get("role") == sys_key]
        inconsistent = [u for u in users_by_id if sys_key and u.get("role") != sys_key]
        roles_summary.append({
            "role_id": rid, "name": name,
            "system_key": sys_key or "(none — custom)",
            "users_with_role_id_matching": len(users_by_id),
            "users_with_role_string_matching": len(users_by_role),
            "inconsistent_users": [
                {"username": u.get("username"), "role": u.get("role"), "role_id": u.get("role_id")}
                for u in inconsistent
            ],
        })
    duplicates = {n: c for n, c in role_name_count.items() if c > 1}
    orphan_users = [
        {"username": u.get("username"), "role": u.get("role"), "role_id": u.get("role_id")}
        for u in all_users
        if u.get("role") not in ["admin", "dean", "department_head", "registrar",
                                  "registration_manager", "employee", "teacher", "student"]
    ]
    return {
        "total_roles": len(all_roles), "total_users": len(all_users),
        "duplicate_role_names": duplicates,
        "users_with_non_standard_role": orphan_users,
        "roles_detail": roles_summary,
    }


@api_router.post("/admin/cleanup-duplicate-roles-now")
async def cleanup_duplicate_roles_now(current_user: dict = Depends(get_current_user)):
    """يُشغل دمج الأدوار المكررة وإصلاح المستخدمين فوراً."""
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح لك — admin فقط")
    await cleanup_duplicate_roles_internal()
    await migrate_broken_user_roles()
    return {"message": "تم تنفيذ دمج الأدوار المكررة وإصلاح المستخدمين. أعد تحميل الصفحة."}


# Include the router in the main app
# إضافة الـ routers المنفصلة (المرحلة 2 من إعادة الهيكلة)
# ملاحظة: api_router يجب أن يُضاف أولاً لأنه يحتوي على routes محددة مثل /departments/dashboard
app.include_router(api_router)

app.include_router(auth_router, prefix="/api")
app.include_router(users_router, prefix="/api")
app.include_router(roles_router, prefix="/api")
app.include_router(departments_router, prefix="/api")
app.include_router(students_router, prefix="/api")
app.include_router(teachers_router, prefix="/api")
app.include_router(courses_router, prefix="/api")
app.include_router(notifications_router, prefix="/api")
app.include_router(teaching_load_router, prefix="/api")
app.include_router(attendance_approval_router, prefix="/api")
app.include_router(weekly_schedule_router, prefix="/api")
app.include_router(study_plans_router, prefix="/api")
app.include_router(admin_tools_router, prefix="/api")
app.include_router(dashboard_router, prefix="/api")
app.include_router(calendar_router, prefix="/api")
app.include_router(global_search_router, prefix="/api")
app.include_router(alumni_router, prefix="/api")
app.include_router(entity_details_router, prefix="/api")
app.include_router(archives_router, prefix="/api")
app.include_router(archive_pdf_router, prefix="/api")
app.include_router(curriculum_router, prefix="/api")
app.include_router(student_status_router, prefix="/api")
app.include_router(student_transfer_router, prefix="/api")

@app.on_event("startup")
async def startup_event():
    from services.firebase_service import init_firebase
    init_firebase()
    # تهيئة خدمة التخزين
    try:
        from services.storage_service import init_storage
        init_storage()
        logging.info("Object storage initialized successfully")
    except Exception as e:
        logging.warning(f"Object storage init failed (non-critical): {e}")
    # إنشاء فهارس MongoDB لتسريع الاستعلامات
    await create_indexes()
    # تحديث صلاحيات الأدوار الافتراضية تلقائياً
    await sync_default_roles()
    # 🔧 Migration لمرة واحدة: استعادة صلاحيات المحاضرات للأدوار التي فقدتها بالخطأ
    await restore_lecture_permissions_v1()
    # 🔧 تنظيف الصلاحيات اليتيمة (التي أُلغيت من النظام) من الأدوار والمستخدمين
    await cleanup_orphan_permissions()
    # 🔧 دمج الأدوار المكررة (مثل عدة "رئيس قسم") قبل مزامنة role
    await cleanup_duplicate_roles_internal()
    # 🔧 Migration دائمة: إصلاح المستخدمين ذوي الأدوار الشاذة تلقائياً
    await migrate_broken_user_roles()
    # 🔧 Migration دائمة: تعبئة semester_id لسجلات teaching_loads المعطّلة
    # يحلّ مشكلة عدم ظهور الإسنادات في صفحة العبء التدريسي
    await backfill_teaching_loads_semester_internal()
    # 🧹 تنظيف تلقائي: حذف teaching_loads اليتيمة (مقرر محذوف أو is_active=False)
    # يحلّ مشكلة الصفوف الفارغة بلا اسم في PDF/Excel
    await cleanup_orphan_teaching_loads_internal()
    # 🧹 إزالة تلقائية للمكررات: نفس (teacher, course, semester) لها سجلات متعددة
    # نتيجة لإصلاحات سابقة دمجت loads بـ semester_id=null مع loads بـ semester_id=active
    await dedup_teaching_loads_internal()
    # 🎓 تعبئة snapshot القسم/الكلية للخريجين الحاليين (يعمل مرة واحدة عملياً)
    await backfill_alumni_department_snapshot_internal()


async def backfill_alumni_department_snapshot_internal():
    """يحفظ snapshot اسم القسم/الكلية داخل graduation_data.department_snapshot
    لكل خريج لا يملك هذا الحقل بعد. يمنع تحوّل عرض القسم لاحقاً بسبب
    تغيير department_id للطالب أو إعادة تسمية القسم.
    """
    try:
        # جلب مرة واحدة كل الأقسام والكليات لأداء أسرع
        depts_map: dict = {}
        async for d in db.departments.find({}):
            depts_map[str(d["_id"])] = {
                "name": (d.get("name") or "").strip(),
                "faculty_id": str(d.get("faculty_id") or "") or None,
            }
        facs_map: dict = {}
        async for f in db.faculties.find({}):
            facs_map[str(f["_id"])] = (f.get("name") or "").strip()

        cursor = db.students.find({
            "is_alumni": True,
            "graduation_data.department_snapshot": {"$exists": False},
        })
        fixed = 0
        async for s in cursor:
            dept_id = str(s.get("department_id") or "") or None
            fac_id = str(s.get("faculty_id") or "") or None
            dept_name = None
            fac_name = None
            if dept_id and dept_id in depts_map:
                dept_name = depts_map[dept_id]["name"]
                if not fac_id:
                    fac_id = depts_map[dept_id]["faculty_id"]
            if fac_id and fac_id in facs_map:
                fac_name = facs_map[fac_id]
            snapshot = {
                "department_id": dept_id,
                "department_name": dept_name,
                "faculty_id": fac_id,
                "faculty_name": fac_name,
                "backfilled": True,
            }
            await db.students.update_one(
                {"_id": s["_id"]},
                {"$set": {"graduation_data.department_snapshot": snapshot}}
            )
            fixed += 1
        if fixed > 0:
            logging.info(f"Alumni department snapshot backfill: added to {fixed} records")
    except Exception as e:
        logging.warning(f"Alumni department snapshot backfill failed (non-critical): {e}")


async def dedup_teaching_loads_internal():
    """🧹 إزالة المكررات في teaching_loads: نفس (teacher_id, course_id, semester_id) له أكثر من سجل.
    
    يحتفظ بأحدث سجل (آخر updated_at أو آخر created_at) ويحذف الباقي.
    آمن للتشغيل المتكرّر — لا يلمس السجلات الفريدة.
    """
    try:
        pipeline = [
            {"$group": {
                "_id": {
                    "teacher_id": "$teacher_id",
                    "course_id": "$course_id",
                    "semester_id": "$semester_id",
                },
                "ids": {"$push": "$_id"},
                "count": {"$sum": 1},
            }},
            {"$match": {"count": {"$gt": 1}}},
        ]
        deleted = 0
        async for group in db.teaching_loads.aggregate(pipeline):
            ids = group["ids"]
            # احتفظ بالأحدث (أكبر ObjectId timestamp) واحذف الباقي
            ids_sorted = sorted(ids, reverse=True)
            keep_id = ids_sorted[0]
            for old_id in ids_sorted[1:]:
                await db.teaching_loads.delete_one({"_id": old_id})
                deleted += 1
            # احتياط: تأكد أن الباقي له updated_at حديث
            await db.teaching_loads.update_one(
                {"_id": keep_id},
                {"$set": {"updated_at": datetime.now(timezone.utc)}}
            )
        if deleted > 0:
            logging.info(f"Teaching loads dedup: removed {deleted} duplicate records")
    except Exception as e:
        logging.warning(f"Teaching loads dedup failed (non-critical): {e}")


async def cleanup_orphan_teaching_loads_internal():
    """🧹 إصلاح تلقائي عند بدء التشغيل: حذف teaching_loads اليتيمة.
    
    يحذف السجلات التي:
    1. تشير لـ course_id لم يعد موجوداً (مقرر محذوف)
    2. تشير لمقرر بـ is_active=False
    
    هذه السجلات هي السبب في الصفوف الفارغة بلا اسم في تقارير PDF/Excel،
    وفي ظهور مقررات شبحية في صفحة العبء التدريسي.
    
    آمن للتشغيل المتكرّر — لا يلمس السجلات السليمة.
    """
    try:
        all_loads = await db.teaching_loads.find({}).to_list(50000)
        if not all_loads:
            return
        deleted_missing = 0
        deleted_inactive = 0
        for load in all_loads:
            try:
                course = await db.courses.find_one({"_id": ObjectId(load["course_id"])})
                if not course:
                    await db.teaching_loads.delete_one({"_id": load["_id"]})
                    deleted_missing += 1
                elif course.get("is_active") is False:
                    await db.teaching_loads.delete_one({"_id": load["_id"]})
                    deleted_inactive += 1
            except Exception:
                # ObjectId غير صالح → احذفه
                try:
                    await db.teaching_loads.delete_one({"_id": load["_id"]})
                    deleted_missing += 1
                except Exception:
                    pass
        if deleted_missing > 0 or deleted_inactive > 0:
            logging.info(
                f"Teaching loads orphan cleanup: deleted {deleted_missing} missing-course, "
                f"{deleted_inactive} inactive-course records"
            )
    except Exception as e:
        logging.warning(f"Teaching loads orphan cleanup failed (non-critical): {e}")


async def backfill_teaching_loads_semester_internal():
    """🔧 إصلاح تلقائي عند بدء التشغيل: يضمن أن teaching_loads.semester_id يطابق course.semester_id.
    
    يصلح حالتين:
    1. السجلات بدون semester_id → يأخذ semester_id من المقرر (وليس الفصل النشط - لتجنّب إظهار سجلات قديمة في الفصل الحالي)
    2. السجلات بـ semester_id خاطئ (لا يطابق course.semester_id) → يصححها
    
    آمن للتشغيل المتكرّر — يفلتر فقط السجلات التي تحتاج إصلاح.
    """
    try:
        # كل السجلات (لأننا نريد التحقق من تطابق semester_id مع course.semester_id)
        all_loads = await db.teaching_loads.find({}).to_list(50000)
        if not all_loads:
            return
        
        fixed = 0
        reset_to_orphan = 0
        for load in all_loads:
            try:
                course = await db.courses.find_one({"_id": ObjectId(load["course_id"])})
                if not course:
                    # المقرر محذوف - لا نلمس السجل (يبقى كما هو)
                    continue
                
                correct_sem_id = course.get("semester_id")
                current_sem_id = load.get("semester_id")
                
                # حالة 1: السجل بدون semester_id والمقرر له واحد → اضبطه
                # حالة 2: السجل بـ semester_id يختلف عن course.semester_id → صحّحه
                if correct_sem_id and current_sem_id != correct_sem_id:
                    await db.teaching_loads.update_one(
                        {"_id": load["_id"]},
                        {"$set": {"semester_id": correct_sem_id, "updated_at": datetime.now(timezone.utc)}}
                    )
                    fixed += 1
                elif not correct_sem_id and current_sem_id:
                    # المقرر بلا فصل لكن السجل عليه فصل → امسحه (تنظيف منطقي)
                    await db.teaching_loads.update_one(
                        {"_id": load["_id"]},
                        {"$set": {"semester_id": None, "updated_at": datetime.now(timezone.utc)}}
                    )
                    reset_to_orphan += 1
            except Exception:
                continue
        
        if fixed > 0 or reset_to_orphan > 0:
            logging.info(f"Teaching loads backfill: aligned {fixed} records, reset {reset_to_orphan} orphans")
    except Exception as e:
        logging.warning(f"Teaching loads backfill failed (non-critical): {e}")


async def restore_lecture_permissions_v1():
    """Migration لمرة واحدة: استعادة صلاحيات المحاضرات للأدوار الافتراضية
    التي فقدتها أثناء orphan cleanup عند تحويل manage_lectures إلى صلاحية مخفية.
    يعمل مرة واحدة فقط (محفوظ في meta.lectures_migration_v1).
    """
    try:
        meta = await db.meta.find_one({"_id": "lectures_migration_v1"})
        if meta and meta.get("done"):
            return  # تم التشغيل سابقاً

        LECTURE_PERMS = [
            Permission.MANAGE_LECTURES, Permission.VIEW_LECTURES,
            Permission.ADD_LECTURE, Permission.EDIT_LECTURE, Permission.DELETE_LECTURE,
            Permission.OVERRIDE_LECTURE_STATUS, Permission.RESCHEDULE_LECTURE,
            Permission.GENERATE_LECTURES,
        ]
        role_map = {
            "admin": UserRole.ADMIN, "teacher": UserRole.TEACHER,
            "employee": UserRole.EMPLOYEE, "dean": UserRole.DEAN,
            "department_head": UserRole.DEPARTMENT_HEAD,
        }
        restored = 0
        for system_key, role_enum in role_map.items():
            default_perms = DEFAULT_PERMISSIONS.get(role_enum, [])
            missing = [p for p in LECTURE_PERMS if p in default_perms]
            if not missing:
                continue
            role = await db.roles.find_one({"system_key": system_key})
            if not role:
                continue
            current = role.get("permissions") or []
            to_add = [p for p in missing if p not in current]
            if to_add:
                new_perms = list(dict.fromkeys(current + to_add))
                await db.roles.update_one({"_id": role["_id"]}, {"$set": {"permissions": new_perms}})
                restored += 1
                logging.info(f"✅ استعادة {len(to_add)} صلاحية محاضرات للدور {system_key}")

        await db.meta.update_one(
            {"_id": "lectures_migration_v1"},
            {"$set": {"done": True, "restored_roles": restored, "at": get_yemen_time()}},
            upsert=True,
        )
        if restored:
            logging.info(f"✅ Lectures migration v1: restored permissions in {restored} role(s).")
    except Exception as e:
        logging.error(f"Lectures migration v1 failed (non-critical): {e}")


async def cleanup_orphan_permissions():
    """يُزيل أي صلاحيات قديمة لم تعد موجودة في النظام من الأدوار والمستخدمين.
    
    يحل مشكلة: ظهور alert "صلاحية غير صالحة" في الواجهة عند تعديل دور لأنه يحتوي
    على صلاحية قديمة (مثل view_curriculum) أُلغيت من الكود.
    
    أيضاً يُحوّل بعض الصلاحيات الملغاة إلى بديلها (مثل view_curriculum → manage_curriculum)
    حتى لا يفقد المستخدمون السلوك المعتمد على تلك الصلاحية.
    """
    try:
        # كل الصلاحيات الفعلية الحالية في النظام (من قائمة ALL_PERMISSIONS)
        valid_keys = {p["key"] for p in ALL_PERMISSIONS}
        # تحويلات: الصلاحية القديمة → البديل الحالي (للحفاظ على نية المنح)
        ALIASES = {
            "view_curriculum": "manage_curriculum",
        }
        
        # 1) الأدوار
        roles_fixed = 0
        async for role in db.roles.find({}):
            perms = role.get("permissions") or []
            new_perms = []
            changed = False
            for p in perms:
                if p in valid_keys:
                    new_perms.append(p)
                elif p in ALIASES:
                    alias = ALIASES[p]
                    if alias not in new_perms:
                        new_perms.append(alias)
                    changed = True
                else:
                    # صلاحية يتيمة لا alias لها — احذفها
                    changed = True
            if changed:
                # أزل أي تكرارات
                new_perms = list(dict.fromkeys(new_perms))
                await db.roles.update_one({"_id": role["_id"]}, {"$set": {"permissions": new_perms}})
                roles_fixed += 1
        
        # 2) المستخدمون (نفس المعالجة)
        users_fixed = 0
        async for user in db.users.find({"permissions": {"$exists": True, "$ne": []}}):
            perms = user.get("permissions") or []
            new_perms = []
            changed = False
            for p in perms:
                if p in valid_keys:
                    new_perms.append(p)
                elif p in ALIASES:
                    alias = ALIASES[p]
                    if alias not in new_perms:
                        new_perms.append(alias)
                    changed = True
                else:
                    changed = True
            if changed:
                new_perms = list(dict.fromkeys(new_perms))
                await db.users.update_one({"_id": user["_id"]}, {"$set": {"permissions": new_perms}})
                users_fixed += 1
        
        if roles_fixed > 0 or users_fixed > 0:
            logging.info(
                f"✅ Orphan permissions cleanup: fixed {roles_fixed} role(s) and {users_fixed} user(s)."
            )
        else:
            logging.info("Orphan permissions cleanup: no orphans found.")
    except Exception as e:
        logging.error(f"Orphan permissions cleanup failed (non-critical): {e}")


async def cleanup_duplicate_roles_internal():
    """يدمج الأدوار المكررة بنفس الاسم، يُبقي على دور النظام (system_key) ويُعيد ربط المستخدمين.
    
    يحل مشكلة: وجود دور نظام "رئيس قسم" (system_key=department_head) ودور مخصص مكرر
    بنفس الاسم. المستخدم يُسند للمكرر فيصبح خارج فلتر النظام.
    """
    try:
        all_roles = await db.roles.find().to_list(1000)
        name_groups = {}
        for role in all_roles:
            name = (role.get("name") or "").strip()
            if not name:
                continue
            name_groups.setdefault(name, []).append(role)

        merged = 0
        for name, roles in name_groups.items():
            if len(roles) <= 1:
                continue
            # الأولوية: دور نظام له system_key صالح، ثم الأقدم
            VALID_KEYS = {"admin", "dean", "department_head", "registrar",
                          "registration_manager", "employee", "teacher", "student"}
            roles.sort(key=lambda r: (
                0 if (r.get("system_key") or "").strip() in VALID_KEYS else 1,
                r.get("created_at") or ""
            ))
            keep = roles[0]
            for dup in roles[1:]:
                # نقل المستخدمين من الدور المكرر إلى الأصلي
                await db.users.update_many(
                    {"role_id": str(dup["_id"])},
                    {"$set": {"role_id": str(keep["_id"])}}
                )
                await db.roles.delete_one({"_id": dup["_id"]})
                merged += 1
                logging.info(f"Cleanup duplicate role '{name}': merged dup {dup['_id']} → {keep['_id']}")
        if merged > 0:
            logging.info(f"✅ Duplicate roles cleanup: merged {merged} duplicate role(s).")
        else:
            logging.info("Duplicate roles cleanup: no duplicates found.")
    except Exception as e:
        logging.error(f"Duplicate roles cleanup failed (non-critical): {e}")


async def migrate_broken_user_roles():
    """يُصلح تلقائياً أي مستخدم role.role لا يطابق role_id.system_key.
    
    يعمل عند كل startup ويصلح ثلاث حالات:
    1. role غير قياسي (custom/null/فارغ) — يستنتج من role_id أو اسم الدور العربي
    2. role قديم لكن role_id يشير لدور آخر — يُحدِّث ليطابق role_id الحالي
    3. مستخدم له role_id صالح لكن role متروك على القيمة القديمة بعد ترقية
    
    Idempotent: لا يُحدث المستخدمين الذين role/role_id متّسقَين.
    """
    try:
        VALID_SYSTEM_ROLES = {"admin", "dean", "department_head", "registrar",
                              "registration_manager", "employee", "teacher", "student"}
        ROLE_NAME_MAP = {
            "رئيس قسم": "department_head", "عميد كلية": "dean", "عميد": "dean",
            "مسجل": "registrar", "مدير التسجيل": "registration_manager",
            "موظف": "employee", "مدير النظام": "admin",
            "Department Head": "department_head", "Dean": "dean", "Registrar": "registrar",
        }
        fixed_count = 0
        # 🔧 نفحص كل مستخدم له role_id ونحدّث user.role ليطابق system_key للدور الحالي
        # (يعالج: ترقية تُحدّث role_id ولا تُحدّث role)
        async for u in db.users.find({"role_id": {"$nin": [None, ""]}}):
            try:
                role = await db.roles.find_one({"_id": ObjectId(u["role_id"])})
                if not role:
                    continue
                sys_key = (role.get("system_key") or "").strip()
                if not sys_key:
                    role_name = (role.get("name") or "").strip()
                    sys_key = ROLE_NAME_MAP.get(role_name, "")
                if not sys_key or sys_key not in VALID_SYSTEM_ROLES:
                    continue
                # 🔒 لا نحوّل تلقائياً لـ teacher/student (شاشات مستقلة)
                if sys_key in ["teacher", "student"]:
                    continue
                current_role = (u.get("role") or "").strip()
                # نُحدِّث فقط إن كان مختلفاً — منع تحديثات لا داعي لها
                if current_role != sys_key:
                    await db.users.update_one(
                        {"_id": u["_id"]},
                        {"$set": {"role": sys_key}}
                    )
                    fixed_count += 1
                    logging.info(f"Migration: '{u.get('username')}' role: '{current_role}' → '{sys_key}'")
            except Exception as e:
                logging.warning(f"Migration: failed to fix user {u.get('username')}: {e}")
        if fixed_count > 0:
            logging.info(f"✅ User role migration: {fixed_count} user(s) fixed automatically.")
        else:
            logging.info("User role migration: all users consistent (no fixes needed).")
    except Exception as e:
        logging.error(f"User role migration failed (non-critical): {e}")

async def _ensure_weekly_schedule_unique_indexes(db):
    """
    🔒 حماية جذرية ضد الحجز المزدوج في weekly_schedule:
    1. نبحث عن أي تكرارات قائمة (نفس teacher/day/slot أو room/day/slot أو dept/level/section/day/slot).
    2. نحتفظ بأقدم نسخة ونحذف الفائض (يتم تسجيله في السجل).
    3. ننشئ فهارس فريدة جزئية (partial unique indexes) بحيث ترفض MongoDB أي إدخال متعارض حتى لو أخطأ الكود.
    """
    try:
        # --- 1) إزالة أي تكرارات قائمة ---
        removed_total = 0
        # (a) نفس المعلم في نفس اليوم/الفترة
        pipe = [
            {"$match": {"teacher_id": {"$exists": True, "$nin": [None, ""]}}},
            {"$group": {
                "_id": {"t": "$teacher_id", "d": "$day", "s": "$slot_number"},
                "docs": {"$push": {"id": "$_id", "created_at": "$created_at"}},
                "count": {"$sum": 1},
            }},
            {"$match": {"count": {"$gt": 1}}},
        ]
        async for group in db.weekly_schedule.aggregate(pipe):
            docs = sorted(group["docs"], key=lambda x: x.get("created_at") or datetime.min.replace(tzinfo=timezone.utc))
            # نحتفظ بالأقدم، نحذف الباقي
            for extra in docs[1:]:
                await db.weekly_schedule.delete_one({"_id": extra["id"]})
                removed_total += 1
                logging.warning(
                    f"[schedule dedupe] حُذف تكرار المعلم: teacher={group['_id']['t']} "
                    f"day={group['_id']['d']} slot={group['_id']['s']} id={extra['id']}"
                )

        # (b) نفس القاعة في نفس اليوم/الفترة
        pipe = [
            {"$match": {"room_id": {"$exists": True, "$nin": [None, ""]}}},
            {"$group": {
                "_id": {"r": "$room_id", "d": "$day", "s": "$slot_number"},
                "docs": {"$push": {"id": "$_id", "created_at": "$created_at"}},
                "count": {"$sum": 1},
            }},
            {"$match": {"count": {"$gt": 1}}},
        ]
        async for group in db.weekly_schedule.aggregate(pipe):
            docs = sorted(group["docs"], key=lambda x: x.get("created_at") or datetime.min.replace(tzinfo=timezone.utc))
            for extra in docs[1:]:
                await db.weekly_schedule.delete_one({"_id": extra["id"]})
                removed_total += 1
                logging.warning(
                    f"[schedule dedupe] حُذف تكرار القاعة: room={group['_id']['r']} "
                    f"day={group['_id']['d']} slot={group['_id']['s']} id={extra['id']}"
                )

        # (c) نفس الشعبة في نفس اليوم/الفترة
        pipe = [
            {"$group": {
                "_id": {
                    "dept": "$department_id", "lvl": "$level",
                    "sec": {"$ifNull": ["$section", ""]},
                    "d": "$day", "s": "$slot_number",
                },
                "docs": {"$push": {"id": "$_id", "created_at": "$created_at"}},
                "count": {"$sum": 1},
            }},
            {"$match": {"count": {"$gt": 1}}},
        ]
        async for group in db.weekly_schedule.aggregate(pipe):
            docs = sorted(group["docs"], key=lambda x: x.get("created_at") or datetime.min.replace(tzinfo=timezone.utc))
            for extra in docs[1:]:
                await db.weekly_schedule.delete_one({"_id": extra["id"]})
                removed_total += 1
                logging.warning(
                    f"[schedule dedupe] حُذف تكرار الشعبة: dept={group['_id']['dept']} "
                    f"level={group['_id']['lvl']} sec={group['_id']['sec']} "
                    f"day={group['_id']['d']} slot={group['_id']['s']} id={extra['id']}"
                )

        if removed_total > 0:
            logging.warning(f"[schedule dedupe] TOTAL removed: {removed_total} slot(s). See warnings above.")

        # --- 2) إنشاء فهارس فريدة جزئية (Partial Unique Indexes) ---
        # MongoDB لا يدعم $ne في partial index. نستخدم $exists+$type فقط.
        # ملاحظة: السجلات ذات teacher_id="" ستُشملها الفهرس، لكن هذا مقبول لأن الحقل نصّي.
        await db.weekly_schedule.create_index(
            [("teacher_id", 1), ("day", 1), ("slot_number", 1)],
            unique=True,
            partialFilterExpression={"teacher_id": {"$exists": True, "$type": "string"}},
            name="uniq_teacher_day_slot",
        )
        await db.weekly_schedule.create_index(
            [("room_id", 1), ("day", 1), ("slot_number", 1)],
            unique=True,
            partialFilterExpression={"room_id": {"$exists": True, "$type": "string"}},
            name="uniq_room_day_slot",
        )
        await db.weekly_schedule.create_index(
            [("department_id", 1), ("level", 1), ("section", 1), ("day", 1), ("slot_number", 1)],
            unique=True,
            partialFilterExpression={"department_id": {"$exists": True, "$type": "string"}},
            name="uniq_section_day_slot",
        )
    except Exception as e:
        logging.warning(f"[weekly_schedule unique indexes] failed: {e}")



async def _ensure_lectures_unique_index(db):
    """
    🔒 حماية جذرية ضد المحاضرات المكررة في lectures:
    1. نبحث عن أي تكرارات قائمة (نفس course_id/date/start_time بحالة نشطة).
    2. نُبقي على المحاضرة الأهم (المنعقدة أولاً ثم الأقدم) ونُلغي الفائض (لا نحذف حفاظاً على سجلات الحضور).
    3. ننشئ فهرس فريد جزئي بحيث ترفض MongoDB أي إدخال مكرر حتى لو أخطأ الكود.
    """
    try:
        active_statuses = [LectureStatus.SCHEDULED, LectureStatus.COMPLETED, LectureStatus.ABSENT]
        cancelled_total = 0
        pipe = [
            {"$match": {"status": {"$in": active_statuses}}},
            {"$group": {
                "_id": {"c": "$course_id", "d": "$date", "s": "$start_time"},
                "docs": {"$push": {"id": "$_id", "created_at": "$created_at", "status": "$status"}},
                "count": {"$sum": 1},
            }},
            {"$match": {"count": {"$gt": 1}}},
        ]
        async for group in db.lectures.aggregate(pipe):
            # نُفضّل الإبقاء على المنعقدة (completed)، ثم الأقدم إنشاءً
            docs = sorted(
                group["docs"],
                key=lambda x: (0 if x.get("status") == LectureStatus.COMPLETED else 1, str(x.get("created_at") or "")),
            )
            for extra in docs[1:]:
                await db.lectures.update_one(
                    {"_id": extra["id"]},
                    {"$set": {
                        "status": LectureStatus.CANCELLED,
                        "cancellation_reason": "أُلغيت تلقائياً: محاضرة مكررة (نفس المقرر والتاريخ والوقت)",
                        "cancelled_at": get_yemen_time(),
                        "cancelled_by_name": "النظام (تنظيف تلقائي)",
                    }},
                )
                cancelled_total += 1
                logging.warning(
                    f"[lectures dedupe] أُلغيت محاضرة مكررة: course={group['_id']['c']} "
                    f"date={group['_id']['d']} start={group['_id']['s']} id={extra['id']}"
                )
        if cancelled_total > 0:
            logging.warning(f"[lectures dedupe] TOTAL cancelled duplicates: {cancelled_total}")

        # فهرس فريد جزئي: يشمل الحالات النشطة فقط (الملغاة مستثناة ليمكن إعادة الإنشاء بعد الإلغاء)
        try:
            await db.lectures.create_index(
                [("course_id", 1), ("date", 1), ("start_time", 1)],
                unique=True,
                partialFilterExpression={"status": {"$in": active_statuses}},
                name="uniq_course_date_start",
            )
        except Exception as idx_err:
            # نسخ MongoDB الأقدم من 6.0 لا تدعم $in في partialFilterExpression — نستخدم scheduled فقط
            logging.warning(f"[lectures unique index] $in filter failed ({idx_err}) — fallback to scheduled-only")
            await db.lectures.create_index(
                [("course_id", 1), ("date", 1), ("start_time", 1)],
                unique=True,
                partialFilterExpression={"status": LectureStatus.SCHEDULED},
                name="uniq_course_date_start",
            )
    except Exception as e:
        logging.warning(f"[lectures unique index] failed: {e}")


async def create_indexes():
    """إنشاء فهارس لتسريع الاستعلامات"""
    try:
        await db.users.create_index("username", unique=True)
        await db.users.create_index("role")
        await db.users.create_index("faculty_id")
        await db.users.create_index("department_id")
        await db.users.create_index("is_active")
        await db.students.create_index("student_id")
        await db.students.create_index("department_id")
        await db.students.create_index("faculty_id")
        await db.students.create_index("is_active")
        await db.teachers.create_index("department_id")
        await db.teachers.create_index("faculty_id")
        await db.courses.create_index("department_id")
        await db.courses.create_index("teacher_id")
        await db.courses.create_index("semester_id")
        await db.lectures.create_index("course_id")
        await db.lectures.create_index("teacher_id")
        await db.lectures.create_index("date")
        await db.lectures.create_index([("teacher_id", 1), ("date", 1)])
        await db.attendance.create_index("lecture_id")
        await db.attendance.create_index("student_id")
        await db.attendance.create_index([("lecture_id", 1), ("student_id", 1)])
        await db.attendance.create_index([("student_id", 1), ("course_id", 1)])  # لتسريع dashboard الطالب
        await db.enrollments.create_index("course_id")
        await db.enrollments.create_index("student_id")
        await db.enrollments.create_index([("course_id", 1), ("student_id", 1)])
        await db.roles.create_index("system_key")
        # فهارس إضافية لتسريع الاستعلامات المتكررة
        await db.notifications.create_index([("recipient_id", 1), ("is_read", 1)])
        await db.students.create_index("user_id")
        await db.teachers.create_index("user_id")
        await db.courses.create_index([("department_id", 1), ("level", 1)])

        # 🔒 حماية على مستوى قاعدة البيانات ضد التعارض المزدوج في الجدول الأسبوعي
        # نطبّق فهارس فريدة جزئية بعد إزالة أي تكرار قائم (نحتفظ بأقدم نسخة)
        await _ensure_weekly_schedule_unique_indexes(db)

        # 🔒 حماية على مستوى قاعدة البيانات ضد المحاضرات المكررة (نفس المقرر/التاريخ/الوقت)
        await _ensure_lectures_unique_index(db)

        logging.info("MongoDB indexes created successfully")
    except Exception as e:
        logging.warning(f"Index creation warning: {e}")

async def sync_default_roles():
    """مزامنة الأدوار الافتراضية - إنشاء المفقودة فقط دون تعديل الموجودة"""
    role_map = {
        "admin": UserRole.ADMIN,
        "teacher": UserRole.TEACHER,
        "employee": UserRole.EMPLOYEE,
        "student": UserRole.STUDENT,
        "dean": UserRole.DEAN,
        "department_head": UserRole.DEPARTMENT_HEAD,
        "registrar": UserRole.REGISTRAR,
        "registration_manager": UserRole.REGISTRATION_MANAGER,
    }
    for system_key, role_enum in role_map.items():
        default_perms = list(DEFAULT_PERMISSIONS.get(role_enum, []))
        if not default_perms:
            continue
        existing = await db.roles.find_one({"system_key": system_key})
        if existing:
            # لا نُعدّل الصلاحيات الموجودة عموماً - المستخدم قد خصصها يدوياً
            # 🔧 استثناء: نضمن أن للعميد صلاحية approve_attendance_changes (ميزة جديدة)
            existing_perms = existing.get("permissions", [])
            if system_key == "dean" and Permission.APPROVE_ATTENDANCE_CHANGES not in existing_perms:
                await db.roles.update_one(
                    {"_id": existing["_id"]},
                    {"$addToSet": {"permissions": Permission.APPROVE_ATTENDANCE_CHANGES}}
                )
                logging.info(f"تمت إضافة approve_attendance_changes إلى دور العميد")
            else:
                logging.info(f"الدور {system_key} موجود مسبقاً - لن يتم تعديله")
        else:
            role_names = {
                "admin": "مدير النظام",
                "teacher": "معلم",
                "employee": "موظف",
                "student": "طالب",
                "dean": "عميد",
                "department_head": "رئيس القسم",
                "registrar": "مسجل",
                "registration_manager": "مدير التسجيل",
            }
            await db.roles.insert_one({
                "name": role_names.get(system_key, system_key),
                "system_key": system_key,
                "permissions": default_perms,
                "created_at": get_yemen_time(),
            })
            logging.info(f"تم إنشاء دور {system_key}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

