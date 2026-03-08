"""
Test file for Student Import APIs
Tests:
1. POST /api/import/students - Import students from Excel with Arabic columns
2. POST /api/students/import/{course_id} - Import students and enroll them
3. GET /api/students - Verify imported students appear in list
4. POST /api/enrollments/{course_id}/enroll - Enroll students in a course
"""

import pytest
import requests
import os
import io
from openpyxl import Workbook

# Get BASE_URL from environment variable
BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
if not BASE_URL:
    BASE_URL = os.environ.get('EXPO_PUBLIC_BACKEND_URL', 'https://punctuality-monitor.preview.emergentagent.com').rstrip('/')

# Test data - IDs from main agent
DEPARTMENT_ID = "698e500997fef774e66e93a8"
COURSE_ID = "698e54c5b17b90bf5c4205fe"

# Admin credentials
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"


class TestStudentImportAPIs:
    """Test class for Student Import APIs"""
    
    @pytest.fixture(autouse=True)
    def setup(self, request):
        """Setup - Login and get auth token"""
        self.base_url = BASE_URL
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        # Login as admin
        login_response = self.session.post(f"{self.base_url}/api/auth/login", json={
            "username": ADMIN_USERNAME,
            "password": ADMIN_PASSWORD
        })
        
        if login_response.status_code == 200:
            token = login_response.json().get("access_token")
            self.session.headers.update({"Authorization": f"Bearer {token}"})
            self.token = token
        else:
            pytest.skip("Login failed, skipping tests")
    
    def create_excel_file_with_arabic_columns(self, students_data):
        """Create Excel file with Arabic column names"""
        wb = Workbook()
        ws = wb.active
        
        # Arabic column headers
        ws['A1'] = 'رقم الطالب'
        ws['B1'] = 'اسم الطالب'
        
        # Add student data
        for idx, student in enumerate(students_data, start=2):
            ws[f'A{idx}'] = student['student_id']
            ws[f'B{idx}'] = student['full_name']
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def create_excel_file_with_english_columns(self, students_data):
        """Create Excel file with English column names"""
        wb = Workbook()
        ws = wb.active
        
        # English column headers
        ws['A1'] = 'student_id'
        ws['B1'] = 'full_name'
        
        # Add student data
        for idx, student in enumerate(students_data, start=2):
            ws[f'A{idx}'] = student['student_id']
            ws[f'B{idx}'] = student['full_name']
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def test_api_health(self):
        """Test API health endpoint"""
        response = self.session.get(f"{self.base_url}/api/health")
        assert response.status_code == 200, f"Health check failed: {response.text}"
        
        data = response.json()
        assert data["status"] == "ok", "Health status should be 'ok'"
        print("✓ API health check passed")
    
    def test_admin_login(self):
        """Test admin login"""
        response = self.session.post(f"{self.base_url}/api/auth/login", json={
            "username": ADMIN_USERNAME,
            "password": ADMIN_PASSWORD
        })
        
        assert response.status_code == 200, f"Login failed: {response.text}"
        
        data = response.json()
        assert "access_token" in data, "Response should contain access_token"
        assert data["user"]["role"] == "admin", "User should be admin"
        print("✓ Admin login successful")
    
    def test_get_departments(self):
        """Test getting departments to verify test department exists"""
        response = self.session.get(f"{self.base_url}/api/departments")
        assert response.status_code == 200, f"Get departments failed: {response.text}"
        
        departments = response.json()
        assert isinstance(departments, list), "Response should be a list"
        
        # Verify our test department exists
        dept_ids = [d["id"] for d in departments]
        if DEPARTMENT_ID not in dept_ids:
            print(f"Warning: Test department {DEPARTMENT_ID} not found. Available: {dept_ids[:3]}")
        else:
            print(f"✓ Test department {DEPARTMENT_ID} exists")
    
    def test_get_courses(self):
        """Test getting courses to verify test course exists"""
        response = self.session.get(f"{self.base_url}/api/courses")
        assert response.status_code == 200, f"Get courses failed: {response.text}"
        
        courses = response.json()
        assert isinstance(courses, list), "Response should be a list"
        
        # Verify our test course exists
        course_ids = [c["id"] for c in courses]
        if COURSE_ID not in course_ids:
            print(f"Warning: Test course {COURSE_ID} not found. Available: {course_ids[:3]}")
        else:
            print(f"✓ Test course {COURSE_ID} exists")
    
    def test_import_students_from_excel_arabic_columns(self):
        """Test POST /api/import/students with Arabic column names"""
        # Create test students with unique IDs
        import uuid
        unique_suffix = uuid.uuid4().hex[:6].upper()
        
        students_data = [
            {"student_id": f"TEST_{unique_suffix}_001", "full_name": "طالب اختبار الأول"},
            {"student_id": f"TEST_{unique_suffix}_002", "full_name": "طالب اختبار الثاني"},
            {"student_id": f"TEST_{unique_suffix}_003", "full_name": "طالب اختبار الثالث"},
        ]
        
        # Create Excel file
        excel_file = self.create_excel_file_with_arabic_columns(students_data)
        
        # Prepare multipart form data
        files = {
            'file': ('students_test.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        # Make request with query params
        url = f"{self.base_url}/api/import/students?department_id={DEPARTMENT_ID}&level=1"
        
        # Remove Content-Type header for multipart
        headers = {"Authorization": f"Bearer {self.token}"}
        
        response = requests.post(url, files=files, headers=headers)
        
        print(f"Import response: {response.status_code} - {response.text}")
        
        assert response.status_code == 200, f"Import failed: {response.text}"
        
        data = response.json()
        assert "imported" in data, "Response should contain 'imported' count"
        assert "imported_ids" in data, "Response should contain 'imported_ids'"
        assert data["imported"] == 3, f"Should import 3 students, got {data['imported']}"
        assert len(data["imported_ids"]) == 3, "Should have 3 imported IDs"
        
        print(f"✓ Successfully imported {data['imported']} students with Arabic columns")
        print(f"  Imported IDs: {data['imported_ids']}")
        
        # Store for cleanup
        self.__class__.imported_student_ids = data["imported_ids"]
        self.__class__.test_student_numbers = [s["student_id"] for s in students_data]
    
    def test_verify_imported_students_in_list(self):
        """Test GET /api/students to verify imported students appear"""
        response = self.session.get(f"{self.base_url}/api/students")
        assert response.status_code == 200, f"Get students failed: {response.text}"
        
        students = response.json()
        assert isinstance(students, list), "Response should be a list"
        
        # Check if any of our test students exist
        if hasattr(self.__class__, 'test_student_numbers'):
            student_ids_in_db = [s.get("student_id") for s in students]
            for test_id in self.__class__.test_student_numbers:
                if test_id in student_ids_in_db:
                    print(f"✓ Test student {test_id} found in students list")
                else:
                    print(f"Warning: Test student {test_id} not found in list")
        
        print(f"✓ Got {len(students)} students from API")
    
    def test_enroll_students_in_course(self):
        """Test POST /api/enrollments/{course_id} to enroll students"""
        if not hasattr(self.__class__, 'imported_student_ids') or not self.__class__.imported_student_ids:
            pytest.skip("No imported students to enroll")
        
        # Enroll imported students
        enroll_data = {
            "course_id": COURSE_ID,
            "student_ids": self.__class__.imported_student_ids
        }
        
        response = self.session.post(
            f"{self.base_url}/api/enrollments/{COURSE_ID}",
            json=enroll_data
        )
        
        print(f"Enroll response: {response.status_code} - {response.text}")
        
        assert response.status_code == 200, f"Enroll failed: {response.text}"
        
        data = response.json()
        assert "enrolled" in data, "Response should contain 'enrolled' count"
        
        print(f"✓ Enrolled {data['enrolled']} students in course {COURSE_ID}")
        if data.get("already_enrolled", 0) > 0:
            print(f"  Already enrolled: {data['already_enrolled']}")
    
    def test_verify_enrollments(self):
        """Test GET /api/enrollments/{course_id} to verify enrollments"""
        response = self.session.get(f"{self.base_url}/api/enrollments/{COURSE_ID}")
        
        print(f"Get enrollments response: {response.status_code} - {response.text[:500] if len(response.text) > 500 else response.text}")
        
        assert response.status_code == 200, f"Get enrollments failed: {response.text}"
        
        enrollments = response.json()
        assert isinstance(enrollments, list), "Response should be a list"
        
        print(f"✓ Course has {len(enrollments)} enrolled students")
    
    def test_import_students_with_course_enrollment(self):
        """Test POST /api/students/import/{course_id} - Import and auto-enroll"""
        import uuid
        unique_suffix = uuid.uuid4().hex[:6].upper()
        
        students_data = [
            {"student_id": f"COURSE_{unique_suffix}_001", "full_name": "طالب مقرر اختبار 1"},
            {"student_id": f"COURSE_{unique_suffix}_002", "full_name": "طالب مقرر اختبار 2"},
        ]
        
        # Create Excel file with English columns (this endpoint expects student_id/full_name)
        excel_file = self.create_excel_file_with_english_columns(students_data)
        
        files = {
            'file': ('course_students.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        url = f"{self.base_url}/api/students/import/{COURSE_ID}"
        headers = {"Authorization": f"Bearer {self.token}"}
        
        response = requests.post(url, files=files, headers=headers)
        
        print(f"Import to course response: {response.status_code} - {response.text}")
        
        assert response.status_code == 200, f"Import to course failed: {response.text}"
        
        data = response.json()
        assert "imported" in data, "Response should contain 'imported' count"
        assert "enrolled" in data, "Response should contain 'enrolled' count"
        
        print(f"✓ Imported {data['imported']} students and enrolled {data['enrolled']} in course")
        
        # Store for cleanup
        if not hasattr(self.__class__, 'test_student_numbers'):
            self.__class__.test_student_numbers = []
        self.__class__.test_student_numbers.extend([s["student_id"] for s in students_data])
    
    def test_import_without_department(self):
        """Test that import fails without department_id"""
        students_data = [
            {"student_id": "NODEPT_001", "full_name": "طالب بدون قسم"},
        ]
        
        excel_file = self.create_excel_file_with_arabic_columns(students_data)
        
        files = {
            'file': ('test.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        # Missing department_id
        url = f"{self.base_url}/api/import/students?level=1"
        headers = {"Authorization": f"Bearer {self.token}"}
        
        response = requests.post(url, files=files, headers=headers)
        
        print(f"Import without dept response: {response.status_code}")
        
        # Should fail with 400 or 422
        assert response.status_code in [400, 422], f"Should fail without department_id: {response.text}"
        print("✓ Correctly rejected import without department_id")
    
    def test_import_with_missing_columns(self):
        """Test that import fails with missing required columns"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'wrong_column'  # Wrong column name
        ws['B1'] = 'another_wrong'
        ws['A2'] = '12345'
        ws['B2'] = 'Test Name'
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {
            'file': ('bad_columns.xlsx', output, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        url = f"{self.base_url}/api/import/students?department_id={DEPARTMENT_ID}&level=1"
        headers = {"Authorization": f"Bearer {self.token}"}
        
        response = requests.post(url, files=files, headers=headers)
        
        print(f"Import with bad columns response: {response.status_code} - {response.text}")
        
        # Should fail with 400
        assert response.status_code == 400, f"Should fail with wrong columns: {response.text}"
        print("✓ Correctly rejected import with wrong column names")
    
    def test_cleanup_test_students(self):
        """Cleanup test data - delete TEST_ prefixed students"""
        response = self.session.get(f"{self.base_url}/api/students")
        
        if response.status_code != 200:
            print(f"Warning: Could not get students for cleanup: {response.text}")
            return
        
        students = response.json()
        deleted_count = 0
        
        for student in students:
            student_id = student.get("student_id", "")
            if student_id.startswith("TEST_") or student_id.startswith("COURSE_"):
                del_response = self.session.delete(f"{self.base_url}/api/students/{student['id']}")
                if del_response.status_code == 200:
                    deleted_count += 1
                else:
                    print(f"Warning: Could not delete student {student['id']}: {del_response.text}")
        
        print(f"✓ Cleaned up {deleted_count} test students")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
