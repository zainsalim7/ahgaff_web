"""
Teaching Load Routes - مسارات إدارة العبء التدريسي
"""
from fastapi import APIRouter, HTTPException, status, Depends, Query
from fastapi.responses import StreamingResponse
from bson import ObjectId
from datetime import datetime, timezone
from typing import Optional
from pydantic import BaseModel

from .deps import get_db, get_current_user, has_permission, log_activity
from models.permissions import Permission

router = APIRouter(tags=["العبء التدريسي"])


class TeachingLoadCreate(BaseModel):
    teacher_id: str
    course_id: str
    weekly_hours: float
    semester_id: Optional[str] = None
    notes: Optional[str] = None


class TeachingLoadUpdate(BaseModel):
    weekly_hours: Optional[float] = None
    notes: Optional[str] = None


async def _ensure_course_in_user_scope(db, current_user: dict, course: dict):
    """🔒 يتأكد أن المقرر ضمن نطاق المستخدم (قسمه/كليته).
    - admin / university_president: سماح كامل
    - مستخدم بـ department_ids (أو department_id قديم): المقرر يجب أن يكون من أحد أقسامه
    - مستخدم بـ faculty_ids/faculty_id فقط: المقرر يجب أن يكون من إحدى كلياته
    - مستخدم بدون نطاق: يُسمح (global مثل رئيس الجامعة إذا لم يكن admin)
    """
    role = current_user.get("role")
    if role in ("admin", "university_president"):
        return
    # دعم multi-department + backward compat مع department_id المفرد
    user_dept_ids = current_user.get("department_ids") or []
    if not user_dept_ids and current_user.get("department_id"):
        user_dept_ids = [current_user.get("department_id")]
    user_dept_ids = [str(x) for x in user_dept_ids if x]

    # دعم multi-faculty + backward compat مع faculty_id المفرد
    user_faculty_ids = current_user.get("faculty_ids") or []
    if not user_faculty_ids and current_user.get("faculty_id"):
        user_faculty_ids = [current_user.get("faculty_id")]
    user_faculty_ids = [str(x) for x in user_faculty_ids if x]

    course_dept_id = str(course.get("department_id") or "")
    course_faculty_id = str(course.get("faculty_id") or "")

    # إذا لم يكن لدى المقرر department/faculty، نستنتج faculty_id من القسم
    if course_dept_id and not course_faculty_id:
        try:
            dept_doc = await db.departments.find_one({"_id": ObjectId(course_dept_id)})
            if dept_doc and dept_doc.get("faculty_id"):
                course_faculty_id = str(dept_doc.get("faculty_id"))
        except Exception:
            pass

    if user_dept_ids:
        # يكفي أن يكون المقرر تابعاً لأحد أقسام المستخدم
        if course_dept_id and course_dept_id not in user_dept_ids:
            raise HTTPException(
                status_code=403,
                detail="لا يمكنك إسناد مقرر لا ينتمي لأحد أقسامك"
            )
    elif user_faculty_ids:
        # يكفي أن يكون المقرر تابعاً لإحدى كليات المستخدم
        if course_faculty_id and course_faculty_id not in user_faculty_ids:
            raise HTTPException(
                status_code=403,
                detail="لا يمكنك إسناد مقرر لا ينتمي لإحدى كلياتك"
            )
    # مستخدم بدون نطاق ولكنه يملك الصلاحية = مستخدم عام → يُسمح


@router.get("/teaching-load")
async def get_teaching_loads(
    department_id: Optional[str] = None,
    teacher_id: Optional[str] = None,
    semester_id: Optional[str] = None,
    all_semesters: bool = False,
    current_user: dict = Depends(get_current_user),
):
    """جلب جدول العبء التدريسي مع تفاصيل المعلم والمقرر.
    افتراضياً يعرض الفصل النشط فقط. تمرير all_semesters=true لعرض كل الفصول،
    أو semester_id لتحديد فصل معيّن (نشط أو مؤرشف).
    """
    if not has_permission(current_user, Permission.VIEW_TEACHING_LOAD) and not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    db = get_db()
    query = {}
    if teacher_id:
        query["teacher_id"] = teacher_id

    # تحديد فلتر الفصل: explicit > active > all
    if semester_id:
        query["semester_id"] = semester_id
    elif not all_semesters:
        active_sem = await db.semesters.find_one({"status": "active"})
        if active_sem:
            query["semester_id"] = str(active_sem["_id"])

    # If department filter, get teacher ids in that department first
    teacher_ids_in_dept = None
    if department_id:
        dept_teachers = await db.teachers.find(
            {"department_id": department_id, "is_active": {"$ne": False}},
            {"_id": 1}
        ).to_list(500)
        teacher_ids_in_dept = [str(t["_id"]) for t in dept_teachers]
        if teacher_id:
            if teacher_id not in teacher_ids_in_dept:
                return {"items": [], "summary": {}}
        else:
            query["teacher_id"] = {"$in": teacher_ids_in_dept}

    loads = await db.teaching_loads.find(query).to_list(500)

    # Collect unique ids for batch lookup
    t_ids = list({l["teacher_id"] for l in loads})
    c_ids = list({l["course_id"] for l in loads})

    # Batch fetch teachers and courses
    teachers_map = {}
    if t_ids:
        obj_ids = []
        for tid in t_ids:
            try:
                obj_ids.append(ObjectId(tid))
            except Exception:
                pass
        teachers_docs = await db.teachers.find({"_id": {"$in": obj_ids}}).to_list(500)
        for t in teachers_docs:
            teachers_map[str(t["_id"])] = t

    courses_map = {}
    if c_ids:
        obj_ids = []
        for cid in c_ids:
            try:
                obj_ids.append(ObjectId(cid))
            except Exception:
                pass
        courses_docs = await db.courses.find({"_id": {"$in": obj_ids}}).to_list(500)
        for c in courses_docs:
            courses_map[str(c["_id"])] = c

    items = []
    teacher_summary = {}  # teacher_id -> total hours

    for load in loads:
        tid = load["teacher_id"]
        cid = load["course_id"]
        teacher = teachers_map.get(tid, {})
        course = courses_map.get(cid, {})

        # 🔧 تجاهل السجلات اليتيمة: المقرر محذوف أو غير نشط
        # (تسبّب في ظهور صفوف بلا اسم في الجدول/PDF — وكذلك بقايا فصول قديمة)
        if not course:
            continue
        if course.get("is_active") is False:
            continue

        wh = load.get("weekly_hours", 0)
        if tid not in teacher_summary:
            teacher_summary[tid] = {
                "teacher_name": teacher.get("full_name", ""),
                "total_hours": 0,
                "courses_count": 0,
            }
        teacher_summary[tid]["total_hours"] += wh
        teacher_summary[tid]["courses_count"] += 1

        items.append({
            "id": str(load["_id"]),
            "teacher_id": tid,
            "teacher_name": teacher.get("full_name", ""),
            "teacher_employee_id": teacher.get("teacher_id", ""),
            "department_id": teacher.get("department_id", ""),
            "course_id": cid,
            "course_name": course.get("name", ""),
            "course_code": course.get("code", ""),
            "course_section": course.get("section", ""),
            "course_level": course.get("level", 1),
            "course_credit_hours": course.get("credit_hours", 3),
            "course_department_id": course.get("department_id", ""),
            "weekly_hours": wh,
            "semester_id": load.get("semester_id"),
            "notes": load.get("notes", ""),
            "created_at": load.get("created_at"),
            "updated_at": load.get("updated_at"),
        })

    return {
        "items": items,
        "summary": teacher_summary,
    }


@router.post("/teaching-load")
async def create_teaching_load(
    data: TeachingLoadCreate,
    current_user: dict = Depends(get_current_user),
):
    """إضافة عبء تدريسي (معلم + مقرر + ساعات أسبوعية)"""
    if not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    db = get_db()

    # Validate teacher exists
    teacher = await db.teachers.find_one({"_id": ObjectId(data.teacher_id)})
    if not teacher:
        raise HTTPException(status_code=404, detail="المعلم غير موجود")

    # Validate course exists
    course = await db.courses.find_one({"_id": ObjectId(data.course_id)})
    if not course:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")

    # 🔒 منع إسناد مقرر خارج نطاق المستخدم (كلية/قسم)
    await _ensure_course_in_user_scope(db, current_user, course)

    # التحقق: هل هذا المقرر (الشعبة) مسند لمعلم آخر؟
    assigned_to_other = await db.teaching_loads.find_one({
        "course_id": data.course_id,
        "teacher_id": {"$ne": data.teacher_id},
    })
    if assigned_to_other:
        other_teacher = await db.teachers.find_one({"_id": ObjectId(assigned_to_other["teacher_id"])})
        other_name = other_teacher.get("full_name", "معلم آخر") if other_teacher else "معلم آخر"
        course_label = f"{course.get('name', '')} ({course.get('code', '')})"
        if course.get("section"):
            course_label += f" - {course['section']}"
        raise HTTPException(
            status_code=409,
            detail=f"المقرر {course_label} مسند بالفعل للمعلم {other_name}. يجب حذف الإسناد الأول قبل إسناده لمعلم آخر"
        )

    # Check if already exists for this teacher
    existing = await db.teaching_loads.find_one({
        "teacher_id": data.teacher_id,
        "course_id": data.course_id,
    })
    if existing:
        # Update existing
        await db.teaching_loads.update_one(
            {"_id": existing["_id"]},
            {"$set": {
                "weekly_hours": data.weekly_hours,
                "notes": data.notes,
                "semester_id": data.semester_id,
                "updated_at": datetime.now(timezone.utc),
                "updated_by": current_user["id"],
            }}
        )
        await log_activity(
            current_user, "update_teaching_load", "teaching_load",
            str(existing["_id"]),
            f"{teacher.get('full_name', '')} - {course.get('name', '')}",
            {"weekly_hours": data.weekly_hours}
        )
        # ربط المعلم بالمقرر في courses
        await db.courses.update_one({"_id": ObjectId(data.course_id)}, {"$set": {"teacher_id": data.teacher_id}})
        return {"id": str(existing["_id"]), "message": "تم تحديث العبء التدريسي بنجاح", "updated": True}

    doc = {
        "teacher_id": data.teacher_id,
        "course_id": data.course_id,
        "weekly_hours": data.weekly_hours,
        "semester_id": data.semester_id,
        "notes": data.notes,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc),
    }
    # 🔧 إن لم يُمرّر semester_id من العميل، نأخذه من المقرر أو من الفصل النشط
    # (وإلا فإن GET /teaching-load الافتراضي يفلتر حسب الفصل النشط ولن يظهر السجل)
    if not doc["semester_id"]:
        doc["semester_id"] = course.get("semester_id")
    if not doc["semester_id"]:
        try:
            active_sem = await db.semesters.find_one({"status": "active"})
            if active_sem:
                doc["semester_id"] = str(active_sem["_id"])
        except Exception:
            pass
    result = await db.teaching_loads.insert_one(doc)

    # ربط المعلم بالمقرر في courses
    await db.courses.update_one({"_id": ObjectId(data.course_id)}, {"$set": {"teacher_id": data.teacher_id}})

    await log_activity(
        current_user, "create_teaching_load", "teaching_load",
        str(result.inserted_id),
        f"{teacher.get('full_name', '')} - {course.get('name', '')}",
        {"weekly_hours": data.weekly_hours}
    )

    return {"id": str(result.inserted_id), "message": "تم إضافة العبء التدريسي بنجاح", "updated": False}


@router.put("/teaching-load/{load_id}")
async def update_teaching_load(
    load_id: str,
    data: TeachingLoadUpdate,
    current_user: dict = Depends(get_current_user),
):
    """تعديل عبء تدريسي"""
    if not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    db = get_db()
    existing = await db.teaching_loads.find_one({"_id": ObjectId(load_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="السجل غير موجود")

    update_fields = {"updated_at": datetime.now(timezone.utc), "updated_by": current_user["id"]}
    if data.weekly_hours is not None:
        update_fields["weekly_hours"] = data.weekly_hours
    if data.notes is not None:
        update_fields["notes"] = data.notes

    await db.teaching_loads.update_one({"_id": ObjectId(load_id)}, {"$set": update_fields})

    await log_activity(
        current_user, "update_teaching_load", "teaching_load",
        load_id, None, {"weekly_hours": data.weekly_hours}
    )

    return {"message": "تم التحديث بنجاح"}


@router.delete("/teaching-load/{load_id}")
async def delete_teaching_load(
    load_id: str,
    current_user: dict = Depends(get_current_user),
):
    """حذف عبء تدريسي - مع التحقق من ملكية القسم/الكلية"""
    if not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    db = get_db()
    existing = await db.teaching_loads.find_one({"_id": ObjectId(load_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="السجل غير موجود")

    # 🔒 منع حذف إسناد مقرر لا ينتمي لقسم/كلية المستخدم (حماية ضد إساءة استخدام cross_university)
    if current_user.get("role") not in ("admin", "university_president"):
        course = await db.courses.find_one({"_id": ObjectId(existing["course_id"])})
        if course:
            # استخدام نفس منطق التحقق (يدعم multi-department/multi-faculty)
            await _ensure_course_in_user_scope(db, current_user, course)

    await db.teaching_loads.delete_one({"_id": ObjectId(load_id)})

    # إزالة ربط المعلم من المقرر
    await db.courses.update_one(
        {"_id": ObjectId(existing["course_id"]), "teacher_id": existing["teacher_id"]},
        {"$set": {"teacher_id": None}}
    )

    await log_activity(
        current_user, "delete_teaching_load", "teaching_load",
        load_id, None, None
    )

    return {"message": "تم الحذف بنجاح"}


@router.get("/teaching-load/teacher/{teacher_id}/courses")
async def get_teacher_courses_for_load(
    teacher_id: str,
    department_id: Optional[str] = None,
    semester_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """جلب جميع المقررات المتاحة لتعيين العبء للمعلم.
    افتراضياً يفلتر بالفصل النشط فقط (يستبعد المقررات من الفصول المؤرشفة).
    """
    if not has_permission(current_user, Permission.VIEW_TEACHING_LOAD) and not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    db = get_db()
    
    # جلب بيانات المعلم لمعرفة قسمه
    teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
    
    # تحديد الفصل المستهدف: لو لم يُمرَّر صراحةً، استخدم الفصل النشط
    target_semester_id = semester_id
    if not target_semester_id:
        active_sem = await db.semesters.find_one({"status": "active"})
        if active_sem:
            target_semester_id = str(active_sem["_id"])
    
    # جلب جميع المقررات النشطة (حسب القسم إذا محدد)
    course_query = {"is_active": True}
    if department_id:
        course_query["department_id"] = department_id
    elif teacher and teacher.get("department_id"):
        course_query["department_id"] = teacher["department_id"]
    
    # فلتر الفصل (يستبعد المقررات من الفصول المؤرشفة)
    if target_semester_id:
        course_query["semester_id"] = target_semester_id
    
    courses = await db.courses.find(course_query).to_list(500)

    # Get existing loads for this teacher (نفس الفصل)
    loads_query = {"teacher_id": teacher_id}
    if target_semester_id:
        loads_query["semester_id"] = target_semester_id
    existing_loads = await db.teaching_loads.find(loads_query).to_list(100)
    loads_map = {l["course_id"]: l for l in existing_loads}

    result = []
    for c in courses:
        cid = str(c["_id"])
        load = loads_map.get(cid)
        dept_name = ""
        if c.get("department_id"):
            try:
                dept = await db.departments.find_one({"_id": ObjectId(c["department_id"])})
                if dept:
                    dept_name = dept.get("name", "")
            except Exception:
                pass

        # اسم المعلم المسند حالياً للمقرر
        current_teacher_name = ""
        if c.get("teacher_id") and c["teacher_id"] != teacher_id:
            try:
                ct = await db.teachers.find_one({"_id": ObjectId(c["teacher_id"])})
                if ct:
                    current_teacher_name = ct.get("full_name", "")
            except Exception:
                pass

        result.append({
            "course_id": cid,
            "course_name": c.get("name", ""),
            "course_code": c.get("code", ""),
            "section": c.get("section", ""),
            "level": c.get("level", 1),
            "department_name": dept_name,
            "credit_hours": c.get("credit_hours", 3),
            "current_teacher_name": current_teacher_name,
            "existing_load_id": str(load["_id"]) if load else None,
            "existing_weekly_hours": load.get("weekly_hours") if load else None,
            "existing_notes": load.get("notes", "") if load else "",
        })

@router.get("/teaching-load/search-courses")
async def search_courses_for_load(
    q: str = "",
    department_id: Optional[str] = None,
    semester_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """بحث في المقررات لإسناد العبء التدريسي.
    افتراضياً يفلتر بالفصل النشط فقط (يستبعد المقررات من الفصول المؤرشفة).
    لو تم تمرير semester_id صراحةً، يُستخدم بدلاً من الفصل النشط.
    """
    if not has_permission(current_user, Permission.VIEW_TEACHING_LOAD) and not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    db = get_db()
    query = {"is_active": True}
    if department_id:
        query["department_id"] = department_id
    if q:
        query["$or"] = [
            {"name": {"$regex": q, "$options": "i"}},
            {"code": {"$regex": q, "$options": "i"}},
        ]

    # تحديد الفصل المستهدف: لو لم يُمرَّر صراحةً، استخدم الفصل النشط
    target_semester_id = semester_id
    if not target_semester_id:
        active_sem = await db.semesters.find_one({"status": "active"})
        if active_sem:
            target_semester_id = str(active_sem["_id"])

    # فلتر بـ semester_id (يستبعد المقررات من الفصول المؤرشفة)
    if target_semester_id:
        query["semester_id"] = target_semester_id

    courses = await db.courses.find(query).limit(20).to_list(20)

    # 🆕 إثراء الـ weekly_hours من curriculum_courses (إن كانت الـ courses القديمة لا تحويها)
    cc_ids = list({c.get("curriculum_course_id") for c in courses if c.get("curriculum_course_id")})
    cc_hours_map = {}
    if cc_ids:
        try:
            cc_obj_ids = [ObjectId(x) for x in cc_ids if x]
            async for cc in db.curriculum_courses.find(
                {"_id": {"$in": cc_obj_ids}}, {"weekly_hours": 1}
            ):
                cc_hours_map[str(cc["_id"])] = cc.get("weekly_hours")
        except Exception:
            pass

    # 🆕 إثراء بأسماء الأقسام والكليات (لتجاوز قصور الـ scope في الـ frontend)
    dept_ids_set = {c.get("department_id") for c in courses if c.get("department_id")}
    dept_map: dict = {}
    faculty_ids_set: set = set()
    if dept_ids_set:
        try:
            dept_obj_ids = [ObjectId(x) for x in dept_ids_set if x]
            async for d in db.departments.find(
                {"_id": {"$in": dept_obj_ids}}, {"name": 1, "faculty_id": 1}
            ):
                dept_map[str(d["_id"])] = d
                if d.get("faculty_id"):
                    faculty_ids_set.add(str(d.get("faculty_id")))
        except Exception:
            pass
    faculty_map: dict = {}
    if faculty_ids_set:
        try:
            fac_obj_ids = [ObjectId(x) for x in faculty_ids_set if x]
            async for f in db.faculties.find(
                {"_id": {"$in": fac_obj_ids}}, {"name": 1}
            ):
                faculty_map[str(f["_id"])] = f.get("name", "")
        except Exception:
            pass

    result = []
    for c in courses:
        teacher_name = ""
        if c.get("teacher_id"):
            try:
                t = await db.teachers.find_one({"_id": ObjectId(c["teacher_id"])})
                if t:
                    teacher_name = t.get("full_name", "")
            except Exception:
                pass
        # weekly_hours: من الـ course أولاً، ثم fallback لـ curriculum
        wh = c.get("weekly_hours")
        if wh is None:
            wh = cc_hours_map.get(c.get("curriculum_course_id", ""))
        # 🆕 استخراج بيانات القسم والكلية للمقرر (يعمل حتى لو القسم خارج نطاق المستخدم)
        c_dept_id = str(c.get("department_id", "") or "")
        c_dept_doc = dept_map.get(c_dept_id) if c_dept_id else None
        c_dept_name = c_dept_doc.get("name", "") if c_dept_doc else ""
        c_fac_id = str(c_dept_doc.get("faculty_id", "")) if c_dept_doc and c_dept_doc.get("faculty_id") else ""
        c_fac_name = faculty_map.get(c_fac_id, "") if c_fac_id else ""
        result.append({
            "course_id": str(c["_id"]),
            "course_name": c.get("name", ""),
            "course_code": c.get("code", ""),
            "section": c.get("section", ""),
            "level": c.get("level", 1),
            "credit_hours": c.get("credit_hours", 3),
            "weekly_hours": wh,  # 🆕 يُستخدم كافتراضي في فورم الإسناد
            "department_id": c_dept_id,
            "department_name": c_dept_name,   # 🆕
            "faculty_id": c_fac_id,            # 🆕
            "faculty_name": c_fac_name,        # 🆕
            "current_teacher_name": teacher_name,
            "semester_id": c.get("semester_id", ""),
        })
    return result


    return result


@router.post("/teaching-load/bulk")
async def bulk_save_teaching_load(
    items: list[TeachingLoadCreate],
    current_user: dict = Depends(get_current_user),
):
    """حفظ مجموعة من سجلات العبء التدريسي دفعة واحدة"""
    if not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    db = get_db()
    created = 0
    updated = 0
    errors = []

    # 🔧 جلب الفصل النشط مرة واحدة لاستخدامه كاحتياطي عند غياب semester_id
    # (الواجهة لا ترسل semester_id لكل عنصر، ولا يجب أن ينتهي السجل بـ null
    # وإلا فإن GET /teaching-load الافتراضي يفلتر بالفصل النشط ولن يظهر السجل)
    active_sem = await db.semesters.find_one({"status": "active"})
    active_sem_id = str(active_sem["_id"]) if active_sem else None

    for item in items:
        # 🔒 منع إسناد مقرر خارج نطاق المستخدم (كلية/قسم)
        course = await db.courses.find_one({"_id": ObjectId(item.course_id)})
        if not course:
            errors.append(f"مقرر غير موجود: {item.course_id}")
            continue
        try:
            await _ensure_course_in_user_scope(db, current_user, course)
        except HTTPException as scope_err:
            course_label = f"{course.get('name', '')} ({course.get('code', '')})"
            errors.append(f"{course_label}: {scope_err.detail}")
            continue

        # 🔧 الإصلاح الحاسم: تحديد الـ semester_id الفعلي للسجل
        # أولوية: (1) ما أرسله العميل صراحةً → (2) course.semester_id → (3) الفصل النشط
        # السبب: إن بقي null، startup backfill سيُحاذي مع course.semester_id ولكن
        # الواجهة لن تُظهر السجل حتى يحدث restart (بعد ~1 ساعة في Cloud Run)
        effective_sem_id = item.semester_id or course.get("semester_id") or active_sem_id

        # 🔧 إضافة: إن لم يكن للمقرر semester_id، اضبطه على الفصل النشط
        # حتى يبقى متّسقاً مع teaching_load ولا يُعكسه startup backfill لاحقاً
        if not course.get("semester_id") and active_sem_id:
            await db.courses.update_one(
                {"_id": ObjectId(item.course_id)},
                {"$set": {"semester_id": active_sem_id}}
            )

        # التحقق: هل المقرر مسند لمعلم آخر؟ (ضمن نفس الفصل فقط)
        check_query = {
            "course_id": item.course_id,
            "teacher_id": {"$ne": item.teacher_id},
        }
        if effective_sem_id:
            check_query["semester_id"] = effective_sem_id
        assigned_to_other = await db.teaching_loads.find_one(check_query)
        if assigned_to_other:
            other_teacher = await db.teachers.find_one({"_id": ObjectId(assigned_to_other["teacher_id"])})
            course_name = course.get("name", "")
            other_name = other_teacher.get("full_name", "معلم آخر") if other_teacher else "معلم آخر"
            errors.append(f"{course_name} مسند لـ {other_name}")
            continue

        existing = await db.teaching_loads.find_one({
            "teacher_id": item.teacher_id,
            "course_id": item.course_id,
            "semester_id": effective_sem_id,
        })
        if existing:
            await db.teaching_loads.update_one(
                {"_id": existing["_id"]},
                {"$set": {
                    "weekly_hours": item.weekly_hours,
                    "notes": item.notes,
                    "semester_id": effective_sem_id,
                    "updated_at": datetime.now(timezone.utc),
                    "updated_by": current_user["id"],
                }}
            )
            updated += 1
        else:
            await db.teaching_loads.insert_one({
                "teacher_id": item.teacher_id,
                "course_id": item.course_id,
                "weekly_hours": item.weekly_hours,
                "semester_id": effective_sem_id,
                "notes": item.notes,
                "created_by": current_user["id"],
                "created_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc),
            })
            created += 1

        # ربط المعلم بالمقرر في courses
        await db.courses.update_one(
            {"_id": ObjectId(item.course_id)},
            {"$set": {"teacher_id": item.teacher_id}}
        )

    await log_activity(
        current_user, "bulk_teaching_load", "teaching_load",
        None, None, {"created": created, "updated": updated}
    )

    msg = f"تم الحفظ بنجاح: {created} جديد، {updated} محدث"
    if errors:
        msg += f"\n\nلم يتم إسناد {len(errors)} مقرر:\n" + "\n".join(errors)

    return {
        "message": msg,
        "created": created,
        "updated": updated,
        "errors": errors,
    }


async def _get_export_data(
    db,
    department_id: str = None,
    start_date: str = None,
    end_date: str = None,
    teacher_id: str = None,
    semester_id: str = None,
):
    """تجميع بيانات التصدير مع حساب إجمالي الساعات للفترة.
    
    Args:
        teacher_id: لتصدير عبء معلم واحد فقط.
        semester_id: لفلترة على فصل دراسي محدد.
    
    Returns:
        (rows, weeks, period_label, grouped_teachers)
        - rows: قائمة مسطّحة (للتوافق مع كود قديم لو وُجد)
        - grouped_teachers: قائمة [{teacher_name, employee_id, courses[], total_weekly, total_semester}]
          مرتّبة حسب اسم المعلم.
    """
    from math import ceil

    query = {}
    if teacher_id:
        query["teacher_id"] = teacher_id
    elif department_id:
        dept_teachers = await db.teachers.find(
            {"department_id": department_id, "is_active": {"$ne": False}}, {"_id": 1}
        ).to_list(500)
        teacher_ids_in_dept = [str(t["_id"]) for t in dept_teachers]
        query["teacher_id"] = {"$in": teacher_ids_in_dept}
    if semester_id:
        query["semester_id"] = semester_id

    loads = await db.teaching_loads.find(query).to_list(1000)
    if not loads:
        return [], 0, "", []

    # حساب الأسابيع في الفترة
    weeks = 16  # افتراضي للفصل الدراسي
    period_label = "فصل دراسي (16 أسبوع)"
    if semester_id:
        try:
            sem_doc = await db.semesters.find_one({"_id": ObjectId(semester_id)})
            if sem_doc:
                period_label = sem_doc.get("name", period_label)
        except Exception:
            pass
    if start_date and end_date:
        try:
            sd = datetime.fromisoformat(start_date)
            ed = datetime.fromisoformat(end_date)
            days = (ed - sd).days + 1
            weeks = max(1, ceil(days / 7))
            period_label = f"من {start_date} إلى {end_date} ({weeks} أسبوع)"
        except Exception:
            pass

    # جلب المعلمين والمقررات
    t_ids = list({l["teacher_id"] for l in loads})
    c_ids = list({l["course_id"] for l in loads})

    teachers_map = {}
    if t_ids:
        docs = await db.teachers.find({"_id": {"$in": [ObjectId(x) for x in t_ids]}}).to_list(500)
        for t in docs:
            teachers_map[str(t["_id"])] = t

    courses_map = {}
    if c_ids:
        docs = await db.courses.find({"_id": {"$in": [ObjectId(x) for x in c_ids]}}).to_list(500)
        for c in docs:
            courses_map[str(c["_id"])] = c

    # خريطة dept→(faculty_id, name) ومخرج names للكليات لاستنتاج كلية المقرر/المعلم عند الحاجة
    dept_info: dict = {}  # dept_id -> {"faculty_id": str, "name": str}
    fac_names: dict = {}  # faculty_id -> name
    all_dept_ids = set()
    for t in teachers_map.values():
        if t.get("department_id"):
            all_dept_ids.add(str(t["department_id"]))
    for c in courses_map.values():
        if c.get("department_id"):
            all_dept_ids.add(str(c["department_id"]))
    if all_dept_ids:
        try:
            async for d in db.departments.find(
                {"_id": {"$in": [ObjectId(x) for x in all_dept_ids if x]}},
                {"faculty_id": 1, "name": 1},
            ):
                dept_info[str(d["_id"])] = {
                    "faculty_id": str(d.get("faculty_id") or ""),
                    "name": (d.get("name", "") or "").strip(),
                }
        except Exception:
            pass
    all_fac_ids = {v["faculty_id"] for v in dept_info.values() if v.get("faculty_id")}
    # أضف faculty_id من سجلات المعلمين/المقررات نفسها (لو كانت موجودة على الكيان)
    for t in teachers_map.values():
        if t.get("faculty_id"):
            all_fac_ids.add(str(t["faculty_id"]))
    for c in courses_map.values():
        if c.get("faculty_id"):
            all_fac_ids.add(str(c["faculty_id"]))
    if all_fac_ids:
        try:
            async for f in db.faculties.find(
                {"_id": {"$in": [ObjectId(x) for x in all_fac_ids if x]}},
                {"name": 1},
            ):
                fac_names[str(f["_id"])] = (f.get("name", "") or "").strip()
        except Exception:
            pass

    def _resolve_faculty(entity: dict) -> str:
        """يحاول إيجاد faculty_id من الكيان مباشرةً أو عبر القسم."""
        fid = entity.get("faculty_id")
        if fid:
            return str(fid)
        did = entity.get("department_id")
        if did and str(did) in dept_info:
            return dept_info[str(did)].get("faculty_id") or ""
        return ""

    # تجميع حسب المعلم
    grouped = {}
    for load in loads:
        tid = load["teacher_id"]
        teacher = teachers_map.get(tid, {})
        course = courses_map.get(load["course_id"], {})

        # 🔧 تجاهل السجلات اليتيمة في التصدير: المقرر محذوف أو غير نشط
        # (هي السبب في ظهور صفوف فارغة بلا اسم في PDF/Excel)
        if not course:
            continue
        if course.get("is_active") is False:
            continue

        wh = load.get("weekly_hours", 0)

        # 🆕 حساب ملاحظات التكليف خارج القسم/الكلية (مع ذكر الاسم)
        note = ""
        t_dept = str(teacher.get("department_id") or "")
        c_dept = str(course.get("department_id") or "")
        t_fac = _resolve_faculty(teacher)
        c_fac = _resolve_faculty(course)
        if t_fac and c_fac and t_fac != c_fac:
            c_fac_name = fac_names.get(c_fac, "")
            note = f"⚑ من كلية: {c_fac_name}" if c_fac_name else "⚑ من كلية أخرى"
        elif t_dept and c_dept and t_dept != c_dept:
            c_dept_name = dept_info.get(c_dept, {}).get("name", "")
            note = f"↻ من قسم: {c_dept_name}" if c_dept_name else "↻ من قسم آخر"

        if tid not in grouped:
            grouped[tid] = {
                "teacher_name": teacher.get("full_name", ""),
                "employee_id": teacher.get("teacher_id", ""),
                "rows": [],
                "total_weekly": 0,
            }
        grouped[tid]["rows"].append({
            "course_name": course.get("name", ""),
            "course_code": course.get("code", ""),
            "section": course.get("section", ""),
            "level": course.get("level"),
            "weekly_hours": wh,
            "total_hours": round(wh * weeks, 2),
            "note": note,
        })
        grouped[tid]["total_weekly"] += wh

    # تسطيح للتصدير (للتوافق مع الكود القديم لو احتاج)
    rows = []
    grouped_list = []
    # ترتيب المعلمين ابجدياً بالاسم
    sorted_items = sorted(grouped.items(), key=lambda kv: kv[1]["teacher_name"] or "")
    for tid, g in sorted_items:
        total_period = round(g["total_weekly"] * weeks, 2)
        # نسخة منظمة للتصدير الجديد (block per teacher)
        grouped_list.append({
            "teacher_name": g["teacher_name"],
            "employee_id": g["employee_id"],
            "courses": list(g["rows"]),
            "total_weekly": round(g["total_weekly"], 2),
            "total_semester": total_period,
        })
        # نسخة مسطّحة قديمة
        for r in g["rows"]:
            rows.append({
                "teacher_name": g["teacher_name"],
                "employee_id": g["employee_id"],
                "course_name": r["course_name"],
                "course_code": r["course_code"],
                "section": r["section"],
                "weekly_hours": r["weekly_hours"],
                "total_hours": r["total_hours"],
            })
        rows.append({
            "teacher_name": g["teacher_name"],
            "employee_id": g["employee_id"],
            "course_name": "*** الإجمالي ***",
            "course_code": "",
            "section": "",
            "weekly_hours": round(g["total_weekly"], 2),
            "total_hours": total_period,
        })

    return rows, weeks, period_label, grouped_list


async def _sync_teaching_loads_for_teachers(db, teacher_ids: list, semester_id: Optional[str] = None) -> int:
    """مزامنة تلقائية: لكل مقرر له `teacher_id` وليس له entry في `teaching_loads`،
    يُنشَأ entry بساعات أسبوعية = `credit_hours` من المقرر (افتراضي 3 إذا غير محدد).

    تُستخدم قبل توليد التقارير لضمان دقة الحساب. آمنة لإعادة الاستدعاء (idempotent).

    Returns: عدد السجلات التي تم إنشاؤها.
    """
    if not teacher_ids:
        return 0

    # جلب كل المقررات التي يدرّسها هؤلاء المعلمون (بصرف النظر عن القسم)
    course_query: dict = {
        "is_active": True,
        "teacher_id": {"$in": teacher_ids},
    }
    if semester_id:
        course_query["semester_id"] = semester_id
    courses = await db.courses.find(course_query).to_list(2000)
    if not courses:
        return 0

    # جلب الـ teaching_loads الموجودة
    load_query: dict = {"teacher_id": {"$in": teacher_ids}}
    if semester_id:
        load_query["semester_id"] = semester_id
    existing_loads = await db.teaching_loads.find(load_query, {"teacher_id": 1, "course_id": 1}).to_list(5000)
    existing_keys = {(l["teacher_id"], l["course_id"]) for l in existing_loads}

    # تحديد السجلات المفقودة
    to_insert = []
    now = datetime.now(timezone.utc)
    for c in courses:
        tid = c.get("teacher_id")
        cid = str(c["_id"])
        if not tid or (tid, cid) in existing_keys:
            continue
        weekly = c.get("credit_hours") or 3
        doc = {
            "teacher_id": tid,
            "course_id": cid,
            "weekly_hours": float(weekly),
            "notes": "تم إنشاؤه تلقائياً بمزامنة المقررات",
            "created_at": now,
            "updated_at": now,
            "auto_synced": True,
        }
        if semester_id:
            doc["semester_id"] = semester_id
        elif c.get("semester_id"):
            doc["semester_id"] = c["semester_id"]
        to_insert.append(doc)

    if to_insert:
        await db.teaching_loads.insert_many(to_insert)
    return len(to_insert)


@router.post("/teaching-load/cleanup-orphans")
async def cleanup_orphan_teaching_loads(current_user: dict = Depends(get_current_user)):
    """🧹 حذف نهائي لسجلات teaching_loads اليتيمة:
    
    1. سجلات تشير إلى course_id لم يعد موجوداً في courses
    2. سجلات لمقررات بـ is_active=False
    
    هذه السجلات هي السبب في ظهور صفوف فارغة بلا اسم في PDF/Excel،
    وفي ظهور مقررات من فصول قديمة/مؤرشفة في الفصل النشط.
    
    آمن للتشغيل المتكرّر — يمسح فقط السجلات اليتيمة.
    """
    if not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    
    all_loads = await db.teaching_loads.find({}).to_list(50000)
    deleted_missing_course = 0
    deleted_inactive_course = 0
    
    for load in all_loads:
        try:
            course = await db.courses.find_one({"_id": ObjectId(load["course_id"])})
            if not course:
                await db.teaching_loads.delete_one({"_id": load["_id"]})
                deleted_missing_course += 1
            elif course.get("is_active") is False:
                await db.teaching_loads.delete_one({"_id": load["_id"]})
                deleted_inactive_course += 1
        except Exception:
            # ObjectId غير صالح أو course_id فاسد → احذفه أيضاً
            try:
                await db.teaching_loads.delete_one({"_id": load["_id"]})
                deleted_missing_course += 1
            except Exception:
                pass
    
    total_deleted = deleted_missing_course + deleted_inactive_course
    await log_activity(
        current_user, "cleanup_orphan_teaching_loads", "teaching_load",
        None,
        f"حذف {total_deleted} سجل عبء يتيم",
        {"missing_course": deleted_missing_course, "inactive_course": deleted_inactive_course}
    )
    return {
        "success": True,
        "deleted_missing_course": deleted_missing_course,
        "deleted_inactive_course": deleted_inactive_course,
        "total_deleted": total_deleted,
        "total_checked": len(all_loads),
        "message": f"تم حذف {total_deleted} سجل يتيم ({deleted_missing_course} مقرر محذوف + {deleted_inactive_course} مقرر غير نشط)"
    }


@router.post("/teaching-load/backfill-semester")
async def backfill_teaching_load_semester(current_user: dict = Depends(get_current_user)):
    """🔧 إصلاح يدوي: محاذاة semester_id لـ teaching_loads مع course.semester_id.
    
    يصلح حالتين: السجلات بدون semester_id، أو بـ semester_id لا يطابق course.semester_id.
    لا يأخذ الفصل النشط كاحتياطي (لتجنّب إظهار سجلات قديمة في الفصل الحالي).
    """
    if not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    
    all_loads = await db.teaching_loads.find({}).to_list(50000)
    fixed = 0
    reset_to_orphan = 0
    course_missing = 0
    for load in all_loads:
        try:
            course = await db.courses.find_one({"_id": ObjectId(load["course_id"])})
            if not course:
                course_missing += 1
                continue
            correct_sem_id = course.get("semester_id")
            current_sem_id = load.get("semester_id")
            if correct_sem_id and current_sem_id != correct_sem_id:
                await db.teaching_loads.update_one(
                    {"_id": load["_id"]},
                    {"$set": {"semester_id": correct_sem_id, "updated_at": datetime.now(timezone.utc)}}
                )
                fixed += 1
            elif not correct_sem_id and current_sem_id:
                await db.teaching_loads.update_one(
                    {"_id": load["_id"]},
                    {"$set": {"semester_id": None, "updated_at": datetime.now(timezone.utc)}}
                )
                reset_to_orphan += 1
        except Exception:
            continue
    
    return {
        "success": True,
        "fixed": fixed,
        "reset_to_orphan": reset_to_orphan,
        "course_missing": course_missing,
        "total_checked": len(all_loads),
        "message": f"تم محاذاة {fixed} سجل مع فصول مقرراتها، تنظيف {reset_to_orphan} سجل، {course_missing} سجل بمقرر محذوف",
    }


@router.post("/teaching-load/sync")
async def sync_teaching_loads_endpoint(
    teacher_id: Optional[str] = None,
    department_id: Optional[str] = None,
    semester_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """مزامنة يدوية: إنشاء teaching_load مفقودة لكل مقرر له teacher_id.

    - بدون فلاتر: مزامنة كل المعلمين النشطين.
    - مع `teacher_id`: لمعلم واحد.
    - مع `department_id`: لجميع معلمي قسم.
    - مع `semester_id`: تقييد بفصل محدد.
    """
    if not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    teacher_query: dict = {"is_active": {"$ne": False}}
    if teacher_id:
        try:
            teacher_query = {"_id": ObjectId(teacher_id)}
        except Exception:
            raise HTTPException(status_code=400, detail="معرّف المعلم غير صحيح")
    elif department_id:
        teacher_query["department_id"] = department_id
    teachers = await db.teachers.find(teacher_query, {"_id": 1}).to_list(1000)
    teacher_ids = [str(t["_id"]) for t in teachers]
    created = await _sync_teaching_loads_for_teachers(db, teacher_ids, semester_id=semester_id)
    try:
        await log_activity(
            current_user, "sync_teaching_loads", "teaching_load",
            "", f"تمت مزامنة أعباء التدريس (تم إنشاء {created} سجل)",
            {"created": created, "teachers": len(teacher_ids)}
        )
    except Exception:
        pass
    return {"success": True, "created": created, "teachers_processed": len(teacher_ids)}


@router.get("/teaching-load/report/advanced")
async def advanced_teaching_load_report(
    department_id: Optional[str] = None,
    faculty_id: Optional[str] = None,
    teacher_id: Optional[str] = None,  # ⭐ تقرير لمعلم واحد
    semester_id: Optional[str] = None,  # ⭐ فلترة حسب الفصل
    term: Optional[int] = Query(None, ge=1, le=2, description="1=أول، 2=ثاني"),  # ⭐ فلترة منطقية
    current_user: dict = Depends(get_current_user),
):
    """تقرير العبء التدريسي المتقدم - مقارنة + فجوات
    الفلاتر المدعومة:
    - department_id: قسم
    - faculty_id: كلية
    - teacher_id: معلم واحد (تقرير فردي)
    - semester_id: فصل أكاديمي محدد
    - term: نوع الفصل (1=أول، 2=ثاني) — يفلتر المقررات حسب term في الخطة
    """
    if not has_permission(current_user, Permission.VIEW_TEACHING_LOAD) and not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    db = get_db()

    # فلتر المعلمين
    teacher_query = {"is_active": {"$ne": False}}
    if department_id:
        teacher_query["department_id"] = department_id
    if teacher_id:
        # عند فلترة معلم واحد، تجاهل الفلاتر الأخرى للمعلم
        try:
            teacher_query = {"_id": ObjectId(teacher_id)}
        except Exception:
            raise HTTPException(status_code=400, detail="معرّف المعلم غير صحيح")

    teachers = await db.teachers.find(teacher_query).to_list(500)
    teacher_ids = [str(t["_id"]) for t in teachers]

    # ⭐ مزامنة تلقائية: إنشاء teaching_load المفقودة للمقررات المسندة
    # هذا يضمن أن الأرقام (assigned_weekly_hours) تتطابق مع courses_count
    # تخطّ المزامنة عند عرض فصل مؤرشف (لا نريد تعديل الأرشيف)
    sem_doc_check = None
    if semester_id:
        try:
            sem_doc_check = await db.semesters.find_one({"_id": ObjectId(semester_id)})
        except Exception:
            pass
    is_archived_sem = bool(sem_doc_check and sem_doc_check.get("status") == "archived")
    if not is_archived_sem and teacher_ids:
        try:
            await _sync_teaching_loads_for_teachers(db, teacher_ids, semester_id=semester_id)
        except Exception as _e:
            # المزامنة best-effort — لا نُفشل التقرير إن فشلت
            pass

    # ⭐ التحقق إذا كان الفصل المختار مؤرشف — نقرأ من semester_archives بدلاً من الـ collection الحية
    is_archived = False
    archived_courses: list[dict] = []
    archived_loads: list[dict] = []
    if semester_id:
        try:
            sem_doc = await db.semesters.find_one({"_id": ObjectId(semester_id)})
            if sem_doc and sem_doc.get("status") == "archived":
                is_archived = True
                archive = await db.semester_archives.find_one({"semester_id": semester_id})
                if archive:
                    archived_courses = archive.get("courses", []) or []
                    # دعم اختياري لـ teaching_loads داخل الأرشيف (لو كانت محفوظة)
                    archived_loads = archive.get("teaching_loads", []) or []
        except Exception:
            pass

    # فلتر المقررات
    # 🔧 ملاحظة: نحتاج قائمتين من المقررات
    #   1. dept_courses: مقررات القسم/الكلية المختار (لحساب "مقررات بدون معلم" + total_courses)
    #   2. teacher_courses: كل مقررات المعلمين المعروضين (للعرض في بطاقة كل معلم)
    # الباغ السابق: كنا نستخدم نفس القائمة المفلترة حسب القسم لكل شيء، فإذا كان لمعلم
    # مقررات في أقسام أخرى لا تظهر له. الإصلاح: تفصلهما.
    dept_course_query = {"is_active": True}
    if department_id and not teacher_id:
        dept_course_query["department_id"] = department_id
    if semester_id:
        dept_course_query["semester_id"] = semester_id
    if term in (1, 2):
        dept_course_query["term"] = term

    teacher_course_query = {"is_active": True}
    if teacher_ids:
        teacher_course_query["teacher_id"] = {"$in": teacher_ids}
    elif teacher_id:
        teacher_course_query["teacher_id"] = teacher_id
    if semester_id:
        teacher_course_query["semester_id"] = semester_id
    if term in (1, 2):
        teacher_course_query["term"] = term

    if is_archived:
        # فلترة المقررات المؤرشفة في الذاكرة
        dept_courses = []
        teacher_courses_list = []
        for c in archived_courses:
            tc_match = True
            if term in (1, 2) and c.get("term") != term:
                tc_match = False
            # dept_courses: محدودة بالقسم المختار
            if tc_match:
                dept_ok = True
                if department_id and not teacher_id and c.get("department_id") != department_id:
                    dept_ok = False
                if dept_ok:
                    dept_courses.append(c)
            # teacher_courses: كل مقررات المعلمين المعروضين
            if tc_match and teacher_ids and c.get("teacher_id") in teacher_ids:
                teacher_courses_list.append(c)
            elif tc_match and teacher_id and c.get("teacher_id") == teacher_id:
                teacher_courses_list.append(c)
        all_courses = dept_courses  # backwards compat for stats
    else:
        dept_courses = await db.courses.find(dept_course_query).to_list(2000)
        teacher_courses_list = await db.courses.find(teacher_course_query).to_list(2000) if (teacher_ids or teacher_id) else []
        all_courses = dept_courses  # backwards compat for stats

    # جلب الأعباء
    if is_archived:
        all_loads = archived_loads
        if teacher_ids:
            all_loads = [l for l in all_loads if l.get("teacher_id") in teacher_ids]
    else:
        load_query = {}
        if teacher_ids:
            load_query["teacher_id"] = {"$in": teacher_ids}
        if semester_id:
            load_query["semester_id"] = semester_id
        all_loads = await db.teaching_loads.find(load_query).to_list(2000)
    loads_by_teacher = {}
    for l in all_loads:
        tid = l["teacher_id"]
        if tid not in loads_by_teacher:
            loads_by_teacher[tid] = []
        loads_by_teacher[tid].append(l)

    # 1. مقارنة أعباء المعلمين
    teacher_comparison = []
    teachers_without_courses = []
    for t in teachers:
        tid = str(t["_id"])
        t_loads = loads_by_teacher.get(tid, [])
        # 🔧 استخدم teacher_courses_list (تشمل مقررات المعلم في جميع الأقسام)
        # بدل all_courses (المحدودة بقسم التقرير) لإصلاح باغ عدم إظهار مقررات المعلم في أقسام أخرى
        course_pool = teacher_courses_list if (teacher_courses_list or teacher_id) else all_courses
        assigned_courses = [c for c in course_pool if c.get("teacher_id") == tid]
        total_weekly = sum(l.get("weekly_hours", 0) for l in t_loads)
        max_hours = t.get("weekly_hours", 12)
        usage_pct = round((total_weekly / max_hours * 100), 1) if max_hours > 0 else 0

        entry = {
            "teacher_id": tid,
            "teacher_name": t.get("full_name", ""),
            "employee_id": t.get("teacher_id", ""),
            "department_id": t.get("department_id", ""),
            "max_weekly_hours": max_hours,
            "assigned_weekly_hours": round(total_weekly, 2),
            "courses_count": len(assigned_courses),
            "usage_percentage": usage_pct,
            "status": "overload" if usage_pct > 100 else ("optimal" if usage_pct >= 70 else ("low" if usage_pct > 0 else "none")),
            "courses": [{"name": c.get("name", ""), "code": c.get("code", ""), "section": c.get("section", ""), "department_id": c.get("department_id", "")} for c in assigned_courses],
        }
        teacher_comparison.append(entry)
        if len(assigned_courses) == 0:
            teachers_without_courses.append(entry)

    # ترتيب حسب الاستخدام تنازلياً
    teacher_comparison.sort(key=lambda x: x["usage_percentage"], reverse=True)

    # 2. مقررات بدون معلم
    courses_without_teacher = []
    for c in all_courses:
        if not c.get("teacher_id"):
            # عدد الطلاب
            students_count = await db.enrollments.count_documents({"course_id": str(c["_id"])})
            courses_without_teacher.append({
                "course_id": str(c["_id"]),
                "course_name": c.get("name", ""),
                "course_code": c.get("code", ""),
                "section": c.get("section", ""),
                "level": c.get("level", 1),
                "credit_hours": c.get("credit_hours", 3),
                "students_count": students_count,
            })

    # 3. إحصائيات عامة
    total_teachers = len(teachers)
    teachers_with_load = len([t for t in teacher_comparison if t["assigned_weekly_hours"] > 0])
    total_courses = len(all_courses)
    courses_assigned = len([c for c in all_courses if c.get("teacher_id")])
    avg_load = round(sum(t["assigned_weekly_hours"] for t in teacher_comparison) / total_teachers, 1) if total_teachers > 0 else 0
    overloaded = len([t for t in teacher_comparison if t["status"] == "overload"])

    # Department name
    dept_name = ""
    if department_id:
        try:
            dept = await db.departments.find_one({"_id": ObjectId(department_id)})
            if dept:
                dept_name = dept.get("name", "")
        except Exception:
            pass  # ObjectId غير صحيح — نتجاهل

    # Teacher name (في وضع المعلم الواحد)
    teacher_name = ""
    if teacher_id and teachers:
        teacher_name = teachers[0].get("full_name", "")

    return {
        "department_name": dept_name,
        "teacher_name": teacher_name,
        "scope": "teacher" if teacher_id else "department",
        "is_archived": is_archived,  # ⭐ للواجهة لمعرفة أن البيانات من الأرشيف
        "summary": {
            "total_teachers": total_teachers,
            "teachers_with_load": teachers_with_load,
            "teachers_without_courses": len(teachers_without_courses),
            "total_courses": total_courses,
            "courses_assigned": courses_assigned,
            "courses_without_teacher": len(courses_without_teacher),
            "average_weekly_load": avg_load,
            "overloaded_teachers": overloaded,
        },
        "teacher_comparison": teacher_comparison,
        "courses_without_teacher": courses_without_teacher,
        "teachers_without_courses": teachers_without_courses,
    }


@router.get("/export/teaching-load/excel")
async def export_teaching_load_excel(
    department_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    teacher_id: Optional[str] = None,
    semester_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """تصدير جدول العبء التدريسي إلى Excel - تنسيق Block per teacher."""
    if not has_permission(current_user, Permission.VIEW_TEACHING_LOAD) and not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    db = get_db()
    _, weeks, period_label, grouped = await _get_export_data(
        db, department_id, start_date, end_date, teacher_id=teacher_id, semester_id=semester_id,
    )

    # جلب الكلية + القسم
    dept_name = ""
    faculty_name = ""
    effective_dept_id = department_id
    # إن لم يوجد department_id لكن teacher_id موجود، نستنتج من المعلم
    if not effective_dept_id and teacher_id:
        try:
            t_doc = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
            if t_doc and t_doc.get("department_id"):
                effective_dept_id = str(t_doc.get("department_id"))
        except Exception:
            pass
    if effective_dept_id:
        try:
            dept = await db.departments.find_one({"_id": ObjectId(effective_dept_id)})
            if dept:
                dept_name = (dept.get("name", "") or "").strip()
                fac_id = dept.get("faculty_id")
                if fac_id:
                    fac = await db.faculties.find_one({"_id": ObjectId(fac_id)})
                    if fac:
                        faculty_name = (fac.get("name", "") or "").strip()
        except Exception:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "العبء التدريسي"
    ws.sheet_view.rightToLeft = True

    # ----- ألوان وحدود ثابتة -----
    BLUE = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    LIGHT_BLUE = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    LIGHT_GREY = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    thin = Side(style='thin', color='B0BEC5')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center', wrap_text=True)
    BOLD = Font(bold=True, size=11)
    WHITE_BOLD = Font(bold=True, color='FFFFFF', size=11)
    TITLE_FONT = Font(bold=True, size=14, color='1565C0')
    SUB_FONT = Font(bold=True, size=12)

    NUM_COLS = 7  # A..G: المقرر | المستوى | الرمز | الشعبة | ساعات أسبوعية | الساعات (16) | ملاحظات

    def style_row(r, fill=None, font=None, align=center, with_border=True):
        for c in range(1, NUM_COLS + 1):
            cell = ws.cell(row=r, column=c)
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            cell.alignment = align
            if with_border:
                cell.border = border

    # ----- Header (Title + Faculty + Department) -----
    title_text = f"جدول العبء التدريسي - {period_label}" if period_label else "جدول العبء التدريسي"
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=NUM_COLS)
    ws.cell(row=1, column=1, value=title_text).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = center
    ws.row_dimensions[1].height = 28

    # Row 2: الكلية | القسم
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=2)
    ws.cell(row=2, column=1, value=f"القسم: {dept_name}" if dept_name else "القسم: -").font = SUB_FONT
    ws.cell(row=2, column=1).alignment = center
    ws.merge_cells(start_row=2, end_row=2, start_column=3, end_column=NUM_COLS)
    ws.cell(row=2, column=3, value=f"الكلية: {faculty_name}" if faculty_name else "الكلية: -").font = SUB_FONT
    ws.cell(row=2, column=3).alignment = center
    ws.row_dimensions[2].height = 22

    # ----- شريط ملخّص (Stats Bar) -----
    total_teachers = len(grouped)
    total_courses = sum(len(b["courses"]) for b in grouped)
    sum_weekly = sum(b["total_weekly"] for b in grouped)
    sum_semester = sum(b["total_semester"] for b in grouped)
    cross_dept_n = sum(1 for b in grouped for c in b["courses"] if (c.get("note") or "").startswith("↻"))
    cross_fac_n = sum(1 for b in grouped for c in b["courses"] if (c.get("note") or "").startswith("⚑"))

    STATS_FILL = PatternFill(start_color='E7F0FE', end_color='E7F0FE', fill_type='solid')
    STATS_FONT = Font(bold=True, size=11, color='1565C0')
    LABEL_FONT = Font(size=10, color='5B6678')
    # كل خانة تحتوي على سطر معاً (label + value) كقيمة نصية واحدة
    stats_items = [
        ("المعلمون", str(total_teachers)),
        ("المقررات", str(total_courses)),
        ("ساعات/أسبوع", str(round(sum_weekly, 2))),
        (f"ساعات/{weeks}", str(round(sum_semester, 2))),
        ("من قسم آخر", str(cross_dept_n)),
        ("من كلية أخرى", str(cross_fac_n)),
    ]
    # نوزع 6 خانات على 7 أعمدة: ندمج العمود الأخير ضمن الأخيرة
    # نتيح خلية واحدة لكل بطاقة بدون دمج (6 خلايا + خلية فارغة)
    for i, (label, val) in enumerate(stats_items, start=1):
        cell = ws.cell(row=3, column=i, value=f"{label}\n{val}")
        cell.font = STATS_FONT
        cell.fill = STATS_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    # العمود السابع (الأخير) فارغ بنفس التنسيق
    last_cell = ws.cell(row=3, column=7, value="")
    last_cell.fill = STATS_FILL
    last_cell.border = border
    ws.row_dimensions[3].height = 36

    cur = 5  # سطر فارغ في R4 ثم نبدأ كتل المعلمين من R5

    # عرض الأعمدة - 7 أعمدة: الرمز | المقرر | المستوى | الشعبة | ساعات أسبوعية | الساعات | ملاحظات
    widths = [13, 26, 11, 11, 15, 16, 22]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = w

    if not grouped:
        ws.merge_cells(start_row=cur, end_row=cur, start_column=1, end_column=NUM_COLS)
        ws.cell(row=cur, column=1, value="لا توجد بيانات").alignment = center
    else:
        for block in grouped:
            # ---- صف عنوان المعلم: خليتان (يمين: المعلم/الاسم, يسار: الرقم الوظيفي/الرقم) ----
            # نقسم الصف على عمود 4 (الشعبة) كنقطة وسطية تقريبية
            mid = 4
            # خلية يمين (المعلم): الأعمدة 1..mid
            ws.merge_cells(start_row=cur, end_row=cur, start_column=1, end_column=mid)
            right_cell = ws.cell(row=cur, column=1, value=f"المعلم : {block['teacher_name'] or '-'}")
            right_cell.font = Font(bold=True, size=12, color='1A2540')
            right_cell.alignment = right
            # خلية يسار (الرقم الوظيفي): الأعمدة mid+1..NUM_COLS
            ws.merge_cells(start_row=cur, end_row=cur, start_column=mid + 1, end_column=NUM_COLS)
            left_cell = ws.cell(row=cur, column=mid + 1, value=f"الرقم الوظيفي : {block['employee_id'] or '-'}")
            left_cell.font = Font(bold=True, size=12, color='1A2540')
            left_cell.alignment = right
            ws.row_dimensions[cur].height = 24
            cur += 1

            # ---- Row: Courses Header (blue) ----
            headers = ["الرمز", "المقرر", "المستوى", "الشعبة", "ساعات أسبوعية", f"الساعات ({weeks})", "ملاحظات"]
            for i, h in enumerate(headers, start=1):
                ws.cell(row=cur, column=i, value=h)
            style_row(cur, fill=BLUE, font=WHITE_BOLD)
            ws.row_dimensions[cur].height = 22
            cur += 1

            # ---- Course rows ----
            for course in block["courses"]:
                lvl = course.get("level")
                lvl_text = f"المستوى {lvl}" if lvl else "-"
                ws.cell(row=cur, column=1, value=course["course_code"] or "-")
                ws.cell(row=cur, column=2, value=course["course_name"] or "-")
                ws.cell(row=cur, column=3, value=lvl_text)
                ws.cell(row=cur, column=4, value=course["section"] or "-")
                ws.cell(row=cur, column=5, value=course["weekly_hours"])
                ws.cell(row=cur, column=6, value=course["total_hours"])
                ws.cell(row=cur, column=7, value=course.get("note") or "")
                style_row(cur)
                ws.cell(row=cur, column=2).alignment = right
                # تلوين خانة الملاحظات إذا كانت غير فارغة
                if course.get("note"):
                    note_cell = ws.cell(row=cur, column=7)
                    note_cell.font = Font(bold=True, color='C62828')
                    note_cell.fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
                cur += 1

            # ---- Row: Total ----
            ws.cell(row=cur, column=1, value="")
            ws.cell(row=cur, column=2, value="*** الإجمالي ***")
            for col_i in range(3, 5):
                ws.cell(row=cur, column=col_i, value="")
            ws.cell(row=cur, column=5, value=block["total_weekly"])
            ws.cell(row=cur, column=6, value=block["total_semester"])
            ws.cell(row=cur, column=7, value="")
            style_row(cur, fill=LIGHT_BLUE, font=BOLD)
            ws.cell(row=cur, column=2).alignment = right
            cur += 1

            # ---- Empty separator row ----
            cur += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=teaching_load.xlsx"}
    )


@router.get("/export/teaching-load/pdf")
async def export_teaching_load_pdf(
    department_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    teacher_id: Optional[str] = None,
    semester_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """تصدير جدول العبء التدريسي إلى PDF - تنسيق Block per teacher.
    
    يدعم تصدير عبء معلم واحد فقط عبر `teacher_id`.
    """
    if not has_permission(current_user, Permission.VIEW_TEACHING_LOAD) and not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import arabic_reshaper
    from bidi.algorithm import get_display
    from io import BytesIO
    from pathlib import Path

    db = get_db()
    _, weeks, period_label, grouped = await _get_export_data(
        db, department_id, start_date, end_date, teacher_id=teacher_id, semester_id=semester_id,
    )

    # جلب الكلية + القسم
    dept_name = ""
    faculty_name = ""
    effective_dept_id = department_id
    # إن لم يوجد department_id لكن teacher_id موجود، نستنتج من المعلم
    if not effective_dept_id and teacher_id:
        try:
            t_doc = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
            if t_doc and t_doc.get("department_id"):
                effective_dept_id = str(t_doc.get("department_id"))
        except Exception:
            pass
    if effective_dept_id:
        try:
            dept = await db.departments.find_one({"_id": ObjectId(effective_dept_id)})
            if dept:
                dept_name = (dept.get("name", "") or "").strip()
                fac_id = dept.get("faculty_id")
                if fac_id:
                    fac = await db.faculties.find_one({"_id": ObjectId(fac_id)})
                    if fac:
                        faculty_name = (fac.get("name", "") or "").strip()
        except Exception:
            pass

    # جلب اسم المعلم في حالة تقرير معلم واحد
    teacher_name_label = None
    if teacher_id:
        try:
            t = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
            if t:
                teacher_name_label = (t.get("full_name", "") or "").strip()
        except Exception:
            pass

    # Register Arabic font
    font_path = Path(__file__).parent.parent / "fonts" / "Amiri-Regular.ttf"
    if font_path.exists():
        try:
            pdfmetrics.registerFont(TTFont('Amiri', str(font_path)))
        except Exception:
            pass
        arabic_font = 'Amiri'
    else:
        arabic_font = 'Helvetica'

    def ar(text):
        reshaped = arabic_reshaper.reshape(str(text or ""))
        return get_display(reshaped)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=15 * mm, leftMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ArabicTitle', parent=styles['Title'],
        fontName=arabic_font, fontSize=18, alignment=TA_CENTER, textColor=colors.HexColor('#1565c0'),
        spaceAfter=4,
    )
    sub_header_style = ParagraphStyle(
        'ArabicSubHeader', parent=styles['Normal'],
        fontName=arabic_font, fontSize=12, alignment=TA_RIGHT, textColor=colors.HexColor('#1a2540'),
    )

    elements = []

    # العنوان الرئيسي
    title_text = f"جدول العبء التدريسي - {period_label}" if period_label else "جدول العبء التدريسي"
    if teacher_name_label:
        title_text = f"عبء المعلم: {teacher_name_label} - {period_label}" if period_label else f"عبء المعلم: {teacher_name_label}"
    elements.append(Paragraph(ar(title_text), title_style))
    elements.append(Spacer(1, 3 * mm))

    # صف الكلية والقسم بجانب بعض (RTL: الكلية يمين، القسم يساره)
    if faculty_name or dept_name:
        meta_table_data = [[
            ar(f"القسم: {dept_name}" if dept_name else "القسم: -"),
            ar(f"الكلية: {faculty_name}" if faculty_name else "الكلية: -"),
        ]]
        meta_table = Table(meta_table_data, colWidths=[doc.width / 2, doc.width / 2])
        meta_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1a2540')),
        ]))
        elements.append(meta_table)
        elements.append(Spacer(1, 7 * mm))

    if not grouped:
        elements.append(Paragraph(ar("لا توجد بيانات"), sub_header_style))
    else:
        # ----- شريط ملخّص (Stats Bar) -----
        total_teachers = len(grouped)
        total_courses = sum(len(b["courses"]) for b in grouped)
        sum_weekly = sum(b["total_weekly"] for b in grouped)
        sum_semester = sum(b["total_semester"] for b in grouped)
        cross_dept_n = sum(1 for b in grouped for c in b["courses"] if (c.get("note") or "").startswith("↻"))
        cross_fac_n = sum(1 for b in grouped for c in b["courses"] if (c.get("note") or "").startswith("⚑"))

        # كل بطاقة = خليّتان (label + value) في عمودين متجاورين
        stats_cards = [
            ("المعلمون", str(total_teachers)),
            ("المقررات", str(total_courses)),
            ("ساعات/أسبوع", str(round(sum_weekly, 2))),
            (f"ساعات/{weeks}", str(round(sum_semester, 2))),
            ("من قسم آخر", str(cross_dept_n)),
            ("من كلية أخرى", str(cross_fac_n)),
        ]
        # نبني صفّاً واحداً (value فوق label لكل بطاقة)
        values_row = [str(v) for _, v in stats_cards]
        labels_row = [ar(lbl) for lbl, _ in stats_cards]
        total_w = doc.width
        card_w = total_w / len(stats_cards)
        stats_tbl = Table(
            [values_row, labels_row],
            colWidths=[card_w] * len(stats_cards),
        )
        stats_tbl.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('FONTSIZE', (0, 1), (-1, 1), 10),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1565c0')),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#5b6678')),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#e7f0fe')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.white),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        # تلوين خاص لبطاقتي خارج النطاق إن وُجدت قيم
        if cross_dept_n > 0:
            stats_tbl.setStyle(TableStyle([
                ('BACKGROUND', (4, 0), (4, 1), colors.HexColor('#fff3e0')),
                ('TEXTCOLOR', (4, 0), (4, 0), colors.HexColor('#e65100')),
            ]))
        if cross_fac_n > 0:
            stats_tbl.setStyle(TableStyle([
                ('BACKGROUND', (5, 0), (5, 1), colors.HexColor('#ffebee')),
                ('TEXTCOLOR', (5, 0), (5, 0), colors.HexColor('#c62828')),
            ]))
        elements.append(stats_tbl)
        elements.append(Spacer(1, 6 * mm))

        # عرض الجدول الموحد لكل المعلمين
        total_w = doc.width
        # ترتيب الأعمدة على PDF (من اليسار → اليمين بصرياً):
        # [ملاحظات | الساعات (16) | ساعات أسبوعية | الشعبة | المستوى | المقرر | الرمز]
        # (الترتيب RTL: الرمز يمين الصفحة، الملاحظات يسارها)
        col_widths = [
            total_w * 0.18,  # ملاحظات
            total_w * 0.10,  # الساعات
            total_w * 0.11,  # ساعات أسبوعية
            total_w * 0.08,  # الشعبة
            total_w * 0.11,  # المستوى
            total_w * 0.30,  # المقرر (اسم)
            total_w * 0.12,  # الرمز
        ]

        for block in grouped:
            block_elements = []

            # --- صف عنوان المعلم: جدول من خليتين (RTL: يمين=المعلم, يسار=الرقم الوظيفي) ---
            # ReportLab يضع index 0 على اليسار، لذا لجعل "المعلم" يظهر يميناً نضعه في index 1
            teacher_heading_row = [[
                ar(f"الرقم الوظيفي : {block['employee_id'] or '-'}"),
                ar(f"المعلم : {block['teacher_name'] or '-'}"),
            ]]
            th_table = Table(teacher_heading_row, colWidths=[total_w * 0.4, total_w * 0.6])
            th_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), arabic_font),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1a2540')),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),   # الرقم الوظيفي يساراً
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),  # المعلم يميناً
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            block_elements.append(th_table)

            # --- جدول المقررات (رأس أزرق + صفوف + إجمالي) ---
            data = []
            # Header (يسار→يمين على PDF)
            data.append([
                ar("ملاحظات"),
                ar(f"الساعات ({weeks})"),
                ar("ساعات أسبوعية"),
                ar("الشعبة"),
                ar("المستوى"),
                ar("المقرر"),
                ar("الرمز"),
            ])
            # Courses
            note_row_indices = []
            for idx, c in enumerate(block["courses"], start=1):
                lvl = c.get("level")
                lvl_text = f"المستوى {lvl}" if lvl else "-"
                note_val = c.get("note") or ""
                if note_val:
                    note_row_indices.append(idx)
                data.append([
                    ar(note_val or "-"),
                    str(c["total_hours"]),
                    str(c["weekly_hours"]),
                    ar(c["section"] or "-"),
                    ar(lvl_text),
                    ar(c["course_name"] or "-"),
                    ar(c["course_code"] or "-"),
                ])
            # Total row
            data.append([
                "",
                str(block["total_semester"]),
                str(block["total_weekly"]),
                "",
                "",
                ar("*** الإجمالي ***"),
                "",
            ])

            tbl = Table(data, colWidths=col_widths)
            n_rows = len(data)
            last_idx = n_rows - 1
            style_cmds = [
                ('FONTNAME', (0, 0), (-1, -1), arabic_font),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565c0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                # Total row
                ('BACKGROUND', (0, last_idx), (-1, last_idx), colors.HexColor('#e3f2fd')),
                ('FONTSIZE', (0, last_idx), (-1, last_idx), 11),
                ('TEXTCOLOR', (0, last_idx), (-1, last_idx), colors.HexColor('#1a2540')),
                # Common
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cfd6e1')),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]
            # إبراز خانة ملاحظات لو فيها قيمة (تلوين برتقالي فاتح ونص أحمر)
            for rr in note_row_indices:
                style_cmds.append(('BACKGROUND', (0, rr), (0, rr), colors.HexColor('#fff3e0')))
                style_cmds.append(('TEXTCOLOR', (0, rr), (0, rr), colors.HexColor('#c62828')))
                style_cmds.append(('FONTSIZE', (0, rr), (0, rr), 10))
            tbl.setStyle(TableStyle(style_cmds))
            block_elements.append(tbl)
            block_elements.append(Spacer(1, 8 * mm))

            # نضمن أن كتلة المعلم لا تنقسم على صفحتين إذا كانت صغيرة
            elements.append(KeepTogether(block_elements))

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=teaching_load.pdf"}
    )


# ============================================================
# Teaching Load Templates - قوالب الأعباء التدريسية
# ============================================================
# الفكرة:
#  - حفظ snapshot لإسنادات فصل دراسي معيّن كـ "قالب"
#  - تطبيق هذا القالب لاحقاً على فصل دراسي آخر
#    (مطابقة المقررات بالكود + المستوى + الشعبة)


class TemplateSaveRequest(BaseModel):
    semester_id: str  # الفصل المصدر الذي ننسخ منه
    template_name: str  # اسم القالب
    term: Optional[str] = None  # first / second / summer (للتنظيم)
    department_id: Optional[str] = None  # لو null = كل الأقسام


class TemplateApplyRequest(BaseModel):
    template_id: str
    target_semester_id: str
    overwrite_existing: bool = False  # هل نستبدل الإسنادات الموجودة؟


@router.get("/teaching-load/templates")
async def list_templates(
    term: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """عرض كل القوالب المحفوظة"""
    if not has_permission(current_user, Permission.VIEW_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    db = get_db()
    query = {}
    if term:
        query["term"] = term

    templates = await db.teaching_load_templates.find(query).sort("created_at", -1).to_list(100)
    result = []
    for t in templates:
        result.append({
            "id": str(t["_id"]),
            "name": t.get("name", ""),
            "term": t.get("term", ""),
            "source_semester_name": t.get("source_semester_name", ""),
            "source_semester_id": t.get("source_semester_id", ""),
            "department_id": t.get("department_id"),
            "department_name": t.get("department_name", ""),
            "items_count": len(t.get("items", [])),
            "created_at": t.get("created_at").isoformat() if t.get("created_at") else None,
            "created_by_name": t.get("created_by_name", ""),
        })
    return result


@router.post("/teaching-load/templates")
async def save_template(
    payload: TemplateSaveRequest,
    current_user: dict = Depends(get_current_user),
):
    """حفظ إسنادات فصل دراسي كقالب"""
    if not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    db = get_db()

    # جلب الفصل المصدر
    src_sem = await db.semesters.find_one({"_id": ObjectId(payload.semester_id)})
    if not src_sem:
        raise HTTPException(status_code=404, detail="الفصل المصدر غير موجود")

    # بناء query للأعباء (مع فلتر القسم لو موجود)
    loads_query = {"semester_id": payload.semester_id}
    if payload.department_id:
        # نحتاج فلترة بالقسم → نجلب المعلمين أو نفلتر بالمقرر
        # الأبسط: نفلتر بالقسم بعد جلب الـ courses
        dept_courses = await db.courses.find(
            {"department_id": payload.department_id, "semester_id": payload.semester_id},
            {"_id": 1}
        ).to_list(2000)
        dept_course_ids = [str(c["_id"]) for c in dept_courses]
        loads_query["course_id"] = {"$in": dept_course_ids}

    loads = await db.teaching_loads.find(loads_query).to_list(2000)

    if not loads:
        raise HTTPException(status_code=400, detail="لا توجد إسنادات في هذا الفصل/القسم لحفظها")

    # بناء snapshot - نخزن بمعرفات يمكن مطابقتها (الكود + المستوى + الشعبة + اسم المعلم)
    items = []
    for load in loads:
        try:
            course = await db.courses.find_one({"_id": ObjectId(load["course_id"])})
            teacher = await db.teachers.find_one({"_id": ObjectId(load["teacher_id"])})
            if not course or not teacher:
                continue
            items.append({
                # معلومات المقرر للمطابقة
                "course_code": course.get("code", ""),
                "course_name": course.get("name", ""),
                "course_level": course.get("level"),
                "course_section": course.get("section", ""),
                "course_department_id": course.get("department_id", ""),
                "curriculum_course_id": course.get("curriculum_course_id"),  # احتياطي للمطابقة
                # معلومات المعلم للمطابقة
                "teacher_employee_id": teacher.get("employee_id", ""),
                "teacher_full_name": teacher.get("full_name", ""),
                "teacher_id_snapshot": str(teacher["_id"]),
                # بيانات الإسناد
                "weekly_hours": load.get("weekly_hours"),
                "notes": load.get("notes", ""),
            })
        except Exception:
            continue

    # جلب اسم القسم لو فلتر بقسم
    department_name = ""
    if payload.department_id:
        dept = await db.departments.find_one({"_id": ObjectId(payload.department_id)})
        if dept:
            department_name = dept.get("name", "")

    template_doc = {
        "name": payload.template_name,
        "term": payload.term or src_sem.get("term") or "",
        "source_semester_id": payload.semester_id,
        "source_semester_name": f"{src_sem.get('name','')} {src_sem.get('academic_year','')}",
        "department_id": payload.department_id,
        "department_name": department_name,
        "items": items,
        "created_at": datetime.now(timezone.utc),
        "created_by": current_user.get("id"),
        "created_by_name": current_user.get("full_name", current_user.get("username", "")),
    }

    result = await db.teaching_load_templates.insert_one(template_doc)

    await log_activity(
        current_user, "save_template", "teaching_load_template",
        str(result.inserted_id), None, {"name": payload.template_name, "items_count": len(items)}
    )

    return {
        "id": str(result.inserted_id),
        "message": f"تم حفظ القالب بنجاح ({len(items)} إسناد)",
        "items_count": len(items),
    }


@router.post("/teaching-load/templates/{template_id}/apply")
async def apply_template(
    template_id: str,
    payload: TemplateApplyRequest,
    current_user: dict = Depends(get_current_user),
):
    """تطبيق قالب على فصل دراسي مستهدف.
    يطابق المقررات بالكود + المستوى + الشعبة
    والمعلمين بـ employee_id (الأكثر استقراراً).
    """
    if not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    db = get_db()

    # جلب القالب
    template = await db.teaching_load_templates.find_one({"_id": ObjectId(template_id)})
    if not template:
        raise HTTPException(status_code=404, detail="القالب غير موجود")

    target_sem = await db.semesters.find_one({"_id": ObjectId(payload.target_semester_id)})
    if not target_sem:
        raise HTTPException(status_code=404, detail="الفصل المستهدف غير موجود")

    items = template.get("items", [])

    # جلب كل مقررات الفصل المستهدف لمرة واحدة
    target_courses = await db.courses.find(
        {"semester_id": payload.target_semester_id, "is_active": True}
    ).to_list(5000)

    # بناء فهرس للمطابقة
    courses_by_key = {}  # (code, level, section, department_id) -> course
    courses_by_curr = {}  # curriculum_course_id -> course
    for c in target_courses:
        key = (
            (c.get("code") or "").strip().upper(),
            c.get("level"),
            (c.get("section") or "").strip(),
            (c.get("department_id") or "").strip(),
        )
        courses_by_key[key] = c
        if c.get("curriculum_course_id"):
            courses_by_curr.setdefault(c["curriculum_course_id"], c)

    # جلب كل المعلمين النشطين لمرة واحدة
    teachers_all = await db.teachers.find({"is_active": {"$ne": False}}).to_list(2000)
    teachers_by_emp = {(t.get("employee_id") or "").strip(): t for t in teachers_all if t.get("employee_id")}
    teachers_by_name = {(t.get("full_name") or "").strip(): t for t in teachers_all if t.get("full_name")}

    # جلب الإسنادات الموجودة في الفصل المستهدف
    existing_loads = await db.teaching_loads.find(
        {"semester_id": payload.target_semester_id}
    ).to_list(5000)
    existing_loads_by_course = {l["course_id"]: l for l in existing_loads}

    # الإحصائيات
    created = 0
    updated = 0
    skipped_existing = 0
    no_course_match = []  # مقررات لم تطابق
    no_teacher_match = []  # معلمون لم يطابقوا (مفصول/متقاعد)
    matched_items = []  # تفاصيل الإسنادات المطبقة

    for item in items:
        # 1. مطابقة المقرر
        course = None
        # محاولة 1: بـ curriculum_course_id
        curr_id = item.get("curriculum_course_id")
        if curr_id and curr_id in courses_by_curr:
            course = courses_by_curr[curr_id]
        # محاولة 2: بـ الكود + المستوى + الشعبة + القسم
        if not course:
            key = (
                (item.get("course_code") or "").strip().upper(),
                item.get("course_level"),
                (item.get("course_section") or "").strip(),
                (item.get("course_department_id") or "").strip(),
            )
            course = courses_by_key.get(key)
        # محاولة 3: بدون قسم (شعبة وكود فقط)
        if not course:
            for c in target_courses:
                if (
                    (c.get("code") or "").strip().upper() == (item.get("course_code") or "").strip().upper()
                    and c.get("level") == item.get("course_level")
                    and (c.get("section") or "").strip() == (item.get("course_section") or "").strip()
                ):
                    course = c
                    break

        if not course:
            no_course_match.append({
                "code": item.get("course_code"),
                "name": item.get("course_name"),
                "level": item.get("course_level"),
                "section": item.get("course_section"),
            })
            continue

        # 2. مطابقة المعلم
        teacher = None
        emp_id = (item.get("teacher_employee_id") or "").strip()
        if emp_id and emp_id in teachers_by_emp:
            teacher = teachers_by_emp[emp_id]
        if not teacher:
            full_name = (item.get("teacher_full_name") or "").strip()
            if full_name and full_name in teachers_by_name:
                teacher = teachers_by_name[full_name]

        if not teacher:
            no_teacher_match.append({
                "course": f"{course.get('code')} {course.get('name')}",
                "teacher_name": item.get("teacher_full_name"),
                "employee_id": item.get("teacher_employee_id"),
            })
            continue

        course_id = str(course["_id"])
        teacher_id = str(teacher["_id"])

        # 3. التحقق من وجود إسناد سابق
        existing = existing_loads_by_course.get(course_id)
        if existing and not payload.overwrite_existing:
            skipped_existing += 1
            continue

        # 4. حفظ الإسناد
        load_doc = {
            "teacher_id": teacher_id,
            "course_id": course_id,
            "semester_id": payload.target_semester_id,
            "weekly_hours": item.get("weekly_hours"),
            "notes": item.get("notes", ""),
            "updated_at": datetime.now(timezone.utc),
            "updated_by": current_user.get("id"),
        }

        if existing:
            await db.teaching_loads.update_one(
                {"_id": existing["_id"]},
                {"$set": load_doc}
            )
            updated += 1
        else:
            load_doc["created_at"] = datetime.now(timezone.utc)
            load_doc["created_by"] = current_user.get("id")
            await db.teaching_loads.insert_one(load_doc)
            created += 1

        matched_items.append({
            "course": f"{course.get('code')} {course.get('name')} - شعبة {course.get('section','')}",
            "teacher": teacher.get("full_name", ""),
            "hours": item.get("weekly_hours"),
        })

    await log_activity(
        current_user, "apply_template", "teaching_load_template",
        template_id, None, {
            "target_semester_id": payload.target_semester_id,
            "created": created, "updated": updated, "skipped": skipped_existing,
        }
    )

    return {
        "message": f"تم تطبيق القالب: {created} إسناد جديد، {updated} تحديث، {skipped_existing} متجاهل",
        "stats": {
            "created": created,
            "updated": updated,
            "skipped_existing": skipped_existing,
            "no_course_match_count": len(no_course_match),
            "no_teacher_match_count": len(no_teacher_match),
            "total_items": len(items),
        },
        "no_course_match": no_course_match[:50],  # أول 50 فقط
        "no_teacher_match": no_teacher_match[:50],
        "applied_count": len(matched_items),
    }


@router.delete("/teaching-load/templates/{template_id}")
async def delete_template(
    template_id: str,
    current_user: dict = Depends(get_current_user),
):
    """حذف قالب"""
    if not has_permission(current_user, Permission.MANAGE_TEACHING_LOAD):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    db = get_db()
    result = await db.teaching_load_templates.delete_one({"_id": ObjectId(template_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="القالب غير موجود")

    await log_activity(
        current_user, "delete_template", "teaching_load_template",
        template_id, None, None
    )
    return {"message": "تم حذف القالب"}
