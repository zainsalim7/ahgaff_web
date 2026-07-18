"""
Curriculum Routes - الخطة الدراسية وإسناد المعلمين
Layer 1: curriculum_courses - تعريفات ثابتة للمقررات (الخطة الدراسية)
Layer 2: teacher_assignments - الإسناد الدائم للمعلم على مقرر من الخطة
Layer 3: courses (موجود) - الجلسات الفعلية في فصل أكاديمي محدد

✨ Auto-Deploy Test: 2026-06-02 - GitHub Actions
"""
from datetime import datetime, timezone
from typing import Optional, List
from io import BytesIO

import pandas as pd
from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from .deps import get_current_user, get_db
from models.permissions import Permission

router = APIRouter(tags=["الخطة الدراسية"])


def _has_manage(user: dict) -> bool:
    """يقبل صلاحية الخطة الدراسية الخاصة أو صلاحية إدارة المقررات العامة."""
    if user.get("role") == "admin":
        return True
    perms = set(user.get("permissions") or [])
    return (
        Permission.MANAGE_CURRICULUM in perms
        or Permission.MANAGE_COURSES in perms
    )


def _has_view(user: dict) -> bool:
    """يقبل صلاحية الخطة (إدارة/عرض)، أو إدارة/عرض المقررات العامة."""
    if user.get("role") == "admin":
        return True
    perms = set(user.get("permissions") or [])
    return any(p in perms for p in [
        Permission.MANAGE_CURRICULUM,
        Permission.VIEW_CURRICULUM,
        Permission.MANAGE_COURSES,
        Permission.VIEW_COURSES,
    ])


def _clean_doc(d: dict) -> dict:
    if not d:
        return d
    out = dict(d)
    if "_id" in out:
        out["id"] = str(out.pop("_id"))
    for k, v in list(out.items()):
        if isinstance(v, ObjectId):
            out[k] = str(v)
    return out


def _now():
    return datetime.now(timezone.utc).isoformat()


# ==================== Models ====================

class CurriculumCourseCreate(BaseModel):
    code: str
    name: str
    name_en: Optional[str] = None
    credit_hours: int = 3
    weekly_hours: Optional[float] = None  # 🆕 الساعات الأسبوعية (تُستخدم كافتراضي عند الإسناد)
    faculty_id: str
    department_id: str
    level: int = Field(..., ge=1, le=10)
    term: int = Field(..., ge=1, le=3)  # 1=أول، 2=ثاني، 3=صيفي
    description: Optional[str] = None
    prerequisites: List[str] = []


class CurriculumCourseUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    name_en: Optional[str] = None
    credit_hours: Optional[int] = None
    weekly_hours: Optional[float] = None  # 🆕
    faculty_id: Optional[str] = None
    department_id: Optional[str] = None
    level: Optional[int] = None
    term: Optional[int] = Field(None, ge=1, le=3)
    description: Optional[str] = None
    prerequisites: Optional[List[str]] = None
    sections_count: Optional[int] = Field(None, ge=0, le=10)  # عدد الشعب المحفوظ (0 = بدون شعب)
    is_active: Optional[bool] = None


class TeacherAssignmentCreate(BaseModel):
    teacher_id: str
    curriculum_course_id: str
    notes: Optional[str] = None


# ==================== Curriculum Endpoints ====================

@router.get("/curriculum/courses")
async def list_curriculum_courses(
    department_id: Optional[str] = None,
    faculty_id: Optional[str] = None,
    level: Optional[int] = None,
    term: Optional[int] = None,
    include_inactive: bool = False,
    current_user: dict = Depends(get_current_user),
):
    """قائمة مقررات الخطة الدراسية مع فلاتر."""
    if not _has_view(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    db = get_db()
    q = {}
    if not include_inactive:
        q["is_active"] = {"$ne": False}
    if department_id:
        q["department_id"] = department_id
    if faculty_id:
        q["faculty_id"] = faculty_id
    if level is not None:
        q["level"] = level
    if term is not None:
        q["term"] = term

    items = []
    async for c in db.curriculum_courses.find(q).sort([("level", 1), ("term", 1), ("name", 1)]):
        items.append(_clean_doc(c))
    return {"items": items, "total": len(items)}


@router.get("/curriculum/by-department/{department_id}")
async def get_curriculum_by_department(
    department_id: str,
    current_user: dict = Depends(get_current_user),
):
    """خطة قسم منظمة في شبكة (مستوى × فصل) مع المعلمين المُسنَدين."""
    if not _has_view(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    db = get_db()

    # جلب القسم والكلية
    try:
        dept = await db.departments.find_one({"_id": ObjectId(department_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرّف القسم غير صحيح")
    if not dept:
        raise HTTPException(status_code=404, detail="القسم غير موجود")

    faculty = None
    levels_count = 5
    if dept.get("faculty_id"):
        try:
            faculty = await db.faculties.find_one({"_id": ObjectId(dept["faculty_id"])})
            if faculty:
                levels_count = int(faculty.get("levels_count") or 5)
        except Exception:
            pass

    # جلب الخطة
    courses = []
    async for c in db.curriculum_courses.find({
        "department_id": department_id, "is_active": {"$ne": False}
    }):
        courses.append(_clean_doc(c))

    # جلب الإسنادات لكل مقررات الخطة في هذا القسم
    course_ids = [c["id"] for c in courses]
    assignments = {}
    if course_ids:
        async for a in db.teacher_assignments.find({
            "curriculum_course_id": {"$in": course_ids}, "is_active": {"$ne": False}
        }):
            assignments.setdefault(a.get("curriculum_course_id"), []).append(_clean_doc(a))

    # جلب أسماء المعلمين
    teacher_ids = list({
        a.get("teacher_id") for arr in assignments.values() for a in arr if a.get("teacher_id")
    })
    teachers_map = {}
    if teacher_ids:
        try:
            tobj = [ObjectId(t) for t in teacher_ids if t]
            async for t in db.teachers.find({"_id": {"$in": tobj}}):
                teachers_map[str(t["_id"])] = {
                    "id": str(t["_id"]), "full_name": t.get("full_name", ""),
                    "teacher_id": t.get("teacher_id", ""),
                    "academic_title": t.get("academic_title"),
                }
        except Exception:
            pass

    # تنظيم في شبكة
    grid = {}  # grid[level][term] = [courses]  (term: 1=أول، 2=ثاني، 3=صيفي)
    for lvl in range(1, levels_count + 1):
        grid[lvl] = {1: [], 2: [], 3: []}
    for c in courses:
        lvl = int(c.get("level") or 0)
        term = int(c.get("term") or 0)
        if lvl in grid and term in grid[lvl]:
            # إثراء بإسناد المعلم
            course_assignments = assignments.get(c["id"], [])
            c["teachers"] = [
                teachers_map.get(a.get("teacher_id"), {"id": a.get("teacher_id"), "full_name": "(غير معروف)"})
                for a in course_assignments
            ]
            grid[lvl][term].append(c)

    # تحويل إلى قائمة مرتبة
    grid_list = []
    for lvl in range(1, levels_count + 1):
        grid_list.append({
            "level": lvl,
            "term1": grid[lvl][1],
            "term2": grid[lvl][2],
            "term3": grid[lvl][3],
        })

    return {
        "department": {
            "id": department_id,
            "name": dept.get("name", ""),
            "code": dept.get("code", ""),
            "faculty_id": dept.get("faculty_id"),
            "faculty_name": (faculty or {}).get("name"),
            "levels_count": levels_count,
        },
        "grid": grid_list,
        "total_courses": len(courses),
    }


@router.get("/curriculum/courses/{course_id}")
async def get_curriculum_course(
    course_id: str,
    current_user: dict = Depends(get_current_user),
):
    """تفاصيل مقرر من الخطة + المُعلمون المُسنَدون + إحصائيات."""
    if not _has_view(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    db = get_db()
    try:
        c = await db.curriculum_courses.find_one({"_id": ObjectId(course_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرّف غير صحيح")
    if not c:
        raise HTTPException(status_code=404, detail="المقرر غير موجود في الخطة")

    course = _clean_doc(c)

    # الإسنادات
    assignments = []
    teacher_ids = []
    async for a in db.teacher_assignments.find({
        "curriculum_course_id": course_id, "is_active": {"$ne": False}
    }):
        ad = _clean_doc(a)
        assignments.append(ad)
        if ad.get("teacher_id"):
            teacher_ids.append(ad["teacher_id"])

    teachers = []
    if teacher_ids:
        try:
            tobj = [ObjectId(t) for t in teacher_ids if t]
            async for t in db.teachers.find({"_id": {"$in": tobj}}):
                teachers.append({
                    "id": str(t["_id"]), "full_name": t.get("full_name", ""),
                    "teacher_id": t.get("teacher_id", ""),
                    "academic_title": t.get("academic_title"),
                })
        except Exception:
            pass

    # عدد الجلسات النشطة في هذا الفصل
    active_offerings = await db.courses.count_documents({"curriculum_course_id": course_id})

    return {
        "course": course,
        "teachers": teachers,
        "assignments": assignments,
        "active_offerings": active_offerings,
    }


@router.post("/curriculum/courses")
async def create_curriculum_course(
    payload: CurriculumCourseCreate,
    current_user: dict = Depends(get_current_user),
):
    if not _has_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    db = get_db()
    # فحص تكرار الكود في نفس القسم
    existing = await db.curriculum_courses.find_one({
        "code": payload.code, "department_id": payload.department_id,
        "is_active": {"$ne": False},
    })
    if existing:
        raise HTTPException(status_code=409, detail=f"الكود '{payload.code}' موجود مسبقاً في هذا القسم")

    doc = payload.dict()
    doc["is_active"] = True
    doc["created_at"] = _now()
    doc["created_by"] = current_user.get("id")
    result = await db.curriculum_courses.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None)
    return {"message": "تم إنشاء المقرر في الخطة", "course": doc}


@router.put("/curriculum/courses/{course_id}")
async def update_curriculum_course(
    course_id: str,
    payload: CurriculumCourseUpdate,
    current_user: dict = Depends(get_current_user),
):
    if not _has_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    db = get_db()
    try:
        oid = ObjectId(course_id)
    except Exception:
        raise HTTPException(status_code=400, detail="معرّف غير صحيح")

    update_data = {k: v for k, v in payload.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="لا توجد بيانات للتحديث")
    update_data["updated_at"] = _now()
    update_data["updated_by"] = current_user.get("id")
    result = await db.curriculum_courses.update_one({"_id": oid}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    return {"message": "تم التحديث"}


@router.delete("/curriculum/courses/{course_id}")
async def delete_curriculum_course(
    course_id: str,
    current_user: dict = Depends(get_current_user),
):
    """حذف ناعم (is_active=False) للحفاظ على الروابط القديمة."""
    if not _has_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    db = get_db()
    try:
        oid = ObjectId(course_id)
    except Exception:
        raise HTTPException(status_code=400, detail="معرّف غير صحيح")
    result = await db.curriculum_courses.update_one(
        {"_id": oid}, {"$set": {"is_active": False, "deleted_at": _now()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="المقرر غير موجود")
    # تعليم الإسنادات كغير نشطة
    await db.teacher_assignments.update_many(
        {"curriculum_course_id": course_id},
        {"$set": {"is_active": False, "ended_at": _now()}}
    )
    return {"message": "تم حذف المقرر من الخطة"}


# ==================== Teacher Assignments ====================

@router.get("/curriculum/assignments")
async def list_assignments(
    teacher_id: Optional[str] = None,
    curriculum_course_id: Optional[str] = None,
    include_inactive: bool = False,
    current_user: dict = Depends(get_current_user),
):
    if not _has_view(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    db = get_db()
    q = {}
    if not include_inactive:
        q["is_active"] = {"$ne": False}
    if teacher_id:
        q["teacher_id"] = teacher_id
    if curriculum_course_id:
        q["curriculum_course_id"] = curriculum_course_id
    items = []
    async for a in db.teacher_assignments.find(q):
        items.append(_clean_doc(a))
    return {"items": items, "total": len(items)}


@router.post("/curriculum/assignments")
async def create_assignment(
    payload: TeacherAssignmentCreate,
    current_user: dict = Depends(get_current_user),
):
    if not _has_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    db = get_db()
    # فحص تكرار
    existing = await db.teacher_assignments.find_one({
        "teacher_id": payload.teacher_id,
        "curriculum_course_id": payload.curriculum_course_id,
        "is_active": {"$ne": False},
    })
    if existing:
        raise HTTPException(status_code=409, detail="هذا الإسناد موجود مسبقاً")

    doc = payload.dict()
    doc["is_active"] = True
    doc["assigned_at"] = _now()
    doc["assigned_by"] = current_user.get("id")
    doc["ended_at"] = None
    result = await db.teacher_assignments.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None)
    return {"message": "تم إسناد المقرر للمعلم", "assignment": doc}


@router.delete("/curriculum/assignments/{assignment_id}")
async def remove_assignment(
    assignment_id: str,
    current_user: dict = Depends(get_current_user),
):
    """إلغاء إسناد (soft - يبقى السجل للتاريخ)."""
    if not _has_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    db = get_db()
    try:
        oid = ObjectId(assignment_id)
    except Exception:
        raise HTTPException(status_code=400, detail="معرّف غير صحيح")
    result = await db.teacher_assignments.update_one(
        {"_id": oid}, {"$set": {"is_active": False, "ended_at": _now()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="الإسناد غير موجود")
    return {"message": "تم إلغاء الإسناد"}


# ==================== Generate Offerings ====================

SECTION_LABELS_AR = ['أ', 'ب', 'ج', 'د', 'ه', 'و', 'ز', 'ح', 'ط', 'ي']


class GenerateOfferingsBody(BaseModel):
    """جسم الطلب الاختياري لتمرير عدد الشعب لكل مقرر."""
    sections_map: Optional[dict[str, int]] = None  # curriculum_course_id → عدد الشعب
    default_sections: Optional[int] = 1  # عدد افتراضي للمقررات غير المحددة


@router.post("/curriculum/generate-offerings")
async def generate_offerings_from_curriculum(
    semester_id: str = Query(..., description="معرّف الفصل الأكاديمي النشط"),
    department_id: Optional[str] = Query(None, description="فلترة قسم محدد"),
    term: Optional[int] = Query(None, ge=1, le=3, description="1=أول، 2=ثاني، 3=صيفي (افتراضي يأخذ من term الفصل)"),
    skip_existing: bool = Query(True, description="تخطي المقررات الموجودة في الفصل"),
    body: Optional[GenerateOfferingsBody] = None,
    current_user: dict = Depends(get_current_user),
):
    """توليد جلسات (Layer 3 courses) من الخطة الدراسية لفصل أكاديمي محدد.
    - يأخذ مقررات الخطة المطابقة للفلاتر
    - يُنشئ courses (Layer 3) لكل واحد بعدد الشعب المحدد (افتراضي: شعبة واحدة "أ")
    - لكل مقرر: ينشئ N courses منفصلة بشعب أ، ب، ج، ...
    - يربطها بـ curriculum_course_id
    - يأخذ المعلم من أحدث teacher_assignment إن وُجد
    """
    if not _has_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    db = get_db()

    # ⭐ خريطة الشعب لكل مقرر (curriculum_course_id → عدد الشعب)
    sections_map = (body.sections_map if body else None) or {}
    default_sections = (body.default_sections if body else 1) or 1
    # تطبيق حدود معقولة
    default_sections = max(1, min(default_sections, len(SECTION_LABELS_AR)))

    # تحقق من الفصل
    try:
        sem = await db.semesters.find_one({"_id": ObjectId(semester_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرّف الفصل غير صحيح")
    if not sem:
        raise HTTPException(status_code=404, detail="الفصل غير موجود")

    # تحديد term: أولوية: query param → semester.term → استنتاج من اسم الفصل (fallback)
    auto_term = sem.get("term")  # ← المصدر الموثوق (يُحدَّد عند إنشاء الفصل)
    if auto_term is None:
        # Fallback مؤقت للفصول القديمة قبل إضافة الحقل
        semester_name = (sem.get("name") or "").strip()
        if "الأول" in semester_name or "الاول" in semester_name or "first" in semester_name.lower():
            auto_term = 1
        elif "الثاني" in semester_name or "ثاني" in semester_name or "second" in semester_name.lower():
            auto_term = 2
        elif "صيف" in semester_name or "summer" in semester_name.lower():
            auto_term = 3
        # حفظ القيمة المستنتجة في الفصل ليصبح موثوقاً للمرات القادمة
        if auto_term is not None:
            try:
                await db.semesters.update_one(
                    {"_id": ObjectId(semester_id)},
                    {"$set": {"term": auto_term}}
                )
            except Exception:
                pass

    # الخطة المُختارة
    q = {"is_active": {"$ne": False}}
    if department_id:
        q["department_id"] = department_id
    # أولوية: term من الـ query (إن مُرّر صراحة)، ثم term المستنتج من الفصل
    effective_term = term if term is not None else auto_term
    # فلترة بـ term عند كل القيم المعروفة (1=أول، 2=ثاني، 3=صيفي)
    if effective_term is not None and effective_term in (1, 2, 3):
        q["term"] = effective_term

    curriculum_list = []
    async for c in db.curriculum_courses.find(q):
        curriculum_list.append(c)

    # الإسنادات النشطة
    cc_ids = [str(c["_id"]) for c in curriculum_list]
    assignments_map = {}
    async for a in db.teacher_assignments.find({
        "curriculum_course_id": {"$in": cc_ids}, "is_active": {"$ne": False}
    }):
        # نأخذ أول إسناد فقط (للبساطة)
        ccid = a.get("curriculum_course_id")
        if ccid and ccid not in assignments_map:
            assignments_map[ccid] = a.get("teacher_id")

    created, skipped, updated = 0, 0, 0
    updated_course_ids: list[str] = []
    created_items = []
    for cc in curriculum_list:
        cc_id = str(cc["_id"])
        # تحديد عدد الشعب لهذا المقرر (الخريطة المرسلة ← ثم المحفوظ على المقرر ← ثم الافتراضي)
        if cc_id in sections_map:
            n_sections = sections_map[cc_id]
        else:
            stored = cc.get("sections_count")
            n_sections = stored if (stored and int(stored) > 1) else default_sections
        n_sections = max(1, min(int(n_sections or 1), len(SECTION_LABELS_AR)))

        for sec_idx in range(n_sections):
            # 🆕 إذا كانت شعبة واحدة فقط، لا نُضيف حرفاً (سيكون section فارغاً)
            section_label = "" if n_sections == 1 else SECTION_LABELS_AR[sec_idx]

            if skip_existing:
                exists = await db.courses.find_one({
                    "semester_id": semester_id,
                    "curriculum_course_id": cc_id,
                    "section": section_label,
                })
                if exists:
                    # 🔧 بدلاً من تجاهل المقرر الموجود، نحدّث الساعات والاسم/الكود من المصدر
                    # حتى تنتشر تعديلات الخطة (مثل تعديل الساعات المعتمدة) إلى المقررات والإسنادات
                    new_credit = cc.get("credit_hours", 3)
                    new_weekly = cc.get("weekly_hours")
                    needs_update = (
                        exists.get("credit_hours") != new_credit
                        or exists.get("weekly_hours") != new_weekly
                        or exists.get("name") != cc.get("name")
                        or exists.get("code") != cc.get("code")
                    )
                    if needs_update:
                        await db.courses.update_one(
                            {"_id": exists["_id"]},
                            {"$set": {
                                "credit_hours": new_credit,
                                "weekly_hours": new_weekly,
                                "name": cc.get("name"),
                                "code": cc.get("code"),
                                "updated_at": _now(),
                            }}
                        )
                        # 🔧 وأيضاً تحديث teaching_loads.weekly_hours لكل المعلمين المرتبطين بهذا المقرر
                        try:
                            await db.teaching_loads.update_many(
                                {"course_id": str(exists["_id"])},
                                {"$set": {"weekly_hours": (new_weekly or new_credit), "updated_at": _now()}}
                            )
                        except Exception:
                            pass
                        updated += 1
                        updated_course_ids.append(str(exists["_id"]))
                    else:
                        skipped += 1
                    continue

            doc = {
                "curriculum_course_id": cc_id,
                "semester_id": semester_id,
                "code": cc.get("code"),
                "name": cc.get("name"),
                "credit_hours": cc.get("credit_hours", 3),
                "weekly_hours": cc.get("weekly_hours"),  # 🆕 نسخ الساعات الأسبوعية من الخطة
                "department_id": cc.get("department_id"),
                "faculty_id": cc.get("faculty_id"),
                "level": cc.get("level"),
                "term": cc.get("term"),
                "section": section_label,
                "room": "",
                # الإسناد التلقائي للشعبة الأولى فقط (تفادي تعارض المقرر الواحد لمعلمين)
                "teacher_id": assignments_map.get(cc_id) if sec_idx == 0 else None,
                "is_active": True,
                "academic_year": sem.get("academic_year") or "",
                "created_at": _now(),
                "created_by": current_user.get("id"),
                "auto_generated": True,
            }
            r = await db.courses.insert_one(doc)
            created += 1
            created_items.append({
                "id": str(r.inserted_id),
                "name": doc["name"],
                "code": doc["code"],
                "section": section_label,
            })

    return {
        "message": f"تم توليد {created} جلسة من الخطة الدراسية" + (f" وتحديث {updated} مقرر موجود" if updated else ""),
        "semester_id": semester_id,
        "semester_name": sem.get("name"),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total_curriculum": len(curriculum_list),
        "items": created_items[:50],  # عينة للعرض
    }


@router.post("/curriculum/fix-auto-generated-visibility")
async def fix_auto_generated_visibility(
    current_user: dict = Depends(get_current_user),
):
    """إصلاح فوري: يضع is_active=True على كل المقررات المُولّدة تلقائياً.
    سبب: نسخة سابقة من generate-offerings لم تضع is_active، فأخفاها فلتر GET /courses.
    """
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="هذه العملية للأدمن فقط")
    db = get_db()

    # عدد المقررات قبل
    total_auto = await db.courses.count_documents({"auto_generated": True})
    hidden = await db.courses.count_documents({
        "auto_generated": True,
        "$or": [{"is_active": {"$exists": False}}, {"is_active": {"$ne": True}}],
    })

    # تحديث
    result = await db.courses.update_many(
        {
            "auto_generated": True,
            "$or": [{"is_active": {"$exists": False}}, {"is_active": {"$ne": True}}],
        },
        {"$set": {"is_active": True}},
    )

    # سجل
    try:
        await db.activity_logs.insert_one({
            "user_id": current_user.get("id"),
            "username": current_user.get("username"),
            "action": "fix_auto_generated_visibility",
            "target_type": "courses",
            "details": f"إصلاح إظهار {result.modified_count} مقرر مولد كانت مخفية",
            "timestamp": _now(),
        })
    except Exception:
        pass

    return {
        "message": f"تم إصلاح {result.modified_count} مقرر — الآن ستظهر في صفحة المقررات",
        "total_auto_generated": total_auto,
        "was_hidden": hidden,
        "now_fixed": result.modified_count,
    }




@router.get("/curriculum/auto-generated-courses")
async def list_auto_generated_courses(
    semester_id: Optional[str] = None,
    department_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """عرض المقررات المُولّدة تلقائياً من الخطة (للتنظيف/المراجعة)."""
    if not _has_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    db = get_db()
    q: dict = {"auto_generated": True}
    if semester_id:
        q["semester_id"] = semester_id
    if department_id:
        q["department_id"] = department_id
    items = []
    async for c in db.courses.find(q).limit(500):
        items.append({
            "id": str(c["_id"]),
            "code": c.get("code"),
            "name": c.get("name"),
            "semester_id": c.get("semester_id"),
            "department_id": c.get("department_id"),
            "curriculum_course_id": c.get("curriculum_course_id"),
            "level": c.get("level"),
            "term": c.get("term"),
            "section": c.get("section"),
        })
    return {"items": items, "total": len(items), "filter": q}


@router.delete("/curriculum/auto-generated-courses")
async def delete_auto_generated_courses(
    semester_id: Optional[str] = None,
    department_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """حذف المقررات المُولّدة تلقائياً (تنظيف بعد التوليد الخاطئ).
    - يحذف فقط المقررات التي auto_generated=True
    - يحترم فلاتر semester_id و department_id
    - لا يحذف المقررات الأصلية اليدوية
    """
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="هذه العملية للأدمن فقط")
    db = get_db()
    q: dict = {"auto_generated": True}
    if semester_id:
        q["semester_id"] = semester_id
    if department_id:
        q["department_id"] = department_id

    # حماية: لازم يكون فلتر على الأقل
    if not semester_id and not department_id:
        raise HTTPException(
            status_code=400,
            detail="يجب تحديد semester_id أو department_id على الأقل لتجنب الحذف الشامل"
        )

    # سجل قبل الحذف
    to_delete = []
    async for c in db.courses.find(q, {"_id": 1, "code": 1, "name": 1}):
        to_delete.append({"id": str(c["_id"]), "code": c.get("code"), "name": c.get("name")})

    result = await db.courses.delete_many(q)

    # سجل النشاط
    try:
        await db.activity_logs.insert_one({
            "user_id": current_user.get("id"),
            "username": current_user.get("username"),
            "action": "cleanup_auto_generated_courses",
            "target_type": "courses",
            "details": f"حذف {result.deleted_count} مقرر مولد تلقائياً",
            "filter": q,
            "deleted_items": to_delete[:50],
            "timestamp": _now(),
        })
    except Exception:
        pass

    return {
        "message": f"تم حذف {result.deleted_count} مقرر مُولّد تلقائياً",
        "deleted_count": result.deleted_count,
        "deleted_items": to_delete,
        "filter": q,
    }




# ==================== Backfill from Active Courses ====================

@router.post("/curriculum/backfill-from-active")
async def backfill_curriculum_from_active_courses(
    current_user: dict = Depends(get_current_user),
):
    """يبني الخطة الدراسية من المقررات الموجودة حالياً في collection `courses`.
    - لكل مقرر مفرَد (code + department_id + level) يُنشَئ curriculum_course
    - يُحدَّث المقرر الأصلي بـ curriculum_course_id ليرتبط بالخطة
    - يُسجَّل teacher_assignment لكل مقرر له معلم
    - لا يكرر الإدخال (skip_existing)
    """
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="هذه العملية متاحة للأدمن فقط")
    db = get_db()

    # المقررات الموجودة في الخطة (لتجنّب التكرار)
    existing = set()
    async for cc in db.curriculum_courses.find({}, {"code": 1, "department_id": 1, "level": 1, "term": 1}):
        existing.add((cc.get("code"), cc.get("department_id"), int(cc.get("level") or 0), int(cc.get("term") or 0)))

    # تجميع المقررات الفعلية بحسب (code, dept, level, term)
    grouped = {}
    async for c in db.courses.find({}):
        code = c.get("code")
        dept = c.get("department_id")
        lvl = int(c.get("level") or 1)
        # محاولة تحديد term من بيانات المقرر أو الفصل النشط (افتراضي 1)
        term = int(c.get("term") or 1)
        if not code or not dept:
            continue
        key = (code, dept, lvl, term)
        if key not in grouped:
            grouped[key] = {"sample": c, "course_ids": []}
        grouped[key]["course_ids"].append(str(c["_id"]))

    created = 0
    skipped = 0
    assignments_created = 0
    linked = 0

    for key, val in grouped.items():
        code, dept, lvl, term = key
        c = val["sample"]
        if key in existing:
            skipped += 1
            # حتى لو موجود، اربط الـ courses الأصلية بـ curriculum_course_id
            cc_existing = await db.curriculum_courses.find_one({
                "code": code, "department_id": dept,
                "level": lvl, "term": term,
            })
            if cc_existing:
                for cid in val["course_ids"]:
                    try:
                        await db.courses.update_one(
                            {"_id": ObjectId(cid), "curriculum_course_id": {"$exists": False}},
                            {"$set": {"curriculum_course_id": str(cc_existing["_id"])}}
                        )
                        linked += 1
                    except Exception:
                        pass
            continue

        # faculty من القسم
        faculty_id = c.get("faculty_id")
        if not faculty_id:
            try:
                d = await db.departments.find_one({"_id": ObjectId(dept)})
                if d:
                    faculty_id = d.get("faculty_id")
            except Exception:
                pass

        doc = {
            "code": code,
            "name": c.get("name") or code,
            "credit_hours": int(c.get("credit_hours") or 3),
            "department_id": dept,
            "faculty_id": faculty_id,
            "level": lvl,
            "term": term,
            "is_active": True,
            "created_at": _now(),
            "created_by": current_user.get("id"),
            "imported_from": "active_courses",
        }
        r = await db.curriculum_courses.insert_one(doc)
        cc_id = str(r.inserted_id)
        created += 1
        existing.add(key)

        # ربط كل المقررات الفعلية بهذا curriculum_course
        for cid in val["course_ids"]:
            try:
                await db.courses.update_one(
                    {"_id": ObjectId(cid)},
                    {"$set": {"curriculum_course_id": cc_id}}
                )
                linked += 1
            except Exception:
                pass

        # إنشاء teacher_assignment للمعلم المرتبط (من أحدث instance)
        teachers_seen = set()
        for cid in val["course_ids"]:
            try:
                cdoc = await db.courses.find_one({"_id": ObjectId(cid)})
                tid = cdoc.get("teacher_id") if cdoc else None
                if tid and tid not in teachers_seen:
                    teachers_seen.add(tid)
                    exists_a = await db.teacher_assignments.find_one({
                        "teacher_id": tid, "curriculum_course_id": cc_id,
                        "is_active": {"$ne": False},
                    })
                    if not exists_a:
                        await db.teacher_assignments.insert_one({
                            "teacher_id": tid,
                            "curriculum_course_id": cc_id,
                            "is_active": True,
                            "assigned_at": _now(),
                            "assigned_by": current_user.get("id"),
                            "ended_at": None,
                            "notes": "مستورد من المقررات النشطة",
                        })
                        assignments_created += 1
            except Exception:
                pass

    return {
        "message": f"تم إنشاء {created} مقرر في الخطة، تخطي {skipped}، ربط {linked} مقرر فعلي",
        "created": created,
        "skipped_existing": skipped,
        "linked_active_courses": linked,
        "assignments_created": assignments_created,
        "total_groups": len(grouped),
    }


# ==================== Import from Archive ====================

@router.post("/curriculum/import-from-archive/{semester_id}")
async def import_curriculum_from_archive(
    semester_id: str,
    current_user: dict = Depends(get_current_user),
):
    """يبني خطة دراسية من فصل مؤرشف (لو فاتك إنشاء الخطة قبل الأرشفة).
    - يقرأ المقررات المؤرشفة في الفصل
    - لكل مقرر فريد (code + dept + level + term افتراضي 1) ينشئ curriculum_course
    - تخطي المقررات الموجودة مسبقاً في الخطة
    """
    if not _has_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    db = get_db()
    archive = await db.semester_archives.find_one({"semester_id": semester_id})
    if not archive:
        raise HTTPException(status_code=404, detail="الفصل غير مؤرشف")

    archived_courses = archive.get("courses", [])
    if not archived_courses:
        return {"message": "لا توجد مقررات في الأرشيف", "imported": 0}

    # المقررات الموجودة في الخطة
    existing_codes = set()
    async for c in db.curriculum_courses.find({}, {"code": 1, "department_id": 1}):
        existing_codes.add((c.get("code"), c.get("department_id")))

    imported = 0
    skipped = 0
    failed = 0
    # محاولة تحديد term من اسم الفصل (افتراضي 1)
    sem_name = (archive.get("semester_name") or "").strip()
    default_term = 2 if any(k in sem_name for k in ["الثاني", "ثاني", "2", "II"]) else 1

    teacher_links = []  # لتطبيق teacher_assignments بعد إنشاء الخطة

    for c in archived_courses:
        code = c.get("code")
        dept_id = c.get("department_id")
        if not code or not dept_id:
            failed += 1
            continue
        if (code, dept_id) in existing_codes:
            skipped += 1
            continue

        # محاولة قراءة faculty من القسم
        faculty_id = None
        try:
            d = await db.departments.find_one({"_id": ObjectId(dept_id)})
            if d:
                faculty_id = d.get("faculty_id")
        except Exception:
            pass

        doc = {
            "code": code,
            "name": c.get("name") or code,
            "credit_hours": int(c.get("credit_hours") or 3),
            "department_id": dept_id,
            "faculty_id": faculty_id,
            "level": int(c.get("level") or 1),
            "term": default_term,
            "is_active": True,
            "created_at": _now(),
            "created_by": current_user.get("id"),
            "imported_from_archive": semester_id,
        }
        r = await db.curriculum_courses.insert_one(doc)
        existing_codes.add((code, dept_id))
        imported += 1
        # تسجيل إسناد المعلم لو موجود
        if c.get("teacher_id"):
            teacher_links.append((str(r.inserted_id), c["teacher_id"]))

    # إنشاء teacher_assignments
    assignments_created = 0
    for cc_id, teacher_id in teacher_links:
        exists = await db.teacher_assignments.find_one({
            "curriculum_course_id": cc_id, "teacher_id": teacher_id,
            "is_active": {"$ne": False},
        })
        if not exists:
            await db.teacher_assignments.insert_one({
                "teacher_id": teacher_id,
                "curriculum_course_id": cc_id,
                "is_active": True,
                "assigned_at": _now(),
                "assigned_by": current_user.get("id"),
                "ended_at": None,
                "notes": f"مستورد من أرشيف الفصل: {sem_name}",
            })
            assignments_created += 1

    return {
        "message": f"تم استيراد {imported} مقرر إلى الخطة الدراسية",
        "imported": imported,
        "skipped_existing": skipped,
        "failed": failed,
        "assignments_created": assignments_created,
        "from_semester": sem_name,
        "default_term_used": default_term,
    }


# ==================== Wipe Department Curriculum ====================

@router.delete("/curriculum/department/{department_id}/wipe")
async def wipe_department_curriculum(
    department_id: str,
    level: Optional[int] = Query(None, description="مسح مستوى معين فقط (اختياري)"),
    term: Optional[int] = Query(None, description="مسح فصل معين فقط 1/2/3 (اختياري)"),
    current_user: dict = Depends(get_current_user),
):
    """مسح خطة قسم (حذف ناعم — يحفظ نسخة في trash للاسترجاع).
    - بدون معاملات: يمسح كل الخطة
    - مع level: يمسح المستوى المحدد فقط
    - مع term: يمسح الفصل المحدد فقط
    - يمكن دمج level و term معاً (مستوى + فصل محدد)
    """
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="هذه العملية متاحة للأدمن فقط")
    db = get_db()
    try:
        dept = await db.departments.find_one({"_id": ObjectId(department_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرف القسم غير صحيح")
    if not dept:
        raise HTTPException(status_code=404, detail="القسم غير موجود")

    # بناء فلتر الاستعلام مع المعاملات الاختيارية
    base_filter: dict = {"department_id": department_id, "is_active": {"$ne": False}}
    scope_label_parts = []
    if level is not None:
        base_filter["level"] = level
        scope_label_parts.append(f"المستوى {level}")
    if term is not None:
        base_filter["term"] = term
        term_name = {1: "الفصل الأول", 2: "الفصل الثاني", 3: "الفصل الصيفي"}.get(term, f"الفصل {term}")
        scope_label_parts.append(term_name)
    scope_label = " - ".join(scope_label_parts) if scope_label_parts else "كامل الخطة"

    # جلب الـ ids قبل التعليم لإلغاء الإسنادات
    cc_ids = []
    async for c in db.curriculum_courses.find(base_filter, {"_id": 1}):
        cc_ids.append(str(c["_id"]))

    if not cc_ids:
        return {"message": f"لا توجد مقررات لمسحها ({scope_label})", "wiped": 0, "assignments_cleared": 0}

    now = _now()
    r1 = await db.curriculum_courses.update_many(
        base_filter,
        {"$set": {"is_active": False, "deleted_at": now, "deleted_by": current_user.get("id"),
                   "wiped_in_bulk": True}}
    )
    r2 = await db.teacher_assignments.update_many(
        {"curriculum_course_id": {"$in": cc_ids}, "is_active": {"$ne": False}},
        {"$set": {"is_active": False, "ended_at": now, "wiped_in_bulk": True}}
    )

    # سجل النشاط
    try:
        await db.activity_logs.insert_one({
            "user_id": current_user.get("id"),
            "username": current_user.get("username"),
            "action": "wipe_curriculum",
            "target_type": "department",
            "target_id": department_id,
            "details": f"مسح خطة قسم {dept.get('name')} ({scope_label}) - {r1.modified_count} مقرر",
            "timestamp": now,
        })
    except Exception:
        pass

    return {
        "message": f"تم مسح {scope_label} لقسم '{dept.get('name')}' - {r1.modified_count} مقرر",
        "wiped": r1.modified_count,
        "assignments_cleared": r2.modified_count,
        "department_name": dept.get("name"),
        "scope": scope_label,
    }


# ==================== Excel Template ====================

@router.get("/template/curriculum")
async def get_curriculum_template(current_user: dict = Depends(get_current_user)):
    """تحميل نموذج Excel لرفع الخطة الدراسية لقسم.
    أعمدة: رمز المقرر | اسم المقرر | الساعات المعتمدة | المستوى | الفصل
    'الفصل' يقبل: 1=الأول، 2=الثاني، 3=الصيفي
    """
    if not _has_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    df = pd.DataFrame({
        "رمز المقرر": ["FIB101", "FIB102", "ARB101", "NAH102", "AQD201", "SUM301"],
        "اسم المقرر": [
            "فقه عبادات (1)",
            "فقه عبادات (2)",
            "الكتابة العربية",
            "النحو (2)",
            "العقيدة الإسلامية",
            "مقرر صيفي مكثّف",
        ],
        "الساعات المعتمدة": [3, 3, 3, 2, 3, 2],
        "المستوى": [1, 1, 1, 1, 2, 2],
        # 1=الفصل الأول، 2=الفصل الثاني، 3=الفصل الصيفي
        "الفصل": [1, 2, 1, 2, "الأول", 3],
    })
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="الخطة الدراسية")
        # تعديل عرض الأعمدة + إضافة ورقة "تعليمات"
        ws = writer.sheets["الخطة الدراسية"]
        widths = {"A": 18, "B": 35, "C": 18, "D": 12, "E": 12}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        # ورقة تعليمات لتوضيح قيم الفصل
        instructions = pd.DataFrame({
            "العمود": ["رمز المقرر", "اسم المقرر", "الساعات المعتمدة", "المستوى", "الفصل"],
            "الوصف": [
                "رمز فريد للمقرر (مثل FIB101)",
                "اسم المقرر بالعربية",
                "عدد الساعات المعتمدة (رقم)",
                "المستوى الدراسي (1-6)",
                "1=الأول، 2=الثاني، 3=الصيفي (يقبل أيضاً النصوص: الأول/الثاني/الصيفي)",
            ],
        })
        instructions.to_excel(writer, index=False, sheet_name="تعليمات")
        ws2 = writer.sheets["تعليمات"]
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 65
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=curriculum_template.xlsx"},
    )


# ==================== Upload Helpers ====================

def _normalize_term(value) -> Optional[int]:
    """يحوّل القيمة إلى رقم فصل (1 أو 2 أو 3) — يقبل النص والرقم.
    1 = الفصل الأول، 2 = الفصل الثاني، 3 = الفصل الصيفي
    """
    if value is None:
        return None
    s = str(value).strip().lower()
    if not s or s in ("nan", "none"):
        return None
    # أرقام
    if s in ("1", "1.0", "اول", "الأول", "الاول", "first", "1st"):
        return 1
    if s in ("2", "2.0", "ثاني", "الثاني", "second", "2nd"):
        return 2
    if s in ("3", "3.0", "صيفي", "الصيفي", "صيف", "الصيف", "summer", "3rd"):
        return 3
    # محاولة كرقم
    try:
        n = int(float(s))
        if n in (1, 2, 3):
            return n
    except Exception:
        pass
    return None


def _parse_curriculum_excel(contents: bytes, filename: str) -> tuple[list, list]:
    """يقرأ ملف Excel/CSV ويُرجع (rows_صالحة, errors).
    كل صف صالح: {code, name, credit_hours, level, term, prerequisites}
    """
    fname = (filename or "").lower()
    try:
        if fname.endswith(".csv"):
            df = pd.read_csv(BytesIO(contents))
        else:
            df = pd.read_excel(BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"تعذر قراءة الملف: {e}")

    # mapping للأعمدة بمرونة
    col_map = {}
    for c in df.columns:
        cl = str(c).strip().lower().replace("ـ", "").replace(" ", "")
        if any(k in cl for k in ["رمز", "كود", "code"]):
            col_map["code"] = c
        elif any(k in cl for k in ["اسمالمقرر", "اسم", "name"]):
            col_map["name"] = c
        elif any(k in cl for k in ["ساعات", "ساعة", "credit", "hours"]):
            col_map["credit_hours"] = c
        elif any(k in cl for k in ["مستوى", "level"]):
            col_map["level"] = c
        elif any(k in cl for k in ["فصل", "term", "semester"]):
            col_map["term"] = c
        elif any(k in cl for k in ["متطلب", "prereq"]):
            col_map["prerequisites"] = c

    required = ["code", "name", "credit_hours", "level", "term"]
    missing = [r for r in required if r not in col_map]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"الأعمدة المطلوبة مفقودة: {', '.join(missing)} - تأكد من النموذج"
        )

    rows = []
    errors = []
    for idx, row in df.iterrows():
        line_num = idx + 2  # +2 لأن الصف الأول هو header و pandas يبدأ من 0
        code = str(row[col_map["code"]] or "").strip()
        name = str(row[col_map["name"]] or "").strip()
        if not code or not name or code == "nan" or name == "nan":
            continue  # صفوف فارغة - تخطي صامت
        try:
            credit_hours = int(float(row[col_map["credit_hours"]]))
        except Exception:
            errors.append(f"السطر {line_num}: ساعات معتمدة غير صحيحة")
            continue
        try:
            level = int(float(row[col_map["level"]]))
        except Exception:
            errors.append(f"السطر {line_num}: المستوى غير صحيح")
            continue
        term = _normalize_term(row[col_map["term"]])
        if term is None:
            errors.append(f"السطر {line_num}: الفصل غير معروف (المقبول: 1=أول، 2=ثاني، 3=صيفي)")
            continue

        prereqs = []
        if "prerequisites" in col_map:
            pv = str(row[col_map["prerequisites"]] or "").strip()
            if pv and pv != "nan":
                prereqs = [x.strip() for x in pv.replace("،", ",").split(",") if x.strip()]

        rows.append({
            "code": code.upper(),
            "name": name,
            "credit_hours": credit_hours,
            "level": level,
            "term": term,
            "prerequisites": prereqs,
            "_line": line_num,
        })
    return rows, errors


# ==================== Upload Preview ====================

@router.post("/curriculum/upload/preview")
async def preview_curriculum_upload(
    department_id: str = Query(..., description="معرف القسم"),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """معاينة ملف الخطة قبل التنفيذ - لا يحفظ شيئاً."""
    if not _has_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    db = get_db()
    try:
        dept = await db.departments.find_one({"_id": ObjectId(department_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرف القسم غير صحيح")
    if not dept:
        raise HTTPException(status_code=404, detail="القسم غير موجود")

    contents = await file.read()
    rows, errors = _parse_curriculum_excel(contents, file.filename or "")

    # فحص حد المستوى من الكلية
    faculty = None
    levels_count = 5
    if dept.get("faculty_id"):
        try:
            faculty = await db.faculties.find_one({"_id": ObjectId(dept["faculty_id"])})
            if faculty:
                levels_count = int(faculty.get("levels_count") or 5)
        except Exception:
            pass

    # فحص الموجود في الخطة
    existing_codes = set()
    async for c in db.curriculum_courses.find(
        {"department_id": department_id, "is_active": {"$ne": False}}, {"code": 1}
    ):
        existing_codes.add(c.get("code"))

    valid = []
    duplicates_in_file = []
    out_of_level = []
    existing_in_db = []
    seen_codes = set()
    for r in rows:
        if r["code"] in seen_codes:
            duplicates_in_file.append(r)
            continue
        seen_codes.add(r["code"])
        if r["level"] > levels_count or r["level"] < 1:
            out_of_level.append({**r, "_reason": f"المستوى {r['level']} خارج نطاق الكلية (1-{levels_count})"})
            continue
        if r["code"] in existing_codes:
            existing_in_db.append(r)
            continue
        valid.append(r)

    return {
        "department": {"id": department_id, "name": dept.get("name"), "levels_count": levels_count},
        "total_rows_read": len(rows),
        "valid_count": len(valid),
        "duplicates_in_file_count": len(duplicates_in_file),
        "existing_in_db_count": len(existing_in_db),
        "out_of_level_count": len(out_of_level),
        "parse_errors": errors,
        "valid_sample": valid[:20],
        "duplicates_in_file": duplicates_in_file[:10],
        "existing_in_db": existing_in_db[:10],
        "out_of_level": out_of_level[:10],
    }


# ==================== Upload Execute ====================

@router.post("/curriculum/upload")
async def upload_curriculum(
    department_id: str = Query(..., description="معرف القسم"),
    mode: str = Query("merge", description="merge=إضافة فقط، replace=استبدال كامل (يمسح القديم أولاً)"),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """رفع الخطة الدراسية لقسم من ملف Excel/CSV.
    - mode=merge: يضيف الجديد فقط، يتخطى الموجود
    - mode=replace: يمسح كل خطة القسم القديمة ثم يضيف الجديد
    """
    if not _has_manage(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح")
    if mode not in ("merge", "replace"):
        raise HTTPException(status_code=400, detail="mode يجب أن يكون merge أو replace")
    db = get_db()
    try:
        dept = await db.departments.find_one({"_id": ObjectId(department_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرف القسم غير صحيح")
    if not dept:
        raise HTTPException(status_code=404, detail="القسم غير موجود")

    contents = await file.read()
    rows, errors = _parse_curriculum_excel(contents, file.filename or "")
    if not rows:
        raise HTTPException(status_code=400, detail="لم يتم العثور على بيانات صالحة في الملف")

    # حد المستوى
    faculty = None
    levels_count = 5
    faculty_id = dept.get("faculty_id")
    if faculty_id:
        try:
            faculty = await db.faculties.find_one({"_id": ObjectId(faculty_id)})
            if faculty:
                levels_count = int(faculty.get("levels_count") or 5)
        except Exception:
            pass

    now = _now()

    # وضع الاستبدال - مسح القديم أولاً
    wiped = 0
    if mode == "replace":
        cc_ids = []
        async for c in db.curriculum_courses.find(
            {"department_id": department_id, "is_active": {"$ne": False}}, {"_id": 1}
        ):
            cc_ids.append(str(c["_id"]))
        if cc_ids:
            r1 = await db.curriculum_courses.update_many(
                {"department_id": department_id, "is_active": {"$ne": False}},
                {"$set": {"is_active": False, "deleted_at": now, "deleted_by": current_user.get("id"),
                          "wiped_in_bulk": True, "wipe_reason": "replace_upload"}}
            )
            wiped = r1.modified_count
            await db.teacher_assignments.update_many(
                {"curriculum_course_id": {"$in": cc_ids}, "is_active": {"$ne": False}},
                {"$set": {"is_active": False, "ended_at": now, "wiped_in_bulk": True}}
            )

    # الموجود في الخطة (بعد المسح إن لزم)
    existing_codes = set()
    async for c in db.curriculum_courses.find(
        {"department_id": department_id, "is_active": {"$ne": False}}, {"code": 1}
    ):
        existing_codes.add(c.get("code"))

    created = 0
    skipped = 0
    out_of_level = 0
    seen = set()
    created_items = []
    for r in rows:
        if r["code"] in seen:
            skipped += 1
            continue
        seen.add(r["code"])
        if r["level"] > levels_count or r["level"] < 1:
            out_of_level += 1
            continue
        if r["code"] in existing_codes:
            skipped += 1
            continue
        doc = {
            "code": r["code"],
            "name": r["name"],
            "credit_hours": r["credit_hours"],
            "department_id": department_id,
            "faculty_id": faculty_id,
            "level": r["level"],
            "term": r["term"],
            "prerequisites": r.get("prerequisites") or [],
            "is_active": True,
            "created_at": now,
            "created_by": current_user.get("id"),
            "imported_from": "excel_upload",
        }
        res = await db.curriculum_courses.insert_one(doc)
        existing_codes.add(r["code"])
        created += 1
        created_items.append({"id": str(res.inserted_id), "code": r["code"], "name": r["name"]})

    # سجل النشاط
    try:
        await db.activity_logs.insert_one({
            "user_id": current_user.get("id"),
            "username": current_user.get("username"),
            "action": "upload_curriculum",
            "target_type": "department",
            "target_id": department_id,
            "details": f"رفع {created} مقرر لقسم {dept.get('name')} (mode={mode})",
            "timestamp": now,
        })
    except Exception:
        pass

    return {
        "message": f"تم رفع {created} مقرر إلى خطة '{dept.get('name')}'",
        "created": created,
        "skipped_duplicates_or_existing": skipped,
        "out_of_level": out_of_level,
        "wiped_in_replace_mode": wiped,
        "parse_errors": errors,
        "department_name": dept.get("name"),
        "mode": mode,
        "sample": created_items[:20],
    }


# =====================================================================
# 📤 تصدير الخطة الدراسية (PDF / Excel) - عربي RTL + محاذاة يمين
# =====================================================================
def _term_label(t):
    return {1: "الفصل الأول", 2: "الفصل الثاني", 3: "الفصل الصيفي"}.get(int(t) if t else 0, "—")


async def _gather_curriculum_rows(db, department_id: str, level=None, term=None):
    """يجمع مقررات الخطة لقسم مع فلاتر اختيارية، مع إثراء أسماء المعلمين."""
    q: dict = {"department_id": department_id, "is_active": {"$ne": False}}
    if level is not None:
        q["level"] = int(level)
    if term is not None:
        q["term"] = int(term)
    docs = await db.curriculum_courses.find(q).sort([("level", 1), ("term", 1), ("name", 1)]).to_list(2000)
    # خريطة معلمين من teacher_assignments الفعَّالة
    cc_ids = [str(d["_id"]) for d in docs]
    teach_map: dict = {}
    if cc_ids:
        async for a in db.teacher_assignments.find(
            {"curriculum_course_id": {"$in": cc_ids}, "is_active": {"$ne": False}}
        ):
            teach_map.setdefault(a["curriculum_course_id"], []).append(a.get("teacher_name", ""))
    rows = []
    for d in docs:
        rows.append({
            "id": str(d["_id"]),
            "level": d.get("level"),
            "term": d.get("term"),
            "code": d.get("code", ""),
            "name": d.get("name", ""),
            "credit_hours": d.get("credit_hours", 0),
            "weekly_hours": d.get("weekly_hours") or 0,
            "teachers": "، ".join([t for t in teach_map.get(str(d["_id"]), []) if t]) or "—",
        })
    return rows


@router.get("/curriculum/department/{department_id}/export")
async def export_curriculum(
    department_id: str,
    format: str = Query("xlsx", description="xlsx أو pdf"),
    level: Optional[int] = None,
    term: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
):
    """تصدير الخطة الدراسية للقسم بصيغة Excel أو PDF مع دعم العربية + RTL."""
    if not _has_manage(current_user) and current_user.get("role") not in ("dean", "department_head"):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية للتصدير")

    db = get_db()
    try:
        dept = await db.departments.find_one({"_id": ObjectId(department_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرف القسم غير صحيح")
    if not dept:
        raise HTTPException(status_code=404, detail="القسم غير موجود")
    fac = None
    if dept.get("faculty_id"):
        try:
            fac = await db.faculties.find_one({"_id": ObjectId(dept["faculty_id"])})
        except Exception:
            pass

    rows = await _gather_curriculum_rows(db, department_id, level=level, term=term)
    dept_name = dept.get("name", "")
    fac_name = (fac or {}).get("name", "")

    scope_parts = []
    if level is not None:
        scope_parts.append(f"المستوى {level}")
    if term is not None:
        scope_parts.append(_term_label(term))
    scope_label = " - ".join(scope_parts) if scope_parts else "كامل الخطة"
    filename_safe = f"curriculum_{department_id}"
    if level is not None:
        filename_safe += f"_L{level}"
    if term is not None:
        filename_safe += f"_T{term}"

    if format.lower() == "pdf":
        return _export_pdf(rows, dept_name, fac_name, scope_label, filename_safe)
    return _export_xlsx(rows, dept_name, fac_name, scope_label, filename_safe)


def _export_xlsx(rows, dept_name, fac_name, scope_label, filename_safe):
    """Excel - RTL + محاذاة يمين + فواصل مستوى/فصل + إجماليات."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الخطة الدراسية"
    ws.sheet_view.rightToLeft = True  # RTL

    # 🆕 حساب الإجماليات
    total_credit = sum(int(r.get("credit_hours") or 0) for r in rows)
    total_weekly = sum(int(r.get("weekly_hours") or 0) for r in rows)

    # العناوين الكبيرة
    ws["A1"] = "الخطة الدراسية"
    ws["A1"].font = Font(name="Arial", size=18, bold=True, color="0d47a1")
    ws["A1"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells("A1:G1")
    ws["A2"] = f"الكلية: {fac_name} - القسم: {dept_name}"
    ws["A2"].font = Font(name="Arial", size=12, bold=True)
    ws["A2"].alignment = Alignment(horizontal="right")
    ws.merge_cells("A2:G2")
    ws["A3"] = (
        f"النطاق: {scope_label}    |    عدد المقررات: {len(rows)}    "
        f"|    إجمالي الساعات المعتمدة: {total_credit}    "
        f"|    إجمالي الساعات الأسبوعية: {total_weekly}"
    )
    ws["A3"].font = Font(name="Arial", size=11, italic=True, color="5a6c7d")
    ws["A3"].alignment = Alignment(horizontal="right")
    ws.merge_cells("A3:G3")

    # رأس الأعمدة
    headers = ["#", "المستوى", "الفصل", "الكود", "اسم المقرر", "الساعات المعتمدة", "الساعات الأسبوعية"]
    header_row = 5
    header_fill = PatternFill("solid", fgColor="1565c0")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="b0bec5")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # البيانات + إجماليات لكل فصل
    row_num = header_row + 1
    last_level = None
    last_term = None
    group_credit = 0   # 🆕 مجموع الفصل الحالي
    group_weekly = 0   # 🆕 مجموع الفصل الحالي
    group_count = 0    # 🆕 عدد مقررات الفصل الحالي

    def _flush_group_totals(target_row, lvl, trm, gc, gw, gcount):
        """يكتب صف إجمالي للفصل المنتهي."""
        if gcount == 0:
            return target_row
        label_cell = ws.cell(row=target_row, column=1,
                              value=f"إجمالي المستوى {lvl} - {_term_label(trm)} ({gcount} مقرر)")
        label_cell.font = Font(name="Arial", size=11, bold=True, color="0d47a1")
        label_cell.fill = PatternFill("solid", fgColor="e3f2fd")
        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        label_cell.border = border
        ws.merge_cells(start_row=target_row, start_column=1, end_row=target_row, end_column=5)
        # عمود ساعات معتمدة
        c1 = ws.cell(row=target_row, column=6, value=gc)
        c1.font = Font(name="Arial", size=11, bold=True, color="0d47a1")
        c1.fill = PatternFill("solid", fgColor="e3f2fd")
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = border
        # عمود ساعات أسبوعية
        c2 = ws.cell(row=target_row, column=7, value=gw)
        c2.font = Font(name="Arial", size=11, bold=True, color="0d47a1")
        c2.fill = PatternFill("solid", fgColor="e3f2fd")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = border
        return target_row + 1

    for idx, r in enumerate(rows, start=1):
        # فاصل مستوى/فصل عند التغيير
        if last_level != r["level"] or last_term != r["term"]:
            # أولاً: اكتب إجمالي الفصل السابق إن وجد
            if last_level is not None:
                row_num = _flush_group_totals(row_num, last_level, last_term, group_credit, group_weekly, group_count)
                group_credit = 0
                group_weekly = 0
                group_count = 0

            sep = ws.cell(row=row_num, column=1,
                         value=f"المستوى {r['level']} - {_term_label(r['term'])}")
            sep.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            sep.fill = PatternFill("solid", fgColor="ef6c00")
            sep.alignment = Alignment(horizontal="right", vertical="center")
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
            row_num += 1
            last_level, last_term = r["level"], r["term"]
        values = [idx, r["level"], _term_label(r["term"]), r["code"], r["name"], r["credit_hours"], r["weekly_hours"]]
        for ci, v in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=ci, value=v)
            cell.font = Font(name="Arial", size=11)
            cell.alignment = Alignment(horizontal="right" if ci == 5 else "center", vertical="center", wrap_text=True)
            cell.border = border
            if row_num % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="f5f7fa")
        # تراكم إجمالي الفصل الحالي
        group_credit += int(r.get("credit_hours") or 0)
        group_weekly += int(r.get("weekly_hours") or 0)
        group_count += 1
        row_num += 1

    # إجمالي آخر مجموعة
    if last_level is not None:
        row_num = _flush_group_totals(row_num, last_level, last_term, group_credit, group_weekly, group_count)

    # 🆕 الإجمالي الكلي
    if rows:
        row_num += 1  # سطر فارغ
        grand_label = ws.cell(row=row_num, column=1,
                               value=f"الإجمالي الكلي للخطة ({len(rows)} مقرر)")
        grand_label.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        grand_label.fill = PatternFill("solid", fgColor="2e7d32")
        grand_label.alignment = Alignment(horizontal="right", vertical="center")
        grand_label.border = border
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
        gc = ws.cell(row=row_num, column=6, value=total_credit)
        gc.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        gc.fill = PatternFill("solid", fgColor="2e7d32")
        gc.alignment = Alignment(horizontal="center", vertical="center")
        gc.border = border
        gw = ws.cell(row=row_num, column=7, value=total_weekly)
        gw.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        gw.fill = PatternFill("solid", fgColor="2e7d32")
        gw.alignment = Alignment(horizontal="center", vertical="center")
        gw.border = border
        ws.row_dimensions[row_num].height = 30

    # عرض الأعمدة
    widths = [5, 10, 14, 16, 40, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[header_row].height = 26

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename_safe}.xlsx"'},
    )


def _export_pdf(rows, dept_name, fac_name, scope_label, filename_safe):
    """PDF - عربي + RTL + كل فصل تحت الآخر + KeepTogether لمنع تقسيم الجدول."""
    import os
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether, PageBreak
    import arabic_reshaper
    from bidi.algorithm import get_display

    font_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "fonts", "Amiri-Regular.ttf")
    if "Amiri" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("Amiri", font_path))

    def ar(s):
        try:
            return get_display(arabic_reshaper.reshape(str(s)))
        except Exception:
            return str(s)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    elements: list = []
    title_style = ParagraphStyle("Title", fontName="Amiri", fontSize=20, leading=30, alignment=2, textColor=colors.HexColor("#0d47a1"), spaceAfter=6)
    sub_style = ParagraphStyle("Sub", fontName="Amiri", fontSize=13, leading=22, alignment=2, textColor=colors.HexColor("#1565c0"), spaceAfter=4)
    info_style = ParagraphStyle("Info", fontName="Amiri", fontSize=11, leading=18, alignment=2, textColor=colors.HexColor("#5a6c7d"), spaceAfter=10)
    totals_style = ParagraphStyle("Totals", fontName="Amiri", fontSize=12, leading=22, alignment=2,
                                    textColor=colors.white, backColor=colors.HexColor("#2e7d32"),
                                    borderPadding=10, spaceBefore=4, spaceAfter=14)

    # 🆕 حساب الإجمالي الكلي
    grand_credit = sum(int(r.get("credit_hours") or 0) for r in rows)
    grand_weekly = sum(int(r.get("weekly_hours") or 0) for r in rows)

    elements.append(Paragraph(ar("الخطة الدراسية"), title_style))
    elements.append(Paragraph(ar(f"الكلية: {fac_name} — القسم: {dept_name}"), sub_style))
    elements.append(Paragraph(ar(f"النطاق: {scope_label}    |    عدد المقررات: {len(rows)}"), info_style))
    # 🆕 شريط الإجماليات الكلية
    elements.append(Paragraph(
        ar(f"الإجمالي الكلي للخطة:  الساعات المعتمدة = {grand_credit}    •    الساعات الأسبوعية = {grand_weekly}"),
        totals_style
    ))
    elements.append(Spacer(1, 6))

    # تجميع حسب المستوى ثم الفصل
    by_level: dict = {}
    for r in rows:
        lv = r.get("level") or 0
        by_level.setdefault(lv, []).append(r)

    level_style = ParagraphStyle("Lvl", fontName="Amiri", fontSize=15, leading=24, alignment=2, textColor=colors.white,
                                  backColor=colors.HexColor("#0d47a1"), borderPadding=10, spaceBefore=2, spaceAfter=8)
    section_style = ParagraphStyle("Sec", fontName="Amiri", fontSize=14, leading=22, alignment=2,
                                    textColor=colors.white, backColor=colors.HexColor("#ef6c00"),
                                    leftIndent=4, rightIndent=4, spaceBefore=8, spaceAfter=6, borderPadding=8)

    sorted_levels = sorted(by_level.keys())
    for li, lv in enumerate(sorted_levels):
        # PageBreak بين المستويات (بعد الأول)
        if li > 0:
            elements.append(PageBreak())
        # العنوان الرئيسي للمستوى
        level_rows = by_level[lv]
        total_in_level = len(level_rows)
        elements.append(Paragraph(ar(f"المستوى {lv}  —  {total_in_level} مقرر"), level_style))
        # تجميع الفصول داخل المستوى
        term_groups: dict = {}
        for r in level_rows:
            tm = r.get("term") or 0
            term_groups.setdefault(tm, []).append(r)
        for tm in sorted(term_groups.keys()):
            grp = term_groups[tm]
            # 🆕 حساب إجمالي هذا (المستوى, الفصل)
            term_credit = sum(int(r.get("credit_hours") or 0) for r in grp)
            term_weekly = sum(int(r.get("weekly_hours") or 0) for r in grp)
            # كل (مستوى, فصل) في كتلة واحدة لا تنقسم عبر الصفحات
            block: list = []
            block.append(Paragraph(ar(f"{_term_label(tm)}  ({len(grp)} مقرر)"), section_style))
            header = [ar(h) for h in ["الساعات الأسبوعية", "الساعات المعتمدة", "اسم المقرر", "الكود", "#"]]
            data = [header]
            for idx, r in enumerate(grp, start=1):
                data.append([
                    ar(str(r["weekly_hours"] or "—")),
                    ar(str(r["credit_hours"] or "—")),
                    ar(r["name"]),
                    ar(r["code"] or "—"),
                    ar(str(idx)),
                ])
            # 🆕 صف الإجمالي
            data.append([
                ar(str(term_weekly)),
                ar(str(term_credit)),
                ar(f"إجمالي {_term_label(tm)}"),
                ar(""),
                ar(""),
            ])
            table = Table(data, colWidths=[22*mm, 22*mm, 80*mm, 30*mm, 10*mm], repeatRows=1)
            table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), "Amiri"),
                ("FONTSIZE", (0, 0), (-1, -1), 11),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565c0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#b0bec5")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#1565c0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f5f7fa")]),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
                # 🆕 تنسيق صف الإجمالي (آخر صف)
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e3f2fd")),
                ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#0d47a1")),
                ("FONTSIZE", (0, -1), (-1, -1), 12),
                ("BOTTOMPADDING", (0, -1), (-1, -1), 6),
                ("TOPPADDING", (0, -1), (-1, -1), 6),
            ]))
            block.append(table)
            block.append(Spacer(1, 4))
            # KeepTogether: عنوان الفصل + جدوله لا ينقسمان عبر صفحتين
            elements.append(KeepTogether(block))

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename_safe}.pdf"'},
    )
