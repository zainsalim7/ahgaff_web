"""
Alumni (الخريجون) Routes
- POST /api/students/{id}/graduate → تخريج طالب
- GET /api/alumni → قائمة الخريجين مع فلاتر (سنة، كلية، قسم، بحث)
- GET /api/alumni/{id} → تفاصيل خريج
- PUT /api/alumni/{id}/restore → استرجاع خريج إلى قائمة الطلاب

ملاحظة معمارية: الخريجون يبقون في نفس collection `students` مع `is_alumni=true`
- يحافظ على الـ student_id، user_id، وكل السجلات السابقة (enrollments, attendance...)
- يُستثنون تلقائياً من GET /students الافتراضي
"""
from typing import Optional
from datetime import datetime, timezone

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel, Field

from .deps import get_current_user, get_db, has_any_permission, get_scope_filter
from models.permissions import Permission

router = APIRouter(tags=["الخريجون"])


# ============================
# Models
# ============================
class GraduateRequest(BaseModel):
    graduation_year: int = Field(..., description="سنة التخرج (الحقل الإلزامي للربط)")
    graduation_date: Optional[str] = Field(None, description="تاريخ التخرج YYYY-MM-DD")
    graduation_semester: Optional[str] = Field(None, description="الفصل: first/second/summer")
    final_gpa: Optional[float] = Field(None, ge=0, le=4.0, description="المعدل التراكمي النهائي")
    total_credit_hours: Optional[int] = Field(None, ge=0)
    certificate_number: Optional[str] = None
    honors: Optional[str] = Field(None, description="مرتبة الشرف / تقدير")
    notes: Optional[str] = None


class BulkGraduateRequest(BaseModel):
    student_ids: list[str] = Field(..., description="قائمة معرفات الطلاب للتخريج")
    graduation_year: int = Field(..., description="سنة التخرج (مشتركة لجميع المختارين)")
    graduation_date: Optional[str] = None
    graduation_semester: Optional[str] = None
    honors: Optional[str] = None
    notes: Optional[str] = None


class AlumniUpdateRequest(BaseModel):
    """تحديث بيانات تخرّج خريج موجود."""
    graduation_year: Optional[int] = Field(None, ge=1900, le=2200)
    graduation_date: Optional[str] = None
    graduation_semester: Optional[str] = None
    final_gpa: Optional[float] = Field(None, ge=0, le=4.0)
    total_credit_hours: Optional[int] = Field(None, ge=0)
    certificate_number: Optional[str] = None
    honors: Optional[str] = None
    notes: Optional[str] = None


def _now():
    return datetime.now(timezone.utc).isoformat()


def _ensure_can_manage_students(current_user: dict):
    if current_user.get("role") == "admin":
        return
    if has_any_permission(current_user, [Permission.MANAGE_STUDENTS]):
        return
    raise HTTPException(status_code=403, detail="ليس لديك صلاحية لإدارة الطلاب")


async def _ensure_student_in_user_scope(db, current_user: dict, student: dict):
    """يتأكد أن الطالب ضمن نطاق كلية/قسم المستخدم."""
    if current_user.get("role") == "admin":
        return
    user_dept_id = current_user.get("department_id")
    user_faculty_id = current_user.get("faculty_id")
    s_dept = str(student.get("department_id") or "")
    s_fac = str(student.get("faculty_id") or "")
    if not s_fac and s_dept:
        try:
            d = await db.departments.find_one({"_id": ObjectId(s_dept)})
            if d and d.get("faculty_id"):
                s_fac = str(d.get("faculty_id"))
        except Exception:
            pass
    if user_dept_id and s_dept != str(user_dept_id):
        raise HTTPException(status_code=403, detail="هذا الطالب ليس من قسمك")
    if user_faculty_id and not user_dept_id and s_fac != str(user_faculty_id):
        raise HTTPException(status_code=403, detail="هذا الطالب ليس من كليتك")


async def _resolve_dept_faculty_snapshot(db, student: dict) -> dict:
    """يبني snapshot باسم القسم والكلية للطالب في لحظة التخرّج.
    
    يحمي البيانات التاريخية للخريج من:
    - تغيير department_id للطالب لاحقاً بسبب خطأ إدخال.
    - إعادة تسمية أو دمج القسم/الكلية بعد التخرّج.
    - أي فقدان مرجع لاحق (department محذوف).
    """
    snapshot = {
        "department_id": str(student.get("department_id") or "") or None,
        "department_name": None,
        "faculty_id": str(student.get("faculty_id") or "") or None,
        "faculty_name": None,
    }
    if snapshot["department_id"]:
        try:
            d = await db.departments.find_one({"_id": ObjectId(snapshot["department_id"])})
            if d:
                snapshot["department_name"] = (d.get("name") or "").strip() or None
                if not snapshot["faculty_id"] and d.get("faculty_id"):
                    snapshot["faculty_id"] = str(d["faculty_id"])
        except Exception:
            pass
    if snapshot["faculty_id"]:
        try:
            f = await db.faculties.find_one({"_id": ObjectId(snapshot["faculty_id"])})
            if f:
                snapshot["faculty_name"] = (f.get("name") or "").strip() or None
        except Exception:
            pass
    return snapshot


# ============================
# 1) Graduate Student
# ============================
@router.post("/students/{student_id}/graduate")
async def graduate_student(
    student_id: str,
    body: GraduateRequest,
    current_user: dict = Depends(get_current_user),
):
    """تخريج طالب: ينقله إلى قائمة الخريجين مع بيانات التخرج المرتبطة بالسنة."""
    _ensure_can_manage_students(current_user)
    db = get_db()

    try:
        student = await db.students.find_one({"_id": ObjectId(student_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرف الطالب غير صحيح")
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    if student.get("is_alumni") is True:
        raise HTTPException(status_code=400, detail="الطالب متخرج بالفعل")

    await _ensure_student_in_user_scope(db, current_user, student)

    # 📸 حفظ snapshot للقسم/الكلية عند التخرج
    snapshot = await _resolve_dept_faculty_snapshot(db, student)

    graduation_data = {
        "year": body.graduation_year,
        "date": body.graduation_date,
        "semester": body.graduation_semester,
        "final_gpa": body.final_gpa,
        "total_credit_hours": body.total_credit_hours,
        "certificate_number": body.certificate_number,
        "honors": body.honors,
        "notes": body.notes,
        "graduated_at": _now(),
        "graduated_by_user_id": current_user.get("id"),
        "graduated_by_username": current_user.get("username"),
        "department_snapshot": snapshot,  # 🔒 قسم/كلية لحظة التخرج
    }

    await db.students.update_one(
        {"_id": ObjectId(student_id)},
        {
            "$set": {
                "is_alumni": True,
                "status": "graduated",
                "graduation_date": body.graduation_date,
                "graduated_from_level": student.get("level"),
                "graduation_data": graduation_data,
                "status_changed_at": _now(),
            }
        },
    )

    # سجل النشاط
    try:
        await db.activity_logs.insert_one({
            "user_id": current_user.get("id"),
            "username": current_user.get("username"),
            "action": "graduate_student",
            "target_type": "student",
            "target_id": student_id,
            "details": f"تخريج الطالب {student.get('full_name', '')} - سنة {body.graduation_year}",
            "timestamp": _now(),
        })
    except Exception:
        pass

    return {
        "message": f"تم تخريج الطالب '{student.get('full_name')}' بنجاح",
        "student_id": student_id,
        "graduation_year": body.graduation_year,
    }


# ============================
# 1.5) Bulk Graduate
# ============================
@router.post("/students/bulk-graduate")
async def bulk_graduate_students(
    body: BulkGraduateRequest,
    current_user: dict = Depends(get_current_user),
):
    """تخريج عدة طلاب دفعة واحدة بنفس السنة المشتركة."""
    _ensure_can_manage_students(current_user)
    db = get_db()

    if not body.student_ids:
        raise HTTPException(status_code=400, detail="قائمة الطلاب فارغة")

    graduated = 0
    failed: list[dict] = []
    skipped: list[dict] = []
    now = _now()

    for sid in body.student_ids:
        try:
            student = await db.students.find_one({"_id": ObjectId(sid)})
        except Exception:
            failed.append({"id": sid, "error": "معرف غير صحيح"})
            continue
        if not student:
            failed.append({"id": sid, "error": "الطالب غير موجود"})
            continue
        if student.get("is_alumni") is True:
            skipped.append({"id": sid, "name": student.get("full_name"), "reason": "متخرج بالفعل"})
            continue
        try:
            await _ensure_student_in_user_scope(db, current_user, student)
        except HTTPException as e:
            failed.append({"id": sid, "name": student.get("full_name"), "error": e.detail})
            continue

        # 📸 حفظ snapshot للقسم/الكلية عند التخرج الجماعي
        snapshot = await _resolve_dept_faculty_snapshot(db, student)
        graduation_data = {
            "year": body.graduation_year,
            "date": body.graduation_date,
            "semester": body.graduation_semester,
            "final_gpa": None,
            "total_credit_hours": None,
            "certificate_number": None,
            "honors": body.honors,
            "notes": body.notes,
            "graduated_at": now,
            "graduated_by_user_id": current_user.get("id"),
            "graduated_by_username": current_user.get("username"),
            "bulk_graduation": True,
            "department_snapshot": snapshot,
        }
        await db.students.update_one(
            {"_id": ObjectId(sid)},
            {"$set": {
                "is_alumni": True,
                "status": "graduated",
                "graduation_date": body.graduation_date,
                "graduated_from_level": student.get("level"),
                "graduation_data": graduation_data,
                "status_changed_at": now,
            }},
        )
        graduated += 1

    try:
        await db.activity_logs.insert_one({
            "user_id": current_user.get("id"),
            "username": current_user.get("username"),
            "action": "bulk_graduate",
            "details": f"تخريج جماعي - سنة {body.graduation_year} - نجح {graduated} من {len(body.student_ids)}",
            "timestamp": now,
        })
    except Exception:
        pass

    return {
        "message": f"تم تخريج {graduated} طالب بنجاح",
        "graduated": graduated,
        "failed": failed,
        "skipped": skipped,
        "graduation_year": body.graduation_year,
    }


# ============================
# 2) List Alumni
# ============================
@router.get("/alumni")
async def list_alumni(
    year: Optional[int] = Query(None, description="فلتر بسنة التخرج"),
    faculty_id: Optional[str] = None,
    department_id: Optional[str] = None,
    q: Optional[str] = Query(None, description="بحث بالاسم أو رقم القيد"),
    current_user: dict = Depends(get_current_user),
):
    """قائمة الخريجين مع فلاتر."""
    db = get_db()
    if not (current_user.get("role") == "admin" or has_any_permission(
        current_user, [Permission.VIEW_STUDENTS, Permission.MANAGE_STUDENTS]
    )):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية لعرض الخريجين")

    query: dict = {"is_alumni": True}
    # تطبيق نطاق المستخدم
    scope = await get_scope_filter(current_user, "students")
    for k, v in scope.items():
        query[k] = v

    if year is not None:
        query["graduation_data.year"] = year
    if faculty_id:
        # يطابق كلاً من الحقل الحالي و snapshot المحفوظ (للسجلات التي تعرّض قسمها للتعديل)
        query["$or"] = query.get("$or", []) + [
            {"faculty_id": faculty_id},
            {"graduation_data.department_snapshot.faculty_id": faculty_id},
        ]
    if department_id:
        or_clauses = [
            {"department_id": department_id},
            {"graduation_data.department_snapshot.department_id": department_id},
        ]
        if "$or" in query:
            existing_or = query.pop("$or")
            query["$and"] = [{"$or": existing_or}, {"$or": or_clauses}]
        else:
            query["$or"] = or_clauses
    if q:
        import re as _re
        rx = {"$regex": _re.escape(q), "$options": "i"}
        q_clauses = [{"full_name": rx}, {"student_id": rx}, {"reference_number": rx}]
        if "$and" in query:
            query["$and"].append({"$or": q_clauses})
        elif "$or" in query:
            existing_or = query.pop("$or")
            query["$and"] = [{"$or": existing_or}, {"$or": q_clauses}]
        else:
            query["$or"] = q_clauses

    cursor = db.students.find(query).sort([("graduation_data.year", -1), ("full_name", 1)])
    alumni = await cursor.to_list(5000)

    # إثراء بأسماء الأقسام والكليات
    dept_ids = list({str(s.get("department_id", "")) for s in alumni if s.get("department_id")})
    dept_map: dict = {}
    fac_ids_set: set = set()
    if dept_ids:
        try:
            obj_ids = [ObjectId(x) for x in dept_ids if x]
            async for d in db.departments.find({"_id": {"$in": obj_ids}}, {"name": 1, "faculty_id": 1}):
                dept_map[str(d["_id"])] = d
                if d.get("faculty_id"):
                    fac_ids_set.add(str(d["faculty_id"]))
        except Exception:
            pass
    fac_map: dict = {}
    if fac_ids_set:
        try:
            obj_ids = [ObjectId(x) for x in fac_ids_set if x]
            async for f in db.faculties.find({"_id": {"$in": obj_ids}}, {"name": 1}):
                fac_map[str(f["_id"])] = f.get("name", "")
        except Exception:
            pass

    result = []
    for s in alumni:
        gd = s.get("graduation_data") or {}
        snap = gd.get("department_snapshot") or {}
        dept_id = str(s.get("department_id", "") or "")
        dept_doc = dept_map.get(dept_id) if dept_id else None
        current_fac_name = ""
        if dept_doc and dept_doc.get("faculty_id"):
            current_fac_name = fac_map.get(str(dept_doc["faculty_id"]), "")
        # 🔒 نُعطي الأولوية لـ snapshot المحفوظ لحظة التخرج (المصدر التاريخي الصحيح)
        # ثم نرجع للقسم/الكلية الحالية كاحتياط لسجلات قديمة قبل إضافة الـ snapshot
        display_dept_id = snap.get("department_id") or dept_id
        display_dept_name = snap.get("department_name") or ((dept_doc or {}).get("name", "") if dept_doc else "")
        display_fac_id = snap.get("faculty_id") or ((dept_doc or {}).get("faculty_id", "") if dept_doc else s.get("faculty_id", ""))
        display_fac_name = snap.get("faculty_name") or current_fac_name
        result.append({
            "id": str(s["_id"]),
            "student_id": s.get("student_id"),
            "reference_number": s.get("reference_number"),
            "full_name": s.get("full_name"),
            "phone": s.get("phone"),
            "email": s.get("email"),
            "department_id": display_dept_id,
            "department_name": display_dept_name,
            "faculty_id": display_fac_id,
            "faculty_name": display_fac_name,
            "level_graduated": s.get("graduated_from_level") or s.get("level"),
            "graduation_year": gd.get("year"),
            "graduation_date": gd.get("date") or s.get("graduation_date"),
            "graduation_semester": gd.get("semester"),
            "final_gpa": gd.get("final_gpa"),
            "total_credit_hours": gd.get("total_credit_hours"),
            "certificate_number": gd.get("certificate_number"),
            "honors": gd.get("honors"),
            "notes": gd.get("notes"),
            "qr_code": s.get("qr_code"),
            "created_at": s.get("created_at"),
        })

    # إحصاءات مختصرة لتسهيل العرض
    years_count: dict = {}
    for r in result:
        y = r.get("graduation_year")
        if y is not None:
            years_count[str(y)] = years_count.get(str(y), 0) + 1

    return {
        "items": result,
        "total": len(result),
        "by_year": years_count,
    }


# ============================
# 3) Get Alumni Detail
# ============================
@router.get("/alumni/{alumni_id}")
async def get_alumni(
    alumni_id: str,
    current_user: dict = Depends(get_current_user),
):
    db = get_db()
    if not (current_user.get("role") == "admin" or has_any_permission(
        current_user, [Permission.VIEW_STUDENTS, Permission.MANAGE_STUDENTS]
    )):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")

    try:
        s = await db.students.find_one({"_id": ObjectId(alumni_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرف غير صحيح")
    if not s or not s.get("is_alumni"):
        raise HTTPException(status_code=404, detail="الخريج غير موجود")

    await _ensure_student_in_user_scope(db, current_user, s)

    gd = s.get("graduation_data") or {}
    return {
        "id": str(s["_id"]),
        "student_id": s.get("student_id"),
        "full_name": s.get("full_name"),
        "phone": s.get("phone"),
        "email": s.get("email"),
        "department_id": s.get("department_id"),
        "faculty_id": s.get("faculty_id"),
        "level_graduated": s.get("graduated_from_level") or s.get("level"),
        "graduation_data": gd,
        "qr_code": s.get("qr_code"),
    }


# ============================
# 4) Restore Alumni → Student
# ============================
@router.put("/alumni/{alumni_id}/restore")
async def restore_alumni_to_student(
    alumni_id: str,
    current_user: dict = Depends(get_current_user),
):
    """استرجاع خريج إلى قائمة الطلاب (في حال خطأ في التخرج)."""
    _ensure_can_manage_students(current_user)
    db = get_db()

    try:
        s = await db.students.find_one({"_id": ObjectId(alumni_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرف غير صحيح")
    if not s:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    if not s.get("is_alumni"):
        raise HTTPException(status_code=400, detail="هذا الطالب ليس خريجاً")

    await _ensure_student_in_user_scope(db, current_user, s)

    # 🔒 استرجاع القسم/الكلية من snapshot إن وُجد (يعيد الطالب لقسمه الأصلي)
    gd = s.get("graduation_data") or {}
    snap = gd.get("department_snapshot") or {}
    set_fields: dict = {
        "is_alumni": False,
        "status": "active",
        "status_changed_at": _now(),
        "status_reason": "استرجاع من قائمة الخريجين",
    }
    if snap.get("department_id") and snap["department_id"] != str(s.get("department_id") or ""):
        set_fields["department_id"] = snap["department_id"]
    if snap.get("faculty_id") and snap["faculty_id"] != str(s.get("faculty_id") or ""):
        set_fields["faculty_id"] = snap["faculty_id"]

    await db.students.update_one(
        {"_id": ObjectId(alumni_id)},
        {
            "$set": set_fields,
            "$unset": {"graduation_data": "", "graduation_date": "", "graduated_from_level": ""},
        },
    )

    try:
        await db.activity_logs.insert_one({
            "user_id": current_user.get("id"),
            "username": current_user.get("username"),
            "action": "restore_alumni",
            "target_type": "student",
            "target_id": alumni_id,
            "details": f"استرجاع الخريج {s.get('full_name', '')} إلى قائمة الطلاب",
            "timestamp": _now(),
        })
    except Exception:
        pass

    return {"message": f"تم استرجاع '{s.get('full_name')}' إلى قائمة الطلاب", "student_id": alumni_id}


# ============================
# 5) Update Alumni Graduation Data
# ============================
@router.put("/alumni/{alumni_id}")
async def update_alumni(
    alumni_id: str,
    body: AlumniUpdateRequest,
    current_user: dict = Depends(get_current_user),
):
    """تحديث بيانات تخرج خريج (السنة، المعدل، الشهادة، التقدير، الفصل، الملاحظات).
    
    الحقول المرسلة فقط هي التي تُحدَّث (partial update). لا يلمس بيانات الطالب الأساسية.
    """
    _ensure_can_manage_students(current_user)
    db = get_db()

    try:
        s = await db.students.find_one({"_id": ObjectId(alumni_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="معرف غير صحيح")
    if not s:
        raise HTTPException(status_code=404, detail="الخريج غير موجود")
    if not s.get("is_alumni"):
        raise HTTPException(status_code=400, detail="هذا الطالب ليس خريجاً")

    await _ensure_student_in_user_scope(db, current_user, s)

    gd = dict(s.get("graduation_data") or {})
    payload = body.dict(exclude_unset=True, exclude_none=False)
    # خريطة من حقول الـ API إلى مفاتيح المخزّن
    field_map = {
        "graduation_year": "year",
        "graduation_date": "date",
        "graduation_semester": "semester",
        "final_gpa": "final_gpa",
        "total_credit_hours": "total_credit_hours",
        "certificate_number": "certificate_number",
        "honors": "honors",
        "notes": "notes",
    }
    for api_key, store_key in field_map.items():
        if api_key in payload:
            gd[store_key] = payload[api_key]
    gd["updated_at"] = _now()
    gd["updated_by_user_id"] = current_user.get("id")
    gd["updated_by_username"] = current_user.get("username")

    update_set: dict = {"graduation_data": gd}
    if "graduation_date" in payload:
        update_set["graduation_date"] = payload["graduation_date"]

    await db.students.update_one(
        {"_id": ObjectId(alumni_id)},
        {"$set": update_set},
    )

    try:
        await db.activity_logs.insert_one({
            "user_id": current_user.get("id"),
            "username": current_user.get("username"),
            "action": "update_alumni",
            "target_type": "student",
            "target_id": alumni_id,
            "details": f"تحديث بيانات تخرج الخريج {s.get('full_name', '')}",
            "timestamp": _now(),
        })
    except Exception:
        pass

    return {"message": "تم تحديث بيانات الخريج بنجاح", "id": alumni_id}


# ============================
# 6) Export Alumni (Excel/PDF)
# ============================
async def _build_alumni_export_rows(
    db, current_user: dict,
    year: Optional[int], faculty_id: Optional[str],
    department_id: Optional[str], q: Optional[str],
):
    """ينفّذ نفس استعلام list_alumni لكن يعيد قائمة مسطّحة جاهزة للتصدير."""
    query: dict = {"is_alumni": True}
    scope = await get_scope_filter(current_user, "students")
    for k, v in scope.items():
        query[k] = v
    if year is not None:
        query["graduation_data.year"] = year
    if faculty_id:
        query["faculty_id"] = faculty_id
    if department_id:
        query["department_id"] = department_id
    if q:
        import re as _re
        rx = {"$regex": _re.escape(q), "$options": "i"}
        query["$or"] = [{"full_name": rx}, {"student_id": rx}, {"reference_number": rx}]

    alumni = await db.students.find(query).sort([("graduation_data.year", -1), ("full_name", 1)]).to_list(10000)
    dept_ids = list({str(s.get("department_id", "")) for s in alumni if s.get("department_id")})
    dept_map: dict = {}
    fac_ids_set: set = set()
    if dept_ids:
        try:
            async for d in db.departments.find(
                {"_id": {"$in": [ObjectId(x) for x in dept_ids if x]}}, {"name": 1, "faculty_id": 1},
            ):
                dept_map[str(d["_id"])] = d
                if d.get("faculty_id"):
                    fac_ids_set.add(str(d["faculty_id"]))
        except Exception:
            pass
    fac_map: dict = {}
    if fac_ids_set:
        try:
            async for f in db.faculties.find(
                {"_id": {"$in": [ObjectId(x) for x in fac_ids_set if x]}}, {"name": 1},
            ):
                fac_map[str(f["_id"])] = (f.get("name", "") or "").strip()
        except Exception:
            pass

    rows = []
    for s in alumni:
        gd = s.get("graduation_data") or {}
        dept_id = str(s.get("department_id", "") or "")
        dept_doc = dept_map.get(dept_id) if dept_id else None
        fac_name = ""
        if dept_doc and dept_doc.get("faculty_id"):
            fac_name = fac_map.get(str(dept_doc["faculty_id"]), "")
        rows.append({
            "student_id": s.get("student_id") or "",
            "full_name": (s.get("full_name") or "").strip(),
            "department": ((dept_doc or {}).get("name") or "").strip() if dept_doc else "",
            "faculty": fac_name,
            "graduation_year": gd.get("year") or "",
            "graduation_semester": gd.get("semester") or "",
            "graduation_date": gd.get("date") or s.get("graduation_date") or "",
            "level_graduated": s.get("graduated_from_level") or s.get("level") or "",
            "final_gpa": gd.get("final_gpa") if gd.get("final_gpa") is not None else "",
            "total_credit_hours": gd.get("total_credit_hours") if gd.get("total_credit_hours") is not None else "",
            "certificate_number": gd.get("certificate_number") or "",
            "honors": gd.get("honors") or "",
            "phone": s.get("phone") or "",
            "email": s.get("email") or "",
        })
    return rows


@router.get("/alumni/export/excel")
async def export_alumni_excel(
    year: Optional[int] = None,
    faculty_id: Optional[str] = None,
    department_id: Optional[str] = None,
    q: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """تصدير قائمة الخريجين إلى Excel."""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    if not (current_user.get("role") == "admin" or has_any_permission(
        current_user, [Permission.VIEW_STUDENTS, Permission.MANAGE_STUDENTS]
    )):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")

    db = get_db()
    rows = await _build_alumni_export_rows(db, current_user, year, faculty_id, department_id, q)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الخريجون"
    ws.sheet_view.rightToLeft = True

    BLUE = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    GREY = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    thin = Side(style='thin', color='B0BEC5')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    BOLD = Font(bold=True, size=11)
    WHITE_BOLD = Font(bold=True, color='FFFFFF', size=11)

    NUM_COLS = 13
    # Title
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=NUM_COLS)
    title_cell = ws.cell(row=1, column=1, value="قائمة الخريجين")
    title_cell.font = Font(bold=True, size=15, color='1565C0')
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # Subtitle (filters summary + total)
    parts = [f"الإجمالي: {len(rows)} خريج"]
    if year:
        parts.append(f"السنة: {year}")
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=NUM_COLS)
    sub = ws.cell(row=2, column=1, value=" · ".join(parts))
    sub.font = Font(bold=True, size=11, color='5B6678')
    sub.alignment = center
    ws.row_dimensions[2].height = 20

    # Header
    headers = [
        "#", "الاسم", "رقم القيد", "القسم", "الكلية", "سنة التخرج",
        "الفصل", "تاريخ التخرج", "المستوى", "المعدل", "الساعات المعتمدة",
        "رقم الشهادة", "التقدير",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = BLUE
        c.font = WHITE_BOLD
        c.alignment = center
        c.border = border
    ws.row_dimensions[4].height = 24

    sem_label = {"first": "الأول", "second": "الثاني", "summer": "الصيفي"}
    cur = 5
    for idx, r in enumerate(rows, start=1):
        vals = [
            idx, r["full_name"], r["student_id"], r["department"], r["faculty"],
            r["graduation_year"],
            sem_label.get(str(r.get("graduation_semester") or ""), r.get("graduation_semester") or ""),
            r["graduation_date"], r["level_graduated"], r["final_gpa"],
            r["total_credit_hours"], r["certificate_number"], r["honors"],
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=cur, column=i, value=v)
            c.alignment = center
            c.border = border
            if idx % 2 == 0:
                c.fill = GREY
        cur += 1

    if not rows:
        ws.merge_cells(start_row=cur, end_row=cur, start_column=1, end_column=NUM_COLS)
        c = ws.cell(row=cur, column=1, value="لا يوجد خريجون مطابقون للفلاتر")
        c.alignment = center
        c.font = Font(italic=True, color='8A95A8')

    widths = [4, 26, 13, 18, 20, 11, 10, 13, 8, 9, 13, 14, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=alumni.xlsx"},
    )


@router.get("/alumni/export/pdf")
async def export_alumni_pdf(
    year: Optional[int] = None,
    faculty_id: Optional[str] = None,
    department_id: Optional[str] = None,
    q: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """تصدير قائمة الخريجين إلى PDF (RTL عربي)."""
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import arabic_reshaper
    from bidi.algorithm import get_display
    from io import BytesIO
    from pathlib import Path

    if not (current_user.get("role") == "admin" or has_any_permission(
        current_user, [Permission.VIEW_STUDENTS, Permission.MANAGE_STUDENTS]
    )):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")

    db = get_db()
    rows = await _build_alumni_export_rows(db, current_user, year, faculty_id, department_id, q)

    font_path = Path(__file__).parent.parent / "fonts" / "Amiri-Regular.ttf"
    if font_path.exists():
        try:
            pdfmetrics.registerFont(TTFont('Amiri', str(font_path)))
        except Exception:
            pass
        arabic_font = 'Amiri'
    else:
        arabic_font = 'Helvetica'

    def ar(t):
        return get_display(arabic_reshaper.reshape(str(t or "")))

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        rightMargin=12 * mm, leftMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'AlumniTitle', parent=styles['Title'], fontName=arabic_font,
        fontSize=18, alignment=TA_CENTER, textColor=colors.HexColor('#1565c0'),
    )
    sub_style = ParagraphStyle(
        'AlumniSub', parent=styles['Normal'], fontName=arabic_font,
        fontSize=11, alignment=TA_CENTER, textColor=colors.HexColor('#5b6678'),
    )

    elements = [Paragraph(ar("قائمة الخريجين"), title_style), Spacer(1, 3 * mm)]
    parts = [f"الإجمالي: {len(rows)} خريج"]
    if year:
        parts.append(f"السنة: {year}")
    elements.append(Paragraph(ar(" · ".join(parts)), sub_style))
    elements.append(Spacer(1, 6 * mm))

    sem_label = {"first": "الأول", "second": "الثاني", "summer": "الصيفي"}

    if not rows:
        elements.append(Paragraph(ar("لا يوجد خريجون مطابقون للفلاتر"), sub_style))
    else:
        # ترتيب الأعمدة من اليسار → اليمين على PDF (RTL محتوى):
        # [التقدير | المعدل | المستوى | تاريخ التخرج | الفصل | سنة التخرج | الكلية | القسم | رقم القيد | الاسم | #]
        header = [
            ar("التقدير"), ar("المعدل"), ar("المستوى"),
            ar("تاريخ التخرج"), ar("الفصل"), ar("سنة التخرج"),
            ar("الكلية"), ar("القسم"), ar("رقم القيد"), ar("الاسم"), ar("#"),
        ]
        data = [header]
        for idx, r in enumerate(rows, start=1):
            data.append([
                ar(r["honors"] or "-"),
                ar(str(r["final_gpa"]) if r["final_gpa"] != "" else "-"),
                ar(str(r["level_graduated"]) if r["level_graduated"] != "" else "-"),
                ar(r["graduation_date"] or "-"),
                ar(sem_label.get(str(r.get("graduation_semester") or ""), r.get("graduation_semester") or "-")),
                ar(str(r["graduation_year"]) if r["graduation_year"] != "" else "-"),
                ar(r["faculty"] or "-"),
                ar(r["department"] or "-"),
                ar(r["student_id"] or "-"),
                ar(r["full_name"] or "-"),
                str(idx),
            ])
        total_w = doc.width
        col_widths = [
            total_w * 0.09, total_w * 0.06, total_w * 0.05,
            total_w * 0.08, total_w * 0.06, total_w * 0.06,
            total_w * 0.12, total_w * 0.12, total_w * 0.08,
            total_w * 0.24, total_w * 0.04,
        ]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, 0), 10.5),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565c0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cfd6e1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fafbfd')]),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(t)

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=alumni.pdf"},
    )
