"""
استيراد الجدول الأسبوعي الشامل من Excel + توليد قالب الاستيراد
"""
import io
import re
from datetime import datetime, timezone

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from bson import ObjectId
from pymongo.errors import DuplicateKeyError

from .deps import get_db, get_current_user, log_activity
from .weekly_schedule import can_manage_schedule, _is_period_unavailable, _build_master_data

router = APIRouter(tags=["استيراد الجدول الأسبوعي"])

AR_ORDINALS = {
    "الاول": 1, "الأول": 1, "الثاني": 2, "الثالث": 3, "الرابع": 4, "الخامس": 5,
    "السادس": 6, "السابع": 7, "الثامن": 8, "التاسع": 9, "العاشر": 10,
}
KNOWN_DAYS = {"السبت", "الأحد", "الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة"}
SECTION_CHARS = {"أ", "ا", "ب", "ج", "د", "هـ", "ه", "و", "ز", "ح"}


def _norm(s) -> str:
    """تطبيع عربي للمطابقة النصية: همزات/تاء مربوطة/ياء + مسافات"""
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r"[\u064B-\u0652\u0640]", "", s)  # تشكيل + تطويل
    s = s.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    s = s.replace("ة", "ه").replace("ى", "ي")
    s = re.sub(r"\s+", " ", s)
    return s


def _norm_day(s) -> str:
    n = _norm(s)
    for d in KNOWN_DAYS:
        if _norm(d) == n:
            return d
    return ""


def _parse_group_label(label: str):
    """يستخرج (المستوى، الشعبة) من تسمية المجموعة مثل: 'المستوى 2 - شعبة أ' أو 'الثاني شريعة أ'"""
    txt = _norm(label)
    if not txt:
        return None, None
    level = None
    m = re.search(r"\d+", txt)
    if m:
        level = int(m.group())
    else:
        for word, val in AR_ORDINALS.items():
            if _norm(word) in txt.split():
                level = val
                break
    section = ""
    m = re.search(r"شعبه\s+(\S+)", txt)
    if m:
        section = m.group(1)
    else:
        last = txt.split()[-1] if txt.split() else ""
        if last in {_norm(c) for c in SECTION_CHARS} and len(txt.split()) > 1:
            section = last
    return level, section


ROW_KINDS = [("course", "مقرر"), ("room", "قاع"), ("teacher", "استاذ"), ("teacher", "معلم")]


def _row_kind(b_label: str):
    n = _norm(b_label)
    for kind, key in ROW_KINDS:
        if key in n:
            return kind
    return None


@router.get("/weekly-schedule/import-template")
async def download_import_template(
    faculty_id: str,
    department_id: str,
    current_user: dict = Depends(get_current_user),
):
    """قالب Excel للاستيراد بنفس بنية النموذج: مجموعات × (أيام × فترات)، كل خلية 3 صفوف (مقرر/قاعة/أستاذ)"""
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = get_db()
    data = await _build_master_data(db, faculty_id, department_id)
    working_days = data["working_days"]
    time_slots = sorted(data["time_slots"], key=lambda x: x.get("slot_number", 0))
    groups = [g for g in data["groups"] if g["department_id"] == department_id]
    if not working_days or not time_slots:
        raise HTTPException(status_code=400, detail="لا توجد أيام عمل أو فترات معرفة لهذه الكلية — عرّفها من الإعدادات أولاً")
    if not groups:
        raise HTTPException(status_code=400, detail="لا توجد مجموعات (مستويات/شعب) لهذا القسم — أضف مقررات بمعلمين أولاً")

    dept = await db.departments.find_one({"_id": ObjectId(department_id)})
    faculty = await db.faculties.find_one({"_id": ObjectId(faculty_id)})

    # خريطة الخلايا الموجودة (لتعبئة القالب بالجدول الحالي)
    cell_map = {}
    for e in data["entries"]:
        if e["department_id"] != department_id:
            continue
        cell_map[(e["level"], e["section"], e["day"], e["slot_number"])] = e

    wb = Workbook()
    ws = wb.active
    ws.title = "الجدول"
    ws.sheet_view.rightToLeft = True

    ns = len(time_slots)
    ncols = 2 + len(working_days) * ns

    thin = Side(style="thin", color="B8C4D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws.cell(row=1, column=1, value=f"قالب استيراد الجدول الأسبوعي — {(faculty or {}).get('name', '')} — قسم {(dept or {}).get('name', '')}")
    tc.font = Font(bold=True, size=13, color="0D2A52")
    tc.alignment = center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ic = ws.cell(row=2, column=1, value="⚠️ استخدم الأسماء كما في ورقة (الأدلة) حرفياً. لا تغيّر تسميات المستويات أو الصفوف. الخلايا الفارغة تُتجاهل.")
    ic.font = Font(size=9, color="B26A00", bold=True)
    ic.alignment = Alignment(horizontal="right", vertical="center")

    # صف الأيام (3) + صف الفترات (4)
    for h, col in (("المستوى / الشعبة", 1), ("", 2)):
        c = ws.cell(row=3, column=col, value=h)
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
        c.fill = PatternFill("solid", fgColor="0D2A52")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = center
    for di, day in enumerate(working_days):
        c1 = 3 + di * ns
        ws.merge_cells(start_row=3, start_column=c1, end_row=3, end_column=c1 + ns - 1)
        dc = ws.cell(row=3, column=c1, value=day)
        dc.fill = PatternFill("solid", fgColor="1565C0")
        dc.font = Font(bold=True, color="FFFFFF", size=11)
        dc.alignment = center
        for si, ts in enumerate(time_slots):
            sc = ws.cell(row=4, column=c1 + si, value=ts.get("name") or f"الفترة {ts.get('slot_number')}")
            sc.fill = PatternFill("solid", fgColor="3D7EDE")
            sc.font = Font(bold=True, color="FFFFFF", size=8)
            sc.alignment = center
            sc.border = border

    row_labels = ["المقرر", "القاعة", "الأستاذ"]
    r = 5
    for g in groups:
        label = f"المستوى {g['level']}" + (f" - شعبة {g['section']}" if g["section"] else "")
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=1)
        gc = ws.cell(row=r, column=1, value=label)
        gc.fill = PatternFill("solid", fgColor="EEF3FA")
        gc.font = Font(bold=True, size=9, color="1A2540")
        gc.alignment = center
        gc.border = border
        for i, rl in enumerate(row_labels):
            bc = ws.cell(row=r + i, column=2, value=rl)
            bc.fill = PatternFill("solid", fgColor="F5F7FA")
            bc.font = Font(size=8, color="5A6B85", bold=True)
            bc.alignment = center
            bc.border = border
            for di, day in enumerate(working_days):
                for si, ts in enumerate(time_slots):
                    col = 3 + di * ns + si
                    cell = ws.cell(row=r + i, column=col)
                    cell.border = border
                    cell.alignment = center
                    cell.font = Font(size=8)
                    e = cell_map.get((g["level"], g["section"], day, ts.get("slot_number")))
                    if e:
                        cell.value = [e["course_name"], e["room_name"], e["teacher_name"]][i]
                        cell.fill = PatternFill("solid", fgColor="FFF8E1")
        r += 3

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 9
    for c in range(3, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # ورقة الأدلة: المقررات والقاعات بالأسماء الدقيقة
    ws2 = wb.create_sheet("الأدلة")
    ws2.sheet_view.rightToLeft = True
    courses = await db.courses.find({"department_id": department_id, "is_active": True}).to_list(2000)
    t_ids = [c.get("teacher_id") for c in courses if c.get("teacher_id")]
    teachers = {str(t["_id"]): t.get("full_name", "") for t in await db.teachers.find({"_id": {"$in": [ObjectId(x) for x in t_ids]}}).to_list(1000)} if t_ids else {}
    rooms = await db.rooms.find({"faculty_id": faculty_id, "is_active": True}).to_list(300)
    heads = ["اسم المقرر (انسخه حرفياً)", "المستوى", "الشعبة", "الأستاذ المسند", "", "القاعات المتاحة"]
    for ci, h in enumerate(heads, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1565C0")
        c.alignment = center
    for ri, cdoc in enumerate(sorted(courses, key=lambda x: (x.get("level") or 0, x.get("section") or "", x.get("name") or "")), 2):
        ws2.cell(row=ri, column=1, value=cdoc.get("name", ""))
        ws2.cell(row=ri, column=2, value=cdoc.get("level") or 1)
        ws2.cell(row=ri, column=3, value=cdoc.get("section") or "-")
        ws2.cell(row=ri, column=4, value=teachers.get(cdoc.get("teacher_id", ""), "⚠️ بلا أستاذ"))
    for ri, rdoc in enumerate(sorted(rooms, key=lambda x: x.get("name", "")), 2):
        ws2.cell(row=ri, column=6, value=rdoc.get("name", ""))
    for ci, w in enumerate([35, 10, 10, 28, 4, 22], 1):
        from openpyxl.utils import get_column_letter as gcl
        ws2.column_dimensions[gcl(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"schedule_import_template_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/weekly-schedule/import-master")
async def import_master_schedule(
    faculty_id: str = Form(...),
    department_id: str = Form(...),
    dry_run: str = Form("1"),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """استيراد الجدول الأسبوعي الشامل من Excel لقسم محدد.
    السياسة: دمج (الخلايا المشغولة تُتخطى) + أخطاء الأسماء تُتخطى مع التقرير + أي تعارض جدولة يوقف الاستيراد كاملاً.
    """
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    is_dry = str(dry_run).strip() not in ("0", "false", "False")

    from openpyxl import load_workbook
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="تعذر قراءة الملف — تأكد أنه ملف Excel (.xlsx) سليم")
    ws = wb.worksheets[0]

    db = get_db()
    dept = await db.departments.find_one({"_id": ObjectId(department_id)})
    if not dept or dept.get("faculty_id") != faculty_id:
        raise HTTPException(status_code=400, detail="القسم غير موجود أو لا يتبع الكلية المحددة")

    settings = await db.schedule_settings.find_one({"_id": f"faculty_{faculty_id}"}) or await db.schedule_settings.find_one({"_id": "global"})
    time_slots = sorted((settings or {}).get("time_slots", []), key=lambda x: x.get("slot_number", 0))
    working_days = (settings or {}).get("working_days", [])
    if not time_slots or not working_days:
        raise HTTPException(status_code=400, detail="لا توجد فترات أو أيام عمل معرفة لهذه الكلية")

    # ===== 1) تحديد صف الأيام وصف الفترات =====
    days_row = None
    for r in range(1, min(11, ws.max_row + 1)):
        found = sum(1 for c in range(1, ws.max_column + 1) if _norm_day(ws.cell(row=r, column=c).value))
        if found >= 2:
            days_row = r
            break
    if not days_row:
        raise HTTPException(status_code=400, detail="لم يتم العثور على صف الأيام (السبت، الأحد...) في الملف — استخدم القالب الرسمي")
    periods_row = days_row + 1

    # خريطة عمود → (يوم، ترتيب الفترة داخل اليوم)
    col_day = {}
    current_day, order = "", 0
    for c in range(1, ws.max_column + 1):
        d = _norm_day(ws.cell(row=days_row, column=c).value)
        if d:
            current_day, order = d, 0
        if current_day and ws.cell(row=periods_row, column=c).value not in (None, ""):
            order += 1
            col_day[c] = (current_day, order)

    errors, conflicts, skipped_existing = [], [], []
    slot_by_order = {i + 1: ts for i, ts in enumerate(time_slots)}
    working_set = set(working_days)

    # ===== 2) بيانات النظام للمطابقة =====
    courses = await db.courses.find({"department_id": department_id, "is_active": True}).to_list(2000)
    courses_by_name = {}
    for cdoc in courses:
        courses_by_name.setdefault(_norm(cdoc.get("name")), []).append(cdoc)
    t_ids = list({c.get("teacher_id") for c in courses if c.get("teacher_id")})
    teachers_map = {str(t["_id"]): t for t in await db.teachers.find({"_id": {"$in": [ObjectId(x) for x in t_ids]}}).to_list(1000)} if t_ids else {}
    rooms = await db.rooms.find({"faculty_id": faculty_id, "is_active": True}).to_list(300)
    rooms_by_name = {_norm(r.get("name")): r for r in rooms}

    existing_slots = await db.weekly_schedule.find({"department_id": department_id}).to_list(5000)
    existing_cells = {(s.get("level"), s.get("section", "") or "", s.get("day"), s.get("slot_number")): s for s in existing_slots}

    # ===== 3) المرور على كتل المجموعات (3 صفوف لكل مجموعة) =====
    to_create = []
    r = periods_row + 1
    while r <= ws.max_row:
        label = ws.cell(row=r, column=1).value
        if label is None or not str(label).strip():
            r += 1
            continue
        level, section = _parse_group_label(str(label))
        block_rows = {"course": r, "room": r + 1, "teacher": r + 2}
        for i in range(3):
            kind = _row_kind(str(ws.cell(row=r + i, column=2).value or ""))
            if kind:
                block_rows[kind] = r + i
        if level is None:
            errors.append(f"صف {r}: تعذر تحديد المستوى من التسمية '{str(label).strip()}' — تُخُطيت المجموعة")
            r += 3
            continue

        for c, (day, order) in col_day.items():
            course_txt = str(ws.cell(row=block_rows["course"], column=c).value or "").strip()
            room_txt = str(ws.cell(row=block_rows["room"], column=c).value or "").strip()
            teacher_txt = str(ws.cell(row=block_rows["teacher"], column=c).value or "").strip()
            if not course_txt and not room_txt and not teacher_txt:
                continue
            loc = f"[المستوى {level}{' شعبة ' + section if section else ''} — {day} فترة {order}]"
            if day not in working_set:
                errors.append(f"{loc} اليوم '{day}' ليس من أيام عمل الكلية — تُخُطيت الخلية")
                continue
            ts = slot_by_order.get(order)
            if not ts:
                errors.append(f"{loc} الفترة رقم {order} غير معرفة في إعدادات الكلية ({len(time_slots)} فترات فقط) — تُخُطيت الخلية")
                continue
            slot_number = ts.get("slot_number")
            if not course_txt:
                errors.append(f"{loc} خلية بلا اسم مقرر — تُخُطيت")
                continue

            # مطابقة المقرر (مقارنة مستوى/شعبة مطبّعة + تمييز الأسماء المتطابقة بالأستاذ)
            candidates = courses_by_name.get(_norm(course_txt), [])
            nsec = _norm(section)
            matches = [x for x in candidates if (x.get("level") or 1) == level and _norm(x.get("section") or "") == nsec]
            if len(matches) > 1 and teacher_txt:
                tmatches = [x for x in matches if _norm(teachers_map.get(x.get("teacher_id", ""), {}).get("full_name", "")) == _norm(teacher_txt)]
                if tmatches:
                    matches = tmatches
            course = matches[0] if matches else None
            if not course:
                if candidates:
                    have = "، ".join(f"م{x.get('level') or 1}{'/' + x.get('section') if x.get('section') else ''}" for x in candidates[:3])
                    errors.append(f"{loc} المقرر '{course_txt}' موجود لكن لمستوى/شعبة مختلفة ({have}) — تُخُطيت الخلية")
                else:
                    errors.append(f"{loc} المقرر '{course_txt}' غير موجود في القسم — تُخُطيت الخلية")
                continue
            section_val = course.get("section") or ""
            if not course.get("teacher_id"):
                errors.append(f"{loc} المقرر '{course_txt}' بلا أستاذ مسند في النظام — تُخُطيت الخلية")
                continue
            teacher = teachers_map.get(course["teacher_id"], {})
            if teacher_txt and _norm(teacher_txt) != _norm(teacher.get("full_name", "")):
                errors.append(f"{loc} اسم الأستاذ في الملف '{teacher_txt}' يختلف عن الأستاذ المسند للمقرر '{teacher.get('full_name', '')}' — تُخُطيت الخلية")
                continue
            if not room_txt:
                errors.append(f"{loc} القاعة مطلوبة — تُخُطيت الخلية")
                continue
            room = rooms_by_name.get(_norm(room_txt))
            if not room:
                errors.append(f"{loc} القاعة '{room_txt}' غير مسجلة في الكلية — تُخُطيت الخلية")
                continue

            # الخلية مشغولة مسبقاً → دمج: تخطٍ مع تنبيه
            ex = existing_cells.get((level, section_val, day, slot_number))
            if ex:
                same = ex.get("course_id") == str(course["_id"])
                skipped_existing.append(f"{loc} الخلية مشغولة مسبقاً بـ'{course_txt if same else 'مقرر آخر'}' — {'مطابقة للملف' if same else 'تم الإبقاء على الموجود'}")
                continue

            to_create.append({
                "faculty_id": faculty_id,
                "department_id": department_id,
                "level": level,
                "section": section_val,
                "day": day,
                "slot_number": slot_number,
                "course_id": str(course["_id"]),
                "teacher_id": course["teacher_id"],
                "room_id": str(room["_id"]),
                "_loc": loc,
                "_course_name": course.get("name", ""),
                "_teacher_name": teacher.get("full_name", ""),
                "_room_name": room.get("name", ""),
            })
        r += 3

    # ===== 4) فحص التعارضات (داخل الملف + مع الجدول القائم عبر كل الأقسام) =====
    all_slots = await db.weekly_schedule.find({}, {"teacher_id": 1, "room_id": 1, "day": 1, "slot_number": 1, "department_id": 1, "level": 1, "section": 1}).to_list(20000)
    busy_teacher = {(s.get("teacher_id"), s.get("day"), s.get("slot_number")) for s in all_slots if s.get("teacher_id")}
    busy_room = {(s.get("room_id"), s.get("day"), s.get("slot_number")) for s in all_slots if s.get("room_id")}
    teacher_daily = {}
    for s in all_slots:
        if s.get("teacher_id"):
            k = (s["teacher_id"], s.get("day"))
            teacher_daily[k] = teacher_daily.get(k, 0) + 1

    prefs_map = {p["teacher_id"]: p for p in await db.teacher_preferences.find({"teacher_id": {"$in": list({x["teacher_id"] for x in to_create})}}).to_list(500)} if to_create else {}

    seen_teacher, seen_room, seen_cell = set(), set(), set()
    for item in to_create:
        loc = item["_loc"]
        tk = (item["teacher_id"], item["day"], item["slot_number"])
        rk = (item["room_id"], item["day"], item["slot_number"])
        ck = (item["level"], item["section"], item["day"], item["slot_number"])
        if ck in seen_cell:
            conflicts.append(f"{loc} خلية مكررة داخل الملف لنفس الشعبة")
        if tk in busy_teacher:
            conflicts.append(f"{loc} تعارض معلم: '{item['_teacher_name']}' لديه محاضرة أخرى في الجدول القائم بنفس (اليوم/الفترة)")
        elif tk in seen_teacher:
            conflicts.append(f"{loc} تعارض معلم داخل الملف: '{item['_teacher_name']}' مذكور في خليتين بنفس (اليوم/الفترة)")
        if rk in busy_room:
            conflicts.append(f"{loc} تعارض قاعة: '{item['_room_name']}' محجوزة في الجدول القائم بنفس (اليوم/الفترة)")
        elif rk in seen_room:
            conflicts.append(f"{loc} تعارض قاعة داخل الملف: '{item['_room_name']}' مذكورة في خليتين بنفس (اليوم/الفترة)")
        pref = prefs_map.get(item["teacher_id"])
        if pref and _is_period_unavailable(pref, item["day"], item["slot_number"]):
            conflicts.append(f"{loc} تعارض تفضيلات: '{item['_teacher_name']}' غير متاح يوم {item['day']} الفترة {item['slot_number']}")
        dk = (item["teacher_id"], item["day"])
        teacher_daily[dk] = teacher_daily.get(dk, 0) + 1
        if pref and teacher_daily[dk] > int(pref.get("max_daily_lectures") or 3):
            conflicts.append(f"{loc} تعارض تفضيلات: '{item['_teacher_name']}' سيتجاوز الحد اليومي ({pref.get('max_daily_lectures', 3)}) يوم {item['day']}")
        seen_teacher.add(tk)
        seen_room.add(rk)
        seen_cell.add(ck)

    can_commit = len(conflicts) == 0 and len(to_create) > 0
    report = {
        "dry_run": is_dry,
        "to_create": len(to_create),
        "created": 0,
        "skipped_existing": skipped_existing,
        "errors": errors,
        "conflicts": conflicts,
        "can_commit": can_commit,
    }

    if conflicts:
        report["message"] = f"🛑 تم إيقاف الاستيراد: يوجد {len(conflicts)} تعارض جدولة — عالج التعارضات ثم أعد المحاولة (لم يُحفظ أي شيء)"
        return report
    if is_dry:
        report["message"] = f"معاينة: سيتم إدراج {len(to_create)} محاضرة" + (f" • تخطي {len(skipped_existing)} مشغولة" if skipped_existing else "") + (f" • {len(errors)} خطأ أسماء" if errors else "")
        return report

    created = 0
    for item in to_create:
        doc = {k: v for k, v in item.items() if not k.startswith("_")}
        doc["created_at"] = datetime.now(timezone.utc)
        doc["created_by"] = current_user["id"]
        doc["imported_from_excel"] = True
        try:
            await db.weekly_schedule.insert_one(doc)
            created += 1
        except DuplicateKeyError:
            conflicts.append(f"{item['_loc']} رُفض من قاعدة البيانات (تعارض فريد لحظي)")

    await log_activity(
        current_user, "import_master_schedule_excel", "weekly_schedule", department_id, None,
        {"faculty_id": faculty_id, "created": created, "errors": len(errors), "skipped_existing": len(skipped_existing)},
    )
    report["created"] = created
    report["message"] = f"✅ تم إدراج {created} محاضرة في الجدول الأسبوعي" + (f" • تخطي {len(skipped_existing)} مشغولة مسبقاً" if skipped_existing else "") + (f" • {len(errors)} خلية بأخطاء أسماء (انظر التقرير)" if errors else "")
    return report
