"""
Weekly Schedule Routes - مسارات الجدول الأسبوعي
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from bson import ObjectId
from datetime import datetime, timezone
from typing import Optional, List
from pydantic import BaseModel
from pymongo.errors import DuplicateKeyError

from .deps import get_db, get_current_user, has_permission, log_activity
from models.permissions import Permission

router = APIRouter(tags=["الجدول الأسبوعي"])

SCHEDULE_ROLES = {"admin", "dean", "department_head", "registrar"}

def can_manage_schedule(user: dict) -> bool:
    if user.get("role") in SCHEDULE_ROLES:
        return True
    if has_permission(user, Permission.MANAGE_SETTINGS):
        return True
    if has_permission(user, Permission.MANAGE_COURSES):
        return True
    return False


# ===== Models =====

class RoomCreate(BaseModel):
    name: str
    faculty_id: str
    capacity: int = 30
    building: str = ""
    floor: str = ""
    room_type: str = "lecture"  # lecture, lab, seminar
    is_active: bool = True


class RoomUpdate(BaseModel):
    name: Optional[str] = None
    faculty_id: Optional[str] = None
    capacity: Optional[int] = None
    building: Optional[str] = None
    floor: Optional[str] = None
    room_type: Optional[str] = None
    is_active: Optional[bool] = None


class TimeSlotCreate(BaseModel):
    slot_number: int
    name: str  # المحاضرة الأولى
    start_time: str  # "08:00"
    end_time: str  # "09:30"


class UnavailablePeriod(BaseModel):
    day: str
    slot_number: int


class TeacherPreferenceUpdate(BaseModel):
    unavailable_days: List[str] = []  # ["الثلاثاء", "الأربعاء"] — يوم كامل
    unavailable_slots: List[int] = []  # [1, 5] = فترات تُطبَّق على كل الأيام (قديم/توافق رجعي)
    unavailable_periods: List[UnavailablePeriod] = []  # 🆕 شبكة يوم×فترة بدقة عالية
    max_daily_lectures: int = 2
    allow_consecutive_lectures: bool = False  # افتراضياً لا نسمح بمحاضرتين متتاليتين لنفس الأستاذ


def _derive_unavailable_periods(pref: dict, working_days: List[str], slot_numbers: List[int]) -> List[dict]:
    """🔄 يستخرج قائمة الخلايا (يوم/فترة) غير المتاحة من كل من:
    - unavailable_periods الجديدة (dosahalt)
    - unavailable_days القديمة (يوم كامل → كل الفترات)
    - unavailable_slots القديمة (فترة عبر كل الأيام)
    ويعيد مجموعة موحدة بدون تكرار.
    """
    result = set()
    # 1) الشبكة الجديدة (إن وُجدت)
    for p in pref.get("unavailable_periods", []) or []:
        if isinstance(p, dict) and p.get("day") and p.get("slot_number") is not None:
            result.add((str(p["day"]), int(p["slot_number"])))
    # 2) توافق رجعي: أيام كاملة قديمة
    for d in pref.get("unavailable_days", []) or []:
        for sn in slot_numbers:
            result.add((str(d), int(sn)))
    # 3) توافق رجعي: فترات قديمة على كل الأيام
    for sn in pref.get("unavailable_slots", []) or []:
        for d in working_days:
            result.add((str(d), int(sn)))
    return [{"day": d, "slot_number": sn} for (d, sn) in result]


def _is_period_unavailable(pref: dict, day: str, slot_number: int) -> bool:
    """يتحقق ما إذا كانت الخلية (يوم/فترة) محظورة، بدمج الشبكة الجديدة + الحقول القديمة."""
    # الشبكة الجديدة
    for p in pref.get("unavailable_periods", []) or []:
        if isinstance(p, dict) and p.get("day") == day and int(p.get("slot_number", -1)) == int(slot_number):
            return True
    # يوم كامل قديم
    if day in (pref.get("unavailable_days", []) or []):
        return True
    # فترة عبر كل الأيام قديمة
    if int(slot_number) in (pref.get("unavailable_slots", []) or []):
        return True
    return False


class ScheduleSlotCreate(BaseModel):
    faculty_id: str
    department_id: str
    level: int
    section: str = ""
    day: str  # السبت, الأحد, ...
    slot_number: int  # 1, 2, 3, 4, 5
    course_id: str
    teacher_id: str
    room_id: str


class ScheduleSlotUpdate(BaseModel):
    course_id: Optional[str] = None
    teacher_id: Optional[str] = None
    room_id: Optional[str] = None


# ===== القاعات =====

@router.get("/rooms")
async def get_rooms(
    faculty_id: Optional[str] = None,
    building: Optional[str] = None,
    is_active: Optional[bool] = True,
    current_user: dict = Depends(get_current_user),
):
    db = get_db()
    query = {}
    if is_active is not None:
        query["is_active"] = is_active
    if faculty_id:
        query["faculty_id"] = faculty_id
    if building:
        query["building"] = building
    rooms = await db.rooms.find(query).sort("name", 1).to_list(500)
    return [{
        "id": str(r["_id"]),
        "name": r.get("name", ""),
        "faculty_id": r.get("faculty_id", ""),
        "capacity": r.get("capacity", 30),
        "building": r.get("building", ""),
        "floor": r.get("floor", ""),
        "room_type": r.get("room_type", "lecture"),
        "is_active": r.get("is_active", True),
    } for r in rooms]


class RoomOccurrence(BaseModel):
    date: str  # YYYY-MM-DD
    start_time: str  # HH:MM
    end_time: str  # HH:MM


class RoomAvailabilityRequest(BaseModel):
    occurrences: List[RoomOccurrence]
    exclude_lecture_id: Optional[str] = None


@router.post("/rooms/availability")
async def check_rooms_availability(data: RoomAvailabilityRequest, current_user: dict = Depends(get_current_user)):
    """🟢🔴 فحص إشغال جميع القاعات المسجلة مقابل مواعيد محددة (محاضرة واحدة أو سلسلة توليد)"""
    db = get_db()
    occurrences = data.occurrences[:200]
    if not occurrences:
        raise HTTPException(status_code=400, detail="لا توجد مواعيد للفحص")

    rooms = await db.rooms.find({"is_active": True}).sort("name", 1).to_list(500)
    dates = list({o.date for o in occurrences})

    lec_query = {
        "date": {"$in": dates},
        "status": {"$ne": "cancelled"},
        "room": {"$nin": ["", None]},
    }
    if data.exclude_lecture_id:
        try:
            lec_query["_id"] = {"$ne": ObjectId(data.exclude_lecture_id)}
        except Exception:
            pass
    lectures = await db.lectures.find(
        lec_query, {"room": 1, "date": 1, "start_time": 1, "end_time": 1, "course_id": 1}
    ).to_list(10000)

    # أسماء المقررات الحاجزة (جلب جماعي)
    course_ids = list({l.get("course_id") for l in lectures if l.get("course_id")})
    course_names = {}
    if course_ids:
        try:
            oids = [ObjectId(cid) for cid in course_ids]
            async for c in db.courses.find({"_id": {"$in": oids}}, {"name": 1}):
                course_names[str(c["_id"])] = c.get("name", "")
        except Exception:
            pass

    # تجميع المحاضرات حسب اسم القاعة (بعد إزالة الفراغات)
    by_room = {}
    for l in lectures:
        rn = (l.get("room") or "").strip()
        if rn:
            by_room.setdefault(rn, []).append(l)

    results = []
    total = len(occurrences)
    for r in rooms:
        rname = (r.get("name") or "").strip()
        conflicts = []
        busy_count = 0
        room_lectures = by_room.get(rname, [])
        for occ in occurrences:
            hit = None
            for l in room_lectures:
                if l.get("date") != occ.date:
                    continue
                ls, le = l.get("start_time", ""), l.get("end_time", "")
                if ls and le and occ.start_time < le and occ.end_time > ls:
                    hit = l
                    break
            if hit:
                busy_count += 1
                if len(conflicts) < 3:
                    conflicts.append({
                        "date": hit.get("date", ""),
                        "start_time": hit.get("start_time", ""),
                        "end_time": hit.get("end_time", ""),
                        "course_name": course_names.get(hit.get("course_id", ""), "مقرر آخر"),
                    })
        results.append({
            "id": str(r["_id"]),
            "name": rname,
            "building": r.get("building", ""),
            "total": total,
            "busy_count": busy_count,
            "conflicts": conflicts,
        })
    return {"results": results, "total_occurrences": total}


@router.post("/rooms")
async def create_room(data: RoomCreate, current_user: dict = Depends(get_current_user)):
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    existing = await db.rooms.find_one({"name": data.name, "faculty_id": data.faculty_id})
    if existing:
        raise HTTPException(status_code=409, detail=f"القاعة '{data.name}' موجودة مسبقاً")
    doc = data.dict()
    doc["created_at"] = datetime.now(timezone.utc)
    result = await db.rooms.insert_one(doc)
    return {"id": str(result.inserted_id), "message": "تم إضافة القاعة بنجاح"}


@router.get("/rooms/template/excel")
async def download_rooms_template(current_user: dict = Depends(get_current_user)):
    """تحميل قالب Excel لاستيراد القاعات"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "القاعات"
    ws.sheet_view.rightToLeft = True

    headers = ['اسم القاعة', 'السعة', 'المبنى', 'الطابق']
    hfill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # أمثلة
    examples = [
        ['ق101', 50, 'المبنى الرئيسي', 'الأول'],
        ['ق202', 40, 'المبنى الرئيسي', 'الثاني'],
        ['معمل حاسوب 1', 30, 'مبنى المعامل', 'الأول'],
    ]
    for i, row in enumerate(examples, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=i, column=col, value=val)

    for col, w in enumerate([18, 10, 22, 12], 1):
        ws.column_dimensions[chr(64 + col)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=rooms_template.xlsx"})


@router.post("/rooms/import")
async def import_rooms(
    faculty_id: str,
    file: bytes = Depends(lambda: None),
    current_user: dict = Depends(get_current_user),
):
    """استيراد قاعات من ملف Excel"""
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    from fastapi import UploadFile, File as FastAPIFile
    raise HTTPException(status_code=400, detail="استخدم /api/rooms/import/upload")


from fastapi import UploadFile, File as FastAPIFile

@router.post("/rooms/import/upload")
async def import_rooms_upload(
    faculty_id: str,
    file: UploadFile = FastAPIFile(...),
    current_user: dict = Depends(get_current_user),
):
    """استيراد قاعات من ملف Excel"""
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    import openpyxl
    from io import BytesIO

    db = get_db()
    content = await file.read()
    filename = file.filename or ""

    rows_data = []
    if filename.endswith('.csv'):
        import csv
        import io
        text = content.decode('utf-8-sig')
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            rows_data.append(row)
        rows_data = rows_data[1:]  # skip header
    elif filename.endswith('.xls'):
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)
        for i in range(1, sheet.nrows):
            rows_data.append([sheet.cell_value(i, j) for j in range(sheet.ncols)])
    else:
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows_data.append(list(row))

    created = 0
    skipped = 0
    errors = []

    for row_num, row in enumerate(rows_data, 2):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        capacity = int(row[1]) if row[1] else 30
        building = str(row[2]).strip() if row[2] else ""
        floor = str(row[3]).strip() if row[3] else ""

        if not name:
            errors.append(f"سطر {row_num}: اسم القاعة فارغ")
            continue

        existing = await db.rooms.find_one({"name": name, "faculty_id": faculty_id})
        if existing:
            skipped += 1
            continue

        await db.rooms.insert_one({
            "name": name,
            "faculty_id": faculty_id,
            "capacity": capacity,
            "building": building,
            "floor": floor,
            "room_type": "lecture",
            "is_active": True,
            "created_at": datetime.now(timezone.utc),
        })
        created += 1

    return {
        "message": f"تم إضافة {created} قاعة جديدة، {skipped} موجودة مسبقاً",
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }



@router.put("/rooms/{room_id}")
async def update_room(room_id: str, data: RoomUpdate, current_user: dict = Depends(get_current_user)):
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    update = {k: v for k, v in data.dict().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="لا توجد بيانات للتحديث")
    await db.rooms.update_one({"_id": ObjectId(room_id)}, {"$set": update})
    return {"message": "تم تحديث القاعة"}


@router.delete("/rooms/{room_id}")
async def delete_room(room_id: str, current_user: dict = Depends(get_current_user)):
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    await db.rooms.delete_one({"_id": ObjectId(room_id)})
    return {"message": "تم حذف القاعة"}


# ===== الفترات الزمنية =====

@router.get("/schedule-settings")
async def get_schedule_settings(
    faculty_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    db = get_db()
    settings_id = f"faculty_{faculty_id}" if faculty_id else "global"
    settings = await db.schedule_settings.find_one({"_id": settings_id})
    if not settings:
        # جلب الإعدادات العامة كقيمة افتراضية
        global_settings = await db.schedule_settings.find_one({"_id": "global"})
        default_slots = [
            {"slot_number": 1, "name": "المحاضرة الأولى", "start_time": "08:00", "end_time": "09:30"},
            {"slot_number": 2, "name": "المحاضرة الثانية", "start_time": "09:45", "end_time": "11:15"},
            {"slot_number": 3, "name": "المحاضرة الثالثة", "start_time": "11:30", "end_time": "13:00"},
            {"slot_number": 4, "name": "المحاضرة الرابعة", "start_time": "13:15", "end_time": "14:45"},
            {"slot_number": 5, "name": "المحاضرة الخامسة", "start_time": "15:00", "end_time": "16:30"},
        ]
        default_days = ["السبت", "الأحد", "الاثنين", "الثلاثاء", "الأربعاء", "الخميس"]
        if global_settings:
            default_slots = global_settings.get("time_slots", default_slots)
            default_days = global_settings.get("working_days", default_days)
        settings = {
            "time_slots": default_slots,
            "working_days": default_days,
        }
    return {
        "time_slots": settings.get("time_slots", []),
        "working_days": settings.get("working_days", []),
    }


@router.put("/schedule-settings")
async def update_schedule_settings(
    time_slots: Optional[List[dict]] = None,
    working_days: Optional[List[str]] = None,
    current_user: dict = Depends(get_current_user),
):
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    update = {}
    if time_slots is not None:
        update["time_slots"] = time_slots
    if working_days is not None:
        update["working_days"] = working_days
    if update:
        await db.schedule_settings.update_one({"_id": "global"}, {"$set": update}, upsert=True)
    return {"message": "تم تحديث الإعدادات"}


DEFAULT_WORKING_DAYS = ["السبت", "الأحد", "الاثنين", "الثلاثاء", "الأربعاء", "الخميس"]


@router.post("/schedule-settings/time-slots")
async def save_time_slots(
    slots: List[TimeSlotCreate],
    faculty_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    settings_id = f"faculty_{faculty_id}" if faculty_id else "global"
    slots_data = [s.dict() for s in slots]
    await db.schedule_settings.update_one(
        {"_id": settings_id},
        {"$set": {"time_slots": slots_data}},
        upsert=True
    )
    # 🛡️ ضمان وجود أيام العمل — يعالج السجلات الناقصة التي كانت تمنع التوليد
    doc = await db.schedule_settings.find_one({"_id": settings_id})
    if not doc.get("working_days"):
        await db.schedule_settings.update_one(
            {"_id": settings_id},
            {"$set": {"working_days": DEFAULT_WORKING_DAYS}}
        )
    return {"message": "تم حفظ الفترات الزمنية"}


class WorkingDaysUpdate(BaseModel):
    days: List[str]


@router.post("/schedule-settings/working-days")
async def save_working_days(
    data: WorkingDaysUpdate,
    faculty_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """💼 حفظ أيام العمل (لكل كلية أو عام)"""
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    if not data.days:
        raise HTTPException(status_code=400, detail="اختر يوم عمل واحداً على الأقل")
    db = get_db()
    settings_id = f"faculty_{faculty_id}" if faculty_id else "global"
    await db.schedule_settings.update_one(
        {"_id": settings_id},
        {"$set": {"working_days": data.days}},
        upsert=True
    )
    return {"message": "تم حفظ أيام العمل"}


# ===== تفضيلات المعلمين =====

@router.get("/teacher-preferences")
async def list_teacher_preferences_summary(department_id: str, current_user: dict = Depends(get_current_user)):
    """📋 ملخص التفضيلات المحفوظة لجميع معلمي قسم معين — لقائمة تبويب تفضيلات المعلمين"""
    db = get_db()
    teachers = await db.teachers.find({"department_id": department_id}).sort("full_name", 1).to_list(1000)
    tids = [str(t["_id"]) for t in teachers]
    prefs_docs = await db.teacher_preferences.find({"teacher_id": {"$in": tids}}).to_list(1000)
    by_tid = {p.get("teacher_id"): p for p in prefs_docs}

    # لعدّ الحقول القديمة (قبل شبكة الفترات) نحتاج عدد الفترات وأيام العمل
    ts_count = await db.time_slots.count_documents({}) or 6
    settings = await db.schedule_settings.find_one({})
    working_days_count = len((settings or {}).get("working_days", []) or []) or 6

    results = []
    for t in teachers:
        tid = str(t["_id"])
        p = by_tid.get(tid)
        unavailable_count = 0
        if p:
            periods = p.get("unavailable_periods") or []
            if periods:
                unavailable_count = len(periods)
            else:
                days_n = len(p.get("unavailable_days") or [])
                slots_n = len(p.get("unavailable_slots") or [])
                unavailable_count = days_n * ts_count + slots_n * max(0, working_days_count - days_n)
        results.append({
            "teacher_id": tid,
            "full_name": t.get("full_name", ""),
            "has_prefs": bool(p),
            "unavailable_count": unavailable_count,
            "max_daily_lectures": (p or {}).get("max_daily_lectures", 2),
            "allow_consecutive_lectures": (p or {}).get("allow_consecutive_lectures", False),
        })
    return results


@router.get("/teacher-preferences/{teacher_id}")
async def get_teacher_preferences(teacher_id: str, current_user: dict = Depends(get_current_user)):
    db = get_db()
    pref = await db.teacher_preferences.find_one({"teacher_id": teacher_id})
    if not pref:
        return {
            "teacher_id": teacher_id,
            "unavailable_days": [],
            "unavailable_slots": [],
            "unavailable_periods": [],
            "max_daily_lectures": 2,
            "allow_consecutive_lectures": False,
        }
    # 🔄 لعرض نظيف: نُرجع الشبكة كما هي إن وُجدت،
    # وإلا نشتقّها من الحقول القديمة (backward compat) لسهولة عمل UI الجديدة
    unavailable_periods = pref.get("unavailable_periods") or []
    if not unavailable_periods and (pref.get("unavailable_days") or pref.get("unavailable_slots")):
        # اشتق الشبكة من الحقول القديمة باستخدام أيام العمل والفترات الحالية
        settings = await db.schedule_settings.find_one({})
        working_days = (settings or {}).get("working_days", ["السبت", "الأحد", "الاثنين", "الثلاثاء", "الأربعاء", "الخميس"])
        ts_docs = await db.time_slots.find({}).to_list(20)
        slot_numbers = sorted([ts.get("slot_number") for ts in ts_docs if ts.get("slot_number")])
        if not slot_numbers:
            slot_numbers = [1, 2, 3, 4, 5]
        unavailable_periods = _derive_unavailable_periods(pref, working_days, slot_numbers)
    return {
        "teacher_id": teacher_id,
        "unavailable_days": pref.get("unavailable_days", []),
        "unavailable_slots": pref.get("unavailable_slots", []),
        "unavailable_periods": unavailable_periods,
        "max_daily_lectures": pref.get("max_daily_lectures", 2),
        "allow_consecutive_lectures": pref.get("allow_consecutive_lectures", False),
    }


@router.put("/teacher-preferences/{teacher_id}")
async def update_teacher_preferences(
    teacher_id: str,
    data: TeacherPreferenceUpdate,
    current_user: dict = Depends(get_current_user),
):
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()

    # 🔄 نطبّع الشبكة (unavailable_periods) ونشتقّ الحقول القديمة للتوافق الرجعي
    # (بحيث كود الجدولة القديم أو النشر السابق على Cloud Run لا ينكسر)
    periods_set = set()
    for p in data.unavailable_periods or []:
        periods_set.add((p.day, int(p.slot_number)))
    # أضف أيضاً كل ما ورد في الحقول القديمة (إن أرسلها الفرونت للتوافق)
    settings = await db.schedule_settings.find_one({})
    working_days = (settings or {}).get("working_days", ["السبت", "الأحد", "الاثنين", "الثلاثاء", "الأربعاء", "الخميس"])
    ts_docs = await db.time_slots.find({}).to_list(20)
    slot_numbers = sorted([ts.get("slot_number") for ts in ts_docs if ts.get("slot_number")])
    if not slot_numbers:
        slot_numbers = [1, 2, 3, 4, 5]
    for d in data.unavailable_days or []:
        for sn in slot_numbers:
            periods_set.add((d, sn))
    for sn in data.unavailable_slots or []:
        for d in working_days:
            periods_set.add((d, sn))

    normalized_periods = [{"day": d, "slot_number": sn} for (d, sn) in sorted(periods_set)]

    # اشتقّ الحقول القديمة المكافئة من الشبكة النهائية:
    # - يوم كامل = يوم كل خلاياه محددة
    # - فترة على كل الأيام = فترة موجودة في كل يوم عمل
    derived_days = []
    for d in working_days:
        if all((d, sn) in periods_set for sn in slot_numbers):
            derived_days.append(d)
    derived_slots = []
    for sn in slot_numbers:
        if all((d, sn) in periods_set for d in working_days):
            derived_slots.append(sn)

    await db.teacher_preferences.update_one(
        {"teacher_id": teacher_id},
        {"$set": {
            "teacher_id": teacher_id,
            "unavailable_days": derived_days,
            "unavailable_slots": derived_slots,
            "unavailable_periods": normalized_periods,
            "max_daily_lectures": data.max_daily_lectures,
            "allow_consecutive_lectures": data.allow_consecutive_lectures,
            "updated_at": datetime.now(timezone.utc),
        }},
        upsert=True
    )
    return {"message": "تم حفظ تفضيلات المعلم"}


# ===== خانات الجدول =====

@router.get("/weekly-schedule")
async def get_weekly_schedule(
    faculty_id: Optional[str] = None,
    department_id: Optional[str] = None,
    level: Optional[int] = None,
    section: Optional[str] = None,
    teacher_id: Optional[str] = None,
    course_id: Optional[str] = None,
    room_id: Optional[str] = None,
    semester_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    db = get_db()
    query = {}
    if semester_id:
        query["semester_id"] = semester_id
    if faculty_id:
        query["faculty_id"] = faculty_id
    if department_id:
        query["department_id"] = department_id
    if level:
        query["level"] = level
    if section:
        query["section"] = section
    if teacher_id:
        query["teacher_id"] = teacher_id
    if course_id:
        query["course_id"] = course_id
    if room_id:
        query["room_id"] = room_id

    slots = await db.weekly_schedule.find(query).to_list(2000)

    # Batch lookup
    course_ids = list({s["course_id"] for s in slots if s.get("course_id")})
    teacher_ids = list({s["teacher_id"] for s in slots if s.get("teacher_id")})
    room_ids = list({s["room_id"] for s in slots if s.get("room_id")})
    dept_ids = list({s["department_id"] for s in slots if s.get("department_id")})

    courses_map = {}
    if course_ids:
        docs = await db.courses.find({"_id": {"$in": [ObjectId(x) for x in course_ids]}}).to_list(500)
        for d in docs:
            courses_map[str(d["_id"])] = d

    teachers_map = {}
    if teacher_ids:
        docs = await db.teachers.find({"_id": {"$in": [ObjectId(x) for x in teacher_ids]}}).to_list(500)
        for d in docs:
            teachers_map[str(d["_id"])] = d

    rooms_map = {}
    if room_ids:
        docs = await db.rooms.find({"_id": {"$in": [ObjectId(x) for x in room_ids]}}).to_list(500)
        for d in docs:
            rooms_map[str(d["_id"])] = d

    depts_map = {}
    if dept_ids:
        docs = await db.departments.find({"_id": {"$in": [ObjectId(x) for x in dept_ids]}}).to_list(100)
        for d in docs:
            depts_map[str(d["_id"])] = d

    result = []
    for s in slots:
        course = courses_map.get(s.get("course_id", ""), {})
        teacher = teachers_map.get(s.get("teacher_id", ""), {})
        room = rooms_map.get(s.get("room_id", ""), {})
        dept = depts_map.get(s.get("department_id", ""), {})
        result.append({
            "id": str(s["_id"]),
            "faculty_id": s.get("faculty_id", ""),
            "department_id": s.get("department_id", ""),
            "department_name": dept.get("name", ""),
            "level": s.get("level"),
            "section": s.get("section", ""),
            "day": s.get("day", ""),
            "slot_number": s.get("slot_number"),
            "course_id": s.get("course_id", ""),
            "course_name": course.get("name", ""),
            "course_code": course.get("code", ""),
            "teacher_id": s.get("teacher_id", ""),
            "teacher_name": teacher.get("full_name", ""),
            "room_id": s.get("room_id", ""),
            "room_name": room.get("name", ""),
        })
    return result


@router.get("/weekly-schedule/conflicts")
async def get_schedule_conflicts(
    faculty_id: Optional[str] = None,
    department_id: Optional[str] = None,
    level: Optional[int] = None,
    section: Optional[str] = None,
    teacher_id: Optional[str] = None,
    semester_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """
    كشف التعارضات في الجدول الأسبوعي الحالي.
    يُرجع قائمة بكل التعارضات: شعبة، معلم، قاعة.
    كل تعارض يحتوي على IDs الـ slots المتعارضة + التفاصيل.
    """
    db = get_db()
    query = {}
    if semester_id:
        query["semester_id"] = semester_id
    if faculty_id:
        query["faculty_id"] = faculty_id
    if department_id:
        query["department_id"] = department_id
    if level:
        query["level"] = level
    if section:
        query["section"] = section
    if teacher_id:
        query["teacher_id"] = teacher_id

    slots = await db.weekly_schedule.find(query).to_list(5000)

    # تجميع الـ slots حسب (اليوم، رقم الفترة)
    by_time: dict = {}  # (day, slot_number) -> [slot_dicts]
    for s in slots:
        key = (s.get("day", ""), s.get("slot_number"))
        by_time.setdefault(key, []).append(s)

    # كشف التعارضات داخل كل فترة زمنية
    section_conflicts: List[dict] = []
    teacher_conflicts: List[dict] = []
    room_conflicts: List[dict] = []
    conflicting_slot_ids: set = set()

    for (day, slot_number), group in by_time.items():
        if len(group) < 2:
            continue  # لا تعارض ممكن

        # 1. تعارض الشعبة: نفس (department + level + section)
        section_groups: dict = {}
        for s in group:
            key = (s.get("department_id", ""), s.get("level"), s.get("section", ""))
            section_groups.setdefault(key, []).append(s)
        for (dept_id, lv, sec), conflicts in section_groups.items():
            if len(conflicts) > 1:
                ids = [str(c["_id"]) for c in conflicts]
                conflicting_slot_ids.update(ids)
                section_conflicts.append({
                    "type": "section",
                    "day": day,
                    "slot_number": slot_number,
                    "department_id": dept_id,
                    "level": lv,
                    "section": sec,
                    "slot_ids": ids,
                    "count": len(conflicts),
                })

        # 2. تعارض المعلم: نفس teacher_id
        teacher_groups: dict = {}
        for s in group:
            tid = s.get("teacher_id", "")
            if tid:
                teacher_groups.setdefault(tid, []).append(s)
        for tid, conflicts in teacher_groups.items():
            if len(conflicts) > 1:
                ids = [str(c["_id"]) for c in conflicts]
                conflicting_slot_ids.update(ids)
                teacher_conflicts.append({
                    "type": "teacher",
                    "day": day,
                    "slot_number": slot_number,
                    "teacher_id": tid,
                    "slot_ids": ids,
                    "count": len(conflicts),
                })

        # 3. تعارض القاعة: نفس room_id
        room_groups: dict = {}
        for s in group:
            rid = s.get("room_id", "")
            if rid:
                room_groups.setdefault(rid, []).append(s)
        for rid, conflicts in room_groups.items():
            if len(conflicts) > 1:
                ids = [str(c["_id"]) for c in conflicts]
                conflicting_slot_ids.update(ids)
                room_conflicts.append({
                    "type": "room",
                    "day": day,
                    "slot_number": slot_number,
                    "room_id": rid,
                    "slot_ids": ids,
                    "count": len(conflicts),
                })

    # جلب أسماء المعلمين والقاعات والأقسام لإثراء التقرير
    teacher_ids = list({c["teacher_id"] for c in teacher_conflicts})
    room_ids = list({c["room_id"] for c in room_conflicts})
    dept_ids = list({c["department_id"] for c in section_conflicts})

    teachers_map: dict = {}
    if teacher_ids:
        docs = await db.teachers.find({"_id": {"$in": [ObjectId(x) for x in teacher_ids]}}).to_list(500)
        teachers_map = {str(d["_id"]): d.get("full_name", "") for d in docs}

    rooms_map: dict = {}
    if room_ids:
        docs = await db.rooms.find({"_id": {"$in": [ObjectId(x) for x in room_ids]}}).to_list(500)
        rooms_map = {str(d["_id"]): d.get("name", "") for d in docs}

    depts_map: dict = {}
    if dept_ids:
        docs = await db.departments.find({"_id": {"$in": [ObjectId(x) for x in dept_ids]}}).to_list(100)
        depts_map = {str(d["_id"]): d.get("name", "") for d in docs}

    for c in section_conflicts:
        c["department_name"] = depts_map.get(c["department_id"], "")
    for c in teacher_conflicts:
        c["teacher_name"] = teachers_map.get(c["teacher_id"], "")
    for c in room_conflicts:
        c["room_name"] = rooms_map.get(c["room_id"], "")

    all_conflicts = section_conflicts + teacher_conflicts + room_conflicts

    return {
        "total_conflicts": len(all_conflicts),
        "total_conflicting_slots": len(conflicting_slot_ids),
        "conflicting_slot_ids": list(conflicting_slot_ids),
        "section_conflicts": section_conflicts,
        "teacher_conflicts": teacher_conflicts,
        "room_conflicts": room_conflicts,
        "all_conflicts": all_conflicts,
    }




@router.post("/weekly-schedule")
async def create_schedule_slot(
    data: ScheduleSlotCreate,
    current_user: dict = Depends(get_current_user),
):
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()

    # === كشف التعارضات ===
    conflicts = []

    # 1. تعارض الشعبة (نفس القسم+المستوى+الشعبة+اليوم+الفترة)
    existing_section = await db.weekly_schedule.find_one({
        "department_id": data.department_id,
        "level": data.level,
        "section": data.section,
        "day": data.day,
        "slot_number": data.slot_number,
    })
    if existing_section:
        c = await db.courses.find_one({"_id": ObjectId(existing_section["course_id"])})
        conflicts.append(f"تعارض شعبة: يوجد مقرر '{c.get('name', '')}' لنفس الشعبة في هذه الفترة")

    # 2. تعارض المعلم
    teacher_busy = await db.weekly_schedule.find_one({
        "teacher_id": data.teacher_id,
        "day": data.day,
        "slot_number": data.slot_number,
    })
    if teacher_busy:
        t = await db.teachers.find_one({"_id": ObjectId(data.teacher_id)})
        c = await db.courses.find_one({"_id": ObjectId(teacher_busy["course_id"])})
        conflicts.append(f"تعارض معلم: '{t.get('full_name', '')}' لديه مقرر '{c.get('name', '')}' في نفس الفترة")

    # 3. تعارض القاعة
    room_busy = await db.weekly_schedule.find_one({
        "room_id": data.room_id,
        "day": data.day,
        "slot_number": data.slot_number,
    })
    if room_busy:
        r = await db.rooms.find_one({"_id": ObjectId(data.room_id)})
        c = await db.courses.find_one({"_id": ObjectId(room_busy["course_id"])})
        conflicts.append(f"تعارض قاعة: '{r.get('name', '')}' مشغولة بمقرر '{c.get('name', '')}' في نفس الفترة")

    # 4. تعارض تفضيلات المعلم
    pref = await db.teacher_preferences.find_one({"teacher_id": data.teacher_id})
    if pref:
        # 🆕 نستخدم _is_period_unavailable الذي يدمج الشبكة الجديدة + الحقول القديمة
        if _is_period_unavailable(pref, data.day, data.slot_number):
            t = await db.teachers.find_one({"_id": ObjectId(data.teacher_id)})
            conflicts.append(f"تعارض تفضيلات: '{t.get('full_name', '')}' غير متاح يوم {data.day} في الفترة {data.slot_number}")
        # Check max daily
        daily_count = await db.weekly_schedule.count_documents({
            "teacher_id": data.teacher_id, "day": data.day
        })
        max_daily = pref.get("max_daily_lectures", 5)
        if daily_count >= max_daily:
            conflicts.append(f"تعارض تفضيلات: المعلم وصل الحد الأقصى ({max_daily} محاضرات) ليوم {data.day}")

    if conflicts:
        raise HTTPException(status_code=409, detail={"message": "يوجد تعارضات", "conflicts": conflicts})

    doc = data.dict()
    doc["created_at"] = datetime.now(timezone.utc)
    doc["created_by"] = current_user["id"]
    try:
        result = await db.weekly_schedule.insert_one(doc)
    except DuplicateKeyError as e:
        # 🔒 خط الدفاع الأخير: MongoDB رفض الحفظ لتعارض فريد.
        # هذا يحدث فقط إن أفلت التعارض من الفحوصات السابقة (race condition نادر أو بيانات قديمة).
        idx_name = str(getattr(e, "details", {}).get("keyPattern", ""))
        if "teacher" in idx_name:
            msg = "تعارض معلم: هذا المعلم لديه محاضرة أخرى في نفس اليوم والفترة"
        elif "room" in idx_name:
            msg = "تعارض قاعة: هذه القاعة محجوزة في نفس اليوم والفترة"
        elif "section" in idx_name or "department" in idx_name:
            msg = "تعارض شعبة: هذه الشعبة لديها محاضرة أخرى في نفس اليوم والفترة"
        else:
            msg = "تعارض في الجدول (تكرار مرفوض من قاعدة البيانات)"
        raise HTTPException(status_code=409, detail={"message": msg, "conflicts": [msg]})
    return {"id": str(result.inserted_id), "message": "تم إضافة المحاضرة في الجدول"}


@router.put("/weekly-schedule/{slot_id}")
async def update_schedule_slot(
    slot_id: str,
    data: ScheduleSlotUpdate,
    current_user: dict = Depends(get_current_user),
):
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    existing = await db.weekly_schedule.find_one({"_id": ObjectId(slot_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="غير موجود")

    update = {k: v for k, v in data.dict().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="لا توجد بيانات")

    # Check conflicts for new values
    # 🔒 نستخدم القيم الجديدة إن أُرسلت، وإلا القيم القديمة
    check_day = update.get("day", existing["day"])
    check_slot = update.get("slot_number", existing["slot_number"])
    check_teacher = update.get("teacher_id", existing.get("teacher_id"))
    check_room = update.get("room_id", existing.get("room_id"))
    check_dept = update.get("department_id", existing.get("department_id"))
    check_level = update.get("level", existing.get("level"))
    check_section = update.get("section", existing.get("section", ""))

    # تعارض الشعبة إن تغيّر أي حقل مؤثر
    if any(k in update for k in ("day", "slot_number", "department_id", "level", "section")):
        section_busy = await db.weekly_schedule.find_one({
            "_id": {"$ne": ObjectId(slot_id)},
            "department_id": check_dept,
            "level": check_level,
            "section": check_section,
            "day": check_day,
            "slot_number": check_slot,
        })
        if section_busy:
            raise HTTPException(status_code=409, detail="تعارض: يوجد محاضرة أخرى لنفس الشعبة في هذه الفترة")

    if check_teacher and any(k in update for k in ("teacher_id", "day", "slot_number")):
        teacher_busy = await db.weekly_schedule.find_one({
            "_id": {"$ne": ObjectId(slot_id)},
            "teacher_id": check_teacher,
            "day": check_day,
            "slot_number": check_slot,
        })
        if teacher_busy:
            raise HTTPException(status_code=409, detail="تعارض: المعلم لديه محاضرة أخرى في نفس الفترة")

    if check_room and any(k in update for k in ("room_id", "day", "slot_number")):
        room_busy = await db.weekly_schedule.find_one({
            "_id": {"$ne": ObjectId(slot_id)},
            "room_id": check_room,
            "day": check_day,
            "slot_number": check_slot,
        })
        if room_busy:
            raise HTTPException(status_code=409, detail="تعارض: القاعة مشغولة في نفس الفترة")

    try:
        await db.weekly_schedule.update_one({"_id": ObjectId(slot_id)}, {"$set": update})
    except DuplicateKeyError as e:
        idx_name = str(getattr(e, "details", {}).get("keyPattern", ""))
        if "teacher" in idx_name:
            msg = "تعارض معلم: هذا المعلم لديه محاضرة أخرى في نفس اليوم والفترة"
        elif "room" in idx_name:
            msg = "تعارض قاعة: هذه القاعة محجوزة في نفس اليوم والفترة"
        elif "section" in idx_name or "department" in idx_name:
            msg = "تعارض شعبة: هذه الشعبة لديها محاضرة أخرى في نفس اليوم والفترة"
        else:
            msg = "تعارض في الجدول (تكرار مرفوض من قاعدة البيانات)"
        raise HTTPException(status_code=409, detail=msg)
    return {"message": "تم التحديث"}


@router.delete("/weekly-schedule/{slot_id}")
async def delete_schedule_slot(
    slot_id: str,
    current_user: dict = Depends(get_current_user),
):
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    await db.weekly_schedule.delete_one({"_id": ObjectId(slot_id)})
    return {"message": "تم الحذف"}


@router.delete("/weekly-schedule/clear/all")
async def clear_schedule(
    faculty_id: Optional[str] = None,
    department_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """مسح كامل الجدول أو جدول كلية/قسم"""
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    query = {}
    if faculty_id:
        query["faculty_id"] = faculty_id
    if department_id:
        query["department_id"] = department_id
    result = await db.weekly_schedule.delete_many(query)
    return {"message": f"تم حذف {result.deleted_count} خانة"}


# ===== التوليد شبه التلقائي =====

@router.post("/weekly-schedule/auto-generate")
async def auto_generate_schedule(
    faculty_id: str,
    department_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """توليد جدول تلقائي مع مراعاة التعارضات والتفضيلات"""
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")

    db = get_db()

    # Get settings (faculty-specific or global)
    settings = await db.schedule_settings.find_one({"_id": f"faculty_{faculty_id}"})
    if not settings:
        settings = await db.schedule_settings.find_one({"_id": "global"})
    if not settings:
        raise HTTPException(status_code=400, detail="يرجى إعداد الفترات الزمنية أولاً")

    time_slots = settings.get("time_slots", [])
    working_days = settings.get("working_days", [])
    if not time_slots:
        raise HTTPException(status_code=400, detail="يرجى إعداد الفترات الزمنية أولاً — تبويب «الفترات» ثم حفظ")
    if not working_days:
        raise HTTPException(status_code=400, detail="أيام العمل غير محددة — افتح تبويب «الفترات»، حدد أيام العمل ثم احفظ")

    slot_numbers = sorted([s["slot_number"] for s in time_slots])

    # Get departments
    dept_query = {"faculty_id": faculty_id, "is_active": {"$ne": False}}
    if department_id:
        dept_query["_id"] = ObjectId(department_id)
    departments = await db.departments.find(dept_query).to_list(100)

    # Get rooms for this faculty only
    rooms = await db.rooms.find({"is_active": True, "faculty_id": faculty_id}).to_list(200)
    if not rooms:
        raise HTTPException(status_code=400, detail="لا توجد قاعات مسجلة لهذه الكلية. أضف قاعات أولاً من تبويب 'القاعات'")

    # Get teacher preferences
    all_prefs = await db.teacher_preferences.find().to_list(500)
    prefs_map = {p["teacher_id"]: p for p in all_prefs}

    # Track occupancy: (day, slot) -> set of teacher_ids, room_ids
    teacher_occupied = {}  # (day, slot) -> set of teacher_ids
    room_occupied = {}     # (day, slot) -> set of room_ids
    section_occupied = {}  # (dept_id, level, section, day, slot) -> True
    teacher_daily_count = {}  # (teacher_id, day) -> count
    teacher_day_slots = {}  # (teacher_id, day) -> set of slot_numbers (to detect consecutive)

    # Load existing schedule to avoid conflicts.
    # 🔒 المعلم قد يُدرِّس في كليات متعددة → نحمّل الجدول عالمياً لتعارض المعلم.
    # القاعات والشعب محلية بالكلية فقط.
    existing_global = await db.weekly_schedule.find({}).to_list(20000)
    for e in existing_global:
        key = (e["day"], e["slot_number"])
        # tracking عالمي للمعلم (يشمل كل الكليات)
        if e.get("teacher_id"):
            teacher_occupied.setdefault(key, set()).add(e["teacher_id"])
            dk = (e["teacher_id"], e["day"])
            teacher_daily_count[dk] = teacher_daily_count.get(dk, 0) + 1
            teacher_day_slots.setdefault(dk, set()).add(e["slot_number"])
        # tracking محلي للقاعة والشعبة (مقيّد بالكلية الحالية)
        if e.get("faculty_id") == faculty_id:
            if e.get("room_id"):
                room_occupied.setdefault(key, set()).add(e["room_id"])
            sk = (e["department_id"], e["level"], e.get("section", ""), e["day"], e["slot_number"])
            section_occupied[sk] = True

    created = 0
    skipped = 0
    errors = []

    for dept in departments:
        dept_id = str(dept["_id"])

        # Get courses for this department
        courses = await db.courses.find({
            "department_id": dept_id, "is_active": True, "teacher_id": {"$ne": None}
        }).to_list(500)

        # Group by (level, section)
        groups = {}
        for c in courses:
            key = (c.get("level", 1), c.get("section", ""))
            groups.setdefault(key, []).append(c)

        for (level, section), group_courses in groups.items():
            for course in group_courses:
                cid = str(course["_id"])
                tid = course.get("teacher_id", "")
                credit = course.get("credit_hours", 3)

                # عدد المحاضرات المطلوبة
                if credit <= 2:
                    needed = 1
                elif credit <= 3:
                    needed = 2
                else:
                    needed = 3

                # Check if already scheduled
                existing_count = await db.weekly_schedule.count_documents({
                    "course_id": cid, "department_id": dept_id,
                    "level": level, "section": section,
                })
                needed -= existing_count
                if needed <= 0:
                    continue

                pref = prefs_map.get(tid, {})
                max_daily = pref.get("max_daily_lectures", 2)
                allow_consecutive = pref.get("allow_consecutive_lectures", False)

                placed = 0
                for day in working_days:
                    if placed >= needed:
                        break
                    # 🆕 تجاوز اليوم إن كانت كل خلاياه محظورة (يوم كامل بلغة الشبكة)
                    if all(_is_period_unavailable(pref, day, sn) for sn in slot_numbers):
                        continue

                    for sn in slot_numbers:
                        if placed >= needed:
                            break
                        # 🆕 فحص الخلية (يوم × فترة) عبر الشبكة الجديدة + توافق رجعي
                        if _is_period_unavailable(pref, day, sn):
                            continue

                        key_ts = (day, sn)
                        sk = (dept_id, level, section, day, sn)

                        # Check section free
                        if sk in section_occupied:
                            continue
                        # Check teacher free
                        if tid in teacher_occupied.get(key_ts, set()):
                            continue
                        # Check teacher daily limit
                        dk = (tid, day)
                        if teacher_daily_count.get(dk, 0) >= max_daily:
                            continue
                        # Check consecutive slots (unless teacher allows)
                        if not allow_consecutive:
                            taken_slots = teacher_day_slots.get(dk, set())
                            if (sn - 1) in taken_slots or (sn + 1) in taken_slots:
                                continue

                        # Find free room
                        busy_rooms = room_occupied.get(key_ts, set())
                        free_room = None
                        for room in rooms:
                            if str(room["_id"]) not in busy_rooms:
                                free_room = room
                                break
                        if not free_room:
                            continue

                        rid = str(free_room["_id"])

                        # Place it!
                        doc = {
                            "faculty_id": faculty_id,
                            "department_id": dept_id,
                            "level": level,
                            "section": section,
                            "day": day,
                            "slot_number": sn,
                            "course_id": cid,
                            "teacher_id": tid,
                            "room_id": rid,
                            "created_at": datetime.now(timezone.utc),
                            "created_by": current_user["id"],
                            "auto_generated": True,
                        }
                        try:
                            await db.weekly_schedule.insert_one(doc)
                        except DuplicateKeyError:
                            # فهرس DB منع التكرار (حماية إضافية) — تجاوز هذه الخلية
                            skipped += 1
                            continue

                        # Update tracking
                        teacher_occupied.setdefault(key_ts, set()).add(tid)
                        room_occupied.setdefault(key_ts, set()).add(rid)
                        section_occupied[sk] = True
                        teacher_daily_count[dk] = teacher_daily_count.get(dk, 0) + 1
                        teacher_day_slots.setdefault(dk, set()).add(sn)

                        created += 1
                        placed += 1

                if placed < needed:
                    remaining = needed - placed
                    errors.append(f"{course.get('name', '')} ({course.get('code', '')}): لم يتم جدولة {remaining} محاضرة")

    return {
        "message": f"تم إنشاء {created} محاضرة في الجدول",
        "created": created,
        "errors": errors,
    }


# ============================================================
# Schedule Drafts - نسخ احتياطية مؤقتة للجدول للمقارنة
# ============================================================


class SaveDraftRequest(BaseModel):
    name: str
    faculty_id: Optional[str] = None
    department_id: Optional[str] = None
    semester_id: Optional[str] = None
    notes: Optional[str] = ""


@router.get("/weekly-schedule/drafts")
async def list_drafts(current_user: dict = Depends(get_current_user)):
    """قائمة نسخ الجدول المحفوظة"""
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    drafts = await db.weekly_schedule_drafts.find({}).sort("created_at", -1).to_list(50)
    result = []
    for d in drafts:
        result.append({
            "id": str(d["_id"]),
            "name": d.get("name", ""),
            "notes": d.get("notes", ""),
            "slots_count": len(d.get("slots", [])),
            "faculty_id": d.get("faculty_id"),
            "department_id": d.get("department_id"),
            "semester_id": d.get("semester_id"),
            "created_at": d.get("created_at").isoformat() if d.get("created_at") else None,
            "created_by_name": d.get("created_by_name", ""),
        })
    return result


@router.post("/weekly-schedule/drafts")
async def save_draft(req: SaveDraftRequest, current_user: dict = Depends(get_current_user)):
    """حفظ snapshot للجدول الحالي كنسخة احتياطية"""
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    query = {}
    if req.faculty_id:
        query["faculty_id"] = req.faculty_id
    if req.department_id:
        query["department_id"] = req.department_id
    if req.semester_id:
        query["semester_id"] = req.semester_id

    slots = await db.weekly_schedule.find(query).to_list(5000)
    clean_slots = []
    for s in slots:
        c = {k: v for k, v in s.items() if k != "_id"}
        c["_orig_id"] = str(s["_id"])
        clean_slots.append(c)

    draft_doc = {
        "name": req.name.strip() or f"نسخة {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}",
        "notes": req.notes or "",
        "slots": clean_slots,
        "faculty_id": req.faculty_id,
        "department_id": req.department_id,
        "semester_id": req.semester_id,
        "created_at": datetime.now(timezone.utc),
        "created_by": current_user.get("id"),
        "created_by_name": current_user.get("full_name", current_user.get("username", "")),
    }
    result = await db.weekly_schedule_drafts.insert_one(draft_doc)

    # حذف النسخ القديمة (الإبقاء على آخر 10 لكل نطاق)
    scope_query = {
        "faculty_id": req.faculty_id,
        "department_id": req.department_id,
        "semester_id": req.semester_id,
    }
    all_in_scope = await db.weekly_schedule_drafts.find(scope_query).sort("created_at", -1).to_list(100)
    if len(all_in_scope) > 10:
        for old in all_in_scope[10:]:
            await db.weekly_schedule_drafts.delete_one({"_id": old["_id"]})

    return {
        "id": str(result.inserted_id),
        "message": f"تم حفظ النسخة ({len(clean_slots)} محاضرة)",
        "slots_count": len(clean_slots),
    }


@router.delete("/weekly-schedule/drafts/{draft_id}")
async def delete_draft(draft_id: str, current_user: dict = Depends(get_current_user)):
    """حذف نسخة احتياطية"""
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    result = await db.weekly_schedule_drafts.delete_one({"_id": ObjectId(draft_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="النسخة غير موجودة")
    return {"message": "تم حذف النسخة"}


@router.post("/weekly-schedule/drafts/{draft_id}/restore")
async def restore_draft(
    draft_id: str,
    backup_current: bool = True,
    backup_name: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """استرجاع نسخة (تحل محل الجدول الحالي في نفس النطاق)
    قبل الاستبدال، يتم حفظ الجدول الحالي تلقائياً كنسخة احتياطية (backup_current=True افتراضياً)،
    حتى لا تُفقد النسخة الحالية أبداً.
    """
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    draft = await db.weekly_schedule_drafts.find_one({"_id": ObjectId(draft_id)})
    if not draft:
        raise HTTPException(status_code=404, detail="النسخة غير موجودة")

    query = {}
    if draft.get("faculty_id"):
        query["faculty_id"] = draft["faculty_id"]
    if draft.get("department_id"):
        query["department_id"] = draft["department_id"]
    if draft.get("semester_id"):
        query["semester_id"] = draft["semester_id"]

    # 1) حفظ الجدول الحالي تلقائياً كنسخة احتياطية قبل الاستبدال
    backup_id = None
    backup_slots_count = 0
    if backup_current:
        current_slots = await db.weekly_schedule.find(query).to_list(5000)
        if current_slots:
            backup_clean = []
            for s in current_slots:
                c = {k: v for k, v in s.items() if k != "_id"}
                c["_orig_id"] = str(s["_id"])
                backup_clean.append(c)
            auto_name = backup_name or f"نسخة تلقائية قبل استعادة \"{draft.get('name','')}\" - {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}"
            backup_doc = {
                "name": auto_name,
                "notes": f"تم إنشاؤها تلقائياً قبل استعادة النسخة {draft.get('name','')}",
                "slots": backup_clean,
                "faculty_id": draft.get("faculty_id"),
                "department_id": draft.get("department_id"),
                "semester_id": draft.get("semester_id"),
                "created_at": datetime.now(timezone.utc),
                "created_by": current_user.get("id"),
                "created_by_name": current_user.get("full_name", current_user.get("username", "")),
                "auto_backup": True,
                "restored_from": str(draft["_id"]),
            }
            res = await db.weekly_schedule_drafts.insert_one(backup_doc)
            backup_id = str(res.inserted_id)
            backup_slots_count = len(backup_clean)

    # 2) استبدال الجدول الحالي بمحتوى النسخة المختارة
    deleted = await db.weekly_schedule.delete_many(query)
    inserted = 0
    for slot in draft.get("slots", []):
        slot_copy = {k: v for k, v in slot.items() if k != "_orig_id"}
        await db.weekly_schedule.insert_one(slot_copy)
        inserted += 1

    return {
        "message": f"تم استرجاع النسخة: حذف {deleted.deleted_count} وإضافة {inserted}",
        "deleted": deleted.deleted_count,
        "inserted": inserted,
        "backup_id": backup_id,
        "backup_slots_count": backup_slots_count,
        "backup_created": backup_id is not None,
    }


@router.get("/weekly-schedule/drafts/{draft_id}/compare")
async def compare_draft(draft_id: str, current_user: dict = Depends(get_current_user)):
    """مقارنة النسخة المحفوظة بالجدول الحالي - مقارنة حقيقية slot-by-slot
    تُرجع:
      - added: محاضرات موجودة في الحالي وغير موجودة في النسخة (مضافة بعد حفظ النسخة)
      - removed: محاضرات موجودة في النسخة وغير موجودة في الحالي (حُذفت)
      - changed: محاضرات بنفس المعرف (course/day/slot/section) لكن تغير فيها شيء (المعلم/القاعة)
      - unchanged_count: عدد المحاضرات المتطابقة
    """
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()
    draft = await db.weekly_schedule_drafts.find_one({"_id": ObjectId(draft_id)})
    if not draft:
        raise HTTPException(status_code=404, detail="النسخة غير موجودة")

    query = {}
    if draft.get("faculty_id"):
        query["faculty_id"] = draft["faculty_id"]
    if draft.get("department_id"):
        query["department_id"] = draft["department_id"]
    if draft.get("semester_id"):
        query["semester_id"] = draft["semester_id"]
    current_slots_raw = await db.weekly_schedule.find(query).to_list(5000)

    # تطبيع: نزع _id وتحويل التواريخ
    def normalize(slots):
        out = []
        for s in slots:
            c = {k: v for k, v in s.items() if k not in ("_id", "_orig_id", "created_at", "updated_at")}
            out.append(c)
        return out

    draft_slots = normalize(draft.get("slots", []))
    current_slots = normalize(current_slots_raw)

    # ====== إثراء الـ slots بالأسماء (الـ DB يخزن IDs فقط) ======
    all_course_ids = set()
    all_teacher_ids = set()
    all_room_ids = set()
    all_dept_ids = set()
    for s in draft_slots + current_slots:
        if s.get("course_id"): all_course_ids.add(s["course_id"])
        if s.get("teacher_id"): all_teacher_ids.add(s["teacher_id"])
        if s.get("room_id"): all_room_ids.add(s["room_id"])
        if s.get("department_id"): all_dept_ids.add(s["department_id"])

    courses_map, teachers_map, rooms_map, depts_map = {}, {}, {}, {}
    if all_course_ids:
        async for d in db.courses.find({"_id": {"$in": [ObjectId(x) for x in all_course_ids]}}):
            courses_map[str(d["_id"])] = d
    if all_teacher_ids:
        async for d in db.teachers.find({"_id": {"$in": [ObjectId(x) for x in all_teacher_ids]}}):
            teachers_map[str(d["_id"])] = d
    if all_room_ids:
        async for d in db.rooms.find({"_id": {"$in": [ObjectId(x) for x in all_room_ids]}}):
            rooms_map[str(d["_id"])] = d
    if all_dept_ids:
        async for d in db.departments.find({"_id": {"$in": [ObjectId(x) for x in all_dept_ids]}}):
            depts_map[str(d["_id"])] = d

    def attach_names(s):
        course = courses_map.get(s.get("course_id", ""), {})
        teacher = teachers_map.get(s.get("teacher_id", ""), {})
        room = rooms_map.get(s.get("room_id", ""), {})
        dept = depts_map.get(s.get("department_id", ""), {})
        s["course_name"] = s.get("course_name") or course.get("name", "")
        s["course_code"] = s.get("course_code") or course.get("code", "")
        s["teacher_name"] = s.get("teacher_name") or teacher.get("full_name", "")
        s["room_name"] = s.get("room_name") or room.get("name", "")
        s["department_name"] = s.get("department_name") or dept.get("name", "")
        return s

    draft_slots = [attach_names(s) for s in draft_slots]
    current_slots = [attach_names(s) for s in current_slots]

    # مفتاح هوية المحاضرة (يحدد "نفس الخانة") - يدعم day أو day_of_week
    def slot_key(s):
        return (
            str(s.get("course_id", "")),
            str(s.get("day") or s.get("day_of_week") or ""),
            int(s.get("slot_number", 0) or 0),
            str(s.get("section", "") or ""),
            int(s.get("level", 0) or 0),
        )

    # حقول مقارنة المحتوى (ما الذي يمكن أن يتغير)
    COMPARE_FIELDS = ["teacher_id", "teacher_name", "room_id", "room_name", "notes"]

    draft_map = {slot_key(s): s for s in draft_slots}
    current_map = {slot_key(s): s for s in current_slots}

    draft_keys = set(draft_map.keys())
    current_keys = set(current_map.keys())

    added_keys = current_keys - draft_keys      # في الحالي فقط = أُضيفت
    removed_keys = draft_keys - current_keys    # في النسخة فقط = حُذفت
    common_keys = draft_keys & current_keys

    def enrich(s):
        return {
            "course_id": s.get("course_id"),
            "course_code": s.get("course_code"),
            "course_name": s.get("course_name"),
            "day_of_week": s.get("day") or s.get("day_of_week"),
            "slot_number": s.get("slot_number"),
            "section": s.get("section"),
            "level": s.get("level"),
            "teacher_id": s.get("teacher_id"),
            "teacher_name": s.get("teacher_name"),
            "room_id": s.get("room_id"),
            "room_name": s.get("room_name"),
            "department_name": s.get("department_name"),
            "faculty_name": s.get("faculty_name"),
        }

    added = [enrich(current_map[k]) for k in added_keys]
    removed = [enrich(draft_map[k]) for k in removed_keys]

    changed = []
    unchanged_count = 0
    for k in common_keys:
        d = draft_map[k]
        c = current_map[k]
        diffs = {}
        for f in COMPARE_FIELDS:
            if (d.get(f) or "") != (c.get(f) or ""):
                diffs[f] = {"draft": d.get(f), "current": c.get(f)}
        if diffs:
            changed.append({**enrich(c), "diffs": diffs})
        else:
            unchanged_count += 1

    # ترتيب موحّد لسهولة العرض
    def sort_key(x):
        return (x.get("day_of_week") or "", x.get("slot_number") or 0, x.get("course_code") or "")
    added.sort(key=sort_key)
    removed.sort(key=sort_key)
    changed.sort(key=sort_key)

    def stats(slots):
        teachers, rooms, courses = set(), set(), set()
        by_day = {}
        for s in slots:
            if s.get("teacher_id"): teachers.add(s["teacher_id"])
            if s.get("room_id"): rooms.add(s["room_id"])
            if s.get("course_id"): courses.add(s["course_id"])
            day = s.get("day") or s.get("day_of_week")
            if day:
                by_day[day] = by_day.get(day, 0) + 1
        return {
            "total_slots": len(slots),
            "teachers_count": len(teachers),
            "rooms_count": len(rooms),
            "courses_count": len(courses),
            "by_day": by_day,
        }

    return {
        "draft_name": draft.get("name", ""),
        "draft_created_at": draft.get("created_at").isoformat() if draft.get("created_at") else None,
        "draft_stats": stats(draft_slots),
        "current_stats": stats(current_slots),
        "diff": {
            "added": added,
            "added_count": len(added),
            "removed": removed,
            "removed_count": len(removed),
            "changed": changed,
            "changed_count": len(changed),
            "unchanged_count": unchanged_count,
        },
    }


# ============================================================
# Visual Export - تصدير مرئي للجدول (PDF + Excel)
# ============================================================

@router.get("/weekly-schedule/export-visual/pdf")
async def export_visual_pdf(
    faculty_id: Optional[str] = None,
    department_id: Optional[str] = None,
    level: Optional[int] = None,
    section: Optional[str] = None,
    teacher_id: Optional[str] = None,
    room_id: Optional[str] = None,
    semester_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """تصدير PDF بتصميم بصري احترافي حسب الفلاتر"""
    try:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import arabic_reshaper
        from bidi.algorithm import get_display
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"مكتبة PDF غير مثبتة: {e}")

    db = get_db()

    try:
        pdfmetrics.registerFont(TTFont("Amiri", "/app/backend/backend/fonts/Amiri-Regular.ttf"))
        font_name = "Amiri"
    except Exception:
        font_name = "Helvetica"

    def ar(text):
        if not text:
            return ""
        try:
            return get_display(arabic_reshaper.reshape(str(text)))
        except Exception:
            return str(text)

    query = {}
    if faculty_id: query["faculty_id"] = faculty_id
    if department_id: query["department_id"] = department_id
    if level is not None: query["level"] = level
    if section: query["section"] = section
    if teacher_id: query["teacher_id"] = teacher_id
    if room_id: query["room_id"] = room_id
    if semester_id: query["semester_id"] = semester_id

    slots = await db.weekly_schedule.find(query).to_list(5000)

    settings = None
    if faculty_id:
        settings = await db.schedule_settings.find_one({"_id": f"faculty_{faculty_id}"})
    if not settings:
        settings = await db.schedule_settings.find_one({"_id": "global"})
    time_slots_cfg = sorted((settings or {}).get("time_slots", []), key=lambda x: x.get("slot_number", 0))
    working_days = (settings or {}).get("working_days", ["السبت", "الأحد", "الاثنين", "الثلاثاء", "الأربعاء"])

    slots_by_cell = {}
    for s in slots:
        key = (s.get("day_of_week"), s.get("slot_number"))
        slots_by_cell.setdefault(key, []).append(s)

    course_ids = {s.get("course_id") for s in slots if s.get("course_id")}
    teacher_ids = {s.get("teacher_id") for s in slots if s.get("teacher_id")}
    room_ids_set = {s.get("room_id") for s in slots if s.get("room_id")}

    courses_map = {}
    if course_ids:
        async for c in db.courses.find({"_id": {"$in": [ObjectId(x) for x in course_ids if x]}}):
            courses_map[str(c["_id"])] = c
    teachers_map = {}
    if teacher_ids:
        async for t in db.teachers.find({"_id": {"$in": [ObjectId(x) for x in teacher_ids if x]}}):
            teachers_map[str(t["_id"])] = t
    rooms_map = {}
    if room_ids_set:
        async for r in db.rooms.find({"_id": {"$in": [ObjectId(x) for x in room_ids_set if x]}}):
            rooms_map[str(r["_id"])] = r

    title_parts = ["الجدول الأسبوعي"]
    if teacher_id and teacher_id in teachers_map:
        title_parts.append(f"للأستاذ: {teachers_map[teacher_id].get('full_name', '')}")
    if room_id and room_id in rooms_map:
        title_parts.append(f"للقاعة: {rooms_map[room_id].get('name', '')}")
    if department_id:
        dept = await db.departments.find_one({"_id": ObjectId(department_id)})
        if dept:
            title_parts.append(f"قسم: {dept.get('name', '')}")
    if level is not None:
        title_parts.append(f"المستوى {level}")
    if section:
        title_parts.append(f"شعبة {section}")
    title = " - ".join(title_parts)
    if semester_id:
        sem = await db.semesters.find_one({"_id": ObjectId(semester_id)})
        if sem:
            title += f" | {sem.get('name', '')} {sem.get('academic_year', '')}"

    import io
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=0.7*cm, leftMargin=0.7*cm, topMargin=0.7*cm, bottomMargin=0.7*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"], fontName=font_name, fontSize=16, alignment=1, textColor=colors.HexColor("#1565c0"))
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], fontName=font_name, fontSize=9, alignment=1, textColor=colors.grey)

    story = []
    story.append(Paragraph(ar(title), title_style))
    story.append(Paragraph(ar(f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  جامعة الأحقاف"), sub_style))
    story.append(Spacer(1, 0.3*cm))

    header_row = [ar("الفترة")] + [ar(d) for d in working_days]
    table_data = [header_row]

    for ts in time_slots_cfg:
        slot_num = ts.get("slot_number")
        time_str = f"{ts.get('start_time','')}\n{ts.get('end_time','')}"
        row = [ar(f"الفترة {slot_num}\n{time_str}")]
        for day in working_days:
            cell_slots = slots_by_cell.get((day, slot_num), [])
            if not cell_slots:
                row.append("")
            else:
                parts = []
                for s in cell_slots:
                    course = courses_map.get(s.get("course_id", ""), {})
                    teacher = teachers_map.get(s.get("teacher_id", ""), {})
                    room = rooms_map.get(s.get("room_id", ""), {})
                    line = course.get("name", "") or course.get("code", "")
                    if teacher.get("full_name") and not teacher_id:
                        line += f"\n{teacher.get('full_name','')}"
                    if room.get("name") and not room_id:
                        line += f"\n[{room.get('name','')}]"
                    sec = s.get("section")
                    if sec and not section:
                        line += f" - شعبة {sec}"
                    parts.append(ar(line))
                row.append("\n\n".join(parts))
        table_data.append(row)

    num_cols = len(header_row)
    col_widths = [3*cm] + [(27 - 3) / (num_cols - 1) * cm] * (num_cols - 1)

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    base_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565c0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#e3f2fd")),
        ("TEXTCOLOR", (0, 1), (0, -1), colors.HexColor("#0d47a1")),
        ("FONTSIZE", (0, 1), (0, -1), 9),
        ("BACKGROUND", (1, 1), (-1, -1), colors.HexColor("#fafafa")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#bdbdbd")),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#1565c0")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]
    for row_idx in range(1, len(table_data)):
        for col_idx in range(1, num_cols):
            if not table_data[row_idx][col_idx]:
                base_style.append(("BACKGROUND", (col_idx, row_idx), (col_idx, row_idx), colors.HexColor("#f5f5f5")))
    tbl.setStyle(TableStyle(base_style))

    story.append(tbl)
    story.append(Spacer(1, 0.4*cm))
    legend_style = ParagraphStyle("Legend", fontName=font_name, fontSize=8, alignment=1, textColor=colors.grey)
    story.append(Paragraph(ar(f"إجمالي المحاضرات: {len(slots)}"), legend_style))

    doc.build(story)
    buf.seek(0)
    filename = f"weekly_schedule_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/weekly-schedule/export-visual/excel")
async def export_visual_excel(
    faculty_id: Optional[str] = None,
    department_id: Optional[str] = None,
    level: Optional[int] = None,
    section: Optional[str] = None,
    teacher_id: Optional[str] = None,
    room_id: Optional[str] = None,
    semester_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """تصدير Excel بنفس فلاتر PDF"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"مكتبة Excel غير مثبتة: {e}")

    db = get_db()
    query = {}
    if faculty_id: query["faculty_id"] = faculty_id
    if department_id: query["department_id"] = department_id
    if level is not None: query["level"] = level
    if section: query["section"] = section
    if teacher_id: query["teacher_id"] = teacher_id
    if room_id: query["room_id"] = room_id
    if semester_id: query["semester_id"] = semester_id

    slots = await db.weekly_schedule.find(query).to_list(5000)

    course_ids = {s.get("course_id") for s in slots if s.get("course_id")}
    teacher_ids = {s.get("teacher_id") for s in slots if s.get("teacher_id")}
    room_ids_set = {s.get("room_id") for s in slots if s.get("room_id")}

    courses_map = {}
    if course_ids:
        async for c in db.courses.find({"_id": {"$in": [ObjectId(x) for x in course_ids if x]}}):
            courses_map[str(c["_id"])] = c
    teachers_map = {}
    if teacher_ids:
        async for t in db.teachers.find({"_id": {"$in": [ObjectId(x) for x in teacher_ids if x]}}):
            teachers_map[str(t["_id"])] = t
    rooms_map = {}
    if room_ids_set:
        async for r in db.rooms.find({"_id": {"$in": [ObjectId(x) for x in room_ids_set if x]}}):
            rooms_map[str(r["_id"])] = r

    settings = None
    if faculty_id:
        settings = await db.schedule_settings.find_one({"_id": f"faculty_{faculty_id}"})
    if not settings:
        settings = await db.schedule_settings.find_one({"_id": "global"})
    time_slots_cfg = sorted((settings or {}).get("time_slots", []), key=lambda x: x.get("slot_number", 0))
    working_days = (settings or {}).get("working_days", ["السبت", "الأحد", "الاثنين", "الثلاثاء", "الأربعاء"])

    wb = Workbook()
    ws = wb.active
    ws.title = "الجدول الأسبوعي"
    ws.sheet_view.rightToLeft = True

    title_parts = ["الجدول الأسبوعي"]
    if teacher_id and teacher_id in teachers_map:
        title_parts.append(f"للأستاذ: {teachers_map[teacher_id].get('full_name','')}")
    if room_id and room_id in rooms_map:
        title_parts.append(f"للقاعة: {rooms_map[room_id].get('name','')}")
    if level is not None:
        title_parts.append(f"م{level}")
    if section:
        title_parts.append(f"شعبة {section}")

    num_cols = len(working_days) + 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1, value=" - ".join(title_parts))
    cell.font = Font(bold=True, size=16, color="1565c0")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    header_fill = PatternFill(start_color="1565c0", end_color="1565c0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="bdbdbd"),
        right=Side(style="thin", color="bdbdbd"),
        top=Side(style="thin", color="bdbdbd"),
        bottom=Side(style="thin", color="bdbdbd"),
    )

    headers = ["الفترة"] + working_days
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
    ws.row_dimensions[3].height = 24

    slots_by_cell = {}
    for s in slots:
        key = (s.get("day_of_week"), s.get("slot_number"))
        slots_by_cell.setdefault(key, []).append(s)

    cell_fill = PatternFill(start_color="fafafa", end_color="fafafa", fill_type="solid")
    empty_fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
    time_fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
    time_font = Font(bold=True, color="0d47a1", size=10)

    for r_idx, ts in enumerate(time_slots_cfg, start=4):
        slot_num = ts.get("slot_number")
        time_str = f"الفترة {slot_num}\n{ts.get('start_time','')}\n{ts.get('end_time','')}"
        c = ws.cell(row=r_idx, column=1, value=time_str)
        c.fill = time_fill
        c.font = time_font
        c.alignment = center
        c.border = border
        for d_idx, day in enumerate(working_days, start=2):
            cell_slots = slots_by_cell.get((day, slot_num), [])
            cell_val = ""
            if cell_slots:
                parts = []
                for s in cell_slots:
                    course = courses_map.get(s.get("course_id", ""), {})
                    teacher = teachers_map.get(s.get("teacher_id", ""), {})
                    room = rooms_map.get(s.get("room_id", ""), {})
                    line = course.get("name", "") or course.get("code", "")
                    if teacher.get("full_name") and not teacher_id:
                        line += f"\n{teacher.get('full_name','')}"
                    if room.get("name") and not room_id:
                        line += f"\n[{room.get('name','')}]"
                    sec = s.get("section")
                    if sec and not section:
                        line += f" - ش/{sec}"
                    parts.append(line)
                cell_val = "\n\n".join(parts)
            c = ws.cell(row=r_idx, column=d_idx, value=cell_val)
            c.fill = empty_fill if not cell_val else cell_fill
            c.alignment = center
            c.border = border
            c.font = Font(size=10)
        ws.row_dimensions[r_idx].height = 60

    ws.column_dimensions["A"].width = 15
    for i in range(2, num_cols + 1):
        ws.column_dimensions[chr(64 + i)].width = 28

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"weekly_schedule_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ============================================================
# Availability Reports - تقارير الفراغ/الإشغال للقاعات والأساتذة
# ============================================================

@router.get("/weekly-schedule/availability/rooms")
async def rooms_availability(
    day_of_week: Optional[str] = None,
    slot_number: Optional[int] = None,
    faculty_id: Optional[str] = None,
    semester_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """تقرير حالة القاعات (فارغة/مشغولة) لكل يوم/فترة.
    - بدون فلاتر: يرجع heatmap كاملة (كل القاعات × كل الأيام × كل الفترات)
    - مع day + slot: يرجع قائمة القاعات الفارغة والمشغولة لتلك اللحظة
    """
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()

    # جلب القاعات
    rooms_query = {"is_active": {"$ne": False}}
    if faculty_id:
        rooms_query["faculty_id"] = faculty_id
    rooms = await db.rooms.find(rooms_query).to_list(500)

    # جلب الجدول
    sched_query = {}
    if faculty_id:
        sched_query["faculty_id"] = faculty_id
    if semester_id:
        sched_query["semester_id"] = semester_id
    if day_of_week:
        sched_query["day_of_week"] = day_of_week
    if slot_number is not None:
        sched_query["slot_number"] = slot_number
    slots = await db.weekly_schedule.find(sched_query).to_list(5000)

    # جلب أسماء المقررات/المعلمين للسياق
    course_ids = {s.get("course_id") for s in slots if s.get("course_id")}
    teacher_ids = {s.get("teacher_id") for s in slots if s.get("teacher_id")}
    courses_map = {}
    if course_ids:
        async for c in db.courses.find({"_id": {"$in": [ObjectId(x) for x in course_ids if x]}}):
            courses_map[str(c["_id"])] = c
    teachers_map = {}
    if teacher_ids:
        async for t in db.teachers.find({"_id": {"$in": [ObjectId(x) for x in teacher_ids if x]}}):
            teachers_map[str(t["_id"])] = t

    # فهرسة slots حسب (room, day, slot)
    busy_map = {}  # (room_id, day, slot) -> slot info
    for s in slots:
        rid = s.get("room_id")
        if not rid:
            continue
        key = (rid, s.get("day_of_week"), s.get("slot_number"))
        course = courses_map.get(s.get("course_id", ""), {})
        teacher = teachers_map.get(s.get("teacher_id", ""), {})
        busy_map[key] = {
            "course_name": course.get("name", ""),
            "course_code": course.get("code", ""),
            "teacher_name": teacher.get("full_name", ""),
            "section": s.get("section"),
            "level": s.get("level"),
        }

    result = []
    for room in rooms:
        rid = str(room["_id"])
        room_info = {
            "id": rid,
            "name": room.get("name", ""),
            "code": room.get("code", ""),
            "capacity": room.get("capacity"),
            "type": room.get("type", "regular"),
        }
        if day_of_week and slot_number is not None:
            # حالة محددة لقاعة في وقت محدد
            busy = busy_map.get((rid, day_of_week, slot_number))
            room_info["status"] = "busy" if busy else "free"
            room_info["details"] = busy
        else:
            # heatmap: كل اليوم/الفترة
            room_info["slots"] = []
            for (br, bd, bs), info in busy_map.items():
                if br == rid:
                    room_info["slots"].append({
                        "day": bd, "slot": bs, **info,
                    })
        result.append(room_info)

    # ترتيب: الفارغة أولاً عند الفلتر المحدد
    if day_of_week and slot_number is not None:
        result.sort(key=lambda r: (0 if r.get("status") == "free" else 1, r.get("name", "")))

    return {
        "filter": {
            "day_of_week": day_of_week,
            "slot_number": slot_number,
            "faculty_id": faculty_id,
            "semester_id": semester_id,
        },
        "total": len(result),
        "free_count": sum(1 for r in result if r.get("status") == "free"),
        "busy_count": sum(1 for r in result if r.get("status") == "busy"),
        "rooms": result,
    }


@router.get("/weekly-schedule/availability/teachers")
async def teachers_availability(
    day_of_week: Optional[str] = None,
    slot_number: Optional[int] = None,
    faculty_id: Optional[str] = None,
    department_id: Optional[str] = None,
    semester_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """تقرير حالة الأساتذة (فارغ/مشغول) مع الإشارة لتفضيلاتهم.
    - مع day + slot: يرجع قائمة الأساتذة الفارغين والمشغولين + ما إذا كان الوقت ضمن تفضيلاتهم
    - بدون فلتر: heatmap كامل
    """
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()

    # جلب الأساتذة
    teachers_query = {"is_active": {"$ne": False}}
    if department_id:
        teachers_query["department_id"] = department_id
    elif faculty_id:
        teachers_query["faculty_id"] = faculty_id
    teachers = await db.teachers.find(teachers_query).to_list(500)

    # جلب الجدول
    sched_query = {}
    if faculty_id:
        sched_query["faculty_id"] = faculty_id
    if department_id:
        sched_query["department_id"] = department_id
    if semester_id:
        sched_query["semester_id"] = semester_id
    if day_of_week:
        sched_query["day_of_week"] = day_of_week
    if slot_number is not None:
        sched_query["slot_number"] = slot_number
    slots = await db.weekly_schedule.find(sched_query).to_list(5000)

    # جلب التفضيلات
    teacher_ids_list = [str(t["_id"]) for t in teachers]
    prefs_list = await db.teacher_preferences.find({"teacher_id": {"$in": teacher_ids_list}}).to_list(500)
    prefs_map = {p["teacher_id"]: p for p in prefs_list}

    # سياق المقرر والقاعة
    course_ids = {s.get("course_id") for s in slots if s.get("course_id")}
    room_ids_set = {s.get("room_id") for s in slots if s.get("room_id")}
    courses_map = {}
    if course_ids:
        async for c in db.courses.find({"_id": {"$in": [ObjectId(x) for x in course_ids if x]}}):
            courses_map[str(c["_id"])] = c
    rooms_map = {}
    if room_ids_set:
        async for r in db.rooms.find({"_id": {"$in": [ObjectId(x) for x in room_ids_set if x]}}):
            rooms_map[str(r["_id"])] = r

    # فهرسة slots
    busy_map = {}  # (teacher_id, day, slot) -> info
    for s in slots:
        tid = s.get("teacher_id")
        if not tid:
            continue
        key = (tid, s.get("day_of_week"), s.get("slot_number"))
        course = courses_map.get(s.get("course_id", ""), {})
        room = rooms_map.get(s.get("room_id", ""), {})
        busy_map[key] = {
            "course_name": course.get("name", ""),
            "course_code": course.get("code", ""),
            "room_name": room.get("name", ""),
            "section": s.get("section"),
            "level": s.get("level"),
        }

    def check_preference(pref: dict, day: str, slot: int):
        """يرجع: 'preferred' | 'unavailable' | 'neutral'"""
        if not pref:
            return "neutral"
        # غير متاح
        unavail = pref.get("unavailable_slots", [])
        for u in unavail:
            if u.get("day") == day and u.get("slot") == slot:
                return "unavailable"
        # مفضّل
        preferred = pref.get("preferred_slots", [])
        for p in preferred:
            if p.get("day") == day and p.get("slot") == slot:
                return "preferred"
        # أيام/أوقات مفضّلة عامة
        preferred_days = pref.get("preferred_days", [])
        if preferred_days and day not in preferred_days:
            return "non_preferred_day"
        return "neutral"

    result = []
    for teacher in teachers:
        tid = str(teacher["_id"])
        pref = prefs_map.get(tid, {})
        teacher_info = {
            "id": tid,
            "full_name": teacher.get("full_name", ""),
            "employee_id": teacher.get("employee_id", ""),
            "department_id": teacher.get("department_id"),
            "has_preferences": bool(pref),
        }
        if day_of_week and slot_number is not None:
            busy = busy_map.get((tid, day_of_week, slot_number))
            teacher_info["status"] = "busy" if busy else "free"
            teacher_info["details"] = busy
            teacher_info["preference_status"] = check_preference(pref, day_of_week, slot_number)
        else:
            teacher_info["slots"] = []
            for (bt, bd, bs), info in busy_map.items():
                if bt == tid:
                    teacher_info["slots"].append({
                        "day": bd, "slot": bs, **info,
                    })
        result.append(teacher_info)

    # ترتيب: الفارغين المفضّلين أولاً
    if day_of_week and slot_number is not None:
        def sort_key(t):
            status_order = 0 if t.get("status") == "free" else 2
            pref_order = {"preferred": 0, "neutral": 1, "non_preferred_day": 2, "unavailable": 3}.get(t.get("preference_status", "neutral"), 1)
            return (status_order + pref_order, t.get("full_name", ""))
        result.sort(key=sort_key)

    free_preferred = sum(1 for t in result if t.get("status") == "free" and t.get("preference_status") == "preferred")
    free_neutral = sum(1 for t in result if t.get("status") == "free" and t.get("preference_status") == "neutral")
    free_unavailable = sum(1 for t in result if t.get("status") == "free" and t.get("preference_status") == "unavailable")
    free_total = sum(1 for t in result if t.get("status") == "free")

    return {
        "filter": {
            "day_of_week": day_of_week,
            "slot_number": slot_number,
            "faculty_id": faculty_id,
            "department_id": department_id,
            "semester_id": semester_id,
        },
        "total": len(result),
        "free_total": free_total,
        "free_preferred": free_preferred,
        "free_neutral": free_neutral,
        "free_unavailable": free_unavailable,
        "busy_count": sum(1 for t in result if t.get("status") == "busy"),
        "teachers": result,
    }


# ===== 🗓️ توليد المحاضرات الفعلية من الجدول الأسبوعي =====

_ARABIC_DAY_TO_WEEKDAY = {
    "السبت": 5, "الأحد": 6, "الاثنين": 0, "الإثنين": 0,
    "الثلاثاء": 1, "الأربعاء": 2, "الخميس": 3, "الجمعة": 4,
}


class GenerateLecturesFromScheduleRequest(BaseModel):
    faculty_id: str
    department_id: Optional[str] = None
    start_date: str  # YYYY-MM-DD
    end_date: str    # YYYY-MM-DD
    holidays: List[str] = []  # تواريخ معطلة تُتخطى
    dry_run: bool = True  # معاينة فقط دون إنشاء


@router.post("/weekly-schedule/generate-lectures")
async def generate_lectures_from_schedule(
    data: GenerateLecturesFromScheduleRequest,
    current_user: dict = Depends(get_current_user),
):
    """
    🗓️ تحويل الجدول الأسبوعي المعتمد إلى محاضرات فعلية بتواريخ محددة.
    - توليد الناقص فقط: المحاضرات الموجودة مسبقاً (من صفحة المقررات أو توليد سابق) تُتخطى ولا تُمس.
    - يتخطى العطلات المحددة.
    - dry_run=True: معاينة (عدد ما سيُنشأ/يُتخطى) دون أي إنشاء.
    """
    if not can_manage_schedule(current_user):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    db = get_db()

    # التحقق من التواريخ
    try:
        start_d = datetime.strptime(data.start_date, "%Y-%m-%d").date()
        end_d = datetime.strptime(data.end_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="صيغة التاريخ غير صحيحة (YYYY-MM-DD)")
    if end_d < start_d:
        raise HTTPException(status_code=400, detail="تاريخ النهاية قبل تاريخ البداية")
    if (end_d - start_d).days > 370:
        raise HTTPException(status_code=400, detail="الفترة أطول من سنة — قلّص النطاق")
    holidays_set = set(data.holidays or [])

    # خانات الجدول ضمن النطاق
    sched_query = {"faculty_id": data.faculty_id}
    if data.department_id:
        sched_query["department_id"] = data.department_id
    slots = await db.weekly_schedule.find(sched_query).to_list(5000)
    if not slots:
        raise HTTPException(status_code=400, detail="لا يوجد جدول أسبوعي في هذا النطاق — ولّد الجدول أولاً")

    # أوقات الفترات (خاصة بالكلية أو العامة)
    settings = await db.schedule_settings.find_one({"_id": f"faculty_{data.faculty_id}"})
    if not settings:
        settings = await db.schedule_settings.find_one({"_id": "global"})
    time_slots = (settings or {}).get("time_slots", [])
    slot_times = {s["slot_number"]: (s.get("start_time", ""), s.get("end_time", "")) for s in time_slots}

    # أسماء القاعات (الجدول يخزن room_id والمحاضرات تخزن الاسم نصاً)
    room_ids = list({s.get("room_id") for s in slots if s.get("room_id")})
    room_names = {}
    if room_ids:
        try:
            async for r in db.rooms.find({"_id": {"$in": [ObjectId(rid) for rid in room_ids]}}):
                room_names[str(r["_id"])] = r.get("name", "")
        except Exception:
            pass

    # بناء المرشحين: لكل خانة × كل تاريخ مطابق لليوم ضمن الفترة
    from datetime import timedelta as _td
    candidates = []
    skipped_no_time = 0
    d = start_d
    while d <= end_d:
        date_str = d.strftime("%Y-%m-%d")
        if date_str not in holidays_set:
            wd = d.weekday()
            for s in slots:
                if _ARABIC_DAY_TO_WEEKDAY.get(s.get("day", "")) != wd:
                    continue
                st_time, en_time = slot_times.get(s.get("slot_number"), ("", ""))
                if not st_time or not en_time:
                    skipped_no_time += 1
                    continue
                candidates.append({
                    "course_id": s.get("course_id", ""),
                    "date": date_str,
                    "start_time": st_time,
                    "end_time": en_time,
                    "room": room_names.get(s.get("room_id", ""), ""),
                })
        d += _td(days=1)

    if not candidates:
        return {"dry_run": data.dry_run, "to_create": 0, "already_exist": 0, "holidays_skipped": len(holidays_set), "courses_count": 0, "message": "لا توجد مواعيد ضمن الفترة المحددة"}

    # المحاضرات الموجودة مسبقاً (نحترمها ونتخطاها) — بأي حالة غير ملغاة
    course_ids = list({c["course_id"] for c in candidates if c["course_id"]})
    existing = set()
    async for lec in db.lectures.find({
        "course_id": {"$in": course_ids},
        "date": {"$gte": data.start_date, "$lte": data.end_date},
        "status": {"$ne": "cancelled"},
    }, {"course_id": 1, "date": 1, "start_time": 1}):
        existing.add((lec.get("course_id"), lec.get("date"), lec.get("start_time")))

    to_create = [c for c in candidates if (c["course_id"], c["date"], c["start_time"]) not in existing]
    already = len(candidates) - len(to_create)

    if data.dry_run:
        return {
            "dry_run": True,
            "to_create": len(to_create),
            "already_exist": already,
            "skipped_no_time": skipped_no_time,
            "courses_count": len(course_ids),
            "schedule_slots": len(slots),
            "date_range": f"{data.start_date} → {data.end_date}",
            "holidays_count": len(holidays_set),
            "message": f"سيتم إنشاء {len(to_create)} محاضرة لـ{len(course_ids)} مقرراً (تخطي {already} موجودة مسبقاً)",
        }

    # التنفيذ الفعلي — الفهرس الفريد uniq_course_date_start شبكة أمان إضافية
    created = 0
    dup_skipped = 0
    now = datetime.now(timezone.utc)
    for c in to_create:
        try:
            await db.lectures.insert_one({
                "course_id": c["course_id"],
                "date": c["date"],
                "start_time": c["start_time"],
                "end_time": c["end_time"],
                "room": c["room"],
                "status": "scheduled",
                "notes": "",
                "created_at": now,
                "created_by": current_user.get("id", ""),
                "generated_from_schedule": True,
            })
            created += 1
        except DuplicateKeyError:
            dup_skipped += 1

    await log_activity(
        current_user, "generate_lectures_from_schedule", "weekly_schedule", data.faculty_id, None,
        {"department_id": data.department_id, "range": f"{data.start_date}→{data.end_date}", "created": created, "skipped": already + dup_skipped},
    )
    return {
        "dry_run": False,
        "created": created,
        "already_exist": already + dup_skipped,
        "courses_count": len(course_ids),
        "message": f"تم إنشاء {created} محاضرة لـ{len(course_ids)} مقرراً (تخطي {already + dup_skipped} موجودة مسبقاً)",
    }
